from asyncio.windows_events import NULL
from PyQt5.QtWidgets import (
    QWidget, QHBoxLayout, QVBoxLayout, QGroupBox, QLineEdit, QPushButton, QListWidget, QTreeWidget,
    QListWidgetItem, QTreeWidgetItem, QSplitter, QTabWidget, QLabel, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QTextEdit, QComboBox, QSpinBox, QDateTimeEdit,
    QMessageBox, QInputDialog, QFileDialog, QMenu, QAction, QDialog, QDialogButtonBox, QFrame,
    QPlainTextEdit, QStackedWidget, QSizePolicy, QCheckBox, QGridLayout, QScrollArea,
    QGraphicsDropShadowEffect, QToolTip, QStyledItemDelegate, QStyle, QStyleOptionViewItem,
    QApplication, QProgressDialog, QShortcut,
)
from PyQt5.QtCore import Qt, QDateTime, pyqtSignal, QTimer, QObject, QThread, QSize, QRect, QRectF, QPointF, QEvent, QSettings
from PyQt5.QtGui import QColor, QPen, QFont, QBrush, QCursor, QPalette, QFontMetrics, QPolygonF, QKeySequence
from datetime import datetime, timedelta
from pages.part_dialog import PartDialog
from collections import deque, Counter, defaultdict
import time
import json
import re

import os
import sys
import subprocess
import csv
from datetime import datetime

from core.session_manager import SessionManager
from core.services.diag_service import DiagService
from core.services.project_service import ProjectService
from core.services.part_file_service import PartFileService
from core.services.managed_file_service import ManagedFileService
from core.services.package_export_service import PackageExportService
from core.services.ui_permission import UIPermissionHelper
from core.services.baseline_service import BaselineService
from core.services.part_doc_ack_service import PartDocAckService
from core.services.issue_service import IssueService
from core.services.traceability_service import TraceabilityService
from core.services.assembly_configuration_service import AssemblyConfigurationService
from core.services.undo_service import UndoService
from core.bom_filter import (
    deduplicate_bom_items_by_id,
    matches_bom_filter_text,
    split_bom_filter_terms,
)
from core.repositories.commit_repository import CommitRepository
from core.services.commit_service import CommitService
from pages.dialogs.package_parts_dialog import PackagePartsDialog
from pages.dialogs.assembly_iteration_compare_dialog import AssemblyIterationCompareDialog
from pages.dialogs.checkout_review_dialog import CheckoutReviewDialog
from pages.dialogs.cad_workspace_dialogs import (
    WorkspaceManagerDialog,
    WorkspaceSelectionDialog,
)
from core.services.cad_workspace_service import CadWorkspaceService
from pages.dialogs.assembly_configuration_dialogs import (
    CreateAssemblyConfigurationDialog,
    ManageAssemblyConfigurationsDialog,
)
from pages.dialogs.windchill_compare_dialog import WindchillCompareSetupDialog
from pages.pdf_viewer_widget import PdfViewerWidget
from pages.rich_text_image_editor import html_to_plain_text, looks_like_html
from utils import safe_exists, safe_startfile
from PyQt5.QtGui import QKeyEvent
from PyQt5.QtGui import QKeyEvent, QPalette, QColor
from PyQt5.QtGui import QPainter, QPixmap, QIcon



class FileDropTable(QTableWidget):
    filesDropped = pyqtSignal(list)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setAcceptDrops(True)
        self.setDragDropMode(QAbstractItemView.DropOnly)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            return
        super().dragEnterEvent(event)

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            return
        super().dragMoveEvent(event)

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            paths = []
            for url in event.mimeData().urls():
                p = url.toLocalFile()
                if p:
                    paths.append(p)
            if paths:
                self.filesDropped.emit(paths)
            event.acceptProposedAction()
            return
        super().dropEvent(event)


class BomTreeWidget(QTreeWidget):
    reorderRequested = pyqtSignal(list, int, int, str)
    folderReorderRequested = pyqtSignal(list, int, str)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setDragEnabled(True)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)
        self.setDragDropMode(QAbstractItemView.InternalMove)
        self.setDefaultDropAction(Qt.MoveAction)
        self._branch_spinner_angle = 0
        self._loading_branch_items = set()
        self._branch_spinner_timer = QTimer(self)
        self._branch_spinner_timer.setInterval(45)
        self._branch_spinner_timer.timeout.connect(self._advance_branch_spinners)
        self._context_menu_selection = []

    def setItemLoading(self, item: QTreeWidgetItem, loading: bool) -> None:
        if item is None:
            return
        item.setData(0, BOM_TREE_LOADING_ROLE, bool(loading))
        item_key = id(item)
        if loading:
            self._loading_branch_items.add(item_key)
            if not self._branch_spinner_timer.isActive():
                self._branch_spinner_timer.start()
        else:
            self._loading_branch_items.discard(item_key)
            if not self._loading_branch_items:
                self._branch_spinner_timer.stop()
        if loading:
            self.viewport().repaint()
        else:
            self.viewport().update()

    def resetLoadingIndicators(self) -> None:
        self._loading_branch_items.clear()
        self._branch_spinner_timer.stop()
        self.viewport().update()

    def mousePressEvent(self, event):
        if event.button() == Qt.RightButton:
            item = self.itemAt(event.pos())
            selected = [row for row in self.selectedItems() if row is not None]
            if item is not None and item in selected:
                self._context_menu_selection = selected
            else:
                self._context_menu_selection = [item] if item is not None else []
            if item is not None and item.isSelected():
                self.setCurrentItem(item)
                event.accept()
                return
        super().mousePressEvent(event)

    def _advance_branch_spinners(self) -> None:
        self._branch_spinner_angle = (self._branch_spinner_angle + 30) % 360
        self.viewport().update()

    def drawBranches(self, painter: QPainter, rect: QRect, index) -> None:
        super().drawBranches(painter, rect, index)
        try:
            item = self.itemFromIndex(index)
            if item is None or not item.data(0, BOM_TREE_LOADING_ROLE):
                return

            size = 10
            center_x = rect.right() - max(5, self.indentation() // 2)
            spinner_rect = QRect(
                center_x - size // 2,
                rect.center().y() - size // 2,
                size,
                size,
            )
            if item.isSelected():
                background = self.palette().brush(QPalette.Highlight)
            elif self.alternatingRowColors() and index.row() % 2:
                background = self.palette().brush(QPalette.AlternateBase)
            else:
                background = self.palette().brush(QPalette.Base)

            painter.save()
            painter.fillRect(spinner_rect.adjusted(-2, -2, 2, 2), background)
            painter.setRenderHint(QPainter.Antialiasing, True)
            pen = QPen(QColor("#2563eb"), 1.7)
            pen.setCapStyle(Qt.RoundCap)
            painter.setPen(pen)
            painter.setBrush(Qt.NoBrush)
            painter.drawArc(
                spinner_rect,
                int(self._branch_spinner_angle * 16),
                int(245 * 16),
            )
            painter.restore()
        except Exception:
            return

    def dropEvent(self, event):
        target = self.itemAt(event.pos())
        selected = [item for item in self.selectedItems() if item is not None]
        if not target or not selected:
            event.ignore()
            return

        indicator = self.dropIndicatorPosition()
        if indicator not in (QAbstractItemView.AboveItem, QAbstractItemView.BelowItem):
            event.ignore()
            return

        target_folder_id = target.data(0, BOM_TREE_FOLDER_ROLE)
        if target_folder_id is not None:
            if target in selected:
                event.ignore()
                return
            visual_parent = target.parent()
            if any(item.parent() is not visual_parent for item in selected):
                event.ignore()
                return
            selected_folder_ids = []
            for item in selected:
                folder_id = item.data(0, BOM_TREE_FOLDER_ROLE)
                if folder_id is None:
                    event.ignore()
                    return
                selected_folder_ids.append(int(folder_id))
            where = "above" if indicator == QAbstractItemView.AboveItem else "below"
            self.folderReorderRequested.emit(
                selected_folder_ids, int(target_folder_id), where
            )
            event.acceptProposedAction()
            return

        target_id = target.data(0, Qt.UserRole)
        target_parent = target.parent()
        if target_parent is None:
            event.ignore()
            return
        engineering_parent = target_parent
        while engineering_parent is not None and engineering_parent.data(0, Qt.UserRole) is None:
            engineering_parent = engineering_parent.parent()
        target_parent_id = (
            engineering_parent.data(0, Qt.UserRole)
            if engineering_parent is not None else None
        )
        if any(item.parent() is not target_parent for item in selected):
            event.ignore()
            return
        if target in selected:
            event.ignore()
            return

        scope = str(self.property("pdmScope") or "").upper()
        selected_ids = []
        if scope == "CAD" or target.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
            parent_cad_id = target_parent.data(0, PDM_CAD_DOCUMENT_ID_ROLE)
            target_member_id = target.data(0, PDM_CAD_MEMBER_ID_ROLE)
            if parent_cad_id is None or target_member_id is None:
                event.ignore()
                return
            for item in selected:
                if item.data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD:
                    event.ignore()
                    return
                member_id = item.data(0, PDM_CAD_MEMBER_ID_ROLE)
                if member_id is None:
                    event.ignore()
                    return
                selected_ids.append(int(member_id))
            target_id = int(target_member_id)
            target_parent_id = int(parent_cad_id)
        elif scope == "EBOM" or target.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_ITEM:
            occurrence = target.data(0, BOM_TREE_OCCURRENCE_ROLE) or {}
            target_usage_id = occurrence.get("usage_id")
            if target_usage_id is None:
                event.ignore()
                return
            for item in selected:
                if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                    event.ignore()
                    return
                usage_id = (item.data(0, BOM_TREE_OCCURRENCE_ROLE) or {}).get("usage_id")
                if usage_id is None:
                    event.ignore()
                    return
                selected_ids.append(int(usage_id))
            target_id = int(target_usage_id)
            target_parent_id = int(target_parent_id)
        else:
            for item in selected:
                try:
                    selected_ids.append(int(item.data(0, Qt.UserRole)))
                except Exception:
                    pass
            if len(selected_ids) != len(selected) or target_id is None or target_parent_id is None:
                event.ignore()
                return

        if indicator == QAbstractItemView.AboveItem:
            where = "above"
        else:
            where = "below"

        self.reorderRequested.emit(selected_ids, int(target_id), int(target_parent_id), where)
        event.acceptProposedAction()


class InlineSpinner(QWidget):
    def __init__(self, parent=None, size: int = 24, color: QColor | None = None):
        super().__init__(parent)
        self._angle = 0
        self._size = int(size)
        self._color = color or QColor(107, 114, 128)
        self.setFixedSize(self._size, self._size)

        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)
        self._timer.start(30)

    def _tick(self):
        self._angle = (self._angle + 25) % 360
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        pen = QPen(self._color, 2)
        pen.setCapStyle(Qt.RoundCap)
        painter.setPen(pen)
        painter.setBrush(Qt.NoBrush)

        pad = 3
        r = self.rect().adjusted(pad, pad, -pad, -pad)
        start = int(self._angle * 16)
        span = int(120 * 16)
        painter.drawArc(r, start, span)


# Role storing the optional "In Work" / "In Work (user)" suffix (painted beside name in column 0).
BOM_TREE_INWORK_ROLE = Qt.UserRole + 33
BOM_TREE_IS_ASSEMBLY_ROLE = Qt.UserRole + 35
# Value: dict with "pdf" / "step" -> tuple (kind: str, tooltip: str); kind in ok|outdated|missing|na
BOM_TREE_FILES_ROLE = Qt.UserRole + 36
# Value: dict { "state": "ok"|"warn", "tooltip": str }
BOM_TREE_INTEGRITY_ROLE = Qt.UserRole + 37
# Value: dict {active_count, total_count, critical_count}
BOM_TREE_ISSUE_ROLE = Qt.UserRole + 38
BOM_TREE_CATEGORY_ROLE = Qt.UserRole + 39
BOM_TREE_FOLDER_ROLE = Qt.UserRole + 40
BOM_TREE_CHILDREN_LOADED_ROLE = Qt.UserRole + 41
BOM_TREE_PATH_ROLE = Qt.UserRole + 42
BOM_TREE_PLACEHOLDER_ROLE = Qt.UserRole + 43
BOM_TREE_LOADING_ROLE = Qt.UserRole + 44
BOM_TREE_REPLACE_CHILDREN_ROLE = Qt.UserRole + 45
BOM_TREE_BINDING_UPDATE_ROLE = Qt.UserRole + 46
BOM_TREE_POLICY_ROLE = Qt.UserRole + 47
BOM_TREE_OCCURRENCE_ROLE = Qt.UserRole + 48
BOM_TREE_PROMOTION_ROLE = Qt.UserRole + 49
BOM_TREE_ITEM_NUMBER_ROLE = Qt.UserRole + 50
BOM_TREE_AES_NUMBER_ROLE = Qt.UserRole + 51
BOM_TREE_FOLDER_SCOPE_ROLE = Qt.UserRole + 52
STRUCTURE_CURRENT_ITERATION_ROLE = Qt.UserRole + 60
STRUCTURE_BOUND_ITERATION_ROLE = Qt.UserRole + 61
STRUCTURE_LATEST_ITERATION_ROLE = Qt.UserRole + 62
# Explicit PDM object identity.  Qt.UserRole remains the legacy Item/BOM id so
# existing Item actions cannot accidentally interpret a CAD Document id as an
# Item id.
PDM_OBJECT_KIND_ROLE = Qt.UserRole + 70
PDM_CAD_DOCUMENT_ID_ROLE = Qt.UserRole + 71
PDM_ASSOCIATION_ID_ROLE = Qt.UserRole + 72
PDM_ASSOCIATION_TYPE_ROLE = Qt.UserRole + 73
PDM_ASSOCIATED_ITEM_ID_ROLE = Qt.UserRole + 74
PDM_CAD_MEMBER_ID_ROLE = Qt.UserRole + 75
PDM_CAD_PAYLOAD_ROLE = Qt.UserRole + 76
PDM_CHILDREN_PAYLOAD_ROLE = Qt.UserRole + 77
PDM_EBOM_ASSOCIATIONS_ROLE = Qt.UserRole + 78
PDM_ASSOCIATIONS_SHOWN_ROLE = Qt.UserRole + 79
PDM_NODE_PAYLOAD_ROLE = Qt.UserRole + 80
COMPARE_ITEM_ID_ROLE = Qt.UserRole + 90
COMPARE_CAD_ID_ROLE = Qt.UserRole + 91
COMPARE_PAYLOAD_ROLE = Qt.UserRole + 92
COMPARE_PARENT_ITEM_ID_ROLE = Qt.UserRole + 93
COMPARE_ITEM_IDS_ROLE = Qt.UserRole + 94

PDM_OBJECT_ITEM = "ITEM"
PDM_OBJECT_CAD = "CAD_DOCUMENT"

_BOM_TYPE_ICON_CACHE = {}
_PDM_OBJECT_ICON_CACHE = {}


def _bom_type_icon(part_type: str) -> QIcon:
    """Return a compact CAD-style icon for assembly and part tree rows."""
    key = str(part_type or "").strip().lower()
    if key in _BOM_TYPE_ICON_CACHE:
        return _BOM_TYPE_ICON_CACHE[key]
    if key == "folder":
        icon = QApplication.style().standardIcon(QStyle.SP_DirIcon)
        _BOM_TYPE_ICON_CACHE[key] = icon
        return icon
    if key not in {"asm", "assembly", "prt", "part"}:
        return QIcon()

    pixmap = QPixmap(14, 14)
    pixmap.fill(Qt.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.Antialiasing, True)

    if key in {"asm", "assembly"}:
        painter.setPen(QPen(QColor("#315b7d"), 1))
        painter.setBrush(QBrush(QColor("#70a6c9")))
        painter.drawRoundedRect(1, 1, 6, 6, 1, 1)
        painter.setBrush(QBrush(QColor("#4f86a8")))
        painter.drawRoundedRect(7, 7, 6, 6, 1, 1)
        painter.setPen(QPen(QColor("#315b7d"), 1.2))
        painter.drawLine(6, 6, 8, 8)
    else:
        painter.setPen(QPen(QColor("#536b3f"), 1))
        painter.setBrush(QBrush(QColor("#8db36f")))
        painter.drawRoundedRect(2, 2, 10, 10, 1, 1)
        painter.setPen(QPen(QColor("#6f9254"), 1))
        painter.drawLine(2, 5, 7, 8)
        painter.drawLine(12, 5, 7, 8)
        painter.drawLine(7, 8, 7, 12)

    painter.end()
    icon = QIcon(pixmap)
    _BOM_TYPE_ICON_CACHE[key] = icon
    return icon


def _pdm_item_icon() -> QIcon:
    """Return an Item/Part master icon that is visually distinct from CAD."""
    key = "pdm_item"
    if key in _PDM_OBJECT_ICON_CACHE:
        return _PDM_OBJECT_ICON_CACHE[key]
    pixmap = QPixmap(16, 16)
    pixmap.fill(Qt.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.Antialiasing, True)
    painter.setPen(QPen(QColor("#8a5a12"), 1))
    painter.setBrush(QBrush(QColor("#f2c56b")))
    painter.drawRoundedRect(2, 2, 12, 12, 2, 2)
    painter.setPen(QPen(QColor("#fff8e7"), 1.4))
    painter.drawLine(5, 6, 11, 6)
    painter.drawLine(5, 9, 10, 9)
    painter.end()
    icon = QIcon(pixmap)
    _PDM_OBJECT_ICON_CACHE[key] = icon
    return icon


def _pdm_cad_icon(category: str) -> QIcon:
    """Return category-specific blue CAD Document icons."""
    category = str(category or "OTHER").strip().upper()
    key = f"pdm_cad:{category}"
    if key in _PDM_OBJECT_ICON_CACHE:
        return _PDM_OBJECT_ICON_CACHE[key]
    pixmap = QPixmap(16, 16)
    pixmap.fill(Qt.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.Antialiasing, True)
    painter.setPen(QPen(QColor("#24577a"), 1))
    if category == "ASSEMBLY":
        painter.setBrush(QBrush(QColor("#75b8df")))
        painter.drawRoundedRect(1, 2, 7, 7, 1, 1)
        painter.setBrush(QBrush(QColor("#438fbe")))
        painter.drawRoundedRect(8, 7, 7, 7, 1, 1)
        painter.drawLine(7, 8, 9, 9)
    elif category == "DRAWING":
        painter.setBrush(QBrush(QColor("#e8f4fb")))
        painter.drawRoundedRect(3, 1, 10, 14, 1, 1)
        painter.setPen(QPen(QColor("#438fbe"), 1))
        painter.drawLine(5, 5, 11, 5)
        painter.drawLine(5, 8, 11, 8)
        painter.drawLine(5, 11, 9, 11)
    else:
        painter.setBrush(QBrush(QColor("#69acd3")))
        painter.drawRoundedRect(2, 2, 12, 12, 2, 2)
        painter.setPen(QPen(QColor("#dff3ff"), 1))
        painter.drawLine(3, 6, 8, 9)
        painter.drawLine(13, 6, 8, 9)
        painter.drawLine(8, 9, 8, 13)
    painter.end()
    icon = QIcon(pixmap)
    _PDM_OBJECT_ICON_CACHE[key] = icon
    return icon


CAD_COL_NAME = 0
# Compatibility alias for integrations that used the former CAD-number column.
CAD_COL_NUMBER = CAD_COL_NAME
CAD_COL_ROW = CAD_COL_NAME
CAD_COL_FILE = CAD_COL_NAME
CAD_COL_DESCRIPTION = 1
CAD_COL_CATEGORY = 2
CAD_COL_REV = 3
CAD_COL_STATE = 4
CAD_COL_ASSOCIATION = 5
CAD_COL_CHECKOUT = 6
CAD_COL_BUILD = 7
CAD_COL_QTY = 8

BOM_COL_ROW = 0
BOM_COL_NAME = 1
BOM_COL_FILES = 2
BOM_COL_AES = 3
BOM_COL_TYPE = 4
BOM_COL_REV = 5
BOM_COL_STATUS = 6
BOM_COL_INTEGRITY = 7
BOM_BASE_TREE_COLUMN_COUNT = 8
EBOM_COL_SOURCE_QTY = 8
EBOM_COL_EFFECTIVE_QTY = 9
EBOM_COL_LEVEL = 10
BOM_COL_EXTRA_START = 11

BOM_EXTRA_COLUMN_SPECS = [
    ("part_number", "Item Number", 115, False, ("part_number", "number")),
    ("aes_number", "AES Number", 95, True, ("aes_number",)),
    ("drawing_number", "Drawing Number", 120, False, ("drawing_number", "drawing")),
    ("filename", "Native File", 135, False, ("filename", "base_file_name")),
    ("base_drw_name", "Drawing File", 135, False, ("base_drw_name",)),
    ("item_type", "Item Type", 115, False, ("item_type",)),
    ("assembly_mode", "Assembly Mode", 105, False, ("assembly_mode",)),
    ("classification", "Classification", 105, False, ("classification",)),
    ("procurement_source", "Source", 82, False, ("procurement_source", "source")),
    ("item_view", "View", 75, False, ("item_view", "view")),
    ("default_unit", "Unit", 58, False, ("default_unit", "unit")),
    ("material", "Material", 105, False, ("material",)),
    ("weight", "Weight", 75, False, ("weight",)),
    ("cad_requirement", "CAD Req.", 85, False, ("cad_requirement",)),
    ("drawing_requirement", "Drawing Req.", 95, False, ("drawing_requirement",)),
    ("cad_control_mode", "CAD Control", 120, False, ("cad_control_mode",)),
    ("default_ebom_behavior", "EBOM Behavior", 115, False, ("default_ebom_behavior",)),
    ("source_quantity", "Source Qty", 74, False, ("source_quantity", "quantity")),
    ("effective_quantity", "Effective Qty", 84, False, ("effective_quantity",)),
    ("level", "Level", 52, False, ("level",)),
    ("usage_id", "Usage ID", 70, False, ("usage_id",)),
    ("represented_part_id", "Represents", 80, False, ("represented_part_id",)),
    ("current_revision_id", "Revision ID", 82, False, ("current_revision_id",)),
    ("current_iteration_id", "Iteration ID", 82, False, ("current_iteration_id",)),
    ("pending_revision_code", "Pending Rev.", 90, False, ("pending_revision_code",)),
    ("released_by", "Released By", 90, False, ("released_by",)),
    ("released_at", "Released At", 125, False, ("released_at",)),
    ("created", "Created", 125, False, ("created", "created_at")),
    ("modified", "Modified", 125, False, ("modified", "modified_at")),
    ("notes", "Notes", 180, False, ("notes",)),
]
BOM_TREE_COLUMN_COUNT = BOM_COL_EXTRA_START + len(BOM_EXTRA_COLUMN_SPECS)

# Inline "In Work" beside name (spec)
_BOM_INWORK_COLOR = QColor("#BA7517")
_BOM_INWORK_GAP_PX = 6
_BOM_TREE_SEL_BG = "#e8eefc"
_BOM_TREE_ROW_TEXT = "#111827"

_FILE_BADGE_STYLES = {
    "ok": {
        "bg": "#EAF3DE", "fg": "#3B6D11", "dot": "#639922", "dash": False,
    },
    "outdated": {
        "bg": "#FAEEDA", "fg": "#854F0B", "dot": "#BA7517", "dash": False,
    },
    "missing": {
        "bg": "#FCEBEB", "fg": "#A32D2D", "dot": "#E24B4A", "dash": False,
    },
    "na": {
        "bg": "#F0F0F0", "fg": "#888888", "dot": "#AAAAAA", "dash": True,
    },
}

_STATUS_BADGE_STYLES = {
    "released": {"bg": "#EAF3DE", "fg": "#3B6D11"},
    "design": {"bg": "#E6F1FB", "fg": "#185FA5"},
    "in work": {"bg": "#FAEEDA", "fg": "#854F0B"},
    "obsolete": {"bg": "#F5F5F5", "fg": "#888888"},
}
# Modest corner radius — full pill (rx = half height) makes long labels look oblong.
_STATUS_BADGE_CORNER_RADIUS = 6


def _normalize_file_badge(doc_key: str, issues: set, doc_info: dict) -> tuple[str, str]:
    """Map PDF/STEP indicator state + integrity issues to badge kind and tooltip."""
    miss = f"missing_{doc_key}"
    out = f"outdated_{doc_key}"
    if miss in issues:
        tip = str((doc_info or {}).get("tooltip") or f"{doc_key.upper()}: missing — no file attached to this revision")
        return "missing", tip
    if out in issues:
        tip = str((doc_info or {}).get("tooltip") or f"{doc_key.upper()}: outdated — file is not the latest in working directory")
        return "outdated", tip
    state = str((doc_info or {}).get("state") or "absent").lower()
    tip = str((doc_info or {}).get("tooltip") or f"{doc_key.upper()}: unknown")
    tl = tip.lower()
    if state == "ok":
        return "ok", tip
    if state == "ack":
        return "ok", tip
    if state == "absent":
        return "na", tip
    if state == "bad":
        if "no attachment" in tl:
            return "missing", tip
        if "outdated" in tl or "newer" in tl or "review" in tl or "not the latest" in tl:
            return "outdated", tip
        if "missing" in tl:
            return "missing", tip
        return "outdated", tip
    return "na", tip


def _file_badges_payload(part_id, issues: set, summary: dict) -> dict:
    pdf_i = (summary or {}).get("pdf") or {}
    step_i = (summary or {}).get("step") or {}
    pk, pt = _normalize_file_badge("pdf", issues, pdf_i)
    sk, st = _normalize_file_badge("step", issues, step_i)
    return {"pdf": (pk, pt), "step": (sk, st)}


def _integrity_payload(part_id, missing_files, missing_ids: set) -> dict:
    lines = []
    try:
        pid = int(part_id)
    except Exception:
        pid = None
    if pid is not None:
        for row in (missing_files or []):
            try:
                bom_id, issue_type, filename = row
                if int(bom_id) != pid:
                    continue
                fn = str(filename or "")
                it = str(issue_type or "")
                if it == "missing_file":
                    lines.append(f"Missing file: {fn}")
                elif it == "outdated_file":
                    lines.append(f"Outdated file: {fn} is not the latest version in working directory.")
                elif it == "missing_drawing":
                    lines.append(f"Missing drawing: {fn}")
                elif it == "missing_pdf":
                    lines.append(f"Missing PDF: {fn}")
                elif it == "missing_step":
                    lines.append(f"Missing STEP: {fn}")
                else:
                    lines.append(f"{it}: {fn}".strip(": "))
            except Exception:
                continue
    warn = bool(lines) or (pid is not None and pid in (missing_ids or set()))
    if not warn:
        tip = "No integrity issues detected for this item."
    elif lines:
        tip = "\n".join(lines)
    else:
        tip = "BOM structure mismatch between PDF drawing and current assembly."
    return {"state": "warn" if warn else "ok", "tooltip": tip}


def _status_badge_key(raw: str) -> str | None:
    s = (raw or "").strip().lower()
    if not s:
        return None
    if "release" in s:
        return "released"
    if "obsolete" in s:
        return "obsolete"
    if "design" in s or "draft" in s:
        return "design"
    if "work" in s or "check" in s or "wip" in s:
        return "in work"
    return None


def _paint_file_pill(painter: QPainter, rect: QRect, label: str, kind: str) -> int:
    st = _FILE_BADGE_STYLES.get(kind, _FILE_BADGE_STYLES["na"])
    bg = QColor(st["bg"])
    fg = QColor(st["fg"])
    dot_col = QColor(st["dot"])

    f = QFont(painter.font())
    f.setPixelSize(10)
    f.setBold(True)
    painter.setFont(f)
    fm = QFontMetrics(f)
    text = (label or "").upper()
    dot_r = 5
    gap_after_dot = 4
    pad_h = 6
    pad_v = 2
    inner_w = dot_r + gap_after_dot + fm.horizontalAdvance(text)
    w = inner_w + pad_h * 2
    h = max(fm.height() + pad_v * 2, dot_r + pad_v * 2)
    pill = QRect(rect.left(), rect.center().y() - h // 2, w, h)

    painter.save()
    painter.setRenderHint(QPainter.Antialiasing, True)
    if st.get("dash"):
        pen = QPen(QColor("#CCCCCC"))
        pen.setStyle(Qt.DashLine)
        pen.setWidthF(0.5)
        painter.setPen(pen)
        painter.setBrush(QColor(st["bg"]))
        painter.drawRoundedRect(pill, 4, 4)
    else:
        painter.setPen(Qt.NoPen)
        painter.setBrush(bg)
        painter.drawRoundedRect(pill, 4, 4)

    cx = pill.left() + pad_h + dot_r // 2
    cy = pill.center().y()
    painter.setPen(Qt.NoPen)
    painter.setBrush(dot_col)
    painter.drawEllipse(QRect(cx - dot_r // 2, cy - dot_r // 2, dot_r, dot_r))

    painter.setPen(fg)
    painter.setFont(f)
    text_rect = QRect(
        pill.left() + pad_h + dot_r + gap_after_dot,
        pill.top(),
        pill.width() - (pad_h * 2 + dot_r + gap_after_dot),
        pill.height(),
    )
    painter.drawText(text_rect, Qt.AlignVCenter | Qt.AlignLeft, text)
    painter.restore()
    return w


def _files_delegate_pill_rects(option_rect: QRect, payload: dict) -> tuple[QRect | None, QRect | None]:
    pdf = payload.get("pdf") or ("na", "")
    step = payload.get("step") or ("na", "")
    r = option_rect.adjusted(4, 0, -4, 0)
    f = QFont()
    f.setPixelSize(10)
    f.setBold(True)
    fm = QFontMetrics(f)

    def _pill_w(kind: str, label: str) -> int:
        text = label.upper()
        dot_r = 5
        gap_after_dot = 4
        pad_h = 6
        pad_v = 2
        inner_w = dot_r + gap_after_dot + fm.horizontalAdvance(text)
        return inner_w + pad_h * 2

    w_pdf = _pill_w(pdf[0], "PDF")
    w_step = _pill_w(step[0], "STEP")
    gap = 4
    h = max(20, fm.height() + 4)
    y = r.center().y() - h // 2
    x0 = r.left()
    pdf_rect = QRect(x0, y, w_pdf, h)
    step_rect = QRect(x0 + w_pdf + gap, y, w_step, h)
    return pdf_rect, step_rect


class _BomTreeNameDelegate(QStyledItemDelegate):
    """Renders part name + optional inline 'In Work' label in column 0."""

    def __init__(self, tree: QTreeWidget, parent=None):
        super().__init__(parent)
        self._tree = tree

    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index):
        if index.column() != BOM_COL_NAME:
            return super().paint(painter, option, index)

        item = self._tree.itemFromIndex(index)
        if item is None:
            return super().paint(painter, option, index)

        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        name = item.text(BOM_COL_NAME) or ""
        suffix = item.data(0, BOM_TREE_INWORK_ROLE) or ""
        issue_summary = item.data(0, BOM_TREE_ISSUE_ROLE) or {}
        active_count = int(issue_summary.get("active_count") or 0)
        total_count = int(issue_summary.get("total_count") or 0)
        direct_active_count = int(issue_summary.get("direct_active_count", active_count) or 0)
        inherited_active_count = int(issue_summary.get("inherited_active_count") or 0)
        direct_total_count = int(issue_summary.get("direct_total_count", total_count) or 0)
        inherited_total_count = int(issue_summary.get("inherited_total_count") or 0)
        binding_update_count = int(item.data(0, BOM_TREE_BINDING_UPDATE_ROLE) or 0)
        policy = item.data(0, BOM_TREE_POLICY_ROLE) or {}
        promotion = item.data(0, BOM_TREE_PROMOTION_ROLE) or []
        widget = opt.widget or self._tree
        style = widget.style()

        opt.text = ""
        painter.save()
        style.drawControl(QStyle.CE_ItemViewItem, opt, painter, widget)
        painter.restore()

        text_rect = style.subElementRect(QStyle.SE_ItemViewItemText, opt, widget)
        if text_rect.width() <= 0:
            text_rect = opt.rect

        # Keep normal (dark) text on selection — do not use HighlightedText (often white on blue).
        name_pen = QColor(_BOM_TREE_ROW_TEXT)

        is_asm = bool(item.data(0, BOM_TREE_IS_ASSEMBLY_ROLE)) or any(
            item.child(index).data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD
            for index in range(item.childCount())
        )
        name_font = QFont(opt.font)
        name_font.setPixelSize(11)
        name_font.setBold(bool(is_asm))
        display_inwork = "In Work" if suffix else ""
        issue_badges = []
        if direct_active_count:
            issue_badges.append((f"!{direct_active_count}", QColor("#c62828")))
        if inherited_active_count:
            issue_badges.append((f"+{inherited_active_count}", QColor("#2563eb")))
        if not issue_badges and total_count:
            if direct_total_count:
                issue_badges.append(("●", QColor("#2e7d32")))
            elif inherited_total_count:
                issue_badges.append(("+", QColor("#2563eb")))
        if binding_update_count:
            issue_badges.append((f"v+{binding_update_count}", QColor("#b45309")))
        classification = str(policy.get("classification") or "PHYSICAL").upper()
        is_representation = bool(policy.get("represented_part_id"))
        is_supplier_package = str(
            policy.get("cad_control_mode") or "CONTROLLED"
        ).upper() == "SUPPLIER_PACKAGE"
        if is_supplier_package:
            issue_badges.append(("SUPPLIER PACKAGE", QColor("#9a3412")))
        if is_representation:
            issue_badges.append(("CAD REP", QColor("#6d28d9")))
        elif classification != "PHYSICAL":
            issue_badges.append((classification.replace("_", " "), QColor("#6d28d9")))
        behavior = str(policy.get("resolved_ebom_behavior") or "NORMAL").upper()
        if is_representation:
            pass
        elif behavior == "FLATTEN":
            issue_badges.append(("FLATTEN", QColor("#0369a1")))
        elif behavior == "EXCLUDE":
            issue_badges.append(("NOT FOR DELIVERY", QColor("#b91c1c")))
        if promotion:
            issue_badges.append(("PROMOTED", QColor("#0f766e")))

        painter.save()
        painter.setFont(name_font)
        painter.setPen(name_pen)

        fm = QFontMetrics(name_font)
        issue_font = QFont(opt.font)
        issue_font.setPixelSize(11)
        issue_font.setBold(True)
        issue_fm = QFontMetrics(issue_font)
        issue_w = sum(issue_fm.horizontalAdvance(label) + 4 for label, _color in issue_badges)
        if len(issue_badges) > 1:
            issue_w += _BOM_INWORK_GAP_PX * (len(issue_badges) - 1)

        if display_inwork:
            suf_font = QFont(opt.font)
            suf_font.setPixelSize(10)
            suf_font.setBold(False)
            suf_font.setWeight(QFont.Normal)
            suf_fm = QFontMetrics(suf_font)
            half = max(48, text_rect.width() // 2)
            suf_elided = suf_fm.elidedText(display_inwork, opt.textElideMode, half)
            suf_w = suf_fm.horizontalAdvance(suf_elided)
            name_max = max(24, text_rect.width() - _BOM_INWORK_GAP_PX - suf_w - (_BOM_INWORK_GAP_PX + issue_w if issue_badges else 0))
            name_elided = fm.elidedText(name, opt.textElideMode, name_max)
        else:
            suf_font = None
            suf_elided = ""
            name_elided = fm.elidedText(
                name, opt.textElideMode,
                max(24, text_rect.width() - (_BOM_INWORK_GAP_PX + issue_w if issue_badges else 0)),
            )

        painter.drawText(text_rect, Qt.AlignVCenter | Qt.AlignLeft, name_elided)
        name_w = fm.horizontalAdvance(name_elided)

        if display_inwork and suf_font is not None:
            painter.setFont(suf_font)
            suf_rect = QRect(
                text_rect.left() + name_w + _BOM_INWORK_GAP_PX,
                text_rect.top(),
                max(0, text_rect.right() - (text_rect.left() + name_w + _BOM_INWORK_GAP_PX)),
                text_rect.height(),
            )
            painter.setPen(_BOM_INWORK_COLOR)
            painter.drawText(suf_rect, Qt.AlignVCenter | Qt.AlignLeft, suf_elided)
            name_w += _BOM_INWORK_GAP_PX + QFontMetrics(suf_font).horizontalAdvance(suf_elided)
        if issue_badges:
            painter.setFont(issue_font)
            x = text_rect.left() + name_w + _BOM_INWORK_GAP_PX
            for label, color in issue_badges:
                label_w = issue_fm.horizontalAdvance(label) + 4
                painter.setPen(color)
                issue_rect = QRect(x, text_rect.top(), label_w, text_rect.height())
                painter.drawText(issue_rect, Qt.AlignVCenter | Qt.AlignLeft, label)
                x += label_w + _BOM_INWORK_GAP_PX
        painter.restore()

    def sizeHint(self, option: QStyleOptionViewItem, index):
        sh = super().sizeHint(option, index)
        if index.column() != BOM_COL_NAME:
            return sh
        item = self._tree.itemFromIndex(index)
        if not item:
            return sh
        suffix = item.data(0, BOM_TREE_INWORK_ROLE) or ""
        if not suffix:
            return sh
        suf_font = QFont(option.font)
        suf_font.setPixelSize(10)
        extra = _BOM_INWORK_GAP_PX + QFontMetrics(suf_font).horizontalAdvance("In Work")
        return QSize(sh.width() + extra, max(sh.height(), 22, QFontMetrics(suf_font).height()))


class _BomTreeFilesDelegate(QStyledItemDelegate):
    """PDF + STEP pill badges in column 1."""

    def __init__(self, tree: QTreeWidget, parent=None):
        super().__init__(parent)
        self._tree = tree

    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index):
        if index.column() != BOM_COL_FILES:
            return super().paint(painter, option, index)
        item = self._tree.itemFromIndex(index)
        if item is None:
            return super().paint(painter, option, index)

        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        opt.text = ""
        opt.icon = QIcon()
        widget = opt.widget or self._tree
        style = widget.style()
        painter.save()
        style.drawControl(QStyle.CE_ItemViewItem, opt, painter, widget)
        painter.restore()

        if item.data(0, BOM_TREE_FOLDER_ROLE):
            return

        if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
            return

        policy = item.data(0, BOM_TREE_POLICY_ROLE) or {}
        if policy.get("represented_part_id"):
            return

        payload = item.data(BOM_COL_FILES, BOM_TREE_FILES_ROLE) or {}
        pdf = payload.get("pdf") or ("na", "")
        step = payload.get("step") or ("na", "")
        r = opt.rect.adjusted(4, 0, -4, 0)
        x = r.left()
        x += _paint_file_pill(painter, QRect(x, r.top(), 0, r.height()), "PDF", pdf[0])
        x += 4
        _paint_file_pill(painter, QRect(x, r.top(), 0, r.height()), "STEP", step[0])

    def helpEvent(self, event, view, option, index):
        if event.type() != QEvent.ToolTip or index.column() != BOM_COL_FILES:
            return super().helpEvent(event, view, option, index)
        item = self._tree.itemFromIndex(index)
        if not item:
            return super().helpEvent(event, view, option, index)
        if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
            return super().helpEvent(event, view, option, index)
        policy = item.data(0, BOM_TREE_POLICY_ROLE) or {}
        if policy.get("represented_part_id"):
            QToolTip.showText(
                event.globalPos(),
                "CAD-only representation: PDF and STEP delivery files belong to the linked physical part.",
                view,
            )
            return True
        payload = item.data(BOM_COL_FILES, BOM_TREE_FILES_ROLE) or {}
        pdf_rect, step_rect = _files_delegate_pill_rects(option.rect, payload)
        try:
            pos = view.viewport().mapFromGlobal(event.globalPos())
        except Exception:
            try:
                pos = event.pos()
            except Exception:
                return super().helpEvent(event, view, option, index)
        pdf = payload.get("pdf") or ("na", "PDF: unknown")
        step = payload.get("step") or ("na", "STEP: unknown")
        tip = ""
        if pdf_rect.contains(pos):
            tip = str(pdf[1] or "")
        elif step_rect.contains(pos):
            tip = str(step[1] or "")
        if tip:
            QToolTip.showText(event.globalPos(), tip, view)
            return True
        return super().helpEvent(event, view, option, index)


class _BomTreeStatusDelegate(QStyledItemDelegate):
    """Status column as a pill badge."""

    def __init__(self, tree: QTreeWidget, parent=None):
        super().__init__(parent)
        self._tree = tree

    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index):
        if index.column() != BOM_COL_STATUS:
            return super().paint(painter, option, index)
        item = self._tree.itemFromIndex(index)
        if item is None:
            return super().paint(painter, option, index)

        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        opt.text = ""
        widget = opt.widget or self._tree
        style = widget.style()
        painter.save()
        style.drawControl(QStyle.CE_ItemViewItem, opt, painter, widget)
        painter.restore()

        if item.data(0, BOM_TREE_FOLDER_ROLE):
            return

        raw = (item.text(BOM_COL_STATUS) or "").strip()
        key = _status_badge_key(raw)
        if not key:
            painter.save()
            painter.setPen(opt.palette.color(QPalette.Text))
            f = QFont(opt.font)
            f.setPixelSize(11)
            painter.setFont(f)
            painter.drawText(opt.rect.adjusted(6, 0, -6, 0), Qt.AlignVCenter | Qt.AlignLeft, raw)
            painter.restore()
            return

        st = _STATUS_BADGE_STYLES[key]
        bg = QColor(st["bg"])
        fg = QColor(st["fg"])
        f = QFont(opt.font)
        f.setPixelSize(11)
        painter.setFont(f)
        fm = QFontMetrics(f)
        label = raw or key.title()
        pad_h, pad_v = 10, 2
        w = fm.horizontalAdvance(label) + pad_h * 2
        h = fm.height() + pad_v * 2
        pill = QRect(
            opt.rect.left() + 6,
            opt.rect.center().y() - h // 2,
            min(w, opt.rect.width() - 12),
            h,
        )
        painter.save()
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setPen(Qt.NoPen)
        painter.setBrush(bg)
        rr = min(_STATUS_BADGE_CORNER_RADIUS, max(2, pill.height() // 2))
        painter.drawRoundedRect(pill, rr, rr)
        painter.setPen(fg)
        painter.setFont(f)
        painter.drawText(pill, Qt.AlignCenter, label)
        painter.restore()

    def helpEvent(self, event, view, option, index):
        if event.type() != QEvent.ToolTip or index.column() != BOM_COL_STATUS:
            return super().helpEvent(event, view, option, index)
        item = self._tree.itemFromIndex(index)
        if item and (item.text(BOM_COL_STATUS) or "").strip():
            QToolTip.showText(event.globalPos(), item.text(BOM_COL_STATUS), view)
            return True
        return super().helpEvent(event, view, option, index)


class _BomTreeIntegrityDelegate(QStyledItemDelegate):
    """Integrity column: checkmark or warning."""

    def __init__(self, tree: QTreeWidget, parent=None):
        super().__init__(parent)
        self._tree = tree

    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index):
        if index.column() != BOM_COL_INTEGRITY:
            return super().paint(painter, option, index)
        item = self._tree.itemFromIndex(index)
        if item is None:
            return super().paint(painter, option, index)

        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        opt.text = ""
        widget = opt.widget or self._tree
        style = widget.style()
        painter.save()
        style.drawControl(QStyle.CE_ItemViewItem, opt, painter, widget)
        painter.restore()

        if item.data(0, BOM_TREE_FOLDER_ROLE):
            return

        if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
            return

        payload = item.data(BOM_COL_INTEGRITY, BOM_TREE_INTEGRITY_ROLE) or {"state": "ok"}
        state = str(payload.get("state") or "ok")
        sym = "✓" if state == "ok" else "⚠"
        col = QColor("#639922") if state == "ok" else QColor("#BA7517")
        f = QFont(opt.font)
        f.setPixelSize(14)
        painter.save()
        painter.setFont(f)
        painter.setPen(col)
        painter.drawText(opt.rect, Qt.AlignCenter, sym)
        painter.restore()

    def helpEvent(self, event, view, option, index):
        if event.type() != QEvent.ToolTip or index.column() != BOM_COL_INTEGRITY:
            return super().helpEvent(event, view, option, index)
        item = self._tree.itemFromIndex(index)
        if not item:
            return super().helpEvent(event, view, option, index)
        if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
            return super().helpEvent(event, view, option, index)
        payload = item.data(BOM_COL_INTEGRITY, BOM_TREE_INTEGRITY_ROLE) or {}
        tip = str(payload.get("tooltip") or "")
        if tip:
            QToolTip.showText(event.globalPos(), tip, view)
            return True
        return super().helpEvent(event, view, option, index)


# ═══════════════════════════════════════════════════════════════════════════
#  EVENT STYLING CONFIG
# ═══════════════════════════════════════════════════════════════════════════
_EVENT_STYLES = {
    "COMMIT":               {"icon": "📦", "color": "#0078d7", "bg": "#e6f2ff", "label": "Commit"},
    "CHECKIN":              {"icon": "🔓", "color": "#16a34a", "bg": "#dcfce7", "label": "Check In"},
    "CHECKOUT":             {"icon": "🔒", "color": "#ea580c", "bg": "#ffedd5", "label": "Check Out"},
    "UNDO_CHECKOUT":        {"icon": "U", "color": "#64748b", "bg": "#f1f5f9", "label": "Undo Checkout"},
    "OBJECT_ITERATION":     {"icon": "I", "color": "#0369a1", "bg": "#e0f2fe", "label": "Iteration"},
    "REVISION_CREATED":     {"icon": "R", "color": "#0f766e", "bg": "#ccfbf1", "label": "New Revision"},
    "PART_RELEASED":        {"icon": "🚀", "color": "#7c3aed", "bg": "#ede9fe", "label": "Released"},
    "ATTACHMENT_VERSION":   {"icon": "📎", "color": "#0891b2", "bg": "#e0f2fe", "label": "Attachment"},
    "ATTACHMENT_RELEASED":  {"icon": "✅", "color": "#059669", "bg": "#d1fae5", "label": "Attach Released"},
}
_DEFAULT_STYLE = {"icon": "📋", "color": "#6b7280", "bg": "#f3f4f6", "label": "Event"}


def _style_for(event_type: str) -> dict:
    return _EVENT_STYLES.get((event_type or "").upper().strip(), _DEFAULT_STYLE)


# ═══════════════════════════════════════════════════════════════════════════
#  KPI CARD
# ═══════════════════════════════════════════════════════════════════════════
class _KpiCard(QFrame):
    """Small metrics card with value, label, and accent color strip."""

    def __init__(self, value: str = "0", label: str = "", accent: str = "#0078d7",
                 icon_text: str = "", parent=None):
        super().__init__(parent)
        self.setFrameShape(QFrame.StyledPanel)
        self.setFixedHeight(60)
        self.setMinimumWidth(110)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setStyleSheet(f"""
            _KpiCard {{
                background: #ffffff;
                border: 1px solid #e5e7eb;
                border-left: 4px solid {accent};
                border-radius: 8px;
                padding: 0px;
            }}
        """)

        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(12)
        shadow.setOffset(0, 2)
        shadow.setColor(QColor(0, 0, 0, 30))
        self.setGraphicsEffect(shadow)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(12, 8, 12, 8)
        lay.setSpacing(2)

        # Top row: icon + value
        top = QHBoxLayout()
        top.setSpacing(6)
        if icon_text:
            ic = QLabel(icon_text)
            ic.setStyleSheet("font-size: 12px; background: transparent; border: none;")
            top.addWidget(ic)
        self._value_lbl = QLabel(str(value))
        self._value_lbl.setStyleSheet(
            f"font-size: 15px; font-weight: bold; color: {accent}; background: transparent; border: none;"
        )
        top.addWidget(self._value_lbl)
        top.addStretch()
        lay.addLayout(top)

        self._label_lbl = QLabel(str(label))
        self._label_lbl.setStyleSheet(
            "font-size: 8px; color: #6b7280; font-weight: 500; background: transparent; border: none;"
        )
        lay.addWidget(self._label_lbl)

    def set_value(self, val: str):
        self._value_lbl.setText(str(val))

    def set_label(self, lbl: str):
        self._label_lbl.setText(str(lbl))


# ═══════════════════════════════════════════════════════════════════════════
#  ACTIVITY SPARKLINE  (mini bar chart in last 30 days)
# ═══════════════════════════════════════════════════════════════════════════
class _ActivitySparkline(QWidget):
    """Tiny 30-day activity bar chart."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(48)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self._bars: list[int] = []
        self._accent = QColor("#0078d7")

    def set_data(self, events: list[dict]):
        """Build 30-day histogram from event timestamps."""
        today = datetime.now().date()
        counts = [0] * 30
        for ev in (events or []):
            ts = ev.get("timestamp", "")
            try:
                dt = datetime.fromisoformat(str(ts).replace("Z", "+00:00")).date()
            except Exception:
                try:
                    dt = datetime.strptime(str(ts)[:19], "%Y-%m-%d %H:%M:%S").date()
                except Exception:
                    continue
            delta = (today - dt).days
            if 0 <= delta < 30:
                counts[29 - delta] += 1
        self._bars = counts
        self.update()

    def paintEvent(self, event):
        if not self._bars:
            return
        from PyQt5.QtGui import QPainter, QColor, QPen
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        w = self.width()
        h = self.height()

        mx = max(self._bars) if self._bars else 1
        mx = max(mx, 1)
        bar_w = max(2, (w - 4) / len(self._bars) - 1)
        gap = 1

        for i, v in enumerate(self._bars):
            x = int(2 + i * (bar_w + gap))
            bar_h = max(1, int((v / mx) * (h - 8)))
            y = h - 4 - bar_h

            if v == 0:
                color = QColor("#e5e7eb")
            else:
                intensity = min(255, 100 + int(155 * (v / mx)))
                color = QColor(0, 120, 215, intensity)
            p.fillRect(int(x), int(y), int(bar_w), int(bar_h), color)
        p.end()


# ═══════════════════════════════════════════════════════════════════════════
#  EVENT-TYPE FILTER CHIP
# ═══════════════════════════════════════════════════════════════════════════
class _FilterChip(QPushButton):
    """Togglable chip for filtering by event type."""
    toggled_custom = pyqtSignal()

    def __init__(self, label: str, icon_text: str, color: str, bg: str, parent=None):
        super().__init__(f"{icon_text} {label}", parent)
        self._active = True
        self._color = color
        self._bg = bg
        self._label = label
        self.setCheckable(True)
        self.setChecked(True)
        self.setCursor(Qt.PointingHandCursor)
        self.setFixedHeight(28)
        self.clicked.connect(lambda: self.toggled_custom.emit())
        self._refresh_style()

    def _refresh_style(self):
        if self.isChecked():
            self.setStyleSheet(f"""
                QPushButton {{
                    background: {self._bg}; color: {self._color};
                    border: 1px solid {self._color}; border-radius: 12px;
                    padding: 2px 10px; font-size: 11px; font-weight: 600;
                }}
                QPushButton:hover {{ background: {self._color}; color: #ffffff; }}
            """)
        else:
            self.setStyleSheet(f"""
                QPushButton {{
                    background: #f3f4f6; color: #475569;
                    border: 1px solid #d1d5db; border-radius: 12px;
                    padding: 2px 10px; font-size: 11px; font-weight: 500;
                }}
                QPushButton:hover {{ background: #e5e7eb; color: #111827; }}
            """)

    def toggle_state(self):
        self.setChecked(not self.isChecked())
        self._refresh_style()

    def is_active(self) -> bool:
        return self.isChecked()

    def mouseReleaseEvent(self, event):
        super().mouseReleaseEvent(event)
        self._refresh_style()


# ═══════════════════════════════════════════════════════════════════════════
#  TIMELINE TABLE DELEGATE  (colored badges in the Event column)
# ═══════════════════════════════════════════════════════════════════════════
class _TimelineTable(QTableWidget):
    """Enhanced table with custom rendering for event badges and relative timestamps."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setColumnCount(7)
        self.setHorizontalHeaderLabels(
            ["", "Timestamp", "Event", "Rev / Iter", "User", "Project", "Details"]
        )
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.setWordWrap(True)
        self.setTextElideMode(Qt.ElideRight)
        self.setAlternatingRowColors(True)
        self.verticalHeader().setVisible(False)
        self.setShowGrid(False)
        hh = self.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Fixed)     # timeline dot
        hh.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # timestamp
        hh.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # event badge
        hh.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # object revision / iteration
        hh.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # user
        hh.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # project
        hh.setSectionResizeMode(6, QHeaderView.Stretch)           # details
        self.setColumnWidth(0, 30)
        self.setStyleSheet("""
            QTableWidget {
                border: 1px solid #e5e7eb;
                border-radius: 8px;
                background: #ffffff;
                alternate-background-color: #fafbfc;
                gridline-color: transparent;
            }
            QTableWidget::item {
                padding: 6px 8px;
                border-bottom: 1px solid #f3f4f6;
            }
            QTableWidget::item:selected {
                background-color: #eff6ff;
                color: #111827;
            }
            QHeaderView::section {
                background: #f8f9fa;
                border: none;
                border-bottom: 2px solid #e5e7eb;
                padding: 8px 6px;
                font-weight: 700;
                font-size: 8pt;
                color: #4b5563;
                text-transform: uppercase;
            }
        """)

    def populate(self, rows: list[dict]):
        """Fill the table with event rows, adding visual timeline and badges."""
        self.setRowCount(0)
        self.setRowCount(len(rows))
        for i, ev in enumerate(rows):
            style = _style_for(ev.get("event", ""))
            ev_data = dict(ev)

            # Col 0: timeline dot
            dot = QLabel(f'<span style="color:{style["color"]}; font-size:18px;">●</span>')
            dot.setAlignment(Qt.AlignCenter)
            dot.setStyleSheet("background: transparent; border: none;")
            self.setCellWidget(i, 0, dot)

            # Col 1: timestamp + relative time
            ts_raw = str(ev.get("timestamp", "") or "")
            rel = _relative_time(ts_raw)
            ts_widget = QWidget()
            ts_lay = QVBoxLayout(ts_widget)
            ts_lay.setContentsMargins(4, 2, 4, 2)
            ts_lay.setSpacing(0)
            ts_main = QLabel(ts_raw[:19] if len(ts_raw) >= 19 else ts_raw)
            ts_main.setStyleSheet("font-size: 8pt; color: #374151; font-weight: 500; background: transparent; border: none;")
            ts_lay.addWidget(ts_main)
            if rel:
                ts_rel = QLabel(rel)
                ts_rel.setStyleSheet("font-size: 9px; color: #9ca3af; background: transparent; border: none;")
                ts_lay.addWidget(ts_rel)
            self.setCellWidget(i, 1, ts_widget)

            # Store the raw event data on a hidden item for retrieval
            hidden_item = QTableWidgetItem("")
            hidden_item.setData(Qt.UserRole, ev_data)
            self.setItem(i, 0, hidden_item)

            # Col 2: event badge
            badge = QLabel(f'  {style["icon"]} {style["label"]}  ')
            badge.setAlignment(Qt.AlignCenter)
            badge.setStyleSheet(f"""
                background: {style['bg']}; color: {style['color']};
                border: 1px solid {style['color']}40;
                border-radius: 10px; padding: 2px 8px;
                font-size: 8pt; font-weight: 600;
            """)
            badge_container = QWidget()
            badge_lay = QHBoxLayout(badge_container)
            badge_lay.setContentsMargins(2, 2, 2, 2)
            badge_lay.addWidget(badge)
            badge_lay.addStretch()
            self.setCellWidget(i, 2, badge_container)

            # Col 3: BOM object revision / iteration
            object_version = str(ev.get("object_version", "") or "")
            object_version_item = QTableWidgetItem(object_version)
            object_version_item.setTextAlignment(Qt.AlignCenter)
            object_version_item.setForeground(QBrush(QColor("#1d4ed8")))
            object_version_font = object_version_item.font()
            object_version_font.setBold(True)
            object_version_item.setFont(object_version_font)
            self.setItem(i, 3, object_version_item)

            # Col 4: user
            user_str = str(ev.get("user", "") or "")
            user_item = QTableWidgetItem(user_str)
            user_item.setForeground(QBrush(QColor("#374151")))
            fnt = user_item.font()
            fnt.setBold(True)
            user_item.setFont(fnt)
            self.setItem(i, 4, user_item)

            # Col 5: project / version
            proj = str(ev.get("project", "") or "")
            ver = str(ev.get("version", "") or "")
            proj_str = f"{proj} ({ver})" if proj and ver else proj or ver
            proj_item = QTableWidgetItem(proj_str)
            proj_item.setForeground(QBrush(QColor("#6b7280")))
            self.setItem(i, 5, proj_item)

            # Col 6: details
            details = html_to_plain_text(str(ev.get("details", "") or ""))
            det_item = QTableWidgetItem(details)
            det_item.setToolTip(details)
            det_item.setForeground(QBrush(QColor("#4b5563")))
            self.setItem(i, 6, det_item)

            # STEP indicator icon in details if applicable
            step_status = str(ev.get("step_diff_status", "") or "").strip()
            cad_type = str(ev.get("cad_type", "") or "").strip()
            if step_status:
                step_suffix = {"BASELINE": " 🟢 STEP Baseline", "COMPARED": " 🔵 STEP Compared"}.get(
                    step_status.upper(), f" ⚪ STEP:{step_status}"
                )
                det_item.setText(details + step_suffix + f" [{cad_type}]" if cad_type else details + step_suffix)
                det_item.setToolTip(details + step_suffix + f" [{cad_type}]" if cad_type else details + step_suffix)
            else:
                    det_item.setText(details + (f" [{cad_type}]" if cad_type else details))
                    det_item.setToolTip(details + (f" [{cad_type}]" if cad_type else details))    

            self.setRowHeight(i, 48)

        try:
            self.resizeRowsToContents()
            # Enforce a minimum row height
            for r in range(self.rowCount()):
                if self.rowHeight(r) < 44:
                    self.setRowHeight(r, 44)
        except Exception:
            pass

    def get_event_data(self, row: int) -> dict:
        """Retrieve the full event dict for a given row."""
        item = self.item(row, 0)
        if item is None:
            return {}
        return item.data(Qt.UserRole) or {}


def _relative_time(ts: str) -> str:
    """Return a human-readable relative time string."""
    if not ts:
        return ""
    try:
        dt = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
    except Exception:
        try:
            dt = datetime.strptime(str(ts)[:19], "%Y-%m-%d %H:%M:%S")
        except Exception:
            return ""
    try:
        delta = datetime.now() - dt.replace(tzinfo=None)
    except Exception:
        delta = datetime.now() - dt
    secs = int(delta.total_seconds())
    if secs < 0:
        return "in the future"
    if secs < 60:
        return "just now"
    if secs < 3600:
        m = secs // 60
        return f"{m}m ago"
    if secs < 86400:
        h = secs // 3600
        return f"{h}h ago"
    days = secs // 86400
    if days == 1:
        return "yesterday"
    if days < 7:
        return f"{days}d ago"
    if days < 30:
        w = days // 7
        return f"{w}w ago"
    if days < 365:
        mo = days // 30
        return f"{mo}mo ago"
    y = days // 365
    return f"{y}y ago"


# ═══════════════════════════════════════════════════════════════════════════
#  USER BREAKDOWN WIDGET
# ═══════════════════════════════════════════════════════════════════════════
class _UserBreakdownBar(QWidget):
    """Horizontal stacked bar showing contribution per user."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(32)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self._segments: list[tuple[str, int, QColor]] = []  # (username, count, color)
        self._total = 0

    _USER_PALETTE = [
        "#0078d7", "#16a34a", "#ea580c", "#7c3aed", "#0891b2",
        "#db2777", "#ca8a04", "#4f46e5", "#0d9488", "#e11d48",
    ]

    def set_data(self, events: list[dict]):
        counter = Counter()
        for ev in (events or []):
            u = str(ev.get("user", "") or "").strip()
            if u:
                counter[u] += 1
        self._segments = []
        for idx, (user, cnt) in enumerate(counter.most_common(10)):
            color = QColor(self._USER_PALETTE[idx % len(self._USER_PALETTE)])
            self._segments.append((user, cnt, color))
        self._total = sum(c for _, c, _ in self._segments) or 1
        self.setToolTip("\n".join(f"{u}: {c} events" for u, c, _ in self._segments))
        self.update()

    def paintEvent(self, event):
        if not self._segments:
            return
        from PyQt5.QtGui import QPainter, QColor
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        w = self.width() - 4
        h = self.height() - 4
        x = 2
        for user, cnt, color in self._segments:
            seg_w = max(4, int((cnt / self._total) * w))
            p.setBrush(QBrush(color))
            p.setPen(Qt.NoPen)
            p.drawRoundedRect(int(x), 2, int(seg_w), int(h), 4, 4)
            if seg_w > 40:
                p.setPen(QPen(QColor("#ffffff")))
                fnt = p.font()
                fnt.setPointSize(8)
                fnt.setBold(True)
                p.setFont(fnt)
                p.drawText(int(x) + 4, 2, int(seg_w) - 8, int(h),
                           Qt.AlignLeft | Qt.AlignVCenter, user[:12])
            x += seg_w + 1
        p.end()


# ═══════════════════════════════════════════════════════════════════════════
#  HISTORY PANEL — THE GENIUS COMPOSITE WIDGET
# ═══════════════════════════════════════════════════════════════════════════
class HistoryPanel(QWidget):
    """Professional history / audit trail panel with KPIs, timeline,
    activity sparkline, user breakdown, and advanced filtering."""

    # Signals for BomPage to connect
    open_details_requested = pyqtSignal(dict)     # full event dict
    open_step_requested = pyqtSignal(dict)        # for STEP viewer
    show_diff_requested = pyqtSignal(dict)        # for STEP diff

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFont(QFont("Segoe UI", 8))
        self._all_rows: list[dict] = []
        self._filtered_rows: list[dict] = []
        self._analytics: dict = {}
        self._build_ui()

    # ── UI Construction ──────────────────────────────────────────────────
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(8)

        # ── KPI cards row ─────────────────────────────────────────────
        self._kpi_area = QWidget()
        kpi_lay = QHBoxLayout(self._kpi_area)
        kpi_lay.setContentsMargins(0, 0, 0, 0)
        kpi_lay.setSpacing(8)

        self._kpi_events = _KpiCard("0", "Total Events", "#374151", "📊")
        self._kpi_commits = _KpiCard("0", "Commits", "#0078d7", "📦")
        self._kpi_checkins = _KpiCard("0", "Check Ins", "#16a34a", "🔓")
        self._kpi_checkouts = _KpiCard("0", "Check Outs", "#ea580c", "🔒")
        self._kpi_releases = _KpiCard("0", "Releases", "#7c3aed", "🚀")
        self._kpi_users = _KpiCard("0", "Contributors", "#0891b2", "👥")

        for card in (self._kpi_events, self._kpi_commits, self._kpi_checkins,
                     self._kpi_checkouts, self._kpi_releases, self._kpi_users):
            kpi_lay.addWidget(card)
        root.addWidget(self._kpi_area)

        # ── Activity sparkline + user breakdown ───────────────────────
        spark_row = QHBoxLayout()
        spark_row.setSpacing(12)

        spark_group = QVBoxLayout()
        spark_lbl = QLabel("📈 Last 30 Days Activity")
        spark_lbl.setStyleSheet("font-size: 10px; color: #6b7280; font-weight: 600; border: none;")
        self._sparkline = _ActivitySparkline()
        spark_group.addWidget(spark_lbl)
        spark_group.addWidget(self._sparkline)
        spark_row.addLayout(spark_group, 3)

        user_group = QVBoxLayout()
        user_lbl = QLabel("👥 Contributors")
        user_lbl.setStyleSheet("font-size: 10px; color: #6b7280; font-weight: 600; border: none;")
        self._user_bar = _UserBreakdownBar()
        user_group.addWidget(user_lbl)
        user_group.addWidget(self._user_bar)
        spark_row.addLayout(user_group, 2)
        root.addLayout(spark_row)

        # ── Filter bar ────────────────────────────────────────────────
        filter_frame = QFrame()
        filter_frame.setStyleSheet("""
            QFrame {
                background: #f8f9fa;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
                padding: 4px;
            }
        """)
        filter_lay = QVBoxLayout(filter_frame)
        filter_lay.setContentsMargins(8, 6, 8, 6)
        filter_lay.setSpacing(6)

        # Text search + export row
        search_row = QHBoxLayout()
        search_row.setSpacing(8)
        self._search = QLineEdit()
        self._search.setPlaceholderText("🔍  Search history (event, user, details, project)…")
        self._search.setStyleSheet("""
            QLineEdit {
                border: 1px solid #d1d5db; border-radius: 6px;
                padding: 5px 10px; font-size: 8pt;
                background: #ffffff;
            }
            QLineEdit:focus { border-color: #0078d7; }
        """)
        self._search.textChanged.connect(self._apply_filters)
        search_row.addWidget(self._search, 1)

        self._export_btn = QPushButton("📥 Export CSV")
        self._export_btn.setObjectName("neutral")
        self._export_btn.setCursor(Qt.PointingHandCursor)
        self._export_btn.setFixedHeight(30)
        search_row.addWidget(self._export_btn)

        self._sort_btn = QPushButton("↕ Sort")
        self._sort_btn.setObjectName("neutral")
        self._sort_btn.setCursor(Qt.PointingHandCursor)
        self._sort_btn.setFixedHeight(30)
        self._sort_btn.setCheckable(True)
        self._sort_btn.setToolTip("Toggle: newest first ↔ oldest first")
        self._sort_btn.clicked.connect(self._toggle_sort)
        search_row.addWidget(self._sort_btn)

        filter_lay.addLayout(search_row)

        # Event-type chips row
        chips_row = QHBoxLayout()
        chips_row.setSpacing(4)
        chips_lbl = QLabel("Filter:")
        chips_lbl.setStyleSheet("font-size: 10px; color: #6b7280; font-weight: 600; border: none; background: transparent;")
        chips_row.addWidget(chips_lbl)

        self._chips: dict[str, _FilterChip] = {}
        for key, sty in _EVENT_STYLES.items():
            chip = _FilterChip(sty["label"], sty["icon"], sty["color"], sty["bg"])
            chip.toggled_custom.connect(self._apply_filters)
            self._chips[key] = chip
            chips_row.addWidget(chip)

        # Select-all / none
        sel_all = QPushButton("All")
        sel_all.setFixedSize(36, 24)
        sel_all.setStyleSheet(
            "font-size: 10px; font-weight: 700; color: #111827; "
            "border-radius: 8px; background: #e5e7eb; border: none;"
        )
        sel_all.setCursor(Qt.PointingHandCursor)
        sel_all.clicked.connect(lambda: self._set_all_chips(True))
        chips_row.addWidget(sel_all)

        sel_none = QPushButton("None")
        sel_none.setFixedSize(40, 24)
        sel_none.setStyleSheet(
            "font-size: 10px; font-weight: 700; color: #111827; "
            "border-radius: 8px; background: #e5e7eb; border: none;"
        )
        sel_none.setCursor(Qt.PointingHandCursor)
        sel_none.clicked.connect(lambda: self._set_all_chips(False))
        chips_row.addWidget(sel_none)

        chips_row.addStretch()

        # Result count label
        self._count_lbl = QLabel("")
        self._count_lbl.setStyleSheet("font-size: 10px; color: #9ca3af; border: none; background: transparent;")
        chips_row.addWidget(self._count_lbl)

        filter_lay.addLayout(chips_row)
        root.addWidget(filter_frame)

        # ── Timeline table ────────────────────────────────────────────
        self._table = _TimelineTable()
        self._table.itemDoubleClicked.connect(self._on_double_click)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_context_menu)
        root.addWidget(self._table, 1)

        # ── Bottom status ─────────────────────────────────────────────
        self._status_lbl = QLabel("Select a part to view history")
        self._status_lbl.setStyleSheet("font-size: 10px; color: #9ca3af; padding: 2px 4px; border: none;")
        root.addWidget(self._status_lbl)

        self._sort_ascending = False   # default newest-first

    # ── Public API ───────────────────────────────────────────────────────
    def set_data(self, rows: list[dict], analytics: dict):
        """Load history data and refresh everything."""
        self._all_rows = list(rows or [])
        self._analytics = dict(analytics or {})
        self._search.setText("")
        self._sort_ascending = False
        self._sort_btn.setChecked(False)

        # KPI cards
        self._kpi_events.set_value(str(analytics.get("events_total", 0)))
        self._kpi_commits.set_value(str(analytics.get("commits", 0)))
        self._kpi_checkins.set_value(str(analytics.get("checkins", 0)))
        self._kpi_checkouts.set_value(str(analytics.get("checkouts", 0)))
        releases = analytics.get("part_releases", 0) + analytics.get("attachment_releases", 0)
        self._kpi_releases.set_value(str(releases))
        self._kpi_users.set_value(str(analytics.get("unique_users", 0)))

        # Activity & user charts
        self._sparkline.set_data(rows)
        self._user_bar.set_data(rows)

        # Last activity
        last = str(analytics.get("last_activity", "") or "")
        rel = _relative_time(last)
        if last:
            self._status_lbl.setText(f"Last activity: {last[:19]}  ({rel})")
        else:
            self._status_lbl.setText("No activity recorded")

        self._apply_filters()

    def clear(self):
        """Reset to empty state."""
        self._all_rows = []
        self._filtered_rows = []
        self._analytics = {}
        self._table.setRowCount(0)
        self._search.setText("")
        self._kpi_events.set_value("0")
        self._kpi_commits.set_value("0")
        self._kpi_checkins.set_value("0")
        self._kpi_checkouts.set_value("0")
        self._kpi_releases.set_value("0")
        self._kpi_users.set_value("0")
        self._sparkline._bars = []
        self._sparkline.update()
        self._user_bar._segments = []
        self._user_bar.update()
        self._count_lbl.setText("")
        self._status_lbl.setText("Select a part to view history")

    def get_all_rows(self) -> list[dict]:
        return list(self._all_rows)

    def get_filtered_rows(self) -> list[dict]:
        return list(self._filtered_rows)

    # ── Filtering ────────────────────────────────────────────────────────
    def _apply_filters(self):
        q = (self._search.text() or "").strip().lower()

        # Active event-type chips
        active_types = set()
        for key, chip in self._chips.items():
            if chip.is_active():
                active_types.add(key)

        filtered = []
        for ev in self._all_rows:
            et = (ev.get("event", "") or "").upper().strip()
            if et not in active_types:
                continue
            if q:
                blob = " ".join(str(ev.get(k, "")) for k in
                                (
                                    "timestamp", "event", "object_version", "user",
                                    "details", "project", "version",
                                )).lower()
                if q not in blob:
                    continue
            filtered.append(ev)

        # Sort
        if self._sort_ascending:
            filtered = list(reversed(filtered))

        self._filtered_rows = filtered
        self._table.populate(filtered)
        total = len(self._all_rows)
        shown = len(filtered)
        if total == shown:
            self._count_lbl.setText(f"{total} events")
        else:
            self._count_lbl.setText(f"{shown} / {total} events")

    def _set_all_chips(self, on: bool):
        for chip in self._chips.values():
            chip.setChecked(on)
            chip._refresh_style()
        self._apply_filters()

    def _toggle_sort(self):
        self._sort_ascending = self._sort_btn.isChecked()
        self._sort_btn.setText("↑ Oldest" if self._sort_ascending else "↕ Sort")
        self._apply_filters()

    # ── Interactions ─────────────────────────────────────────────────────
    def _on_double_click(self, item):
        row = item.row()
        ev_data = self._table.get_event_data(row)
        if ev_data:
            self.open_details_requested.emit(ev_data)

    def _on_context_menu(self, pos):
        item = self._table.itemAt(pos)
        if not item:
            return
        row = item.row()
        ev_data = self._table.get_event_data(row)
        if not ev_data:
            return

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background: #ffffff; border: 1px solid #d1d5db;
                border-radius: 8px; padding: 4px 0;
            }
            QMenu::item {
                padding: 6px 20px 6px 12px; font-size: 12px;
            }
            QMenu::item:selected {
                background: #eff6ff; color: #0078d7;
            }
            QMenu::separator {
                height: 1px; background: #e5e7eb; margin: 4px 8px;
            }
        """)

        style = _style_for(ev_data.get("event", ""))
        et = (ev_data.get("event") or "").upper()

        details_act = menu.addAction(f"📋  View Full Details")
        details_act.triggered.connect(lambda: self.open_details_requested.emit(ev_data))

        # Copy details to clipboard
        copy_act = menu.addAction("📝  Copy Details to Clipboard")
        copy_act.triggered.connect(lambda: self._copy_to_clipboard(ev_data))

        menu.addSeparator()

        step_path = str(ev_data.get("step_file_path") or "").strip()
        if step_path:
            step_act = menu.addAction("🔬  Open STEP in 3D Viewer")
            step_act.triggered.connect(lambda: self.open_step_requested.emit(ev_data))

        step_status = str(ev_data.get("step_diff_status") or "").strip().upper()
        if step_status == "COMPARED":
            diff_act = menu.addAction("🔍  Show STEP Diff Zones")
            diff_act.triggered.connect(lambda: self.show_diff_requested.emit(ev_data))

        if et == "COMMIT":
            menu.addSeparator()
            commit_id = str(ev_data.get("commit_id") or "")
            if commit_id:
                cid_act = menu.addAction(f"🏷  Copy Commit ID: {commit_id[:12]}…")
                _cid_val = commit_id  # capture in closure
                def _copy_cid(_checked, _v=_cid_val):
                    from PyQt5.QtWidgets import QApplication as _QApp
                    _QApp.clipboard().setText(_v)
                cid_act.triggered.connect(_copy_cid)

        menu.exec_(self._table.viewport().mapToGlobal(pos))

    def _copy_to_clipboard(self, ev_data: dict):
        from PyQt5.QtWidgets import QApplication
        lines = []
        for k in ("timestamp", "event", "user", "project", "version", "details",
                   "commit_id", "step_diff_status", "step_diff_summary"):
            v = ev_data.get(k)
            if v:
                lines.append(f"{k}: {v}")
        QApplication.clipboard().setText("\n".join(lines))


class _TreeLoadWorker(QObject):
    finished = pyqtSignal(int, object, object)  # seq, tree_data, missing_map
    failed = pyqtSignal(int, str)

    def __init__(self, seq: int, bom_service, project_id: int, missing_files: list):
        super().__init__()
        self._seq = int(seq)
        self._bom_service = bom_service
        self._project_id = int(project_id)
        self._missing_files = list(missing_files or [])
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    def run(self):
        try:
            if self._cancelled:
                return
            tree_data = self._bom_service.get_bom_lazy_tree(self._project_id) or {}

            missing_map: dict[int, set] = {}
            try:
                for bom_id, issue_type, _filename in self._missing_files:
                    try:
                        missing_map.setdefault(int(bom_id), set()).add(str(issue_type))
                    except Exception:
                        continue
            except Exception:
                missing_map = {}

            if self._cancelled:
                return

            self.finished.emit(self._seq, tree_data, missing_map)
        except Exception as e:
            self.failed.emit(self._seq, str(e))


class _LazyChildrenWorker(QObject):
    finished = pyqtSignal(int, object)
    failed = pyqtSignal(int, str)
    done = pyqtSignal()

    def __init__(self, request_id: int, bom_service, project_id: int, parent_id: int, parent_path: str):
        super().__init__()
        self._request_id = int(request_id)
        self._bom_service = bom_service
        self._project_id = int(project_id)
        self._parent_id = int(parent_id)
        self._parent_path = str(parent_path or "")
        self._cancelled = False

    def cancel(self) -> None:
        self._cancelled = True

    def run(self) -> None:
        try:
            if self._cancelled:
                return
            nodes = self._bom_service.get_bom_lazy_children(
                self._project_id,
                self._parent_id,
                self._parent_path,
            ) or []
            if self._cancelled:
                return
            self.finished.emit(self._request_id, nodes)
        except Exception as exc:
            self.failed.emit(self._request_id, str(exc))
        finally:
            self.done.emit()


class _SearchWorker(QObject):
    finished = pyqtSignal(int, object)  # seq, results
    failed = pyqtSignal(int, str)

    def __init__(self, seq: int, bom_service, query: str):
        super().__init__()
        self._seq = int(seq)
        self._bom_service = bom_service
        self._query = str(query or "")
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    def run(self):
        try:
            if self._cancelled:
                return
            results = self._bom_service.search_parts(self._query) or []
            if self._cancelled:
                return
            self.finished.emit(self._seq, results)
        except Exception as e:
            self.failed.emit(self._seq, str(e))


class _InitialDiagWorker(QObject):
    finished = pyqtSignal(object, object)  # missing_files, working_result
    failed = pyqtSignal(str)

    def __init__(self, working_dir: str):
        super().__init__()
        self._working_dir = str(working_dir or "")
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    def run(self):
        try:
            if self._cancelled:
                return
            diag_service = DiagService()
            missing_files = []
            working_result = None
            if self._working_dir and os.path.isdir(self._working_dir):
                missing_files = diag_service.sync_bom_files(self._working_dir) or []
                if self._cancelled:
                    return
                working_result = diag_service.check_working_directory(self._working_dir)
            if self._cancelled:
                return
            self.finished.emit(missing_files, working_result)
        except Exception as e:
            self.failed.emit(str(e))


class _ExportWorker(QObject):
    progress = pyqtSignal(str, int, int)
    finished = pyqtSignal(object)
    failed = pyqtSignal(str)

    def __init__(self, operation):
        super().__init__()
        self._operation = operation

    def run(self):
        try:
            result = self._operation(
                lambda message, value, maximum:
                    self.progress.emit(str(message or "Exporting..."), int(value or 0), int(maximum or 0))
            )
            self.finished.emit(result)
        except Exception as exc:
            self.failed.emit(str(exc))


class _RelationshipGraphCanvas(QWidget):
    nodeSelected = pyqtSignal(dict)
    nodeContextMenuRequested = pyqtSignal(dict, object)

    def __init__(self, graph: dict, parent=None):
        super().__init__(parent)
        self.graph = graph or {"nodes": [], "edges": []}
        self.edge_filters = {
            "structure": True,
            "association": True,
            "drawing": True,
        }
        self.scale = 1.0
        self.node_rects: dict[str, QRectF] = {}
        self.manual_positions: dict[str, QPointF] = {}
        self.selected_node_id = None
        self._drag_node_id = None
        self._drag_offset = QPointF(0, 0)
        self.setMinimumSize(900, 560)
        self._layout_graph()

    def set_graph_scale(self, scale: float) -> None:
        self.scale = max(0.45, min(1.8, float(scale or 1.0)))
        self._layout_graph()
        self.update()

    def set_edge_filter(self, filter_name: str, enabled: bool) -> None:
        if filter_name not in self.edge_filters:
            return
        self.edge_filters[filter_name] = bool(enabled)
        visible_node_ids = {
            str(node.get("id")) for node in self._visible_nodes()
        }
        if self.selected_node_id and self.selected_node_id not in visible_node_ids:
            self.selected_node_id = None
            self.nodeSelected.emit({})
        self._layout_graph()
        self.update()

    def _edge_category(self, edge: dict, nodes_by_id: dict[str, dict]) -> str:
        source_node = nodes_by_id.get(str(edge.get("source")), {})
        target_node = nodes_by_id.get(str(edge.get("target")), {})
        label = str(edge.get("label") or "").upper()
        kind = str(edge.get("kind") or "").lower()
        if label == "DRW" or target_node.get("type") == "drawing":
            return "drawing"
        if kind == "association":
            return "association"
        return "structure"

    def _visible_edges(self) -> list:
        nodes_by_id = {str(node.get("id")): node for node in self.graph.get("nodes") or []}
        visible = []
        for edge in self.graph.get("edges") or []:
            category = self._edge_category(edge, nodes_by_id)
            if self.edge_filters.get(category, True):
                visible.append(edge)
        return visible

    def _visible_nodes(self) -> list:
        nodes = list(self.graph.get("nodes") or [])
        nodes_by_id = {str(node.get("id")): node for node in nodes}
        visible_edges = self._visible_edges()
        visible_ids = {
            str(edge.get("source")) for edge in visible_edges
        } | {
            str(edge.get("target")) for edge in visible_edges
        }
        for node in nodes:
            node_id = str(node.get("id"))
            if node.get("type") == "focus" or node_id == self.selected_node_id:
                visible_ids.add(node_id)
            if node.get("type") == "drawing" and not self.edge_filters.get("drawing", True):
                visible_ids.discard(node_id)
        return [node for node in nodes if str(node.get("id")) in visible_ids and str(node.get("id")) in nodes_by_id]

    def _layout_graph(self) -> None:
        nodes = self._visible_nodes()
        edges = self._visible_edges()
        nodes_by_id = {str(node.get("id")): node for node in nodes}
        max_width = 0
        max_height = 0
        rects = {}

        cad_nodes = [node for node in nodes if node.get("lane") == "cad"]
        cad_levels = defaultdict(list)
        for node in cad_nodes:
            cad_levels[int(node.get("level") or 0)].append(node)
        for level, level_nodes in cad_levels.items():
            level_nodes.sort(key=lambda row: (
                1 if row.get("type") == "drawing" else 0,
                str(row.get("label") or "").lower(),
            ))
            for row_index, node in enumerate(level_nodes):
                x = 38 + level * 230
                y = 54 + row_index * 72
                rect = QRectF(x, y, 170, 44)
                rects[str(node["id"])] = rect
                max_width = max(max_width, int(rect.right()) + 70)
                max_height = max(max_height, int(rect.bottom()) + 90)

        item_nodes = [node for node in nodes if node.get("lane") == "item"]
        item_children = defaultdict(list)
        item_has_parent = set()
        for edge in edges:
            if str(edge.get("kind") or "") != "structure":
                continue
            source = str(edge.get("source"))
            target = str(edge.get("target"))
            if (
                source in nodes_by_id
                and target in nodes_by_id
                and nodes_by_id[source].get("lane") == "item"
                and nodes_by_id[target].get("lane") == "item"
            ):
                item_children[source].append(target)
                item_has_parent.add(target)
        item_roots = [
            node for node in item_nodes
            if str(node.get("id")) not in item_has_parent
        ]
        item_roots.sort(key=lambda row: (
            0 if row.get("type") == "focus" else 1,
            str(row.get("label") or "").lower(),
        ))
        if not item_roots:
            item_roots = sorted(
                item_nodes, key=lambda row: str(row.get("label") or "").lower()
            )
        cad_bottom = max(
            [rect.bottom() for key, rect in rects.items()
             if nodes_by_id.get(key, {}).get("lane") == "cad"] or [220]
        )
        item_lane_top = max(254, int(cad_bottom) + 42)
        root_spacing = 230
        root_start_x = 270 if len(item_roots) <= 3 else 140
        item_y0 = item_lane_top + 86
        placed = set()

        def place_item_subtree(node_id: str, x: int, y: int, depth: int = 0) -> int:
            node = nodes_by_id.get(node_id)
            if not node or node_id in placed:
                return y
            rect = QRectF(x + depth * 18, y, 170, 44)
            rects[node_id] = rect
            placed.add(node_id)
            bottom = int(rect.bottom()) + 26
            nonlocal max_width, max_height
            max_width = max(max_width, int(rect.right()) + 70)
            max_height = max(max_height, int(rect.bottom()) + 90)
            child_y = y + 76
            for child_id in item_children.get(node_id, [])[:14]:
                child_bottom = place_item_subtree(child_id, x, child_y, depth + 1)
                child_y = max(child_y + 64, child_bottom + 10)
                bottom = max(bottom, child_bottom)
            return bottom

        for root_index, root in enumerate(item_roots):
            root_id = str(root["id"])
            x = root_start_x + root_index * root_spacing
            place_item_subtree(root_id, x, item_y0)
        overflow_index = 0
        for node in item_nodes:
            node_id = str(node["id"])
            if node_id in placed:
                continue
            rect = QRectF(38 + overflow_index * 210, item_y0 + 76, 170, 44)
            rects[node_id] = rect
            overflow_index += 1
            max_width = max(max_width, int(rect.right()) + 70)
            max_height = max(max_height, int(rect.bottom()) + 90)

        for node_id, position in self.manual_positions.items():
            if node_id not in rects:
                continue
            rects[node_id].moveTo(position)
            max_width = max(max_width, int(rects[node_id].right()) + 70)
            max_height = max(max_height, int(rects[node_id].bottom()) + 90)

        item_bottom = max(
            [rect.bottom() for key, rect in rects.items()
             if nodes_by_id.get(key, {}).get("lane") == "item"] or [514]
        )
        self.lane_rects = {
            "cad": QRectF(18, 16, max(840, max_width - 36), max(210, cad_bottom - 16 + 30)),
            "item": QRectF(
                18, item_lane_top, max(840, max_width - 36),
                max(260, item_bottom - item_lane_top + 30),
            ),
        }
        self.node_rects = rects
        self.setMinimumSize(
            int(max(900, max_width) * self.scale),
            int(max(560, max_height) * self.scale),
        )

    def _scaled_rect(self, rect: QRectF) -> QRectF:
        return QRectF(
            rect.x() * self.scale, rect.y() * self.scale,
            rect.width() * self.scale, rect.height() * self.scale,
        )

    def _scaled_point(self, point: QPointF) -> QPointF:
        return QPointF(point.x() * self.scale, point.y() * self.scale)

    def _edge_path_points(
        self, src: QRectF, dst: QRectF,
        source_node: dict, target_node: dict, edge: dict,
    ) -> list[QPointF]:
        source_lane = source_node.get("lane")
        target_lane = target_node.get("lane")
        label = str(edge.get("label") or "").upper()
        is_association = str(edge.get("kind") or "") == "association"
        is_drawing = label == "DRW" or target_node.get("type") == "drawing"

        if source_lane == "cad" and target_lane == "item" and is_association:
            start = QPointF(src.center().x(), src.bottom())
            end = QPointF(dst.center().x(), dst.top())
            mid_y = start.y() + max(28, (end.y() - start.y()) * 0.45)
            return [start, QPointF(start.x(), mid_y), QPointF(end.x(), mid_y), end]

        if source_lane == "item" and target_lane == "cad" and is_association:
            start = QPointF(src.center().x(), src.top())
            end = QPointF(dst.center().x(), dst.bottom())
            mid_y = end.y() + max(28, (start.y() - end.y()) * 0.45)
            return [start, QPointF(start.x(), mid_y), QPointF(end.x(), mid_y), end]

        if source_lane == target_lane == "item":
            start = QPointF(src.center().x(), src.bottom())
            end = QPointF(dst.center().x(), dst.top())
            if abs(start.x() - end.x()) < 12:
                return [start, end]
            mid_y = (start.y() + end.y()) / 2
            return [start, QPointF(start.x(), mid_y), QPointF(end.x(), mid_y), end]

        if is_drawing:
            start = QPointF(src.right(), src.center().y())
            end = QPointF(dst.left(), dst.center().y())
            if dst.left() <= src.right():
                lane_y = min(src.top(), dst.top()) - 24
                return [start, QPointF(start.x(), lane_y), QPointF(end.x(), lane_y), end]
            return [start, end]

        if dst.center().x() >= src.center().x():
            start = QPointF(src.right(), src.center().y())
            end = QPointF(dst.left(), dst.center().y())
        else:
            start = QPointF(src.left(), src.center().y())
            end = QPointF(dst.right(), dst.center().y())
        if abs(start.y() - end.y()) < 18:
            return [start, end]
        mid_x = (start.x() + end.x()) / 2
        return [start, QPointF(mid_x, start.y()), QPointF(mid_x, end.y()), end]

    def _draw_arrow(
        self, painter: QPainter, start: QPointF, end: QPointF,
        color: QColor, dashed: bool = False,
    ) -> None:
        self._draw_polyline_arrow(painter, [start, end], color, dashed)

    def _draw_polyline_arrow(
        self, painter: QPainter, points: list[QPointF],
        color: QColor, dashed: bool = False,
    ) -> None:
        if len(points) < 2:
            return
        pen = QPen(color, max(1, int(1.2 * self.scale)))
        if dashed:
            pen.setStyle(Qt.DashLine)
        painter.setPen(pen)
        scaled_points = [self._scaled_point(point) for point in points]
        painter.drawPolyline(QPolygonF(scaled_points))
        start = scaled_points[-2]
        end = scaled_points[-1]
        dx = end.x() - start.x()
        dy = end.y() - start.y()
        length = max((dx * dx + dy * dy) ** 0.5, 1)
        ux, uy = dx / length, dy / length
        size = 8 * self.scale
        left = QPointF(
            end.x() - ux * size - uy * size * 0.45,
            end.y() - uy * size + ux * size * 0.45,
        )
        right = QPointF(
            end.x() - ux * size + uy * size * 0.45,
            end.y() - uy * size - ux * size * 0.45,
        )
        painter.setBrush(color)
        painter.drawPolygon(QPolygonF([end, left, right]))

    def mousePressEvent(self, event):
        pos = QPointF(event.pos().x() / max(self.scale, 0.1), event.pos().y() / max(self.scale, 0.1))
        nodes_by_id = {str(node["id"]): node for node in self._visible_nodes()}
        for node_id, rect in self.node_rects.items():
            if rect.contains(pos):
                self.selected_node_id = node_id
                node = dict(nodes_by_id.get(node_id, {}))
                self.nodeSelected.emit(node)
                if event.button() == Qt.RightButton:
                    self._drag_node_id = None
                    self.nodeContextMenuRequested.emit(node, event.globalPos())
                    self.update()
                    return
                if event.button() == Qt.LeftButton:
                    self._drag_node_id = node_id
                    self._drag_offset = pos - rect.topLeft()
                    self.setCursor(Qt.ClosedHandCursor)
                self.update()
                return
        self.selected_node_id = None
        self._drag_node_id = None
        self.nodeSelected.emit({})
        self.update()

    def mouseMoveEvent(self, event):
        if not self._drag_node_id or not (event.buttons() & Qt.LeftButton):
            super().mouseMoveEvent(event)
            return
        pos = QPointF(event.pos().x() / max(self.scale, 0.1), event.pos().y() / max(self.scale, 0.1))
        rect = self.node_rects.get(self._drag_node_id)
        if not rect:
            return
        new_top_left = pos - self._drag_offset
        new_top_left.setX(max(24, new_top_left.x()))
        new_top_left.setY(max(34, new_top_left.y()))
        rect.moveTo(new_top_left)
        self.node_rects[self._drag_node_id] = rect
        self.manual_positions[self._drag_node_id] = QPointF(rect.x(), rect.y())
        self.setMinimumSize(
            int(max(self.minimumWidth() / max(self.scale, 0.1), rect.right() + 70) * self.scale),
            int(max(self.minimumHeight() / max(self.scale, 0.1), rect.bottom() + 90) * self.scale),
        )
        self.update()

    def mouseReleaseEvent(self, event):
        if self._drag_node_id:
            self._drag_node_id = None
            self.unsetCursor()
            event.accept()
            return
        super().mouseReleaseEvent(event)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.fillRect(self.rect(), QColor("#f6f7e3"))

        lane_rects = getattr(self, "lane_rects", {}) or {}
        base_width = max(840, self.width() / max(self.scale, 0.1) - 36)
        for lane_rect, title in (
            (lane_rects.get("cad", QRectF(18, 16, base_width, 210)), "CAD Structure"),
            (lane_rects.get("item", QRectF(18, 254, base_width, 260)), "EBOM / Item Structure"),
        ):
            rect = self._scaled_rect(lane_rect)
            painter.setPen(QPen(QColor("#c9bd64"), 1))
            painter.setBrush(QColor("#fffbd5"))
            painter.drawRect(rect)
            painter.setPen(QColor("#1f2933"))
            painter.setFont(QFont("Segoe UI", max(8, int(9 * self.scale))))
            painter.drawText(
                rect.adjusted(0, 6 * self.scale, 0, 0),
                Qt.AlignHCenter | Qt.AlignTop,
                title,
            )

        nodes_by_id = {str(node["id"]): node for node in self._visible_nodes()}
        for edge in self._visible_edges():
            src = self.node_rects.get(str(edge.get("source")))
            dst = self.node_rects.get(str(edge.get("target")))
            if not src or not dst:
                continue
            source_node = nodes_by_id.get(str(edge.get("source")), {})
            target_node = nodes_by_id.get(str(edge.get("target")), {})
            same_lane = source_node.get("lane") == target_node.get("lane")
            path_points = self._edge_path_points(src, dst, source_node, target_node, edge)
            dashed = str(edge.get("kind") or "") == "association"
            self._draw_polyline_arrow(
                painter, path_points,
                QColor("#4b5563") if dashed else QColor("#5f6670"),
                dashed=dashed,
            )
            label = str(edge.get("label") or "")
            if label:
                scaled_points = [self._scaled_point(point) for point in path_points]
                mid_point_index = max(0, len(scaled_points) // 2 - 1)
                mid_start = scaled_points[mid_point_index]
                mid_end = scaled_points[min(mid_point_index + 1, len(scaled_points) - 1)]
                mid = QPointF((mid_start.x() + mid_end.x()) / 2, (mid_start.y() + mid_end.y()) / 2)
                painter.setPen(QColor("#111827"))
                painter.setFont(QFont("Segoe UI", max(7, int(8 * self.scale))))
                painter.drawText(
                    QRectF(mid.x() - 48, mid.y() - 18, 96, 16),
                    Qt.AlignCenter,
                    label,
                )

        for node_id, rect in self.node_rects.items():
            node = nodes_by_id.get(node_id, {})
            scaled = self._scaled_rect(rect)
            node_type = str(node.get("type") or "")
            fill = QColor("#ebe7ff") if node.get("lane") == "cad" else QColor("#eee9ff")
            border = QColor("#9b8cff")
            if node_type == "drawing":
                fill = QColor("#edf2f7")
                border = QColor("#94a3b8")
            elif node_type == "focus":
                fill = QColor("#dff7e8")
                border = QColor("#22a06b")
            if node_id == self.selected_node_id:
                border = QColor("#dc2626")
            painter.setPen(QPen(border, 2.4 if node_id == self.selected_node_id else 1.2))
            painter.setBrush(fill)
            painter.drawRect(scaled)
            painter.setPen(QColor("#0f172a"))
            painter.setFont(QFont("Segoe UI", max(7, int(8 * self.scale))))
            painter.drawText(
                scaled.adjusted(6, 4, -6, -4),
                Qt.AlignCenter | Qt.TextWordWrap,
                str(node.get("label") or ""),
            )


class BomPage(QWidget):
    """Full-featured BOM Management Page (restored features)"""

    RELATIONSHIP_GRAPH_PATH_LIMIT = 6
    RELATIONSHIP_GRAPH_ALL_PATH_LIMIT = 10000

    initial_tree_ready = pyqtSignal()
    issue_requested = pyqtSignal(int)
    create_issue_requested = pyqtSignal(int)

    def __init__(self, bom_service):
        super().__init__()
        self.setFont(QFont("Segoe UI", 8))
        self.bom_service = bom_service
        self.diag_service = DiagService()
        self.project_service = ProjectService()
        self.session = SessionManager()
        self.perm = UIPermissionHelper()
        self.part_file_service = PartFileService()
        self.managed_file_service = ManagedFileService(
            part_file_service=self.part_file_service
        )
        self.part_doc_ack_service = PartDocAckService()
        self.package_export_service = PackageExportService()
        self.baseline_service = BaselineService()
        self.commit_repo = CommitRepository()
        self.issue_service = IssueService()
        self.traceability_service = TraceabilityService()
        self.assembly_configuration_service = AssemblyConfigurationService()
        self.undo_service = UndoService()
        self._issue_summary_cache = {}

        self.working_dir = None
        self.commits_dir = None

        if self.session.project_id:
            self.working_dir = self.get_working_dir()
            if self.working_dir:
                self.commits_dir = self.working_dir + "/commits"
        
        self.current_part_aes = None
        self.current_part_id = None
        self._tree_load_seq = 0
        self._initial_tree_ready_emitted = False
        self._tree_thread = None
        self._tree_worker = None
        self._search_thread = None
        self._search_worker = None
        self._diag_thread = None
        self._diag_worker = None
        self._lazy_level_request_seq = 0
        self._lazy_level_requests = {}
        self._retired_lazy_level_requests = []

        self._tree_build_seq = 0
        self._tree_build_queue = deque()
        self._tree_build_missing_map = {}
        self._tree_build_timer = QTimer(self)
        self._tree_build_timer.setInterval(0)
        self._tree_build_timer.timeout.connect(self._tree_build_step)
        self._indicator_refresh_queue = deque()
        self._indicator_refresh_timer = QTimer(self)
        self._indicator_refresh_timer.setInterval(0)
        self._indicator_refresh_timer.timeout.connect(self._indicator_refresh_step)

        self._search_build_seq = 0
        self._search_build_queue = deque()
        self._search_build_timer = QTimer(self)
        self._search_build_timer.setInterval(0)
        self._search_build_timer.timeout.connect(self._search_build_step)

        self._last_tree_node_count = 0
        self._full_tree_cached = False
        self._cached_tree_data = None
        self._cached_missing_map = None
        self._bom_row_numbers = {}
        self._bom_folder_row_numbers = {}
        self._bom_folder_path_rows = {}
        self._lazy_tree_active = True
        self._lazy_tree_materialized = False
        self._in_search_mode = False     # True while _search_tree (index 2) is visible
        self._advanced_filter_flat_mode = False
        self._ebom_filter_flat_mode = False
        self._bom_advanced_filters = self._default_bom_advanced_filters()
        self._active_saved_filter_id = None
        self._active_saved_filter_name = ""
        self._advanced_filter_dialog = None
        self._bom_mode = "ebom"
        self._pdm_cad_roots = []
        self._pdm_cad_scope_path = []
        self._pdm_ebom_roots = []
        self._pdm_ebom_scope_path = []
        self._ebom_associations_by_item = defaultdict(list)
        self.init_ui()
        self._install_undo_shortcuts()
        self.advanced_filter_btn.setEnabled(True)
        self.saved_filters_btn.setEnabled(True)

        # Pre-render indicator icons (fast + consistent colors)
        self._indicator_icon_cache = {}
        self._icon_pdf_ok = self._make_indicator_icon(pdf_ok=True, step_ok=False)
        self._icon_pdf_bad = self._make_indicator_icon(pdf_ok=False, step_ok=False)
        self._icon_step_ok = self._make_indicator_icon(pdf_ok=False, step_ok=True)
        self._icon_step_bad = self._make_indicator_icon(pdf_ok=False, step_ok=False, step_present=True)
        self._icon_pdf_step_ok = self._make_indicator_icon(pdf_ok=True, step_ok=True)
        self._icon_pdf_bad_step_ok = self._make_indicator_icon(pdf_ok=False, step_ok=True)
        self._icon_pdf_ok_step_bad = self._make_indicator_icon(pdf_ok=True, step_ok=False, step_present=True)
        self._icon_pdf_step_bad = self._make_indicator_icon(pdf_ok=False, step_ok=False, step_present=True)

        self.missing_files = []
        self.missing_ids = set()
        self._doc_indicator_cache = {}

        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(250)
        self._search_timer.timeout.connect(self._perform_search_now)

        # Start initial diagnostics and first tree load without blocking UI thread.
        if self.session.project_id and self.working_dir and os.path.isdir(self.working_dir):
            QTimer.singleShot(0, self._start_initial_diagnostics_and_load)
        else:
            try:
                self.show_alert("No project loaded. Select a project from the toolbar to begin.", "error")
            except Exception:
                pass
            self.load_tree()
            self._load_released_ebom_tree()

        try:
            self.destroyed.connect(lambda *_: self._cancel_background_work())
        except Exception:
            pass

    def _install_undo_shortcuts(self) -> None:
        shortcut = QShortcut(QKeySequence("Ctrl+Z"), self)
        shortcut.setContext(Qt.WidgetWithChildrenShortcut)
        shortcut.activated.connect(self._undo_last_nexus_action)
        self._undo_shortcut = shortcut

    def _undo_last_nexus_action(self) -> None:
        # Let focused editors keep their native text undo.  Nexus-level undo
        # handles committed app actions such as delete/remove.
        focused = QApplication.focusWidget()
        if isinstance(focused, (QLineEdit, QTextEdit, QPlainTextEdit, QSpinBox, QComboBox)):
            try:
                if hasattr(focused, "undo") and focused.isEnabled():
                    focused.undo()
                    return
            except Exception:
                pass
        try:
            record = self.undo_service.undo_last()
        except Exception as exc:
            QMessageBox.information(self, "Undo", str(exc))
            return
        try:
            self._refresh_after_undo(record.label)
        except Exception:
            self.load_tree()
        try:
            self.window().statusBar().showMessage(f"Undone: {record.label}", 5000)
        except Exception:
            pass

    def _refresh_after_undo(self, label: str = "") -> None:
        # Prefer contextual refresh when the tree is already in PDM mode; fall
        # back to the normal tree loader for older/legacy views.
        try:
            if str(getattr(self, "_bom_mode", "")).lower() == "cad":
                self._load_pdm_cad_tree()
                return
            if str(getattr(self, "_bom_mode", "")).lower() == "ebom":
                self._load_pdm_ebom_tree()
                return
        except Exception:
            pass
        self.load_tree()

    def _cancel_background_work(self) -> None:
        self._cancel_lazy_level_requests(wait=True)
        for worker in (getattr(self, "_tree_worker", None), getattr(self, "_search_worker", None), getattr(self, "_diag_worker", None)):
            try:
                if worker is not None:
                    worker.cancel()
            except Exception:
                pass

        for thread in (getattr(self, "_tree_thread", None), getattr(self, "_search_thread", None), getattr(self, "_diag_thread", None)):
            try:
                if thread is not None:
                    thread.quit()
            except Exception:
                pass

        try:
            self._tree_build_timer.stop()
        except Exception:
            pass
        try:
            self._search_build_timer.stop()
        except Exception:
            pass
        try:
            self._indicator_refresh_timer.stop()
            self._indicator_refresh_queue.clear()
        except Exception:
            pass

    def _cancel_lazy_level_requests(self, wait: bool = False) -> None:
        requests = list((getattr(self, "_lazy_level_requests", {}) or {}).values())
        self._lazy_level_requests = {}
        if wait:
            requests.extend(getattr(self, "_retired_lazy_level_requests", []) or [])
            self._retired_lazy_level_requests = []
        for request in requests:
            item = request.get("item")
            try:
                self.tree.setItemLoading(item, False)
            except Exception:
                pass
            try:
                request["worker"].cancel()
            except Exception:
                pass
            try:
                request["thread"].quit()
            except Exception:
                pass
            if wait:
                try:
                    request["thread"].wait(1500)
                except Exception:
                    pass
            elif request not in self._retired_lazy_level_requests:
                self._retired_lazy_level_requests.append(request)
                try:
                    request["thread"].finished.connect(
                        lambda _request=request: self._forget_retired_lazy_request(_request)
                    )
                except Exception:
                    pass
        try:
            self.tree.resetLoadingIndicators()
        except Exception:
            pass

    def _forget_retired_lazy_request(self, request: dict) -> None:
        try:
            self._retired_lazy_level_requests.remove(request)
        except (ValueError, AttributeError):
            pass

    def _cancel_lazy_requests_for_parts(self, part_ids) -> None:
        wanted = {int(value) for value in (part_ids or [])}
        for request_id, request in list((getattr(self, "_lazy_level_requests", {}) or {}).items()):
            if int(request.get("part_id") or -1) not in wanted:
                continue
            self._lazy_level_requests.pop(int(request_id), None)
            try:
                self.tree.setItemLoading(request.get("item"), False)
            except Exception:
                pass
            try:
                request["worker"].cancel()
                request["thread"].quit()
            except Exception:
                pass
            if request not in self._retired_lazy_level_requests:
                self._retired_lazy_level_requests.append(request)
                try:
                    request["thread"].finished.connect(
                        lambda _request=request: self._forget_retired_lazy_request(_request)
                    )
                except Exception:
                    pass

    def _set_tree_loading(self, loading: bool) -> None:
        try:
            stack = getattr(self, "_tree_stack", None)
            if stack is None:
                return
            if loading:
                stack.setCurrentIndex(0)
            elif getattr(self, "_bom_mode", "cad") == "ebom":
                target_tree = (
                    getattr(self, "_ebom_filter_tree", None)
                    if getattr(self, "_ebom_filter_flat_mode", False)
                    else getattr(self, "_ebom_tree", None)
                )
                if target_tree is not None:
                    stack.setCurrentWidget(target_tree)
            else:
                target_tree = getattr(self, "_cad_tree", None)
                if target_tree is not None:
                    stack.setCurrentWidget(target_tree)
        except Exception:
            pass

    def _on_bom_mode_changed(self, _index: int = 0) -> None:
        mode = str(self.bom_mode_selector.currentData() or "cad")
        self._bom_mode = mode
        try:
            self.advanced_filter_btn.setEnabled(mode == "ebom")
            self.saved_filters_btn.setEnabled(mode == "ebom")
            self.bom_columns_btn.setEnabled(mode == "ebom")
            self.search_input.setPlaceholderText(
                "Search Item Number, name, or AES..."
                if mode == "ebom"
                else "Search CAD file, name, description, or related drawing..."
            )
        except Exception:
            pass
        if mode == "ebom":
            self._in_search_mode = False
            self._load_released_ebom_tree()
        else:
            self._load_pdm_cad_tree()
        self.clear_details()
        self._set_engineering_item_actions_enabled(False)
        read_only = False
        try:
            self.add_part_btn.setEnabled(
                mode == "ebom" and self.perm.can("manage_parts")
            )
            self.add_folder_btn.setEnabled(
                self.perm.can("manage_parts")
            )
        except Exception:
            pass
        if getattr(self, "_current_part_details", None):
            self._update_lifecycle_action_states(self._current_part_details)
        self._update_pdm_scope_bar()
        self._sync_visual_action_states()

    def _pdm_cad_checkout_text(self, document: dict) -> str:
        owner = document.get("checked_out_by")
        if owner is None:
            return "Checked in"
        username = str(document.get("checked_out_by_username") or "").strip()
        if not username and self.session.user_id is not None:
            try:
                if int(owner) == int(self.session.user_id):
                    username = "you"
            except Exception:
                pass
        return f"In Work ({username or owner})"

    @staticmethod
    def _windchill_item_label(info: dict) -> str:
        number = str(info.get("part_number") or "").strip()
        name = str(info.get("name") or info.get("id") or "Item").strip()
        version = str(
            info.get("current_version")
            or info.get("version_label")
            or ""
        ).strip()
        if not version:
            revision = str(info.get("revision") or "").strip()
            iteration = str(info.get("iteration_number") or "").strip()
            version = f"{revision}.{iteration}" if revision and iteration else revision
        view = str(info.get("item_view") or info.get("view") or "Design").strip()
        identity = ", ".join(value for value in (number, name) if value)
        if version:
            identity = f"{identity}, {version}" if identity else version
        if view:
            identity = f"{identity} ({view.title()})" if identity else f"({view.title()})"
        return identity or name or "Item"

    def _pdm_creo_file_text(self, document: dict) -> str:
        source_name = str(
            document.get("latest_creo_file_name")
            or document.get("source_file_name")
            or ""
        ).strip()
        try:
            version = document.get("latest_creo_file_version")
            if version is None:
                version = document.get("creo_file_version")
            version = int(version) if version is not None and str(version).strip() else None
        except Exception:
            version = None
        if source_name:
            return source_name
        if version is not None:
            base_name = str(document.get("file_name") or "").strip()
            return f"{base_name}.{version}" if base_name else f"Creo .{version}"
        return ""

    @staticmethod
    def _creo_name_parts(name: str) -> tuple[str, int | None]:
        """Return (base Creo file, numeric file suffix) for PRT/ASM/DRW names."""
        cleaned = os.path.basename(str(name or "").replace("\\", "/")).strip()
        match = re.match(
            r"^(?P<base>.+\.(?:prt|asm|drw))(?:\.(?P<version>\d+))?$",
            cleaned,
            flags=re.IGNORECASE,
        )
        if not match:
            return cleaned, None
        version = match.group("version")
        return match.group("base"), int(version) if version else None

    def _build_pdm_cad_working_file_index(self) -> dict:
        """Index Creo files physically present in the project working directory."""
        root = str(self.get_working_dir() or self.working_dir or "").strip()
        if root:
            self.working_dir = root
        index: dict[str, list[dict]] = {}
        if not root or not safe_exists(root):
            return index
        skip_dirs = {
            ".git", ".nexus", "__pycache__", "commits", "snapshots",
            "baselines", "exports",
        }
        try:
            walker = os.walk(root)
            for folder, dirs, files in walker:
                dirs[:] = [
                    d for d in dirs
                    if d not in skip_dirs and not d.startswith(".nexus")
                ]
                for file_name in files:
                    base, version = self._creo_name_parts(file_name)
                    if not base or not re.search(r"\.(?:prt|asm|drw)$", base, re.IGNORECASE):
                        continue
                    key = base.casefold()
                    index.setdefault(key, []).append({
                        "base": base,
                        "version": version,
                        "name": file_name,
                        "path": os.path.join(folder, file_name),
                    })
        except Exception:
            return index
        for rows in index.values():
            rows.sort(
                key=lambda row: (
                    row.get("version") is None,
                    int(row.get("version") or -1),
                    str(row.get("name") or "").casefold(),
                )
            )
        return index

    def _pdm_cad_iteration_warnings(self, document: dict) -> list[str]:
        """Detect local Creo file/version problems for a CAD Document row."""
        category = str(document.get("category") or "").upper()
        if category not in {"ASSEMBLY", "COMPONENT", "DRAWING"}:
            return []
        index = getattr(self, "_pdm_cad_working_file_index", None)
        if index is None:
            index = self._build_pdm_cad_working_file_index()
            self._pdm_cad_working_file_index = index

        file_name = str(document.get("file_name") or "").strip()
        expected_source = str(
            document.get("latest_creo_file_name")
            or document.get("source_file_name")
            or ""
        ).strip()
        expected_base, source_version = self._creo_name_parts(
            expected_source or file_name
        )
        if not expected_base:
            return ["CAD file name is missing."]

        try:
            expected_version = document.get("latest_creo_file_version")
            if expected_version is None:
                expected_version = document.get("creo_file_version")
            expected_version = (
                int(expected_version)
                if expected_version is not None and str(expected_version).strip()
                else source_version
            )
        except Exception:
            expected_version = source_version

        rows = list(index.get(expected_base.casefold()) or [])
        warnings = []
        if expected_version is None:
            warnings.append(
                "No approved Creo file iteration is recorded for this CAD Document."
            )
        if not rows:
            warnings.append(
                f"{expected_base} does not exist in the project working directory."
            )
        else:
            numbered_versions = sorted({
                int(row["version"])
                for row in rows
                if row.get("version") is not None
            })
            if expected_version is not None and numbered_versions:
                if int(expected_version) not in numbered_versions:
                    warnings.append(
                        f"Approved Creo file {expected_base}.{expected_version} is missing from the working directory."
                    )
                newer = [value for value in numbered_versions if value > int(expected_version)]
                older = [value for value in numbered_versions if value < int(expected_version)]
                if newer:
                    warnings.append(
                        f"Newer local Creo version exists: {expected_base}.{max(newer)}."
                    )
                if older:
                    warnings.append(
                        f"Older local Creo version(s) exist: "
                        + ", ".join(f"{expected_base}.{value}" for value in older[-3:])
                        + (" ..." if len(older) > 3 else "")
                    )
            elif expected_version is not None and not numbered_versions:
                warnings.append(
                    f"{expected_base} exists only without a Creo numeric suffix; expected .{expected_version}."
                )
        for drawing in document.get("related_drawings") or []:
            for drawing_warning in self._pdm_cad_iteration_warnings(drawing):
                drawing_name = str(
                    drawing.get("file_name") or drawing.get("name") or "DRW"
                )
                warnings.append(f"Drawing {drawing_name}: {drawing_warning}")
        return warnings

    def _decorate_pdm_cad_iteration_health(
        self, item: QTreeWidgetItem, document: dict
    ) -> list[str]:
        warnings = self._pdm_cad_iteration_warnings(document)
        base_revision = self._pdm_cad_revision_text(document)
        item.setText(CAD_COL_REV, f"⚠ {base_revision}" if warnings else base_revision)
        tooltip = "\n".join(warnings) if warnings else "CAD file iteration is consistent with the working directory."
        item.setToolTip(CAD_COL_REV, tooltip)
        item.setData(CAD_COL_REV, Qt.UserRole, list(warnings))
        color = QColor("#b45309") if warnings else QColor("#0f172a")
        item.setForeground(CAD_COL_REV, QBrush(color))
        if warnings:
            item.setForeground(CAD_COL_NAME, QBrush(QColor("#b45309")))
        else:
            item.setForeground(CAD_COL_NAME, QBrush(QColor("#0f172a")))
        return warnings

    def _pdm_cad_revision_text(self, document: dict) -> str:
        revision = f"CAD {document.get('revision') or 'A'}.{int(document.get('iteration') or 1)}"
        creo_file = self._pdm_creo_file_text(document)
        return f"{revision} | Creo {creo_file}" if creo_file else revision

    @staticmethod
    def _pdm_document_associations(document: dict) -> list[dict]:
        """Return every Item association while accepting the legacy flat payload."""
        associations = [
            dict(row) for row in (document.get("associations") or []) if row
        ]
        if not associations and document.get("association_id") is not None:
            associations = [{
                "id": document.get("association_id"),
                "association_id": document.get("association_id"),
                "item_id": document.get("item_id"),
                "association_type": document.get("association_type"),
                "item_number": document.get("item_number"),
                "item_name": document.get("item_name"),
                "item_aes_number": document.get("item_aes_number"),
                "item_version_label": document.get("item_version_label"),
                "is_primary_drawing": document.get("is_primary_drawing"),
            }]
        unique = []
        seen = set()
        for association in associations:
            key = (
                association.get("association_id") or association.get("id"),
                association.get("item_id"),
                str(association.get("association_type") or "").upper(),
            )
            if key in seen:
                continue
            seen.add(key)
            unique.append(association)
        return unique

    @classmethod
    def _pdm_association_for_item(
        cls, document: dict, item_id: int
    ) -> dict | None:
        for association in cls._pdm_document_associations(document):
            try:
                if int(association.get("item_id")) == int(item_id):
                    return association
            except Exception:
                continue
        return None

    @staticmethod
    def _pdm_selected_drawings(document: dict) -> list[dict]:
        """Drawings explicitly selected for the Item represented by this row."""
        return [
            dict(drawing)
            for drawing in (document.get("related_drawings") or [])
            if drawing and (
                bool(drawing.get("selected_for_item"))
                or drawing.get("drawing_association_id") is not None
                or bool(drawing.get("is_primary_drawing"))
            )
        ]

    @staticmethod
    def _pdm_item_association_label(association: dict) -> str:
        association_type = str(
            association.get("association_type") or "CONTENT"
        ).strip().upper().replace("_", " ")
        identity = " - ".join(
            value for value in (
                str(association.get("item_number") or "").strip(),
                str(association.get("item_name") or "").strip(),
            ) if value
        ) or f"Item {association.get('item_id') or '-'}"
        version = str(association.get("item_version_label") or "").strip()
        return f"{association_type}: {identity}" + (
            f" | Item {version}" if version else ""
        )

    @classmethod
    def _pdm_item_association_lines(cls, document: dict) -> list[str]:
        return [
            cls._pdm_item_association_label(association)
            for association in cls._pdm_document_associations(document)
        ]

    def _pdm_related_item_text(self, document: dict) -> str:
        associations = self._pdm_document_associations(document)
        if not associations:
            return "Unassociated"
        if len(associations) == 1:
            return self._pdm_item_association_label(associations[0])
        owner_count = sum(
            1 for association in associations
            if str(association.get("association_type") or "").upper() == "OWNER"
        )
        suffix = f" | OWNER: {owner_count}" if owner_count else ""
        return f"{len(associations)} Item associations{suffix}"

    @staticmethod
    def _cad_representation_role(document: dict) -> str:
        return str(document.get("association_type") or "CONTENT").upper().replace(
            "_", " "
        )

    def _cad_representation_drawings_text(self, document: dict) -> str:
        assigned = self._pdm_selected_drawings(document)
        drawings = assigned or list(document.get("related_drawings") or [])
        if not drawings:
            return "No DRW"
        labels = []
        for drawing in drawings:
            label = str(
                drawing.get("file_name") or drawing.get("name") or drawing.get("id")
            )
            if drawing.get("is_primary_drawing"):
                label += " [PRIMARY]"
            labels.append(label)
        return ", ".join(labels)

    def _cad_payload_identity(self, payload: dict) -> str:
        return str(payload.get("file_name") or payload.get("name") or payload.get("id") or "CAD")

    def _cad_representation_paths(self, cad_document_id: int) -> list[str]:
        wanted = int(cad_document_id)
        paths = []

        def walk_payload(node: dict, ancestors: list[str]) -> None:
            if not node:
                return
            label = self._cad_payload_identity(node)
            current = [*ancestors, label]
            try:
                if int(node.get("id")) == wanted:
                    paths.append(" > ".join(current))
            except Exception:
                pass
            for child in node.get("children") or []:
                walk_payload(child, current)

        for root in list(getattr(self, "_pdm_cad_roots", []) or []):
            walk_payload(root, [])
        if paths:
            return list(dict.fromkeys(paths))

        def walk_item(item: QTreeWidgetItem, ancestors: list[str]) -> None:
            payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
            label = self._cad_payload_identity(payload) or item.text(CAD_COL_NAME)
            current = [*ancestors, label]
            try:
                if int(item.data(0, PDM_CAD_DOCUMENT_ID_ROLE)) == wanted:
                    paths.append(" > ".join(current))
            except Exception:
                pass
            for index in range(item.childCount()):
                child = item.child(index)
                if not self._is_lazy_placeholder(child):
                    walk_item(child, current)

        tree = getattr(self, "_cad_tree", None)
        if tree is not None:
            for index in range(tree.topLevelItemCount()):
                walk_item(tree.topLevelItem(index), [])
        return list(dict.fromkeys(paths))

    def _cad_representation_path_text(self, document: dict, max_paths: int = 2) -> str:
        try:
            paths = self._cad_representation_paths(int(document["id"]))
        except Exception:
            paths = []
        if not paths:
            return "CAD BOM location: not loaded"
        shown = paths[:max(1, int(max_paths))]
        if len(paths) > len(shown):
            shown.append(f"+{len(paths) - len(shown)} more location(s)")
        return "\n".join(shown)

    def _cad_representation_summary_line(self, document: dict) -> str:
        role = self._cad_representation_role(document)
        file_name = self._cad_payload_identity(document)
        drawings = self._cad_representation_drawings_text(document)
        path = self._cad_representation_path_text(document, max_paths=1)
        return f"{role}: {file_name} | DRW: {drawings} | Path: {path}"

    def _add_pdm_cad_node(
        self, document: dict, parent_item: QTreeWidgetItem | None = None
    ) -> QTreeWidgetItem:
        associations = self._pdm_document_associations(document)
        primary_association = associations[0] if associations else {}
        association_type = str(
            primary_association.get("association_type")
            or document.get("association_type") or ""
        ).strip().upper()
        related_item = self._pdm_related_item_text(document)
        revision = self._pdm_cad_revision_text(document)
        excluded = bool(
            document.get("member_build_excluded")
            or document.get("build_excluded")
        )
        values = [
            str(document.get("file_name") or ""),
            str(document.get("name") or ""),
            str(document.get("category") or ""),
            revision,
            str(document.get("lifecycle_state") or ""),
            related_item,
            self._pdm_cad_checkout_text(document),
            "Excluded" if excluded else "Included",
            str(max(1, int(document.get("quantity") or 1))),
        ]
        item = QTreeWidgetItem(values)
        item.setData(0, PDM_OBJECT_KIND_ROLE, PDM_OBJECT_CAD)
        item.setData(0, PDM_CAD_DOCUMENT_ID_ROLE, int(document["id"]))
        item.setData(0, PDM_ASSOCIATION_ID_ROLE, document.get("association_id"))
        item.setData(0, PDM_ASSOCIATION_TYPE_ROLE, association_type or None)
        item.setData(0, PDM_ASSOCIATED_ITEM_ID_ROLE, document.get("item_id"))
        item.setData(0, PDM_CAD_MEMBER_ID_ROLE, document.get("member_id"))
        item.setData(0, PDM_CAD_PAYLOAD_ROLE, dict(document))
        item.setData(0, PDM_NODE_PAYLOAD_ROLE, dict(document))
        child_payloads = list(document.get("children") or [])
        item.setData(0, PDM_CHILDREN_PAYLOAD_ROLE, child_payloads)
        item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, bool(child_payloads))
        item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, not bool(child_payloads))
        item.setData(0, Qt.UserRole, document.get("item_id"))
        item.setIcon(CAD_COL_NUMBER, _pdm_cad_icon(document.get("category")))
        drawing_files = ", ".join(
            str(drawing.get("file_name") or drawing.get("name") or "")
            for drawing in document.get("related_drawings") or []
        ) or "None"
        association_lines = self._pdm_item_association_lines(document)
        iteration_warnings = self._decorate_pdm_cad_iteration_health(item, document)
        item.setToolTip(
            CAD_COL_NUMBER,
            "CAD Document\n"
            f"File: {document.get('file_name') or '-'}\n"
            f"Creo file version: {self._pdm_creo_file_text(document) or '-'}\n"
            f"CAD iteration health:\n"
            + ("\n".join(f"  ⚠ {line}" for line in iteration_warnings) if iteration_warnings else "  OK")
            + "\n"
            f"Related drawings: {drawing_files}\n"
            f"CAD BOM location:\n{self._cad_representation_path_text(document)}\n"
            f"Item associations ({len(association_lines)}):\n"
            + ("\n".join(f"  {line}" for line in association_lines) if association_lines else "  None")
            + "\n"
            f"Checkout: {self._pdm_cad_checkout_text(document)}",
        )
        if not associations:
            item.setForeground(CAD_COL_ASSOCIATION, QBrush(QColor("#b45309")))
        if excluded:
            item.setForeground(CAD_COL_BUILD, QBrush(QColor("#b45309")))
        if parent_item is None:
            self._cad_tree.addTopLevelItem(item)
        else:
            parent_item.addChild(item)
        self._ensure_pdm_lazy_placeholder(item)
        return item

    def _apply_pdm_cad_tree_item_data(self, item: QTreeWidgetItem, document: dict) -> None:
        payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
        payload.update(dict(document or {}))
        associations = self._pdm_document_associations(payload)
        primary_association = associations[0] if associations else {}
        association_type = str(
            primary_association.get("association_type")
            or payload.get("association_type") or ""
        ).strip().upper()
        related_item = self._pdm_related_item_text(payload)
        revision = self._pdm_cad_revision_text(payload)
        excluded = bool(payload.get("member_build_excluded") or payload.get("build_excluded"))
        item.setText(CAD_COL_NAME, str(payload.get("file_name") or ""))
        item.setText(CAD_COL_DESCRIPTION, str(payload.get("name") or ""))
        item.setText(CAD_COL_CATEGORY, str(payload.get("category") or ""))
        item.setText(CAD_COL_REV, revision)
        item.setText(CAD_COL_STATE, str(payload.get("lifecycle_state") or ""))
        item.setText(CAD_COL_ASSOCIATION, related_item)
        item.setText(CAD_COL_CHECKOUT, self._pdm_cad_checkout_text(payload))
        item.setText(CAD_COL_BUILD, "Excluded" if excluded else "Included")
        item.setText(CAD_COL_QTY, str(max(1, int(payload.get("quantity") or 1))))
        item.setData(0, PDM_CAD_PAYLOAD_ROLE, payload)
        item.setData(0, PDM_NODE_PAYLOAD_ROLE, payload)
        item.setData(0, PDM_ASSOCIATION_ID_ROLE, payload.get("association_id"))
        item.setData(0, PDM_ASSOCIATION_TYPE_ROLE, association_type or None)
        item.setData(0, PDM_ASSOCIATED_ITEM_ID_ROLE, payload.get("item_id"))
        item.setData(0, Qt.UserRole, payload.get("item_id"))
        item.setIcon(CAD_COL_NUMBER, _pdm_cad_icon(payload.get("category")))
        iteration_warnings = self._decorate_pdm_cad_iteration_health(item, payload)
        drawing_files = ", ".join(
            str(drawing.get("file_name") or drawing.get("name") or "")
            for drawing in payload.get("related_drawings") or []
        ) or "None"
        association_lines = self._pdm_item_association_lines(payload)
        item.setToolTip(
            CAD_COL_NUMBER,
            "CAD Document\n"
            f"File: {payload.get('file_name') or '-'}\n"
            f"Creo file version: {self._pdm_creo_file_text(payload) or '-'}\n"
            f"CAD iteration health:\n"
            + ("\n".join(f"  ⚠ {line}" for line in iteration_warnings) if iteration_warnings else "  OK")
            + "\n"
            f"Related drawings: {drawing_files}\n"
            f"CAD BOM location:\n{self._cad_representation_path_text(payload)}\n"
            f"Item associations ({len(association_lines)}):\n"
            + ("\n".join(f"  {line}" for line in association_lines) if association_lines else "  None")
            + "\n"
            f"Checkout: {self._pdm_cad_checkout_text(payload)}",
        )

    def _apply_ebom_cad_tree_item_data(self, item: QTreeWidgetItem, document: dict) -> None:
        payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
        payload.update(dict(document or {}))
        association_type = str(
            payload.get("association_type")
            or item.data(0, PDM_ASSOCIATION_TYPE_ROLE)
            or "CONTENT"
        ).upper()
        item_id = (
            payload.get("item_id")
            or item.data(0, PDM_ASSOCIATED_ITEM_ID_ROLE)
            or item.data(0, Qt.UserRole)
        )
        revision = self._pdm_cad_revision_text(payload)
        cad_name = str(
            payload.get("file_name") or payload.get("name") or "CAD Document"
        )
        item.setText(BOM_COL_NAME, cad_name)
        item.setText(BOM_COL_AES, association_type.replace("_", " "))
        item.setText(BOM_COL_TYPE, f"CAD {str(payload.get('category') or '').title()}")
        item.setText(BOM_COL_REV, revision)
        item.setText(BOM_COL_STATUS, self._pdm_cad_checkout_text(payload))
        try:
            if item_id is not None:
                item_id = int(item_id)
                item.setData(0, Qt.UserRole, item_id)
                item.setData(0, PDM_ASSOCIATED_ITEM_ID_ROLE, item_id)
        except Exception:
            pass
        item.setData(0, PDM_ASSOCIATION_ID_ROLE, payload.get("association_id"))
        item.setData(0, PDM_ASSOCIATION_TYPE_ROLE, association_type)
        item.setData(0, PDM_CAD_PAYLOAD_ROLE, payload)
        item.setData(0, PDM_NODE_PAYLOAD_ROLE, payload)
        item.setIcon(BOM_COL_NAME, _pdm_cad_icon(payload.get("category")))
        item.setToolTip(
            BOM_COL_NAME,
            "Associated CAD Document (not an EBOM usage)\n"
            f"File: {payload.get('file_name') or '-'}\n"
            f"Association: {association_type.replace('_', ' ')}\n"
            f"CAD version: {payload.get('revision') or 'A'}.{int(payload.get('iteration') or 1)}\n"
            f"Creo file version: {self._pdm_creo_file_text(payload) or '-'}\n"
            f"Lifecycle: {payload.get('lifecycle_state') or '-'}\n"
            f"Checkout: {self._pdm_cad_checkout_text(payload)}",
        )

    def _refresh_pdm_cad_filter(self) -> int:
        tree = getattr(self, "_cad_tree", None)
        if tree is None:
            return 0
        query = str(self.search_input.text() or "").strip().casefold()
        visible_count = 0

        def recurse(item: QTreeWidgetItem) -> bool:
            nonlocal visible_count
            if self._is_lazy_placeholder(item):
                item.setHidden(False)
                return False
            own_text = " ".join(
                str(item.text(column) or "") for column in range(tree.columnCount())
            )
            payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
            drawing_text = " ".join(
                " ".join(
                    str(value or "")
                    for value in (
                        drawing.get("file_name"),
                        drawing.get("name"),
                        self._pdm_creo_file_text(drawing),
                    )
                )
                for drawing in payload.get("related_drawings") or []
            )
            warning_text = " ".join(self._pdm_cad_iteration_warnings(payload))
            own_text = f"{own_text} {drawing_text} {warning_text}".casefold()
            own_match = not query or query in own_text
            child_match = False
            for index in range(item.childCount()):
                child_match = recurse(item.child(index)) or child_match
            show = own_match or child_match
            item.setHidden(not show)
            if own_match:
                visible_count += 1
            if query and child_match:
                item.setExpanded(True)
            return show

        for index in range(tree.topLevelItemCount()):
            recurse(tree.topLevelItem(index))
        return visible_count

    def _load_pdm_cad_tree(self) -> None:
        tree = getattr(self, "_cad_tree", None)
        if tree is None:
            return
        self._set_tree_loading(True)
        try:
            self._pdm_cad_working_file_index = self._build_pdm_cad_working_file_index()
            data = (
                self.bom_service.get_pdm_cad_structure()
                if self.session.project_id else {"roots": [], "document_count": 0}
            )
            roots = list(data.get("roots") or [])
            self._pdm_cad_roots = roots
            self._pdm_folders("CAD", refresh=True)
            self._pdm_cad_scope_path, render_roots = self._pdm_roots_for_reload_scope(
                "cad", roots, getattr(self, "_pdm_cad_scope_path", []) or []
            )
            self._render_pdm_cad_roots(render_roots)
            all_documents = list(self.bom_service.list_pdm_cad_documents() or []) if self.session.project_id else []
            models = [
                row for row in all_documents
                if str(row.get("category") or "").upper()
                in {"ASSEMBLY", "COMPONENT"}
            ]
            unassociated = sum(1 for row in models if not row.get("association_id"))
            drawing_count = int(data.get("drawing_count") or 0)
            unbound_drawings = int(data.get("unbound_drawing_count") or 0)
            cad_iteration_warning_count = sum(
                1
                for row in all_documents
                if self._pdm_cad_iteration_warnings(row)
            )
            self.bom_health_label.setText(
                f"Models: {int(data.get('document_count') or len(models))}  |  "
                f"Drawings: {drawing_count}  |  Unassociated: {unassociated}"
                + (f"  |  Drawings to bind: {unbound_drawings}" if unbound_drawings else "")
                + (f"  |  CAD file warnings: {cad_iteration_warning_count}" if cad_iteration_warning_count else "")
            )
            self.bom_health_label.setStyleSheet(
                "font-size:8pt;font-weight:700;"
                f"color:{'#b45309' if cad_iteration_warning_count else '#475569'};"
                "background:transparent;border:none;"
            )
            self.bom_mode_selector.setToolTip(
                "Native CAD Document assembly/member structure. Item delivery files and "
                "Item integrity are intentionally shown only in EBOM / Item Structure."
            )
        except Exception as exc:
            tree.clear()
            self.show_alert(f"CAD Structure could not be loaded: {exc}", "error")
        finally:
            try:
                tree.setUpdatesEnabled(True)
            except Exception:
                pass
            self._set_tree_loading(False)

    def _add_ebom_associated_cad_node(
        self, document: dict, item_parent: QTreeWidgetItem, item_id: int,
        insert_index: int | None = None,
    ) -> QTreeWidgetItem:
        association_type = str(document.get("association_type") or "CONTENT").upper()
        assigned_drawings = self._pdm_selected_drawings(document)
        revision = self._pdm_cad_revision_text(document)
        cad_name = str(
            document.get("file_name") or document.get("name") or "CAD Document"
        )
        cad_item = QTreeWidgetItem([""] * self._ebom_tree.columnCount())
        role_text = association_type.replace("_", " ")
        drawing_names = self._cad_representation_drawings_text(document)
        location_text = self._cad_representation_path_text(document)
        cad_item.setText(BOM_COL_NAME, f"{role_text}  {cad_name}")
        cad_item.setText(BOM_COL_AES, f"DRW: {drawing_names}")
        cad_item.setText(
            BOM_COL_TYPE,
            "CAD Representation" if association_type in {"IMAGE", "CONTRIBUTING_IMAGE"} else f"CAD {str(document.get('category') or '').title()}",
        )
        cad_item.setText(BOM_COL_REV, revision)
        cad_item.setText(BOM_COL_STATUS, self._pdm_cad_checkout_text(document))
        cad_item.setData(0, Qt.UserRole, int(item_id))
        cad_item.setData(0, PDM_OBJECT_KIND_ROLE, PDM_OBJECT_CAD)
        cad_item.setData(0, PDM_CAD_DOCUMENT_ID_ROLE, int(document["id"]))
        cad_item.setData(0, PDM_ASSOCIATION_ID_ROLE, document.get("association_id"))
        cad_item.setData(0, PDM_ASSOCIATION_TYPE_ROLE, association_type)
        cad_item.setData(0, PDM_ASSOCIATED_ITEM_ID_ROLE, int(item_id))
        cad_item.setData(0, PDM_CAD_PAYLOAD_ROLE, dict(document))
        cad_item.setData(0, PDM_NODE_PAYLOAD_ROLE, dict(document))
        cad_item.setIcon(BOM_COL_NAME, _pdm_cad_icon(document.get("category")))
        cad_item.setForeground(BOM_COL_NAME, QBrush(QColor("#355a73")))
        cad_item.setToolTip(
            BOM_COL_NAME,
            "Associated CAD Document (not an EBOM usage)\n"
            f"File: {document.get('file_name') or '-'}\n"
            f"Association: {role_text}\n"
            f"CAD version: {document.get('revision') or 'A'}.{int(document.get('iteration') or 1)}\n"
            f"Creo file version: {self._pdm_creo_file_text(document) or '-'}\n"
            f"Item drawing assignment: {drawing_names}\n"
            f"CAD BOM location:\n{location_text}\n"
            f"Lifecycle: {document.get('lifecycle_state') or '-'}\n"
            f"Checkout: {self._pdm_cad_checkout_text(document)}",
        )
        cad_item.setToolTip(
            BOM_COL_AES,
            "CAD-to-Item association type. Right-click this CAD row to change or remove it.",
        )
        if insert_index is None:
            item_parent.addChild(cad_item)
        else:
            item_parent.insertChild(max(0, int(insert_index)), cad_item)
        for drawing in assigned_drawings:
            drawing_payload = dict(drawing)
            drawing_payload["item_id"] = int(item_id)
            drawing_payload["association_id"] = (
                drawing.get("drawing_association_id")
                or drawing.get("association_id")
            )
            drawing_payload["association_type"] = str(
                drawing.get("drawing_association_type")
                or drawing.get("association_type") or "CONTENT"
            ).upper()
            drawing_item = QTreeWidgetItem([""] * self._ebom_tree.columnCount())
            drawing_item.setText(
                BOM_COL_NAME,
                str(drawing.get("file_name") or drawing.get("name") or "CAD Drawing"),
            )
            drawing_item.setText(
                BOM_COL_AES,
                "PRIMARY DRAWING" if drawing.get("is_primary_drawing") else "DRAWING",
            )
            drawing_item.setText(BOM_COL_TYPE, "CAD Drawing")
            drawing_item.setText(
                BOM_COL_REV, self._pdm_cad_revision_text(drawing)
            )
            drawing_item.setText(
                BOM_COL_STATUS, self._pdm_cad_checkout_text(drawing)
            )
            drawing_item.setData(0, Qt.UserRole, int(item_id))
            drawing_item.setData(0, PDM_OBJECT_KIND_ROLE, PDM_OBJECT_CAD)
            drawing_item.setData(
                0, PDM_CAD_DOCUMENT_ID_ROLE, int(drawing["id"])
            )
            drawing_item.setData(
                0, PDM_ASSOCIATION_ID_ROLE, drawing_payload.get("association_id")
            )
            drawing_item.setData(
                0, PDM_ASSOCIATION_TYPE_ROLE,
                drawing_payload.get("association_type"),
            )
            drawing_item.setData(0, PDM_ASSOCIATED_ITEM_ID_ROLE, int(item_id))
            drawing_item.setData(0, PDM_CAD_PAYLOAD_ROLE, drawing_payload)
            drawing_item.setData(0, PDM_NODE_PAYLOAD_ROLE, drawing_payload)
            drawing_item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, False)
            drawing_item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, True)
            drawing_item.setIcon(BOM_COL_NAME, _pdm_cad_icon("DRAWING"))
            drawing_item.setForeground(BOM_COL_NAME, QBrush(QColor("#526f85")))
            drawing_item.setToolTip(
                BOM_COL_NAME,
                "Drawing explicitly assigned to this Item\n"
                f"Model: {cad_name}\n"
                f"Role: {'Primary drawing' if drawing.get('is_primary_drawing') else 'Supporting drawing'}\n"
                f"Creo file version: {self._pdm_creo_file_text(drawing) or '-'}",
            )
            cad_item.addChild(drawing_item)
        return cad_item

    def _add_released_ebom_node(
        self, info: dict, parent_item: QTreeWidgetItem | None = None,
        associations_by_item: dict | None = None,
    ) -> QTreeWidgetItem | None:
        payload = dict(info or {})
        if self._is_deleted_bom_payload(payload):
            return None
        payload["id"] = int(payload.get("bom_id") or payload.get("id"))
        payload["current_version"] = str(
            payload.get("version_label") or payload.get("current_version") or ""
        )
        payload["status"] = str(payload.get("state") or payload.get("status") or "")
        payload["relation_parent_id"] = payload.get("effective_parent_bom_id")
        payload["usage_id"] = payload.get("item_usage_id") or payload.get("usage_id")
        payload["quantity"] = int(payload.get("source_quantity") or 1)
        payload["_has_children"] = bool(payload.get("children"))
        item = QTreeWidgetItem([""] * self._ebom_tree.columnCount())
        self._apply_tree_item_data(item, payload)
        item.setText(BOM_COL_AES, str(payload.get("part_number") or ""))
        item.setToolTip(
            BOM_COL_AES,
            "Item Number (PLM identity)"
            + (
                f"\nAES delivery reference: {payload.get('aes_number')}"
                if payload.get("aes_number")
                else "\nAES delivery reference: not assigned"
            ),
        )
        item.setData(0, PDM_OBJECT_KIND_ROLE, PDM_OBJECT_ITEM)
        item.setData(0, PDM_NODE_PAYLOAD_ROLE, dict(payload))
        child_payloads = list(payload.get("children") or [])
        item.setData(0, PDM_CHILDREN_PAYLOAD_ROLE, child_payloads)
        item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, not bool(child_payloads))
        item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, bool(child_payloads))
        item.setData(
            0,
            PDM_EBOM_ASSOCIATIONS_ROLE,
            list((associations_by_item or {}).get(int(payload["id"]), [])),
        )
        item.setData(0, PDM_ASSOCIATIONS_SHOWN_ROLE, False)
        item.setIcon(BOM_COL_NAME, _pdm_item_icon())
        item.setText(EBOM_COL_SOURCE_QTY, str(int(payload.get("source_quantity") or 1)))
        item.setText(
            EBOM_COL_EFFECTIVE_QTY,
            str(int(payload.get("effective_quantity") or 1)),
        )
        item.setText(EBOM_COL_LEVEL, str(int(payload.get("level") or 0)))
        promotion = list(payload.get("promoted_through") or [])
        if promotion:
            labels = " > ".join(
                str(
                    value.get("part_number")
                    or value.get("aes_number")
                    or value.get("name")
                    or value.get("bom_id")
                )
                for value in promotion
            )
            item.setToolTip(
                EBOM_COL_EFFECTIVE_QTY,
                f"Promoted through {labels}; flattened quantities are multiplied.",
            )
        if parent_item is None:
            self._ebom_tree.addTopLevelItem(item)
        else:
            parent_item.addChild(item)
        self._ensure_pdm_lazy_placeholder(item)
        return item

    @staticmethod
    def _is_deleted_bom_payload(payload: dict | None) -> bool:
        data = payload or {}
        if data.get("deleted_at"):
            return True
        for key in ("status", "state", "lifecycle_state", "revision_state"):
            if str(data.get(key) or "").strip().casefold() == "deleted":
                return True
        return False

    def _load_released_ebom_tree(self) -> None:
        self._set_tree_loading(True)
        try:
            QApplication.processEvents()
            data = self.bom_service.get_released_ebom_project(
                int(self.session.project_id)
            ) if self.session.project_id else {"roots": []}
            associations_by_item = defaultdict(list)
            if self.session.project_id:
                for document in self.bom_service.list_pdm_cad_documents() or []:
                    if str(document.get("category") or "").upper() == "DRAWING":
                        continue
                    for association in self._pdm_document_associations(document):
                        if association.get("item_id") is None:
                            continue
                        item_document = dict(document)
                        item_document.update(association)
                        item_document["association_id"] = (
                            association.get("association_id")
                            or association.get("id")
                        )
                        item_document["item_id"] = int(association["item_id"])
                        item_document["association_type"] = str(
                            association.get("association_type") or "CONTENT"
                        ).upper()
                        # Global CAD data contains every model-bound drawing. Item-specific
                        # drawing assignments are loaded only when the user chooses Show CAD.
                        item_document["related_drawings"] = []
                        associations_by_item[int(association["item_id"])].append(
                            item_document
                        )
            self._ebom_associations_by_item = associations_by_item
            visible_roots = list(data.get("roots") or [])
            excluded_roots = list(data.get("excluded_roots") or [])
            flattened_roots = list(data.get("flattened_roots") or [])
            self._pdm_ebom_roots = visible_roots
            self._pdm_folders("EBOM", refresh=True)
            # Status/type choices for Advanced Filter are derived lazily from
            # these cached payloads. Reset the lightweight choice cache when
            # the EBOM payload changes; do not materialize every Qt tree row
            # merely to open the filter dialog.
            self._ebom_filter_choice_cache = {}
            self._pdm_ebom_scope_path, render_roots = self._pdm_roots_for_reload_scope(
                "ebom", visible_roots, getattr(self, "_pdm_ebom_scope_path", []) or []
            )
            self._render_pdm_ebom_roots(render_roots)
            try:
                score = max(
                    0,
                    self.issue_service.health_score()
                    - len(getattr(self, "missing_ids", set()) or set()),
                )
                color = "#2e7d32" if score >= 85 else ("#a16207" if score >= 65 else "#b91c1c")
                self.bom_health_label.setText(f"Item health: {score}/100")
                self.bom_health_label.setStyleSheet(
                    f"font-size:11px;font-weight:700;color:{color};background:transparent;border:none;"
                )
            except Exception:
                self.bom_health_label.setText(
                    "Items: " + str(sum(
                        1 for row in self._iter_tree_items(self._ebom_tree)
                        if row.data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD
                    ))
                )
            self.bom_mode_selector.setToolTip(
                "Persisted EBOM contains Item usages. Associated CAD Documents are "
                "available from the Visual ribbon as indented relation rows and "
                "are not EBOM usages. "
                f"Visible roots: {len(visible_roots)}."
            )
            if not visible_roots and (excluded_roots or flattened_roots):
                self.show_alert(
                    "Released EBOM has no visible deliverable root. In CAD Structure, "
                    "review items marked NOT FOR DELIVERY or FLATTEN.",
                    "warning",
                )
            else:
                self.hide_alert()
        except Exception as exc:
            self._ebom_tree.clear()
            self.show_alert(f"Item Structure could not be loaded: {exc}", "error")
        finally:
            self._ebom_tree.setUpdatesEnabled(True)
            self._set_tree_loading(False)

    def _pdm_current_item_id(self) -> int | None:
        value = getattr(self, "current_part_id", None)
        if value is None:
            QMessageBox.information(
                self, "PDM", "Select an Item in CAD Structure or Item Structure first."
            )
            return None
        return int(value)

    def manage_pdm_cad_structure(self) -> None:
        dialog = QDialog(self)
        dialog.setWindowTitle("CAD Document Structure")
        dialog.resize(920, 620)
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel(
            "Managed CAD structure. CAD Documents may exist without an Item association."
        ))
        tree = QTreeWidget()
        tree.setHeaderLabels([
            "CAD File", "Name", "Category", "Association", "Rev/Iter",
            "State", "Checkout", "Qty", "Build",
        ])
        tree.header().setSectionResizeMode(QHeaderView.Stretch)
        tree.setAlternatingRowColors(True)
        layout.addWidget(tree, 1)

        editor = QGridLayout()
        parent_input = QComboBox()
        child_input = QComboBox()
        quantity_input = QSpinBox()
        quantity_input.setRange(1, 100000)
        excluded_input = QCheckBox("Exclude this CAD member from Item Structure build")
        editor.addWidget(QLabel("Parent assembly"), 0, 0)
        editor.addWidget(parent_input, 0, 1)
        editor.addWidget(QLabel("Child CAD Document"), 1, 0)
        editor.addWidget(child_input, 1, 1)
        editor.addWidget(QLabel("Quantity"), 0, 2)
        editor.addWidget(quantity_input, 0, 3)
        editor.addWidget(excluded_input, 1, 2, 1, 2)
        layout.addLayout(editor)
        action_row = QHBoxLayout()
        add_btn = QPushButton("Add or update CAD member")
        add_btn.setObjectName("primary")
        remove_btn = QPushButton("Remove selected member")
        remove_btn.setObjectName("secondary")
        checkout_btn = QPushButton("Check Out")
        checkin_btn = QPushButton("Check In")
        revise_btn = QPushButton("New Revision")
        release_btn = QPushButton("Release")
        action_row.addWidget(add_btn)
        action_row.addWidget(remove_btn)
        action_row.addWidget(checkout_btn)
        action_row.addWidget(checkin_btn)
        action_row.addWidget(revise_btn)
        action_row.addWidget(release_btn)
        action_row.addStretch()
        layout.addLayout(action_row)
        close_buttons = QDialogButtonBox(QDialogButtonBox.Close)
        close_buttons.rejected.connect(dialog.reject)
        layout.addWidget(close_buttons)
        can_manage = self.perm.can("manage_parts")
        add_btn.setEnabled(can_manage)
        remove_btn.setEnabled(can_manage)
        checkout_btn.setEnabled(can_manage)
        checkin_btn.setEnabled(can_manage)
        revise_btn.setEnabled(can_manage)
        release_btn.setEnabled(can_manage)

        def add_node(node, parent=None):
            association = str(node.get("association_type") or "UNASSOCIATED")
            item = QTreeWidgetItem([
                str(node.get("file_name") or node.get("name") or ""),
                str(node.get("name") or ""),
                str(node.get("category") or ""),
                association,
                f"{node.get('revision') or 'A'}.{int(node.get('iteration') or 1)}",
                str(node.get("lifecycle_state") or ""),
                str(node.get("checked_out_by") or "—"),
                str(int(node.get("quantity") or 1)),
                "EXCLUDED" if node.get("member_build_excluded") or node.get("build_excluded") else "INCLUDED",
            ])
            item.setData(0, Qt.UserRole, int(node["id"]))
            item.setData(0, Qt.UserRole + 1, node.get("member_id"))
            if association == "UNASSOCIATED":
                item.setForeground(3, QBrush(QColor("#b45309")))
            if parent is None:
                tree.addTopLevelItem(item)
            else:
                parent.addChild(item)
            for child in node.get("children") or []:
                add_node(child, item)

        def refresh():
            documents = self.bom_service.list_pdm_cad_documents() or []
            previous_parent = parent_input.currentData()
            previous_child = child_input.currentData()
            parent_input.clear()
            child_input.clear()
            for document in documents:
                label = str(document.get("file_name") or "")
                name = str(document.get("name") or "").strip()
                if name and name.casefold() != label.casefold():
                    label = f"{label} - {name}"
                if str(document.get("category") or "").upper() == "ASSEMBLY":
                    parent_input.addItem(label, int(document["id"]))
                child_input.addItem(label, int(document["id"]))
            for combo, previous in ((parent_input, previous_parent), (child_input, previous_child)):
                if previous is not None:
                    index = combo.findData(previous)
                    if index >= 0:
                        combo.setCurrentIndex(index)
            tree.clear()
            for root in (self.bom_service.get_pdm_cad_structure().get("roots") or []):
                add_node(root)
            tree.expandToDepth(1)

        def add_member():
            parent_id = parent_input.currentData()
            child_id = child_input.currentData()
            if parent_id is None or child_id is None:
                QMessageBox.information(dialog, "CAD Structure", "Select parent and child CAD Documents.")
                return
            try:
                self.bom_service.add_pdm_cad_member(
                    int(parent_id), int(child_id), quantity_input.value(),
                    excluded_input.isChecked(),
                )
            except Exception as exc:
                QMessageBox.critical(dialog, "CAD Structure", str(exc))
                return
            refresh()

        def remove_member():
            selected = tree.currentItem()
            member_id = selected.data(0, Qt.UserRole + 1) if selected else None
            if member_id is None:
                QMessageBox.information(dialog, "CAD Structure", "Select a child CAD member, not a root.")
                return
            undo_record = self.undo_service.snapshot_cad_member_remove(
                int(member_id), f"Remove CAD Component {selected.text(0)}"
            )
            self.bom_service.remove_pdm_cad_member(int(member_id))
            self.undo_service.push(undo_record)
            refresh()
            try:
                self.window().statusBar().showMessage("CAD component removed. Press Ctrl+Z to undo.", 6000)
            except Exception:
                pass

        def selected_cad_id():
            selected = tree.currentItem()
            return int(selected.data(0, Qt.UserRole)) if selected else None

        def checkout_document():
            cad_id = selected_cad_id()
            if cad_id is None:
                QMessageBox.information(dialog, "CAD Document", "Select a CAD Document.")
                return
            try:
                document = self.bom_service.pdm_service.repo.get_cad_document(cad_id) or {}
                self._checkout_pdm_cad_document(cad_id, document)
                refresh()
            except Exception as exc:
                QMessageBox.warning(dialog, "Check Out", str(exc))

        def checkin_document():
            cad_id = selected_cad_id()
            if cad_id is None:
                QMessageBox.information(dialog, "CAD Document", "Select a CAD Document.")
                return
            path, _filter = QFileDialog.getOpenFileName(
                dialog, "CAD Document Check In", str(getattr(self, "working_dir", "") or ""),
                "Creo CAD (*.asm *.asm.* *.prt *.prt.* *.drw *.drw.*);;All files (*)",
            )
            if not path:
                return
            note, accepted = QInputDialog.getText(
                dialog, "CAD Document Check In", "Check-in note:"
            )
            if not accepted:
                return
            try:
                self.bom_service.checkin_pdm_cad_document(cad_id, path, note)
                refresh()
            except Exception as exc:
                QMessageBox.warning(dialog, "Check In", str(exc))

        def revise_document():
            cad_id = selected_cad_id()
            if cad_id is None:
                QMessageBox.information(dialog, "CAD Document", "Select a CAD Document.")
                return
            try:
                self.bom_service.revise_pdm_cad_document(cad_id)
                refresh()
            except Exception as exc:
                QMessageBox.warning(dialog, "New CAD Revision", str(exc))

        def release_document():
            cad_id = selected_cad_id()
            if cad_id is None:
                QMessageBox.information(dialog, "CAD Document", "Select a CAD Document.")
                return
            if QMessageBox.question(
                dialog, "Release CAD Document", "Release the selected CAD Document iteration?",
                QMessageBox.Yes | QMessageBox.No,
            ) != QMessageBox.Yes:
                return
            try:
                self.bom_service.release_pdm_cad_document(cad_id)
                refresh()
            except Exception as exc:
                QMessageBox.warning(dialog, "Release CAD Document", str(exc))

        add_btn.clicked.connect(add_member)
        remove_btn.clicked.connect(remove_member)
        checkout_btn.clicked.connect(checkout_document)
        checkin_btn.clicked.connect(checkin_document)
        revise_btn.clicked.connect(revise_document)
        release_btn.clicked.connect(release_document)
        refresh()
        dialog.exec_()

    def register_cad_document(
        self, item_id: int | None = None, drawing_owner_cad_id: int | None = None
    ) -> int | None:
        if isinstance(item_id, bool):
            item_id = None
        if not self.perm.can("manage_parts"):
            QMessageBox.warning(self, "Permission denied", "You cannot register CAD Documents.")
            return
        drawing_only = drawing_owner_cad_id is not None
        path, _selected_filter = QFileDialog.getOpenFileName(
            self,
            "Register Related Drawing" if drawing_only else "Register CAD Document",
            str(getattr(self, "working_dir", "") or ""),
            (
                "Creo Drawing (*.drw *.drw.*)"
                if drawing_only else
                "Creo CAD (*.asm *.asm.* *.prt *.prt.* *.drw *.drw.*);;All files (*)"
            ),
        )
        if not path:
            return
        file_name = os.path.basename(path)
        versionless = re.sub(r"\.\d+$", "", file_name)
        stem, extension = os.path.splitext(versionless)
        category = {
            ".asm": "ASSEMBLY", ".prt": "COMPONENT", ".drw": "DRAWING"
        }.get(extension.casefold(), "OTHER")
        if category == "OTHER":
            QMessageBox.warning(
                self,
                "Register CAD Document",
                "Only Creo .prt, .asm, and .drw files can be registered here.",
            )
            return
        if drawing_only and category != "DRAWING":
            QMessageBox.warning(
                self, "Register Related Drawing", "Select a Creo .drw file."
            )
            return
        drawing_owner = None
        if category == "DRAWING":
            models = [
                document
                for document in (self.bom_service.list_pdm_cad_documents() or [])
                if str(document.get("category") or "").upper()
                in {"ASSEMBLY", "COMPONENT"}
            ]
            if item_id is not None:
                models = [
                    document for document in models
                    if document.get("item_id") is not None
                    and int(document.get("item_id")) == int(item_id)
                ]
            if drawing_owner_cad_id is not None:
                drawing_owner = next(
                    (
                        document for document in models
                        if int(document.get("id") or 0) == int(drawing_owner_cad_id)
                    ),
                    None,
                )
            else:
                if not models:
                    QMessageBox.warning(
                        self,
                        "Register Drawing",
                        "Register or associate the owning PRT/ASM model first. "
                        "A drawing cannot be registered as an isolated CAD node.",
                    )
                    return
                labels = [
                    f"{document.get('file_name') or document.get('name')}"
                    + (f" - {document.get('name')}" if document.get("name") else "")
                    for document in models
                ]
                selected, accepted = QInputDialog.getItem(
                    self,
                    "Bind Drawing to Model",
                    "Owning PRT/ASM model:",
                    labels,
                    0,
                    False,
                )
                if not accepted:
                    return
                drawing_owner = models[labels.index(selected)]
            if drawing_owner is None:
                QMessageBox.warning(
                    self, "Register Drawing", "The owning PRT/ASM model was not found."
                )
                return
        name, accepted = QInputDialog.getText(
            self, "Register CAD Document", "CAD Document name:", text=stem
        )
        if not accepted:
            return
        try:
            cad_document_id = self.bom_service.create_pdm_cad_document(
                number=versionless, name=name, file_name=file_name,
                category=category, authoring_application="CREO",
                drawing_owner_cad_document_id=(
                    int(drawing_owner["id"]) if drawing_owner else None
                ),
            )
        except Exception as exc:
            QMessageBox.critical(self, "Register CAD Document", str(exc))
            return None
        if item_id is not None and category == "DRAWING":
            try:
                self.bom_service.associate_cad_document(
                    int(item_id), int(cad_document_id), "CONTENT"
                )
            except Exception as exc:
                QMessageBox.warning(
                    self,
                    "Drawing Registered",
                    "The drawing was bound to its model, but its Item CONTENT "
                    f"association could not be created:\n{exc}",
                )
            else:
                QMessageBox.information(
                    self,
                    "Drawing Registered",
                    f"{file_name} is now related to "
                    f"{drawing_owner.get('file_name') or drawing_owner.get('name')}. "
                    "It is available in the model details and is not a CAD Structure node.",
                )
        elif item_id is not None:
            types = [
                ("Owner (drives structure and attributes)", "OWNER"),
                ("Contributing image", "CONTRIBUTING_IMAGE"),
                ("Image / alternate CAD representation", "IMAGE"),
                ("Contributing content", "CONTRIBUTING_CONTENT"),
                ("Content / supporting CAD", "CONTENT"),
            ]
            label, accepted = QInputDialog.getItem(
                self, "Register and Associate CAD",
                "Association to the selected Item:",
                [value[0] for value in types], 0, False,
            )
            if accepted:
                association_type = next(value for text, value in types if text == label)
                try:
                    self.bom_service.associate_cad_document(
                        int(item_id), int(cad_document_id), association_type
                    )
                except Exception as exc:
                    QMessageBox.warning(
                        self, "CAD Registered",
                        "The CAD Document was registered, but its Item association "
                        f"could not be created:\n{exc}",
                    )
                    self._refresh_pdm_context_rows(cad_ids=[int(cad_document_id)])
                    return int(cad_document_id)
                QMessageBox.information(
                    self, "Register and Associate CAD",
                    "The CAD Document is now managed and appears beneath the selected Item.",
                )
            else:
                QMessageBox.information(
                    self, "CAD Registered",
                    "The CAD Document was registered without an Item association.",
                )
        elif category == "DRAWING":
            QMessageBox.information(
                self,
                "Drawing Registered",
                f"{file_name} is now related to "
                f"{drawing_owner.get('file_name') or drawing_owner.get('name')}. "
                "It is shown from the model details, not as a separate tree node.",
            )
        else:
            QMessageBox.information(
                self, "Register CAD Document",
                "The file is now a managed CAD Document and may remain CAD-only until associated from an Item.",
            )
        if category == "DRAWING" and drawing_owner is not None:
            self._refresh_pdm_context_rows(
                cad_ids=[int(drawing_owner["id"]), int(cad_document_id)],
                item_ids=[int(item_id)] if item_id is not None else [],
            )
            self._reselect_cad_in_current_view(int(drawing_owner["id"]))
        else:
            try:
                document = self.bom_service.pdm_service.repo.get_cad_document(int(cad_document_id)) or {}
                if (
                    document
                    and getattr(self, "_cad_tree", None) is not None
                    and not getattr(self, "_pdm_cad_scope_path", [])
                    and not self._find_pdm_cad_items([int(cad_document_id)])
                ):
                    self._add_pdm_cad_node(document)
                    self._refresh_pdm_cad_filter()
            except Exception:
                pass
            self._refresh_pdm_context_rows(
                cad_ids=[int(cad_document_id)],
                item_ids=[int(item_id)] if item_id is not None else [],
            )
            if item_id is not None:
                self._refresh_ebom_association_rows_for_item(int(item_id))
        return int(cad_document_id)

    @staticmethod
    def _item_identity_text(details: dict) -> str:
        number = str(details.get("part_number") or "").strip()
        name = str(details.get("name") or details.get("id") or "Item").strip()
        aes = str(details.get("aes_number") or "").strip()
        identity = " — ".join(value for value in (number or "No Number", name) if value)
        return f"{identity}  |  AES {aes}" if aes else identity

    def manage_cad_item_associations(
        self, item_id: int | None = None, focus_cad_id: int | None = None
    ) -> None:
        if isinstance(item_id, bool):
            item_id = None
        item_id = int(item_id) if item_id is not None else self._pdm_current_item_id()
        if item_id is None:
            return
        if not self._ensure_item_checked_out_for_pdm_change(int(item_id)):
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Associations")
        dialog.resize(1180, 650)
        dialog.setStyleSheet("""
            QDialog { background:#e9edf1; color:#1d2935; font:9pt 'Segoe UI'; }
            QLabel#associationTitle { color:#172635; font-size:12pt; font-weight:600; }
            QLabel#associationHint { color:#526577; font-size:9pt; }
            QLabel#associationStatus { color:#2b5878; font-weight:600; }
            QLineEdit, QComboBox {
                background:white; border:1px solid #9eaab5; border-radius:0;
                min-height:23px; padding:1px 5px;
            }
            QTableWidget {
                background:white; alternate-background-color:#f2f5f7;
                border:1px solid #9eaab5; gridline-color:#d4dbe1;
            }
            QHeaderView::section {
                background:#d8e0e6; color:#22313e; border:0;
                border-right:1px solid #aeb8c2; border-bottom:1px solid #9eaab5;
                padding:4px 6px; font-weight:600;
            }
            QPushButton {
                border:1px solid #8d99a4; border-radius:0; background:#e2e7eb;
                min-height:24px; padding:2px 10px;
            }
            QPushButton#primary { background:#246a9b; color:white; border-color:#1d587f; }
        """)
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(7)
        details = self.bom_service.get_part_details(item_id) or {}
        association_title = QLabel("Edit CAD–Item Associations")
        association_title.setObjectName("associationTitle")
        layout.addWidget(association_title)
        layout.addWidget(QLabel(f"Item: {self._item_identity_text(details)}"))
        association_hint = QLabel(
            "Associate an existing PRT/ASM with this Item, select its relationship type, "
            "and explicitly choose which related DRW documents belong to this Item."
        )
        association_hint.setObjectName("associationHint")
        association_hint.setWordWrap(True)
        layout.addWidget(association_hint)
        search = QLineEdit()
        search.setPlaceholderText(
            "Search CAD file, name, category, association type, drawing, or related Item..."
        )
        search.setClearButtonEnabled(True)
        layout.addWidget(search)
        table = QTableWidget(0, 9)
        table.setHorizontalHeaderLabels([
            "", "Status", "CAD File", "Name", "Object Type", "Rev/Iter",
            "Association Type", "Item Drawing Assignment", "Other Item Associations",
        ])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        for column in (0, 1, 2, 4, 5, 6):
            table.horizontalHeader().setSectionResizeMode(
                column, QHeaderView.ResizeToContents
            )
        for column in (3, 7, 8):
            table.horizontalHeader().setSectionResizeMode(column, QHeaderView.Stretch)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)

        controls = QHBoxLayout()
        association_status = QLabel("")
        association_status.setObjectName("associationStatus")
        controls.addWidget(association_status, 1)
        associate_btn = QPushButton("Apply Selected")
        associate_btn.setObjectName("primary")
        controls.addWidget(associate_btn)
        remove_btn = QPushButton("Remove Selected")
        remove_btn.setObjectName("secondary")
        controls.addWidget(remove_btn)
        can_manage = self.perm.can("manage_parts")
        associate_btn.setEnabled(can_manage)
        remove_btn.setEnabled(can_manage)
        layout.addLayout(controls)
        layout.addWidget(table, 1)
        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        row_states = {}
        association_types = (
            "OWNER", "CONTRIBUTING_IMAGE", "IMAGE",
            "CONTRIBUTING_CONTENT", "CONTENT",
        )

        def checked_rows():
            return [
                row for row in range(table.rowCount())
                if table.item(row, 0) is not None
                and table.item(row, 0).checkState() == Qt.Checked
            ]

        def update_drawing_button(state: dict) -> None:
            button = state["drawing_button"]
            drawings = list(state.get("drawings") or [])
            selected_ids = set(state.get("selected_drawing_ids") or set())
            primary_id = state.get("primary_drawing_id")
            if not drawings:
                button.setText("No related DRW")
                button.setToolTip(
                    "Bind one or more related DRW CAD Documents to this model first."
                )
                button.setEnabled(False)
                return
            names = {
                int(drawing["id"]): str(
                    drawing.get("file_name") or drawing.get("name") or drawing["id"]
                )
                for drawing in drawings if drawing.get("id") is not None
            }
            if not selected_ids:
                button.setText("Select related drawing(s)...")
                button.setToolTip(
                    "No drawing is explicitly assigned to this Item for this CAD model."
                )
            else:
                primary_name = (
                    names.get(int(primary_id), "") if primary_id is not None else ""
                )
                if len(selected_ids) == 1:
                    button.setText(
                        primary_name or names.get(next(iter(selected_ids)), "1 drawing")
                    )
                else:
                    button.setText(
                        f"{len(selected_ids)} drawings | Primary: "
                        f"{primary_name or 'not set'}"
                    )
                button.setToolTip("\n".join(
                    ("Primary: " if drawing_id == primary_id else "Supporting: ")
                    + names.get(drawing_id, str(drawing_id))
                    for drawing_id in sorted(selected_ids)
                ))
            button.setEnabled(
                can_manage and not state.get("cad_checked_out", False)
            )

        def choose_drawings(cad_id: int) -> None:
            state = row_states.get(int(cad_id))
            if not state:
                return
            drawings = list(state.get("drawings") or [])
            if not drawings:
                return
            drawing_dialog = QDialog(dialog)
            drawing_dialog.setWindowTitle("Select Item Drawings")
            drawing_dialog.resize(760, 430)
            drawing_layout = QVBoxLayout(drawing_dialog)
            model_name = str(
                state["document"].get("file_name")
                or state["document"].get("name") or cad_id
            )
            heading = QLabel(f"Related drawings for {model_name}")
            heading.setObjectName("associationTitle")
            drawing_layout.addWidget(heading)
            hint = QLabel(
                "Select the drawings that define this EBOM Item. Exactly one selected "
                "drawing is the primary drawing; the others are supporting content."
            )
            hint.setObjectName("associationHint")
            hint.setWordWrap(True)
            drawing_layout.addWidget(hint)
            drawing_table = QTableWidget(len(drawings), 6)
            drawing_table.setHorizontalHeaderLabels([
                "Use", "Primary", "Drawing File", "Name", "Rev/Iter", "Lifecycle"
            ])
            drawing_table.horizontalHeader().setSectionResizeMode(
                QHeaderView.ResizeToContents
            )
            drawing_table.horizontalHeader().setSectionResizeMode(
                2, QHeaderView.Stretch
            )
            drawing_table.horizontalHeader().setSectionResizeMode(
                3, QHeaderView.Stretch
            )
            drawing_table.setSelectionBehavior(QAbstractItemView.SelectRows)
            drawing_table.setAlternatingRowColors(True)
            drawing_table.verticalHeader().setVisible(False)
            drawing_table.blockSignals(True)
            selected_ids = set(state.get("selected_drawing_ids") or set())
            primary_id = state.get("primary_drawing_id")
            for drawing_row, drawing in enumerate(drawings):
                drawing_id = int(drawing["id"])
                use_item = QTableWidgetItem("")
                use_item.setFlags(use_item.flags() | Qt.ItemIsUserCheckable)
                use_item.setCheckState(
                    Qt.Checked if drawing_id in selected_ids else Qt.Unchecked
                )
                use_item.setData(Qt.UserRole, drawing_id)
                primary_item = QTableWidgetItem("")
                primary_item.setFlags(primary_item.flags() | Qt.ItemIsUserCheckable)
                primary_item.setCheckState(
                    Qt.Checked if drawing_id == primary_id else Qt.Unchecked
                )
                primary_item.setData(Qt.UserRole, drawing_id)
                drawing_table.setItem(drawing_row, 0, use_item)
                drawing_table.setItem(drawing_row, 1, primary_item)
                values = (
                    drawing.get("file_name"), drawing.get("name"),
                    self._pdm_cad_revision_text(drawing),
                    drawing.get("lifecycle_state"),
                )
                for column, value in enumerate(values, start=2):
                    drawing_table.setItem(
                        drawing_row, column, QTableWidgetItem(str(value or "-"))
                    )
            drawing_table.blockSignals(False)

            def drawing_item_changed(changed_item: QTableWidgetItem) -> None:
                if changed_item.column() not in {0, 1}:
                    return
                drawing_table.blockSignals(True)
                try:
                    row = changed_item.row()
                    use_item = drawing_table.item(row, 0)
                    primary_item = drawing_table.item(row, 1)
                    if (
                        changed_item.column() == 1
                        and primary_item.checkState() == Qt.Checked
                    ):
                        use_item.setCheckState(Qt.Checked)
                        for other_row in range(drawing_table.rowCount()):
                            if other_row != row:
                                drawing_table.item(other_row, 1).setCheckState(
                                    Qt.Unchecked
                                )
                    elif (
                        changed_item.column() == 0
                        and use_item.checkState() != Qt.Checked
                    ):
                        primary_item.setCheckState(Qt.Unchecked)
                finally:
                    drawing_table.blockSignals(False)

            drawing_table.itemChanged.connect(drawing_item_changed)
            drawing_layout.addWidget(drawing_table, 1)
            drawing_buttons = QDialogButtonBox(
                QDialogButtonBox.Ok | QDialogButtonBox.Cancel
            )
            drawing_buttons.accepted.connect(drawing_dialog.accept)
            drawing_buttons.rejected.connect(drawing_dialog.reject)
            drawing_layout.addWidget(drawing_buttons)
            if drawing_dialog.exec_() != QDialog.Accepted:
                return
            new_selected = {
                int(drawing_table.item(row, 0).data(Qt.UserRole))
                for row in range(drawing_table.rowCount())
                if drawing_table.item(row, 0).checkState() == Qt.Checked
            }
            new_primary = next((
                int(drawing_table.item(row, 1).data(Qt.UserRole))
                for row in range(drawing_table.rowCount())
                if drawing_table.item(row, 1).checkState() == Qt.Checked
            ), None)
            if new_selected and new_primary not in new_selected:
                new_primary = sorted(new_selected)[0]
            state["selected_drawing_ids"] = new_selected
            state["primary_drawing_id"] = new_primary
            table.item(state["row"], 0).setCheckState(Qt.Checked)
            update_drawing_button(state)

        def populate():
            row_states.clear()
            all_documents = list(self.bom_service.list_pdm_cad_documents() or [])
            try:
                item_documents = list(
                    self.bom_service.list_item_cad_associations(int(item_id)) or []
                )
            except Exception:
                item_documents = []
            item_models = {
                int(document["id"]): document
                for document in item_documents
                if document.get("id") is not None
                and str(document.get("category") or "").upper() != "DRAWING"
            }
            target_item_has_owner = any(
                str(document.get("association_type") or "").upper() == "OWNER"
                for document in item_models.values()
            )
            documents = []
            seen_document_ids = set()
            for document in all_documents:
                if str(document.get("category") or "").upper() == "DRAWING":
                    continue
                cad_id = int(document["id"])
                if cad_id in seen_document_ids:
                    continue
                seen_document_ids.add(cad_id)
                merged = dict(document)
                item_document = item_models.get(cad_id)
                if item_document:
                    merged["related_drawings"] = list(
                        item_document.get("related_drawings") or []
                    )
                documents.append((merged, item_document))
            table.setRowCount(len(documents))
            for row, (document, item_document) in enumerate(documents):
                cad_id = int(document["id"])
                target_association = self._pdm_association_for_item(document, item_id)
                if target_association is None and item_document is not None:
                    target_association = self._pdm_association_for_item(
                        item_document, item_id
                    ) or {
                        "id": item_document.get("association_id"),
                        "association_id": item_document.get("association_id"),
                        "item_id": item_id,
                        "association_type": item_document.get("association_type"),
                    }
                associations = self._pdm_document_associations(document)
                other_associations = []
                for association in associations:
                    try:
                        if int(association.get("item_id")) == int(item_id):
                            continue
                    except Exception:
                        pass
                    other_associations.append(association)
                drawings = list(document.get("related_drawings") or [])
                selected_drawings = self._pdm_selected_drawings(document)
                selected_drawing_ids = {
                    int(drawing["id"])
                    for drawing in selected_drawings if drawing.get("id") is not None
                }
                primary_drawing_id = next((
                    int(drawing["id"])
                    for drawing in selected_drawings
                    if drawing.get("id") is not None
                    and bool(drawing.get("is_primary_drawing"))
                ), None)
                if selected_drawing_ids and primary_drawing_id is None:
                    primary_drawing_id = sorted(selected_drawing_ids)[0]
                check = QTableWidgetItem("")
                check.setFlags(check.flags() | Qt.ItemIsUserCheckable)
                check.setCheckState(Qt.Unchecked)
                check.setData(Qt.UserRole, cad_id)
                check.setData(
                    Qt.UserRole + 1,
                    (target_association or {}).get("association_id")
                    or (target_association or {}).get("id"),
                )
                check.setData(
                    Qt.UserRole + 2,
                    item_id if target_association is not None else None,
                )
                cad_checked_out = document.get("checked_out_by") is not None
                if cad_checked_out:
                    check.setFlags(check.flags() & ~Qt.ItemIsEnabled)
                    check.setToolTip(
                        "Check in or undo this CAD Document before changing its Item association."
                    )
                table.setItem(row, 0, check)
                status_item = QTableWidgetItem(
                    "Associated" if target_association is not None else "Available"
                )
                status_item.setForeground(QBrush(QColor(
                    "#246a46" if target_association is not None else "#526577"
                )))
                table.setItem(row, 1, status_item)
                values = (
                    document.get("file_name"), document.get("name"),
                    document.get("category"), self._pdm_cad_revision_text(document),
                )
                for column, value in enumerate(values, start=2):
                    table.setItem(row, column, QTableWidgetItem(str(value or "")))
                type_combo = QComboBox()
                for kind in association_types:
                    type_combo.addItem(kind.replace("_", " "), kind)
                current_type = str(
                    (target_association or {}).get("association_type") or ""
                ).upper()
                if not current_type:
                    has_owner_elsewhere = any(
                        str(association.get("association_type") or "").upper()
                        == "OWNER"
                        for association in other_associations
                    )
                    current_type = (
                        "IMAGE"
                        if has_owner_elsewhere or target_item_has_owner else "OWNER"
                    )
                type_index = type_combo.findData(current_type)
                type_combo.setCurrentIndex(max(0, type_index))
                type_combo.setEnabled(can_manage and not cad_checked_out)
                table.setCellWidget(row, 6, type_combo)
                drawing_button = QPushButton()
                drawing_button.setObjectName("secondary")
                table.setCellWidget(row, 7, drawing_button)
                other_text = (
                    f"{len(other_associations)} other Item(s)"
                    if other_associations else "None"
                )
                other_item = QTableWidgetItem(other_text)
                other_item.setToolTip("\n".join(
                    self._pdm_item_association_label(association)
                    for association in other_associations
                ) or "This CAD Document has no association with another Item.")
                table.setItem(row, 8, other_item)
                state = {
                    "row": row,
                    "document": document,
                    "association": target_association,
                    "drawings": drawings,
                    "selected_drawing_ids": selected_drawing_ids,
                    "primary_drawing_id": primary_drawing_id,
                    "drawing_button": drawing_button,
                    "type_combo": type_combo,
                    "cad_checked_out": cad_checked_out,
                }
                row_states[cad_id] = state
                update_drawing_button(state)
                drawing_button.clicked.connect(
                    lambda _checked=False, value=cad_id: choose_drawings(value)
                )
                type_combo.currentIndexChanged.connect(
                    lambda _index, row_value=row: table.item(
                        row_value, 0
                    ).setCheckState(Qt.Checked)
                )
                if focus_cad_id is not None and int(focus_cad_id) == cad_id:
                    check.setCheckState(Qt.Checked)
                    table.selectRow(row)
                    table.scrollToItem(check)
            association_status.setText(
                f"{sum(1 for state in row_states.values() if state.get('association'))} "
                "CAD association(s) currently belong to this Item."
            )
            apply_filter()

        def apply_filter(*_args):
            query = str(search.text() or "").strip().casefold()
            for row in range(table.rowCount()):
                text = " ".join(
                    str(table.item(row, column).text() or "")
                    for column in range(1, table.columnCount())
                    if table.item(row, column) is not None
                ).casefold()
                state = row_states.get(int(table.item(row, 0).data(Qt.UserRole)))
                if state:
                    text += " " + str(state["type_combo"].currentText()).casefold()
                    text += " " + str(state["drawing_button"].text()).casefold()
                    text += " " + " ".join(
                        str(drawing.get("file_name") or drawing.get("name") or "")
                        for drawing in state.get("drawings") or []
                    ).casefold()
                table.setRowHidden(row, bool(query and query not in text))

        def apply_associations():
            rows = checked_rows()
            if not rows:
                QMessageBox.information(dialog, "PDM", "Check one or more CAD Documents.")
                return
            errors = []
            updated = 0
            for row in rows:
                cad_id = int(table.item(row, 0).data(Qt.UserRole))
                state = row_states[cad_id]
                try:
                    self.bom_service.associate_cad_document(
                        item_id, cad_id,
                        str(state["type_combo"].currentData()),
                    )
                    set_drawings = getattr(
                        self.bom_service, "set_item_model_drawings", None
                    )
                    if set_drawings is None and (
                        state.get("selected_drawing_ids")
                        or self._pdm_selected_drawings(state["document"])
                    ):
                        raise RuntimeError(
                            "The Item drawing-assignment service is not available."
                        )
                    if set_drawings is not None:
                        set_drawings(
                            int(item_id), cad_id,
                            sorted(state.get("selected_drawing_ids") or set()),
                            primary_drawing_id=state.get("primary_drawing_id"),
                        )
                    updated += 1
                except Exception as exc:
                    errors.append(
                        f"{state['document'].get('file_name') or cad_id}: {exc}"
                    )
            populate()
            if errors:
                QMessageBox.warning(dialog, "Association results", "\n".join(dict.fromkeys(errors)))
            else:
                association_status.setText(
                    f"Updated {updated} CAD association(s) and their Item drawing assignments."
                )

        def remove_associations():
            association_ids = [
                table.item(row, 0).data(Qt.UserRole + 1)
                for row in checked_rows()
                if table.item(row, 0).data(Qt.UserRole + 1) is not None
                and table.item(row, 0).data(Qt.UserRole + 2) is not None
                and int(table.item(row, 0).data(Qt.UserRole + 2)) == int(item_id)
            ]
            if not association_ids:
                QMessageBox.information(
                    dialog, "PDM", "Check CAD Documents associated with this Item first."
                )
                return
            if QMessageBox.question(
                dialog,
                "Remove CAD Associations",
                f"Remove {len(association_ids)} selected CAD association(s) from this Item?\n\n"
                "The CAD Documents remain managed. Associations to other Items are unchanged.",
                QMessageBox.Yes | QMessageBox.Cancel,
                QMessageBox.Cancel,
            ) != QMessageBox.Yes:
                return
            errors = []
            for association_id in association_ids:
                try:
                    self.bom_service.remove_cad_item_association(int(association_id))
                except Exception as exc:
                    errors.append(str(exc))
            populate()
            if errors:
                QMessageBox.warning(
                    dialog, "Association results", "\n".join(dict.fromkeys(errors))
                )
            else:
                association_status.setText(
                    f"Removed {len(association_ids)} CAD association(s) from this Item."
                )

        search.textChanged.connect(apply_filter)
        associate_btn.clicked.connect(apply_associations)
        remove_btn.clicked.connect(remove_associations)
        populate()
        dialog.exec_()
        if getattr(self, "current_part_id", None):
            self._refresh_pdm_context_rows(item_ids=[int(self.current_part_id)])

    def auto_associate_cad_documents(self) -> None:
        if not self.perm.can("manage_parts"):
            QMessageBox.warning(self, "Permission denied", "You cannot change CAD–Item associations.")
            return
        try:
            result = self.bom_service.auto_associate_cad_documents()
        except Exception as exc:
            QMessageBox.critical(self, "Auto Associate", str(exc))
            return
        associated = len(result.get("associated") or [])
        unresolved = list(result.get("unresolved") or [])
        counts = Counter(str(row.get("status") or "UNKNOWN") for row in unresolved)
        unresolved_text = ", ".join(f"{key}: {value}" for key, value in sorted(counts.items()))
        QMessageBox.information(
            self, "Auto Associate",
            f"Associated: {associated}\nUnresolved: {len(unresolved)}"
            + (f"\n{unresolved_text}" if unresolved_text else ""),
        )

    def build_item_structure_from_cad(self) -> None:
        if not self.perm.can("manage_parts"):
            QMessageBox.warning(self, "Permission denied", "You cannot build the Item Structure.")
            return
        item_id = self._pdm_current_item_id()
        if item_id is None:
            return
        if not self._ensure_item_checked_out_for_pdm_change(int(item_id)):
            return
        answer = QMessageBox.question(
            self, "Build Item Structure",
            "Build the multi-level Item Structure from the selected Item's OWNER CAD assembly?\n\n"
            "Manual Item usages will be preserved.",
            QMessageBox.Yes | QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return
        try:
            result = self.bom_service.build_item_structure_from_cad(item_id, True)
        except Exception as exc:
            QMessageBox.critical(self, "Build Item Structure", str(exc))
            return
        QMessageBox.information(
            self, "Build Item Structure",
            "Build completed.\n"
            f"Created: {result.get('created', 0)}\n"
            f"Updated: {result.get('updated', 0)}\n"
            f"Removed: {result.get('removed', 0)}\n"
            f"No related Item: {result.get('no_related_item', 0)}\n"
            f"Excluded: {result.get('excluded', 0)}\n"
            f"Conflicts: {result.get('conflicts', 0)}",
        )
        self.bom_mode_selector.setCurrentIndex(
            self.bom_mode_selector.findData("ebom")
        )
        self._refresh_loaded_part_branch(int(item_id))
        self.display_details(int(item_id))

    @staticmethod
    def _find_payload_node_by_id(nodes, wanted_id: int, key: str = "id") -> dict | None:
        for node in nodes or []:
            try:
                candidate = node.get(key)
                if candidate is None and key == "bom_id":
                    candidate = node.get("id")
                if candidate is None and key == "id":
                    candidate = node.get("bom_id")
                if candidate is not None and int(candidate) == int(wanted_id):
                    return node
            except Exception:
                pass
            found = BomPage._find_payload_node_by_id(
                node.get("children") or [], int(wanted_id), key
            )
            if found is not None:
                return found
        return None

    @staticmethod
    def _cad_label(payload: dict) -> str:
        return str(
            payload.get("file_name")
            or payload.get("name")
            or payload.get("id")
            or "CAD Document"
        )

    def _cad_payload_paths_to_document(self, cad_id: int) -> list[list[dict]]:
        wanted = int(cad_id)
        paths: list[list[dict]] = []

        def walk(node: dict, path: list[dict]) -> None:
            current = [*path, dict(node or {})]
            try:
                if int((node or {}).get("id")) == wanted:
                    paths.append(current)
            except Exception:
                pass
            for child in (node or {}).get("children") or []:
                walk(child, current)

        roots = list(getattr(self, "_pdm_cad_roots", []) or [])
        if not roots:
            try:
                roots = list(
                    (self.bom_service.get_pdm_cad_structure() or {}).get("roots") or []
                )
                self._pdm_cad_roots = roots
            except Exception:
                roots = []
        for root in roots:
            walk(root, [])
        return paths

    def _item_payload_by_id(self, item_id: int) -> dict:
        roots = list(getattr(self, "_pdm_ebom_roots", []) or [])
        if not roots:
            try:
                roots = list(
                    (self.bom_service.get_released_ebom_project(
                        int(self.session.project_id)
                    ) or {}).get("roots") or []
                )
                self._pdm_ebom_roots = roots
            except Exception:
                roots = []
        return self._find_payload_node_by_id(roots, int(item_id), "bom_id") or (
            self.bom_service.get_part_details(int(item_id)) or {"id": int(item_id)}
        )

    def _relationship_graph_for_item(
        self, item_id: int, depth: int = 1, path_limit: int | None = None,
    ) -> dict:
        item_id = int(item_id)
        path_limit = int(path_limit or self.RELATIONSHIP_GRAPH_PATH_LIMIT)
        nodes = {}
        edges = []
        hidden_parent_paths = defaultdict(int)

        def add_node(
            node_id, lane, label, level, node_type="normal",
            object_kind=None, object_id=None,
        ):
            key = str(node_id)
            if key not in nodes:
                nodes[key] = {
                    "id": key, "lane": lane, "label": str(label),
                    "level": int(level), "type": node_type,
                    "object_kind": object_kind, "object_id": object_id,
                }
            else:
                nodes[key]["level"] = min(int(nodes[key].get("level") or 0), int(level))
                if node_type == "focus":
                    nodes[key]["type"] = "focus"
            return key

        def add_edge(source, target, label="", kind="structure"):
            edge = {
                "source": str(source), "target": str(target),
                "label": str(label or ""), "kind": str(kind or "structure"),
            }
            if edge not in edges:
                edges.append(edge)

        item_root = self._item_payload_by_id(item_id)
        focus_key = add_node(
            f"item:{item_id}", "item", self._windchill_item_label(item_root),
            1, "focus", "item", item_id,
        )

        def add_item_children(parent_payload: dict, parent_key: str, parent_level: int, remaining: int):
            if remaining <= 0:
                return
            for child in list(parent_payload.get("children") or [])[:30]:
                child_id = child.get("bom_id") or child.get("id")
                if child_id is None:
                    continue
                child_key = add_node(
                    f"item:{int(child_id)}", "item",
                    self._windchill_item_label(child), parent_level + 1,
                    object_kind="item", object_id=int(child_id),
                )
                add_edge(parent_key, child_key, "", "structure")
                add_item_children(child, child_key, parent_level + 1, remaining - 1)

        add_item_children(item_root, focus_key, 1, max(1, int(depth)))

        try:
            cad_documents = [
                row for row in (self.bom_service.list_item_cad_associations(item_id) or [])
                if str(row.get("category") or "").upper() != "DRAWING"
            ]
        except Exception:
            cad_documents = []
        for document in cad_documents[:16]:
            cad_id = document.get("id")
            if cad_id is None:
                continue
            cad_id = int(cad_id)
            paths = self._cad_payload_paths_to_document(cad_id) or [[document]]
            hidden_parent_paths[cad_id] += max(0, len(paths) - path_limit)
            for path in paths[:path_limit]:
                previous_key = None
                for level, cad_payload in enumerate(path):
                    node_cad_id = cad_payload.get("id")
                    if node_cad_id is None:
                        continue
                    key = add_node(
                        f"cad:{int(node_cad_id)}:{level}", "cad",
                        self._cad_payload_identity(cad_payload), level,
                        "focus" if int(node_cad_id) == cad_id else "normal",
                        "cad", int(node_cad_id),
                    )
                    if previous_key is not None:
                        add_edge(previous_key, key, "", "structure")
                    previous_key = key
                    if int(node_cad_id) == cad_id:
                        role = self._cad_representation_role(document)
                        add_edge(key, focus_key, role, "association")
                        for drawing in self._pdm_selected_drawings(document)[:4]:
                            drawing_id = drawing.get("id")
                            if drawing_id is None:
                                continue
                            drawing_key = add_node(
                                f"cad-drawing:{int(drawing_id)}", "cad",
                                str(drawing.get("file_name") or drawing.get("name") or "DRW"),
                                level + 1, "drawing", "cad", int(drawing_id),
                            )
                            add_edge(key, drawing_key, "DRW", "association")
                        break
        hidden = max(0, len(cad_documents) - 16)
        return self._normalize_relationship_graph_cad_nodes({
            "nodes": list(nodes.values()), "edges": edges, "hidden": hidden,
            "hidden_parent_paths": dict(hidden_parent_paths),
            "path_limit": path_limit,
        })

    def _relationship_graph_for_cad(
        self, cad_id: int, payload: dict | None = None, path_limit: int | None = None,
    ) -> dict:
        cad_id = int(cad_id)
        path_limit = int(path_limit or self.RELATIONSHIP_GRAPH_PATH_LIMIT)
        payload = dict(payload or {})
        document = self.bom_service.pdm_service.repo.get_cad_document(cad_id) or {}
        document = dict(document or {})
        document.update(payload)
        paths = self._cad_payload_paths_to_document(cad_id) or [[document]]
        terminal_payload = dict(paths[0][-1] if paths and paths[0] else document)
        terminal_payload.update(document)

        associations = self._pdm_document_associations(terminal_payload)
        if not associations:
            try:
                associations = list(
                    self.bom_service.list_cad_item_associations(cad_id) or []
                )
            except Exception:
                associations = []
        nodes = {}
        edges = []

        def add_node(
            node_id, lane, label, level, node_type="normal",
            object_kind=None, object_id=None,
        ):
            key = str(node_id)
            if key not in nodes:
                nodes[key] = {
                    "id": key, "lane": lane, "label": str(label),
                    "level": int(level), "type": node_type,
                    "object_kind": object_kind, "object_id": object_id,
                }
            else:
                nodes[key]["level"] = min(int(nodes[key].get("level") or 0), int(level))
                if node_type == "focus":
                    nodes[key]["type"] = "focus"
            return key

        def add_edge(source, target, label="", kind="structure"):
            edge = {
                "source": str(source), "target": str(target),
                "label": str(label or ""), "kind": str(kind or "structure"),
            }
            if edge not in edges:
                edges.append(edge)

        selected_cad_keys = []
        hidden_parent_paths = {
            cad_id: max(0, len(paths) - path_limit)
        }
        for path_index, path in enumerate(paths[:path_limit]):
            previous_key = None
            for level, cad_payload in enumerate(path):
                node_cad_id = cad_payload.get("id")
                if node_cad_id is None:
                    continue
                is_focus = int(node_cad_id) == cad_id
                key = add_node(
                    f"cad:{int(node_cad_id)}:{path_index}:{level}", "cad",
                    self._cad_payload_identity(cad_payload), level,
                    "focus" if is_focus else "normal", "cad", int(node_cad_id),
                )
                if previous_key is not None:
                    add_edge(previous_key, key, "", "structure")
                previous_key = key
                if is_focus:
                    selected_cad_keys.append(key)
                    break
        if not selected_cad_keys:
            selected_cad_keys.append(
                add_node(
                    f"cad:{cad_id}:0:0", "cad",
                    self._cad_payload_identity(terminal_payload), 0, "focus",
                    "cad", cad_id,
                )
            )

        seen_item_ids = set()
        for association in associations[:24]:
            try:
                item_id = int(association.get("item_id"))
            except Exception:
                continue
            if item_id in seen_item_ids:
                continue
            seen_item_ids.add(item_id)
            item_payload = self._item_payload_by_id(item_id)
            item_key = add_node(
                f"item:{item_id}", "item",
                self._windchill_item_label(item_payload), 1, "focus",
                "item", item_id,
            )
            role = str(
                association.get("association_type")
                or terminal_payload.get("association_type") or "CONTENT"
            ).upper().replace("_", " ")
            for cad_key in selected_cad_keys:
                add_edge(cad_key, item_key, role, "association")
            for child in list(item_payload.get("children") or [])[:8]:
                child_id = child.get("bom_id") or child.get("id")
                if child_id is None:
                    continue
                child_key = add_node(
                    f"item:{int(child_id)}", "item",
                    self._windchill_item_label(child), 2,
                    object_kind="item", object_id=int(child_id),
                )
                add_edge(item_key, child_key, "", "structure")

        drawings = self._pdm_selected_drawings(terminal_payload) or list(
            terminal_payload.get("related_drawings") or []
        )
        for drawing in drawings[:8]:
            drawing_id = drawing.get("id")
            if drawing_id is None:
                continue
            drawing_key = add_node(
                f"cad-drawing:{int(drawing_id)}", "cad",
                str(drawing.get("file_name") or drawing.get("name") or "DRW"),
                3, "drawing", "cad", int(drawing_id),
            )
            for cad_key in selected_cad_keys:
                add_edge(cad_key, drawing_key, "DRW", "association")

        hidden = max(0, len(associations) - 24)
        return self._normalize_relationship_graph_cad_nodes({
            "nodes": list(nodes.values()), "edges": edges, "hidden": hidden,
            "hidden_parent_paths": hidden_parent_paths,
            "path_limit": path_limit,
        })

    def _normalize_relationship_graph_cad_nodes(self, graph: dict) -> dict:
        """Display each CAD Document once while preserving all occurrence links."""
        graph = {
            "nodes": [dict(node) for node in (graph.get("nodes") or [])],
            "edges": [dict(edge) for edge in (graph.get("edges") or [])],
            "hidden": graph.get("hidden", 0),
            "hidden_parent_paths": dict(graph.get("hidden_parent_paths") or {}),
            "path_limit": graph.get("path_limit", self.RELATIONSHIP_GRAPH_PATH_LIMIT),
        }
        canonical_by_cad_id = {}
        redirect = {}
        merged_nodes = []

        def should_replace_canonical(current: dict, candidate: dict) -> bool:
            current_type = str(current.get("type") or "")
            candidate_type = str(candidate.get("type") or "")
            if candidate_type == "focus" and current_type != "focus":
                return True
            if current_type == "drawing" and candidate_type != "drawing":
                return True
            return False

        for node in graph["nodes"]:
            node_id = str(node.get("id"))
            try:
                is_cad_document = (
                    str(node.get("object_kind") or "").lower() == "cad"
                    and node.get("object_id") is not None
                )
                cad_object_id = int(node.get("object_id")) if is_cad_document else None
            except Exception:
                is_cad_document = False
                cad_object_id = None
            if not is_cad_document:
                merged_nodes.append(node)
                continue

            canonical = canonical_by_cad_id.get(cad_object_id)
            if canonical is None:
                canonical_by_cad_id[cad_object_id] = node
                merged_nodes.append(node)
                continue

            if should_replace_canonical(canonical, node):
                canonical_id = str(canonical.get("id"))
                redirect[canonical_id] = node_id
                canonical_by_cad_id[cad_object_id] = node
                for index, existing in enumerate(merged_nodes):
                    if str(existing.get("id")) == canonical_id:
                        merged_nodes[index] = node
                        break
            else:
                redirect[node_id] = str(canonical.get("id"))
                canonical["level"] = min(
                    int(canonical.get("level") or 0),
                    int(node.get("level") or 0),
                )
                if node.get("type") == "focus":
                    canonical["type"] = "focus"

        normalized_edges = []
        seen_edges = set()
        for edge in graph["edges"]:
            source = str(edge.get("source"))
            target = str(edge.get("target"))
            while source in redirect and redirect[source] != source:
                source = redirect[source]
            while target in redirect and redirect[target] != target:
                target = redirect[target]
            if source == target:
                continue
            edge_key = (
                source, target,
                str(edge.get("kind") or "structure"),
                str(edge.get("label") or ""),
            )
            if edge_key in seen_edges:
                continue
            normalized_edge = dict(edge)
            normalized_edge["source"] = source
            normalized_edge["target"] = target
            normalized_edges.append(normalized_edge)
            seen_edges.add(edge_key)

        graph["nodes"] = merged_nodes
        graph["edges"] = normalized_edges
        return graph

    def _expand_relationship_graph_node(
        self, graph: dict, selected_node: dict, categories: set[str] | None = None,
    ) -> tuple[dict, int]:
        """Add the selected CAD/EBOM node's local structure, associations, and drawings."""
        categories = set(categories or {"structure", "association", "drawing"})
        graph = {
            "nodes": [dict(node) for node in (graph.get("nodes") or [])],
            "edges": [dict(edge) for edge in (graph.get("edges") or [])],
            "hidden": graph.get("hidden", 0),
            "hidden_parent_paths": dict(graph.get("hidden_parent_paths") or {}),
            "path_limit": graph.get("path_limit", self.RELATIONSHIP_GRAPH_PATH_LIMIT),
        }
        selected_node = dict(selected_node or {})
        selected_graph_id = str(selected_node.get("id") or "")
        object_kind = str(selected_node.get("object_kind") or "").lower()
        object_id = selected_node.get("object_id")
        if not selected_graph_id or object_id is None:
            return graph, 0

        nodes_by_id = {str(node.get("id")): node for node in graph["nodes"]}
        edge_keys = {
            (
                str(edge.get("source")), str(edge.get("target")),
                str(edge.get("kind") or "structure"), str(edge.get("label") or ""),
            )
            for edge in graph["edges"]
        }
        added = 0

        def add_node(node_id, lane, label, level, node_type="normal", object_kind=None, object_id=None):
            nonlocal added
            key = str(node_id)
            if key not in nodes_by_id:
                node = {
                    "id": key, "lane": lane, "label": str(label),
                    "level": int(level or 0), "type": node_type,
                    "object_kind": object_kind, "object_id": object_id,
                }
                graph["nodes"].append(node)
                nodes_by_id[key] = node
                added += 1
            return key

        def add_edge(source, target, label="", kind="structure"):
            nonlocal added
            key = (str(source), str(target), str(kind or "structure"), str(label or ""))
            if key not in edge_keys:
                graph["edges"].append({
                    "source": str(source), "target": str(target),
                    "label": str(label or ""), "kind": str(kind or "structure"),
                })
                edge_keys.add(key)
                added += 1

        def existing_object_node(kind: str, oid: int):
            for node in graph["nodes"]:
                try:
                    if (
                        str(node.get("object_kind") or "").lower() == kind
                        and int(node.get("object_id")) == int(oid)
                    ):
                        return str(node.get("id"))
                except Exception:
                    continue
            return None

        def add_cad_drawings(cad_key: str, cad_payload: dict, cad_level: int) -> None:
            drawings = self._pdm_selected_drawings(cad_payload) or list(
                cad_payload.get("related_drawings") or []
            )
            for drawing in drawings[:12]:
                drawing_id = drawing.get("id")
                if drawing_id is None:
                    continue
                drawing_key = existing_object_node("cad", int(drawing_id)) or add_node(
                    f"cad-drawing:{int(drawing_id)}", "cad",
                    str(drawing.get("file_name") or drawing.get("name") or "DRW"),
                    cad_level + 1, "drawing", "cad", int(drawing_id),
                )
                add_edge(cad_key, drawing_key, "DRW", "association")

        def add_cad_occurrence_paths(cad_id: int, document_payload: dict, stem: str) -> list[str]:
            terminal_keys = []
            try:
                paths = self._cad_payload_paths_to_document(int(cad_id)) or []
            except Exception:
                paths = []
            if not paths:
                paths = [[document_payload]]
            path_limit = int(graph.get("path_limit") or self.RELATIONSHIP_GRAPH_PATH_LIMIT)
            graph["hidden_parent_paths"][int(cad_id)] = max(0, len(paths) - path_limit)
            for path_index, path in enumerate(paths[:path_limit]):
                previous_key = None
                terminal_key = None
                for level, cad_payload in enumerate(path):
                    node_cad_id = cad_payload.get("id")
                    if node_cad_id is None:
                        continue
                    node_cad_id = int(node_cad_id)
                    key = existing_object_node("cad", node_cad_id) or add_node(
                        f"cad:{node_cad_id}:{stem}:{path_index}:{level}", "cad",
                        self._cad_payload_identity(cad_payload), level,
                        "normal", "cad", node_cad_id,
                    )
                    if previous_key is not None:
                        add_edge(previous_key, key, "", "structure")
                    previous_key = key
                    if node_cad_id == int(cad_id):
                        terminal_key = key
                        break
                if terminal_key:
                    terminal_keys.append(terminal_key)
            if not terminal_keys:
                terminal_keys.append(
                    existing_object_node("cad", int(cad_id)) or add_node(
                        f"cad:{int(cad_id)}:{stem}:terminal", "cad",
                        self._cad_payload_identity(document_payload),
                        int(selected_node.get("level") or 0),
                        "normal", "cad", int(cad_id),
                    )
                )
            return terminal_keys

        def merge_duplicate_cad_nodes() -> None:
            canonical_by_cad_id = {}
            redirect = {}
            merged_nodes = []
            for node in graph["nodes"]:
                node_id = str(node.get("id"))
                try:
                    is_cad_document = (
                        str(node.get("object_kind") or "").lower() == "cad"
                        and node.get("object_id") is not None
                    )
                    cad_object_id = int(node.get("object_id")) if is_cad_document else None
                except Exception:
                    is_cad_document = False
                    cad_object_id = None
                if not is_cad_document:
                    merged_nodes.append(node)
                    continue
                canonical = canonical_by_cad_id.get(cad_object_id)
                if canonical is None:
                    canonical_by_cad_id[cad_object_id] = node
                    merged_nodes.append(node)
                    continue
                canonical_id = str(canonical.get("id"))
                redirect[node_id] = canonical_id
                canonical["level"] = min(
                    int(canonical.get("level") or 0),
                    int(node.get("level") or 0),
                )
                if node.get("type") == "focus":
                    canonical["type"] = "focus"

            normalized_edges = []
            seen_edges = set()
            for edge in graph["edges"]:
                source = redirect.get(str(edge.get("source")), str(edge.get("source")))
                target = redirect.get(str(edge.get("target")), str(edge.get("target")))
                if source == target:
                    continue
                edge_key = (
                    source, target,
                    str(edge.get("kind") or "structure"),
                    str(edge.get("label") or ""),
                )
                if edge_key in seen_edges:
                    continue
                normalized_edge = dict(edge)
                normalized_edge["source"] = source
                normalized_edge["target"] = target
                normalized_edges.append(normalized_edge)
                seen_edges.add(edge_key)
            graph["nodes"] = merged_nodes
            graph["edges"] = normalized_edges

        if object_kind == "item":
            try:
                item_payload = self._item_payload_by_id(int(object_id))
            except Exception:
                item_payload = {}
            parent_level = int(selected_node.get("level") or 1)
            if "structure" in categories:
                for child in list(item_payload.get("children") or [])[:30]:
                    child_id = child.get("bom_id") or child.get("id")
                    if child_id is None:
                        continue
                    child_key = add_node(
                        f"item:{int(child_id)}", "item",
                        self._windchill_item_label(child), parent_level + 1,
                        "normal", "item", int(child_id),
                    )
                    add_edge(selected_graph_id, child_key, "", "structure")
            if {"association", "drawing"} & categories:
                try:
                    cad_documents = [
                        row for row in (self.bom_service.list_item_cad_associations(int(object_id)) or [])
                        if str(row.get("category") or "").upper() != "DRAWING"
                    ]
                except Exception:
                    cad_documents = []
                for document in cad_documents[:24]:
                    cad_id = document.get("id")
                    if cad_id is None:
                        continue
                    cad_id = int(cad_id)
                    document_payload = dict(
                        self.bom_service.pdm_service.repo.get_cad_document(cad_id) or {}
                    )
                    document_payload.update(document)
                    cad_keys = add_cad_occurrence_paths(
                        cad_id, document_payload,
                        f"assoc:{selected_graph_id}",
                    )
                    role = self._cad_representation_role(document_payload)
                    for cad_key in cad_keys:
                        if "association" in categories:
                            add_edge(cad_key, selected_graph_id, role, "association")
                        if "drawing" in categories:
                            cad_node = nodes_by_id.get(str(cad_key), {})
                            add_cad_drawings(
                                cad_key, document_payload,
                                int(cad_node.get("level") if cad_node else parent_level),
                            )
        elif object_kind == "cad":
            payload = self._find_payload_node_by_id(
                getattr(self, "_pdm_cad_roots", []) or [], int(object_id), "id"
            ) or self.bom_service.pdm_service.repo.get_cad_document(int(object_id)) or {}
            parent_level = int(selected_node.get("level") or 0)
            if "structure" in categories:
                for child in list(payload.get("children") or [])[:30]:
                    child_id = child.get("id")
                    if child_id is None:
                        continue
                    child_key = add_node(
                        f"cad:{int(child_id)}:expanded:{selected_graph_id}", "cad",
                        self._cad_payload_identity(child), parent_level + 1,
                        "normal", "cad", int(child_id),
                    )
                    add_edge(selected_graph_id, child_key, "", "structure")
            document_payload = dict(
                self.bom_service.pdm_service.repo.get_cad_document(int(object_id)) or {}
            )
            document_payload.update(payload)
            if "drawing" in categories:
                add_cad_drawings(selected_graph_id, document_payload, parent_level)
            if "association" in categories:
                try:
                    associations = list(
                        self.bom_service.list_cad_item_associations(int(object_id)) or []
                    )
                except Exception:
                    associations = []
                for association in associations[:24]:
                    try:
                        item_id = int(association.get("item_id"))
                    except Exception:
                        continue
                    item_payload = self._item_payload_by_id(item_id)
                    item_key = existing_object_node("item", item_id) or add_node(
                        f"item:{item_id}", "item",
                        self._windchill_item_label(item_payload), parent_level + 1,
                        "normal", "item", item_id,
                    )
                    role = str(
                        association.get("association_type")
                        or document_payload.get("association_type") or "CONTENT"
                    ).upper().replace("_", " ")
                    add_edge(selected_graph_id, item_key, role, "association")
        merge_duplicate_cad_nodes()
        return self._normalize_relationship_graph_cad_nodes(graph), added

    def _collapse_relationship_graph_node(
        self, graph: dict, selected_node: dict, categories: set[str],
    ) -> tuple[dict, int]:
        """Remove selected relationship layers around one graph node."""
        graph = {
            "nodes": [dict(node) for node in (graph.get("nodes") or [])],
            "edges": [dict(edge) for edge in (graph.get("edges") or [])],
            "hidden": graph.get("hidden", 0),
            "hidden_parent_paths": dict(graph.get("hidden_parent_paths") or {}),
            "path_limit": graph.get("path_limit", self.RELATIONSHIP_GRAPH_PATH_LIMIT),
        }
        selected_id = str((selected_node or {}).get("id") or "")
        if not selected_id:
            return graph, 0
        nodes_by_id = {str(node.get("id")): node for node in graph["nodes"]}

        def edge_category(edge: dict) -> str:
            target_node = nodes_by_id.get(str(edge.get("target")), {})
            label = str(edge.get("label") or "").upper()
            kind = str(edge.get("kind") or "").lower()
            if label == "DRW" or target_node.get("type") == "drawing":
                return "drawing"
            if kind == "association":
                return "association"
            return "structure"

        remove_edges = set()
        if "structure" in categories:
            queue = deque([selected_id])
            visited = set()
            while queue:
                current = queue.popleft()
                if current in visited:
                    continue
                visited.add(current)
                for index, edge in enumerate(graph["edges"]):
                    if edge_category(edge) != "structure":
                        continue
                    if str(edge.get("source")) == current:
                        remove_edges.add(index)
                        queue.append(str(edge.get("target")))
        selected_kind = str((selected_node or {}).get("object_kind") or "").lower()
        if selected_kind == "item" and "drawing" in categories:
            related_cad_node_ids = {
                str(edge.get("source"))
                for edge in graph["edges"]
                if edge_category(edge) == "association"
                and str(edge.get("target")) == selected_id
            }
            if not related_cad_node_ids:
                try:
                    related_cad_ids = {
                        int(row.get("id"))
                        for row in (
                            self.bom_service.list_item_cad_associations(
                                int((selected_node or {}).get("object_id"))
                            ) or []
                        )
                        if row.get("id") is not None
                    }
                except Exception:
                    related_cad_ids = set()
                related_cad_node_ids = {
                    str(node.get("id"))
                    for node in graph["nodes"]
                    if str(node.get("object_kind") or "").lower() == "cad"
                    and node.get("object_id") is not None
                    and int(node.get("object_id")) in related_cad_ids
                }
            for index, edge in enumerate(graph["edges"]):
                if (
                    edge_category(edge) == "drawing"
                    and str(edge.get("source")) in related_cad_node_ids
                ):
                    remove_edges.add(index)
        for index, edge in enumerate(graph["edges"]):
            category = edge_category(edge)
            if category not in categories or category == "structure":
                continue
            if str(edge.get("source")) == selected_id or str(edge.get("target")) == selected_id:
                remove_edges.add(index)

        if not remove_edges:
            return graph, 0

        graph["edges"] = [
            edge for index, edge in enumerate(graph["edges"])
            if index not in remove_edges
        ]

        keep_ids = {selected_id}
        for node in graph["nodes"]:
            if str(node.get("type") or "") == "focus":
                keep_ids.add(str(node.get("id")))

        changed = True
        while changed:
            changed = False
            connected_ids = set()
            for edge in graph["edges"]:
                connected_ids.add(str(edge.get("source")))
                connected_ids.add(str(edge.get("target")))
            removable_ids = {
                str(node.get("id"))
                for node in graph["nodes"]
                if str(node.get("id")) not in keep_ids
                and str(node.get("id")) not in connected_ids
            }
            if removable_ids:
                graph["nodes"] = [
                    node for node in graph["nodes"]
                    if str(node.get("id")) not in removable_ids
                ]
                changed = True

        return graph, len(remove_edges)

    def show_relationship_graph(
        self, item_id: int | None = None, cad_id: int | None = None,
        payload: dict | None = None,
    ) -> None:
        try:
            if item_id is not None:
                graph = self._relationship_graph_for_item(int(item_id), depth=1)
                title = "Relationship Graph - EBOM Item"
            elif cad_id is not None:
                graph = self._relationship_graph_for_cad(int(cad_id), payload)
                title = "Relationship Graph - CAD Document"
            else:
                return
        except Exception as exc:
            QMessageBox.warning(self, "Relationship Graph", str(exc))
            return
        if not graph.get("nodes"):
            QMessageBox.information(self, "Relationship Graph", "No relationship graph data is available.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.resize(980, 650)
        layout = QVBoxLayout(dialog)
        toolbar = QHBoxLayout()
        info = QLabel(
            f"Focused graph: {len(graph.get('nodes') or [])} node(s), "
            f"{len(graph.get('edges') or [])} link(s)"
            + (f", {graph.get('hidden')} CAD representation(s) hidden" if graph.get("hidden") else "")
        )
        zoom_out = QPushButton("-")
        zoom_reset = QPushButton("100%")
        zoom_in = QPushButton("+")
        expand_button = QPushButton("Expand")
        isolate_button = QPushButton("Isolate")
        show_all_parents_button = QPushButton("Show All Parents")
        structure_filter = QCheckBox("Structure")
        association_filter = QCheckBox("Associations")
        drawing_filter = QCheckBox("Drawings")
        for checkbox in (structure_filter, association_filter, drawing_filter):
            checkbox.setChecked(True)
            checkbox.setToolTip("Show or hide this relationship layer in the graph.")
        expand_button.setEnabled(False)
        isolate_button.setEnabled(False)
        show_all_parents_button.setEnabled(False)
        show_all_parents_button.setToolTip(
            "Show every CAD parent occurrence for the selected CAD Document."
        )
        for button in (
            expand_button, isolate_button, show_all_parents_button,
            zoom_out, zoom_reset, zoom_in,
        ):
            button.setFixedHeight(24)
        toolbar.addWidget(info)
        toolbar.addStretch(1)
        toolbar.addWidget(structure_filter)
        toolbar.addWidget(association_filter)
        toolbar.addWidget(drawing_filter)
        toolbar.addSpacing(10)
        toolbar.addWidget(expand_button)
        toolbar.addWidget(isolate_button)
        toolbar.addWidget(show_all_parents_button)
        toolbar.addWidget(zoom_out)
        toolbar.addWidget(zoom_reset)
        toolbar.addWidget(zoom_in)
        layout.addLayout(toolbar)
        canvas = _RelationshipGraphCanvas(graph)
        scroll = QScrollArea()
        scroll.setWidget(canvas)
        scroll.setWidgetResizable(False)
        layout.addWidget(scroll, 1)
        scale_state = {"value": 1.0}
        state = {"graph": graph, "selected": {}}

        def graph_caption(current_graph: dict) -> str:
            hidden_parent_total = sum(
                int(value or 0)
                for value in (current_graph.get("hidden_parent_paths") or {}).values()
            )
            return (
                f"Focused graph: {len(current_graph.get('nodes') or [])} node(s), "
                f"{len(current_graph.get('edges') or [])} link(s)"
                + (
                    f", {current_graph.get('hidden')} hidden"
                    if current_graph.get("hidden") else ""
                )
                + (
                    f", {hidden_parent_total} parent occurrence(s) hidden"
                    if hidden_parent_total else ""
                )
            )

        def set_graph(current_graph: dict, keep_positions: bool = False) -> None:
            state["graph"] = current_graph
            state["selected"] = {}
            canvas.graph = current_graph
            if not keep_positions:
                canvas.manual_positions.clear()
            canvas.selected_node_id = None
            canvas._layout_graph()
            canvas.update()
            info.setText(graph_caption(current_graph))
            expand_button.setEnabled(False)
            isolate_button.setEnabled(False)
            show_all_parents_button.setEnabled(False)

        def apply_graph_filters() -> None:
            canvas.set_edge_filter("structure", structure_filter.isChecked())
            canvas.set_edge_filter("association", association_filter.isChecked())
            canvas.set_edge_filter("drawing", drawing_filter.isChecked())
            visible_nodes = canvas._visible_nodes()
            visible_edges = canvas._visible_edges()
            info.setText(
                f"Visible graph: {len(visible_nodes)} node(s), "
                f"{len(visible_edges)} link(s)"
            )

        def on_node_selected(node: dict) -> None:
            state["selected"] = dict(node or {})
            can_navigate = bool(
                state["selected"].get("object_kind")
                and state["selected"].get("object_id") is not None
            )
            expand_button.setEnabled(can_navigate)
            isolate_button.setEnabled(can_navigate)
            show_all_parents_button.setEnabled(False)
            try:
                if str(state["selected"].get("object_kind") or "").lower() == "cad":
                    cad_object_id = int(state["selected"].get("object_id"))
                    hidden_parent_paths = state["graph"].get("hidden_parent_paths") or {}
                    hidden_count = int(
                        hidden_parent_paths.get(cad_object_id)
                        or hidden_parent_paths.get(str(cad_object_id))
                        or 0
                    )
                    show_all_parents_button.setEnabled(hidden_count > 0)
                    if hidden_count > 0:
                        show_all_parents_button.setText(
                            f"Show All Parents (+{hidden_count})"
                        )
                    else:
                        show_all_parents_button.setText("Show All Parents")
            except Exception:
                show_all_parents_button.setText("Show All Parents")

        def expand_selected() -> None:
            expand_selected_categories({"structure", "association", "drawing"})

        def expand_selected_categories(categories: set[str]) -> None:
            selected = dict(state.get("selected") or {})
            if not selected:
                return
            expanded_graph, added = self._expand_relationship_graph_node(
                state["graph"], selected, categories=categories,
            )
            set_graph(expanded_graph, keep_positions=True)
            if not added:
                QMessageBox.information(
                    dialog, "Relationship Graph",
                    "No additional relationship was found for the selected node.",
                )

        def collapse_selected_categories(categories: set[str]) -> None:
            selected = dict(state.get("selected") or {})
            if not selected:
                return
            collapsed_graph, removed = self._collapse_relationship_graph_node(
                state["graph"], selected, categories,
            )
            set_graph(collapsed_graph, keep_positions=True)
            if not removed:
                QMessageBox.information(
                    dialog, "Relationship Graph",
                    "No visible relationship of that type was found on the selected node.",
                )

        def isolate_selected() -> None:
            selected = dict(state.get("selected") or {})
            object_kind = str(selected.get("object_kind") or "").lower()
            object_id = selected.get("object_id")
            if object_id is None:
                return
            try:
                if object_kind == "item":
                    set_graph(self._relationship_graph_for_item(int(object_id), depth=1))
                elif object_kind == "cad":
                    set_graph(self._relationship_graph_for_cad(int(object_id)))
            except Exception as exc:
                QMessageBox.warning(dialog, "Relationship Graph", str(exc))

        def show_all_parents_selected() -> None:
            selected = dict(state.get("selected") or {})
            if str(selected.get("object_kind") or "").lower() != "cad":
                return
            object_id = selected.get("object_id")
            if object_id is None:
                return
            try:
                set_graph(
                    self._relationship_graph_for_cad(
                        int(object_id),
                        path_limit=self.RELATIONSHIP_GRAPH_ALL_PATH_LIMIT,
                    )
                )
            except Exception as exc:
                QMessageBox.warning(dialog, "Relationship Graph", str(exc))

        def show_node_context_menu(node: dict, global_pos) -> None:
            state["selected"] = dict(node or {})
            on_node_selected(state["selected"])
            if not state["selected"]:
                return
            menu = QMenu(dialog)
            isolate_action = menu.addAction("Isolate")
            menu.addSeparator()
            expand_all_action = menu.addAction("Expand All")
            expand_menu = menu.addMenu("Expand")
            expand_structure_action = expand_menu.addAction("Structure")
            expand_association_action = expand_menu.addAction("Associations")
            expand_drawing_action = expand_menu.addAction("Drawings / DRW")
            collapse_all_action = menu.addAction("Collapse All")
            collapse_menu = menu.addMenu("Collapse")
            collapse_structure_action = collapse_menu.addAction("Structure")
            collapse_association_action = collapse_menu.addAction("Associations")
            collapse_drawing_action = collapse_menu.addAction("Drawings / DRW")
            selected_action = menu.exec_(global_pos)
            if selected_action == isolate_action:
                isolate_selected()
            elif selected_action == expand_all_action:
                expand_selected_categories({"structure", "association", "drawing"})
            elif selected_action == expand_structure_action:
                expand_selected_categories({"structure"})
            elif selected_action == expand_association_action:
                expand_selected_categories({"association"})
            elif selected_action == expand_drawing_action:
                expand_selected_categories({"drawing"})
            elif selected_action == collapse_all_action:
                collapse_selected_categories({"structure", "association", "drawing"})
            elif selected_action == collapse_structure_action:
                collapse_selected_categories({"structure"})
            elif selected_action == collapse_association_action:
                collapse_selected_categories({"association"})
            elif selected_action == collapse_drawing_action:
                collapse_selected_categories({"drawing"})

        canvas.nodeSelected.connect(on_node_selected)
        canvas.nodeContextMenuRequested.connect(show_node_context_menu)
        expand_button.clicked.connect(expand_selected)
        isolate_button.clicked.connect(isolate_selected)
        show_all_parents_button.clicked.connect(show_all_parents_selected)
        structure_filter.toggled.connect(lambda _checked: apply_graph_filters())
        association_filter.toggled.connect(lambda _checked: apply_graph_filters())
        drawing_filter.toggled.connect(lambda _checked: apply_graph_filters())
        zoom_out.clicked.connect(
            lambda: (
                scale_state.update(value=scale_state["value"] - 0.1),
                canvas.set_graph_scale(scale_state["value"]),
            )
        )
        zoom_in.clicked.connect(
            lambda: (
                scale_state.update(value=scale_state["value"] + 0.1),
                canvas.set_graph_scale(scale_state["value"]),
            )
        )
        zoom_reset.clicked.connect(
            lambda: (
                scale_state.update(value=1.0),
                canvas.set_graph_scale(1.0),
            )
        )
        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        dialog.exec_()

    @staticmethod
    def _cad_item_label(payload: dict) -> str:
        number = str(payload.get("item_number") or "").strip()
        name = str(payload.get("item_name") or "").strip()
        aes = str(payload.get("item_aes_number") or "").strip()
        label = " - ".join(value for value in (number, name) if value)
        if aes:
            label = f"{label} | AES {aes}" if label else f"AES {aes}"
        return label or "No EBOM Item"

    @staticmethod
    def _normalized_cad_category(cad_document: dict) -> str:
        category = str(cad_document.get("category") or "").strip().upper()
        if category:
            return category
        file_name = str(cad_document.get("file_name") or "").lower()
        if ".asm" in file_name:
            return "ASSEMBLY"
        if ".prt" in file_name:
            return "COMPONENT"
        if ".drw" in file_name:
            return "DRAWING"
        return ""

    def _cad_create_ebom_item_disabled_reason(self, cad_document: dict) -> str:
        if not self.perm.can("manage_parts"):
            return "You do not have permission to create EBOM Items."
        category = self._normalized_cad_category(cad_document)
        if category == "DRAWING":
            return "A DRW is a related drawing of a PRT/ASM; create the EBOM Item from the owning model."
        if category not in {"ASSEMBLY", "COMPONENT"}:
            return "Create EBOM Items only from PRT/ASM CAD Documents."
        return ""

    def _related_drawings_for_cad_document(self, cad_document: dict) -> list[dict]:
        cad_id = cad_document.get("id") or cad_document.get("child_cad_document_id")
        drawings = [dict(row) for row in (cad_document.get("related_drawings") or []) if row]
        if drawings or cad_id is None:
            return drawings
        try:
            for document in self.bom_service.list_pdm_cad_documents() or []:
                try:
                    if int(document.get("id") or 0) != int(cad_id):
                        continue
                except Exception:
                    continue
                return [
                    dict(row)
                    for row in (document.get("related_drawings") or [])
                    if row
                ]
        except Exception:
            return []
        return []

    def _select_drawings_for_new_cad_item(
        self, cad_document: dict
    ) -> tuple[list[int], int | None] | None:
        drawings = self._related_drawings_for_cad_document(cad_document)
        if not drawings:
            return ([], None)

        dialog = QDialog(self)
        dialog.setWindowTitle("Select Item Drawing")
        dialog.resize(620, 420)
        layout = QVBoxLayout(dialog)
        title = QLabel(
            "Select the drawing(s) that correspond to the new EBOM Item."
        )
        title.setStyleSheet("font-weight:700;color:#172635;")
        layout.addWidget(title)
        hint = QLabel(
            "The selected DRW will be associated to the new Item as CONTENT. "
            "If several drawings are selected, choose the primary drawing."
        )
        hint.setWordWrap(True)
        hint.setStyleSheet("color:#526577;")
        layout.addWidget(hint)

        drawing_list = QListWidget()
        drawing_list.setAlternatingRowColors(True)
        drawing_list.setSelectionMode(QAbstractItemView.SingleSelection)
        primary_combo = QComboBox()
        row_by_id = {}
        for drawing in drawings:
            try:
                drawing_id = int(drawing.get("id"))
            except Exception:
                continue
            label = str(
                drawing.get("file_name")
                or drawing.get("name")
                or f"Drawing {drawing_id}"
            )
            creo_text = self._pdm_creo_file_text(drawing)
            if creo_text:
                label = f"{label}  |  Creo {creo_text}"
            row = QListWidgetItem(label)
            row.setFlags(row.flags() | Qt.ItemIsUserCheckable)
            row.setCheckState(
                Qt.Checked if bool(drawing.get("is_primary_drawing")) else Qt.Unchecked
            )
            row.setData(Qt.UserRole, drawing_id)
            row.setToolTip(
                "CAD Drawing\n"
                f"File: {drawing.get('file_name') or '-'}\n"
                f"CAD version: {drawing.get('revision') or 'A'}.{int(drawing.get('iteration') or 1)}\n"
                f"Creo file: {creo_text or '-'}"
            )
            drawing_list.addItem(row)
            row_by_id[drawing_id] = row

        if drawing_list.count() == 1:
            drawing_list.item(0).setCheckState(Qt.Checked)

        def selected_ids() -> list[int]:
            result = []
            for index in range(drawing_list.count()):
                row = drawing_list.item(index)
                if row.checkState() == Qt.Checked:
                    try:
                        result.append(int(row.data(Qt.UserRole)))
                    except Exception:
                        pass
            return result

        def refresh_primary_combo() -> None:
            previous = primary_combo.currentData()
            primary_combo.blockSignals(True)
            primary_combo.clear()
            for index in range(drawing_list.count()):
                row = drawing_list.item(index)
                if row.checkState() != Qt.Checked:
                    continue
                primary_combo.addItem(row.text(), int(row.data(Qt.UserRole)))
            if previous is not None:
                match_index = primary_combo.findData(previous)
                if match_index >= 0:
                    primary_combo.setCurrentIndex(match_index)
            primary_combo.setEnabled(primary_combo.count() > 1)
            primary_combo.blockSignals(False)

        drawing_list.itemChanged.connect(lambda _row: refresh_primary_combo())
        refresh_primary_combo()

        layout.addWidget(drawing_list, 1)
        layout.addWidget(QLabel("Primary drawing:"))
        layout.addWidget(primary_combo)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)

        def accept_dialog() -> None:
            ids = selected_ids()
            if not ids:
                QMessageBox.warning(
                    dialog,
                    "Select Item Drawing",
                    "Select at least one drawing for this new Item, or cancel creation.",
                )
                return
            dialog.accept()

        buttons.accepted.connect(accept_dialog)
        buttons.rejected.connect(dialog.reject)
        if dialog.exec_() != QDialog.Accepted:
            return None

        ids = selected_ids()
        primary = primary_combo.currentData()
        try:
            primary_id = int(primary) if primary is not None else (ids[0] if len(ids) == 1 else None)
        except Exception:
            primary_id = ids[0] if len(ids) == 1 else None
        return (ids, primary_id)

    def _create_ebom_item_from_cad_document(
        self,
        cad_document: dict,
        parent_item_id: int | None = None,
        quantity: int = 1,
    ) -> int | None:
        if not self.perm.can("manage_parts"):
            QMessageBox.warning(self, "Permission denied", "You cannot create EBOM Items.")
            return None
        if not cad_document:
            return None
        cad_id = cad_document.get("id") or cad_document.get("child_cad_document_id")
        if cad_id is None:
            QMessageBox.warning(self, "Create Item", "The CAD Document was not found.")
            return None
        category = self._normalized_cad_category(cad_document)
        if category not in {"ASSEMBLY", "COMPONENT"}:
            QMessageBox.warning(
                self,
                "Create Item",
                "Create EBOM Items only from PRT/ASM CAD Documents.",
            )
            return None
        associations = self._pdm_document_associations(cad_document)
        try:
            list_for_cad = getattr(
                self.bom_service, "list_cad_item_associations", None
            )
            if list_for_cad is not None:
                fresh_associations = list(list_for_cad(int(cad_id)) or [])
                if fresh_associations:
                    associations = fresh_associations
            else:
                fresh_document = next((
                    row for row in (self.bom_service.list_pdm_cad_documents() or [])
                    if int(row.get("id") or 0) == int(cad_id)
                ), None)
                if fresh_document:
                    associations = self._pdm_document_associations(fresh_document)
        except Exception:
            pass
        has_owner = any(
            str(association.get("association_type") or "").upper() == "OWNER"
            for association in associations
        )
        association_choices = list(self._pdm_association_types())
        if has_owner:
            association_choices = [
                choice for choice in association_choices if choice[1] != "OWNER"
            ]
        default_type = "IMAGE" if has_owner else "OWNER"
        default_index = next((
            index for index, (_label, value) in enumerate(association_choices)
            if value == default_type
        ), 0)
        selected_type_label, accepted = QInputDialog.getItem(
            self,
            "Create EBOM Item from CAD",
            (
                "Association to the new Item:\n"
                + (
                    "This CAD Document already has an OWNER. Select a shared or "
                    "supporting association."
                    if has_owner else
                    "Select how this CAD Document represents the new Item."
                )
            ),
            [label for label, _value in association_choices],
            default_index,
            False,
        )
        if not accepted:
            return None
        association_type = next(
            value for label, value in association_choices
            if label == selected_type_label
        )
        drawing_selection = self._select_drawings_for_new_cad_item(cad_document)
        if drawing_selection is None:
            return None
        selected_drawing_ids, primary_drawing_id = drawing_selection
        file_name = str(cad_document.get("file_name") or "").strip()
        stem = os.path.splitext(file_name)[0] if file_name else ""
        part_data = {
            "name": str(cad_document.get("name") or stem or file_name or "New Item").strip(),
            "type": "asm" if category == "ASSEMBLY" else "prt",
            "item_type": "MECHANICAL_PART",
            "assembly_mode": "SEPARABLE" if category == "ASSEMBLY" else "COMPONENT",
            "procurement_source": "MAKE",
            "item_view": "DESIGN",
            "default_unit": "EA",
            "classification": "PHYSICAL",
            "cad_control_mode": "CONTROLLED",
            "default_ebom_behavior": "NORMAL",
            "cad_requirement": "REQUIRED",
            "drawing_requirement": "OPTIONAL",
            "notes": f"Created from CAD Document {file_name or cad_id}.",
        }
        dialog = PartDialog(self, part_data)
        dialog.setWindowTitle("Create EBOM Item from CAD")
        if dialog.exec_() != QDialog.Accepted:
            return None
        try:
            new_item_id = self.bom_service.add_part(dialog.get_data())
            if not isinstance(new_item_id, int):
                raise ValueError("The Item could not be created.")
            self.bom_service.pdm_service.associate(
                int(self.session.project_id), int(new_item_id), int(cad_id),
                association_type, self.bom_service.user_id,
            )
            if selected_drawing_ids:
                self.bom_service.pdm_service.set_item_model_drawings(
                    int(new_item_id),
                    int(cad_id),
                    selected_drawing_ids,
                    primary_drawing_id=primary_drawing_id,
                    actor_id=self.bom_service.user_id,
                )
        except Exception as exc:
            QMessageBox.warning(self, "Create EBOM Item from CAD", str(exc))
            return None

        usage_message = ""
        if parent_item_id is not None and int(parent_item_id) != int(new_item_id):
            if QMessageBox.question(
                self,
                "Add to Item Structure",
                "Add the new Item as a child usage under the corresponding parent EBOM Item?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            ) == QMessageBox.Yes:
                if self._ensure_item_checked_out_for_pdm_change(int(parent_item_id)):
                    try:
                        self.bom_service.add_manual_item_usage(
                            int(parent_item_id), int(new_item_id),
                            max(1, int(quantity or 1)),
                        )
                        usage_message = "\nA child usage was added to the parent Item."
                    except Exception as exc:
                        usage_message = (
                            "\nThe Item was created and associated, but the child usage "
                            f"could not be added: {exc}"
                        )
        affected_items = []
        if parent_item_id is not None:
            affected_items.append(int(parent_item_id))
        affected_items.append(int(new_item_id))
        self._add_part_to_tree(int(new_item_id))
        if parent_item_id is not None:
            self._refresh_loaded_part_branch(int(parent_item_id))
        else:
            self._refresh_pdm_ebom_structure_branch(int(new_item_id))
        self._refresh_pdm_context_rows(
            item_ids=affected_items,
            cad_ids=sorted({int(cad_id), *[int(value) for value in selected_drawing_ids]}),
        )
        QMessageBox.information(
            self,
            "Create EBOM Item from CAD",
            f"Created Item from {file_name or cad_id} and associated it as "
            f"{association_type.replace('_', ' ')}."
            + (
                "\nSelected drawing(s): "
                + ", ".join(str(value) for value in selected_drawing_ids)
                if selected_drawing_ids else ""
            )
            + usage_message,
        )
        return int(new_item_id)

    def _top_level_ebom_item_id(self) -> int | None:
        for root in list(getattr(self, "_pdm_ebom_roots", []) or []):
            value = root.get("bom_id") or root.get("id")
            try:
                if value is not None:
                    return int(value)
            except (TypeError, ValueError):
                continue
        tree = getattr(self, "_ebom_tree", None)
        if tree is not None:
            for index in range(tree.topLevelItemCount()):
                item = tree.topLevelItem(index)
                if (
                    item is None
                    or self._is_folder_tree_item(item)
                    or item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD
                ):
                    continue
                value = item.data(0, Qt.UserRole)
                try:
                    if value is not None:
                        return int(value)
                except (TypeError, ValueError):
                    continue
        return None

    def compare_top_level_cad_to_item_structure(self) -> None:
        item_id = self._top_level_ebom_item_id()
        if item_id is None:
            QMessageBox.information(
                self,
                "Compare CAD to Item Structure",
                "No top-level EBOM Item is available to compare.",
            )
            return
        self.compare_cad_to_item_structure(item_id)

    def compare_cad_to_item_structure(self, item_id: int | None = None) -> None:
        if item_id is None:
            item_id = self._pdm_current_item_id()
        if item_id is None:
            return
        item_id = int(item_id)
        try:
            comparison = self.bom_service.compare_cad_to_item_structure(item_id)
            associations = self.bom_service.list_item_cad_associations(item_id)
            owner = next(
                (
                    row for row in associations or []
                    if str(row.get("association_type") or "").upper() == "OWNER"
                    and str(row.get("category") or "").upper() != "DRAWING"
                ),
                None,
            )
            if not owner:
                raise ValueError("This Item has no OWNER CAD Document.")
            cad_data = self.bom_service.get_pdm_cad_structure()
            cad_root = self._find_payload_node_by_id(
                cad_data.get("roots") or [], int(owner["id"]), "id"
            )
            if cad_root is None:
                cad_root = dict(owner, children=[])
            ebom_data = self.bom_service.get_released_ebom_project(
                int(self.session.project_id)
            )
            item_root = self._find_payload_node_by_id(
                ebom_data.get("roots") or [], int(item_id), "bom_id"
            )
            if item_root is None:
                item_root = self.bom_service.get_part_details(int(item_id)) or {}
                item_root["children"] = []
        except Exception as exc:
            QMessageBox.warning(self, "Compare CAD to Item Structure", str(exc))
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Compare CAD Structure to EBOM Item Structure")
        dialog.resize(1180, 680)
        dialog.setStyleSheet("""
            QDialog { background:#e9edf1; color:#1d2935; font:9pt 'Segoe UI'; }
            QLabel#compareTitle { color:#172635; font-size:12pt; font-weight:600; }
            QLabel#compareSubtitle { color:#536270; font-size:8.5pt; }
            QTreeWidget {
                background:white; alternate-background-color:#f4f6f8;
                border:1px solid #9eaab5; gridline-color:#d4dbe1;
            }
            QHeaderView::section {
                background:#d8e0e6; color:#22313e; border:0;
                border-right:1px solid #aeb8c2; border-bottom:1px solid #9eaab5;
                padding:4px 6px; font-weight:600;
            }
            QPushButton {
                border:1px solid #8d99a4; border-radius:0; background:#e2e7eb;
                min-height:24px; padding:2px 10px;
            }
            QPushButton#compareToolButton {
                min-width:26px; max-width:26px; min-height:24px; max-height:24px;
                padding:0px; font-weight:600;
            }
            QPushButton#compareCommandButton {
                min-height:22px; padding:1px 8px;
            }
            QFrame#compareToolbar {
                background:#dce3e8; border:1px solid #aab5bf;
            }
            QFrame#compareBridgeToolbar {
                background:#dce3e8; border:1px solid #aab5bf;
                min-width:34px; max-width:34px;
            }
            QCheckBox { color:#233241; }
        """)
        layout = QVBoxLayout(dialog)
        title = QLabel("CAD / EBOM Structure Compare")
        title.setObjectName("compareTitle")
        subtitle = QLabel(
            "Left: OWNER CAD structure. Right: persisted EBOM Item structure. "
            "Shared PRT/ASM documents are evaluated against every structure-participating "
            "Item association. Selecting a row locates the corresponding CAD/Item row."
        )
        subtitle.setObjectName("compareSubtitle")
        subtitle.setWordWrap(True)
        layout.addWidget(title)
        layout.addWidget(subtitle)

        command_bar = QFrame()
        command_bar.setObjectName("compareToolbar")
        command_layout = QHBoxLayout(command_bar)
        command_layout.setContentsMargins(4, 3, 4, 3)
        command_layout.setSpacing(3)
        refresh_btn = QPushButton("Refresh")
        refresh_btn.setObjectName("compareCommandButton")
        previous_diff_btn = QPushButton("Previous Difference")
        previous_diff_btn.setObjectName("compareCommandButton")
        next_diff_btn = QPushButton("Next Difference")
        next_diff_btn.setObjectName("compareCommandButton")
        expand_btn = QPushButton("Expand")
        expand_btn.setObjectName("compareCommandButton")
        collapse_btn = QPushButton("Collapse")
        collapse_btn.setObjectName("compareCommandButton")
        show_differences_only = QCheckBox("Show differences only")
        for button in (
            refresh_btn, previous_diff_btn, next_diff_btn,
            expand_btn, collapse_btn,
        ):
            command_layout.addWidget(button)
        command_layout.addSpacing(8)
        command_layout.addWidget(show_differences_only)
        command_layout.addStretch(1)
        layout.addWidget(command_bar)

        splitter = QSplitter(Qt.Horizontal)
        cad_panel = QGroupBox("CAD Structure")
        cad_layout = QVBoxLayout(cad_panel)
        cad_commands = QHBoxLayout()
        cad_commands.setSpacing(3)
        cad_register_btn = QPushButton("Register")
        cad_add_btn = QPushButton("Add")
        cad_edit_btn = QPushButton("Edit")
        cad_delete_btn = QPushButton("Delete")
        cad_open_btn = QPushButton("Open")
        for button in (
            cad_register_btn, cad_add_btn, cad_edit_btn,
            cad_delete_btn, cad_open_btn,
        ):
            button.setObjectName("compareCommandButton")
            cad_commands.addWidget(button)
        cad_commands.addStretch(1)
        cad_layout.addLayout(cad_commands)
        cad_tree = QTreeWidget()
        cad_tree.setColumnCount(5)
        cad_tree.setHeaderLabels([
            "CAD Item", "Associated EBOM Item", "Association", "Qty", "Build Status"
        ])
        cad_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        cad_tree.setSelectionMode(QAbstractItemView.SingleSelection)
        cad_tree.setAlternatingRowColors(True)
        cad_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        cad_header = cad_tree.header()
        cad_header.setSectionResizeMode(QHeaderView.Interactive)
        cad_header.setStretchLastSection(False)
        cad_header.setMinimumSectionSize(45)
        for column, width in enumerate((320, 260, 110, 55, 130)):
            cad_tree.setColumnWidth(column, width)
        cad_layout.addWidget(cad_tree)

        item_panel = QGroupBox("EBOM Item Structure")
        item_layout = QVBoxLayout(item_panel)
        item_commands = QHBoxLayout()
        item_commands.setSpacing(3)
        item_add_btn = QPushButton("Add")
        item_add_usage_btn = QPushButton("Add Usage")
        item_edit_btn = QPushButton("Edit")
        item_delete_btn = QPushButton("Delete")
        item_associate_btn = QPushButton("Associate")
        item_open_btn = QPushButton("Open")
        for button in (
            item_add_btn, item_add_usage_btn, item_edit_btn,
            item_delete_btn, item_associate_btn, item_open_btn,
        ):
            button.setObjectName("compareCommandButton")
            item_commands.addWidget(button)
        item_commands.addStretch(1)
        item_layout.addLayout(item_commands)
        item_tree = QTreeWidget()
        item_tree.setColumnCount(4)
        item_tree.setHeaderLabels(["EBOM Item", "Qty", "Source", "Compare Status"])
        item_tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        item_tree.setSelectionMode(QAbstractItemView.SingleSelection)
        item_tree.setAlternatingRowColors(True)
        item_header = item_tree.header()
        item_header.setSectionResizeMode(QHeaderView.Interactive)
        item_header.setStretchLastSection(False)
        item_header.setMinimumSectionSize(45)
        for column, width in enumerate((300, 55, 100, 150)):
            item_tree.setColumnWidth(column, width)
        item_layout.addWidget(item_tree)

        bridge_toolbar = QFrame()
        bridge_toolbar.setObjectName("compareBridgeToolbar")
        bridge_layout = QVBoxLayout(bridge_toolbar)
        bridge_layout.setContentsMargins(3, 4, 3, 4)
        bridge_layout.setSpacing(4)

        def bridge_button(text: str, tooltip: str) -> QPushButton:
            button = QPushButton(text)
            button.setObjectName("compareToolButton")
            button.setToolTip(tooltip)
            bridge_layout.addWidget(button)
            return button

        locate_pair_btn = bridge_button("<>", "Locate the associated row on the opposite side")
        create_item_btn = bridge_button("+I", "Create an EBOM Item from the selected CAD Document")
        associate_btn = bridge_button("=>", "Manage CAD-Item association for the selected row")
        open_btn = bridge_button("O", "Open the selected CAD Document or EBOM Item in Nexus")
        bridge_layout.addSpacing(8)
        previous_bridge_btn = bridge_button("^", "Previous compare difference")
        next_bridge_btn = bridge_button("v", "Next compare difference")
        bridge_layout.addSpacing(8)
        differences_btn = bridge_button("!", "Toggle show differences only")
        recompare_btn = bridge_button("R", "Refresh this comparison")
        bridge_layout.addStretch(1)

        splitter.addWidget(cad_panel)
        splitter.addWidget(bridge_toolbar)
        splitter.addWidget(item_panel)
        splitter.setSizes([560, 36, 560])
        layout.addWidget(splitter, 1)

        cad_rows_by_item: dict[int, list[QTreeWidgetItem]] = defaultdict(list)
        item_rows_by_item: dict[int, list[QTreeWidgetItem]] = defaultdict(list)
        item_edges: set[tuple[int, int]] = set()
        expected_edges: set[tuple[int, int]] = set()
        cad_relation_edge_groups: list[set[tuple[int, int]]] = []
        cad_item_ids: set[int] = set()

        def collect_item_edges(node: dict, parent_id=None):
            node_id = node.get("bom_id") or node.get("id")
            try:
                node_id = int(node_id)
            except Exception:
                node_id = None
            if parent_id is not None and node_id is not None:
                item_edges.add((int(parent_id), int(node_id)))
            for child in node.get("children") or []:
                collect_item_edges(child, node_id)

        collect_item_edges(item_root)

        def structure_item_associations(node: dict) -> list[dict]:
            """All active Item projections that participate in CAD structure.

            CAD Structure stores one CAD Document node even when the same PRT
            or ASM represents several EBOM Items.  The legacy flattened
            ``item_id`` is only a display compatibility field and must not be
            used as the complete comparison identity.
            """
            associations = self._pdm_document_associations(node)
            participating = []
            for association in associations:
                item_value = association.get("item_id")
                try:
                    item_value = int(item_value)
                except Exception:
                    continue
                value = association.get("participates_in_structure")
                if value is None:
                    association_type = str(
                        association.get("association_type") or ""
                    ).upper()
                    participates = association_type in {
                        "OWNER", "CONTRIBUTING_IMAGE", "IMAGE"
                    }
                else:
                    try:
                        participates = bool(int(value))
                    except Exception:
                        participates = str(value).strip().lower() not in {
                            "", "0", "false", "no", "none",
                        }
                if participates:
                    row = dict(association)
                    row["item_id"] = item_value
                    participating.append(row)

            # Old projects may expose only the flattened association fields.
            if not associations and node.get("item_id") is not None:
                try:
                    item_value = int(node.get("item_id"))
                except Exception:
                    item_value = None
                if item_value is not None:
                    participates = node.get("participates_in_structure")
                    try:
                        participates = (
                            True if participates is None else bool(int(participates))
                        )
                    except Exception:
                        participates = str(participates).strip().lower() not in {
                            "", "0", "false", "no", "none",
                        }
                    if participates:
                        participating.append({
                            "item_id": item_value,
                            "item_number": node.get("item_number"),
                            "item_name": node.get("item_name"),
                            "item_aes_number": node.get("item_aes_number"),
                            "association_type": node.get("association_type"),
                            "participates_in_structure": 1,
                        })

            unique = []
            seen = set()
            for association in participating:
                item_value = int(association["item_id"])
                if item_value in seen:
                    continue
                seen.add(item_value)
                unique.append(association)
            return unique

        def structure_item_ids(node: dict) -> list[int]:
            return [
                int(association["item_id"])
                for association in structure_item_associations(node)
            ]

        def representation_only_node(node: dict) -> bool:
            associations = structure_item_associations(node)
            return bool(associations) and all(
                str(association.get("association_type") or "").upper()
                in {"IMAGE", "CONTRIBUTING_IMAGE"}
                for association in associations
            )

        def collect_expected_edges(node: dict, parent_items=None):
            item_values = structure_item_ids(node)
            cad_item_ids.update(item_values)
            relation_edges = {
                (int(parent_value), int(item_value))
                for parent_value in (parent_items or [])
                for item_value in item_values
            }
            if not representation_only_node(node):
                expected_edges.update(relation_edges)
            if (
                relation_edges
                and not representation_only_node(node)
                and not node.get("member_build_excluded")
                and not node.get("build_excluded")
                and not node.get("document_build_excluded")
            ):
                cad_relation_edge_groups.append(relation_edges)
            for child in node.get("children") or []:
                collect_expected_edges(child, item_values)

        collect_expected_edges(cad_root)

        def status_color(status: str) -> QColor:
            normalized = str(status or "").upper()
            if normalized in {"OWNER", "ROOT", "MATCHED", "MATCHED CAD", "COMPLETED"}:
                return QColor("#166534")
            if normalized in {"NO EBOM ITEM", "MISSING IN EBOM", "STRUCTURE MISMATCH"}:
                return QColor("#b45309")
            if normalized in {
                "EXTRA EBOM ITEM", "EXCLUDED", "NOT PARTICIPATING",
                "REPRESENTATION",
            }:
                return QColor("#64748b")
            return QColor("#334155")

        def add_cad_node(
            node: dict,
            parent: QTreeWidgetItem | None = None,
            parent_item_ids=None,
        ):
            item_associations = structure_item_associations(node)
            all_item_associations = self._pdm_document_associations(node)
            item_values = [
                int(association["item_id"])
                for association in item_associations
            ]
            quantity = max(1, int(node.get("quantity") or 1))
            association_types = []
            associated_labels = []
            for item_association in item_associations:
                association_type = str(
                    item_association.get("association_type") or "IMAGE"
                ).upper().replace("_", " ")
                if association_type not in association_types:
                    association_types.append(association_type)
                associated_labels.append(
                    self._pdm_item_association_label(item_association)
                )
            association = ", ".join(association_types) or "UNASSOCIATED"
            if len(associated_labels) == 1:
                associated_item_text = associated_labels[0]
            elif associated_labels:
                associated_item_text = f"{len(associated_labels)} associated EBOM Items"
            else:
                associated_item_text = "No EBOM Item"
            representation_only = representation_only_node(node)
            if parent is None:
                status = "OWNER"
            elif node.get("member_build_excluded") or node.get("build_excluded") or node.get("document_build_excluded"):
                status = "EXCLUDED"
            elif representation_only:
                status = "REPRESENTATION"
            elif not item_values:
                status = (
                    "NOT PARTICIPATING"
                    if all_item_associations or node.get("item_id") is not None
                    else "NO EBOM ITEM"
                )
            elif not parent_item_ids:
                status = "PARENT NOT ASSOCIATED"
            elif any(
                (int(parent_value), int(item_value)) in item_edges
                for parent_value in parent_item_ids
                for item_value in item_values
            ):
                status = "MATCHED"
            else:
                status = "MISSING IN EBOM"
            row = QTreeWidgetItem([
                self._cad_label(node),
                associated_item_text,
                association,
                str(quantity),
                status,
            ])
            row.setData(
                0, COMPARE_ITEM_ID_ROLE,
                item_values[0] if item_values else None,
            )
            row.setData(0, COMPARE_ITEM_IDS_ROLE, list(item_values))
            row.setData(0, COMPARE_CAD_ID_ROLE, node.get("id"))
            row.setData(
                0, COMPARE_PARENT_ITEM_ID_ROLE,
                parent_item_ids[0] if parent_item_ids else None,
            )
            row.setData(0, COMPARE_PAYLOAD_ROLE, dict(node))
            row.setIcon(0, _pdm_cad_icon(node.get("category")))
            row.setForeground(4, QBrush(status_color(status)))
            if associated_labels:
                row.setToolTip(
                    1,
                    "All structure-participating associations for this shared CAD Document:\n"
                    + "\n".join(associated_labels),
                )
            for item_value in item_values:
                cad_rows_by_item[item_value].append(row)
            if parent is None:
                cad_tree.addTopLevelItem(row)
            else:
                parent.addChild(row)
            for child in node.get("children") or []:
                add_cad_node(child, row, item_values)
            row.setExpanded(True)
            return row

        def add_item_node(node: dict, parent: QTreeWidgetItem | None = None, parent_item_id=None):
            node_id = node.get("bom_id") or node.get("id")
            try:
                node_id = int(node_id)
            except Exception:
                node_id = None
            quantity = max(1, int(node.get("source_quantity") or node.get("quantity") or 1))
            if parent is None:
                status = "ROOT"
            elif node_id in cad_item_ids and parent_item_id is not None and (int(parent_item_id), int(node_id)) in expected_edges:
                status = "MATCHED CAD"
            elif node_id in cad_item_ids:
                status = "STRUCTURE MISMATCH"
            else:
                status = "EXTRA EBOM ITEM"
            row = QTreeWidgetItem([
                self._windchill_item_label(node),
                str(quantity),
                str(node.get("source") or "MANUAL"),
                status,
            ])
            row.setData(0, COMPARE_ITEM_ID_ROLE, node_id)
            row.setData(0, COMPARE_PAYLOAD_ROLE, dict(node))
            row.setIcon(0, _pdm_item_icon())
            row.setForeground(3, QBrush(status_color(status)))
            if node_id is not None:
                item_rows_by_item[node_id].append(row)
            if parent is None:
                item_tree.addTopLevelItem(row)
            else:
                parent.addChild(row)
            for child in node.get("children") or []:
                add_item_node(child, row, node_id)
            row.setExpanded(True)
            return row

        add_cad_node(cad_root)
        add_item_node(item_root)

        syncing = {"active": False}

        def sync_from(source_tree: QTreeWidget, target_tree: QTreeWidget, target_map: dict):
            if syncing["active"]:
                return
            source_item = source_tree.currentItem()
            if source_item is None:
                return
            item_values = list(
                source_item.data(0, COMPARE_ITEM_IDS_ROLE) or []
            )
            if not item_values:
                item_value = source_item.data(0, COMPARE_ITEM_ID_ROLE)
                if item_value is not None:
                    item_values = [item_value]
            if not item_values:
                return
            current_target = target_tree.currentItem()
            current_target_id = (
                current_target.data(0, COMPARE_ITEM_ID_ROLE)
                if current_target is not None else None
            )
            preferred_values = []
            if current_target_id is not None:
                try:
                    current_target_id = int(current_target_id)
                    if current_target_id in {int(value) for value in item_values}:
                        preferred_values.append(current_target_id)
                except Exception:
                    pass
            preferred_values.extend(
                int(value) for value in item_values
                if int(value) not in preferred_values
            )
            target = next((
                candidate
                for item_value in preferred_values
                for candidate in (target_map.get(int(item_value)) or [])
            ), None)
            if target is None:
                return
            syncing["active"] = True
            try:
                target_tree.setCurrentItem(target)
                target_tree.scrollToItem(target)
            finally:
                syncing["active"] = False

        cad_tree.itemSelectionChanged.connect(
            lambda: sync_from(cad_tree, item_tree, item_rows_by_item)
        )
        item_tree.itemSelectionChanged.connect(
            lambda: sync_from(item_tree, cad_tree, cad_rows_by_item)
        )

        def recompare_dialog() -> None:
            dialog.accept()
            QTimer.singleShot(
                0,
                lambda value=int(item_id): self.compare_cad_to_item_structure(value),
            )

        def row_status(row: QTreeWidgetItem | None) -> str:
            if row is None:
                return ""
            tree = row.treeWidget()
            column = 4 if tree is cad_tree else 3
            return str(row.text(column) or "").upper()

        def is_difference_row(row: QTreeWidgetItem | None) -> bool:
            status = row_status(row)
            return bool(status and status not in {
                "OWNER", "ROOT", "MATCHED", "MATCHED CAD", "COMPLETED",
                "REPRESENTATION",
            })

        def all_compare_rows() -> list[QTreeWidgetItem]:
            rows = []

            def collect(parent: QTreeWidgetItem):
                rows.append(parent)
                for child_index in range(parent.childCount()):
                    collect(parent.child(child_index))

            for tree in (cad_tree, item_tree):
                for index in range(tree.topLevelItemCount()):
                    collect(tree.topLevelItem(index))
            return rows

        def set_difference_filter(enabled: bool) -> None:
            def apply_row(row: QTreeWidgetItem) -> bool:
                child_visible = False
                for child_index in range(row.childCount()):
                    if apply_row(row.child(child_index)):
                        child_visible = True
                visible = not enabled or is_difference_row(row) or child_visible
                row.setHidden(not visible)
                return visible

            for tree in (cad_tree, item_tree):
                for index in range(tree.topLevelItemCount()):
                    apply_row(tree.topLevelItem(index))

        def selected_cad_row() -> QTreeWidgetItem | None:
            row = cad_tree.currentItem()
            if row is not None:
                return row
            item_row = item_tree.currentItem()
            item_value = (
                item_row.data(0, COMPARE_ITEM_ID_ROLE)
                if item_row is not None else None
            )
            try:
                candidates = cad_rows_by_item.get(int(item_value), [])
            except Exception:
                candidates = []
            return candidates[0] if candidates else None

        def selected_item_row() -> QTreeWidgetItem | None:
            row = item_tree.currentItem()
            if row is not None:
                return row
            cad_row = cad_tree.currentItem()
            item_values = (
                list(cad_row.data(0, COMPARE_ITEM_IDS_ROLE) or [])
                if cad_row is not None else []
            )
            for value in item_values:
                candidates = item_rows_by_item.get(int(value), [])
                if candidates:
                    return candidates[0]
            return None

        def selected_item_value() -> int | None:
            row = selected_item_row()
            value = row.data(0, COMPARE_ITEM_ID_ROLE) if row is not None else None
            try:
                return int(value)
            except Exception:
                return None

        def selected_cad_value() -> int | None:
            row = selected_cad_row()
            value = row.data(0, COMPARE_CAD_ID_ROLE) if row is not None else None
            try:
                return int(value)
            except Exception:
                return None

        def selected_cad_payload() -> dict:
            row = selected_cad_row()
            return dict(row.data(0, COMPARE_PAYLOAD_ROLE) or {}) if row is not None else {}

        def compare_action(action) -> None:
            dialog.accept()
            QTimer.singleShot(0, action)

        def reopen_after(action) -> None:
            dialog.accept()

            def run_and_reopen() -> None:
                action()
                self.compare_cad_to_item_structure(int(item_id))

            QTimer.singleShot(0, run_and_reopen)

        def locate_pair() -> None:
            if cad_tree.hasFocus() or cad_tree.currentItem() is not None:
                sync_from(cad_tree, item_tree, item_rows_by_item)
            if item_tree.hasFocus() or item_tree.currentItem() is not None:
                sync_from(item_tree, cad_tree, cad_rows_by_item)

        def open_selected_object() -> None:
            cad_row = cad_tree.currentItem()
            if cad_tree.hasFocus() and cad_row is not None:
                cad_value = cad_row.data(0, COMPARE_CAD_ID_ROLE)
                if cad_value is not None:
                    dialog.accept()
                    QTimer.singleShot(
                        0,
                        lambda value=int(cad_value): self._select_cad_in_structure(value),
                    )
                    return
            item_value = selected_item_value()
            if item_value is not None:
                dialog.accept()
                QTimer.singleShot(
                    0,
                    lambda value=int(item_value): self._select_item_in_ebom(value),
                )

        def open_selected_cad() -> None:
            cad_value = selected_cad_value()
            if cad_value is None:
                QMessageBox.information(dialog, "Open CAD Document", "Select a CAD row first.")
                return
            compare_action(lambda value=int(cad_value): self._select_cad_in_structure(value))

        def open_selected_item() -> None:
            item_value = selected_item_value()
            if item_value is None:
                QMessageBox.information(dialog, "Open EBOM Item", "Select an EBOM row first.")
                return
            compare_action(lambda value=int(item_value): self._select_item_in_ebom(value))

        def manage_selected_association() -> None:
            target_item_id = selected_item_value()
            cad_row = selected_cad_row()
            cad_value = (
                cad_row.data(0, COMPARE_CAD_ID_ROLE)
                if cad_row is not None else None
            )
            if target_item_id is None:
                QMessageBox.information(
                    dialog,
                    "CAD-Item Association",
                    "Select an EBOM Item or a CAD row that already has an Item projection.",
                )
                return
            dialog.accept()
            QTimer.singleShot(
                0,
                lambda item_value=int(target_item_id), focus_value=(
                    int(cad_value) if cad_value is not None else None
                ): self.manage_cad_item_associations(
                    item_value, focus_cad_id=focus_value
                ),
            )

        def associate_selected_cad_to_item() -> None:
            cad_value = selected_cad_value()
            item_value = selected_item_value()
            if cad_value is None and item_value is None:
                QMessageBox.information(
                    dialog,
                    "CAD-Item Association",
                    "Select a CAD row and/or an EBOM Item row first.",
                )
                return
            if cad_value is not None and item_value is not None:
                compare_action(
                    lambda c=int(cad_value), i=int(item_value):
                    self._associate_specific_cad_to_item(c, i)
                )
                return
            if cad_value is not None:
                compare_action(
                    lambda value=int(cad_value): self._associate_cad_to_an_item(value)
                )
                return
            compare_action(
                lambda value=int(item_value): self.manage_cad_item_associations(value)
            )

        def add_ebom_item_from_selected_parent() -> None:
            if not self.perm.can("manage_parts"):
                QMessageBox.warning(
                    dialog, "Permission", "You do not have permission to create Items."
                )
                return
            parent_row = selected_item_row()
            parent_value = selected_item_value()
            if parent_row is None or parent_value is None:
                QMessageBox.information(
                    dialog,
                    "Add EBOM Item",
                    "Select the EBOM Item that will receive the new child Item.",
                )
                return
            item_dialog = PartDialog(dialog)
            if item_dialog.exec_() != QDialog.Accepted:
                return
            part_data = item_dialog.get_data()
            if not part_data.get("name"):
                QMessageBox.warning(dialog, "Validation Error", "Name is required.")
                return
            quantity, accepted = QInputDialog.getInt(
                dialog, "Add EBOM Item", "Usage quantity:", 1, 1, 100000, 1
            )
            if not accepted:
                return
            try:
                added = self.bom_service.add_part(part_data)
                if not isinstance(added, int):
                    raise ValueError("The Item could not be created.")
                self.bom_service.add_manual_item_usage(
                    int(parent_value), int(added), int(quantity)
                )
                info = self._add_part_to_tree(int(added)) or {}
                if not info:
                    info = self.bom_service.get_part_details(int(added)) or {"id": int(added)}
                info = dict(info)
                info.setdefault("id", int(added))
                info.setdefault("bom_id", int(added))
                info["source_quantity"] = int(quantity)
                info["quantity"] = int(quantity)
                info["source"] = "MANUAL"
                info.setdefault("children", [])
                row = add_item_node(info, parent_row, int(parent_value))
                parent_row.setExpanded(True)
                item_tree.setCurrentItem(row)
                item_tree.scrollToItem(row)
                try:
                    self.window().statusBar().showMessage(
                        f"Item {info.get('part_number') or added} added under {parent_row.text(0)}.",
                        6000,
                    )
                except Exception:
                    pass
                try:
                    self.display_details(int(added))
                except Exception:
                    pass
            except Exception as exc:
                QMessageBox.critical(
                    dialog, "New Item", f"Could not create Item:\n{exc}"
                )

        def selected_item_for_main_tree():
            value = selected_item_value()
            if value is None:
                return None
            tree = getattr(self, "_ebom_tree", None)
            if tree is None:
                return None
            self._materialize_pdm_tree_for_search(tree)
            for candidate in self._iter_tree_items(tree):
                if candidate.data(0, Qt.UserRole) == int(value):
                    return candidate
            return None

        def edit_selected_ebom_item() -> None:
            item_value = selected_item_value()
            if item_value is None:
                QMessageBox.information(dialog, "Edit EBOM Item", "Select an EBOM row first.")
                return
            reopen_after(lambda value=int(item_value): self.edit_part(value))

        def delete_selected_ebom_item() -> None:
            item_value = selected_item_value()
            if item_value is None:
                QMessageBox.information(dialog, "Delete EBOM Item", "Select an EBOM row first.")
                return
            reopen_after(lambda value=int(item_value): self.delete_part(value))

        def add_manual_usage_to_selected_item() -> None:
            item_value = selected_item_value()
            if item_value is None:
                QMessageBox.information(dialog, "Add Item Usage", "Select a parent EBOM Item first.")
                return
            self.current_part_id = int(item_value)
            reopen_after(self.add_manual_item_usage)

        def register_cad_from_compare() -> None:
            compare_action(self.register_cad_document)

        def add_cad_component_from_compare() -> None:
            cad_value = selected_cad_value()
            if cad_value is None:
                QMessageBox.information(dialog, "Add CAD Component", "Select a CAD assembly row first.")
                return
            payload = selected_cad_payload()
            if str(payload.get("category") or "").upper() != "ASSEMBLY":
                QMessageBox.information(
                    dialog, "Add CAD Component", "Select a CAD assembly row first."
                )
                return
            reopen_after(lambda value=int(cad_value): self._add_cad_member_from_tree(value))

        def edit_cad_occurrence_from_compare() -> None:
            row = selected_cad_row()
            if row is None:
                QMessageBox.information(dialog, "Edit CAD Occurrence", "Select a CAD row first.")
                return
            parent_row = row.parent()
            if parent_row is None:
                QMessageBox.information(
                    dialog,
                    "Edit CAD Occurrence",
                    "Select a child CAD occurrence, not the root CAD Document.",
                )
                return
            parent_cad_id = parent_row.data(0, COMPARE_CAD_ID_ROLE)
            child_cad_id = row.data(0, COMPARE_CAD_ID_ROLE)
            payload = selected_cad_payload()
            if parent_cad_id is None or child_cad_id is None:
                return

            edit_dialog = QDialog(dialog)
            edit_dialog.setWindowTitle("Edit CAD Occurrence")
            edit_layout = QGridLayout(edit_dialog)
            quantity = QSpinBox()
            quantity.setRange(1, 100000)
            try:
                quantity.setValue(max(1, int(payload.get("quantity") or row.text(3) or 1)))
            except Exception:
                quantity.setValue(1)
            excluded = QCheckBox("Exclude this occurrence from Item Structure build")
            excluded.setChecked(bool(
                payload.get("member_build_excluded") or payload.get("build_excluded")
            ))
            edit_layout.addWidget(QLabel("CAD component"), 0, 0)
            edit_layout.addWidget(QLabel(row.text(0)), 0, 1)
            edit_layout.addWidget(QLabel("Quantity"), 1, 0)
            edit_layout.addWidget(quantity, 1, 1)
            edit_layout.addWidget(excluded, 2, 0, 1, 2)
            buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
            buttons.accepted.connect(edit_dialog.accept)
            buttons.rejected.connect(edit_dialog.reject)
            edit_layout.addWidget(buttons, 3, 0, 1, 2)
            if edit_dialog.exec_() != QDialog.Accepted:
                return
            try:
                self.bom_service.add_pdm_cad_member(
                    int(parent_cad_id), int(child_cad_id),
                    quantity.value(), excluded.isChecked(),
                )
            except Exception as exc:
                QMessageBox.warning(dialog, "Edit CAD Occurrence", str(exc))
                return
            recompare_dialog()

        def delete_selected_cad_from_compare() -> None:
            row = selected_cad_row()
            if row is None:
                QMessageBox.information(dialog, "Delete CAD", "Select a CAD row first.")
                return
            cad_value = row.data(0, COMPARE_CAD_ID_ROLE)
            parent_row = row.parent()
            if parent_row is not None:
                member_id = selected_cad_payload().get("member_id")
                parent_cad_id = parent_row.data(0, COMPARE_CAD_ID_ROLE)
                if member_id is not None and parent_cad_id is not None:
                    reopen_after(
                        lambda value=int(member_id), label=row.text(0),
                               parent_value=int(parent_cad_id):
                        self._remove_cad_member_from_tree(value, label, parent_value)
                    )
                    return
            if cad_value is None:
                return
            payload = selected_cad_payload()
            reopen_after(
                lambda value=int(cad_value), data=payload:
                self._delete_selected_pdm_cad_document(value, data)
            )

        def select_difference(step: int) -> None:
            rows = [row for row in all_compare_rows() if is_difference_row(row)]
            if not rows:
                return
            current = cad_tree.currentItem() or item_tree.currentItem()
            try:
                current_index = rows.index(current)
            except ValueError:
                current_index = -1 if step > 0 else 0
            target = rows[(current_index + step) % len(rows)]
            target_tree = target.treeWidget()
            if target_tree is not None:
                target_tree.setCurrentItem(target)
                target_tree.scrollToItem(target)

        def create_from_selected_cad_row(selected: QTreeWidgetItem, data: dict) -> None:
            try:
                quantity = max(1, int(data.get("quantity") or selected.text(3) or 1))
            except Exception:
                quantity = 1
            new_item_id = self._create_ebom_item_from_cad_document(
                data,
                parent_item_id=selected.data(0, COMPARE_PARENT_ITEM_ID_ROLE),
                quantity=quantity,
            )
            if new_item_id is None:
                return
            dialog.accept()
            QTimer.singleShot(
                0,
                lambda value=int(item_id): self.compare_cad_to_item_structure(value),
            )

        def create_from_current_cad_row() -> None:
            row = selected_cad_row()
            if row is None:
                QMessageBox.information(
                    dialog, "Create EBOM Item from CAD", "Select a CAD row first."
                )
                return
            payload = dict(row.data(0, COMPARE_PAYLOAD_ROLE) or {})
            create_reason = self._cad_create_ebom_item_disabled_reason(payload)
            if create_reason:
                QMessageBox.information(
                    dialog, "Create EBOM Item from CAD", create_reason
                )
                return
            create_from_selected_cad_row(row, payload)

        refresh_btn.clicked.connect(recompare_dialog)
        recompare_btn.clicked.connect(recompare_dialog)
        expand_btn.clicked.connect(lambda: (cad_tree.expandAll(), item_tree.expandAll()))
        collapse_btn.clicked.connect(lambda: (cad_tree.collapseAll(), item_tree.collapseAll()))
        show_differences_only.toggled.connect(set_difference_filter)
        differences_btn.clicked.connect(
            lambda: show_differences_only.setChecked(
                not show_differences_only.isChecked()
            )
        )
        previous_diff_btn.clicked.connect(lambda: select_difference(-1))
        previous_bridge_btn.clicked.connect(lambda: select_difference(-1))
        next_diff_btn.clicked.connect(lambda: select_difference(1))
        next_bridge_btn.clicked.connect(lambda: select_difference(1))
        locate_pair_btn.clicked.connect(locate_pair)
        create_item_btn.clicked.connect(create_from_current_cad_row)
        associate_btn.clicked.connect(manage_selected_association)
        open_btn.clicked.connect(open_selected_object)
        cad_register_btn.clicked.connect(register_cad_from_compare)
        cad_add_btn.clicked.connect(add_cad_component_from_compare)
        cad_edit_btn.clicked.connect(edit_cad_occurrence_from_compare)
        cad_delete_btn.clicked.connect(delete_selected_cad_from_compare)
        cad_open_btn.clicked.connect(open_selected_cad)
        item_add_btn.clicked.connect(add_ebom_item_from_selected_parent)
        item_add_usage_btn.clicked.connect(add_manual_usage_to_selected_item)
        item_edit_btn.clicked.connect(edit_selected_ebom_item)
        item_delete_btn.clicked.connect(delete_selected_ebom_item)
        item_associate_btn.clicked.connect(associate_selected_cad_to_item)
        item_open_btn.clicked.connect(open_selected_item)

        def cad_context_menu(position):
            row = cad_tree.itemAt(position)
            if row is None:
                return
            cad_tree.setCurrentItem(row)
            payload = dict(row.data(0, COMPARE_PAYLOAD_ROLE) or {})
            menu = QMenu(dialog)
            create_action = menu.addAction("Create EBOM Item from CAD...")
            create_reason = self._cad_create_ebom_item_disabled_reason(payload)
            create_action.setEnabled(not create_reason)
            if create_reason:
                create_action.setToolTip(create_reason)
            create_action.triggered.connect(
                lambda _checked=False, selected=row, data=payload:
                create_from_selected_cad_row(selected, data)
            )
            open_action = menu.addAction("Open CAD Document in CAD Structure")
            open_action.triggered.connect(
                lambda _checked=False, value=row.data(0, COMPARE_CAD_ID_ROLE):
                self._select_cad_in_structure(int(value)) if value is not None else None
            )
            menu.exec_(cad_tree.viewport().mapToGlobal(position))

        cad_tree.customContextMenuRequested.connect(cad_context_menu)

        missing_edges = sum(
            1 for supported_edges in cad_relation_edge_groups
            if not supported_edges.intersection(item_edges)
        )
        summary = QLabel(
            f"CAD-Item projections: {sum(len(rows) for rows in cad_rows_by_item.values())}  |  "
            f"CAD relations without an EBOM usage: {missing_edges}  |  "
            f"Legacy compare rows: {len(comparison.get('rows') or [])}"
        )
        summary.setObjectName("compareSubtitle")
        layout.addWidget(summary)
        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        dialog.exec_()

    def add_manual_item_usage(self) -> None:
        if not self.perm.can("manage_parts"):
            QMessageBox.warning(self, "Permission denied", "You cannot edit the Item Structure.")
            return
        parent_id = self._pdm_current_item_id()
        if parent_id is None:
            return
        if not self._ensure_item_checked_out_for_pdm_change(int(parent_id)):
            return
        items = [
            row for row in (self.bom_service.list_pdm_items() or [])
            if int(row.get("id") or 0) != int(parent_id)
        ]
        if not items:
            QMessageBox.information(self, "Manual Item Usage", "No other Items are available.")
            return
        labels = [self._item_identity_text(row) for row in items]
        selected, accepted = QInputDialog.getItem(
            self, "Manual Item Usage", "Child Item:", labels, 0, False
        )
        if not accepted:
            return
        quantity, accepted = QInputDialog.getInt(
            self, "Manual Item Usage", "Quantity:", 1, 1, 100000, 1
        )
        if not accepted:
            return
        child = items[labels.index(selected)]
        try:
            self.bom_service.add_manual_item_usage(
                int(parent_id), int(child["id"]), int(quantity)
            )
        except Exception as exc:
            QMessageBox.critical(self, "Manual Item Usage", str(exc))
            return
        self._refresh_loaded_part_branch(int(parent_id))
        self._refresh_part_in_tree(int(child["id"]))

    def _refresh_ebom_filters(self) -> int:
        tree = getattr(self, "_ebom_tree", None)
        if tree is None:
            return 0
        query = str(self.search_input.text() or "").strip()
        filters = self._bom_advanced_filters or self._default_bom_advanced_filters()
        advanced_active = not self._is_default_bom_advanced_filter(filters)
        if query or advanced_active or str(filters.get("text") or "").strip():
            return self._apply_ebom_db_flat_filter(query=query, filters=filters)
        show_parents = bool(filters.get("show_parent_matches", True))
        flat_results = bool(advanced_active) and (
            not show_parents or bool(filters.get("remove_duplicates", False))
        )
        visible_count = 0

        def item_matches(item):
            if self._is_lazy_placeholder(item):
                return False, False
            haystack = " ".join(
                str(item.text(column) or "")
                for column in range(tree.columnCount())
            ) + " " + " ".join(
                (
                    str(item.toolTip(BOM_COL_NAME) or ""),
                    str(item.toolTip(BOM_COL_AES) or ""),
                    str(item.data(0, BOM_TREE_ITEM_NUMBER_ROLE) or ""),
                    str(item.data(0, BOM_TREE_AES_NUMBER_ROLE) or ""),
                )
            )
            basic_match = matches_bom_filter_text(haystack, query)
            is_related_cad = item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD
            if is_related_cad:
                return bool(query) and basic_match, True
            advanced_match = (
                self._bom_tree_item_matches_advanced_filter(item, filters)
                if advanced_active else True
            )
            return basic_match and advanced_match, False

        if flat_results:
            result_tree = getattr(self, "_ebom_filter_tree", None)
            if result_tree is None:
                return 0
            matches = []
            seen_item_ids = set()
            for item in self._iter_tree_items_visual_order(tree):
                self_match, is_related_cad = item_matches(item)
                if not self_match or is_related_cad:
                    continue
                if filters.get("remove_duplicates", False):
                    item_id = item.data(0, Qt.UserRole)
                    if item_id in seen_item_ids:
                        continue
                    seen_item_ids.add(item_id)
                matches.append(item)

            result_tree.setUpdatesEnabled(False)
            try:
                result_tree.clear()
                for source_item in matches:
                    clone = source_item.clone()
                    while clone.childCount():
                        clone.takeChild(0)
                    clone.setHidden(False)
                    clone.setExpanded(False)
                    result_tree.addTopLevelItem(clone)
            finally:
                result_tree.setUpdatesEnabled(True)
            self._ebom_filter_flat_mode = True
            self._tree_stack.setCurrentWidget(result_tree)
            return len(matches)

        self._ebom_filter_flat_mode = False
        try:
            getattr(self, "_ebom_filter_tree", None).clear()
        except Exception:
            pass

        def recurse(item):
            nonlocal visible_count
            if self._is_lazy_placeholder(item):
                item.setHidden(False)
                return False
            self_match, is_related_cad = item_matches(item)
            child_match = False
            for index in range(item.childCount()):
                child_match = recurse(item.child(index)) or child_match
            show = self_match or (show_parents and child_match)
            item.setHidden(not show)
            if show and not query and not is_related_cad:
                # Keep related rows visible beneath an Item that itself passed;
                # their neutral return value above prevents filter bypass.
                for index in range(item.childCount()):
                    child = item.child(index)
                    if child.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                        child.setHidden(False)
            if self_match:
                visible_count += 1
            if child_match and (query or advanced_active):
                item.setExpanded(True)
            return show

        for index in range(tree.topLevelItemCount()):
            recurse(tree.topLevelItem(index))
        self._renumber_tree_rows(tree)
        if getattr(self, "_bom_mode", "cad") == "ebom":
            self._tree_stack.setCurrentWidget(tree)
        return visible_count

    def _apply_ebom_db_flat_filter(self, query: str = "", filters: dict | None = None) -> int:
        """Show EBOM filter results from SQL without expanding the lazy tree.

        When parent branches are enabled, the result is a lightweight filtered
        EBOM tree: every matching descendant is shown under its real ancestor
        path.  This fixes lazy BOM filtering where matches below collapsed
        subassemblies were previously invisible because only loaded siblings
        were inspected.
        """
        result_tree = getattr(self, "_ebom_filter_tree", None)
        if result_tree is None or not self.session.project_id:
            return 0
        filters = filters or self._default_bom_advanced_filters()
        q = str(query or "").strip()
        where = ["b.project_id=?"]
        params = [int(self.session.project_id)]
        table_names = set()
        bom_columns = set()
        relation_table = "bom_children"
        relation_parent = "parent_id"
        relation_child = "child_id"
        relation_project_clause = (
            "EXISTS (SELECT 1 FROM bom bp WHERE bp.id=u.parent_id AND bp.project_id=b.project_id)"
        )
        relation_order_expr = "u.id"
        try:
            with self.bom_service.bom_repo.get_conn() as conn:
                table_names = {
                    str(row[0])
                    for row in conn.execute(
                        "SELECT name FROM sqlite_master WHERE type IN ('table','view')"
                    ).fetchall()
                }
                bom_columns = {
                    str(row[1])
                    for row in conn.execute("PRAGMA table_info(bom)").fetchall()
                }
                if "item_usages" in table_names:
                    relation_table = "item_usages"
                    relation_parent = "parent_item_id"
                    relation_child = "child_item_id"
                    relation_project_clause = "u.project_id=b.project_id"
                relation_columns = {
                    str(row[1])
                    for row in conn.execute(f"PRAGMA table_info({relation_table})").fetchall()
                }
                if "sort_order" in relation_columns:
                    relation_order_expr = "COALESCE(u.sort_order, u.id)"
        except Exception:
            pass

        def bom_expr(column: str, default: str = "") -> str:
            if column in bom_columns:
                return f"COALESCE(b.{column}, {default!r})"
            return f"{default!r}"

        def table_exists(name: str) -> bool:
            return name in table_names

        if "represented_part_id" in bom_columns:
            where.append("b.represented_part_id IS NULL")
        if "deleted_at" in bom_columns:
            where.append("b.deleted_at IS NULL")
        if "status" in bom_columns:
            where.append("lower(COALESCE(b.status,''))<>'deleted'")
        if "lifecycle_state" in bom_columns:
            where.append("lower(COALESCE(b.lifecycle_state,''))<>'deleted'")

        if q:
            like = f"%{q.lower()}%"
            searchable_columns = [
                column for column in (
                    "part_number", "aes_number", "name", "type",
                    "item_type", "lifecycle_state", "status",
                    "drawing_number", "material", "procurement_source",
                )
                if column in bom_columns
            ]
            if searchable_columns:
                where.append(
                    "(" + " OR ".join(
                        f"lower(COALESCE(b.{column},'')) LIKE ?"
                        for column in searchable_columns
                    ) + ")"
                )
                params.extend([like] * len(searchable_columns))
        text = str(filters.get("text") or "").strip()
        if text:
            like = f"%{text.lower()}%"
            text_columns = [
                column for column in ("part_number", "aes_number", "name", "drawing_number")
                if column in bom_columns
            ]
            if text_columns:
                where.append(
                    "(" + " OR ".join(
                        f"lower(COALESCE(b.{column},'')) LIKE ?"
                        for column in text_columns
                    ) + ")"
                )
                params.extend([like] * len(text_columns))
        status = str(filters.get("status") or "All").strip()
        if status and status != "All":
            where.append(f"lower(COALESCE({bom_expr('lifecycle_state')}, {bom_expr('status')}, ''))=?")
            params.append(status.lower())
        part_type = str(filters.get("type") or "All").strip()
        if part_type and part_type != "All" and "type" in bom_columns:
            where.append("lower(COALESCE(b.type,''))=?")
            params.append(part_type.lower())
        revision = str(filters.get("revision") or "").strip()
        if revision and "revision" in bom_columns:
            where.append("lower(COALESCE(b.revision,'')) LIKE ?")
            params.append(f"%{revision.lower()}%")

        structure = str(filters.get("structure") or "Any")
        if structure == "Assemblies only":
            where.append(
                f"EXISTS (SELECT 1 FROM {relation_table} u WHERE {relation_project_clause} AND u.{relation_parent}=b.id)"
            )
        elif structure == "Leaf parts only":
            where.append(
                f"NOT EXISTS (SELECT 1 FROM {relation_table} u WHERE {relation_project_clause} AND u.{relation_parent}=b.id)"
            )
        selected_categories = [
            str(value).strip()
            for value in (filters.get("categories") or [])
            if str(value).strip()
        ]
        legacy_category = str(filters.get("category") or "").strip()
        if not selected_categories and legacy_category not in ("", "All"):
            selected_categories = [legacy_category]
        if selected_categories and table_exists("bom_categories") and table_exists("bom_item_categories"):
            category_keys = {value.casefold() for value in selected_categories}
            category_values = [value for value in selected_categories if value.casefold() != "uncategorized"]
            clauses = []
            if category_values:
                placeholders = ",".join("?" for _ in category_values)
                clauses.append(
                    "EXISTS ("
                    "SELECT 1 FROM bom_item_categories bic "
                    "JOIN bom_categories bc ON bc.id=bic.category_id "
                    "WHERE bic.bom_id=b.id AND bc.project_id=b.project_id "
                    f"AND lower(bc.name) IN ({placeholders})"
                    ")"
                )
                params.extend([value.lower() for value in category_values])
            if "uncategorized" in category_keys:
                clauses.append(
                    "NOT EXISTS (SELECT 1 FROM bom_item_categories bic WHERE bic.bom_id=b.id)"
                )
            if clauses:
                where.append("(" + " OR ".join(clauses) + ")")

        for doc_key in ("pdf", "step"):
            desired = str(filters.get(doc_key) or "Any")
            if desired == "Any" or not table_exists("part_files"):
                continue
            file_type = doc_key.upper()
            exists_sql = (
                "EXISTS ("
                "SELECT 1 FROM part_files pf "
                "WHERE pf.part_id=b.id AND upper(COALESCE(pf.file_type,''))=? "
                "AND COALESCE(pf.deleted_at,'')=''"
                ")"
            )
            if desired in ("OK", "Outdated"):
                where.append(exists_sql)
                params.append(file_type)
            elif desired in ("Missing", "Not attached"):
                where.append("NOT " + exists_sql)
                params.append(file_type)

        work_state = str(filters.get("work_state") or "Any")
        owner = str(filters.get("work_owner") or "All").strip()
        if (work_state != "Any" or (owner and owner != "All")) and table_exists("locks"):
            if work_state == "In Work":
                where.append("EXISTS (SELECT 1 FROM locks l WHERE l.part_id=b.id)")
            elif work_state == "Checked In":
                where.append("NOT EXISTS (SELECT 1 FROM locks l WHERE l.part_id=b.id)")
            if owner and owner != "All" and table_exists("users"):
                where.append(
                    "EXISTS ("
                    "SELECT 1 FROM locks l LEFT JOIN users lu ON lu.id=l.user_id "
                    "WHERE l.part_id=b.id AND lower(COALESCE(lu.username,'')) LIKE ?"
                    ")"
                )
                params.append(f"%{owner.lower()}%")

        issues_filter = str(filters.get("issues") or "Any")
        if issues_filter != "Any" and table_exists("issues") and table_exists("issue_parts"):
            active_issue_sql = (
                "EXISTS ("
                "SELECT 1 FROM issue_parts ip JOIN issues i ON i.id=ip.issue_id "
                "WHERE ip.part_id=b.id AND COALESCE(i.archived,0)=0 "
                "AND lower(COALESCE(i.status,'')) NOT IN ('closed','rejected')"
                ")"
            )
            any_issue_sql = (
                "EXISTS (SELECT 1 FROM issue_parts ip JOIN issues i ON i.id=ip.issue_id "
                "WHERE ip.part_id=b.id AND COALESCE(i.archived,0)=0)"
            )
            if issues_filter == "Active issues":
                where.append(active_issue_sql)
            elif issues_filter == "Any linked issue":
                where.append(any_issue_sql)
            elif issues_filter == "No linked issues":
                where.append("NOT " + any_issue_sql)

        sql = f"""
            SELECT b.id, b.part_number, b.aes_number, b.name, b.type,
                   {bom_expr('drawing_number')} AS drawing_number,
                   {bom_expr('filename')} AS filename,
                   {bom_expr('base_file_name')} AS base_file_name,
                   {bom_expr('base_drw_name')} AS base_drw_name,
                   {bom_expr('material')} AS material,
                   {bom_expr('weight')} AS weight,
                   {bom_expr('notes')} AS notes,
                   {bom_expr('revision', 'A')} AS revision,
                   COALESCE({bom_expr('lifecycle_state')}, {bom_expr('status')}, '') AS lifecycle_state,
                   {bom_expr('item_type')} AS item_type,
                   {bom_expr('assembly_mode')} AS assembly_mode,
                   {bom_expr('classification')} AS classification,
                   {bom_expr('default_ebom_behavior')} AS default_ebom_behavior,
                   {bom_expr('cad_requirement')} AS cad_requirement,
                   {bom_expr('drawing_requirement')} AS drawing_requirement,
                   {bom_expr('cad_control_mode')} AS cad_control_mode,
                   {bom_expr('procurement_source')} AS procurement_source,
                   {bom_expr('item_view')} AS item_view,
                   {bom_expr('default_unit')} AS default_unit,
                   {('b.represented_part_id' if 'represented_part_id' in bom_columns else 'NULL')} AS represented_part_id,
                   {('b.current_revision_id' if 'current_revision_id' in bom_columns else 'NULL')} AS current_revision_id,
                   {('b.current_iteration_id' if 'current_iteration_id' in bom_columns else 'NULL')} AS current_iteration_id,
                   {bom_expr('pending_revision_code')} AS pending_revision_code,
                   {('b.released_by' if 'released_by' in bom_columns else 'NULL')} AS released_by,
                   {bom_expr('released_at')} AS released_at,
                   {bom_expr('created')} AS created,
                   {bom_expr('modified')} AS modified,
                   EXISTS (
                       SELECT 1 FROM {relation_table} u
                       WHERE {relation_project_clause} AND u.{relation_parent}=b.id
                   ) AS has_children
            FROM bom b
            WHERE {' AND '.join(where)}
            ORDER BY lower(COALESCE(b.part_number,'')), lower(b.name), b.id
            LIMIT 500
        """
        show_parent_branches = bool(filters.get("show_parent_matches", True))
        remove_duplicates = bool(filters.get("remove_duplicates", False))
        try:
            with self.bom_service.bom_repo.get_conn() as conn:
                rows = [dict(row) for row in conn.execute(sql, params).fetchall()]
                match_ids = [int(row["id"]) for row in rows if row.get("id") is not None]

                parent_by_child: dict[int, list[int]] = {}
                if show_parent_branches and not remove_duplicates and match_ids:
                    if relation_table == "item_usages":
                        rel_sql = f"""
                            SELECT u.{relation_parent} AS parent_id,
                                   u.{relation_child} AS child_id
                            FROM {relation_table} u
                            JOIN bom p ON p.id=u.{relation_parent}
                            JOIN bom c ON c.id=u.{relation_child}
                            WHERE u.project_id=? AND p.project_id=? AND c.project_id=?
                              {"AND p.deleted_at IS NULL AND c.deleted_at IS NULL" if "deleted_at" in bom_columns else ""}
                              {"AND lower(COALESCE(p.status,''))<>'deleted' AND lower(COALESCE(c.status,''))<>'deleted'" if "status" in bom_columns else ""}
                              {"AND lower(COALESCE(p.lifecycle_state,''))<>'deleted' AND lower(COALESCE(c.lifecycle_state,''))<>'deleted'" if "lifecycle_state" in bom_columns else ""}
                            ORDER BY {relation_order_expr}, u.id
                        """
                        rel_params = (int(self.session.project_id),) * 3
                    else:
                        rel_sql = f"""
                            SELECT u.{relation_parent} AS parent_id,
                                   u.{relation_child} AS child_id
                            FROM {relation_table} u
                            JOIN bom p ON p.id=u.{relation_parent}
                            JOIN bom c ON c.id=u.{relation_child}
                            WHERE p.project_id=? AND c.project_id=?
                              {"AND p.deleted_at IS NULL AND c.deleted_at IS NULL" if "deleted_at" in bom_columns else ""}
                              {"AND lower(COALESCE(p.status,''))<>'deleted' AND lower(COALESCE(c.status,''))<>'deleted'" if "status" in bom_columns else ""}
                              {"AND lower(COALESCE(p.lifecycle_state,''))<>'deleted' AND lower(COALESCE(c.lifecycle_state,''))<>'deleted'" if "lifecycle_state" in bom_columns else ""}
                            ORDER BY {relation_order_expr}, u.id
                        """
                        rel_params = (int(self.session.project_id), int(self.session.project_id))
                    for rel in conn.execute(rel_sql, rel_params).fetchall():
                        parent_id = int(rel["parent_id"])
                        child_id = int(rel["child_id"])
                        parent_by_child.setdefault(child_id, []).append(parent_id)

                    include_ids = set(match_ids)
                    paths: set[tuple[int, ...]] = set()

                    def ancestor_paths(item_id: int, trail: tuple[int, ...] = ()) -> list[tuple[int, ...]]:
                        if item_id in trail:
                            return [(item_id,)]
                        parents = parent_by_child.get(int(item_id), [])
                        if not parents:
                            return [(int(item_id),)]
                        built = []
                        next_trail = trail + (int(item_id),)
                        for parent_id in parents:
                            include_ids.add(int(parent_id))
                            for parent_path in ancestor_paths(int(parent_id), next_trail):
                                built.append(parent_path + (int(item_id),))
                        return built or [(int(item_id),)]

                    for match_id in match_ids:
                        for path in ancestor_paths(int(match_id)):
                            for depth in range(1, len(path) + 1):
                                paths.add(path[:depth])

                    if include_ids:
                        id_list = sorted(include_ids)
                        row_by_id = {}
                        for offset in range(0, len(id_list), 800):
                            chunk = id_list[offset:offset + 800]
                            placeholders = ",".join("?" for _ in chunk)
                            fetch_sql = f"""
                                SELECT b.id, b.part_number, b.aes_number, b.name, b.type,
                                       {bom_expr('drawing_number')} AS drawing_number,
                                       {bom_expr('filename')} AS filename,
                                       {bom_expr('base_file_name')} AS base_file_name,
                                       {bom_expr('base_drw_name')} AS base_drw_name,
                                       {bom_expr('material')} AS material,
                                       {bom_expr('weight')} AS weight,
                                       {bom_expr('notes')} AS notes,
                                       {bom_expr('revision', 'A')} AS revision,
                                       COALESCE({bom_expr('lifecycle_state')}, {bom_expr('status')}, '') AS lifecycle_state,
                                       {bom_expr('item_type')} AS item_type,
                                       {bom_expr('assembly_mode')} AS assembly_mode,
                                       {bom_expr('classification')} AS classification,
                                       {bom_expr('default_ebom_behavior')} AS default_ebom_behavior,
                                       {bom_expr('cad_requirement')} AS cad_requirement,
                                       {bom_expr('drawing_requirement')} AS drawing_requirement,
                                       {bom_expr('cad_control_mode')} AS cad_control_mode,
                                       {bom_expr('procurement_source')} AS procurement_source,
                                       {bom_expr('item_view')} AS item_view,
                                       {bom_expr('default_unit')} AS default_unit,
                                       {('b.represented_part_id' if 'represented_part_id' in bom_columns else 'NULL')} AS represented_part_id,
                                       {('b.current_revision_id' if 'current_revision_id' in bom_columns else 'NULL')} AS current_revision_id,
                                       {('b.current_iteration_id' if 'current_iteration_id' in bom_columns else 'NULL')} AS current_iteration_id,
                                       {bom_expr('pending_revision_code')} AS pending_revision_code,
                                       {('b.released_by' if 'released_by' in bom_columns else 'NULL')} AS released_by,
                                       {bom_expr('released_at')} AS released_at,
                                       {bom_expr('created')} AS created,
                                       {bom_expr('modified')} AS modified,
                                       EXISTS (
                                           SELECT 1 FROM {relation_table} u
                                           WHERE {relation_project_clause} AND u.{relation_parent}=b.id
                                       ) AS has_children
                                FROM bom b
                                WHERE b.project_id=?
                                  {"AND b.represented_part_id IS NULL" if "represented_part_id" in bom_columns else ""}
                                  {"AND b.deleted_at IS NULL" if "deleted_at" in bom_columns else ""}
                                  {"AND lower(COALESCE(b.status,''))<>'deleted'" if "status" in bom_columns else ""}
                                  {"AND lower(COALESCE(b.lifecycle_state,''))<>'deleted'" if "lifecycle_state" in bom_columns else ""}
                                  AND b.id IN ({placeholders})
                            """
                            for row in conn.execute(fetch_sql, [int(self.session.project_id), *chunk]).fetchall():
                                row_by_id[int(row["id"])] = dict(row)

                        rows = []
                        for path in sorted(paths, key=lambda p: (len(p), p)):
                            leaf_id = int(path[-1])
                            row = dict(row_by_id.get(leaf_id) or {})
                            if not row:
                                continue
                            row["_filter_path"] = path
                            row["_filter_is_match"] = leaf_id in set(match_ids)
                            rows.append(row)
        except Exception:
            rows = []
        result_tree.setUpdatesEnabled(False)
        try:
            result_tree.clear()
            path_items: dict[tuple[int, ...], QTreeWidgetItem] = {}
            match_count = 0
            for row in rows:
                part_id = int(row["id"])
                if row.get("_filter_is_match", True):
                    match_count += 1
                info = {
                    "id": part_id,
                    "bom_id": part_id,
                    "part_number": row.get("part_number") or "",
                    "aes_number": row.get("aes_number") or "",
                    "name": row.get("name") or "",
                    "type": row.get("type") or "",
                    "current_version": row.get("revision") or "A",
                    "revision": row.get("revision") or "A",
                    "status": row.get("lifecycle_state") or "",
                    "lifecycle_state": row.get("lifecycle_state") or "",
                    "item_type": row.get("item_type") or "",
                    "drawing_number": row.get("drawing_number") or "",
                    "filename": row.get("filename") or row.get("base_file_name") or "",
                    "base_file_name": row.get("base_file_name") or "",
                    "base_drw_name": row.get("base_drw_name") or "",
                    "material": row.get("material") or "",
                    "weight": row.get("weight") or "",
                    "notes": row.get("notes") or "",
                    "assembly_mode": row.get("assembly_mode") or "",
                    "classification": row.get("classification") or "",
                    "default_ebom_behavior": row.get("default_ebom_behavior") or "",
                    "cad_requirement": row.get("cad_requirement") or "",
                    "drawing_requirement": row.get("drawing_requirement") or "",
                    "cad_control_mode": row.get("cad_control_mode") or "",
                    "procurement_source": row.get("procurement_source") or "",
                    "item_view": row.get("item_view") or "",
                    "default_unit": row.get("default_unit") or "",
                    "represented_part_id": row.get("represented_part_id"),
                    "current_revision_id": row.get("current_revision_id"),
                    "current_iteration_id": row.get("current_iteration_id"),
                    "pending_revision_code": row.get("pending_revision_code") or "",
                    "released_by": row.get("released_by"),
                    "released_at": row.get("released_at") or "",
                    "created": row.get("created") or "",
                    "modified": row.get("modified") or "",
                    "_has_children": bool(row.get("has_children")),
                    "_defer_indicators": True,
                    "children": [],
                }
                item = QTreeWidgetItem([""] * result_tree.columnCount())
                self._apply_tree_item_data(item, info)
                item.setText(BOM_COL_AES, str(info.get("part_number") or ""))
                item.setData(0, PDM_OBJECT_KIND_ROLE, PDM_OBJECT_ITEM)
                item.setData(0, PDM_NODE_PAYLOAD_ROLE, dict(info))
                item.setData(0, PDM_CHILDREN_PAYLOAD_ROLE, [])
                item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, bool(row.get("has_children")))
                # This is a filtered SQL projection.  Do not mark it as a live
                # lazy branch or add a Loading... placeholder; the real EBOM
                # tree owns lazy loading.  Otherwise a searched-but-unloaded
                # assembly can get stuck showing a loading spinner.
                item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, True)
                item.setData(0, BOM_TREE_LOADING_ROLE, False)
                item.setData(
                    0,
                    PDM_EBOM_ASSOCIATIONS_ROLE,
                    list((getattr(self, "_ebom_associations_by_item", {}) or {}).get(part_id, [])),
                )
                item.setData(0, PDM_ASSOCIATIONS_SHOWN_ROLE, False)
                item.setIcon(BOM_COL_NAME, _pdm_item_icon())
                path = tuple(row.get("_filter_path") or (part_id,))
                item.setData(0, BOM_TREE_PATH_ROLE, " > ".join(str(value) for value in path))
                path_items[path] = item
                parent_path = path[:-1]
                parent_item = path_items.get(parent_path) if parent_path else None
                if parent_item is not None:
                    parent_item.addChild(item)
                    parent_item.setExpanded(True)
                else:
                    result_tree.addTopLevelItem(item)
        finally:
            result_tree.setUpdatesEnabled(True)
        self._ebom_filter_flat_mode = True
        self._tree_stack.setCurrentWidget(result_tree)
        self._sync_search_tree_row_numbers()
        return match_count if show_parent_branches and not remove_duplicates else len(rows)

    def _show_tree_placeholder(self, _message: str) -> None:
        # Backward compatible; loader is now a spinner overlay.
        self._set_tree_loading(True)
        
    def auto_check_wd(self):

        if not self.session.project_id or not self.working_dir or not os.path.isdir(self.working_dir):
            try:
                self.show_alert("No project loaded. Select a project from the toolbar to begin.", "error")
            except Exception:
                pass
            return
        
        working_result = self.diag_service.check_working_directory(self.working_dir)
        if working_result == "error_no_snapshot":
            self.show_alert("⚠️ No snapshots found for this project. Please create a snapshot to enable integrity checks.", "error")
            return
        elif working_result:
            self.show_alert("⚠️ Integrity issue detected in working directory. Run a diagnostic before proceeding.", "error")
        else:
            self.hide_alert()

    def _start_initial_diagnostics_and_load(self) -> None:
        # Display the lazy root level immediately. Diagnostics are independent and
        # update badges later; disk scanning must never gate BOM navigation.
        self.load_tree()
        self._load_released_ebom_tree()
        try:
            if self._diag_worker is not None:
                self._diag_worker.cancel()
        except Exception:
            pass

        worker = _InitialDiagWorker(str(self.working_dir or ""))
        thread = QThread(self)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(self._on_initial_diag_finished)
        worker.failed.connect(self._on_initial_diag_failed)
        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        self._diag_worker = worker
        self._diag_thread = thread
        thread.start()

    def _on_initial_diag_finished(self, missing_files: object, working_result: object) -> None:
        try:
            self.missing_files = list(missing_files or [])
        except Exception:
            self.missing_files = []

        try:
            ids = set()
            for row in (self.missing_files or []):
                try:
                    bom_id = row[0]
                    ids.add(int(bom_id))
                except Exception:
                    continue
            self.missing_ids = ids
        except Exception:
            self.missing_ids = set()

        # Apply integrity banner result (same behavior as auto_check_wd).
        try:
            if working_result == "error_no_snapshot":
                self.show_alert("⚠️ No snapshots found for this project. Please create a snapshot to enable integrity checks.", "error")
            elif working_result:
                self.show_alert("⚠️ Integrity issue detected in working directory. Run a diagnostic before proceeding.", "error")
            else:
                self.hide_alert()
        except Exception:
            pass

        self._refresh_loaded_integrity_indicators()

    def _on_initial_diag_failed(self, _err: str) -> None:
        try:
            self.missing_files = []
            self.missing_ids = set()
        except Exception:
            pass

    def _refresh_loaded_integrity_indicators(self) -> None:
        self._invalidate_doc_indicator()
        for tree in (
            getattr(self, "tree", None), getattr(self, "_search_tree", None),
            getattr(self, "_ebom_tree", None),
        ):
            if tree is None:
                continue
            for item in self._iter_tree_items(tree):
                if self._is_folder_tree_item(item) or self._is_lazy_placeholder(item):
                    continue
                if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                    continue
                part_id = item.data(0, Qt.UserRole)
                item.setData(
                    BOM_COL_INTEGRITY,
                    BOM_TREE_INTEGRITY_ROLE,
                    _integrity_payload(part_id, self.missing_files, self.missing_ids),
                )
                self._indicator_refresh_queue.append((item, part_id))
        if self._indicator_refresh_queue and not self._indicator_refresh_timer.isActive():
            self._indicator_refresh_timer.start()



    def get_working_dir(self):
        if not self.session.project_id:
            return None
        project = self.project_service.get_project_by_id(self.session.project_id) or {}
        project_working_dir = project.get("working_directory")
        if project_working_dir:
            return project_working_dir

    def _bom_tree_headers(self, ebom: bool = False) -> list[str]:
        base_name = "Item / Related CAD" if ebom else "Name"
        number_label = "Number / Association" if ebom else "AES Number"
        return [
            "#", base_name, "Files", number_label, "Object", "Rev/Iter",
            "Status", "Integrity", "Source Qty", "Effective Qty", "Level",
        ] + [label for _key, label, _width, _default, _aliases in BOM_EXTRA_COLUMN_SPECS]

    def _default_visible_bom_column_keys(self) -> set[str]:
        return {
            key for key, _label, _width, default_visible, _aliases in BOM_EXTRA_COLUMN_SPECS
            if default_visible
        }

    def _visible_bom_column_keys(self) -> set[str]:
        settings = QSettings("Nexus", "NexusPDM")
        raw = str(settings.value("bom/visible_extra_columns", "") or "").strip()
        if not raw:
            return self._default_visible_bom_column_keys()
        try:
            loaded = json.loads(raw)
            return {
                str(value)
                for value in (loaded or [])
                if str(value) in {spec[0] for spec in BOM_EXTRA_COLUMN_SPECS}
            }
        except Exception:
            return self._default_visible_bom_column_keys()

    def _save_visible_bom_column_keys(self, keys: set[str]) -> None:
        QSettings("Nexus", "NexusPDM").setValue(
            "bom/visible_extra_columns",
            json.dumps(sorted(str(key) for key in keys)),
        )

    def _configure_bom_tree_columns(self, tree: QTreeWidget | None, ebom: bool = False) -> None:
        if tree is None:
            return
        tree.setHeaderLabels(self._bom_tree_headers(ebom=ebom))
        widths = {
            BOM_COL_ROW: 38,
            BOM_COL_NAME: 260 if ebom else 280,
            BOM_COL_FILES: 100,
            BOM_COL_AES: 125 if ebom else 90,
            BOM_COL_TYPE: 70,
            BOM_COL_REV: 70,
            BOM_COL_STATUS: 85,
            BOM_COL_INTEGRITY: 55,
            EBOM_COL_SOURCE_QTY: 74,
            EBOM_COL_EFFECTIVE_QTY: 84,
            EBOM_COL_LEVEL: 45,
        }
        for column, width in widths.items():
            try:
                tree.setColumnWidth(column, width)
            except Exception:
                pass
        for offset, (_key, _label, width, _default, _aliases) in enumerate(BOM_EXTRA_COLUMN_SPECS):
            try:
                tree.setColumnWidth(BOM_COL_EXTRA_START + offset, int(width))
            except Exception:
                pass
        self._apply_bom_column_visibility(tree, ebom=ebom)
        try:
            tree.setTreePosition(BOM_COL_NAME)
        except Exception:
            pass
        try:
            header = tree.header()
            if not bool(header.property("nexusColumnMenuInstalled")):
                header.setContextMenuPolicy(Qt.CustomContextMenu)
                header.customContextMenuRequested.connect(
                    lambda _pos, self=self: self.show_bom_column_dialog()
                )
                header.setProperty("nexusColumnMenuInstalled", True)
                header.setToolTip("Right-click to manage visible tree columns.")
        except Exception:
            pass

    def _apply_bom_column_visibility(self, tree: QTreeWidget | None = None, ebom: bool | None = None) -> None:
        trees = []
        if tree is not None:
            trees = [(tree, bool(ebom))]
        else:
            trees = [
                (getattr(self, "tree", None), False),
                (getattr(self, "_search_tree", None), False),
                (getattr(self, "_ebom_tree", None), True),
                (getattr(self, "_ebom_filter_tree", None), True),
            ]
        visible_keys = self._visible_bom_column_keys()
        for target_tree, is_ebom in trees:
            if target_tree is None:
                continue
            # EBOM-specific quantity/level columns stay visible in EBOM only.
            for column in (EBOM_COL_SOURCE_QTY, EBOM_COL_EFFECTIVE_QTY, EBOM_COL_LEVEL):
                try:
                    target_tree.setColumnHidden(column, not is_ebom)
                except Exception:
                    pass
            for offset, (key, _label, _width, _default, _aliases) in enumerate(BOM_EXTRA_COLUMN_SPECS):
                try:
                    target_tree.setColumnHidden(BOM_COL_EXTRA_START + offset, key not in visible_keys)
                except Exception:
                    pass

    def _extra_bom_column_value(self, info: dict, spec: tuple) -> str:
        key, _label, _width, _default, aliases = spec
        for alias in aliases:
            value = info.get(alias)
            if value not in (None, ""):
                if isinstance(value, (list, tuple, set)):
                    return ", ".join(str(v) for v in value if str(v).strip())
                return str(value)
        if key == "effective_quantity":
            value = info.get("effective_quantity")
            if value in (None, ""):
                value = info.get("quantity")
            return "" if value in (None, "") else str(value)
        return ""

    def _apply_extra_bom_columns(self, item: QTreeWidgetItem, info: dict) -> None:
        for offset, spec in enumerate(BOM_EXTRA_COLUMN_SPECS):
            column = BOM_COL_EXTRA_START + offset
            try:
                value = self._extra_bom_column_value(info or {}, spec)
                item.setText(column, value)
                if value:
                    item.setToolTip(column, value)
            except Exception:
                pass

    def show_bom_column_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Tree Columns")
        dialog.resize(520, 620)
        layout = QVBoxLayout(dialog)
        title = QLabel("Select BOM Tree Columns")
        title.setStyleSheet("font-weight:700;font-size:13px;color:#1f3347;")
        layout.addWidget(title)
        help_label = QLabel(
            "Choose the metadata columns shown in the Product Structure tree. "
            "The main tree column remains fixed, like Creo/Windchill."
        )
        help_label.setWordWrap(True)
        layout.addWidget(help_label)

        list_widget = QListWidget()
        list_widget.setAlternatingRowColors(True)
        visible = self._visible_bom_column_keys()
        for key, label, _width, _default, _aliases in BOM_EXTRA_COLUMN_SPECS:
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, key)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if key in visible else Qt.Unchecked)
            if key == "aes_number":
                item.setToolTip("AES delivery reference. Kept available even when Item Number is the primary PLM number.")
            list_widget.addItem(item)
        layout.addWidget(list_widget, 1)

        quick_row = QHBoxLayout()
        select_all = QPushButton("Select All")
        clear_all = QPushButton("Clear Optional")
        default_btn = QPushButton("Default")
        quick_row.addWidget(select_all)
        quick_row.addWidget(clear_all)
        quick_row.addWidget(default_btn)
        quick_row.addStretch()
        layout.addLayout(quick_row)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)

        def set_all(state):
            for row in range(list_widget.count()):
                list_widget.item(row).setCheckState(state)

        select_all.clicked.connect(lambda: set_all(Qt.Checked))
        clear_all.clicked.connect(lambda: set_all(Qt.Unchecked))

        def reset_default():
            defaults = self._default_visible_bom_column_keys()
            for row in range(list_widget.count()):
                item = list_widget.item(row)
                item.setCheckState(Qt.Checked if item.data(Qt.UserRole) in defaults else Qt.Unchecked)

        default_btn.clicked.connect(reset_default)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        if dialog.exec_() != QDialog.Accepted:
            return
        selected = set()
        for row in range(list_widget.count()):
            item = list_widget.item(row)
            if item.checkState() == Qt.Checked:
                selected.add(str(item.data(Qt.UserRole)))
        self._save_visible_bom_column_keys(selected)
        self._apply_bom_column_visibility()
        try:
            self.window().statusBar().showMessage("BOM tree columns updated.")
        except Exception:
            pass

    def init_ui(self):
        self.setObjectName("pdmWorkspace")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 5, 6, 6)
        layout.setSpacing(5)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setObjectName("pdmWorkspaceSplitter")
        splitter.setHandleWidth(4)
        splitter.setChildrenCollapsible(False)
        layout.addWidget(splitter, 1)

        # Left panel
        left_widget = QWidget()
        left_widget.setObjectName("structureWorkspace")
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 2, 0)
        left_layout.setSpacing(5)

        # Alert section (hidden by default)
        self.alert_frame = QFrame()
        self.alert_frame.setFrameShape(QFrame.StyledPanel)
        self.alert_frame.setFrameShadow(QFrame.Raised)
        self.alert_frame.setStyleSheet("""
            QFrame {
                background-color: #fff4f2;
                border: 1px solid #d98b82;
                border-left: 4px solid #b42318;
                border-radius: 0;
            }
            QLabel {
                color: #7a271a;
                font-weight: 600;
                padding: 4px 6px;
            }
        """)
        self.alert_label = QLabel("")
        alert_layout = QHBoxLayout(self.alert_frame)
        alert_layout.addWidget(self.alert_label)
        self.alert_frame.hide()  # Hidden by default
        left_layout.addWidget(self.alert_frame)

        

        # Search group
        search_group = QGroupBox("FIND IN STRUCTURE")
        search_group.setObjectName("structureSearch")
        search_layout = QHBoxLayout(search_group)
        search_layout.setContentsMargins(7, 8, 7, 6)
        search_layout.setSpacing(5)
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search Item Number, name, or AES...")
        try:
            self.search_input.textChanged.connect(self._schedule_search)
        except Exception:
            pass
        self.search_input.returnPressed.connect(self._perform_search_now)
        self.search_btn = QPushButton("Search")
        self.search_btn.setObjectName("primary")
        self.search_btn.setFixedHeight(25)
        self.search_btn.clicked.connect(self._perform_search_now)
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.search_btn)
        left_layout.addWidget(search_group)

        filter_row = QHBoxLayout()
        self.advanced_filter_btn = QPushButton("Advanced Filter")
        self.advanced_filter_btn.setProperty("structureTool", True)
        self.advanced_filter_btn.clicked.connect(self.show_advanced_filter_dialog)
        self.saved_filters_btn = QPushButton("Saved Filters")
        self.saved_filters_btn.setProperty("structureTool", True)
        self.saved_filters_btn.clicked.connect(self.show_saved_bom_filters_menu)
        self.clear_filter_btn = QPushButton("Clear")
        self.clear_filter_btn.setProperty("structureTool", True)
        self.clear_filter_btn.clicked.connect(self.clear_bom_tree_filter)
        self.clear_filter_btn.setEnabled(False)
        self.bom_columns_btn = QPushButton("Columns")
        self.bom_columns_btn.setProperty("structureTool", True)
        self.bom_columns_btn.setToolTip("Manage visible Product Structure tree columns.")
        self.bom_columns_btn.clicked.connect(self.show_bom_column_dialog)
        filter_row.addWidget(self.advanced_filter_btn)
        filter_row.addWidget(self.saved_filters_btn)
        filter_row.addWidget(self.clear_filter_btn)
        filter_row.addWidget(self.bom_columns_btn)
        left_layout.addLayout(filter_row)

        # Tree (BOM structure)
        tree_group = QFrame()
        tree_group.setObjectName("structureBrowser")
        tree_layout = QVBoxLayout(tree_group)
        tree_layout.setContentsMargins(7, 6, 7, 7)
        tree_layout.setSpacing(5)
        bom_header = QLabel("PRODUCT STRUCTURE")
        bom_header.setStyleSheet("""
            QLabel {
                font-size: 8pt;
                font-weight: 700;
                color: #3f5368;
                letter-spacing: 0.08em;
                background: transparent;
                border: none;
                padding: 0px;
            }
        """)
        bom_header_row = QHBoxLayout()
        bom_header_row.setSpacing(5)
        bom_header_row.addWidget(bom_header)
        view_caption = QLabel("VIEW")
        view_caption.setObjectName("structureViewCaption")
        bom_header_row.addWidget(view_caption)
        self.bom_mode_selector = QComboBox()
        self.bom_mode_selector.addItem("CAD Structure", "cad")
        self.bom_mode_selector.addItem("EBOM / Item Structure", "ebom")
        self.bom_mode_selector.setCurrentIndex(
            self.bom_mode_selector.findData("ebom")
        )
        self.bom_mode_selector.setObjectName("structureViewSelector")
        self.bom_mode_selector.setFixedWidth(190)
        self.bom_mode_selector.setFixedHeight(24)
        self.bom_mode_selector.setToolTip(
            "Switch between the native CAD assembly and the independent Item/EBOM structure."
        )
        self.bom_mode_selector.currentIndexChanged.connect(
            self._on_bom_mode_changed
        )
        bom_header_row.addWidget(self.bom_mode_selector)
        self.bom_export_btn = QPushButton("Export")
        self.bom_export_btn.setObjectName("neutral")
        self.bom_export_btn.setFixedHeight(24)
        self.bom_export_btn.setProperty("structureTool", True)
        self.bom_export_btn.clicked.connect(self.export_bom)
        bom_header_row.addWidget(self.bom_export_btn)
        self.bom_compare_btn = QPushButton("Compare")
        self.bom_compare_btn.setObjectName("neutral")
        self.bom_compare_btn.setFixedHeight(24)
        self.bom_compare_btn.setProperty("structureTool", True)
        self.bom_compare_btn.setToolTip(
            "Compare the top-level EBOM Item against its OWNER CAD Document structure."
        )
        self.bom_compare_btn.clicked.connect(
            self.compare_top_level_cad_to_item_structure
        )
        bom_header_row.addWidget(self.bom_compare_btn)
        self.pdm_actions_btn = QPushButton("PDM")
        self.pdm_actions_btn.setObjectName("neutral")
        self.pdm_actions_btn.setFixedHeight(24)
        self.pdm_actions_btn.setToolTip(
            "Manage CAD–Item associations, compare structures, and build the Item Structure."
        )
        pdm_menu = QMenu(self.pdm_actions_btn)
        register_cad_action = pdm_menu.addAction("Register CAD Document...")
        register_cad_action.triggered.connect(self.register_cad_document)
        cad_structure_action = pdm_menu.addAction("CAD Document Structure...")
        cad_structure_action.triggered.connect(self.manage_pdm_cad_structure)
        manage_associations_action = pdm_menu.addAction("CAD–Item Associations...")
        manage_associations_action.triggered.connect(self.manage_cad_item_associations)
        auto_associate_action = pdm_menu.addAction("Auto Associate CAD Documents")
        auto_associate_action.triggered.connect(self.auto_associate_cad_documents)
        pdm_menu.addSeparator()
        compare_pdm_action = pdm_menu.addAction("Compare CAD to Item Structure...")
        compare_pdm_action.triggered.connect(self.compare_cad_to_item_structure)
        build_pdm_action = pdm_menu.addAction("Build Item Structure from CAD")
        build_pdm_action.triggered.connect(self.build_item_structure_from_cad)
        pdm_menu.addSeparator()
        manual_usage_action = pdm_menu.addAction("Add Manual Item Usage...")
        manual_usage_action.triggered.connect(self.add_manual_item_usage)
        self.pdm_actions_btn.setMenu(pdm_menu)
        bom_header_row.addWidget(self.pdm_actions_btn)
        # Association/build commands now live on the selected Item or CAD row.
        # Keep the legacy object for compatibility with older integrations, but
        # do not expose a detached global PDM command bucket in the new UX.
        self.pdm_actions_btn.hide()
        bom_header_row.addStretch()
        self.bom_health_label = QLabel("Health: --")
        self.bom_health_label.setStyleSheet(
            "font-size:8pt;font-weight:600;color:#40566d;background:#edf2f6;"
            "border:1px solid #c7d1dc;padding:2px 6px;"
        )
        bom_header_row.addWidget(self.bom_health_label)
        tree_layout.addLayout(bom_header_row)

        self._bom_scope_frame = QFrame()
        self._bom_scope_frame.setObjectName("bomScopeBar")
        scope_layout = QHBoxLayout(self._bom_scope_frame)
        scope_layout.setContentsMargins(6, 3, 6, 3)
        scope_layout.setSpacing(4)
        self._bom_scope_top_btn = QPushButton("Top Level")
        self._bom_scope_top_btn.setObjectName("scopeTop")
        self._bom_scope_top_btn.setFixedHeight(22)
        self._bom_scope_top_btn.clicked.connect(self._clear_pdm_isolation)
        scope_layout.addWidget(self._bom_scope_top_btn)
        self._bom_scope_path_layout = QHBoxLayout()
        self._bom_scope_path_layout.setContentsMargins(0, 0, 0, 0)
        self._bom_scope_path_layout.setSpacing(3)
        scope_layout.addLayout(self._bom_scope_path_layout, 1)
        self._bom_scope_frame.setStyleSheet("""
            QFrame#bomScopeBar {
                background: #edf2f6;
                border: 1px solid #bcc8d4;
                border-left: 3px solid #4f81a8;
                border-radius: 0;
            }
            QPushButton#scopeTop, QPushButton[scopeCrumb="true"] {
                background: transparent;
                border: 1px solid transparent;
                color: #174f78;
                font-size: 10px;
                font-weight: 600;
                padding: 1px 6px;
            }
            QPushButton#scopeTop:hover, QPushButton[scopeCrumb="true"]:hover {
                background: #d9e8f4;
                border-color: #8fb2ce;
            }
            QLabel[scopeSeparator="true"] {
                color: #64748b;
                background: transparent;
                font-size: 10px;
                padding: 0 1px;
            }
        """)
        self._bom_scope_frame.hide()
        tree_layout.addWidget(self._bom_scope_frame)

        self.tree = BomTreeWidget()
        try:
            self.tree.setIconSize(QSize(12, 12))
        except Exception:
            pass
        try:
            self.tree.setIndentation(14)
            self.tree.setAnimated(True)
            self.tree.setAlternatingRowColors(True)
            self.tree.setUniformRowHeights(True)
            self.tree.setMouseTracking(True)
            self.tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
            self.tree.itemEntered.connect(self._on_tree_item_entered)
        except Exception:
            pass
        self._configure_bom_tree_columns(self.tree, ebom=False)
        self.tree.itemClicked.connect(self.on_tree_item_clicked)
        self.tree.itemExpanded.connect(self._on_bom_item_expanded)
        try:
            self.tree.itemDoubleClicked.connect(self._on_tree_item_double_clicked)
        except Exception:
            pass
        self.tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tree.customContextMenuRequested.connect(self.show_tree_context_menu)
        self.tree.reorderRequested.connect(self._handle_tree_drag_reorder)

        # Tree loading overlay (spinner) inside the tree area
        self._tree_stack = QStackedWidget()
        self._tree_loading_widget = QWidget()
        _loading_layout = QVBoxLayout(self._tree_loading_widget)
        _loading_layout.setContentsMargins(0, 0, 0, 0)
        _loading_layout.setAlignment(Qt.AlignCenter)
        _loading_layout.addWidget(InlineSpinner(size=48))
        self._tree_stack.addWidget(self._tree_loading_widget)  # index 0
        self._tree_stack.addWidget(self.tree)                  # index 1

        # Search-results tree — shown instead of self.tree while a search is active.
        # self.tree is never cleared or rebuilt due to search; switching is a pure UI swap.
        self._search_tree = QTreeWidget()
        try:
            self._search_tree.setIconSize(QSize(12, 12))
        except Exception:
            pass
        try:
            self._search_tree.setIndentation(14)
            self._search_tree.setAnimated(True)
            self._search_tree.setAlternatingRowColors(True)
            self._search_tree.setUniformRowHeights(True)
            self._search_tree.setMouseTracking(True)
            self._search_tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
            self._search_tree.itemEntered.connect(self._on_tree_item_entered)
        except Exception:
            pass
        self._configure_bom_tree_columns(self._search_tree, ebom=False)
        self._search_tree.itemClicked.connect(self.on_tree_item_clicked)
        try:
            self._search_tree.itemDoubleClicked.connect(self._on_tree_item_double_clicked)
        except Exception:
            pass
        self._search_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self._search_tree.customContextMenuRequested.connect(self.show_tree_context_menu)
        self._tree_stack.addWidget(self._search_tree)          # index 2

        # The persisted Item Structure has its own tree so CAD Document and Item
        # selection/editing state remain independent.
        self._ebom_tree = BomTreeWidget()
        self._ebom_tree.setProperty("pdmScope", "EBOM")
        self._configure_bom_tree_columns(self._ebom_tree, ebom=True)
        self._ebom_tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._ebom_tree.setAlternatingRowColors(True)
        self._ebom_tree.setUniformRowHeights(True)
        self._ebom_tree.setIndentation(14)
        self._ebom_tree.setAnimated(True)
        self._ebom_tree.setMouseTracking(True)
        self._ebom_tree.itemClicked.connect(self.on_tree_item_clicked)
        self._ebom_tree.itemExpanded.connect(self._on_pdm_tree_item_expanded)
        self._ebom_tree.itemSelectionChanged.connect(self._sync_visual_action_states)
        self._ebom_tree.itemDoubleClicked.connect(self._on_tree_item_double_clicked)
        self._ebom_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self._ebom_tree.customContextMenuRequested.connect(
            self._show_pdm_tree_context_menu
        )
        self._ebom_tree.reorderRequested.connect(self._handle_pdm_tree_drag_reorder)
        self._ebom_tree.folderReorderRequested.connect(
            self._handle_pdm_folder_drag_reorder
        )
        self._tree_stack.addWidget(self._ebom_tree)             # index 3

        # Flat advanced-filter results live in a separate view.  A child row
        # cannot remain visible inside QTreeWidget when its parent is hidden,
        # so this preserves the source hierarchy while showing exact matches.
        self._ebom_filter_tree = BomTreeWidget()
        self._ebom_filter_tree.setDragEnabled(False)
        self._ebom_filter_tree.setAcceptDrops(False)
        self._ebom_filter_tree.setDragDropMode(QAbstractItemView.NoDragDrop)
        self._configure_bom_tree_columns(self._ebom_filter_tree, ebom=True)
        self._ebom_filter_tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._ebom_filter_tree.setAlternatingRowColors(True)
        self._ebom_filter_tree.setUniformRowHeights(True)
        self._ebom_filter_tree.setIndentation(14)
        self._ebom_filter_tree.setMouseTracking(True)
        self._ebom_filter_tree.itemClicked.connect(self.on_tree_item_clicked)
        self._ebom_filter_tree.itemSelectionChanged.connect(
            self._sync_visual_action_states
        )
        self._ebom_filter_tree.itemDoubleClicked.connect(
            self._on_tree_item_double_clicked
        )
        self._ebom_filter_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self._ebom_filter_tree.customContextMenuRequested.connect(
            self._show_pdm_tree_context_menu
        )
        self._tree_stack.addWidget(self._ebom_filter_tree)      # index 4

        # Managed CAD Documents have their own native structure browser.  This
        # schema intentionally has no Item PDF/STEP or integrity columns.
        self._cad_tree = BomTreeWidget()
        self._cad_tree.setProperty("pdmScope", "CAD")
        self._cad_tree.setHeaderLabels([
            "CAD Name", "Description", "Category", "CAD / Creo Ver",
            "Lifecycle", "Related Item", "Checkout", "Build", "Qty",
        ])
        self._cad_tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._cad_tree.setAlternatingRowColors(True)
        self._cad_tree.setUniformRowHeights(True)
        self._cad_tree.setIndentation(14)
        self._cad_tree.setAnimated(True)
        self._cad_tree.setMouseTracking(True)
        self._cad_tree.setIconSize(QSize(14, 14))
        self._cad_tree.setTreePosition(CAD_COL_NUMBER)
        for column, width in enumerate((220, 175, 72, 130, 76, 185, 100, 65, 42)):
            self._cad_tree.setColumnWidth(column, width)
        self._cad_tree.itemClicked.connect(self.on_tree_item_clicked)
        self._cad_tree.itemExpanded.connect(self._on_pdm_tree_item_expanded)
        self._cad_tree.itemSelectionChanged.connect(self._sync_visual_action_states)
        self._cad_tree.itemDoubleClicked.connect(self._on_tree_item_double_clicked)
        self._cad_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self._cad_tree.customContextMenuRequested.connect(
            self._show_pdm_tree_context_menu
        )
        self._cad_tree.reorderRequested.connect(self._handle_pdm_tree_drag_reorder)
        self._cad_tree.folderReorderRequested.connect(
            self._handle_pdm_folder_drag_reorder
        )
        self._tree_stack.addWidget(self._cad_tree)              # index 5

        _bom_tree_qss = f"""
            QTreeWidget {{
                background: #FFFFFF;
                border: 1px solid #aeb9c5;
                border-radius: 0;
                font-size: 10px;
                gridline-color: #d9dfe6;
                font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display", "Segoe UI", system-ui, sans-serif;
                font-weight: 400;
                letter-spacing: 0;
                text-transform: none;
                show-decoration-selected: 1;
            }}
            QHeaderView::section {{
                background-color: #e7ebef;
                font-size: 10px;
                color: #2f4152;
                font-weight: 600;
                border-top: 1px solid #f7f9fa;
                border-bottom: 1px solid #aeb9c5;
                padding: 4px 5px;
                border-right: 1px solid #c4cdd6;
            }}
            QTreeWidget::item {{
                height: 23px;
                border: none;
                border-bottom: 1px solid #e5e9ed;
                background: #FFFFFF;
                color: {_BOM_TREE_ROW_TEXT};
            }}
            QTreeWidget::item:alternate {{
                background: #f7f8fa;
            }}
            QTreeWidget::item:hover {{
                background: #eaf2f8;
            }}
            QTreeWidget::item:selected {{
                background: {_BOM_TREE_SEL_BG};
                color: {_BOM_TREE_ROW_TEXT};
            }}
            QTreeWidget::item:selected:active {{
                background: {_BOM_TREE_SEL_BG};
                color: {_BOM_TREE_ROW_TEXT};
            }}
            QTreeWidget::branch:selected {{
                background: {_BOM_TREE_SEL_BG};
            }}
            QTreeWidget::branch:selected:alternate {{
                background: {_BOM_TREE_SEL_BG};
            }}
        """
        self.tree.setStyleSheet(_bom_tree_qss)
        self._search_tree.setStyleSheet(_bom_tree_qss)
        self._ebom_tree.setStyleSheet(_bom_tree_qss)
        self._ebom_filter_tree.setStyleSheet(_bom_tree_qss)
        self._cad_tree.setStyleSheet(_bom_tree_qss)

        def _bom_tree_sel_palette(w):
            try:
                pal = w.palette()
                cbg = QColor(_BOM_TREE_SEL_BG)
                cfg = QColor(_BOM_TREE_ROW_TEXT)
                for cg in (QPalette.Active, QPalette.Inactive):
                    pal.setColor(cg, QPalette.Highlight, cbg)
                    pal.setColor(cg, QPalette.HighlightedText, cfg)
                w.setPalette(pal)
            except Exception:
                pass

        for _tw in (
            self.tree, self._search_tree, self._ebom_tree,
            self._ebom_filter_tree, self._cad_tree,
        ):
            try:
                _tw.setShowDecorationSelected(True)
            except Exception:
                pass
            _bom_tree_sel_palette(_tw)

        self.tree.setItemDelegateForColumn(BOM_COL_NAME, _BomTreeNameDelegate(self.tree, self.tree))
        self.tree.setItemDelegateForColumn(BOM_COL_FILES, _BomTreeFilesDelegate(self.tree, self.tree))
        self.tree.setItemDelegateForColumn(BOM_COL_STATUS, _BomTreeStatusDelegate(self.tree, self.tree))
        self.tree.setItemDelegateForColumn(BOM_COL_INTEGRITY, _BomTreeIntegrityDelegate(self.tree, self.tree))
        self._search_tree.setItemDelegateForColumn(BOM_COL_NAME, _BomTreeNameDelegate(self._search_tree, self._search_tree))
        self._search_tree.setItemDelegateForColumn(BOM_COL_FILES, _BomTreeFilesDelegate(self._search_tree, self._search_tree))
        self._search_tree.setItemDelegateForColumn(BOM_COL_STATUS, _BomTreeStatusDelegate(self._search_tree, self._search_tree))
        self._search_tree.setItemDelegateForColumn(BOM_COL_INTEGRITY, _BomTreeIntegrityDelegate(self._search_tree, self._search_tree))
        self._ebom_tree.setItemDelegateForColumn(BOM_COL_NAME, _BomTreeNameDelegate(self._ebom_tree, self._ebom_tree))
        self._ebom_tree.setItemDelegateForColumn(BOM_COL_FILES, _BomTreeFilesDelegate(self._ebom_tree, self._ebom_tree))
        self._ebom_tree.setItemDelegateForColumn(BOM_COL_STATUS, _BomTreeStatusDelegate(self._ebom_tree, self._ebom_tree))
        self._ebom_tree.setItemDelegateForColumn(BOM_COL_INTEGRITY, _BomTreeIntegrityDelegate(self._ebom_tree, self._ebom_tree))
        self._ebom_filter_tree.setItemDelegateForColumn(BOM_COL_NAME, _BomTreeNameDelegate(self._ebom_filter_tree, self._ebom_filter_tree))
        self._ebom_filter_tree.setItemDelegateForColumn(BOM_COL_FILES, _BomTreeFilesDelegate(self._ebom_filter_tree, self._ebom_filter_tree))
        self._ebom_filter_tree.setItemDelegateForColumn(BOM_COL_STATUS, _BomTreeStatusDelegate(self._ebom_filter_tree, self._ebom_filter_tree))
        self._ebom_filter_tree.setItemDelegateForColumn(BOM_COL_INTEGRITY, _BomTreeIntegrityDelegate(self._ebom_filter_tree, self._ebom_filter_tree))

        self._tree_stack.setCurrentWidget(self._ebom_tree)

        tree_layout.addWidget(self._tree_stack)
        left_layout.addWidget(tree_group)

        # Right panel
        right_widget = QWidget()
        right_widget.setObjectName("objectWorkspace")
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(2, 0, 0, 0)
        right_layout.setSpacing(5)

        # Tabs for details
        self.tabs = QTabWidget()
        self.tabs.setObjectName("objectWorkspaceTabs")
        self.tabs.setDocumentMode(True)

        # Details tab: compact engineering summary with an on-demand full attribute view.
        details_tab = QWidget()
        details_layout = QVBoxLayout(details_tab)
        details_layout.setContentsMargins(8, 8, 8, 8)
        details_layout.setSpacing(8)
        self._current_part_details = {}

        # Alert section (hidden by default)
        self.details_alert_frame = QFrame()
        self.details_alert_frame.setFrameShape(QFrame.StyledPanel)
        self.details_alert_frame.setFrameShadow(QFrame.Raised)
        self.details_alert_frame.setStyleSheet("""
            QFrame {
                background-color: #fff4f2;
                border: 1px solid #d98b82;
                border-left: 4px solid #b42318;
                border-radius: 0;
            }
            QLabel {
                color: #7a271a;
                font-weight: 600;
                padding: 4px 6px;
            }
        """)
        self.details_alert_label = QLabel("")
        details_alert_layout = QHBoxLayout(self.details_alert_frame)
        details_alert_layout.addWidget(self.details_alert_label)
        self.details_alert_frame.hide()  # Hidden by default

        details_layout.addWidget(self.details_alert_frame)

        self.details_summary_card = QFrame()
        self.details_summary_card.setObjectName("bomDetailsSummary")
        self.details_summary_card.setStyleSheet("""
            QFrame#bomDetailsSummary {
                background: #ffffff;
                border: 1px solid #aeb8c2;
                border-left: 3px solid #2d6f9f;
                border-radius: 0;
            }
            QLabel#bomDetailsTitle {
                color: #172033;
                font-size: 13px;
                font-weight: 700;
                background: transparent;
            }
            QLabel#bomDetailsIdentity {
                color: #526071;
                font-size: 10px;
                background: transparent;
            }
            QLabel#bomDetailsField {
                color: #66758a;
                font-size: 9px;
                font-weight: 700;
                background: transparent;
            }
            QLabel#bomDetailsValue {
                color: #172033;
                font-size: 10px;
                font-weight: 600;
                background: transparent;
            }
        """)
        summary_layout = QVBoxLayout(self.details_summary_card)
        summary_layout.setContentsMargins(11, 9, 11, 9)
        summary_layout.setSpacing(7)

        summary_heading = QHBoxLayout()
        heading_text = QVBoxLayout()
        heading_text.setSpacing(2)
        self.details_name_label = QLabel("Select a BOM item")
        self.details_name_label.setObjectName("bomDetailsTitle")
        self.details_identity_label = QLabel("Select a row in the BOM structure to view its summary.")
        self.details_identity_label.setObjectName("bomDetailsIdentity")
        heading_text.addWidget(self.details_name_label)
        heading_text.addWidget(self.details_identity_label)
        summary_heading.addLayout(heading_text, 1)
        self.view_full_details_btn = QPushButton("View Full Details")
        self.view_full_details_btn.setObjectName("neutral")
        self.view_full_details_btn.setEnabled(False)
        self.view_full_details_btn.clicked.connect(self._open_full_bom_details)
        self.edit_categories_btn = QPushButton("Edit Categories")
        self.edit_categories_btn.setObjectName("neutral")
        self.edit_categories_btn.setEnabled(False)
        self.edit_categories_btn.clicked.connect(self._edit_current_part_categories)
        summary_heading.addWidget(self.edit_categories_btn, 0, Qt.AlignTop)
        summary_heading.addWidget(self.view_full_details_btn, 0, Qt.AlignTop)
        summary_layout.addLayout(summary_heading)

        self.details_summary_grid = QGridLayout()
        self.details_summary_grid.setHorizontalSpacing(28)
        self.details_summary_grid.setVerticalSpacing(10)
        self._details_summary_fields = {}
        summary_fields = [
            ("part_number", "Number", ("part_number",)),
            ("aes_number", "AES Number", ("aes_number",)),
            ("item_type", "Item Type", ("item_type",)),
            ("source", "Source", ("procurement_source",)),
            ("view", "View", ("item_view",)),
            ("unit", "Default Unit", ("default_unit",)),
            ("drawing", "DRW Number", ("drawing_number",)),
            ("type", "Type", ("type",)),
            ("revision", "Revision / Iteration", ("current_version", "revision")),
            ("state", "Lifecycle", ("status", "state", "lifecycle_state")),
            ("material", "Material", ("material",)),
            ("categories", "Categories", ("categories",)),
        ]
        for index, (field_key, label_text, detail_keys) in enumerate(summary_fields):
            row = (index // 2) * 2
            column = (index % 2) * 2
            field_label = QLabel(label_text.upper())
            field_label.setObjectName("bomDetailsField")
            value_label = QLabel("-")
            value_label.setObjectName("bomDetailsValue")
            value_label.setWordWrap(True)
            self.details_summary_grid.addWidget(field_label, row, column)
            self.details_summary_grid.addWidget(value_label, row + 1, column)
            self._details_summary_fields[field_key] = (field_label, value_label, detail_keys)
        self.details_summary_grid.setColumnStretch(0, 1)
        self.details_summary_grid.setColumnStretch(2, 1)
        summary_layout.addLayout(self.details_summary_grid)
        details_layout.addWidget(self.details_summary_card)

        self.associated_files_card = QFrame()
        self.associated_files_card.setObjectName("associatedFilesCard")
        self.associated_files_card.setStyleSheet("""
            QFrame#associatedFilesCard {
                background: #ffffff;
                border: 1px solid #c4cdd6;
                border-radius: 0;
            }
            QLabel#associatedFilesTitle {
                color: #172033;
                font-size: 11px;
                font-weight: 700;
                background: transparent;
            }
            QLabel#associatedFilesLabel {
                color: #66758a;
                font-size: 9px;
                font-weight: 700;
                background: transparent;
            }
            QLabel#associatedFilesValue {
                color: #172033;
                font-size: 10px;
                font-weight: 600;
                background: transparent;
            }
        """)
        files_card_layout = QVBoxLayout(self.associated_files_card)
        files_card_layout.setContentsMargins(11, 8, 11, 8)
        files_card_layout.setSpacing(6)
        files_heading = QHBoxLayout()
        files_heading.addWidget(
            self._details_card_label("Associated CAD and Item Drawings", "associatedFilesTitle")
        )
        files_heading.addStretch()
        association_hint = QLabel("Use Edit Associations to assign model-specific drawings")
        association_hint.setStyleSheet("color:#64748b;font-size:9px;background:transparent;")
        files_heading.addWidget(association_hint)
        files_card_layout.addLayout(files_heading)

        associated_files_grid = QGridLayout()
        associated_files_grid.setHorizontalSpacing(28)
        associated_files_grid.setVerticalSpacing(4)
        associated_files_grid.addWidget(self._details_card_label("CAD DOCUMENTS", "associatedFilesLabel"), 0, 0)
        associated_files_grid.addWidget(self._details_card_label("ASSIGNED ITEM DRAWINGS", "associatedFilesLabel"), 0, 1)
        self.associated_cad_file_label = self._details_card_label("Not linked", "associatedFilesValue")
        self.associated_drawing_file_label = self._details_card_label("Not linked", "associatedFilesValue")
        self.associated_cad_file_label.setWordWrap(True)
        self.associated_drawing_file_label.setWordWrap(True)
        associated_files_grid.addWidget(self.associated_cad_file_label, 1, 0)
        associated_files_grid.addWidget(self.associated_drawing_file_label, 1, 1)
        associated_files_grid.setColumnStretch(0, 1)
        associated_files_grid.setColumnStretch(1, 1)
        files_card_layout.addLayout(associated_files_grid)
        details_layout.addWidget(self.associated_files_card)
        details_layout.addStretch(1)
        self.tabs.addTab(details_tab, "Details")

        

        # Windchill-style structure workspace: recursive Uses and direct Where Used.
        structure_tab = QWidget()
        structure_layout = QVBoxLayout(structure_tab)
        structure_layout.setContentsMargins(8, 8, 8, 8)
        structure_layout.setSpacing(6)
        self.structure_summary_label = QLabel("Select a BOM item")
        self.structure_summary_label.setStyleSheet(
            "font-size:10px;font-weight:700;color:#526071;background:transparent;"
        )
        structure_layout.addWidget(self.structure_summary_label)

        self.structure_views = QTabWidget()
        self.structure_views.setDocumentMode(True)
        self.uses_tree = self._create_structure_relation_tree()
        self.where_used_tree = self._create_structure_relation_tree()
        self.effective_where_used_tree = self._create_structure_relation_tree()
        self.uses_tree.itemDoubleClicked.connect(self._open_structure_tree_item)
        self.where_used_tree.itemDoubleClicked.connect(self._open_structure_tree_item)
        self.effective_where_used_tree.itemDoubleClicked.connect(
            self._open_structure_tree_item
        )
        for relation_tree in (
            self.uses_tree, self.where_used_tree, self.effective_where_used_tree
        ):
            relation_tree.setContextMenuPolicy(Qt.CustomContextMenu)
            relation_tree.customContextMenuRequested.connect(self._show_structure_context_menu)
        self.structure_views.addTab(self.uses_tree, "Uses")
        self.structure_views.addTab(self.where_used_tree, "CAD Where Used")
        self.structure_views.addTab(
            self.effective_where_used_tree, "Effective EBOM Where Used"
        )
        structure_layout.addWidget(self.structure_views, 1)
        structure_actions = QHBoxLayout()
        self.compare_iterations_btn = QPushButton("Compare Iterations")
        self.compare_iterations_btn.setObjectName("neutral")
        self.compare_iterations_btn.clicked.connect(self.compare_assembly_iterations)
        self.compare_iterations_btn.setEnabled(False)
        self.update_child_versions_btn = QPushButton("Update Child Versions")
        self.update_child_versions_btn.setObjectName("neutral")
        self.update_child_versions_btn.clicked.connect(self.update_child_versions)
        self.update_child_versions_btn.setEnabled(False)
        structure_actions.addStretch()
        structure_actions.addWidget(self.compare_iterations_btn)
        structure_actions.addWidget(self.update_child_versions_btn)
        structure_layout.addLayout(structure_actions)
        self.tabs.addTab(structure_tab, "Structure")

        # Notes tab
        notes_tab = QWidget()
        notes_layout = QVBoxLayout(notes_tab)
        self.notes_view = QTextEdit()
        self.notes_view.setReadOnly(True)
        notes_layout.addWidget(QLabel("Notes:"))
        notes_layout.addWidget(self.notes_view)
        self.tabs.addTab(notes_tab, "Notes")

        # Engineering issues tab
        issues_tab = QWidget()
        issues_layout = QVBoxLayout(issues_tab)
        self.part_issues_table = QTableWidget()
        self.part_issues_table.setColumnCount(4)
        self.part_issues_table.setHorizontalHeaderLabels(["Issue", "Title", "Status", "Priority"])
        self.part_issues_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.part_issues_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.part_issues_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.part_issues_table.itemDoubleClicked.connect(lambda *_: self._open_current_part_issues())
        issue_actions = QHBoxLayout()
        open_issues_btn = QPushButton("Open Issue List")
        open_issues_btn.clicked.connect(self._open_current_part_issues)
        create_issue_btn = QPushButton("Create Issue")
        create_issue_btn.setObjectName("primary")
        create_issue_btn.clicked.connect(self._create_issue_for_current_part)
        issue_actions.addWidget(open_issues_btn)
        issue_actions.addWidget(create_issue_btn)
        issue_actions.addStretch()
        issues_layout.addWidget(self.part_issues_table)
        issues_layout.addLayout(issue_actions)
        self.tabs.addTab(issues_tab, "Issues")

        # Managed files: Creo content, generated outputs, and document history.
        files_tab = QWidget()
        self.files_tab = files_tab
        files_layout = QVBoxLayout(files_tab)

        files_summary = QFrame()
        files_summary.setObjectName("detailCard")
        summary_layout = QGridLayout(files_summary)
        self.files_summary_labels = {}
        for column, (key, title) in enumerate((
            ("item", "BOM Object"),
            ("version", "Rev / Iter"),
            ("state", "Lifecycle"),
            ("lock", "Checkout"),
        )):
            title_label = QLabel(title)
            title_label.setObjectName("mutedLabel")
            value_label = QLabel("-")
            value_label.setStyleSheet("font-weight: 600;")
            summary_layout.addWidget(title_label, 0, column)
            summary_layout.addWidget(value_label, 1, column)
            self.files_summary_labels[key] = value_label
        files_layout.addWidget(files_summary)

        files_layout.addWidget(QLabel("Current Managed Content"))
        self.files_table = FileDropTable()
        self.files_table.setColumnCount(10)
        self.files_table.setHorizontalHeaderLabels([
            "Role", "File", "File Revision", "Creo Ver", "Rev / Iter", "Source", "State", "Health", "Created By", "Updated"
        ])
        self.files_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.files_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.files_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.files_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.files_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.files_table.itemSelectionChanged.connect(self.on_attachment_selected)
        self.files_table.filesDropped.connect(self._on_files_dropped)
        self.files_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.files_table.customContextMenuRequested.connect(self.show_files_context_menu)
        files_layout.addWidget(self.files_table)

        files_actions = QHBoxLayout()
        self.add_attachment_btn = QPushButton("Attach Document")
        self.add_attachment_btn.setObjectName("primary")
        self.add_attachment_btn.clicked.connect(self.add_attachment)
        self.add_version_btn = QPushButton("Add Document Version")
        self.add_version_btn.setObjectName("neutral")
        self.add_version_btn.clicked.connect(self.add_attachment_version)
        self.set_active_btn = QPushButton("Use in Working Iteration")
        self.set_active_btn.setObjectName("neutral")
        self.set_active_btn.clicked.connect(self.set_active_version)
        self.open_active_btn = QPushButton("Open")
        self.open_active_btn.setObjectName("neutral")
        self.open_active_btn.clicked.connect(self.open_active_attachment)
        self.preview_file_btn = QPushButton("Preview")
        self.preview_file_btn.setObjectName("neutral")
        self.preview_file_btn.setCheckable(True)
        self.preview_file_btn.setToolTip("Show or collapse the selected PDF preview")
        self.preview_file_btn.toggled.connect(self._toggle_managed_preview)
        self.preview_file_btn.setEnabled(False)
        self.open_folder_btn = QPushButton("Open Folder")
        self.open_folder_btn.setObjectName("neutral")
        self.open_folder_btn.clicked.connect(self.open_active_attachment_folder)
        self.remove_attachment_btn = QPushButton("Obsolete Document")
        self.remove_attachment_btn.setObjectName("danger")
        self.remove_attachment_btn.clicked.connect(self.remove_attachment)

        self.export_package_btn = QPushButton("Export Package")
        self.export_package_btn.setObjectName("primary")
        self.export_package_btn.clicked.connect(self.export_package)

        self.create_baseline_btn = QPushButton("Create Baseline")
        self.create_baseline_btn.setObjectName("neutral")
        self.create_baseline_btn.clicked.connect(self.create_baseline)

        self.export_baseline_btn = QPushButton("Export Baseline")
        self.export_baseline_btn.setObjectName("neutral")
        self.export_baseline_btn.clicked.connect(self.export_baseline)
        self.link_file_issue_btn = QPushButton("Link to Issue")
        self.link_file_issue_btn.setObjectName("neutral")
        self.link_file_issue_btn.clicked.connect(self.link_selected_file_to_issue)
        files_actions.addWidget(self.add_attachment_btn)
        files_actions.addWidget(self.add_version_btn)
        files_actions.addWidget(self.set_active_btn)
        files_actions.addWidget(self.open_active_btn)
        files_actions.addWidget(self.preview_file_btn)
        files_actions.addWidget(self.open_folder_btn)
        files_actions.addWidget(self.remove_attachment_btn)
        files_actions.addWidget(self.link_file_issue_btn)
        files_actions.addWidget(self.export_package_btn)
        files_actions.addWidget(self.create_baseline_btn)
        files_actions.addWidget(self.export_baseline_btn)
        files_layout.addLayout(files_actions)

        files_layout.addWidget(QLabel("Content History"))
        self.versions_table = QTableWidget()
        self.versions_table.setColumnCount(11)
        self.versions_table.setHorizontalHeaderLabels([
            "Rev / Iter", "Role", "File", "File Revision", "Creo Ver", "Source", "State", "Health", "Created By", "Created", "Note"
        ])
        self.versions_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.versions_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.versions_table.horizontalHeader().setSectionResizeMode(10, QHeaderView.Stretch)
        self.versions_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.versions_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.versions_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.versions_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.versions_table.customContextMenuRequested.connect(self.show_versions_context_menu)
        self.versions_table.itemSelectionChanged.connect(self._on_version_selection_changed)
        files_layout.addWidget(self.versions_table)

        files_layout.addWidget(QLabel("Related Issues:"))
        self.file_related_issues_table = QTableWidget()
        self.file_related_issues_table.setColumnCount(5)
        self.file_related_issues_table.setHorizontalHeaderLabels(["Issue", "Title", "Status", "Role", "Linked"])
        self.file_related_issues_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.file_related_issues_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        files_layout.addWidget(self.file_related_issues_table)

        versions_actions = QHBoxLayout()
        self.open_version_btn = QPushButton("Open Historical Content")
        self.open_version_btn.setObjectName("neutral")
        self.open_version_btn.clicked.connect(self.open_selected_version)

        _can_release = self.perm.can("release_files")
        self.add_attachment_btn.setEnabled(_can_release)
        self.add_version_btn.setEnabled(_can_release)
        self.set_active_btn.setEnabled(_can_release)
        self.remove_attachment_btn.setEnabled(_can_release)
        self.create_baseline_btn.setEnabled(_can_release)

        self.release_version_btn = QPushButton("Approve Document Version")
        self.release_version_btn.setObjectName("primary")
        self.release_version_btn.clicked.connect(self.release_selected_version)
        self.release_version_btn.setEnabled(_can_release)

        self.delete_version_btn = QPushButton("Obsolete Version")
        self.delete_version_btn.setObjectName("danger")
        self.delete_version_btn.clicked.connect(self.delete_selected_version)
        self.delete_version_btn.setEnabled(_can_release)
        versions_actions.addWidget(self.open_version_btn)
        versions_actions.addWidget(self.release_version_btn)
        versions_actions.addWidget(self.delete_version_btn)
        files_layout.addLayout(versions_actions)

        # ── Embedded PDF preview ──────────────────────────────────────
        self.pdf_viewer = PdfViewerWidget()
        self.pdf_viewer.setMinimumHeight(250)
        self.pdf_viewer.setVisible(False)
        files_layout.addWidget(self.pdf_viewer)

        self.tabs.addTab(files_tab, "Managed Files")

        # ── History tab (genius panel) ────────────────────────────────
        self.history_panel = HistoryPanel(self)
        self.history_panel.open_details_requested.connect(self._open_history_details_dialog)
        self.history_panel.open_step_requested.connect(self._open_associated_step_for_history_event)
        self.history_panel.show_diff_requested.connect(self._show_step_diff_for_history_event)
        self.history_panel._export_btn.clicked.connect(self.export_history_csv)
        self.tabs.addTab(self.history_panel, "History")

        right_layout.addWidget(self.tabs)

        # Contextual object ribbon.  Its groups follow the selected Item or CAD
        # Document and remain part of the workspace chrome, above both panes.
        action_ribbon = QFrame()
        action_ribbon.setObjectName("actionRibbon")
        action_ribbon.setFixedHeight(58)
        action_layout = QHBoxLayout(action_ribbon)
        action_layout.setContentsMargins(4, 2, 4, 2)
        action_layout.setSpacing(1)
        self.add_part_btn = QPushButton("New Item")
        self.add_part_btn.setObjectName("primary")
        self.add_part_btn.clicked.connect(self.add_part)
        self.add_folder_btn = QPushButton("Add Folder")
        self.add_folder_btn.setObjectName("neutral")
        self.add_folder_btn.clicked.connect(self.add_bom_folder)
        self.edit_part_btn = QPushButton("Edit Attributes")
        self.edit_part_btn.setObjectName("neutral")
        self.edit_part_btn.clicked.connect(self.edit_part)
        self.delete_part_btn = QPushButton("Delete Item")
        self.delete_part_btn.setObjectName("danger")
        self.delete_part_btn.clicked.connect(self.delete_part)
        self.undo_checkout_btn = QPushButton("Undo Checkout")
        self.undo_checkout_btn.setObjectName("neutral")
        self.undo_checkout_btn.clicked.connect(self.undo_checkout)
        self.checkout_part_btn = QPushButton("Check Out")
        self.checkout_part_btn.setObjectName("neutral")
        self.checkout_part_btn.clicked.connect(self.checkout_part)
        self.checkin_part_btn = QPushButton("Check In")
        self.checkin_part_btn.setObjectName("neutral")
        self.checkin_part_btn.clicked.connect(self.checkin_part)
        self.add_child_btn = QPushButton("Add Child")
        self.add_child_btn.setObjectName("primary")
        self.add_child_btn.clicked.connect(self.add_child)
        self.compare_structure_btn = QPushButton("Compare Structure")
        self.compare_structure_btn.setObjectName("neutral")
        self.compare_structure_btn.clicked.connect(lambda: self.compare_part_structure(whole_bom=True))
        self.create_configuration_btn = QPushButton("Create Configuration")
        self.create_configuration_btn.setObjectName("primary")
        self.create_configuration_btn.clicked.connect(self.create_assembly_configuration)
        self.create_configuration_btn.setEnabled(False)
        self.manage_configurations_btn = QPushButton("Manage Configurations")
        self.manage_configurations_btn.setObjectName("neutral")
        self.manage_configurations_btn.clicked.connect(self.manage_assembly_configurations)
        self.manage_configurations_btn.setEnabled(bool(self.session.project_id))
        self.register_cad_btn = QPushButton("Register CAD")
        self.register_cad_btn.setObjectName("neutral")
        self.register_cad_btn.clicked.connect(self.register_cad_document)
        self.register_cad_btn.setEnabled(False)
        self.add_cad_component_btn = QPushButton("Add Component")
        self.add_cad_component_btn.setObjectName("neutral")
        self.add_cad_component_btn.clicked.connect(self._add_cad_component_from_ribbon)
        self.add_cad_component_btn.setEnabled(False)
        self.delete_cad_btn = QPushButton("Delete CAD")
        self.delete_cad_btn.setObjectName("danger")
        self.delete_cad_btn.clicked.connect(
            lambda: self._delete_selected_pdm_cad_document()
        )
        self.delete_cad_btn.setEnabled(False)
        self.show_associated_cad_btn = QPushButton("Show CAD")
        self.show_associated_cad_btn.setObjectName("neutral")
        self.show_associated_cad_btn.clicked.connect(self._toggle_selected_ebom_cad_associations)
        self.show_associated_cad_btn.setEnabled(False)
        self.clear_pdm_scope_btn = QPushButton("Top Level")
        self.clear_pdm_scope_btn.setObjectName("neutral")
        self.clear_pdm_scope_btn.clicked.connect(self._clear_pdm_isolation)
        self.clear_pdm_scope_btn.setEnabled(False)

        _can_manage = self.perm.can("manage_parts")
        self.add_part_btn.setEnabled(_can_manage)
        self.add_folder_btn.setEnabled(_can_manage)
        self.edit_part_btn.setEnabled(_can_manage)
        self.delete_part_btn.setEnabled(_can_manage)
        self.add_child_btn.setEnabled(_can_manage)

        self.set_revision_btn = QPushButton("Create New Revision")
        self.set_revision_btn.setObjectName("neutral")
        self.set_revision_btn.clicked.connect(self.set_part_revision)
        self.set_revision_btn.setEnabled(self.perm.can("set_revision"))
        self.release_revision_btn = QPushButton("Release Revision")
        self.release_revision_btn.setObjectName("neutral")
        self.release_revision_btn.clicked.connect(self.release_part_revision)
        self.release_revision_btn.setEnabled(self.perm.can("set_revision"))

        action_buttons = (
            self.add_part_btn, self.add_child_btn, self.add_folder_btn,
            self.edit_part_btn, self.compare_structure_btn, self.delete_part_btn,
            self.checkout_part_btn, self.checkin_part_btn, self.undo_checkout_btn,
            self.set_revision_btn, self.release_revision_btn,
            self.create_configuration_btn, self.manage_configurations_btn,
            self.register_cad_btn, self.add_cad_component_btn, self.delete_cad_btn,
            self.show_associated_cad_btn, self.clear_pdm_scope_btn,
        )
        for button in action_buttons:
            button.setFixedHeight(30)
            button.setIconSize(QSize(14, 14))
            button.setProperty("ribbonAction", True)
            button.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)

        style = self.style()
        action_specs = (
            (self.add_part_btn, "New", QStyle.SP_FileIcon, "Create an Item master"),
            (self.edit_part_btn, "Edit", QStyle.SP_FileDialogDetailedView, "Edit the selected Item attributes"),
            (self.delete_part_btn, "Delete", QStyle.SP_TrashIcon, "Delete the selected item"),
            (self.checkout_part_btn, "Check Out", QStyle.SP_ArrowForward, "Check out the selected item"),
            (self.checkin_part_btn, "Check In", QStyle.SP_DialogApplyButton, "Review and finish the selected checkout"),
            (self.undo_checkout_btn, "Undo", QStyle.SP_ArrowBack, "Undo checkout"),
            (self.add_child_btn, "Add Child", QStyle.SP_ArrowDown, "Add a child to the selected assembly"),
            (self.add_folder_btn, "Folder", QStyle.SP_DirIcon, "Add an organization folder"),
            (self.compare_structure_btn, "Compare", QStyle.SP_FileDialogContentsView, "Compare BOM structures"),
            (self.set_revision_btn, "Create New Revision", QStyle.SP_FileDialogNewFolder, "Create a new revision"),
            (self.release_revision_btn, "Release Revision", QStyle.SP_DialogApplyButton, "Release the current revision"),
            (self.create_configuration_btn, "Create Configuration", QStyle.SP_FileDialogNewFolder, "Create an assembly configuration"),
            (self.manage_configurations_btn, "Manage Configurations", QStyle.SP_ComputerIcon, "Manage assembly configurations"),
            (self.register_cad_btn, "Register", QStyle.SP_FileIcon, "Register a managed PRT or ASM CAD Document"),
            (self.add_cad_component_btn, "Add Component", QStyle.SP_ArrowDown, "Add a CAD component to the selected checked-out ASM"),
            (self.delete_cad_btn, "Delete", QStyle.SP_TrashIcon, "Delete the selected CAD Document"),
            (self.show_associated_cad_btn, "Show CAD", QStyle.SP_FileDialogDetailedView, "Show or hide associated CAD under the selected Item"),
            (self.clear_pdm_scope_btn, "Top Level", QStyle.SP_ArrowUp, "Return the active BOM browser to the top level"),
        )
        for button, text_value, icon_type, tooltip in action_specs:
            button.setText(text_value)
            button.setIcon(style.standardIcon(icon_type))
            button.setToolTip(tooltip)

        self._action_ribbon_menu_bindings = []

        def make_menu_button(text_value, tooltip, icon_type, entries):
            button = QPushButton(text_value)
            button.setProperty("ribbonAction", True)
            button.setFixedHeight(30)
            button.setIconSize(QSize(14, 14))
            button.setIcon(style.standardIcon(icon_type))
            button.setToolTip(tooltip)
            button.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
            menu = QMenu(button)
            menu.setStyleSheet("QMenu { font-size: 10px; } QMenu::item { padding: 4px 18px 4px 22px; }")
            bindings = []
            for label, target in entries:
                action = QAction(target.icon(), label, menu)
                action.triggered.connect(lambda _checked=False, target=target: target.click())
                menu.addAction(action)
                bindings.append((action, target))
            self._action_ribbon_menu_bindings.append((button, bindings))
            menu.aboutToShow.connect(self._sync_action_ribbon_menus)
            button.setMenu(menu)
            return button

        self.revision_actions_btn = make_menu_button(
            "Revision",
            "Revision actions",
            QStyle.SP_FileDialogNewFolder,
            (
                ("Create New Revision", self.set_revision_btn),
                ("Release Revision", self.release_revision_btn),
            ),
        )
        self.configuration_actions_btn = make_menu_button(
            "Config",
            "Configuration actions",
            QStyle.SP_ComputerIcon,
            (
                ("Create Configuration", self.create_configuration_btn),
                ("Manage Configurations", self.manage_configurations_btn),
            ),
        )

        def add_action_category(title, buttons):
            category = QFrame()
            category.setObjectName("actionCategory")
            category_layout = QVBoxLayout(category)
            category_layout.setContentsMargins(4, 0, 4, 0)
            category_layout.setSpacing(0)
            title_label = QLabel(title)
            title_label.setObjectName("actionCategoryTitle")
            title_label.setAlignment(Qt.AlignCenter)
            title_label.setFixedHeight(15)
            command_layout = QHBoxLayout()
            command_layout.setContentsMargins(0, 0, 0, 0)
            command_layout.setSpacing(1)
            for button in buttons:
                command_layout.addWidget(button)
            category_layout.addLayout(command_layout)
            category_layout.addWidget(title_label)
            action_layout.addWidget(category)

        add_action_category("Editing", (self.add_part_btn, self.edit_part_btn, self.delete_part_btn))
        add_action_category(
            "Check Out/In",
            (self.checkout_part_btn, self.checkin_part_btn, self.undo_checkout_btn),
        )
        add_action_category("New/Add To", (self.add_child_btn, self.add_folder_btn))
        add_action_category("Lifecycle", (self.revision_actions_btn,))
        add_action_category("CAD", (self.register_cad_btn, self.add_cad_component_btn, self.delete_cad_btn))
        add_action_category("Tools", (self.compare_structure_btn, self.configuration_actions_btn))
        add_action_category("Visual", (self.show_associated_cad_btn, self.clear_pdm_scope_btn))
        action_layout.addStretch(1)
        action_ribbon.setStyleSheet(
            "QFrame#actionRibbon { background: #eef1f4; border: 1px solid #aeb9c5; border-radius: 0; }"
            "QFrame#actionCategory { background: transparent; border: 0; border-right: 1px solid #bcc6d0; }"
            "QLabel#actionCategoryTitle { color: #4d6276; font-size: 9px; font-weight: 600; border: 0; }"
            "QPushButton[ribbonAction=\"true\"] { background: transparent; color: #1f3447; border: 1px solid transparent; "
            "border-radius: 0; font-size: 10px; padding: 2px 5px; }"
            "QPushButton[ribbonAction=\"true\"]:hover { background: #dce9f3; border-color: #8eafc8; }"
            "QPushButton[ribbonAction=\"true\"]:pressed { background: #cbddea; border-color: #6f98b6; }"
            "QPushButton[ribbonAction=\"true\"]:disabled { color: #98a3ad; background: transparent; border-color: transparent; }"
        )
        self._sync_action_ribbon_menus()
        self._sync_visual_action_states()
        layout.insertWidget(0, action_ribbon)

        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget)

        # Set minimum widths
        left_widget.setMinimumWidth(540)
        right_widget.setMinimumWidth(460)

        splitter.setStretchFactor(0, 6)
        splitter.setStretchFactor(1, 5)
        splitter.setSizes([720, 580])

        self.setStyleSheet("""
            QWidget#pdmWorkspace {
                background: #dfe4e9;
            }
            QWidget#structureWorkspace,
            QWidget#objectWorkspace {
                background: #f4f5f6;
            }
            QSplitter#pdmWorkspaceSplitter::handle {
                background: #aeb9c5;
            }
            QSplitter#pdmWorkspaceSplitter::handle:hover {
                background: #7695ad;
            }
            QGroupBox#structureSearch {
                background: #eef1f4;
                border: 1px solid #aeb9c5;
                border-radius: 0;
                margin-top: 7px;
                color: #3f5368;
                font-size: 9px;
                font-weight: 700;
            }
            QGroupBox#structureSearch::title {
                subcontrol-origin: margin;
                left: 7px;
                padding: 0 3px;
            }
            QGroupBox#structureSearch QLineEdit {
                min-height: 23px;
                padding: 0 6px;
                background: #ffffff;
                border: 1px solid #9daab7;
                border-radius: 0;
            }
            QGroupBox#structureSearch QLineEdit:focus {
                border: 1px solid #2f75a4;
            }
            QGroupBox#structureSearch QPushButton#primary {
                min-height: 23px;
                background: #2f75a4;
                color: #ffffff;
                border: 1px solid #245f86;
                border-radius: 0;
                padding: 0 12px;
            }
            QGroupBox#structureSearch QPushButton#primary:hover {
                background: #3e86b5;
            }
            QFrame#structureBrowser {
                background: #eef1f4;
                border: 1px solid #aeb9c5;
            }
            QLabel#structureViewCaption {
                color: #607286;
                font-size: 9px;
                font-weight: 700;
                padding-left: 8px;
            }
            QComboBox#structureViewSelector {
                background: #ffffff;
                border: 1px solid #9daab7;
                border-radius: 0;
                padding: 1px 6px;
                color: #1f3447;
                font-weight: 600;
            }
            QPushButton[structureTool="true"] {
                min-height: 22px;
                background: #eef1f4;
                border: 1px solid #aeb9c5;
                border-radius: 0;
                color: #253b4f;
                padding: 1px 8px;
            }
            QPushButton[structureTool="true"]:hover {
                background: #dce9f3;
                border-color: #7fa4bf;
            }
            QPushButton[structureTool="true"]:disabled {
                color: #99a3ad;
                background: #eceff2;
                border-color: #ccd3da;
            }
            QTabWidget#objectWorkspaceTabs::pane {
                border: 1px solid #aeb9c5;
                background: #ffffff;
                top: -1px;
            }
            QTabWidget#objectWorkspaceTabs QTabBar::tab {
                background: #dfe4e9;
                border: 1px solid #aeb9c5;
                border-radius: 0;
                padding: 6px 12px;
                color: #344a5f;
                font-weight: 600;
            }
            QTabWidget#objectWorkspaceTabs QTabBar::tab:selected {
                background: #ffffff;
                color: #173f5e;
                border-top: 2px solid #2f75a4;
                border-bottom-color: #ffffff;
            }
            QTabWidget#objectWorkspaceTabs QTabBar::tab:hover:!selected {
                background: #eaf0f4;
            }
        """)

    def _sync_action_ribbon_menus(self) -> None:
        for menu_button, bindings in getattr(self, "_action_ribbon_menu_bindings", []):
            has_enabled_action = False
            for action, target_button in bindings:
                enabled = target_button.isEnabled()
                action.setEnabled(enabled)
                has_enabled_action = has_enabled_action or enabled
            menu_button.setEnabled(has_enabled_action)

    def _pdm_mode_for_tree(self, tree: QTreeWidget | None) -> str | None:
        if tree is None:
            return None
        if tree is getattr(self, "_cad_tree", None):
            return "cad"
        if tree in (
            getattr(self, "_ebom_tree", None),
            getattr(self, "_ebom_filter_tree", None),
        ):
            return "ebom"
        return None

    def _current_pdm_tree(self) -> QTreeWidget | None:
        if str(getattr(self, "_bom_mode", "cad")) == "ebom":
            filtered = getattr(self, "_ebom_filter_tree", None)
            if (
                getattr(self, "_ebom_filter_flat_mode", False)
                and filtered is not None
            ):
                return filtered
            return getattr(self, "_ebom_tree", None)
        return getattr(self, "_cad_tree", None)

    def _pdm_name_column_for_tree(self, tree: QTreeWidget | None) -> int:
        return CAD_COL_NAME if tree is getattr(self, "_cad_tree", None) else BOM_COL_NAME

    def _ensure_pdm_lazy_placeholder(self, item: QTreeWidgetItem) -> None:
        if item is None or not item.data(0, BOM_TREE_IS_ASSEMBLY_ROLE):
            return
        tree = item.treeWidget()
        if item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE):
            return
        for index in range(item.childCount()):
            if self._is_lazy_placeholder(item.child(index)):
                return
        column_count = tree.columnCount() if tree is not None else BOM_TREE_COLUMN_COUNT
        placeholder = QTreeWidgetItem([""] * column_count)
        placeholder.setText(self._pdm_name_column_for_tree(tree), "Loading...")
        placeholder.setData(0, BOM_TREE_PLACEHOLDER_ROLE, True)
        placeholder.setDisabled(True)
        item.addChild(placeholder)

    def _remove_pdm_lazy_placeholders(self, item: QTreeWidgetItem) -> None:
        for index in range(item.childCount() - 1, -1, -1):
            if self._is_lazy_placeholder(item.child(index)):
                item.takeChild(index)

    def _render_pdm_cad_roots(self, roots: list[dict]) -> None:
        tree = getattr(self, "_cad_tree", None)
        if tree is None:
            return
        tree.setUpdatesEnabled(False)
        try:
            tree.resetLoadingIndicators()
            tree.clear()
            for root in list(roots or []):
                self._add_pdm_cad_node(root)
            if not getattr(self, "_pdm_cad_scope_path", []):
                self._render_pdm_folder_context("CAD", None, [tree])
            tree.collapseAll()
            # Filtering may expand ancestors of matching descendants, so it
            # must run after the default collapsed state is established.
            self._refresh_pdm_cad_filter()
        finally:
            tree.setUpdatesEnabled(True)
        self._update_pdm_scope_bar()
        self._sync_visual_action_states()

    def _render_pdm_ebom_roots(self, roots: list[dict]) -> None:
        tree = getattr(self, "_ebom_tree", None)
        if tree is None:
            return
        tree.setUpdatesEnabled(False)
        try:
            tree.resetLoadingIndicators()
            tree.clear()
            for root in list(roots or []):
                if self._is_deleted_bom_payload(root):
                    continue
                self._add_released_ebom_node(
                    root,
                    associations_by_item=getattr(self, "_ebom_associations_by_item", {}),
                )
            if not getattr(self, "_pdm_ebom_scope_path", []):
                self._render_pdm_folder_context("EBOM", None, [tree])
            self._renumber_tree_rows(tree)
            tree.collapseAll()
            # Preserve the expansions created by search/advanced filtering.
            self._refresh_ebom_filters()
        finally:
            tree.setUpdatesEnabled(True)
        self._update_pdm_scope_bar()
        self._sync_visual_action_states()

    def _on_pdm_tree_item_expanded(self, item: QTreeWidgetItem) -> None:
        if (
            item is None
            or self._is_lazy_placeholder(item)
            or item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE)
            or item.data(0, BOM_TREE_LOADING_ROLE)
        ):
            return
        tree = item.treeWidget()
        if tree is getattr(self, "_ebom_filter_tree", None):
            # The filter tree is a lightweight SQL projection, not the live lazy
            # EBOM tree. Its nodes intentionally do not carry complete child
            # payloads, so expanding them must never start lazy loading/spinners.
            item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, True)
            item.setData(0, BOM_TREE_LOADING_ROLE, False)
            try:
                tree.setItemLoading(item, False)
            except Exception:
                pass
            return
        if not isinstance(tree, BomTreeWidget):
            self._load_pdm_lazy_children_for_item(item, refresh_filters=True)
            return

        # Paint the branch spinner where the expand triangle normally appears,
        # then populate the cached direct children on the next event-loop turn.
        # Keeping the branch closed avoids flashing the temporary Loading row.
        previous_signal_state = tree.blockSignals(True)
        try:
            item.setExpanded(False)
        finally:
            tree.blockSignals(previous_signal_state)
        tree.setItemLoading(item, True)
        QTimer.singleShot(
            0,
            lambda tree=tree, item=item: self._finish_pdm_tree_item_expansion(
                tree, item
            ),
        )

    def _finish_pdm_tree_item_expansion(
        self, tree: BomTreeWidget, item: QTreeWidgetItem
    ) -> None:
        try:
            if item.treeWidget() is not tree:
                return
            self._load_pdm_lazy_children_for_item(item, refresh_filters=True)
        finally:
            try:
                tree.setItemLoading(item, False)
            except Exception:
                pass
        try:
            if item.treeWidget() is tree and item.childCount():
                item.setExpanded(True)
        except Exception:
            pass

    def _load_pdm_lazy_children_for_item(
        self, item: QTreeWidgetItem, refresh_filters: bool = False
    ) -> None:
        tree = item.treeWidget()
        mode = self._pdm_mode_for_tree(tree)
        if mode not in {"cad", "ebom"}:
            return
        children = list(item.data(0, PDM_CHILDREN_PAYLOAD_ROLE) or [])
        tree.setUpdatesEnabled(False)
        try:
            self._remove_pdm_lazy_placeholders(item)
            item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, True)
            if mode == "cad":
                for child in children:
                    self._add_pdm_cad_node(child, item)
                parent_id = item.data(0, PDM_CAD_DOCUMENT_ID_ROLE)
                if parent_id is not None:
                    self._render_pdm_folder_context(
                        "CAD", int(parent_id), [item]
                    )
            else:
                for child in children:
                    if self._is_deleted_bom_payload(child):
                        continue
                    self._add_released_ebom_node(
                        child,
                        item,
                        getattr(self, "_ebom_associations_by_item", {}),
                    )
                parent_id = item.data(0, Qt.UserRole)
                if parent_id is not None:
                    self._render_pdm_folder_context(
                        "EBOM", int(parent_id), [item]
                    )
                self._renumber_tree_rows(tree)
        finally:
            tree.setUpdatesEnabled(True)
            tree.viewport().update()
        if refresh_filters:
            if mode == "cad":
                self._refresh_pdm_cad_filter()
            else:
                self._refresh_ebom_filters()
        self._sync_visual_action_states()

    def _materialize_pdm_tree_for_search(self, tree: QTreeWidget | None) -> None:
        mode = self._pdm_mode_for_tree(tree)
        if tree is None or mode not in {"cad", "ebom"}:
            return

        def visit(item: QTreeWidgetItem) -> None:
            if self._is_lazy_placeholder(item):
                return
            if not item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE):
                self._load_pdm_lazy_children_for_item(item, refresh_filters=False)
            for index in range(item.childCount()):
                visit(item.child(index))

        tree.setUpdatesEnabled(False)
        try:
            for index in range(tree.topLevelItemCount()):
                visit(tree.topLevelItem(index))
        finally:
            tree.setUpdatesEnabled(True)
            tree.viewport().update()

    def _remove_direct_ebom_cad_rows(self, item: QTreeWidgetItem) -> None:
        for index in range(item.childCount() - 1, -1, -1):
            child = item.child(index)
            if child.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                item.takeChild(index)

    def _show_ebom_cad_associations(
        self, item: QTreeWidgetItem, refresh_filters: bool = True
    ) -> None:
        if item is None or item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
            return
        associations = list(item.data(0, PDM_EBOM_ASSOCIATIONS_ROLE) or [])
        try:
            item_id = int(item.data(0, Qt.UserRole))
        except (TypeError, ValueError):
            return
        try:
            item_documents = [
                document
                for document in (
                    self.bom_service.list_item_cad_associations(item_id) or []
                )
                if str(document.get("category") or "").upper() != "DRAWING"
            ]
            if item_documents:
                associations = item_documents
                item.setData(
                    0, PDM_EBOM_ASSOCIATIONS_ROLE, list(item_documents)
                )
        except Exception:
            # The global association summary still permits Show CAD on older data.
            pass
        if not associations:
            return
        tree = item.treeWidget()
        if tree is not None:
            tree.setUpdatesEnabled(False)
        try:
            self._remove_direct_ebom_cad_rows(item)
            for document in reversed(associations):
                self._add_ebom_associated_cad_node(
                    document, item, item_id, insert_index=0
                )
            item.setData(0, PDM_ASSOCIATIONS_SHOWN_ROLE, True)
        finally:
            if tree is not None:
                tree.setUpdatesEnabled(True)
                tree.viewport().update()
        try:
            item.setExpanded(True)
        except Exception:
            pass
        if refresh_filters:
            self._renumber_tree_rows(getattr(self, "_ebom_tree", None))
            self._refresh_ebom_filters()
            self._sync_visual_action_states()

    def _hide_ebom_cad_associations(
        self, item: QTreeWidgetItem, refresh_filters: bool = True
    ) -> None:
        if item is None:
            return
        tree = item.treeWidget()
        if tree is not None:
            tree.setUpdatesEnabled(False)
        try:
            self._remove_direct_ebom_cad_rows(item)
            item.setData(0, PDM_ASSOCIATIONS_SHOWN_ROLE, False)
        finally:
            if tree is not None:
                tree.setUpdatesEnabled(True)
                tree.viewport().update()
        if refresh_filters:
            self._renumber_tree_rows(getattr(self, "_ebom_tree", None))
            self._refresh_ebom_filters()
            self._sync_visual_action_states()

    def _selected_ebom_item_for_visual_action(self) -> QTreeWidgetItem | None:
        if str(getattr(self, "_bom_mode", "cad")) != "ebom":
            return None
        tree = getattr(self, "_ebom_tree", None)
        item = tree.currentItem() if tree is not None else None
        if item is not None and item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
            item = item.parent()
        if (
            item is None
            or self._is_lazy_placeholder(item)
            or item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD
        ):
            return None
        return item

    def _selected_cad_item_for_ribbon(self) -> QTreeWidgetItem | None:
        mode = str(getattr(self, "_bom_mode", "cad"))
        tree = (
            getattr(self, "_cad_tree", None)
            if mode == "cad" else
            getattr(self, "_ebom_tree", None)
        )
        item = tree.currentItem() if tree is not None else None
        if (
            item is None
            or self._is_lazy_placeholder(item)
            or item.data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD
        ):
            return None
        return item

    def _add_cad_component_from_ribbon(self) -> None:
        item = self._selected_cad_item_for_ribbon()
        if item is None:
            QMessageBox.information(self, "CAD Structure", "Select an ASM CAD Document.")
            return
        payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
        if str(payload.get("category") or item.text(CAD_COL_CATEGORY)).upper() != "ASSEMBLY":
            QMessageBox.information(self, "CAD Structure", "Only an ASM CAD Document can contain components.")
            return
        try:
            checked_out_by_me = (
                self.session.user_id is not None
                and payload.get("checked_out_by") is not None
                and int(payload["checked_out_by"]) == int(self.session.user_id)
            )
        except Exception:
            checked_out_by_me = False
        if not (checked_out_by_me or self.perm.can("merge")):
            QMessageBox.information(
                self,
                "CAD Structure",
                "Check out the ASM CAD Document before adding a component.",
            )
            return
        cad_id = item.data(0, PDM_CAD_DOCUMENT_ID_ROLE)
        if cad_id is not None:
            self._add_cad_member_from_tree(int(cad_id))

    def _delete_selected_pdm_cad_document(
        self,
        cad_id: int | None = None,
        payload: dict | None = None,
        *,
        reselect_cad_id: int | None = None,
    ) -> None:
        if isinstance(cad_id, bool):
            cad_id = None
        if not self.perm.can("manage_parts"):
            QMessageBox.warning(self, "Permission denied", "You cannot delete CAD Documents.")
            return
        item = None
        if cad_id is None:
            item = self._selected_cad_item_for_ribbon()
            if item is None:
                QMessageBox.information(self, "CAD Structure", "Select a CAD Document to delete.")
                return
            cad_id = int(item.data(0, PDM_CAD_DOCUMENT_ID_ROLE))
            payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
        else:
            cad_id = int(cad_id)
            payload = dict(payload or {})
        label = str(
            payload.get("file_name")
            or payload.get("name")
            or f"CAD Document {cad_id}"
        )
        related_drawings = list(payload.get("related_drawings") or [])
        delete_related_drawings = False
        if related_drawings:
            drawing_names = "\n".join(
                f"- {drawing.get('file_name') or drawing.get('name') or drawing.get('id')}"
                for drawing in related_drawings[:12]
            )
            if len(related_drawings) > 12:
                drawing_names += f"\n- ... and {len(related_drawings) - 12} more"
            message = (
                f"Delete CAD Document {label} and its related drawing CAD Document(s)?\n\n"
                f"{drawing_names}\n\n"
                "The EBOM Items are not deleted, but CAD associations and CAD-build usages are removed."
            )
            if QMessageBox.question(
                self,
                "Delete CAD Document",
                message,
                QMessageBox.Yes | QMessageBox.Cancel,
                QMessageBox.Cancel,
            ) != QMessageBox.Yes:
                return
            delete_related_drawings = True
        else:
            if QMessageBox.question(
                self,
                "Delete CAD Document",
                f"Delete CAD Document {label}?\n\n"
                "The EBOM Item is not deleted, but CAD associations and CAD-build usages are removed.",
                QMessageBox.Yes | QMessageBox.Cancel,
                QMessageBox.Cancel,
            ) != QMessageBox.Yes:
                return
        try:
            result = self.bom_service.delete_pdm_cad_document(
                int(cad_id), delete_related_drawings=delete_related_drawings
            )
        except Exception as exc:
            QMessageBox.warning(self, "Delete CAD Document", str(exc))
            return
        deleted_ids = list((result or {}).get("deleted_ids") or [int(cad_id)])
        self._remove_pdm_cad_documents_from_trees(deleted_ids)
        self._refresh_pdm_context_rows()
        if reselect_cad_id is not None:
            self._reselect_cad_in_current_view(int(reselect_cad_id))
        else:
            self.clear_details()
        count = int((result or {}).get("deleted_count") or 0)
        status_message = (
            f"Deleted {count} CAD Document{'s' if count != 1 else ''}."
            if count else
            "No CAD Document was deleted."
        )
        self.window().statusBar().showMessage(status_message, 6000)

    def _toggle_selected_ebom_cad_associations(self) -> None:
        item = self._selected_ebom_item_for_visual_action()
        if item is None:
            return
        if item.data(0, PDM_ASSOCIATIONS_SHOWN_ROLE):
            self._hide_ebom_cad_associations(item)
        else:
            self._show_ebom_cad_associations(item)

    def _pdm_scope_path_for_mode(self, mode: str | None = None) -> list:
        mode = str(mode or getattr(self, "_bom_mode", "cad"))
        if mode == "ebom":
            return list(getattr(self, "_pdm_ebom_scope_path", []) or [])
        return list(getattr(self, "_pdm_cad_scope_path", []) or [])

    @staticmethod
    def _pdm_int_or_none(value):
        try:
            if value in (None, ""):
                return None
            return int(value)
        except Exception:
            return None

    def _pdm_scope_identity(self, payload: dict, mode: str):
        payload = dict(payload or {})
        if mode == "cad":
            cad_id = self._pdm_int_or_none(
                payload.get("id")
                or payload.get("cad_document_id")
                or payload.get("child_cad_document_id")
            )
            member_id = self._pdm_int_or_none(payload.get("member_id"))
            if cad_id is None:
                return None
            return ("cad", cad_id, member_id)
        item_id = self._pdm_int_or_none(
            payload.get("bom_id")
            or payload.get("id")
            or payload.get("item_id")
        )
        if item_id is None:
            return None
        return ("ebom", item_id)

    def _pdm_payload_matches_scope_identity(self, payload: dict, identity, mode: str) -> bool:
        candidate = self._pdm_scope_identity(payload, mode)
        if candidate is None or identity is None:
            return False
        if mode == "cad":
            _kind, wanted_cad_id, wanted_member_id = identity
            _candidate_kind, candidate_cad_id, candidate_member_id = candidate
            if wanted_cad_id != candidate_cad_id:
                return False
            if wanted_member_id is not None and candidate_member_id is not None:
                return wanted_member_id == candidate_member_id
            return True
        return candidate == identity

    def _find_pdm_payload_path_for_scope(
        self, nodes: list[dict], identities: list, mode: str, depth: int = 0
    ) -> list[dict] | None:
        if not identities or depth >= len(identities):
            return None
        for node in nodes or []:
            if not self._pdm_payload_matches_scope_identity(node, identities[depth], mode):
                continue
            if depth == len(identities) - 1:
                return [dict(node)]
            child_path = self._find_pdm_payload_path_for_scope(
                list(node.get("children") or []), identities, mode, depth + 1
            )
            if child_path is not None:
                return [dict(node)] + child_path
        return None

    def _find_pdm_payload_path_to_identity(
        self, nodes: list[dict], identity, mode: str
    ) -> list[dict] | None:
        for node in nodes or []:
            if self._pdm_payload_matches_scope_identity(node, identity, mode):
                return [dict(node)]
            child_path = self._find_pdm_payload_path_to_identity(
                list(node.get("children") or []), identity, mode
            )
            if child_path is not None:
                return [dict(node)] + child_path
        return None

    def _pdm_scope_entry_from_payload(self, payload: dict, mode: str) -> dict:
        clean_payload = dict(payload or {})
        return {
            "label": self._pdm_payload_label(clean_payload, mode),
            "payload": clean_payload,
        }

    def _pdm_roots_for_reload_scope(
        self, mode: str, roots: list[dict], previous_scope_path: list
    ) -> tuple[list, list[dict]]:
        previous_scope_path = list(previous_scope_path or [])
        roots = list(roots or [])
        if not previous_scope_path:
            return [], roots

        identities = [
            self._pdm_scope_identity(entry.get("payload") or {}, mode)
            for entry in previous_scope_path
        ]
        identities = [identity for identity in identities if identity is not None]
        fresh_path = None
        if identities:
            fresh_path = self._find_pdm_payload_path_for_scope(roots, identities, mode)
            if fresh_path is None:
                fresh_path = self._find_pdm_payload_path_to_identity(
                    roots, identities[-1], mode
                )
        if fresh_path:
            refreshed_scope_path = [
                self._pdm_scope_entry_from_payload(payload, mode)
                for payload in fresh_path
            ]
            return refreshed_scope_path, [dict(fresh_path[-1])]

        # If a refresh cannot locate the scoped object in the new root data,
        # keep the user's isolated working scope instead of silently jumping
        # back to top level.  The user can explicitly press Top Level/change
        # breadcrumb scope when they want to leave it.
        stale_payload = dict(previous_scope_path[-1].get("payload") or {})
        return previous_scope_path, ([stale_payload] if stale_payload else [])

    def _pdm_payload_label(self, payload: dict, mode: str) -> str:
        if mode == "cad":
            label = str(
                payload.get("file_name")
                or payload.get("name")
                or "CAD Document"
            )
        else:
            label = str(
                payload.get("name")
                or payload.get("part_number")
                or payload.get("aes_number")
                or "Item"
            )
        return " ".join(label.split()) or ("CAD Document" if mode == "cad" else "Item")

    def _pdm_scope_entry_from_item(self, item: QTreeWidgetItem, mode: str) -> dict | None:
        payload = dict(item.data(0, PDM_NODE_PAYLOAD_ROLE) or {})
        if not payload:
            payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
        if not payload:
            return None
        label = self._pdm_payload_label(payload, mode)
        return {"label": label, "payload": payload}

    def _pdm_path_from_item(self, item: QTreeWidgetItem, mode: str) -> list:
        path = []
        current = item
        while current is not None:
            if not self._is_lazy_placeholder(current):
                if not (mode == "ebom" and current.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD):
                    entry = self._pdm_scope_entry_from_item(current, mode)
                    if entry:
                        path.append(entry)
            current = current.parent()
        path.reverse()
        return path

    def _isolate_pdm_tree_item(self, item: QTreeWidgetItem) -> None:
        tree = item.treeWidget() if item is not None else None
        mode = self._pdm_mode_for_tree(tree)
        if mode not in {"cad", "ebom"}:
            return
        if mode == "ebom" and item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
            item = item.parent()
            if item is None:
                return
        path = self._pdm_path_from_item(item, mode)
        if not path:
            return
        if mode == "cad":
            self._pdm_cad_scope_path = path
            self._render_pdm_cad_roots([dict(path[-1]["payload"])])
            target_tree = self._cad_tree
        else:
            self._pdm_ebom_scope_path = path
            self._render_pdm_ebom_roots([dict(path[-1]["payload"])])
            target_tree = self._ebom_tree
        try:
            root_item = target_tree.topLevelItem(0)
            if root_item is not None:
                target_tree.setCurrentItem(root_item)
                root_item.setSelected(True)
                self.on_tree_item_clicked(root_item, self._pdm_name_column_for_tree(target_tree))
        except Exception:
            pass

    def _open_pdm_scope_at(self, mode: str, index: int) -> None:
        path = self._pdm_scope_path_for_mode(mode)
        if index < 0 or index >= len(path):
            return
        path = path[: index + 1]
        payload = dict(path[-1].get("payload") or {})
        if payload.get("_folder_scope"):
            self._open_pdm_folder_snapshot(path[-1], mode)
            return
        if mode == "cad":
            self._pdm_cad_scope_path = path
            self._render_pdm_cad_roots([dict(path[-1]["payload"])])
            target_tree = self._cad_tree
        else:
            self._pdm_ebom_scope_path = path
            self._render_pdm_ebom_roots([dict(path[-1]["payload"])])
            target_tree = self._ebom_tree
        try:
            root_item = target_tree.topLevelItem(0)
            if root_item is not None:
                target_tree.setCurrentItem(root_item)
                root_item.setSelected(True)
                self.on_tree_item_clicked(root_item, self._pdm_name_column_for_tree(target_tree))
        except Exception:
            pass

    def _open_pdm_folder_snapshot(self, entry: dict, mode: str) -> None:
        payload = dict((entry or {}).get("payload") or {})
        children = list(payload.get("_folder_children") or [])
        tree = self._cad_tree if mode == "cad" else self._ebom_tree
        if mode == "cad":
            self._pdm_cad_scope_path = [entry]
        else:
            self._pdm_ebom_scope_path = [entry]
        tree.setUpdatesEnabled(False)
        try:
            tree.clear()
            for child in children:
                try:
                    tree.addTopLevelItem(child.clone())
                except Exception:
                    pass
            if mode == "cad":
                self._refresh_pdm_cad_filter()
            else:
                self._renumber_tree_rows(tree)
                self._refresh_ebom_filters()
        finally:
            tree.setUpdatesEnabled(True)
            tree.viewport().update()
        self._update_pdm_scope_bar()
        self._sync_visual_action_states()

    def _open_pdm_folder_item(self, item: QTreeWidgetItem) -> None:
        if item is None or not self._is_folder_tree_item(item):
            return
        tree = item.treeWidget()
        mode = self._pdm_mode_for_tree(tree)
        if mode not in {"cad", "ebom"}:
            return
        folder_id = item.data(0, BOM_TREE_FOLDER_ROLE)
        label = item.text(self._pdm_name_column_for_tree(tree)) or "Folder"
        entry = {
            "label": str(label),
            "payload": {
                "_folder_scope": mode,
                "_folder_id": int(folder_id),
                "_folder_children": [
                    item.child(index).clone()
                    for index in range(item.childCount())
                    if not self._is_lazy_placeholder(item.child(index))
                ],
            },
        }
        self._open_pdm_folder_snapshot(entry, mode)

    def _clear_pdm_isolation(self) -> None:
        mode = str(getattr(self, "_bom_mode", "cad"))
        if mode == "ebom":
            self._pdm_ebom_scope_path = []
            self._render_pdm_ebom_roots(getattr(self, "_pdm_ebom_roots", []) or [])
        else:
            self._pdm_cad_scope_path = []
            self._render_pdm_cad_roots(getattr(self, "_pdm_cad_roots", []) or [])
        self.clear_details()

    def _update_pdm_scope_bar(self) -> None:
        frame = getattr(self, "_bom_scope_frame", None)
        layout = getattr(self, "_bom_scope_path_layout", None)
        if frame is None or layout is None:
            return
        mode = str(getattr(self, "_bom_mode", "cad"))
        path = self._pdm_scope_path_for_mode(mode)
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
        frame.setVisible(bool(path))
        if not path:
            return
        for index, entry in enumerate(path):
            if index:
                separator = QLabel(">")
                separator.setProperty("scopeSeparator", True)
                layout.addWidget(separator)
            label = str(entry.get("label") or "")
            display_label = label if len(label) <= 34 else label[:31] + "..."
            button = QPushButton(display_label)
            button.setProperty("scopeCrumb", True)
            button.setFixedHeight(22)
            button.setToolTip(label)
            button.clicked.connect(
                lambda _checked=False, i=index, m=mode: self._open_pdm_scope_at(m, i)
            )
            layout.addWidget(button)
        layout.addStretch(1)

    def _sync_visual_action_states(self) -> None:
        try:
            mode = str(getattr(self, "_bom_mode", "cad"))
            scoped = bool(self._pdm_scope_path_for_mode(mode))
            if hasattr(self, "clear_pdm_scope_btn"):
                self.clear_pdm_scope_btn.setEnabled(scoped)
            item = self._selected_ebom_item_for_visual_action()
            associations = list(item.data(0, PDM_EBOM_ASSOCIATIONS_ROLE) or []) if item is not None else []
            shown = bool(item and item.data(0, PDM_ASSOCIATIONS_SHOWN_ROLE))
            if hasattr(self, "show_associated_cad_btn"):
                self.show_associated_cad_btn.setText("Hide CAD" if shown else "Show CAD")
                self.show_associated_cad_btn.setEnabled(bool(item and associations))
                self.show_associated_cad_btn.setToolTip(
                    "Hide associated CAD rows for the selected Item"
                    if shown else
                    "Show associated CAD rows under the selected EBOM Item"
                )
            cad_mode = mode == "cad"
            can_manage = self.perm.can("manage_parts")
            cad_item = self._selected_cad_item_for_ribbon()
            payload = dict(cad_item.data(0, PDM_CAD_PAYLOAD_ROLE) or {}) if cad_item is not None else {}
            checked_out_by_me = False
            try:
                checked_out_by_me = (
                    self.session.user_id is not None
                    and payload.get("checked_out_by") is not None
                    and int(payload["checked_out_by"]) == int(self.session.user_id)
                )
            except Exception:
                checked_out_by_me = False
            cad_checkout_editable = checked_out_by_me or self.perm.can("merge")
            is_assembly = str(payload.get("category") or "").upper() == "ASSEMBLY"
            cad_context = cad_item is not None
            if hasattr(self, "register_cad_btn"):
                self.register_cad_btn.setEnabled(bool(cad_mode and can_manage and self.session.project_id))
            if hasattr(self, "add_cad_component_btn"):
                self.add_cad_component_btn.setEnabled(bool(cad_mode and cad_context and can_manage and is_assembly and cad_checkout_editable))
            if hasattr(self, "delete_cad_btn"):
                checked_out = payload.get("checked_out_by") is not None
                self.delete_cad_btn.setEnabled(
                    bool(cad_context and can_manage and not checked_out)
                )
                self.delete_cad_btn.setToolTip(
                    "Check in or undo checkout before deleting this CAD Document"
                    if checked_out else
                    "Delete the selected CAD Document"
                )
        except Exception:
            pass

    # -------------------------
    # Context menu / tree actions
    # -------------------------
    def _pdm_folders(self, scope: str, refresh: bool = False) -> list[dict]:
        scope = str(scope or "EBOM").upper()
        attribute = "_cad_folders_cache" if scope == "CAD" else "_bom_folders_cache"
        cached = getattr(self, attribute, None)
        if cached is not None and not refresh:
            return list(cached)
        try:
            folders = (
                self.bom_service.list_cad_folders()
                if scope == "CAD"
                else self.bom_service.list_bom_folders()
            ) or []
        except Exception:
            folders = []
        for folder in folders:
            folder["scope"] = scope
        setattr(self, attribute, list(folders))
        return list(folders)

    @staticmethod
    def _pdm_folder_context_key(scope: str) -> str:
        return (
            "effective_parent_cad_document_id"
            if str(scope).upper() == "CAD"
            else "effective_parent_bom_id"
        )

    @staticmethod
    def _pdm_folder_members_key(scope: str) -> str:
        return "cad_document_ids" if str(scope).upper() == "CAD" else "item_ids"

    def _make_pdm_folder_tree_item(
        self, folder: dict, tree: QTreeWidget, scope: str
    ) -> QTreeWidgetItem:
        item = QTreeWidgetItem([""] * tree.columnCount())
        name_column = CAD_COL_NAME if str(scope).upper() == "CAD" else BOM_COL_NAME
        item.setText(name_column, str(folder.get("name") or "Folder"))
        item.setData(0, Qt.UserRole, None)
        item.setData(0, BOM_TREE_FOLDER_ROLE, int(folder["id"]))
        item.setData(0, BOM_TREE_FOLDER_SCOPE_ROLE, str(scope).upper())
        item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, True)
        item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, True)
        item.setIcon(name_column, _bom_type_icon("folder"))
        item.setToolTip(
            name_column,
            f"{str(scope).upper()} organizational folder; engineering structure is unchanged.",
        )
        return item

    def _pdm_folder_object_id(
        self, item: QTreeWidgetItem, scope: str
    ) -> int | None:
        if item is None or self._is_folder_tree_item(item):
            return None
        value = (
            item.data(0, PDM_CAD_DOCUMENT_ID_ROLE)
            if str(scope).upper() == "CAD"
            else item.data(0, Qt.UserRole)
        )
        if str(scope).upper() == "EBOM" and (
            item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD
        ):
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _render_pdm_folder_context(
        self, scope: str, parent_object_id, containers: list
    ) -> None:
        scope = str(scope or "EBOM").upper()
        folders = self._pdm_folders(scope)
        context_key = self._pdm_folder_context_key(scope)
        context_folders = [
            folder
            for folder in folders
            if folder.get(context_key) == parent_object_id
        ]
        if not context_folders:
            return
        roots = []
        children_by_folder = defaultdict(list)
        for folder in context_folders:
            parent_folder_id = folder.get("parent_folder_id")
            if parent_folder_id is None:
                roots.append(folder)
            else:
                children_by_folder[int(parent_folder_id)].append(folder)
        order = lambda row: (int(row.get("sort_order") or 0), int(row["id"]))
        roots.sort(key=order)
        for children in children_by_folder.values():
            children.sort(key=order)

        for container in list(containers or []):
            tree = (
                container
                if isinstance(container, QTreeWidget)
                else container.treeWidget()
            )
            if tree is None:
                continue

            def add_folder(folder: dict, display_parent) -> None:
                folder_item = self._make_pdm_folder_tree_item(
                    folder, tree, scope
                )
                self._container_add(display_parent, folder_item)
                assigned = {
                    int(value)
                    for value in (
                        folder.get(self._pdm_folder_members_key(scope)) or []
                    )
                }
                candidates = []
                search_container = display_parent
                for index in range(self._container_count(search_container)):
                    candidate = self._container_item(search_container, index)
                    candidate_id = self._pdm_folder_object_id(candidate, scope)
                    if candidate_id is not None and candidate_id in assigned:
                        candidates.append(candidate_id)
                for object_id in candidates:
                    for index in range(self._container_count(search_container)):
                        candidate = self._container_item(search_container, index)
                        if self._pdm_folder_object_id(candidate, scope) == object_id:
                            folder_item.addChild(
                                self._container_take(search_container, index)
                            )
                            break
                for child_folder in children_by_folder.get(
                    int(folder["id"]), []
                ):
                    add_folder(child_folder, folder_item)
                folder_item.setExpanded(False)

            for folder in roots:
                add_folder(folder, container)

    def _reload_pdm_folder_scope(self, scope: str) -> None:
        scope = str(scope or "EBOM").upper()
        self._pdm_folders(scope, refresh=True)
        tree = getattr(self, "_cad_tree", None) if scope == "CAD" else getattr(self, "_ebom_tree", None)
        if tree is None:
            return

        def unwrap_folders(container) -> None:
            index = self._container_count(container) - 1
            while index >= 0:
                child = self._container_item(container, index)
                if child is None:
                    index -= 1
                    continue
                if self._is_folder_tree_item(child):
                    unwrap_folders(child)
                    promoted = []
                    while child.childCount():
                        promoted.append(child.takeChild(0))
                    self._container_take(container, index)
                    for promoted_item in reversed(promoted):
                        if isinstance(container, QTreeWidget):
                            container.insertTopLevelItem(index, promoted_item)
                        else:
                            container.insertChild(index, promoted_item)
                else:
                    unwrap_folders(child)
                index -= 1

        tree.setUpdatesEnabled(False)
        try:
            unwrap_folders(tree)
            self._render_pdm_folder_context(scope, None, [tree])
            for item in list(self._iter_tree_items(tree)):
                if self._is_folder_tree_item(item) or self._is_lazy_placeholder(item):
                    continue
                if scope == "CAD":
                    if item.data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD:
                        continue
                    parent_id = item.data(0, PDM_CAD_DOCUMENT_ID_ROLE)
                else:
                    if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                        continue
                    parent_id = item.data(0, Qt.UserRole)
                if parent_id is None or not item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE):
                    continue
                self._render_pdm_folder_context(scope, int(parent_id), [item])
            if scope == "CAD":
                self._refresh_pdm_cad_filter()
            else:
                self._renumber_tree_rows(tree)
                self._refresh_ebom_filters()
        finally:
            tree.setUpdatesEnabled(True)
            tree.viewport().update()
        self._sync_visual_action_states()

    def _folder_record(self, folder_id: int) -> dict:
        for folder in (
            list(getattr(self, "_bom_folders_cache", []) or [])
            + list(getattr(self, "_cad_folders_cache", []) or [])
        ):
            try:
                if int(folder.get("id")) == int(folder_id):
                    return dict(folder)
            except Exception:
                continue
        try:
            folders = self._pdm_folders("EBOM", refresh=True) + self._pdm_folders(
                "CAD", refresh=True
            )
            return next(
                (dict(folder) for folder in folders if int(folder.get("id")) == int(folder_id)),
                {},
            )
        except Exception:
            return {}

    def add_bom_folder(self, parent_item=None, parent_folder_id=None) -> None:
        if not self.perm.can("manage_parts"):
            return
        scope = str(getattr(self, "_bom_mode", "EBOM") or "EBOM").upper()
        if parent_folder_id is not None:
            scope = str(
                self._folder_record(int(parent_folder_id)).get("scope") or scope
            ).upper()
        tree = self._current_pdm_tree()
        selected = (
            parent_item
            if isinstance(parent_item, QTreeWidgetItem)
            else None
        )
        parent_bom_id = None
        parent_cad_document_id = None
        if parent_folder_id is None and selected is not None:
            selected_folder_id = selected.data(0, BOM_TREE_FOLDER_ROLE)
            if selected_folder_id:
                parent_folder_id = int(selected_folder_id)
                scope = str(
                    selected.data(0, BOM_TREE_FOLDER_SCOPE_ROLE) or scope
                ).upper()
            else:
                if scope == "CAD":
                    payload = dict(selected.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
                    if str(payload.get("category") or "").upper() == "ASSEMBLY":
                        parent_cad_document_id = int(
                            selected.data(0, PDM_CAD_DOCUMENT_ID_ROLE)
                        )
                else:
                    selected_id = selected.data(0, Qt.UserRole)
                    selected_type = str(
                        selected.text(BOM_COL_TYPE) or ""
                    ).strip().lower()
                    if (
                        selected_id is not None
                        and selected_type in {"asm", "assembly"}
                    ):
                        parent_bom_id = int(selected_id)

        if parent_folder_id is not None:
            parent_folder = self._folder_record(int(parent_folder_id))
            location = f"inside folder '{parent_folder.get('name') or 'Folder'}'"
        elif parent_cad_document_id is not None:
            location = (
                "inside CAD assembly '"
                + str(
                    selected.text(CAD_COL_NAME)
                    if selected is not None
                    else parent_cad_document_id
                )
                + "'"
            )
        elif parent_bom_id is not None:
            parent = self.bom_service.get_part_details(int(parent_bom_id)) or {}
            location = f"inside assembly '{parent.get('name') or parent_bom_id}'"
        else:
            location = "at the project root"
        name, accepted = QInputDialog.getText(
            self, "Add BOM Folder", f"Folder name ({location}):"
        )
        if not accepted:
            return
        try:
            if scope == "CAD":
                self.bom_service.create_cad_folder(
                    name,
                    parent_cad_document_id=parent_cad_document_id,
                    parent_folder_id=parent_folder_id,
                )
            else:
                self.bom_service.create_bom_folder(
                    name,
                    parent_bom_id=parent_bom_id,
                    parent_folder_id=parent_folder_id,
                )
            self._reload_pdm_folder_scope(scope)
        except Exception as exc:
            QMessageBox.critical(self, "Add BOM Folder", f"Could not create folder:\n{exc}")

    def _assign_bom_folder_items(self, folder_id: int) -> None:
        folder = self._folder_record(int(folder_id))
        scope = str(folder.get("scope") or "EBOM").upper()
        try:
            eligible = (
                self.bom_service.eligible_cad_folder_documents(int(folder_id))
                if scope == "CAD"
                else self.bom_service.eligible_bom_folder_items(int(folder_id))
            ) or []
        except Exception as exc:
            QMessageBox.critical(self, "Folder Items", f"Could not load eligible items:\n{exc}")
            return
        dialog = QDialog(self)
        noun = "CAD Documents" if scope == "CAD" else "Items"
        dialog.setWindowTitle(f"{noun} in {folder.get('name') or 'Folder'}")
        dialog.resize(560, 480)
        layout = QVBoxLayout(dialog)
        layout.addWidget(
            QLabel(
                f"Select the direct {noun} to display in this organizational folder."
            )
        )
        item_list = QListWidget()
        item_list.setAlternatingRowColors(True)
        for part in eligible:
            if scope == "CAD":
                label = (
                    f"{part.get('file_name') or part.get('name') or part.get('id')}  "
                    f"[{part.get('category') or 'CAD'}]"
                )
            else:
                label = (
                    f"{part.get('part_number') or 'No Number'}  "
                    f"{part.get('name') or ''}  [{part.get('type') or ''}]"
                )
            list_item = QListWidgetItem(label)
            list_item.setData(Qt.UserRole, int(part["id"]))
            list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable)
            list_item.setCheckState(Qt.Checked if part.get("assigned") else Qt.Unchecked)
            item_list.addItem(list_item)
        layout.addWidget(item_list, 1)
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)

        def save_items():
            selected_ids = [
                int(item_list.item(row).data(Qt.UserRole))
                for row in range(item_list.count())
                if item_list.item(row).checkState() == Qt.Checked
            ]
            try:
                if scope == "CAD":
                    self.bom_service.set_cad_folder_documents(
                        int(folder_id), selected_ids
                    )
                else:
                    self.bom_service.set_bom_folder_items(
                        int(folder_id), selected_ids
                    )
                self._reload_pdm_folder_scope(scope)
                dialog.accept()
            except Exception as exc:
                QMessageBox.critical(dialog, "Folder Items", f"Could not save folder items:\n{exc}")

        buttons.accepted.connect(save_items)
        buttons.rejected.connect(dialog.reject)
        dialog.exec_()

    def _rename_bom_folder(self, folder_id: int) -> None:
        folder = self._folder_record(int(folder_id))
        name, accepted = QInputDialog.getText(
            self, "Rename BOM Folder", "Folder name:", text=str(folder.get("name") or "")
        )
        if not accepted:
            return
        try:
            self.bom_service.rename_bom_folder(int(folder_id), name)
            self._reload_pdm_folder_scope(
                str(folder.get("scope") or "EBOM").upper()
            )
        except Exception as exc:
            QMessageBox.critical(self, "Rename BOM Folder", f"Could not rename folder:\n{exc}")

    def _delete_bom_folder(self, folder_id: int) -> None:
        folder = self._folder_record(int(folder_id))
        answer = QMessageBox.question(
            self,
            "Delete Organizational Folder",
            f"Delete folder '{folder.get('name') or 'Folder'}'?\n\n"
            "Its subfolders will also be deleted. Contained objects will return "
            "to their normal structure position; no Item, CAD Document, or "
            "engineering relation will be deleted.",
            QMessageBox.Yes | QMessageBox.Cancel,
            QMessageBox.Cancel,
        )
        if answer != QMessageBox.Yes:
            return
        scope = str(folder.get("scope") or "EBOM").upper()
        try:
            self.bom_service.delete_bom_folder(int(folder_id))
            self._reload_pdm_folder_scope(scope)
            self._clear_folder_selection()
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Delete Organizational Folder",
                f"Could not delete folder:\n{exc}",
            )

    def _show_folder_context_menu(self, tree: QTreeWidget, item: QTreeWidgetItem, folder_id: int) -> None:
        menu = QMenu(self)
        folder = self._folder_record(int(folder_id))
        scope = str(folder.get("scope") or "EBOM").upper()
        open_action = menu.addAction("Open")
        menu.addSeparator()
        assign_action = menu.addAction(
            "Assign CAD Documents..." if scope == "CAD" else "Assign Items..."
        )
        subfolder_action = menu.addAction("Add Subfolder")
        rename_action = menu.addAction("Rename Folder")
        menu.addSeparator()
        delete_action = menu.addAction("Delete Folder")
        can_manage = self.perm.can("manage_parts")
        for action in (assign_action, subfolder_action, rename_action, delete_action):
            action.setEnabled(can_manage)
        open_action.triggered.connect(lambda: self._open_pdm_folder_item(item))
        assign_action.triggered.connect(lambda: self._assign_bom_folder_items(int(folder_id)))
        subfolder_action.triggered.connect(
            lambda: self.add_bom_folder(parent_item=item, parent_folder_id=int(folder_id))
        )
        rename_action.triggered.connect(lambda: self._rename_bom_folder(int(folder_id)))
        delete_action.triggered.connect(lambda: self._delete_bom_folder(int(folder_id)))
        menu.exec_(tree.viewport().mapToGlobal(tree.visualItemRect(item).bottomLeft()))

    def _folder_context_for_pdm_item(
        self, item: QTreeWidgetItem, scope: str
    ):
        parent = item.parent()
        if parent is None:
            return None
        if self._is_folder_tree_item(parent):
            folder = self._folder_record(int(parent.data(0, BOM_TREE_FOLDER_ROLE)))
            return folder.get(self._pdm_folder_context_key(scope))
        return self._pdm_folder_object_id(parent, scope)

    def _selected_pdm_folder_objects(
        self, item: QTreeWidgetItem, scope: str
    ) -> list[QTreeWidgetItem]:
        tree = item.treeWidget() if item is not None else None
        selected = []
        if tree is not None:
            cached = list(getattr(tree, "_context_menu_selection", []) or [])
            if item in cached:
                selected = cached
            else:
                selected = list(tree.selectedItems())
        if item not in selected:
            selected = [item]
        result = []
        seen = set()
        for candidate in selected:
            object_id = self._pdm_folder_object_id(candidate, scope)
            if object_id is None or object_id in seen:
                continue
            seen.add(object_id)
            result.append(candidate)
        return result or [item]

    @staticmethod
    def _pdm_tree_item_contains(
        parent: QTreeWidgetItem | None, child: QTreeWidgetItem | None
    ) -> bool:
        current = child.parent() if child is not None else None
        while current is not None:
            if current is parent:
                return True
            current = current.parent()
        return False

    def _nested_pdm_folder_selection_pairs(
        self, items: list[QTreeWidgetItem], scope: str
    ) -> list[tuple[QTreeWidgetItem, QTreeWidgetItem]]:
        scope = str(scope or "EBOM").upper()
        object_items = [
            item
            for item in list(items or [])
            if self._pdm_folder_object_id(item, scope) is not None
        ]
        pairs = []
        for index, first in enumerate(object_items):
            for second in object_items[index + 1:]:
                if self._pdm_tree_item_contains(first, second):
                    pairs.append((first, second))
                elif self._pdm_tree_item_contains(second, first):
                    pairs.append((second, first))
        return pairs

    def _move_pdm_objects_to_folder(
        self,
        items: list[QTreeWidgetItem],
        scope: str,
        target_folder_id: int | None,
    ) -> None:
        scope = str(scope or "EBOM").upper()
        valid_items = [
            item
            for item in list(items or [])
            if self._pdm_folder_object_id(item, scope) is not None
        ]
        if not valid_items:
            return
        nested_pairs = self._nested_pdm_folder_selection_pairs(valid_items, scope)
        if nested_pairs:
            parent, child = nested_pairs[0]
            QMessageBox.warning(
                self,
                "Move to Folder",
                "The selection contains a parent and one of its children:\n\n"
                f"{parent.text(self._pdm_name_column_for_tree(parent.treeWidget()))} > "
                f"{child.text(self._pdm_name_column_for_tree(child.treeWidget()))}\n\n"
                "Move only objects from the same direct assembly level.",
            )
            return
        contexts = {
            self._folder_context_for_pdm_item(item, scope)
            for item in valid_items
        }
        if len(contexts) != 1:
            QMessageBox.warning(
                self,
                "Move to Folder",
                "Select objects from the same assembly level. A folder cannot "
                "contain objects from different structural parents.",
            )
            return
        object_ids = {
            int(self._pdm_folder_object_id(item, scope))
            for item in valid_items
        }
        try:
            if target_folder_id is not None:
                target = self._folder_record(int(target_folder_id))
                members_key = self._pdm_folder_members_key(scope)
                selected = {
                    int(value) for value in (target.get(members_key) or [])
                }
                selected.update(object_ids)
                if scope == "CAD":
                    self.bom_service.set_cad_folder_documents(
                        int(target_folder_id), sorted(selected)
                    )
                else:
                    self.bom_service.set_bom_folder_items(
                        int(target_folder_id), sorted(selected)
                    )
            else:
                members_key = self._pdm_folder_members_key(scope)
                source_ids = {
                    int(item.parent().data(0, BOM_TREE_FOLDER_ROLE))
                    for item in valid_items
                    if item.parent() is not None
                    and self._is_folder_tree_item(item.parent())
                }
                for source_id in source_ids:
                    source = self._folder_record(source_id)
                    remaining = [
                        int(value)
                        for value in (source.get(members_key) or [])
                        if int(value) not in object_ids
                    ]
                    if scope == "CAD":
                        self.bom_service.set_cad_folder_documents(
                            source_id, remaining
                        )
                    else:
                        self.bom_service.set_bom_folder_items(
                            source_id, remaining
                        )
            self._reload_pdm_folder_scope(scope)
        except Exception as exc:
            QMessageBox.warning(self, "Move to Folder", str(exc))

    def _add_move_to_folder_menu(
        self, menu: QMenu, item: QTreeWidgetItem, scope: str
    ) -> None:
        scope = str(scope or "EBOM").upper()
        selected_items = self._selected_pdm_folder_objects(item, scope)
        contexts = {
            self._folder_context_for_pdm_item(selected, scope)
            for selected in selected_items
        }
        context = self._folder_context_for_pdm_item(item, scope)
        context_key = self._pdm_folder_context_key(scope)
        folders = [
            folder
            for folder in self._pdm_folders(scope)
            if folder.get(context_key) == context
        ]
        if not folders:
            return
        count = len(selected_items)
        move_menu = menu.addMenu(
            f"Move {count} Selected to Folder" if count > 1 else "Move to Folder"
        )
        move_menu.setEnabled(len(contexts) == 1)
        if len(contexts) != 1:
            move_menu.setToolTip(
                "Selected objects must belong to the same assembly level."
            )
        current_folder_ids = {
            int(selected.parent().data(0, BOM_TREE_FOLDER_ROLE))
            for selected in selected_items
            if selected.parent() is not None
            and self._is_folder_tree_item(selected.parent())
        }
        no_folder = move_menu.addAction("No Folder")
        no_folder.setEnabled(bool(current_folder_ids))
        no_folder.triggered.connect(
            lambda _checked=False, rows=list(selected_items), value=scope:
            self._move_pdm_objects_to_folder(rows, value, None)
        )
        move_menu.addSeparator()

        children_by_folder = defaultdict(list)
        roots = []
        folder_ids = {int(folder["id"]) for folder in folders}
        for folder in folders:
            parent_folder_id = folder.get("parent_folder_id")
            if (
                parent_folder_id is not None
                and int(parent_folder_id) in folder_ids
            ):
                children_by_folder[int(parent_folder_id)].append(folder)
            else:
                roots.append(folder)
        folder_order = lambda row: (
            int(row.get("sort_order") or 0),
            str(row.get("name") or "").lower(),
            int(row["id"]),
        )
        roots.sort(key=folder_order)
        for children in children_by_folder.values():
            children.sort(key=folder_order)

        def add_folder_target(
            parent_menu: QMenu, folder: dict, visited: set[int] | None = None
        ) -> None:
            folder_id = int(folder["id"])
            child_folders = list(children_by_folder.get(folder_id) or [])
            target_menu = (
                parent_menu.addMenu(str(folder.get("name") or "Folder"))
                if child_folders
                else None
            )
            action_parent = target_menu if target_menu is not None else parent_menu
            action = action_parent.addAction(
                "Move Here" if target_menu is not None else str(folder.get("name") or "Folder")
            )
            action.setEnabled(
                not (
                    len(current_folder_ids) == 1
                    and folder_id in current_folder_ids
                    and all(
                        selected.parent() is not None
                        and self._is_folder_tree_item(selected.parent())
                        for selected in selected_items
                    )
                )
            )
            action.triggered.connect(
                lambda _checked=False, rows=list(selected_items), value=scope,
                       target_folder_id=folder_id:
                self._move_pdm_objects_to_folder(rows, value, target_folder_id)
            )
            if not child_folders:
                return
            target_menu.addSeparator()
            next_visited = set(visited or set())
            next_visited.add(folder_id)
            for child_folder in child_folders:
                child_id = int(child_folder["id"])
                if child_id in next_visited:
                    continue
                add_folder_target(target_menu, child_folder, next_visited)

        for folder in roots:
            add_folder_target(move_menu, folder)

    @staticmethod
    def _pdm_association_types() -> list[tuple[str, str]]:
        return [
            ("Owner - drives structure and attributes", "OWNER"),
            ("Contributing image", "CONTRIBUTING_IMAGE"),
            ("Image - alternate CAD representation", "IMAGE"),
            ("Contributing content", "CONTRIBUTING_CONTENT"),
            ("Content - supporting CAD", "CONTENT"),
        ]

    def _prompt_pdm_association_type(
        self, current: str = "OWNER", title: str = "CAD-Item Association"
    ) -> str | None:
        values = self._pdm_association_types()
        current = str(current or "OWNER").upper()
        index = next(
            (position for position, (_label, value) in enumerate(values) if value == current),
            0,
        )
        selected, accepted = QInputDialog.getItem(
            self, title, "Association type:",
            [label for label, _value in values], index, False,
        )
        if not accepted:
            return None
        return next(value for label, value in values if label == selected)

    def _ensure_item_checked_out_for_pdm_change(self, item_id: int) -> bool:
        """Association/structure edits are Item data and require an Item checkout."""
        details = self.bom_service.get_part_details(int(item_id)) or {}
        if details.get("locked"):
            lock_user_id = details.get("locked_by_user_id")
            if (
                lock_user_id is not None and self.session.user_id is not None
                and int(lock_user_id) != int(self.session.user_id)
            ):
                QMessageBox.warning(
                    self, "Item Checked Out",
                    f"This Item is checked out by {details.get('locked_by_username') or 'another user'}.",
                )
                return False
            if str(details.get("checkout_origin") or "ITEM").upper() == "CAD":
                try:
                    self.bom_service.checkout_item(int(item_id))
                except Exception as exc:
                    QMessageBox.warning(self, "Item Checkout", str(exc))
                    return False
            return True
        answer = QMessageBox.question(
            self,
            "Check Out Item",
            "This operation changes the selected Item. Check out the Item now?\n\n"
            "The Item data is reserved. You can check out its OWNER CAD Document separately when needed.",
            QMessageBox.Yes | QMessageBox.Cancel,
            QMessageBox.Yes,
        )
        if answer != QMessageBox.Yes:
            return False
        state = str(
            details.get("revision_state") or details.get("lifecycle_state") or ""
        ).strip().lower()
        revision_code = None
        if state == "released":
            try:
                suggested = self.bom_service.suggest_next_revision(int(item_id))
            except Exception:
                suggested = ""
            revision_code, accepted = QInputDialog.getText(
                self,
                "Check Out Released Item",
                "The Released Item remains immutable. Enter the new revision to create:",
                QLineEdit.Normal,
                suggested,
            )
            if not accepted or not str(revision_code or "").strip():
                return False
        try:
            self.bom_service.checkout_item(
                int(item_id), released_revision_code=revision_code,
                include_owner_cad=False,
            )
        except Exception as exc:
            QMessageBox.warning(self, "Check Out Item", str(exc))
            return False
        return True

    def _reload_pdm_structure_views(self) -> None:
        """Refresh the current PDM context without rebuilding either tree.

        This method used to be the common "reload everything" escape hatch.
        Action handlers call it after small mutations, so keep it intentionally
        narrow: update selected/touched Item rows, visible CAD rows, detail
        panels and filters.  Startup, mode switches and explicit user Refresh
        still call the full loaders directly.
        """
        item_ids = set()
        cad_ids = set()
        try:
            if getattr(self, "current_part_id", None) is not None:
                item_ids.add(int(self.current_part_id))
        except Exception:
            pass
        try:
            if getattr(self, "current_cad_document_id", None) is not None:
                cad_ids.add(int(self.current_cad_document_id))
        except Exception:
            pass
        for tree in (getattr(self, "_cad_tree", None), getattr(self, "_ebom_tree", None)):
            try:
                item = tree.currentItem() if tree is not None else None
                if item is None:
                    continue
                if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                    value = item.data(0, PDM_CAD_DOCUMENT_ID_ROLE)
                    if value is not None:
                        cad_ids.add(int(value))
                else:
                    value = item.data(0, Qt.UserRole)
                    if value is not None:
                        item_ids.add(int(value))
            except Exception:
                continue
        self._refresh_pdm_context_rows(item_ids=item_ids, cad_ids=cad_ids)

    def _refresh_pdm_context_rows(
        self,
        *,
        item_ids=None,
        cad_ids=None,
        refresh_cad_branches: bool = False,
    ) -> None:
        """Targeted repaint for affected EBOM Items and CAD Documents."""
        normalized_item_ids = set()
        normalized_cad_ids = set()
        for value in item_ids or []:
            try:
                normalized_item_ids.add(int(value))
            except Exception:
                continue
        for value in cad_ids or []:
            try:
                normalized_cad_ids.add(int(value))
            except Exception:
                continue

        for item_id in sorted(normalized_item_ids):
            try:
                self._refresh_part_in_tree(int(item_id))
                self._refresh_ebom_association_rows_for_item(int(item_id))
                self._invalidate_doc_indicator(int(item_id))
            except Exception:
                pass
        if normalized_item_ids:
            try:
                self._renumber_full_bom_tree_rows()
                self._sync_search_tree_row_numbers()
            except Exception:
                pass

        if normalized_cad_ids:
            if refresh_cad_branches:
                for cad_id in sorted(normalized_cad_ids):
                    self._refresh_pdm_cad_structure_branch(int(cad_id))
            self.refresh_cad_documents_after_merge(normalized_cad_ids)

        try:
            current_part = getattr(self, "current_part_id", None)
            if current_part is not None and int(current_part) in normalized_item_ids:
                self.display_details(int(current_part))
        except Exception:
            pass
        try:
            current_cad = getattr(self, "current_cad_document_id", None)
            if current_cad is not None and int(current_cad) in normalized_cad_ids:
                self._reselect_cad_in_current_view(int(current_cad))
        except Exception:
            pass
        self._sync_visual_action_states()

    def _refresh_ebom_association_rows_for_item(self, item_id: int) -> None:
        try:
            rows = list(self.bom_service.list_item_cad_associations(int(item_id)) or [])
        except Exception:
            rows = []
        try:
            self._ebom_associations_by_item[int(item_id)] = rows
        except Exception:
            pass
        tree = getattr(self, "_ebom_tree", None)
        if tree is None:
            return
        for item in list(self._find_tree_items(int(item_id), tree)):
            try:
                if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                    continue
                was_shown = bool(item.data(0, PDM_ASSOCIATIONS_SHOWN_ROLE))
                was_expanded = bool(item.isExpanded())
                item.setData(0, PDM_EBOM_ASSOCIATIONS_ROLE, list(rows))
                if was_shown:
                    self._remove_direct_ebom_cad_rows(item)
                    self._show_ebom_cad_associations(item, refresh_filters=False)
                    item.setExpanded(was_expanded)
            except Exception:
                continue
        try:
            self._refresh_ebom_filters()
        except Exception:
            pass

    def _refresh_pdm_ebom_structure_branch(self, item_id: int) -> None:
        """Refresh one loaded EBOM branch after add/delete/usage changes."""
        tree = getattr(self, "_ebom_tree", None)
        if tree is None or not self.session.project_id:
            return
        try:
            data = self.bom_service.get_released_ebom_project(int(self.session.project_id)) or {}
            roots = list(data.get("roots") or [])
            self._pdm_ebom_roots = roots
            associations_by_item = defaultdict(list)
            for document in self.bom_service.list_pdm_cad_documents() or []:
                if str(document.get("category") or "").upper() == "DRAWING":
                    continue
                for association in self._pdm_document_associations(document):
                    if association.get("item_id") is None:
                        continue
                    item_document = dict(document)
                    item_document.update(association)
                    item_document["association_id"] = (
                        association.get("association_id") or association.get("id")
                    )
                    item_document["item_id"] = int(association["item_id"])
                    item_document["association_type"] = str(
                        association.get("association_type") or "CONTENT"
                    ).upper()
                    item_document["related_drawings"] = []
                    associations_by_item[int(association["item_id"])].append(item_document)
            self._ebom_associations_by_item = associations_by_item
            node = self._find_payload_node_by_id(roots, int(item_id), "bom_id")
        except Exception:
            node = None
        if not node:
            self._refresh_part_in_tree(int(item_id))
            return

        matches = list(self._find_tree_items(int(item_id), tree))
        if not matches and not getattr(self, "_pdm_ebom_scope_path", []):
            try:
                self._add_released_ebom_node(node, associations_by_item=associations_by_item)
                matches = list(self._find_tree_items(int(item_id), tree))
            except Exception:
                matches = []
        for item in matches:
            try:
                was_loaded = bool(item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE))
                was_expanded = bool(item.isExpanded())
                associations_shown = bool(item.data(0, PDM_ASSOCIATIONS_SHOWN_ROLE))
                payload = dict(node)
                payload["id"] = int(payload.get("bom_id") or payload.get("id"))
                payload["current_version"] = str(
                    payload.get("version_label") or payload.get("current_version") or ""
                )
                payload["status"] = str(payload.get("state") or payload.get("status") or "")
                payload["relation_parent_id"] = payload.get("effective_parent_bom_id")
                payload["usage_id"] = payload.get("item_usage_id") or payload.get("usage_id")
                payload["quantity"] = int(payload.get("source_quantity") or 1)
                payload["_has_children"] = bool(payload.get("children"))
                self._apply_tree_item_data(item, payload)
                item.setText(BOM_COL_AES, str(payload.get("part_number") or ""))
                item.setData(0, PDM_OBJECT_KIND_ROLE, PDM_OBJECT_ITEM)
                item.setData(0, PDM_NODE_PAYLOAD_ROLE, dict(payload))
                children = list(payload.get("children") or [])
                item.setData(0, PDM_CHILDREN_PAYLOAD_ROLE, children)
                item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, bool(children))
                item.setData(
                    0,
                    PDM_EBOM_ASSOCIATIONS_ROLE,
                    list(associations_by_item.get(int(payload["id"]), [])),
                )
                item.setText(EBOM_COL_SOURCE_QTY, str(int(payload.get("source_quantity") or 1)))
                item.setText(
                    EBOM_COL_EFFECTIVE_QTY,
                    str(int(payload.get("effective_quantity") or 1)),
                )
                item.setText(EBOM_COL_LEVEL, str(int(payload.get("level") or 0)))
                if was_loaded:
                    item.takeChildren()
                    if associations_shown:
                        item.setData(0, PDM_ASSOCIATIONS_SHOWN_ROLE, False)
                        self._show_ebom_cad_associations(item, refresh_filters=False)
                    for child in children:
                        if self._is_deleted_bom_payload(child):
                            continue
                        self._add_released_ebom_node(child, item, associations_by_item)
                    self._render_pdm_folder_context("EBOM", int(item_id), [item])
                    item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, True)
                    item.setExpanded(was_expanded and bool(item.childCount()))
                else:
                    item.takeChildren()
                    item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, not bool(children))
                    self._ensure_pdm_lazy_placeholder(item)
            except Exception:
                continue
        try:
            self._renumber_tree_rows(tree)
            self._refresh_ebom_filters()
        except Exception:
            pass

    def _refresh_pdm_cad_structure_branch(self, cad_id: int) -> None:
        """Refresh one loaded CAD branch after member/quantity changes."""
        try:
            data = self.bom_service.get_pdm_cad_structure() or {}
            roots = list(data.get("roots") or [])
            self._pdm_cad_roots = roots
            node = self._find_payload_node_by_id(roots, int(cad_id), "id")
        except Exception:
            node = None
        if not node:
            self.refresh_cad_documents_after_merge([int(cad_id)])
            return

        for item in list(self._find_pdm_cad_items([int(cad_id)])):
            if item.treeWidget() is not getattr(self, "_cad_tree", None):
                continue
            try:
                was_loaded = bool(item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE))
                was_expanded = bool(item.isExpanded())
                children = list(node.get("children") or [])
                item.setData(0, PDM_CHILDREN_PAYLOAD_ROLE, children)
                item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, bool(children))
                self._apply_pdm_cad_tree_item_data(item, dict(node))
                if was_loaded:
                    item.takeChildren()
                    for child in children:
                        self._add_pdm_cad_node(child, item)
                    self._render_pdm_folder_context("CAD", int(cad_id), [item])
                    item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, True)
                    item.setExpanded(was_expanded and bool(item.childCount()))
                else:
                    item.takeChildren()
                    item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, not bool(children))
                    self._ensure_pdm_lazy_placeholder(item)
            except Exception:
                continue
        try:
            self._refresh_pdm_cad_filter()
        except Exception:
            pass

    def _remove_pdm_cad_documents_from_trees(self, cad_ids) -> None:
        wanted = set()
        for value in cad_ids or []:
            try:
                wanted.add(int(value))
            except Exception:
                continue
        if not wanted:
            return
        for item in list(self._find_pdm_cad_items(wanted)):
            try:
                parent = item.parent()
                tree = item.treeWidget()
                if parent is not None:
                    parent.removeChild(item)
                elif tree is not None:
                    index = tree.indexOfTopLevelItem(item)
                    if index >= 0:
                        tree.takeTopLevelItem(index)
            except Exception:
                continue
        for tree in (getattr(self, "_cad_tree", None), getattr(self, "_ebom_tree", None)):
            try:
                tree.viewport().update()
            except Exception:
                pass
        try:
            self._refresh_pdm_cad_filter()
            self._refresh_ebom_filters()
        except Exception:
            pass

    def _associate_specific_cad_to_item(self, cad_id: int, item_id: int) -> None:
        if not self._ensure_item_checked_out_for_pdm_change(int(item_id)):
            return
        association_type = self._prompt_pdm_association_type()
        if association_type is None:
            return
        try:
            self.bom_service.associate_cad_document(
                int(item_id), int(cad_id), association_type
            )
        except Exception as exc:
            QMessageBox.warning(self, "Associate CAD Document", str(exc))
            return
        self._refresh_pdm_context_rows(item_ids=[int(item_id)], cad_ids=[int(cad_id)])

    def _register_and_associate_cad(self, item_id: int) -> None:
        if self._ensure_item_checked_out_for_pdm_change(int(item_id)):
            self.register_cad_document(int(item_id))

    def _associate_cad_to_an_item(self, cad_id: int) -> None:
        items = list(self.bom_service.list_pdm_items() or [])
        if not items:
            QMessageBox.information(self, "Associate CAD Document", "No Items are available.")
            return
        labels = [self._item_identity_text(row) for row in items]
        selected, accepted = QInputDialog.getItem(
            self, "Associate CAD Document", "Target Item:", labels, 0, False
        )
        if not accepted:
            return
        self.manage_cad_item_associations(
            int(items[labels.index(selected)]["id"]),
            focus_cad_id=int(cad_id),
        )

    def _selected_pdm_rows(self, tree: QTreeWidget, fallback: QTreeWidgetItem) -> list[QTreeWidgetItem]:
        """Return the stable right-click selection, preserving multi-select."""
        rows = []
        try:
            cached = list(getattr(tree, "_context_menu_selection", []) or [])
            selected = cached or list(tree.selectedItems())
        except Exception:
            selected = []
        for row in selected or []:
            if row is None or self._is_lazy_placeholder(row) or self._is_folder_tree_item(row):
                continue
            if row not in rows:
                rows.append(row)
        if fallback is not None and fallback not in rows:
            rows.append(fallback)
        return rows

    def _pdm_ebom_relation_parent_for_item(self, item: QTreeWidgetItem) -> int | None:
        if item is None:
            return None
        try:
            occurrence = item.data(0, BOM_TREE_OCCURRENCE_ROLE) or {}
            parent_id = occurrence.get("parent_id")
            if parent_id is not None:
                return int(parent_id)
        except Exception:
            pass
        try:
            payload = item.data(0, PDM_NODE_PAYLOAD_ROLE) or {}
            parent_id = payload.get("effective_parent_bom_id") or payload.get("relation_parent_id")
            if parent_id is not None:
                return int(parent_id)
        except Exception:
            pass
        parent = item.parent()
        while parent is not None:
            if self._is_folder_tree_item(parent):
                parent = parent.parent()
                continue
            if parent.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                parent = parent.parent()
                continue
            try:
                value = parent.data(0, Qt.UserRole)
                return int(value) if value is not None else None
            except Exception:
                return None
        return None

    def _pdm_ebom_relation_selection_for_item(self, item: QTreeWidgetItem) -> dict | None:
        if item is None or item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
            return None
        try:
            child_id = int(item.data(0, Qt.UserRole))
        except Exception:
            return None
        return {
            "child_id": child_id,
            "child_name": str(item.text(BOM_COL_NAME) or child_id),
            "source_parent_id": self._pdm_ebom_relation_parent_for_item(item),
        }

    def _apply_pdm_ebom_structure_operation(
        self,
        selected_items: list[QTreeWidgetItem],
        mode: str,
    ) -> None:
        selections = []
        for row in selected_items or []:
            payload = self._pdm_ebom_relation_selection_for_item(row)
            if payload is not None:
                selections.append(payload)
        if not selections:
            return QMessageBox.warning(self, "Item Structure", "Select one or more EBOM Items.")
        self._start_pdm_structure_target_selection(
            "EBOM",
            str(mode).lower(),
            selections,
            "Select the target EBOM assembly in the tree. You can search, filter, expand, or isolate first. Press Esc to cancel.",
        )

    def _finish_pdm_ebom_structure_operation(
        self,
        target_parent_id: int,
        selections: list[dict],
        mode: str,
    ) -> None:
        try:
            result = self.bom_service.apply_child_relation_operation(
                int(target_parent_id), selections, str(mode).lower()
            )
        except Exception as exc:
            QMessageBox.warning(self, "Item Structure", str(exc))
            return
        affected = {int(result.get("target_parent_id") or target_parent_id)}
        affected.update(int(value) for value in (result.get("source_parent_ids") or []))
        affected.update(int(value) for value in (result.get("child_ids") or []))
        for item_id in sorted(affected):
            self._refresh_pdm_ebom_structure_branch(int(item_id))
        try:
            self._refresh_ebom_filters()
        except Exception:
            pass
        verb = "moved" if str(mode).lower() == "move" else "copied"
        try:
            self.window().statusBar().showMessage(
                f"{len(result.get('child_ids') or [])} Item usage(s) {verb}."
            )
        except Exception:
            pass

    def _remove_pdm_ebom_children_from_tree(
        self,
        parent_item_id: int,
        parent_name: str,
        selected_items: list[QTreeWidgetItem],
    ) -> None:
        child_ids = []
        for row in selected_items or []:
            selection = self._pdm_ebom_relation_selection_for_item(row)
            if not selection:
                continue
            if selection.get("source_parent_id") != int(parent_item_id):
                continue
            child_ids.append(int(selection["child_id"]))
        child_ids = sorted(set(child_ids))
        if not child_ids:
            return
        if QMessageBox.question(
            self,
            "Remove EBOM Children",
            f"Remove {len(child_ids)} child usage(s) from {parent_name or parent_item_id}?\n\n"
            "The Item masters are not deleted.",
            QMessageBox.Yes | QMessageBox.Cancel,
            QMessageBox.Cancel,
        ) != QMessageBox.Yes:
            return
        try:
            result = self.bom_service.remove_children_from_parent(
                int(parent_item_id), child_ids
            )
        except Exception as exc:
            QMessageBox.warning(self, "Remove EBOM Children", str(exc))
            return
        affected = {int(parent_item_id), *child_ids}
        for item_id in sorted(affected):
            self._refresh_pdm_ebom_structure_branch(int(item_id))
        moved_items = result.get("moved_to_root_items") or []
        if moved_items:
            QMessageBox.information(
                self,
                "Moved to Top Level",
                "Some removed children have no remaining parent, so they are now top-level EBOM Items.",
            )

    def _cad_structure_selection_for_item(self, item: QTreeWidgetItem) -> dict | None:
        if item is None or item.data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD:
            return None
        try:
            child_cad_id = int(item.data(0, PDM_CAD_DOCUMENT_ID_ROLE))
        except Exception:
            return None
        payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
        member_id = item.data(0, PDM_CAD_MEMBER_ID_ROLE)
        parent_item = item.parent()
        source_parent_id = None
        try:
            if parent_item is not None:
                source_parent_id = int(parent_item.data(0, PDM_CAD_DOCUMENT_ID_ROLE))
        except Exception:
            source_parent_id = None
        return {
            "member_id": int(member_id) if member_id is not None else None,
            "child_cad_document_id": child_cad_id,
            "source_parent_cad_id": source_parent_id,
            "quantity": max(1, int(payload.get("quantity") or 1)),
            "build_excluded": bool(
                payload.get("member_build_excluded") or payload.get("build_excluded")
            ),
        }

    def _apply_pdm_cad_structure_operation(
        self,
        selected_items: list[QTreeWidgetItem],
        mode: str,
    ) -> None:
        selections = []
        for row in selected_items or []:
            payload = self._cad_structure_selection_for_item(row)
            if payload is not None:
                selections.append(payload)
        if not selections:
            return QMessageBox.warning(self, "CAD Structure", "Select one or more CAD Documents.")
        self._start_pdm_structure_target_selection(
            "CAD",
            str(mode).lower(),
            selections,
            "Select the target CAD assembly in the CAD Structure tree. You can search, filter, expand, or isolate first. Press Esc to cancel.",
        )

    def _finish_pdm_cad_structure_operation(
        self,
        target_parent_cad_id: int,
        selections: list[dict],
        mode: str,
    ) -> None:
        try:
            result = self.bom_service.apply_pdm_cad_member_operation(
                int(target_parent_cad_id), selections, str(mode).lower()
            )
        except Exception as exc:
            QMessageBox.warning(self, "CAD Structure", str(exc))
            return
        affected = {int(result.get("target_parent_cad_id") or target_parent_cad_id)}
        affected.update(int(value) for value in (result.get("source_parent_cad_ids") or []))
        affected.update(int(value) for value in (result.get("child_cad_document_ids") or []))
        for cad_id in sorted(affected):
            self._refresh_pdm_cad_structure_branch(int(cad_id))
        try:
            self._refresh_pdm_cad_filter()
        except Exception:
            pass

    def _start_pdm_structure_target_selection(
        self,
        scope: str,
        mode: str,
        selections: list[dict],
        message: str,
    ) -> None:
        self._exit_pdm_structure_target_selection(clear_status=False)
        self._pdm_structure_target_state = {
            "scope": str(scope or "").upper(),
            "mode": str(mode or "").lower(),
            "selections": list(selections or []),
        }
        trees = (
            (getattr(self, "_cad_tree", None),)
            if str(scope or "").upper() == "CAD"
            else (getattr(self, "_ebom_tree", None), getattr(self, "_ebom_filter_tree", None))
        )
        for tree in trees:
            if tree is None:
                continue
            try:
                tree.itemClicked.connect(self._on_pdm_structure_target_clicked)
            except Exception:
                pass
        try:
            self.setCursor(Qt.PointingHandCursor)
            self.window().statusBar().showMessage(message)
        except Exception:
            pass

    def _exit_pdm_structure_target_selection(self, clear_status: bool = True) -> None:
        for tree in (
            getattr(self, "_cad_tree", None),
            getattr(self, "_ebom_tree", None),
            getattr(self, "_ebom_filter_tree", None),
        ):
            if tree is None:
                continue
            try:
                tree.itemClicked.disconnect(self._on_pdm_structure_target_clicked)
            except Exception:
                pass
        self._pdm_structure_target_state = None
        try:
            self.setCursor(Qt.ArrowCursor)
            if clear_status:
                self.window().statusBar().showMessage("Structure operation cancelled.")
        except Exception:
            pass

    def _on_pdm_structure_target_clicked(self, item: QTreeWidgetItem, _column: int = 0) -> None:
        state = getattr(self, "_pdm_structure_target_state", None)
        if not state or item is None:
            return
        scope = str(state.get("scope") or "").upper()
        selections = list(state.get("selections") or [])
        mode = str(state.get("mode") or "copy").lower()
        if scope == "CAD":
            if item.data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD:
                self.window().statusBar().showMessage("Select an ASM CAD Document as target.")
                return
            try:
                target_cad_id = int(item.data(0, PDM_CAD_DOCUMENT_ID_ROLE))
            except Exception:
                return
            payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
            category = str(payload.get("category") or item.text(CAD_COL_CATEGORY) or "").upper()
            if category != "ASSEMBLY":
                self.window().statusBar().showMessage("The target CAD Document must be an ASM.")
                return
            if target_cad_id in {
                int(row.get("child_cad_document_id"))
                for row in selections
                if row.get("child_cad_document_id") is not None
            }:
                self.window().statusBar().showMessage("A CAD assembly cannot be selected as its own target.")
                return
            self._exit_pdm_structure_target_selection(clear_status=False)
            self._finish_pdm_cad_structure_operation(target_cad_id, selections, mode)
            return

        if self._is_folder_tree_item(item) or item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
            self.window().statusBar().showMessage("Select an EBOM assembly Item as target.")
            return
        try:
            target_item_id = int(item.data(0, Qt.UserRole))
        except Exception:
            return
        if str(item.text(BOM_COL_TYPE) or "").strip().lower() not in {"asm", "assembly"}:
            self.window().statusBar().showMessage("The target EBOM Item must be an assembly.")
            return
        if target_item_id in {
            int(row.get("child_id"))
            for row in selections
            if row.get("child_id") is not None
        }:
            self.window().statusBar().showMessage("An Item cannot be selected as its own target parent.")
            return
        self._exit_pdm_structure_target_selection(clear_status=False)
        self._finish_pdm_ebom_structure_operation(target_item_id, selections, mode)

    def _remove_pdm_cad_members_from_tree(
        self,
        selected_items: list[QTreeWidgetItem],
    ) -> None:
        member_rows = []
        for row in selected_items or []:
            member_id = row.data(0, PDM_CAD_MEMBER_ID_ROLE)
            if member_id is None:
                continue
            parent = row.parent()
            parent_id = None
            try:
                parent_id = int(parent.data(0, PDM_CAD_DOCUMENT_ID_ROLE)) if parent is not None else None
            except Exception:
                parent_id = None
            if parent_id is None:
                continue
            member_rows.append((int(member_id), parent_id, row.text(CAD_COL_FILE)))
        if not member_rows:
            return
        if QMessageBox.question(
            self,
            "Remove CAD Components",
            f"Remove {len(member_rows)} CAD occurrence(s) from their assemblies?\n\n"
            "The CAD Documents themselves are not deleted.",
            QMessageBox.Yes | QMessageBox.Cancel,
            QMessageBox.Cancel,
        ) != QMessageBox.Yes:
            return
        affected = set()
        undo_records = []
        try:
            for member_id, parent_id, _label in member_rows:
                undo_records.append(
                    self.undo_service.snapshot_cad_member_remove(
                        int(member_id), f"Remove CAD Component {_label}"
                    )
                )
                self.bom_service.remove_pdm_cad_member(int(member_id))
                affected.add(int(parent_id))
            for record in undo_records:
                self.undo_service.push(record)
        except Exception as exc:
            QMessageBox.warning(self, "Remove CAD Components", str(exc))
            return
        for cad_id in sorted(affected):
            self._refresh_pdm_cad_structure_branch(int(cad_id))
        try:
            self.window().statusBar().showMessage("CAD component(s) removed. Press Ctrl+Z to undo.", 6000)
        except Exception:
            pass

    def _selected_pdm_sibling_reorder_context(self, scope: str):
        scope = str(scope or "").upper()
        tree = getattr(self, "_cad_tree", None) if scope == "CAD" else getattr(self, "_ebom_tree", None)
        if tree is None:
            raise ValueError("The structure tree is not available.")
        selected = [
            item for item in tree.selectedItems()
            if item is not None
            and not self._is_lazy_placeholder(item)
            and not self._is_folder_tree_item(item)
        ]
        if not selected:
            current = tree.currentItem()
            selected = [current] if current is not None else []
        if not selected:
            raise ValueError("Select one or more sibling rows to reorder.")
        visual_parent = selected[0].parent()
        if visual_parent is None:
            raise ValueError("Top-level rows cannot be reordered here.")
        if any(item.parent() is not visual_parent for item in selected):
            raise ValueError("Reorder works only for rows under the same parent.")

        if scope == "CAD":
            parent_id = visual_parent.data(0, PDM_CAD_DOCUMENT_ID_ROLE)
            if parent_id is None:
                raise ValueError("The parent CAD assembly was not found.")
            selected_ids = []
            for item in selected:
                if item.data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD:
                    raise ValueError("Only CAD component rows can be reordered in CAD Structure.")
                member_id = item.data(0, PDM_CAD_MEMBER_ID_ROLE)
                if member_id is None:
                    raise ValueError("Top-level CAD Documents cannot be reordered here.")
                selected_ids.append(int(member_id))
            current_order = [
                int(value) for value in self.bom_service.ordered_pdm_cad_member_ids(int(parent_id))
            ]
        else:
            parent_id = self._pdm_ebom_relation_parent_for_item(selected[0])
            if parent_id is None:
                raise ValueError("Top-level EBOM Items cannot be reordered here.")
            selected_ids = []
            for item in selected:
                if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                    raise ValueError("Associated CAD rows are not EBOM usage rows.")
                usage_id = (item.data(0, BOM_TREE_OCCURRENCE_ROLE) or {}).get("usage_id")
                if usage_id is None:
                    raise ValueError("This EBOM row has no persisted usage to reorder.")
                selected_ids.append(int(usage_id))
            current_order = [
                int(value) for value in self.bom_service.ordered_pdm_item_usage_ids(int(parent_id))
            ]

        selected_set = set(selected_ids)
        selected_in_order = [value for value in current_order if value in selected_set]
        if not selected_in_order:
            raise ValueError("No valid selected rows to reorder.")
        return int(parent_id), current_order, selected_in_order

    def _apply_pdm_reordered_children(self, scope: str, parent_id: int, ordered_ids: list[int]) -> None:
        scope = str(scope or "").upper()
        if scope == "CAD":
            self.bom_service.reorder_pdm_cad_members(int(parent_id), ordered_ids)
            self._refresh_pdm_cad_structure_branch(int(parent_id))
            self._refresh_pdm_cad_filter()
        else:
            self.bom_service.reorder_pdm_item_usages(int(parent_id), ordered_ids)
            self._refresh_pdm_ebom_structure_branch(int(parent_id))
            self._refresh_ebom_filters()
        try:
            self.window().statusBar().showMessage(
                "CAD child order updated." if scope == "CAD" else "EBOM child order updated."
            )
        except Exception:
            pass

    def _reorder_selected_pdm_siblings(self, scope: str, mode: str) -> None:
        if not self.perm.can("manage_parts"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to reorder structure rows.")
        try:
            parent_id, current_order, selected_in_order = self._selected_pdm_sibling_reorder_context(scope)
            selected_set = set(selected_in_order)
            remaining = [value for value in current_order if value not in selected_set]
            if mode == "top":
                new_order = selected_in_order + remaining
            elif mode == "bottom":
                new_order = remaining + selected_in_order
            elif mode == "up":
                new_order = list(current_order)
                positions = [index for index, value in enumerate(new_order) if value in selected_set]
                if positions and positions[0] > 0:
                    before = new_order.pop(positions[0] - 1)
                    new_order.insert(positions[-1], before)
            elif mode == "down":
                new_order = list(current_order)
                positions = [index for index, value in enumerate(new_order) if value in selected_set]
                if positions and positions[-1] < len(new_order) - 1:
                    after = new_order.pop(positions[-1] + 1)
                    new_order.insert(positions[0], after)
            elif mode == "position":
                pos, ok = QInputDialog.getInt(
                    self,
                    "Move To Position",
                    f"New position for selected row(s), 1 to {len(current_order)}:",
                    1,
                    1,
                    max(1, len(current_order)),
                    1,
                )
                if not ok:
                    return
                insert_at = max(0, min(int(pos) - 1, len(remaining)))
                new_order = remaining[:insert_at] + selected_in_order + remaining[insert_at:]
            else:
                return
            if new_order != current_order:
                self._apply_pdm_reordered_children(scope, parent_id, new_order)
        except Exception as exc:
            QMessageBox.warning(self, "Reorder", str(exc))

    def _handle_pdm_tree_drag_reorder(
        self, selected_ids: list, target_id: int, target_parent_id: int, where: str
    ) -> None:
        if not self.perm.can("manage_parts"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to reorder structure rows.")
        tree = self.sender()
        scope = str(tree.property("pdmScope") or "").upper() if tree is not None else ""
        try:
            selected_set = {int(value) for value in selected_ids or []}
            current_order = (
                [int(value) for value in self.bom_service.ordered_pdm_cad_member_ids(int(target_parent_id))]
                if scope == "CAD" else
                [int(value) for value in self.bom_service.ordered_pdm_item_usage_ids(int(target_parent_id))]
            )
            if int(target_id) not in current_order or not selected_set.issubset(set(current_order)):
                raise ValueError("Drag reorder works only between sibling rows under the same parent.")
            if int(target_id) in selected_set:
                return
            selected_in_order = [value for value in current_order if value in selected_set]
            remaining = [value for value in current_order if value not in selected_set]
            target_index = remaining.index(int(target_id))
            if where == "below":
                target_index += 1
            new_order = remaining[:target_index] + selected_in_order + remaining[target_index:]
            if new_order != current_order:
                self._apply_pdm_reordered_children(scope, int(target_parent_id), new_order)
        except Exception as exc:
            QMessageBox.warning(self, "Reorder", str(exc))

    def _handle_pdm_folder_drag_reorder(
        self, selected_folder_ids: list, target_folder_id: int, where: str
    ) -> None:
        if not self.perm.can("manage_parts"):
            return QMessageBox.warning(
                self, "Permission",
                "You do not have permission to reorder folders.",
            )
        try:
            selected_set = {int(value) for value in (selected_folder_ids or [])}
            target_folder_id = int(target_folder_id)
            if target_folder_id in selected_set:
                return
            target = self._folder_record(target_folder_id)
            if not target:
                raise ValueError("Target folder was not found.")
            scope = str(target.get("scope") or "EBOM").upper()

            def same_location(folder: dict) -> bool:
                return (
                    str(folder.get("scope") or "EBOM").upper() == scope
                    and folder.get("parent_bom_id") == target.get("parent_bom_id")
                    and folder.get("parent_cad_document_id") == target.get("parent_cad_document_id")
                    and folder.get("parent_folder_id") == target.get("parent_folder_id")
                )

            sibling_folders = [
                dict(folder)
                for folder in self._pdm_folders(scope)
                if same_location(dict(folder))
            ]
            sibling_folders.sort(
                key=lambda row: (
                    int(row.get("sort_order") or 0),
                    str(row.get("name") or "").casefold(),
                    int(row["id"]),
                )
            )
            current_order = [int(folder["id"]) for folder in sibling_folders]
            if (
                target_folder_id not in current_order
                or not selected_set.issubset(set(current_order))
            ):
                raise ValueError("Folders can be reordered only with sibling folders.")
            selected_in_order = [value for value in current_order if value in selected_set]
            remaining = [value for value in current_order if value not in selected_set]
            target_index = remaining.index(target_folder_id)
            if str(where or "").lower() == "below":
                target_index += 1
            new_order = (
                remaining[:target_index]
                + selected_in_order
                + remaining[target_index:]
            )
            if new_order != current_order:
                self.bom_service.reorder_bom_folders(new_order)
                self._reload_pdm_folder_scope(scope)
        except Exception as exc:
            QMessageBox.warning(self, "Reorder Folder", str(exc))

    def _add_pdm_reorder_menu(self, menu: QMenu, scope: str) -> None:
        reorder_menu = menu.addMenu("Reorder")
        actions = (
            ("Move Up", "up"),
            ("Move Down", "down"),
            ("Move To Top", "top"),
            ("Move To Bottom", "bottom"),
            ("Move To Position...", "position"),
        )
        for label, mode in actions:
            action = reorder_menu.addAction(label)
            action.setEnabled(self.perm.can("manage_parts"))
            action.triggered.connect(
                lambda _checked=False, value=mode, current_scope=str(scope).upper():
                self._reorder_selected_pdm_siblings(current_scope, value)
            )

    def _change_cad_item_association(
        self, cad_id: int, item_id: int, association_id, current_type: str
    ) -> None:
        if not self._ensure_item_checked_out_for_pdm_change(int(item_id)):
            return
        association_type = self._prompt_pdm_association_type(
            current_type, "Change CAD-Item Association"
        )
        if association_type is None or association_type == str(current_type or "").upper():
            return
        try:
            self.bom_service.associate_cad_document(
                int(item_id), int(cad_id), association_type
            )
        except Exception as exc:
            QMessageBox.warning(self, "Change Association", str(exc))
            return
        self._refresh_pdm_context_rows(item_ids=[int(item_id)], cad_ids=[int(cad_id)])

    def _remove_cad_item_association(
        self, association_id: int, item_id: int, cad_label: str
    ) -> None:
        if not self._ensure_item_checked_out_for_pdm_change(int(item_id)):
            return
        if QMessageBox.question(
            self, "Remove CAD Association",
            f"Remove the association between this Item and {cad_label}?\n\n"
            "The CAD Document remains managed; its associations to other Items are unchanged.",
            QMessageBox.Yes | QMessageBox.Cancel,
            QMessageBox.Cancel,
        ) != QMessageBox.Yes:
            return
        try:
            self.bom_service.remove_cad_item_association(int(association_id))
        except Exception as exc:
            QMessageBox.warning(self, "Remove CAD Association", str(exc))
            return
        self._refresh_pdm_context_rows(item_ids=[int(item_id)])

    def _add_cad_member_from_tree(self, parent_cad_id: int) -> None:
        documents = [
            row for row in (self.bom_service.list_pdm_cad_documents() or [])
            if int(row.get("id") or 0) != int(parent_cad_id)
            and str(row.get("category") or "").upper()
            in {"ASSEMBLY", "COMPONENT"}
        ]
        if not documents:
            QMessageBox.information(self, "Add CAD Component", "No other CAD Documents are available.")
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("Add CAD Component")
        layout = QGridLayout(dialog)
        child_combo = QComboBox()
        for document in documents:
            label = str(document.get("file_name") or "")
            name = str(document.get("name") or "").strip()
            if name and name.casefold() != label.casefold():
                label = f"{label} - {name}"
            child_combo.addItem(
                label,
                int(document["id"]),
            )
        quantity = QSpinBox()
        quantity.setRange(1, 100000)
        excluded = QCheckBox("Exclude this CAD occurrence from Item Structure build")
        layout.addWidget(QLabel("CAD component"), 0, 0)
        layout.addWidget(child_combo, 0, 1)
        layout.addWidget(QLabel("Quantity"), 1, 0)
        layout.addWidget(quantity, 1, 1)
        layout.addWidget(excluded, 2, 0, 1, 2)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons, 3, 0, 1, 2)
        if dialog.exec_() != QDialog.Accepted:
            return
        try:
            self.bom_service.add_pdm_cad_member(
                int(parent_cad_id), int(child_combo.currentData()),
                quantity.value(), excluded.isChecked(),
            )
        except Exception as exc:
            QMessageBox.warning(self, "Add CAD Component", str(exc))
            return
        self._refresh_pdm_context_rows(
            cad_ids=[int(parent_cad_id), int(child_combo.currentData())],
            refresh_cad_branches=True,
        )
        self._reselect_cad_in_current_view(int(parent_cad_id))

    def _bind_existing_drawing_to_model(self, model_cad_id: int) -> None:
        drawings = [
            document
            for document in (self.bom_service.list_pdm_cad_documents() or [])
            if str(document.get("category") or "").upper() == "DRAWING"
            and document.get("drawing_owner_cad_document_id") is None
        ]
        if not drawings:
            QMessageBox.information(
                self, "Bind Drawing", "There are no unbound managed drawings."
            )
            return
        labels = [
            str(document.get("file_name") or document.get("name") or document["id"])
            for document in drawings
        ]
        selected, accepted = QInputDialog.getItem(
            self,
            "Bind Existing Drawing",
            "Drawing to bind to this PRT/ASM:",
            labels,
            0,
            False,
        )
        if not accepted:
            return
        drawing = drawings[labels.index(selected)]
        try:
            self.bom_service.bind_pdm_drawing_to_model(
                int(drawing["id"]), int(model_cad_id)
            )
        except Exception as exc:
            QMessageBox.warning(self, "Bind Drawing", str(exc))
            return
        self._refresh_pdm_context_rows(
            cad_ids=[int(model_cad_id), int(drawing["id"])],
        )
        self._reselect_cad_in_current_view(int(model_cad_id))

    def _remove_cad_member_from_tree(
        self, member_id: int, label: str, parent_cad_id: int
    ) -> None:
        if QMessageBox.question(
            self, "Remove CAD Component",
            f"Remove {label} from this CAD assembly?\n\n"
            "The CAD Document itself will remain managed.",
            QMessageBox.Yes | QMessageBox.Cancel,
            QMessageBox.Cancel,
        ) != QMessageBox.Yes:
            return
        try:
            undo_record = self.undo_service.snapshot_cad_member_remove(
                int(member_id), f"Remove CAD Component {label}"
            )
            self.bom_service.remove_pdm_cad_member(int(member_id))
            self.undo_service.push(undo_record)
        except Exception as exc:
            QMessageBox.warning(self, "Remove CAD Component", str(exc))
            return
        self._refresh_pdm_context_rows(
            cad_ids=[int(parent_cad_id)],
            refresh_cad_branches=True,
        )
        self._reselect_cad_in_current_view(int(parent_cad_id))
        try:
            self.window().statusBar().showMessage("CAD component removed. Press Ctrl+Z to undo.", 6000)
        except Exception:
            pass

    def _edit_cad_member_from_tree(self, item: QTreeWidgetItem) -> None:
        parent = item.parent()
        if parent is None:
            return
        parent_cad_id = parent.data(0, PDM_CAD_DOCUMENT_ID_ROLE)
        child_cad_id = item.data(0, PDM_CAD_DOCUMENT_ID_ROLE)
        if parent_cad_id is None or child_cad_id is None:
            return
        payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit CAD Occurrence")
        layout = QGridLayout(dialog)
        quantity = QSpinBox()
        quantity.setRange(1, 100000)
        quantity.setValue(max(1, int(payload.get("quantity") or 1)))
        excluded = QCheckBox("Exclude this occurrence from Item Structure build")
        excluded.setChecked(bool(
            payload.get("member_build_excluded") or payload.get("build_excluded")
        ))
        layout.addWidget(QLabel("CAD component"), 0, 0)
        layout.addWidget(QLabel(str(payload.get("file_name") or item.text(CAD_COL_FILE))), 0, 1)
        layout.addWidget(QLabel("Quantity"), 1, 0)
        layout.addWidget(quantity, 1, 1)
        layout.addWidget(excluded, 2, 0, 1, 2)
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons, 3, 0, 1, 2)
        if dialog.exec_() != QDialog.Accepted:
            return
        try:
            self.bom_service.add_pdm_cad_member(
                int(parent_cad_id), int(child_cad_id), quantity.value(),
                excluded.isChecked(),
            )
        except Exception as exc:
            QMessageBox.warning(self, "Edit CAD Occurrence", str(exc))
            return
        self._refresh_pdm_context_rows(
            cad_ids=[int(parent_cad_id), int(child_cad_id)],
            refresh_cad_branches=True,
        )
        self._reselect_cad_in_current_view(int(parent_cad_id))

    def _revise_pdm_cad_document(self, cad_id: int) -> None:
        try:
            result = self.bom_service.revise_pdm_cad_document(int(cad_id))
        except Exception as exc:
            QMessageBox.warning(self, "Create CAD Revision", str(exc))
            return
        QMessageBox.information(
            self, "Create CAD Revision",
            f"Created CAD revision {result.get('revision') or '-'}.1.",
        )
        self._refresh_pdm_context_rows(cad_ids=[int(cad_id)])
        self._reselect_cad_in_current_view(int(cad_id))

    def _release_pdm_cad_document(self, cad_id: int) -> None:
        if QMessageBox.question(
            self, "Release CAD Document",
            "Release this CAD Document iteration? Released CAD is immutable and must be revised before further work.",
            QMessageBox.Yes | QMessageBox.Cancel,
            QMessageBox.Cancel,
        ) != QMessageBox.Yes:
            return
        try:
            self.bom_service.release_pdm_cad_document(int(cad_id))
        except Exception as exc:
            QMessageBox.warning(self, "Release CAD Document", str(exc))
            return
        self._refresh_pdm_context_rows(cad_ids=[int(cad_id)])
        self._reselect_cad_in_current_view(int(cad_id))

    def _local_cad_workspaces(self) -> CadWorkspaceService:
        service = getattr(self, "_cad_workspace_service", None)
        if service is None:
            service = CadWorkspaceService()
            self._cad_workspace_service = service
        return service

    def _choose_local_cad_workspace(self, title="Select CAD Workspace"):
        return WorkspaceSelectionDialog.choose(
            self._local_cad_workspaces(), self, title=title
        )

    def manage_local_cad_workspaces(self) -> None:
        WorkspaceManagerDialog(self._local_cad_workspaces(), self).exec_()

    def _select_checked_rows_dialog(
        self,
        title: str,
        message: str,
        rows: list[dict],
        label_fn,
        *,
        checked_ids=None,
    ) -> list[int] | None:
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.resize(560, 420)
        layout = QVBoxLayout(dialog)
        label = QLabel(message)
        label.setWordWrap(True)
        layout.addWidget(label)
        list_widget = QListWidget()
        list_widget.setAlternatingRowColors(True)
        checked = {int(value) for value in (checked_ids or [])}
        for row in rows:
            try:
                row_id = int(row["id"])
            except Exception:
                continue
            item = QListWidgetItem(str(label_fn(row)))
            item.setData(Qt.UserRole, row_id)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(
                Qt.Checked if not checked or row_id in checked else Qt.Unchecked
            )
            list_widget.addItem(item)
        layout.addWidget(list_widget, 1)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        if dialog.exec_() != QDialog.Accepted:
            return None
        selected = []
        for index in range(list_widget.count()):
            item = list_widget.item(index)
            if item.checkState() == Qt.Checked:
                selected.append(int(item.data(Qt.UserRole)))
        return selected

    def _associated_item_rows_for_cad_checkout(self, cad_id: int, payload=None) -> list[dict]:
        payload = dict(payload or {})
        try:
            item_ids = list(
                self.bom_service.pdm_service.checkout_target_item_ids(int(cad_id))
                or []
            )
        except Exception:
            item_ids = []
        if not item_ids:
            legacy_item_id = payload.get("item_id") or payload.get("associated_item_id")
            if legacy_item_id is not None:
                item_ids = [legacy_item_id]
        rows = []
        seen = set()
        for item_id in item_ids:
            try:
                item_id = int(item_id)
            except Exception:
                continue
            if item_id in seen:
                continue
            seen.add(item_id)
            details = self.bom_service.get_part_details(item_id) or {"id": item_id}
            details.setdefault("id", item_id)
            rows.append(details)
        return rows

    def _prompt_released_item_revision_codes(self, item_rows: list[dict]) -> dict | None:
        revision_codes = {}
        for details in item_rows or []:
            try:
                item_id = int(details.get("id"))
            except Exception:
                continue
            if (
                str(
                    details.get("revision_state")
                    or details.get("lifecycle_state") or ""
                ).lower() != "released"
                or details.get("locked")
            ):
                continue
            try:
                suggested = self.bom_service.suggest_next_revision(item_id)
            except Exception:
                suggested = ""
            revision_code, accepted = QInputDialog.getText(
                self,
                "Check Out Released Item",
                (
                    f"{self._item_identity_text(details)} is Released.\n"
                    "Enter the next Item revision to create:"
                ),
                QLineEdit.Normal,
                suggested,
            )
            if not accepted or not str(revision_code or "").strip():
                return None
            revision_codes[item_id] = str(revision_code).strip()
        return revision_codes

    def _prompt_actor_user_id(
        self,
        title: str,
        prompt: str,
        *,
        default_user_id: int | None = None,
        only_for_admin: bool = True,
    ) -> int | None:
        """Return the effective user for a lifecycle operation, or None when cancelled."""
        current_user_id = getattr(self.session, "user_id", None)
        if only_for_admin and not self.perm.can("merge"):
            return int(current_user_id) if current_user_id is not None else None
        if not self.perm.can("merge"):
            return int(current_user_id) if current_user_id is not None else None
        users = self.project_service.get_users_for_project(self.session.project_id) or []
        choices = []
        choice_to_id = {}
        for user in users:
            try:
                user_id = int(user.get("id"))
            except Exception:
                continue
            name = str(user.get("username") or "").strip()
            email = str(user.get("email") or "").strip()
            if not name:
                continue
            label = f"{name} ({email})" if email else name
            choices.append(label)
            choice_to_id[label] = user_id
        if not choices:
            return int(default_user_id or current_user_id) if (default_user_id or current_user_id) is not None else None
        preferred = int(default_user_id or current_user_id or 0)
        default_label = choices[0]
        for label, user_id in choice_to_id.items():
            if preferred and int(user_id) == preferred:
                default_label = label
                break
        selected, ok = QInputDialog.getItem(
            self,
            title,
            prompt,
            choices,
            choices.index(default_label) if default_label in choices else 0,
            False,
        )
        if not ok:
            return None
        return choice_to_id.get(selected)

    def _prompt_cad_workspace_copy(self, title: str) -> tuple[bool, dict | None]:
        """Ask whether CAD files should be copied to a local workspace."""
        dialog = QMessageBox(self)
        dialog.setIcon(QMessageBox.Question)
        dialog.setWindowTitle(title)
        dialog.setText("Copy the checked-out CAD file(s) to a workspace?")
        dialog.setInformativeText(
            "Choose Yes when the physical Creo files will be edited. "
            "Choose No when you only need the CAD reservation and Item binding."
        )
        yes_button = dialog.addButton("Yes, copy to workspace", QMessageBox.AcceptRole)
        no_button = dialog.addButton("No, reserve only", QMessageBox.ActionRole)
        dialog.addButton(QMessageBox.Cancel)
        dialog.setDefaultButton(yes_button)
        dialog.exec_()
        if dialog.clickedButton() == yes_button:
            workspace = self._choose_local_cad_workspace(title)
            if not workspace:
                return False, None
            return True, workspace
        if dialog.clickedButton() == no_button:
            return False, {}
        return False, None

    def _metadata_checkout_for_cad_document(self, cad_id: int, payload=None) -> None:
        item_rows = self._associated_item_rows_for_cad_checkout(cad_id, payload)
        if not item_rows:
            QMessageBox.information(
                self,
                "CAD Metadata Checkout",
                "This CAD Document has no related EBOM Item to check out for metadata changes.",
            )
            return
        selected_ids = self._select_checked_rows_dialog(
            "CAD Metadata Checkout",
            "Select the related EBOM Item data to check out. Use this for CAD metadata or structure edits such as adding components or changing quantities. No CAD workspace copy will be created.",
            item_rows,
            lambda row: self._item_identity_text(row),
        )
        if selected_ids is None:
            return
        if not selected_ids:
            QMessageBox.information(
                self, "CAD Metadata Checkout", "Select at least one related Item."
            )
            return
        selected_rows = [
            row for row in item_rows if int(row.get("id") or 0) in set(selected_ids)
        ]
        revision_codes = self._prompt_released_item_revision_codes(selected_rows)
        if revision_codes is None:
            return
        try:
            for row in selected_rows:
                item_id = int(row["id"])
                self.bom_service.checkout_item(
                    item_id,
                    released_revision_code=revision_codes.get(item_id),
                    include_owner_cad=False,
                )
        except Exception as exc:
            QMessageBox.warning(self, "CAD Metadata Checkout", str(exc))
            return
        self._refresh_pdm_context_rows(
            item_ids=[int(row["id"]) for row in selected_rows if row.get("id") is not None],
            cad_ids=[int(cad_id)],
        )
        self._reselect_cad_in_current_view(int(cad_id))
        QMessageBox.information(
            self,
            "CAD Metadata Checkout",
            "Related Item data is checked out. CAD files remain checked in; workspace checkout was not created.",
        )

    def _checkout_pdm_cad_document(self, cad_id: int, payload: dict | None = None) -> None:
        payload = dict(payload or {})
        document = self.bom_service.pdm_service.repo.get_cad_document(int(cad_id)) or {}
        for key, value in document.items():
            payload.setdefault(key, value)
        as_user_id = self._prompt_actor_user_id(
            "Check Out CAD As",
            "Check out this CAD Document as:",
        )
        if as_user_id is None:
            return
        copy_to_workspace, workspace = self._prompt_cad_workspace_copy("CAD Checkout Workspace")
        if workspace is None:
            return
        workspace_service = self._local_cad_workspaces() if copy_to_workspace else None
        workspace_descriptor = (
            workspace_service.checkout_descriptor(workspace["id"])
            if copy_to_workspace and workspace_service and workspace else {}
        )
        needs_revision = (
            str(payload.get("lifecycle_state") or "").upper() == "RELEASED"
        )
        try:
            associated_item_ids = list(
                self.bom_service.pdm_service.checkout_target_item_ids(int(cad_id))
                or []
            )
        except Exception:
            associated_item_ids = []
        if not associated_item_ids:
            legacy_item_id = (
                payload.get("item_id") or payload.get("associated_item_id")
            )
            if legacy_item_id is not None:
                associated_item_ids = [int(legacy_item_id)]
        associated_item_ids = sorted({
            int(value) for value in associated_item_ids if value is not None
        })
        released_revision_codes = {}
        associated_item_labels = []
        for item_id in associated_item_ids:
            details = self.bom_service.get_part_details(int(item_id)) or {}
            associated_item_labels.append(
                self._item_identity_text(details or {"id": item_id})
            )
            if (
                str(
                    details.get("revision_state")
                    or details.get("lifecycle_state") or ""
                ).lower() == "released"
                and not details.get("locked")
            ):
                try:
                    suggested = self.bom_service.suggest_next_revision(
                        int(item_id)
                    )
                except Exception:
                    suggested = ""
                revision_code, accepted = QInputDialog.getText(
                    self,
                    "Check Out Related Released Item",
                    (
                        f"{self._item_identity_text(details or {'id': item_id})} is Released.\n"
                        "Enter the next Item revision to create for this shared CAD checkout:"
                    ),
                    QLineEdit.Normal,
                    suggested,
                )
                if not accepted or not str(revision_code or "").strip():
                    return
                released_revision_codes[int(item_id)] = str(
                    revision_code
                ).strip()
        message = (
            "Released CAD iterations are immutable. Create the next CAD revision and check it out?"
            if needs_revision else "Check out this CAD Document?"
        )
        if associated_item_ids:
            message += (
                f"\n\n{len(associated_item_ids)} associated Item"
                f"{'s' if len(associated_item_ids) != 1 else ''} will be checked "
                "out automatically; other CAD Documents remain checked in."
            )
            if associated_item_labels:
                message += "\n\n" + "\n".join(
                    f"- {label}" for label in associated_item_labels
                )
        if QMessageBox.question(
            self, "Check Out CAD Document", message,
            QMessageBox.Yes | QMessageBox.Cancel,
            QMessageBox.Yes,
        ) != QMessageBox.Yes:
            return
        revised = None
        try:
            if needs_revision:
                revised = self.bom_service.revise_pdm_cad_document(int(cad_id)) or {}
            result = self.bom_service.checkout_pdm_cad_document(
                int(cad_id),
                released_item_revision_codes=released_revision_codes,
                as_user_id=as_user_id,
                **workspace_descriptor,
            )
            materialized_files = (
                workspace_service.materialize_cad_document_package(
                    workspace["id"], int(cad_id), preserve_existing=True
                )
                if copy_to_workspace and workspace_service and workspace else []
            )
            materialized = materialized_files[0] if materialized_files else None
        except Exception as exc:
            try:
                current = self.bom_service.pdm_service.repo.get_cad_document(int(cad_id)) or {}
                if (
                    current.get("checked_out_by") is not None
                    and copy_to_workspace and workspace
                    and str(current.get("checkout_workspace_id") or "") == workspace["id"]
                ):
                    self.bom_service.undo_checkout_pdm_cad_document(
                        int(cad_id), "Workspace materialization failed", as_user_id=as_user_id
                    )
            except Exception:
                pass
            if revised is not None:
                self._refresh_pdm_context_rows(cad_ids=[int(cad_id)])
                self._reselect_cad_in_current_view(int(cad_id))
                QMessageBox.warning(
                    self,
                    "CAD Revision Created; Checkout Failed",
                    f"CAD revision {revised.get('revision') or '-'}.1 was created, "
                    f"but it could not be checked out:\n{exc}",
                )
            else:
                QMessageBox.warning(self, "Check Out CAD Document", str(exc))
            return
        suffix = ""
        if (result or {}).get("item_checkout_auto_created"):
            suffix = "\nThe related Item was checked out automatically."
        QMessageBox.information(
            self,
            "Check Out CAD Document",
            "CAD Document checked out." + suffix
            + (
                f"\n\nWorkspace: {workspace['name']}\nFiles:\n"
                + "\n".join(f"- {row['path']}" for row in materialized_files)
                if materialized else
                "\n\nNo workspace copy was created."
            ),
        )
        self._refresh_pdm_context_rows(
            item_ids=associated_item_ids,
            cad_ids=[int(cad_id)] + [
                int(value) for value in (result or {}).get("related_drawing_checkout_ids", [])
            ],
        )
        self._reselect_cad_in_current_view(int(cad_id))

    def _checkin_pdm_cad_document(self, cad_id: int, payload: dict | None = None) -> None:
        payload = dict(payload or {})
        actor_user_id = self._prompt_actor_user_id(
            "Check In CAD As",
            "Check in this CAD Document as:",
            default_user_id=payload.get("checked_out_by"),
        )
        if actor_user_id is None:
            return
        initial = str(getattr(self, "working_dir", "") or "")
        file_name = str(payload.get("file_name") or "")
        if initial and file_name:
            candidate = os.path.join(initial, file_name)
            initial = candidate if os.path.exists(candidate) else initial
        path, _selected_filter = QFileDialog.getOpenFileName(
            self, "Check In CAD Document", initial,
            "Creo CAD (*.asm *.asm.* *.prt *.prt.* *.drw *.drw.*);;All files (*)",
        )
        if not path:
            return
        note, accepted = QInputDialog.getText(
            self, "Check In CAD Document", "Check-in comment:"
        )
        if not accepted or not str(note or "").strip():
            if accepted:
                QMessageBox.warning(self, "Check In CAD Document", "A check-in comment is required.")
            return
        try:
            result = self.bom_service.checkin_pdm_cad_document(
                int(cad_id), path, str(note).strip(), as_user_id=actor_user_id
            )
        except Exception as exc:
            QMessageBox.warning(self, "Check In CAD Document", str(exc))
            return
        message = "A new CAD Document iteration was created."
        item_checkout = str((result or {}).get("item_checkout") or "")
        if item_checkout in {"RETAINED_WITH_CHANGES", "RETAINED_UNVERIFIED", "RETAINED_EXPLICIT"}:
            message += "\nThe related Item remains checked out because it has independent Item changes."
        elif item_checkout == "AUTO_RELEASED":
            message += "\nThe automatic Item checkout was released without creating an Item iteration."
        QMessageBox.information(self, "Check In CAD Document", message)
        affected_items = []
        try:
            affected_items = list((result or {}).get("affected_part_ids") or [])
        except Exception:
            affected_items = []
        if not affected_items:
            try:
                affected_items = list(
                    self.bom_service.pdm_service.checkout_target_item_ids(int(cad_id))
                    or []
                )
            except Exception:
                affected_items = []
        self._refresh_pdm_context_rows(item_ids=affected_items, cad_ids=[int(cad_id)])
        self._reselect_cad_in_current_view(int(cad_id))

    def _undo_pdm_cad_checkout(self, cad_id: int) -> None:
        payload = self.bom_service.pdm_service.repo.get_cad_document(int(cad_id)) or {}
        actor_user_id = self._prompt_actor_user_id(
            "Undo CAD Checkout As",
            "Undo this CAD checkout as:",
            default_user_id=payload.get("checked_out_by"),
        )
        if actor_user_id is None:
            return
        if QMessageBox.question(
            self, "Undo CAD Checkout",
            "Discard the CAD working state and undo this checkout?\n\n"
            "The associated Item checkout will stay active.",
            QMessageBox.Yes | QMessageBox.Cancel,
            QMessageBox.Cancel,
        ) != QMessageBox.Yes:
            return
        try:
            self.bom_service.undo_checkout_pdm_cad_document(
                int(cad_id), "Discarded from CAD Structure", as_user_id=actor_user_id
            )
        except Exception as exc:
            QMessageBox.warning(self, "Undo CAD Checkout", str(exc))
            return
        try:
            item_ids = self.bom_service.pdm_service.checkout_target_item_ids(int(cad_id)) or []
        except Exception:
            item_ids = []
        self._refresh_pdm_context_rows(item_ids=item_ids, cad_ids=[int(cad_id)])
        self._reselect_cad_in_current_view(int(cad_id))

    def _reselect_cad_in_current_view(self, cad_id: int) -> bool:
        """Restore the CAD inspector and lifecycle buttons after a tree reload."""
        tree = (
            self._ebom_tree
            if str(getattr(self, "_bom_mode", "cad")) == "ebom"
            else self._cad_tree
        )
        for item in self._iter_tree_items(tree):
            if (
                item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD
                and item.data(0, PDM_CAD_DOCUMENT_ID_ROLE) == int(cad_id)
            ):
                tree.setCurrentItem(item)
                item.setSelected(True)
                tree.scrollToItem(item)
                self._show_pdm_cad_selection(item)
                return True
        self.clear_details()
        return False

    def _select_item_in_ebom(self, item_id: int) -> None:
        """Navigate from a related CAD row to its owning EBOM Item context."""
        target_index = self.bom_mode_selector.findData("ebom")
        if target_index >= 0:
            self.bom_mode_selector.setCurrentIndex(target_index)
        self._materialize_pdm_tree_for_search(getattr(self, "_ebom_tree", None))
        for item in self._iter_tree_items(self._ebom_tree):
            if (
                item.data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD
                and item.data(0, Qt.UserRole) == int(item_id)
            ):
                self._ebom_tree.setCurrentItem(item)
                item.setSelected(True)
                self._ebom_tree.scrollToItem(item)
                self.display_details(int(item_id))
                return
        self.display_details(int(item_id))

    def _select_cad_in_structure(self, cad_id: int) -> None:
        self.bom_mode_selector.setCurrentIndex(self.bom_mode_selector.findData("cad"))
        self._materialize_pdm_tree_for_search(getattr(self, "_cad_tree", None))
        for item in self._iter_tree_items(self._cad_tree):
            if item.data(0, PDM_CAD_DOCUMENT_ID_ROLE) == int(cad_id):
                self._cad_tree.setCurrentItem(item)
                item.setSelected(True)
                self._cad_tree.scrollToItem(item)
                self._show_pdm_cad_selection(item)
                return

    def _show_pdm_tree_context_menu(self, position) -> None:
        tree = self.sender()
        if not isinstance(tree, QTreeWidget):
            return
        item = tree.itemAt(position)
        if not hasattr(tree, "_context_menu_selection"):
            tree._context_menu_selection = []
        if item is not None:
            cached_selection = list(tree._context_menu_selection or [])
            if item not in cached_selection:
                selected = list(tree.selectedItems())
                tree._context_menu_selection = (
                    selected if item in selected else [item]
                )
            elif len(cached_selection) > 1:
                tree.clearSelection()
                for cached_item in cached_selection:
                    if cached_item is not None:
                        cached_item.setSelected(True)
                tree.setCurrentItem(item)
        else:
            tree._context_menu_selection = []
        menu = QMenu(self)
        can_manage = self.perm.can("manage_parts")
        if item is None:
            if can_manage:
                folder_action = menu.addAction("Create Top-Level Folder")
                folder_action.triggered.connect(lambda: self.add_bom_folder())
                if tree is self._cad_tree:
                    menu.addSeparator()
                    action = menu.addAction("Register CAD Document...")
                    action.triggered.connect(lambda: self.register_cad_document())
                menu.exec_(tree.viewport().mapToGlobal(position))
            tree._context_menu_selection = []
            return
        if item not in tree.selectedItems():
            tree.clearSelection()
            item.setSelected(True)
        tree.setCurrentItem(item)
        folder_id = item.data(0, BOM_TREE_FOLDER_ROLE)
        if folder_id is not None:
            self._show_folder_context_menu(tree, item, int(folder_id))
            tree._context_menu_selection = []
            return
        kind = item.data(0, PDM_OBJECT_KIND_ROLE) or PDM_OBJECT_ITEM
        if kind == PDM_OBJECT_CAD:
            cad_id = int(item.data(0, PDM_CAD_DOCUMENT_ID_ROLE))
            item_id = item.data(0, PDM_ASSOCIATED_ITEM_ID_ROLE)
            association_id = item.data(0, PDM_ASSOCIATION_ID_ROLE)
            association_type = str(item.data(0, PDM_ASSOCIATION_TYPE_ROLE) or "")
            payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
            payload.setdefault("id", cad_id)
            payload.setdefault("item_id", item_id)
            associations = self._pdm_document_associations(payload)
            parent_item_id = None
            if item.parent() is not None:
                parent_item_id = (
                    item.parent().data(0, PDM_ASSOCIATED_ITEM_ID_ROLE)
                    or item.parent().data(0, Qt.UserRole)
                )

            def create_item_from_cad_context(data: dict, parent_value) -> None:
                try:
                    quantity = max(1, int(data.get("quantity") or 1))
                except Exception:
                    quantity = 1
                self._create_ebom_item_from_cad_document(
                    data,
                    parent_item_id=parent_value,
                    quantity=quantity,
                )

            view_action = menu.addAction("View CAD Document Information")
            view_action.triggered.connect(lambda: self._show_pdm_cad_selection(item))
            graph_action = menu.addAction("Show Relationship Graph")
            graph_action.triggered.connect(
                lambda _checked=False, value=cad_id, data=payload:
                self.show_relationship_graph(cad_id=value, payload=data)
            )
            isolate_action = menu.addAction("Isolate")
            isolate_action.triggered.connect(
                lambda _checked=False, row=item: self._isolate_pdm_tree_item(row)
            )
            if tree is self._cad_tree and can_manage:
                self._add_move_to_folder_menu(menu, item, "CAD")
                if str(payload.get("category") or "").upper() == "ASSEMBLY":
                    add_folder_action = menu.addAction("Add Folder Here")
                    add_folder_action.triggered.connect(
                        lambda _checked=False, row=item:
                        self.add_bom_folder(parent_item=row)
                    )
            if tree is self._ebom_tree:
                locate_action = menu.addAction("Open in CAD Structure")
                locate_action.triggered.connect(
                    lambda _checked=False, value=cad_id: self._select_cad_in_structure(value)
                )
            if len(associations) == 1 and associations[0].get("item_id") is not None:
                item_action = menu.addAction("Open Associated Item")
                item_action.triggered.connect(
                    lambda _checked=False, value=int(associations[0]["item_id"]):
                    self._select_item_in_ebom(value)
                )
            elif len(associations) > 1:
                item_menu = menu.addMenu(
                    f"Open Associated Item ({len(associations)})"
                )
                for association in associations:
                    target_item_id = association.get("item_id")
                    if target_item_id is None:
                        continue
                    item_action = item_menu.addAction(
                        self._pdm_item_association_label(association)
                    )
                    item_action.triggered.connect(
                        lambda _checked=False, value=int(target_item_id):
                        self._select_item_in_ebom(value)
                    )
            menu.addSeparator()
            checked_out = payload.get("checked_out_by") is not None
            try:
                checked_out_by_me = (
                    checked_out and self.session.user_id is not None
                    and int(payload.get("checked_out_by")) == int(self.session.user_id)
                )
            except Exception:
                checked_out_by_me = False
            cad_checkout_editable = checked_out_by_me or self.perm.can("merge")
            checkout_action = menu.addAction(
                "Revise and Check Out CAD..."
                if str(payload.get("lifecycle_state") or "").upper() == "RELEASED"
                else "Check Out CAD"
            )
            checkout_action.setEnabled(can_manage and not checked_out)
            checkout_action.triggered.connect(
                lambda _checked=False, value=cad_id, data=payload:
                self._checkout_pdm_cad_document(value, data)
            )
            can_finish_cad_checkout = checked_out_by_me or self.perm.can("merge")
            checkin_action = menu.addAction("Check In CAD...")
            checkin_action.setEnabled(can_manage and checked_out and can_finish_cad_checkout)
            checkin_action.triggered.connect(
                lambda _checked=False, value=cad_id, data=payload:
                self._checkin_pdm_cad_document(value, data)
            )
            undo_action = menu.addAction("Undo CAD Checkout")
            undo_action.setEnabled(can_manage and checked_out and can_finish_cad_checkout)
            undo_action.triggered.connect(
                lambda _checked=False, value=cad_id: self._undo_pdm_cad_checkout(value)
            )
            lifecycle_menu = menu.addMenu("CAD Lifecycle")
            revise_action = lifecycle_menu.addAction("Create New CAD Revision")
            revise_action.setEnabled(can_manage and not checked_out)
            revise_action.triggered.connect(
                lambda _checked=False, value=cad_id: self._revise_pdm_cad_document(value)
            )
            release_action = lifecycle_menu.addAction("Release CAD Iteration")
            release_action.setEnabled(
                can_manage and not checked_out
                and str(payload.get("lifecycle_state") or "").upper() != "RELEASED"
            )
            release_action.triggered.connect(
                lambda _checked=False, value=cad_id: self._release_pdm_cad_document(value)
            )
            delete_document_action = menu.addAction("Delete CAD Document")
            delete_document_action.setEnabled(can_manage and not checked_out)
            if checked_out:
                delete_document_action.setToolTip(
                    "Check in or undo checkout before deleting this CAD Document."
                )
            delete_document_action.triggered.connect(
                lambda _checked=False, value=cad_id, data=payload:
                self._delete_selected_pdm_cad_document(value, data)
            )
            menu.addSeparator()
            category = str(
                payload.get("category") or item.text(CAD_COL_CATEGORY)
            ).upper()
            if category in {"ASSEMBLY", "COMPONENT"}:
                drawing_menu = menu.addMenu("Related Drawings")
                related_drawings = list(payload.get("related_drawings") or [])
                if related_drawings:
                    for drawing in related_drawings:
                        drawing_label = (
                            f"{drawing.get('file_name') or drawing.get('name')}  "
                            f"[{drawing.get('revision') or 'A'}."
                            f"{int(drawing.get('iteration') or 1)} - "
                            f"{drawing.get('lifecycle_state') or 'IN_WORK'}]"
                        )
                        drawing_row_menu = drawing_menu.addMenu(drawing_label)
                        info_action = drawing_row_menu.addAction("Managed DRW CAD Document")
                        info_action.setEnabled(False)
                        delete_drawing_action = drawing_row_menu.addAction("Delete Drawing CAD Document")
                        delete_drawing_action.setEnabled(
                            can_manage and drawing.get("checked_out_by") is None
                        )
                        if drawing.get("checked_out_by") is not None:
                            delete_drawing_action.setToolTip(
                                "Check in or undo this drawing checkout before deleting it."
                            )
                        delete_drawing_action.triggered.connect(
                            lambda _checked=False, drawing_row=dict(drawing), owner_id=cad_id:
                            self._delete_selected_pdm_cad_document(
                                int(drawing_row["id"]),
                                drawing_row,
                                reselect_cad_id=owner_id,
                            )
                        )
                    drawing_menu.addSeparator()
                else:
                    none_action = drawing_menu.addAction("No drawing linked")
                    none_action.setEnabled(False)
                    drawing_menu.addSeparator()
                register_drawing_action = drawing_menu.addAction(
                    "Register Related Drawing..."
                )
                register_drawing_action.setEnabled(can_manage)
                register_drawing_action.triggered.connect(
                    lambda _checked=False, value=cad_id:
                    self.register_cad_document(drawing_owner_cad_id=value)
                )
                bind_drawing_action = drawing_menu.addAction(
                    "Bind Existing Drawing..."
                )
                bind_drawing_action.setEnabled(can_manage)
                bind_drawing_action.triggered.connect(
                    lambda _checked=False, value=cad_id:
                    self._bind_existing_drawing_to_model(value)
                )
                menu.addSeparator()
            association_menu = menu.addMenu(
                f"CAD-Item Associations ({len(associations)})"
            )
            for association in associations:
                target_item_id = association.get("item_id")
                target_association_id = (
                    association.get("association_id") or association.get("id")
                )
                if target_item_id is None or target_association_id is None:
                    continue
                association_row_menu = association_menu.addMenu(
                    self._pdm_item_association_label(association)
                )
                open_item_action = association_row_menu.addAction("Open Item")
                open_item_action.triggered.connect(
                    lambda _checked=False, value=int(target_item_id):
                    self._select_item_in_ebom(value)
                )
                edit_item_association = association_row_menu.addAction(
                    "Edit Type and Item Drawings..."
                )
                edit_item_association.setEnabled(can_manage and not checked_out)
                edit_item_association.triggered.connect(
                    lambda _checked=False, value=int(target_item_id), focus=cad_id:
                    self.manage_cad_item_associations(value, focus_cad_id=focus)
                )
                remove_action = association_row_menu.addAction("Remove Association")
                remove_action.setEnabled(can_manage and not checked_out)
                remove_action.triggered.connect(
                    lambda _checked=False, aid=int(target_association_id),
                           iid=int(target_item_id),
                           label=str(payload.get("file_name") or payload.get("name") or "CAD Document"):
                    self._remove_cad_item_association(aid, iid, label)
                )
            if associations:
                association_menu.addSeparator()
            if category == "DRAWING":
                drawing_item_id = item_id or parent_item_id
                drawing_model_id = payload.get("drawing_owner_cad_document_id")
                edit_drawing_assignment = association_menu.addAction(
                    "Edit Item Drawing Assignment..."
                )
                edit_drawing_assignment.setEnabled(
                    can_manage and not checked_out
                    and drawing_item_id is not None and drawing_model_id is not None
                )
                if drawing_item_id is not None and drawing_model_id is not None:
                    edit_drawing_assignment.triggered.connect(
                        lambda _checked=False, value=int(drawing_item_id),
                               focus=int(drawing_model_id):
                        self.manage_cad_item_associations(value, focus_cad_id=focus)
                    )
            else:
                associate_action = association_menu.addAction("Add Association...")
                associate_action.setEnabled(can_manage and not checked_out)
                associate_action.triggered.connect(
                    lambda _checked=False, value=cad_id:
                    self._associate_cad_to_an_item(value)
                )
            create_item_action = association_menu.addAction("Create EBOM Item from CAD...")
            create_reason = self._cad_create_ebom_item_disabled_reason(payload)
            create_item_action.setEnabled(not create_reason)
            if create_reason:
                create_item_action.setToolTip(create_reason)
            create_item_action.triggered.connect(
                lambda _checked=False, data=payload, parent_value=parent_item_id:
                create_item_from_cad_context(data, parent_value)
            )
            if tree is self._cad_tree:
                selected_cad_rows = [
                    row for row in self._selected_pdm_rows(tree, item)
                    if row.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD
                ]
                if category == "ASSEMBLY":
                    menu.addSeparator()
                    add_member_action = menu.addAction("Add CAD Component...")
                    add_member_action.setEnabled(can_manage and cad_checkout_editable)
                    add_member_action.setToolTip(
                        "The parent CAD assembly must be checked out by you."
                    )
                    add_member_action.triggered.connect(
                        lambda _checked=False, value=cad_id: self._add_cad_member_from_tree(value)
                    )
                structure_menu = menu.addMenu("CAD Structure")
                copy_occurrence_action = structure_menu.addAction(
                    "Copy Selected Occurrence(s) To..."
                )
                copy_occurrence_action.setEnabled(can_manage and bool(selected_cad_rows))
                copy_occurrence_action.triggered.connect(
                    lambda _checked=False, rows=list(selected_cad_rows):
                    self._apply_pdm_cad_structure_operation(rows, "copy")
                )
                move_occurrence_action = structure_menu.addAction(
                    "Move Selected Occurrence(s) To..."
                )
                move_occurrence_action.setEnabled(can_manage and bool(selected_cad_rows))
                move_occurrence_action.triggered.connect(
                    lambda _checked=False, rows=list(selected_cad_rows):
                    self._apply_pdm_cad_structure_operation(rows, "move")
                )
                removable_cad_rows = [
                    row for row in selected_cad_rows
                    if row.data(0, PDM_CAD_MEMBER_ID_ROLE) is not None
                ]
                remove_occurrences_action = structure_menu.addAction(
                    "Remove Selected Occurrence(s)"
                )
                remove_occurrences_action.setEnabled(
                    can_manage and bool(removable_cad_rows)
                )
                remove_occurrences_action.triggered.connect(
                    lambda _checked=False, rows=list(removable_cad_rows):
                    self._remove_pdm_cad_members_from_tree(rows)
                )
                structure_menu.addSeparator()
                self._add_pdm_reorder_menu(structure_menu, "CAD")
                member_id = item.data(0, PDM_CAD_MEMBER_ID_ROLE)
                if member_id is not None:
                    parent_item = item.parent()
                    parent_payload = dict(
                        parent_item.data(0, PDM_CAD_PAYLOAD_ROLE) or {}
                    ) if parent_item is not None else {}
                    parent_cad_id = (
                        parent_item.data(0, PDM_CAD_DOCUMENT_ID_ROLE)
                        if parent_item is not None else None
                    )
                    try:
                        parent_checked_out_by_me = (
                            self.session.user_id is not None
                            and parent_cad_id is not None
                            and parent_payload.get("checked_out_by") is not None
                            and int(parent_payload["checked_out_by"]) == int(self.session.user_id)
                        )
                    except Exception:
                        parent_checked_out_by_me = False
                    parent_cad_checkout_editable = parent_checked_out_by_me or self.perm.can("merge")
                    edit_member_action = menu.addAction("Edit CAD Occurrence...")
                    edit_member_action.setEnabled(
                        can_manage and parent_item is not None and parent_cad_checkout_editable
                    )
                    edit_member_action.triggered.connect(
                        lambda _checked=False, row=item: self._edit_cad_member_from_tree(row)
                    )
                    remove_member_action = menu.addAction("Remove from CAD Assembly")
                    remove_member_action.setEnabled(
                        can_manage and parent_cad_id is not None and parent_cad_checkout_editable
                    )
                    if parent_cad_id is None:
                        remove_member_action.setToolTip(
                            "Open the parent CAD assembly context before removing this occurrence."
                        )
                    remove_member_action.triggered.connect(
                        lambda _checked=False, value=int(member_id), label=item.text(CAD_COL_FILE),
                               parent_id=int(parent_cad_id) if parent_cad_id is not None else None:
                        self._remove_cad_member_from_tree(value, label, parent_id)
                        if parent_id is not None else None
                    )
            menu.exec_(tree.viewport().mapToGlobal(position))
            tree._context_menu_selection = []
            return

        item_id = item.data(0, Qt.UserRole)
        if item_id is None:
            tree._context_menu_selection = []
            return
        item_id = int(item_id)
        view_action = menu.addAction("View Item Details")
        view_action.triggered.connect(lambda: self.display_details(item_id))
        graph_action = menu.addAction("Show Relationship Graph")
        graph_action.triggered.connect(
            lambda _checked=False, value=item_id:
            self.show_relationship_graph(item_id=value)
        )
        isolate_action = menu.addAction("Isolate")
        isolate_action.triggered.connect(
            lambda _checked=False, row=item: self._isolate_pdm_tree_item(row)
        )
        if can_manage:
            self._add_move_to_folder_menu(menu, item, "EBOM")
            if str(item.text(BOM_COL_TYPE) or "").strip().lower() in {
                "asm", "assembly"
            }:
                add_folder_action = menu.addAction("Add Folder Here")
                add_folder_action.triggered.connect(
                    lambda _checked=False, row=item:
                    self.add_bom_folder(parent_item=row)
                )
        compare_top_action = menu.addAction("Compare with OWNER CAD...")
        compare_top_action.triggered.connect(
            lambda _checked=False, value=item_id: (
                setattr(self, "current_part_id", value),
                self.compare_cad_to_item_structure(),
            )
        )
        menu.addSeparator()
        details = self.bom_service.get_part_details(item_id) or {}
        locked = bool(details.get("locked"))
        try:
            active_cad = self.bom_service.checked_out_cad_for_item(item_id) or []
        except Exception:
            active_cad = []
        checkout_origin = str(details.get("checkout_origin") or "ITEM").upper()
        lock_owner_id = details.get("locked_by_user_id")
        lock_is_mine = (
            lock_owner_id is None or self.session.user_id is None
            or int(lock_owner_id) == int(self.session.user_id)
        )
        can_finish_checkout = lock_is_mine or self.perm.can("merge")
        checkout_action = menu.addAction("Check Out Item")
        if locked and checkout_origin == "CAD":
            checkout_action.setText("Make Item Checkout Explicit")
        checkout_action.setEnabled(
            not locked or (checkout_origin == "CAD" and can_finish_checkout)
        )
        checkout_action.triggered.connect(lambda: self.checkout_part(item_id))
        checkin_action = menu.addAction("Check In Item...")
        checkin_action.setEnabled(locked and can_finish_checkout)
        if active_cad:
            checkin_action.setToolTip("Check in Item data only; associated CAD Documents stay checked out.")
        checkin_action.triggered.connect(lambda: self.checkin_part(item_id))
        undo_action = menu.addAction("Undo Item Checkout")
        undo_action.setEnabled(locked and can_finish_checkout)
        if active_cad:
            undo_action.setToolTip("Undo Item checkout and automatically undo associated CAD checkouts.")
        undo_action.triggered.connect(lambda: self.undo_checkout(item_id))
        menu.addSeparator()
        cad_menu = menu.addMenu("Associated CAD Documents")
        associate_action = cad_menu.addAction("Associate Existing CAD...")
        associate_action.setEnabled(can_manage)
        associate_action.triggered.connect(
            lambda _checked=False, value=item_id: self.manage_cad_item_associations(value)
        )
        register_action = cad_menu.addAction("Register and Associate CAD...")
        register_action.setEnabled(can_manage)
        register_action.triggered.connect(
            lambda _checked=False, value=item_id: self._register_and_associate_cad(value)
        )
        structure_menu = menu.addMenu("Item Structure")
        selected_ebom_rows = [
            row for row in self._selected_pdm_rows(tree, item)
            if row.data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD
        ]
        add_usage_action = structure_menu.addAction("Add Manual Item Usage...")
        add_usage_action.setEnabled(can_manage)
        add_usage_action.triggered.connect(
            lambda _checked=False, value=item_id: (
                setattr(self, "current_part_id", value), self.add_manual_item_usage()
            )
        )
        copy_usage_action = structure_menu.addAction("Copy Selected Usage(s) To...")
        copy_usage_action.setEnabled(can_manage and bool(selected_ebom_rows))
        copy_usage_action.triggered.connect(
            lambda _checked=False, rows=list(selected_ebom_rows):
            self._apply_pdm_ebom_structure_operation(rows, "copy")
        )
        move_usage_action = structure_menu.addAction("Move Selected Usage(s) To...")
        move_usage_action.setEnabled(can_manage and bool(selected_ebom_rows))
        move_usage_action.triggered.connect(
            lambda _checked=False, rows=list(selected_ebom_rows):
            self._apply_pdm_ebom_structure_operation(rows, "move")
        )
        try:
            relation_parent_id = self._pdm_ebom_relation_parent_for_item(item)
        except Exception:
            relation_parent_id = None
        if relation_parent_id is not None:
            same_parent_rows = [
                row for row in selected_ebom_rows
                if self._pdm_ebom_relation_parent_for_item(row) == int(relation_parent_id)
            ]
            remove_usage_action = structure_menu.addAction(
                "Remove Selected Usage(s)" if len(same_parent_rows) > 1 else "Remove Usage"
            )
            remove_usage_action.setEnabled(can_manage and bool(same_parent_rows))
            remove_usage_action.triggered.connect(
                lambda _checked=False, parent_id=int(relation_parent_id),
                       parent_name=str(item.parent().text(BOM_COL_NAME) if item.parent() is not None else ""),
                       rows=list(same_parent_rows):
                self._remove_pdm_ebom_children_from_tree(parent_id, parent_name, rows)
            )
        structure_menu.addSeparator()
        self._add_pdm_reorder_menu(structure_menu, "EBOM")
        compare_action = structure_menu.addAction("Compare with OWNER CAD...")
        compare_action.triggered.connect(
            lambda _checked=False, value=item_id: (
                setattr(self, "current_part_id", value), self.compare_cad_to_item_structure()
            )
        )
        build_action = structure_menu.addAction("Build Item Structure from OWNER CAD")
        build_action.setEnabled(can_manage)
        build_action.triggered.connect(
            lambda _checked=False, value=item_id: (
                setattr(self, "current_part_id", value), self.build_item_structure_from_cad()
            )
        )
        files_menu = menu.addMenu("Item Delivery Files")
        preview_pdf = files_menu.addAction("Preview PDF")
        preview_pdf.triggered.connect(lambda: self.preview_part_pdf(item_id))
        open_pdf = files_menu.addAction("Open PDF")
        open_pdf.triggered.connect(lambda: self.open_part_pdf(item_id))
        open_step = files_menu.addAction("Open STEP")
        open_step.triggered.connect(lambda: self.open_part_step(item_id))
        issues_action = menu.addAction("Open Item Issues")
        issues_action.triggered.connect(
            lambda _checked=False, value=item_id: self.issue_requested.emit(value)
        )
        menu.exec_(tree.viewport().mapToGlobal(position))
        tree._context_menu_selection = []

    def show_tree_context_menu(self, position):
        tree = self.sender()
        if not isinstance(tree, QTreeWidget):
            tree = self.tree
        item = tree.itemAt(position)
        if not item:
            if tree is self.tree and self.perm.can("manage_parts"):
                menu = QMenu(self)
                create_folder_action = menu.addAction("Create Top-Level Folder")
                create_folder_action.triggered.connect(lambda: self.add_bom_folder())
                menu.exec_(tree.viewport().mapToGlobal(position))
            return
        try:
            if item not in tree.selectedItems():
                tree.clearSelection()
                item.setSelected(True)
                tree.setCurrentItem(item)
        except Exception:
            pass

        folder_id = item.data(0, BOM_TREE_FOLDER_ROLE)
        if folder_id:
            self._show_folder_context_menu(tree, item, int(folder_id))
            return

        item_id = item.data(0, Qt.UserRole)
        menu = QMenu()

        open_pdf_act = QAction("Open PDF", self)
        open_pdf_act.triggered.connect(lambda _=False, pid=item_id: self.open_part_pdf(int(pid)))
        preview_pdf_act = QAction("Preview PDF", self)
        preview_pdf_act.triggered.connect(lambda _=False, pid=item_id: self.preview_part_pdf(int(pid)))
        open_step_act = QAction("Open STEP", self)
        open_step_act.triggered.connect(lambda _=False, pid=item_id: self.open_part_step(int(pid)))
        view_action = QAction("View Item Details", self)
        view_action.triggered.connect(lambda: self.display_details(item_id))
        current_version = str(item.text(BOM_COL_REV) or "").strip()
        view_cad_files_action = QAction(
            f"View Latest CAD Files ({current_version})" if current_version else "View Latest CAD Files",
            self,
        )
        view_cad_files_action.triggered.connect(
            lambda _=False, pid=item_id: self._show_latest_bom_cad_files(int(pid))
        )
        compare_iterations_action = QAction("Compare Assembly Iterations...", self)
        compare_iterations_action.triggered.connect(
            lambda _=False, pid=item_id: self.compare_assembly_iterations(int(pid))
        )
        create_configuration_action = QAction("Create Configuration...", self)
        create_configuration_action.triggered.connect(
            lambda _=False, pid=item_id: self.create_assembly_configuration(int(pid))
        )
        compare_action = QAction("Compare to Part Structure", self)
        compare_action.triggered.connect(lambda _=False, pid=item_id: self.compare_part_structure(int(pid)))
        refresh_files_act = QAction("Refresh Files", self)
        refresh_files_act.triggered.connect(lambda _=False, pid=item_id: self._refresh_part_in_tree(int(pid)))

        menu.addAction(preview_pdf_act)
        menu.addAction(open_pdf_act)
        menu.addAction(open_step_act)
        menu.addAction(view_action)
        menu.addAction(view_cad_files_action)
        if str(item.text(BOM_COL_TYPE) or "").strip().lower() in {"asm", "assembly"}:
            menu.addAction(compare_iterations_action)
            menu.addAction(create_configuration_action)
        menu.addAction(compare_action)
        menu.addSeparator()
        menu.addAction(refresh_files_act)
        if (
            tree is self.tree
            and item.parent() is not None
            and not self._is_folder_tree_item(item.parent())
        ):
            reorder_menu = menu.addMenu("Reorder")
            move_up_act = QAction("Move Up", self)
            move_down_act = QAction("Move Down", self)
            move_top_act = QAction("Move To Top", self)
            move_bottom_act = QAction("Move To Bottom", self)
            move_position_act = QAction("Move To Position...", self)
            move_up_act.triggered.connect(lambda _=False: self._reorder_selected_siblings("up"))
            move_down_act.triggered.connect(lambda _=False: self._reorder_selected_siblings("down"))
            move_top_act.triggered.connect(lambda _=False: self._reorder_selected_siblings("top"))
            move_bottom_act.triggered.connect(lambda _=False: self._reorder_selected_siblings("bottom"))
            move_position_act.triggered.connect(lambda _=False: self._reorder_selected_siblings("position"))
            for action in (move_up_act, move_down_act, move_top_act, move_bottom_act, move_position_act):
                action.setEnabled(self.perm.can("manage_parts"))
                reorder_menu.addAction(action)
        menu.addSeparator()
        edit_action = QAction("Edit Attributes", self)
        edit_action.triggered.connect(lambda: self.edit_part(item_id))
        delete_action = QAction("Delete Item", self)
        delete_action.triggered.connect(lambda: self.delete_part(item_id))
        add_child_action = QAction("Add Child", self)
        add_child_action.triggered.connect(
            lambda _checked=False, parent_item=item: self.add_child(parent_item)
        )
        selected_relation_items = [
            selected for selected in tree.selectedItems()
            if selected is not None
            and not self._is_folder_tree_item(selected)
            and not self._is_lazy_placeholder(selected)
            and selected.data(0, Qt.UserRole) is not None
        ]
        if item not in selected_relation_items:
            selected_relation_items = [item]
        add_to_parent_action = QAction("Add to Parent", self)
        add_to_parent_action.triggered.connect(
            lambda _checked=False, selected=list(selected_relation_items): self.add_selected_to_parent(selected)
        )
        add_folder_action = QAction("Add Folder Here", self)
        add_folder_action.triggered.connect(lambda: self.add_bom_folder(parent_item=item))
        add_dwg_action = QAction("Associate Drawing", self)
        add_dwg_action.triggered.connect(lambda: self.add_dwg_to_part(item_id))

        view_3d_action = QAction("🔬 View STEP in 3D Viewer", self)
        view_3d_action.triggered.connect(lambda: self._open_step_in_3d_viewer(item_id))

        policy = item.data(0, BOM_TREE_POLICY_ROLE) or {}
        is_cad_representation = bool(policy.get("represented_part_id"))

        if self.perm.can("manage_parts"):
            menu.addAction(edit_action)
            menu.addAction(delete_action)
            if str(item.text(BOM_COL_TYPE) or "").strip().lower() in {"asm", "assembly"}:
                menu.addAction(add_child_action)
                menu.addAction(add_folder_action)
            if selected_relation_items:
                menu.addAction(add_to_parent_action)
        if not is_cad_representation:
            menu.addAction(add_dwg_action)

        occurrence = item.data(0, BOM_TREE_OCCURRENCE_ROLE) or {}
        if (
            getattr(self, "_bom_mode", "cad") == "cad"
            and occurrence.get("usage_id") is not None
            and occurrence.get("parent_id") is not None
            and self.perm.can("manage_parts")
        ):
            edit_ebom_behavior_action = QAction("Edit EBOM Behavior", self)
            edit_ebom_behavior_action.triggered.connect(
                lambda _checked=False, parent_id=int(occurrence["parent_id"]),
                       usage_id=int(occurrence["usage_id"]), row_item=item:
                    self._edit_occurrence_ebom_behavior(
                        parent_id, usage_id, row_item
                    )
            )
            menu.addAction(edit_ebom_behavior_action)

        try:
            summary = self._indicator_summary_for_part(int(item_id), self._issues_for_part(int(item_id)))
            ack_actions = []
            for doc_key, doc_type in (("pdf", "PDF"), ("step", "STEP")):
                doc = summary.get(doc_key) or {}
                tip = str(doc.get("tooltip") or "").lower()
                if doc.get("state") == "bad" and ("newer commit" in tip or "needs review" in tip):
                    action = QAction(f"Acknowledge {doc_type} as safe", self)

                    def _ack_doc(_checked=False, _doc_type=doc_type, _part_id=item_id):
                        self.part_doc_ack_service.mark_up_to_date(
                            int(_part_id),
                            _doc_type,
                            self._doc_ack_target(int(_part_id), _doc_type),
                        )
                        self._refresh_part_in_tree(int(_part_id))
                        try:
                            if getattr(self, "current_part_id", None) and int(self.current_part_id) == int(_part_id):
                                self.display_details(int(_part_id))
                        except Exception:
                            pass

                    action.triggered.connect(_ack_doc)
                    ack_actions.append(action)

            if ack_actions:
                menu.addSeparator()
                for action in ack_actions:
                    menu.addAction(action)
        except Exception:
            pass

        # Allow removing a child relation when right-clicking a child node
        try:
            parent_item = item.parent()
            relation_parent_id = parent_item.data(0, Qt.UserRole) if parent_item is not None else None
            relation_parent_name = parent_item.text(BOM_COL_NAME) if parent_item is not None else ""
            if self._is_folder_tree_item(parent_item):
                parent_folder = self._folder_record(int(parent_item.data(0, BOM_TREE_FOLDER_ROLE)))
                relation_parent_id = parent_folder.get("effective_parent_bom_id")
                if relation_parent_id is not None:
                    relation_parent = self.bom_service.get_part_details(int(relation_parent_id)) or {}
                    relation_parent_name = str(relation_parent.get("name") or relation_parent_id)
            if relation_parent_id is not None and self.perm.can("manage_parts"):
                same_parent_items = [
                    selected for selected in selected_relation_items
                    if self._direct_relation_parent_for_item(selected) == int(relation_parent_id)
                ]
                if len(same_parent_items) == len(selected_relation_items):
                    remove_child_action = QAction(
                        "Remove Selected Children" if len(same_parent_items) > 1 else "Remove Child",
                        self,
                    )
                    remove_child_action.triggered.connect(
                        lambda _checked=False, parent_id=int(relation_parent_id),
                               parent_name=str(relation_parent_name), selected=list(same_parent_items):
                            self.remove_selected_children(parent_id, parent_name, selected)
                    )
                    menu.addAction(remove_child_action)
        except Exception:
            pass

        # Only show 3D viewer action if part has a committed STEP file
        try:
            latest_step = self.commit_repo.get_latest_step_commit_for_part(
                int(item_id), int(self.session.project_id)
            )
            if (
                not is_cad_representation
                and latest_step
                and latest_step.step_file_path
                and os.path.exists(latest_step.step_file_path)
            ):
                menu.addSeparator()
                menu.addAction(view_3d_action)
        except Exception:
            pass

        menu.exec_(tree.viewport().mapToGlobal(position))

    def _edit_occurrence_ebom_behavior(
        self, parent_id: int, usage_id: int, item: QTreeWidgetItem
    ) -> None:
        policy = item.data(0, BOM_TREE_POLICY_ROLE) or {}
        current = str(policy.get("ebom_behavior") or "INHERIT").upper()
        choices = [
            ("INHERIT", "INHERIT — use the child object's default"),
            ("NORMAL", "NORMAL — deliver this occurrence"),
            ("FLATTEN", "FLATTEN — hide this occurrence, promote its children"),
            ("EXCLUDE", "EXCLUDE — not for delivery in this parent"),
        ]
        current_index = next(
            (index for index, value in enumerate(choices) if value[0] == current),
            0,
        )
        selected_label, ok = QInputDialog.getItem(
            self,
            "Edit EBOM Behavior",
            "Occurrence delivery behavior:",
            [value[1] for value in choices],
            current_index,
            False,
        )
        if not ok:
            return
        selected = next(
            value for value, label in choices if label == selected_label
        )
        try:
            relation = self.bom_service.set_occurrence_ebom_behavior(
                int(parent_id), int(usage_id), str(selected)
            )
            updated = dict(policy)
            updated["ebom_behavior"] = str(relation["ebom_behavior"])
            updated["resolved_ebom_behavior"] = (
                str(updated.get("default_ebom_behavior") or "NORMAL")
                if updated["ebom_behavior"] == "INHERIT"
                else updated["ebom_behavior"]
            )
            item.setData(0, BOM_TREE_POLICY_ROLE, updated)
            item.treeWidget().viewport().update()
            QMessageBox.information(
                self,
                "EBOM Behavior",
                "The checked-out parent structure was updated. The rule will be "
                "frozen in its next iteration on check-in.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Edit EBOM Behavior", str(exc))

    def _default_bom_advanced_filters(self) -> dict:
        return {
            "text": "",
            "text_match_mode": "normal",
            "work_state": "Any",
            "work_owner": "All",
            "status": "All",
            "type": "All",
            "revision": "",
            "categories": [],
            "structure": "Any",
            "pdf": "Any",
            "step": "Any",
            "integrity": "Any",
            "issues": "Any",
            "remove_duplicates": False,
            "show_parent_matches": True,
            "expand_matches": True,
        }

    def _is_default_bom_advanced_filter(self, filters: dict | None = None) -> bool:
        filters = filters or getattr(self, "_bom_advanced_filters", {}) or {}
        defaults = self._default_bom_advanced_filters()
        legacy_category = str(filters.get("category") or "").strip()
        if legacy_category not in ("", "All"):
            return False
        return all(
            (key == "text_match_mode" and not str(filters.get("text") or "").strip())
            or filters.get(key, default) == default
            for key, default in defaults.items()
        )

    def _normalize_bom_filter_definition(self, definition: dict | None) -> dict:
        defaults = self._default_bom_advanced_filters()
        source = definition if isinstance(definition, dict) else {}
        normalized = dict(defaults)
        for key in defaults:
            if key in source:
                normalized[key] = source[key]
        categories = normalized.get("categories") or []
        if isinstance(categories, str):
            categories = [categories]
        normalized["categories"] = [
            str(value).strip() for value in categories
            if str(value).strip()
        ]
        legacy_category = str(source.get("category") or "").strip()
        if not normalized["categories"] and legacy_category not in ("", "All"):
            normalized["categories"] = [legacy_category]
        normalized["text"] = ", ".join(split_bom_filter_terms(normalized.get("text")))
        match_mode_aliases = {
            "normal": "normal",
            "normal filter": "normal",
            "whole_word": "whole_word",
            "match whole word": "whole_word",
        }
        normalized["text_match_mode"] = match_mode_aliases.get(
            str(normalized.get("text_match_mode") or "").strip().casefold(),
            defaults["text_match_mode"],
        )
        for key in ("remove_duplicates", "show_parent_matches", "expand_matches"):
            normalized[key] = bool(normalized.get(key, defaults[key]))
        return normalized

    def _prompt_saved_filter_details(self, title: str, name: str = "", shared: bool = False):
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.setMinimumWidth(380)
        layout = QVBoxLayout(dialog)
        form = QGridLayout()
        name_input = QLineEdit(str(name or ""))
        name_input.setPlaceholderText("Filter name")
        shared_check = QCheckBox("Share with everyone in this project")
        shared_check.setChecked(bool(shared))
        form.addWidget(QLabel("Name"), 0, 0)
        form.addWidget(name_input, 0, 1)
        form.addWidget(shared_check, 1, 1)
        layout.addLayout(form)
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        name_input.selectAll()
        name_input.setFocus()
        if dialog.exec_() != QDialog.Accepted:
            return None
        clean_name = " ".join(name_input.text().split())
        if not clean_name:
            QMessageBox.warning(self, "Saved Filter", "Enter a name for the filter.")
            return None
        return clean_name, shared_check.isChecked()

    def _save_bom_filter_definition(self, definition: dict, suggested_name: str = ""):
        normalized = self._normalize_bom_filter_definition(definition)
        if self._is_default_bom_advanced_filter(normalized):
            QMessageBox.information(self, "Saved Filter", "Set at least one filter criterion before saving.")
            return None
        details = self._prompt_saved_filter_details("Save BOM Filter", suggested_name)
        if not details:
            return None
        try:
            return self.bom_service.create_saved_bom_filter(details[0], normalized, details[1])
        except Exception as exc:
            QMessageBox.warning(self, "Saved Filter", str(exc))
            return None

    def save_current_bom_filter(self):
        saved = self._save_bom_filter_definition(self._bom_advanced_filters)
        if saved:
            self._active_saved_filter_id = int(saved["id"])
            self._active_saved_filter_name = str(saved.get("name") or "")
            self._update_advanced_filter_button_state()
        return saved

    def apply_saved_bom_filter(self, filter_id: int):
        try:
            saved = self.bom_service.get_saved_bom_filter(int(filter_id))
            definition = self._normalize_bom_filter_definition(saved.get("definition"))
            active_dialog = getattr(self, "_advanced_filter_dialog", None)
            if active_dialog is not None:
                active_dialog.close()
            self.apply_bom_tree_filter(definition)
            if self._is_default_bom_advanced_filter():
                return
            self._active_saved_filter_id = int(saved["id"])
            self._active_saved_filter_name = str(saved.get("name") or "")
            self._update_advanced_filter_button_state()
        except Exception as exc:
            QMessageBox.warning(self, "Saved Filter", str(exc))

    def show_saved_bom_filters_menu(self):
        menu = QMenu(self)
        try:
            saved_filters = self.bom_service.list_saved_bom_filters()
        except Exception as exc:
            QMessageBox.warning(self, "Saved Filters", str(exc))
            return
        if saved_filters:
            for saved in saved_filters:
                owner_id = int(saved.get("owner_user_id") or 0)
                own = owner_id == int(self.session.user_id or 0)
                suffix = ""
                if saved.get("is_shared"):
                    suffix = " [Shared]" if own else f" [Shared by {saved.get('owner_name') or 'user'}]"
                action = menu.addAction(f"{saved.get('name') or 'Unnamed'}{suffix}")
                action.setCheckable(True)
                action.setChecked(int(saved["id"]) == int(self._active_saved_filter_id or 0))
                action.triggered.connect(
                    lambda _checked=False, filter_id=int(saved["id"]): self.apply_saved_bom_filter(filter_id)
                )
        else:
            empty_action = menu.addAction("No saved filters")
            empty_action.setEnabled(False)
        menu.addSeparator()
        menu.addAction("Save Current Filter...", self.save_current_bom_filter)
        menu.addAction("Manage Saved Filters...", self.show_saved_bom_filters_manager)
        button = self.saved_filters_btn
        menu.exec_(button.mapToGlobal(button.rect().bottomLeft()))

    def show_saved_bom_filters_manager(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Saved BOM Filters")
        dialog.resize(720, 430)
        layout = QVBoxLayout(dialog)
        info = QLabel("Private filters are visible only to you. Shared filters can be applied by everyone in this project.")
        info.setWordWrap(True)
        info.setStyleSheet("color:#475569;")
        layout.addWidget(info)

        table = QTableWidget(0, 4)
        table.setHorizontalHeaderLabels(["Name", "Owner", "Visibility", "Updated"])
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        layout.addWidget(table)

        command_row = QHBoxLayout()
        apply_btn = QPushButton("Apply")
        save_btn = QPushButton("Save Current As...")
        update_btn = QPushButton("Update from Current")
        rename_btn = QPushButton("Rename")
        duplicate_btn = QPushButton("Duplicate")
        share_btn = QPushButton("Share")
        for button in (apply_btn, save_btn, update_btn, rename_btn, duplicate_btn, share_btn):
            command_row.addWidget(button)
        layout.addLayout(command_row)

        order_row = QHBoxLayout()
        up_btn = QPushButton("Move Up")
        down_btn = QPushButton("Move Down")
        delete_btn = QPushButton("Delete")
        delete_btn.setObjectName("danger")
        close_btn = QPushButton("Close")
        order_row.addWidget(up_btn)
        order_row.addWidget(down_btn)
        order_row.addStretch()
        order_row.addWidget(delete_btn)
        order_row.addWidget(close_btn)
        layout.addLayout(order_row)

        rows_by_id = {}

        def selected_filter():
            row = table.currentRow()
            if row < 0:
                return None
            item = table.item(row, 0)
            return rows_by_id.get(int(item.data(Qt.UserRole) or 0)) if item else None

        def refresh(preferred_id=None):
            nonlocal rows_by_id
            try:
                rows = self.bom_service.list_saved_bom_filters()
            except Exception as exc:
                QMessageBox.warning(dialog, "Saved Filters", str(exc))
                rows = []
            rows_by_id = {int(row["id"]): row for row in rows}
            table.setRowCount(len(rows))
            selected_row = -1
            for index, saved in enumerate(rows):
                filter_id = int(saved["id"])
                values = (
                    str(saved.get("name") or ""),
                    str(saved.get("owner_name") or ""),
                    "Shared" if saved.get("is_shared") else "Private",
                    str(saved.get("updated_at") or ""),
                )
                for column, value in enumerate(values):
                    item = QTableWidgetItem(value)
                    if column == 0:
                        item.setData(Qt.UserRole, filter_id)
                    table.setItem(index, column, item)
                if filter_id == int(preferred_id or 0):
                    selected_row = index
            if selected_row < 0 and rows:
                selected_row = 0
            if selected_row >= 0:
                table.selectRow(selected_row)
            update_actions()

        def update_actions():
            saved = selected_filter()
            own = bool(saved) and int(saved.get("owner_user_id") or 0) == int(self.session.user_id or 0)
            for button in (apply_btn, duplicate_btn):
                button.setEnabled(bool(saved))
            for button in (update_btn, rename_btn, share_btn, up_btn, down_btn, delete_btn):
                button.setEnabled(own)
            share_btn.setText("Make Private" if own and saved.get("is_shared") else "Share")

        def apply_selected():
            saved = selected_filter()
            if saved:
                dialog.accept()
                self.apply_saved_bom_filter(int(saved["id"]))

        def save_current():
            saved = self.save_current_bom_filter()
            if saved:
                refresh(saved["id"])

        def update_selected():
            saved = selected_filter()
            if not saved:
                return
            answer = QMessageBox.question(
                dialog, "Update Saved Filter",
                f"Replace the criteria in '{saved['name']}' with the current BOM filter?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
            )
            if answer != QMessageBox.Yes:
                return
            try:
                updated = self.bom_service.update_saved_bom_filter(
                    int(saved["id"]), definition=self._normalize_bom_filter_definition(self._bom_advanced_filters)
                )
                refresh(updated["id"])
            except Exception as exc:
                QMessageBox.warning(dialog, "Saved Filter", str(exc))

        def rename_selected():
            saved = selected_filter()
            if not saved:
                return
            name, ok = QInputDialog.getText(dialog, "Rename Saved Filter", "Name:", text=str(saved["name"]))
            if not ok:
                return
            try:
                updated = self.bom_service.update_saved_bom_filter(int(saved["id"]), name=name)
                if int(saved["id"]) == int(self._active_saved_filter_id or 0):
                    self._active_saved_filter_name = str(updated.get("name") or "")
                    self._update_advanced_filter_button_state()
                refresh(updated["id"])
            except Exception as exc:
                QMessageBox.warning(dialog, "Saved Filter", str(exc))

        def duplicate_selected():
            saved = selected_filter()
            if not saved:
                return
            details = self._prompt_saved_filter_details("Duplicate BOM Filter", f"Copy of {saved['name']}")
            if not details:
                return
            try:
                created = self.bom_service.duplicate_saved_bom_filter(
                    int(saved["id"]), details[0], details[1]
                )
                refresh(created["id"])
            except Exception as exc:
                QMessageBox.warning(dialog, "Saved Filter", str(exc))

        def toggle_share():
            saved = selected_filter()
            if not saved:
                return
            try:
                updated = self.bom_service.update_saved_bom_filter(
                    int(saved["id"]), is_shared=not bool(saved.get("is_shared"))
                )
                refresh(updated["id"])
            except Exception as exc:
                QMessageBox.warning(dialog, "Saved Filter", str(exc))

        def move_selected(direction):
            saved = selected_filter()
            if not saved:
                return
            try:
                self.bom_service.move_saved_bom_filter(int(saved["id"]), direction)
                refresh(saved["id"])
            except Exception as exc:
                QMessageBox.warning(dialog, "Saved Filter", str(exc))

        def delete_selected():
            saved = selected_filter()
            if not saved:
                return
            answer = QMessageBox.question(
                dialog, "Delete Saved Filter", f"Delete '{saved['name']}'?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
            )
            if answer != QMessageBox.Yes:
                return
            try:
                self.bom_service.delete_saved_bom_filter(int(saved["id"]))
                if int(saved["id"]) == int(self._active_saved_filter_id or 0):
                    self._active_saved_filter_id = None
                    self._active_saved_filter_name = ""
                    self._update_advanced_filter_button_state()
                refresh()
            except Exception as exc:
                QMessageBox.warning(dialog, "Saved Filter", str(exc))

        table.itemSelectionChanged.connect(update_actions)
        table.itemDoubleClicked.connect(lambda _item: apply_selected())
        apply_btn.clicked.connect(apply_selected)
        save_btn.clicked.connect(save_current)
        update_btn.clicked.connect(update_selected)
        rename_btn.clicked.connect(rename_selected)
        duplicate_btn.clicked.connect(duplicate_selected)
        share_btn.clicked.connect(toggle_share)
        up_btn.clicked.connect(lambda: move_selected(-1))
        down_btn.clicked.connect(lambda: move_selected(1))
        delete_btn.clicked.connect(delete_selected)
        close_btn.clicked.connect(dialog.reject)
        refresh(self._active_saved_filter_id)
        dialog.exec_()

    def _collect_tree_column_values(self, column: int) -> list[str]:
        values = set()
        # Advanced filtering belongs to the persisted EBOM view.  Mixing in
        # legacy-tree values (for example Design versus WIP, or Part versus
        # prt) presents choices that cannot match any row in the new EBOM.
        tree = self._current_tree_for_filtering()
        if tree is not None and tree is getattr(self, "_ebom_tree", None):
            cache = getattr(self, "_ebom_filter_choice_cache", {}) or {}
            if column in cache:
                return list(cache[column])

            # The EBOM payload already contains the complete lazy structure.
            # Reading these two fields from dictionaries is substantially
            # cheaper than creating every QTreeWidgetItem merely to open the
            # Advanced Filter dialog.
            payload_fields = {
                BOM_COL_STATUS: ("status", "state", "lifecycle_state"),
                BOM_COL_TYPE: ("type",),
            }.get(column, ())
            stack = list(reversed(getattr(self, "_pdm_ebom_roots", []) or []))
            while stack:
                node = stack.pop()
                if not isinstance(node, dict):
                    continue
                for field in payload_fields:
                    value = str(node.get(field) or "").strip()
                    if value:
                        values.add(value)
                        break
                stack.extend(reversed(node.get("children") or []))

            # Include any currently rendered values without forcing unloaded
            # branches to materialize.
            try:
                for item in self._iter_tree_items(tree):
                    if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                        continue
                    value = str(item.text(column) or "").strip()
                    if value:
                        values.add(value)
            except Exception:
                pass
            result = sorted(values, key=lambda s: s.lower())
            cache[column] = tuple(result)
            self._ebom_filter_choice_cache = cache
            return result
        for tree in (tree,):
            if tree is None:
                continue
            try:
                for item in self._iter_tree_items(tree):
                    if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                        continue
                    value = str(item.text(column) or "").strip()
                    if value:
                        values.add(value)
            except Exception:
                continue
        return sorted(values, key=lambda s: s.lower())

    def _current_tree_for_filtering(self) -> QTreeWidget:
        if getattr(self, "_bom_mode", "cad") == "ebom":
            return getattr(self, "_ebom_tree", self.tree)
        if getattr(self, "_bom_mode", "cad") == "cad":
            return getattr(self, "_cad_tree", self.tree)
        if getattr(self, "_in_search_mode", False):
            return getattr(self, "_search_tree", self.tree)
        return self.tree

    def _file_badge_kind(self, item: QTreeWidgetItem, doc_key: str) -> str:
        payload = item.data(BOM_COL_FILES, BOM_TREE_FILES_ROLE) or {}
        value = payload.get(doc_key)
        if isinstance(value, (tuple, list)) and value:
            return str(value[0] or "na").lower()
        return "na"

    def _bom_tree_item_matches_advanced_filter(self, item: QTreeWidgetItem, filters: dict) -> bool:
        if self._is_folder_tree_item(item):
            return False
        text = str(filters.get("text") or "").strip()
        if text:
            payload = item.data(0, PDM_NODE_PAYLOAD_ROLE) or {}
            if not isinstance(payload, dict):
                payload = {}
            haystack = " ".join([
                str(item.text(col) or "") for col in range(BOM_COL_NAME, BOM_COL_STATUS + 1)
            ] + [
                str(item.data(0, BOM_TREE_ITEM_NUMBER_ROLE) or ""),
                str(item.data(0, BOM_TREE_AES_NUMBER_ROLE) or ""),
                str(item.data(0, BOM_TREE_INWORK_ROLE) or ""),
                ", ".join(item.data(0, BOM_TREE_CATEGORY_ROLE) or []),
                str(item.toolTip(BOM_COL_NAME) or ""),
                str(item.toolTip(BOM_COL_AES) or ""),
                str(item.toolTip(BOM_COL_FILES) or ""),
                str(payload.get("part_number") or ""),
                str(payload.get("aes_number") or ""),
            ])
            whole_word = str(filters.get("text_match_mode") or "normal") == "whole_word"
            if not matches_bom_filter_text(haystack, text, whole_word=whole_word):
                return False

        locked_txt = str(item.data(0, BOM_TREE_INWORK_ROLE) or "")
        locked_l = locked_txt.lower()
        work_state = str(filters.get("work_state") or "Any")
        if work_state == "In Work" and "in work" not in locked_l:
            return False
        if work_state == "Checked In" and "in work" in locked_l:
            return False

        owner = str(filters.get("work_owner") or "All").strip()
        if owner and owner != "All" and owner.lower() not in locked_l:
            return False

        status = str(filters.get("status") or "All").strip()
        if status and status != "All" and str(item.text(BOM_COL_STATUS) or "").strip().lower() != status.lower():
            return False

        part_type = str(filters.get("type") or "All").strip()
        if part_type and part_type != "All" and str(item.text(BOM_COL_TYPE) or "").strip().lower() != part_type.lower():
            return False

        revision = str(filters.get("revision") or "").strip().lower()
        if revision and revision not in str(item.text(BOM_COL_REV) or "").strip().lower():
            return False

        selected_categories = [
            str(value).strip()
            for value in (filters.get("categories") or [])
            if str(value).strip()
        ]
        # Compatibility with a filter created before category multi-selection.
        legacy_category = str(filters.get("category") or "").strip()
        if not selected_categories and legacy_category not in ("", "All"):
            selected_categories = [legacy_category]
        item_categories = [str(value).strip() for value in (item.data(0, BOM_TREE_CATEGORY_ROLE) or [])]
        if selected_categories:
            selected_keys = {value.casefold() for value in selected_categories}
            matches_uncategorized = "uncategorized" in selected_keys and not item_categories
            matches_category = any(value.casefold() in selected_keys for value in item_categories)
            if not (matches_uncategorized or matches_category):
                return False

        structure = str(filters.get("structure") or "Any")
        is_assembly = bool(item.data(0, BOM_TREE_IS_ASSEMBLY_ROLE)) or any(
            item.child(index).data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD
            for index in range(item.childCount())
        )
        if structure == "Assemblies only" and not is_assembly:
            return False
        if structure == "Leaf parts only" and is_assembly:
            return False

        doc_map = {
            "OK": "ok",
            "Outdated": "outdated",
            "Missing": "missing",
            "Not attached": "na",
        }
        for doc_key in ("pdf", "step"):
            desired = str(filters.get(doc_key) or "Any")
            if desired != "Any" and self._file_badge_kind(item, doc_key) != doc_map.get(desired, desired.lower()):
                return False

        integrity = str(filters.get("integrity") or "Any")
        integrity_state = str((item.data(BOM_COL_INTEGRITY, BOM_TREE_INTEGRITY_ROLE) or {}).get("state") or "ok").lower()
        if integrity == "Healthy only" and integrity_state != "ok":
            return False
        if integrity == "Has integrity issues" and integrity_state != "warn":
            return False

        issues = str(filters.get("issues") or "Any")
        issue_summary = item.data(0, BOM_TREE_ISSUE_ROLE) or {}
        active_count = int(issue_summary.get("active_count") or 0)
        total_count = int(issue_summary.get("total_count") or 0)
        if issues == "Active issues" and active_count <= 0:
            return False
        if issues == "Any linked issue" and total_count <= 0:
            return False
        if issues == "No linked issues" and total_count > 0:
            return False

        return True

    def show_advanced_filter_dialog(self):
        existing = getattr(self, "_advanced_filter_dialog", None)
        if existing is not None and existing.isVisible():
            existing.raise_()
            existing.activateWindow()
            return

        dialog = QDialog(self, Qt.Tool)
        dialog.setWindowModality(Qt.NonModal)
        dialog.setModal(False)
        dialog.setAttribute(Qt.WA_DeleteOnClose, True)
        self._advanced_filter_dialog = dialog

        def forget_dialog(*_):
            if getattr(self, "_advanced_filter_dialog", None) is dialog:
                self._advanced_filter_dialog = None

        dialog.destroyed.connect(forget_dialog)
        dialog.setWindowTitle("Advanced BOM Filter")
        dialog.resize(660, 550)
        layout = QVBoxLayout(dialog)
        layout.setSpacing(10)

        intro = QLabel(
            "Filter the visible BOM by text, lifecycle state, owner, document health, issues, and integrity. "
            "Separate text values with commas to match any value."
        )
        intro.setWordWrap(True)
        intro.setStyleSheet("color:#475569;")
        layout.addWidget(intro)

        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(8)
        current = dict(getattr(self, "_bom_advanced_filters", {}) or self._default_bom_advanced_filters())

        text_input = QLineEdit(str(current.get("text") or ""))
        text_input.setPlaceholderText("Multiple values: MA00, MA01")
        text_input.setToolTip("Separate values with commas. An item is shown when it matches any value.")
        text_match_combo = QComboBox()
        text_match_combo.addItem("Normal filter", "normal")
        text_match_combo.addItem("Match whole word", "whole_word")
        current_match_mode = str(current.get("text_match_mode") or "normal")
        current_match_index = text_match_combo.findData(current_match_mode)
        text_match_combo.setCurrentIndex(max(0, current_match_index))
        text_match_combo.setToolTip(
            "Normal filter matches text inside a larger value. Match whole word requires word boundaries."
        )
        grid.addWidget(QLabel("Contains"), 0, 0)
        grid.addWidget(text_input, 0, 1, 1, 2)
        grid.addWidget(text_match_combo, 0, 3)

        work_combo = QComboBox()
        work_combo.addItems(["Any", "In Work", "Checked In"])
        work_combo.setCurrentText(str(current.get("work_state") or "Any"))
        grid.addWidget(QLabel("Work state"), 1, 0)
        grid.addWidget(work_combo, 1, 1)

        owner_combo = QComboBox()
        owner_combo.addItem("All")
        users = self.project_service.get_users_for_project(self.session.project_id) or []
        for u in users:
            label = str(u.get("username") or "").strip()
            if not label:
                continue
            owner_combo.addItem(label)
        if owner_combo.findText(str(current.get("work_owner") or "All")) >= 0:
            owner_combo.setCurrentText(str(current.get("work_owner") or "All"))
        grid.addWidget(QLabel("Owner"), 1, 2)
        grid.addWidget(owner_combo, 1, 3)

        status_combo = QComboBox()
        status_combo.addItem("All")
        status_combo.addItems(self._collect_tree_column_values(BOM_COL_STATUS))
        if status_combo.findText(str(current.get("status") or "All")) >= 0:
            status_combo.setCurrentText(str(current.get("status") or "All"))
        grid.addWidget(QLabel("Status"), 2, 0)
        grid.addWidget(status_combo, 2, 1)

        type_combo = QComboBox()
        type_combo.addItem("All")
        type_combo.addItems(self._collect_tree_column_values(BOM_COL_TYPE))
        if type_combo.findText(str(current.get("type") or "All")) >= 0:
            type_combo.setCurrentText(str(current.get("type") or "All"))
        grid.addWidget(QLabel("Type"), 2, 2)
        grid.addWidget(type_combo, 2, 3)

        revision_input = QLineEdit(str(current.get("revision") or ""))
        revision_input.setPlaceholderText("Revision contains...")
        grid.addWidget(QLabel("Revision"), 3, 0)
        grid.addWidget(revision_input, 3, 1)

        structure_combo = QComboBox()
        structure_combo.addItems(["Any", "Assemblies only", "Leaf parts only"])
        structure_combo.setCurrentText(str(current.get("structure") or "Any"))
        grid.addWidget(QLabel("Structure"), 3, 2)
        grid.addWidget(structure_combo, 3, 3)

        category_list = QListWidget()
        category_list.setMaximumHeight(92)
        category_list.setToolTip("Select one or more categories. Items matching any selection are shown.")

        category_names = ["Uncategorized"]
        try:
            category_names.extend(row["name"] for row in self.bom_service.list_categories())
        except Exception:
            pass
        current_categories = {
            str(value).casefold() for value in (current.get("categories") or [])
        }
        legacy_category = str(current.get("category") or "").strip()
        if not current_categories and legacy_category not in ("", "All"):
            current_categories.add(legacy_category.casefold())
        for category_name in category_names:
            item = QListWidgetItem(str(category_name))
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(
                Qt.Checked if str(category_name).casefold() in current_categories else Qt.Unchecked
            )
            category_list.addItem(item)
        grid.addWidget(QLabel("Category"), 4, 0)
        grid.addWidget(category_list, 4, 1, 1, 3)

        pdf_combo = QComboBox()
        pdf_combo.addItems(["Any", "OK", "Outdated", "Missing", "Not attached"])
        pdf_combo.setCurrentText(str(current.get("pdf") or "Any"))
        grid.addWidget(QLabel("PDF"), 5, 0)
        grid.addWidget(pdf_combo, 5, 1)

        step_combo = QComboBox()
        step_combo.addItems(["Any", "OK", "Outdated", "Missing", "Not attached"])
        step_combo.setCurrentText(str(current.get("step") or "Any"))
        grid.addWidget(QLabel("STEP"), 5, 2)
        grid.addWidget(step_combo, 5, 3)

        integrity_combo = QComboBox()
        integrity_combo.addItems(["Any", "Healthy only", "Has integrity issues"])
        integrity_combo.setCurrentText(str(current.get("integrity") or "Any"))
        grid.addWidget(QLabel("Integrity"), 6, 0)
        grid.addWidget(integrity_combo, 6, 1)

        issue_combo = QComboBox()
        issue_combo.addItems(["Any", "Active issues", "Any linked issue", "No linked issues"])
        issue_combo.setCurrentText(str(current.get("issues") or "Any"))
        grid.addWidget(QLabel("Issues"), 6, 2)
        grid.addWidget(issue_combo, 6, 3)

        layout.addLayout(grid)

        show_parent_check = QCheckBox("Show parent branches for matching child items")
        show_parent_check.setChecked(bool(current.get("show_parent_matches", True)))
        layout.addWidget(show_parent_check)

        expand_check = QCheckBox("Expand matching branches after applying")
        expand_check.setChecked(bool(current.get("expand_matches", True)))
        layout.addWidget(expand_check)

        remove_duplicates_check = QCheckBox("Remove duplicate items (keep first occurrence)")
        remove_duplicates_check.setChecked(bool(current.get("remove_duplicates", False)))
        remove_duplicates_check.setToolTip(
            "Show a flat result list with one row per BOM item ID, keeping its first occurrence."
        )
        layout.addWidget(remove_duplicates_check)

        def update_tree_option_state(remove_duplicates: bool):
            show_parent_check.setEnabled(not remove_duplicates)
            expand_check.setEnabled(not remove_duplicates)

        remove_duplicates_check.toggled.connect(update_tree_option_state)
        update_tree_option_state(remove_duplicates_check.isChecked())

        buttons = QDialogButtonBox()
        apply_btn = buttons.addButton("Apply", QDialogButtonBox.ApplyRole)
        save_btn = buttons.addButton("Save As...", QDialogButtonBox.ActionRole)
        clear_btn = buttons.addButton("Clear Filter", QDialogButtonBox.ResetRole)
        close_btn = buttons.addButton(QDialogButtonBox.Close)
        layout.addWidget(buttons)

        def collect_filters():
            return {
                "text": text_input.text().strip(),
                "text_match_mode": text_match_combo.currentData(),
                "work_state": work_combo.currentText(),
                "work_owner": owner_combo.currentText(),
                "status": status_combo.currentText(),
                "type": type_combo.currentText(),
                "revision": revision_input.text().strip(),
                "categories": [
                    category_list.item(row).text()
                    for row in range(category_list.count())
                    if category_list.item(row).checkState() == Qt.Checked
                ],
                "structure": structure_combo.currentText(),
                "pdf": pdf_combo.currentText(),
                "step": step_combo.currentText(),
                "integrity": integrity_combo.currentText(),
                "issues": issue_combo.currentText(),
                "remove_duplicates": remove_duplicates_check.isChecked(),
                "show_parent_matches": show_parent_check.isChecked(),
                "expand_matches": expand_check.isChecked(),
            }

        def on_apply():
            filters = collect_filters()
            self._active_saved_filter_id = None
            self._active_saved_filter_name = ""
            self.apply_bom_tree_filter(filters)

        def on_save_as():
            saved = self._save_bom_filter_definition(collect_filters())
            if not saved:
                return
            self._active_saved_filter_id = int(saved["id"])
            self._active_saved_filter_name = str(saved.get("name") or "")
            self.apply_bom_tree_filter(saved.get("definition") or {})
            self._active_saved_filter_id = int(saved["id"])
            self._active_saved_filter_name = str(saved.get("name") or "")
            self._update_advanced_filter_button_state()

        def on_clear():
            self.clear_bom_tree_filter()
            dialog.close()

        apply_btn.clicked.connect(on_apply)
        save_btn.clicked.connect(on_save_as)
        clear_btn.clicked.connect(on_clear)
        close_btn.clicked.connect(dialog.close)
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()

    def clear_bom_tree_filter(self):
        was_flat_filter = bool(getattr(self, "_advanced_filter_flat_mode", False))
        self._advanced_filter_flat_mode = False
        self._bom_advanced_filters = self._default_bom_advanced_filters()
        self._active_saved_filter_id = None
        self._active_saved_filter_name = ""
        for tree in (
            getattr(self, "tree", None),
            getattr(self, "_search_tree", None),
            getattr(self, "_ebom_tree", None),
            getattr(self, "_ebom_filter_tree", None),
        ):
            if tree is None:
                continue
            try:
                for item in self._iter_tree_items(tree):
                    item.setHidden(False)
            except Exception:
                pass
        if getattr(self, "_bom_mode", "cad") == "cad":
            self._refresh_pdm_cad_filter()
            self._tree_stack.setCurrentWidget(self._cad_tree)
        elif getattr(self, "_bom_mode", "cad") == "ebom":
            self._refresh_ebom_filters()
        elif was_flat_filter:
            try:
                if self.search_input.text().strip():
                    self._perform_search_now()
                else:
                    self._exit_search_mode()
            except Exception:
                self._exit_search_mode()
        self._refresh_bom_row_numbers()
        self._update_advanced_filter_button_state(visible_count=None)

    def _update_advanced_filter_button_state(self, visible_count: int | None = None):
        active = not self._is_default_bom_advanced_filter()
        try:
            self.clear_filter_btn.setEnabled(active)
        except Exception:
            pass
        try:
            if active:
                suffix = f" ({visible_count} shown)" if visible_count is not None else ""
                saved_name = str(getattr(self, "_active_saved_filter_name", "") or "").strip()
                label = saved_name if saved_name else "Active"
                if len(label) > 24:
                    label = label[:21] + "..."
                self.advanced_filter_btn.setText(f"Advanced Filter: {label}{suffix}")
            else:
                self.advanced_filter_btn.setText("Advanced Filter")
        except Exception:
            pass
        try:
            saved_name = str(getattr(self, "_active_saved_filter_name", "") or "").strip()
            self.saved_filters_btn.setToolTip(
                f"Applied saved filter: {saved_name}" if saved_name else "Apply or manage saved BOM filters"
            )
        except Exception:
            pass

    def _apply_bom_tree_filter_to_tree(self, tree: QTreeWidget, filters: dict) -> int:
        visible_count = 0
        show_parents = bool(filters.get("show_parent_matches", True))

        def recurse(item):
            nonlocal visible_count
            show_self = self._bom_tree_item_matches_advanced_filter(item, filters)
            child_match_visible = False
            for i in range(item.childCount()):
                child = item.child(i)
                child_show = recurse(child)
                child_match_visible = child_match_visible or child_show
            show = show_self or (show_parents and child_match_visible)
            item.setHidden(not show)
            if show_self:
                visible_count += 1
                if filters.get("expand_matches", True):
                    try:
                        item.setExpanded(True)
                    except Exception:
                        pass
            return show

        try:
            tree.setUpdatesEnabled(False)
            for i in range(tree.topLevelItemCount()):
                recurse(tree.topLevelItem(i))
        finally:
            try:
                tree.setUpdatesEnabled(True)
            except Exception:
                pass
        return visible_count

    def _clone_tree_item_shallow(self, item: QTreeWidgetItem) -> QTreeWidgetItem:
        clone = QTreeWidgetItem([str(item.text(col) or "") for col in range(BOM_TREE_COLUMN_COUNT)])
        clone.setTextAlignment(BOM_COL_ROW, Qt.AlignCenter)
        for col in range(BOM_TREE_COLUMN_COUNT):
            try:
                clone.setToolTip(col, item.toolTip(col))
            except Exception:
                pass
        for column, role in (
            (0, Qt.UserRole),
            (0, BOM_TREE_ITEM_NUMBER_ROLE),
            (0, BOM_TREE_AES_NUMBER_ROLE),
            (0, BOM_TREE_INWORK_ROLE),
            (0, BOM_TREE_IS_ASSEMBLY_ROLE),
            (0, BOM_TREE_ISSUE_ROLE),
            (0, BOM_TREE_CATEGORY_ROLE),
            (0, BOM_TREE_BINDING_UPDATE_ROLE),
            (0, BOM_TREE_POLICY_ROLE),
            (0, BOM_TREE_OCCURRENCE_ROLE),
            (0, BOM_TREE_PROMOTION_ROLE),
            (0, PDM_NODE_PAYLOAD_ROLE),
            (BOM_COL_FILES, BOM_TREE_FILES_ROLE),
            (BOM_COL_INTEGRITY, BOM_TREE_INTEGRITY_ROLE),
        ):
            try:
                clone.setData(column, role, item.data(column, role))
            except Exception:
                pass
        try:
            clone.setIcon(BOM_COL_NAME, item.icon(BOM_COL_NAME))
        except Exception:
            pass
        return clone

    def _source_tree_for_flat_bom_filter(self) -> QTreeWidget:
        query = ""
        try:
            query = self.search_input.text().strip()
        except Exception:
            query = ""
        if query and getattr(self, "_in_search_mode", False) and not getattr(self, "_advanced_filter_flat_mode", False):
            return getattr(self, "_search_tree", self.tree)
        return self.tree

    def _iter_tree_items_visual_order(self, tree_widget: QTreeWidget):
        stack = []
        try:
            for index in range(tree_widget.topLevelItemCount() - 1, -1, -1):
                stack.append(tree_widget.topLevelItem(index))
        except Exception:
            return
        while stack:
            item = stack.pop()
            yield item
            try:
                for index in range(item.childCount() - 1, -1, -1):
                    stack.append(item.child(index))
            except Exception:
                continue

    def _apply_bom_tree_filter_flat(self, filters: dict) -> int:
        source_tree = self._source_tree_for_flat_bom_filter()
        remove_duplicates = bool(filters.get("remove_duplicates", False))
        source_items = (
            self._iter_tree_items_visual_order(source_tree)
            if remove_duplicates
            else self._iter_tree_items(source_tree)
        )
        matches = [
            item for item in source_items
            if self._bom_tree_item_matches_advanced_filter(item, filters)
        ]
        if remove_duplicates:
            matches = deduplicate_bom_items_by_id(
                matches,
                lambda item: item.data(0, Qt.UserRole),
            )
        # The source may be _search_tree, which _enter_search_mode() clears. Build
        # detached rows first so deleted Qt item wrappers are never accessed.
        match_clones = [self._clone_tree_item_shallow(item) for item in matches]
        self._advanced_filter_flat_mode = True
        self._enter_search_mode()
        try:
            self._search_tree.setUpdatesEnabled(False)
            for clone in match_clones:
                clone.setHidden(False)
                self._search_tree.addTopLevelItem(clone)
        finally:
            try:
                self._search_tree.setUpdatesEnabled(True)
            except Exception:
                pass
        self._sync_search_tree_row_numbers()
        return len(match_clones)

    def apply_bom_tree_filter(self, filters: dict | None = None):
        self._bom_advanced_filters = self._normalize_bom_filter_definition(
            filters or self._bom_advanced_filters or self._default_bom_advanced_filters()
        )
        if self._is_default_bom_advanced_filter():
            self.clear_bom_tree_filter()
            return
        if getattr(self, "_bom_mode", "cad") == "ebom":
            # The advanced dialog has its own Contains field.  Do not silently
            # AND it with a basic CAD/EBOM search left in the shared search box.
            try:
                self._search_timer.stop()
                previous_signal_state = self.search_input.blockSignals(True)
                try:
                    self.search_input.clear()
                finally:
                    self.search_input.blockSignals(previous_signal_state)
                self.search_btn.setText("Search")
            except Exception:
                pass
            total_visible = self._refresh_ebom_filters()
            self._update_advanced_filter_button_state(total_visible)
            return
        # Let an incremental basic-search build finish before filtering its rows.
        # _search_build_step() reapplies the current advanced filter on completion.
        try:
            if self.search_input.text().strip() and self._search_build_timer.isActive():
                self._update_advanced_filter_button_state(visible_count=None)
                return
        except Exception:
            pass
        # Advanced predicates must inspect the full BOM. Materialize only when the
        # user explicitly filters; normal startup and browsing remain level-lazy.
        if not getattr(self, "_in_search_mode", False):
            self._materialize_all_lazy_branches()
        if (
            bool(self._bom_advanced_filters.get("remove_duplicates", False))
            or not bool(self._bom_advanced_filters.get("show_parent_matches", True))
        ):
            total_visible = self._apply_bom_tree_filter_flat(self._bom_advanced_filters)
            self._update_advanced_filter_button_state(total_visible)
            return
        if getattr(self, "_advanced_filter_flat_mode", False):
            self._advanced_filter_flat_mode = False
            try:
                if self.search_input.text().strip():
                    self._perform_search_now()
                    return
                self._exit_search_mode()
            except Exception:
                self._exit_search_mode()
        total_visible = 0
        for tree in (getattr(self, "tree", None), getattr(self, "_search_tree", None)):
            if tree is None:
                continue
            try:
                total_visible += self._apply_bom_tree_filter_to_tree(tree, self._bom_advanced_filters)
            except Exception:
                pass
        self._refresh_bom_row_numbers()
        self._update_advanced_filter_button_state(total_visible)

    # -------------------------
    # Files (Vault / PLM-like)
    # -------------------------
    def _set_files_tab_enabled(self, enabled: bool):
        # Read-only controls: follow enabled state only
        for w in (
            getattr(self, "files_table", None),
            getattr(self, "versions_table", None),
            getattr(self, "open_active_btn", None),
            getattr(self, "open_folder_btn", None),
            getattr(self, "export_package_btn", None),
            getattr(self, "export_baseline_btn", None),
            getattr(self, "open_version_btn", None),
        ):
            if w is not None:
                w.setEnabled(enabled)

        # Write controls: also require release_files permission
        can_release = self.perm.can("release_files")
        released = False
        if enabled and getattr(self, "current_part_id", None):
            try:
                details = self.bom_service.get_part_details(int(self.current_part_id)) or {}
                state = str(details.get("revision_state") or details.get("lifecycle_state") or "")
                released = state.strip().lower() == "released"
            except Exception:
                released = False
        for w in (
            getattr(self, "add_attachment_btn", None),
            getattr(self, "add_version_btn", None),
            getattr(self, "set_active_btn", None),
            getattr(self, "remove_attachment_btn", None),
            getattr(self, "delete_version_btn", None),
        ):
            if w is not None:
                w.setEnabled(bool(enabled) and can_release and not released)
        for w in (
            getattr(self, "create_baseline_btn", None),
            getattr(self, "release_version_btn", None),
        ):
            if w is not None:
                w.setEnabled(bool(enabled) and can_release)

    def _current_bom_view_item_ids_for_selector(self) -> set[int]:
        """Return EBOM Item IDs visible in the current structure browser."""
        if str(getattr(self, "_bom_mode", "cad")) != "ebom":
            return set()
        try:
            query = str(self.search_input.text() or "").strip()
            advanced_text = str(
                (self._bom_advanced_filters or {}).get("text") or ""
            ).strip()
            if query or advanced_text:
                return self._ebom_db_filter_item_ids(query=query)
        except Exception:
            pass
        tree = self._current_pdm_tree()
        if tree is None:
            return set()
        ids = set()
        try:
            for item in self._iter_tree_items(tree):
                if (
                    item.isHidden()
                    or self._is_folder_tree_item(item)
                    or self._is_lazy_placeholder(item)
                    or item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD
                ):
                    continue
                part_id = item.data(0, Qt.UserRole)
                if part_id is not None:
                    ids.add(int(part_id))
        except Exception:
            return set()
        return ids

    def _ebom_db_filter_item_ids(self, query: str = "") -> set[int]:
        if not self.session.project_id:
            return set()
        filters = self._bom_advanced_filters or self._default_bom_advanced_filters()
        where = ["project_id=?", "represented_part_id IS NULL"]
        params = [int(self.session.project_id)]
        for text in (str(query or "").strip(), str(filters.get("text") or "").strip()):
            if not text:
                continue
            like = f"%{text.lower()}%"
            where.append(
                "("
                "lower(COALESCE(part_number,'')) LIKE ? OR "
                "lower(COALESCE(aes_number,'')) LIKE ? OR "
                "lower(COALESCE(name,'')) LIKE ? OR "
                "lower(COALESCE(type,'')) LIKE ?"
                ")"
            )
            params.extend([like] * 4)
        status = str(filters.get("status") or "All").strip()
        if status and status != "All":
            where.append("lower(COALESCE(lifecycle_state,status,''))=?")
            params.append(status.lower())
        part_type = str(filters.get("type") or "All").strip()
        if part_type and part_type != "All":
            where.append("lower(COALESCE(type,''))=?")
            params.append(part_type.lower())
        revision = str(filters.get("revision") or "").strip()
        if revision:
            where.append("lower(COALESCE(revision,'')) LIKE ?")
            params.append(f"%{revision.lower()}%")
        try:
            with self.bom_service.bom_repo.get_conn() as conn:
                rows = conn.execute(
                    f"SELECT id FROM bom WHERE {' AND '.join(where)}",
                    params,
                ).fetchall()
            return {int(row["id"]) for row in rows}
        except Exception:
            return set()

    def _has_active_bom_view_filter_for_selector(self) -> bool:
        if str(getattr(self, "_bom_mode", "cad")) != "ebom":
            return False
        if bool(self._pdm_scope_path_for_mode("ebom")):
            return True
        if str(self.search_input.text() or "").strip():
            return True
        if getattr(self, "_ebom_filter_flat_mode", False):
            return True
        return not self._is_default_bom_advanced_filter()

    def _run_export_with_progress(
        self,
        *,
        title: str,
        initial_message: str,
        operation,
        on_success,
        error_title: str,
    ) -> None:
        progress = QProgressDialog(initial_message, "Cancel", 0, 0, self)
        progress.setCancelButton(None)
        progress.setWindowTitle(title)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setAutoClose(False)
        progress.setAutoReset(False)
        progress.setValue(0)
        progress.show()

        thread = QThread(self)
        worker = _ExportWorker(operation)
        worker.moveToThread(thread)

        def update_progress(message: str, value: int, maximum: int) -> None:
            if maximum > 0:
                progress.setRange(0, maximum)
                progress.setValue(max(0, min(value, maximum)))
            else:
                progress.setRange(0, 0)
            progress.setLabelText(message or "Exporting...")
            QApplication.processEvents()

        def cleanup() -> None:
            try:
                thread.quit()
            except Exception:
                pass
            try:
                progress.close()
            except Exception:
                pass

        def finished(result) -> None:
            cleanup()
            try:
                on_success(result)
            finally:
                try:
                    worker.deleteLater()
                    thread.deleteLater()
                except Exception:
                    pass

        def failed(message: str) -> None:
            cleanup()
            try:
                worker.deleteLater()
                thread.deleteLater()
            except Exception:
                pass
            QMessageBox.critical(self, error_title, message)

        thread.started.connect(worker.run)
        worker.progress.connect(update_progress)
        worker.finished.connect(finished)
        worker.failed.connect(failed)
        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)
        thread.start()
        self._active_export_thread = thread
        self._active_export_worker = worker
        self._active_export_progress = progress

    def create_baseline(self):
        if not self.perm.can("release_files"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to manage files.")
        if not getattr(self, "current_part_id", None):
            return QMessageBox.warning(self, "Select", "Select a part first.")

        name, ok = QInputDialog.getText(self, "Create Baseline", "Baseline name:")
        if not ok or not (name or "").strip():
            return

        include_children = (
            QMessageBox.question(
                self,
                "Include Children",
                "Include child parts in this baseline?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            == QMessageBox.Yes
        )

        # Let user pick parts (defaults to current part checked)
        current_filter_ids = self._current_bom_view_item_ids_for_selector()
        use_current_filter = self._has_active_bom_view_filter_for_selector()
        dlg = PackagePartsDialog(
            self,
            project_id=self.session.project_id,
            preselected_ids=[int(self.current_part_id)],
            current_filter_ids=current_filter_ids,
            use_current_filter=use_current_filter,
            title="Create Baseline - Select Items",
            subtitle="Filter the EBOM scope, select all filtered Items when needed, then freeze their active PDF/STEP versions.",
            kicker="BASELINE",
        )
        part_ids = [int(self.current_part_id)]
        if dlg.exec_() == QDialog.Accepted:
            picked = dlg.selected_part_ids()
            if picked:
                part_ids = picked

        try:
            res = self.baseline_service.create_baseline(name=name.strip(), part_ids=part_ids, include_children=include_children)
            missing = res.get("missing") or []
            QMessageBox.information(
                self,
                "Baseline Created",
                f"Baseline #{res.get('baseline_id')} created.\n"
                f"Parts: {len(res.get('expanded_part_ids') or [])}\n"
                f"Issues: {len(missing)} (missing or not released)",
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create baseline:\n{e}")

    def export_baseline(self):
        baselines = self.baseline_service.list_baselines()
        if not baselines:
            return QMessageBox.warning(self, "No Baselines", "No baselines exist for this project.")

        labels = []
        id_by_label = {}
        for b in baselines:
            label = f"{b['id']} - {b['name']} ({b.get('created_at') or ''})"
            labels.append(label)
            id_by_label[label] = int(b["id"])

        choice, ok = QInputDialog.getItem(self, "Export Baseline", "Select baseline:", labels, 0, False)
        if not ok or not choice:
            return

        baseline_id = id_by_label.get(choice)
        if not baseline_id:
            return

        dest_dir = QFileDialog.getExistingDirectory(self, "Select Destination Folder")
        if not dest_dir:
            return

        def do_export(progress_callback):
            return self.baseline_service.export_baseline(
                baseline_id,
                dest_dir,
                progress_callback=progress_callback,
            )

        def show_export_result(manifest):
            missing = manifest.get("missing") or []
            out_dir = (manifest.get("package") or {}).get("output_dir")
            QMessageBox.information(
                self,
                "Baseline Exported",
                f"Export complete.\nOutput: {out_dir}\nMissing: {len(missing)}",
            )

        self._run_export_with_progress(
            title="Export Baseline",
            initial_message="Preparing baseline export...",
            operation=do_export,
            on_success=show_export_result,
            error_title="Export Baseline Failed",
        )

    def refresh_files_tab(self):
        if not getattr(self, "current_part_id", None):
            if hasattr(self, "files_table"):
                self.files_table.setRowCount(0)
            if hasattr(self, "versions_table"):
                self.versions_table.setRowCount(0)
            if hasattr(self, "file_related_issues_table"):
                self.file_related_issues_table.setRowCount(0)
            if hasattr(self, "pdf_viewer"):
                self.pdf_viewer.close_preview()
                self.pdf_viewer.setVisible(False)
            if hasattr(self, "preview_file_btn"):
                self.preview_file_btn.blockSignals(True)
                self.preview_file_btn.setChecked(False)
                self.preview_file_btn.setText("Preview")
                self.preview_file_btn.blockSignals(False)
            self._set_files_tab_enabled(False)
            return

        self._set_files_tab_enabled(True)
        details = {}
        try:
            managed_files = self.managed_file_service.list_current_files(
                int(self.current_part_id)
            )
        except Exception as exc:
            managed_files = []
            if hasattr(self, "pdf_viewer"):
                self.pdf_viewer.show_error(f"Unable to load managed files:\n{exc}")

        try:
            details = self.bom_service.get_part_details(int(self.current_part_id)) or {}
            context = self.managed_file_service.revision_repo.get_current_context(
                int(self.current_part_id)
            )
            lock = self.managed_file_service.part_file_service.lock_repo.get_by_part(
                int(self.current_part_id)
            )
            values = {
                "item": self._item_identity_text(details),
                "version": str(context.get("version_label") or "-"),
                "state": str(context.get("state") or "-"),
                "lock": self.managed_file_service._username(lock.user_id) if lock else "Not checked out",
            }
            for key, label in getattr(self, "files_summary_labels", {}).items():
                label.setText(values.get(key, "-"))
        except Exception:
            pass

        self.files_table.setRowCount(len(managed_files))
        first_pdf_row = None
        for i, entry in enumerate(managed_files):
            role_item = QTableWidgetItem(str(entry.get("role_label") or ""))
            role_item.setData(Qt.UserRole, entry.get("file_id"))
            role_item.setData(Qt.UserRole + 1, entry)
            self.files_table.setItem(i, 0, role_item)
            values = [
                entry.get("filename"),
                entry.get("file_revision"),
                entry.get("creo_iteration"),
                entry.get("bound_to"),
                entry.get("source"),
                entry.get("state"),
                entry.get("health"),
                entry.get("created_by_label"),
                entry.get("updated"),
            ]
            for column, value in enumerate(values, start=1):
                self.files_table.setItem(i, column, QTableWidgetItem(str(value or "")))
            if first_pdf_row is None and str(entry.get("file_type") or "").upper() == "PDF":
                first_pdf_row = i

        # Clear versions view until an attachment is selected
        self.versions_table.setRowCount(0)
        if hasattr(self, "file_related_issues_table"):
            self.file_related_issues_table.setRowCount(0)
        if first_pdf_row is not None:
            self.files_table.selectRow(first_pdf_row)
            self.on_attachment_selected()
        elif hasattr(self, "pdf_viewer"):
            self._collapse_managed_preview()

    def _selected_attachment_id(self):
        items = getattr(self, "files_table", None).selectedItems() if hasattr(self, "files_table") else []
        if not items:
            return None
        file_id = self.files_table.item(self.files_table.currentRow(), 0).data(Qt.UserRole)
        return file_id

    def _selected_managed_file(self):
        if not hasattr(self, "files_table") or self.files_table.currentRow() < 0:
            return None
        item = self.files_table.item(self.files_table.currentRow(), 0)
        return item.data(Qt.UserRole + 1) if item else None

    def _selected_file_type(self):
        entry = self._selected_managed_file() or {}
        return str(entry.get("file_type") or "").strip().upper()

    def _selected_history_row(self):
        if not hasattr(self, "versions_table") or self.versions_table.currentRow() < 0:
            return None
        item = self.versions_table.item(self.versions_table.currentRow(), 0)
        return item.data(Qt.UserRole + 1) if item else None

    def _selected_version_id(self):
        items = getattr(self, "versions_table", None).selectedItems() if hasattr(self, "versions_table") else []
        if not items:
            return None
        return self.versions_table.item(self.versions_table.currentRow(), 0).data(Qt.UserRole)

    def on_attachment_selected(self):
        selected = self._selected_managed_file()
        file_id = self._selected_attachment_id()
        if not selected:
            self.versions_table.setRowCount(0)
            if hasattr(self, "file_related_issues_table"):
                self.file_related_issues_table.setRowCount(0)
            if hasattr(self, "pdf_viewer"):
                self.pdf_viewer.close_preview()
            self._collapse_managed_preview()
            return
        history = self.managed_file_service.list_file_history(
            int(self.current_part_id),
            str(selected.get("role") or "document"),
            file_id=selected.get("related_file_ids") or file_id,
        )
        self.versions_table.setRowCount(len(history))
        for i, row in enumerate(history):
            version_item = QTableWidgetItem(str(row.get("bound_to") or ""))
            version_item.setData(Qt.UserRole, row.get("version_id"))
            version_item.setData(Qt.UserRole + 1, row)
            self.versions_table.setItem(i, 0, version_item)
            values = [
                row.get("role_label"), row.get("filename"), row.get("file_revision"), row.get("creo_iteration"),
                row.get("source"), row.get("state"), row.get("health"),
                row.get("created_by_label"), row.get("created_at"), row.get("note"),
            ]
            for column, value in enumerate(values, start=1):
                self.versions_table.setItem(i, column, QTableWidgetItem(str(value or "")))

        # Auto-preview the active version if it is a PDF
        self._try_pdf_preview_for_active(file_id)
        if file_id:
            self._load_related_issues_for_file(file_id, self._selected_version_id())
        elif hasattr(self, "file_related_issues_table"):
            self.file_related_issues_table.setRowCount(0)
        self._update_managed_file_actions()

    def _update_managed_file_actions(self):
        entry = self._selected_managed_file() or {}
        is_document = bool(entry.get("file_id"))
        can_manage = bool(self.perm.can("release_files") and is_document)
        for button in (
            getattr(self, "add_version_btn", None),
            getattr(self, "set_active_btn", None),
            getattr(self, "remove_attachment_btn", None),
            getattr(self, "link_file_issue_btn", None),
            getattr(self, "release_version_btn", None),
            getattr(self, "delete_version_btn", None),
        ):
            if button:
                button.setEnabled(can_manage)
        for button in (
            getattr(self, "open_active_btn", None),
            getattr(self, "open_folder_btn", None),
            getattr(self, "open_version_btn", None),
        ):
            if button:
                button.setEnabled(bool(entry))
        if hasattr(self, "preview_file_btn"):
            is_pdf = str(entry.get("file_type") or "").upper() == "PDF"
            self.preview_file_btn.setEnabled(is_pdf)
            if not is_pdf:
                self._collapse_managed_preview()

    def _load_related_issues_for_file(self, file_id, version_id=None):
        if not hasattr(self, "file_related_issues_table"):
            return
        try:
            issues = self.traceability_service.issues_for_engineering_file(
                int(file_id),
                int(version_id) if version_id else None,
            )
        except Exception:
            issues = []
        self.file_related_issues_table.setRowCount(len(issues))
        for row, issue in enumerate(issues):
            values = [
                issue.get("issue_number"),
                issue.get("title"),
                issue.get("status"),
                issue.get("file_role"),
                issue.get("linked_at"),
            ]
            for col, value in enumerate(values):
                cell = QTableWidgetItem(str(value or ""))
                if col == 0 and issue.get("id") is not None:
                    cell.setData(Qt.UserRole, int(issue["id"]))
                self.file_related_issues_table.setItem(row, col, cell)

    # ── PDF preview helpers ─────────────────────────────────────────
    def _collapse_managed_preview(self):
        if hasattr(self, "preview_file_btn"):
            self.preview_file_btn.blockSignals(True)
            self.preview_file_btn.setChecked(False)
            self.preview_file_btn.setText("Preview")
            self.preview_file_btn.blockSignals(False)
        if hasattr(self, "pdf_viewer"):
            self.pdf_viewer.close_preview()
            self.pdf_viewer.setVisible(False)

    def _toggle_managed_preview(self, checked):
        if not hasattr(self, "pdf_viewer"):
            return
        if not checked:
            self._collapse_managed_preview()
            return
        entry = self._selected_managed_file() or {}
        if str(entry.get("file_type") or "").upper() != "PDF":
            self._collapse_managed_preview()
            return
        self.preview_file_btn.setText("Hide Preview")
        self.pdf_viewer.setVisible(True)
        history = self._selected_history_row() or {}
        if str(history.get("filename") or "").lower().endswith(".pdf"):
            self._on_version_selection_changed()
        else:
            self._try_pdf_preview_for_active(self._selected_attachment_id())

    def _try_pdf_preview_for_active(self, file_id):
        """Load the selected managed PDF into the preview."""
        if not hasattr(self, "pdf_viewer"):
            return
        if not getattr(self, "preview_file_btn", None) or not self.preview_file_btn.isChecked():
            return
        try:
            entry = self._selected_managed_file() or {}
            if str(entry.get("file_type") or "").upper() != "PDF":
                self.pdf_viewer.close_preview()
                return
            path = str(entry.get("path") or "")
            resolved = self._resolve_file_path(path)
            if resolved:
                self.pdf_viewer.load_pdf(resolved)
            else:
                self.pdf_viewer.show_error("The selected PDF attachment has no active version.")
        except Exception as exc:
            self.pdf_viewer.show_error(f"Unable to preview the active PDF:\n{exc}")

    def _on_version_selection_changed(self):
        """When a specific version row is selected, preview it if it is a PDF."""
        file_id = self._selected_attachment_id()
        version_id = self._selected_version_id()
        if file_id:
            self._load_related_issues_for_file(file_id, version_id)
        if not hasattr(self, "pdf_viewer"):
            return
        self._update_managed_file_actions()
        if not getattr(self, "preview_file_btn", None) or not self.preview_file_btn.isChecked():
            return
        try:
            history = self._selected_history_row() or {}
            if not history:
                return
            fname = str(history.get("filename") or "").lower()
            if not fname.endswith(".pdf"):
                self._collapse_managed_preview()
                return
            path = str(history.get("path") or "")
            resolved = self._resolve_file_path(path)
            if resolved:
                self.pdf_viewer.load_pdf(resolved)
            else:
                self.pdf_viewer.show_error("Unable to resolve the selected PDF version path.")
        except Exception as exc:
            self.pdf_viewer.show_error(f"Unable to preview the selected PDF version:\n{exc}")

    def release_selected_version(self):
        if not self.perm.can("release_files"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to release files.")

        file_id = self._selected_attachment_id()
        version_id = self._selected_version_id()
        if not file_id or not version_id:
            return QMessageBox.warning(self, "Select", "Select an attachment and a version.")

        try:
            self.part_file_service.release_version(version_id)
            self.refresh_files_tab()
            self.on_attachment_selected()
            if getattr(self, "current_part_id", None):
                self._refresh_part_in_tree(int(self.current_part_id))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to approve document version:\n{e}")

    def set_part_revision(self):
        if not getattr(self, "current_part_id", None):
            return
        if not self.perm.can("set_revision"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to set revisions.")

        current = ""
        try:
            details = self.bom_service.get_part_details(self.current_part_id) or {}
            current = str(details.get("revision") or "").strip().upper()
        except Exception:
            current = ""

        revision, ok = QInputDialog.getText(
            self,
            "Create New Revision",
            "New revision code (examples: B, A1, A010):",
        )
        if not ok:
            return
        revision = (revision or "").strip().upper()
        note, ok = QInputDialog.getMultiLineText(
            self,
            "Create New Revision",
            f"Reason for creating revision {revision} from {current or 'the released revision'}:",
        )
        if not ok:
            return

        try:
            created = self.bom_service.create_revision(self.current_part_id, revision, note=note)
            self._refresh_lock_family_rows(int(self.current_part_id))
            self.display_details(self.current_part_id)
            QMessageBox.information(
                self,
                "New Revision",
                f"Created {created.get('version_label') or revision + '.1'} from the released configuration.",
            )
        except Exception as e:
            QMessageBox.critical(self, "New Revision", f"Failed to create revision:\n{e}")

    def release_part_revision(self):
        if not getattr(self, "current_part_id", None):
            return
        if not self.perm.can("set_revision"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to release revisions.")
        details = self.bom_service.get_part_details(int(self.current_part_id)) or {}
        version = str(details.get("current_version") or details.get("revision") or "")
        note, ok = QInputDialog.getMultiLineText(
            self,
            "Release Revision",
            f"Release note for {version}:",
        )
        if not ok:
            return
        answer = QMessageBox.question(
            self,
            "Release Revision",
            f"Release {version}? The revision and its exact child configuration will become immutable.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return
        try:
            self.bom_service.release_part(int(self.current_part_id), note=note)
            self._refresh_lock_family_rows(int(self.current_part_id))
            self.display_details(int(self.current_part_id))
            QMessageBox.information(self, "Release Revision", f"{version} is now released.")
        except Exception as exc:
            QMessageBox.critical(self, "Release Revision", str(exc))

    def add_attachment(self):
        if not self.perm.can("release_files"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to manage files.")
        if not getattr(self, "current_part_id", None):
            return

        file_type, ok = QInputDialog.getItem(
            self, "Document Type", "Select document type:",
            ["PDF", "STEP", "DWG", "XLSX", "OTHER"], 0, False
        )
        if not ok:
            return

        display_name, ok = QInputDialog.getText(self, "Display Name", "Enter attachment name:")
        if not ok or not display_name.strip():
            return

        file_revision, ok = QInputDialog.getText(
            self, "File Revision", "File revision (optional, e.g. A010 or A020):"
        )
        if not ok:
            return

        note, _ = QInputDialog.getText(self, "Version Note", "Version note (optional):")

        source_path, _ = QFileDialog.getOpenFileName(self, "Select file")
        if not source_path:
            return
        if self._is_native_creo_path(source_path):
            return QMessageBox.warning(
                self, "Native Creo Content",
                "Native PRT, ASM, and DRW files are managed through checkout and commit."
            )

        try:
            self.part_file_service.create_attachment(
                part_id=self.current_part_id,
                file_type=file_type,
                display_name=display_name.strip(),
                description="",
                source_path=source_path,
                note=note or "",
                revision_override=str(file_revision or "").strip().upper(),
                file_role=self.managed_file_service.role_for_type(file_type),
            )
            self.refresh_files_tab()
            self._refresh_part_in_tree(int(self.current_part_id))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add attachment:\n{e}")

    def link_selected_file_to_issue(self):
        file_id = self._selected_attachment_id()
        if not file_id:
            return QMessageBox.warning(self, "Select", "Select a vaulted file first.")
        try:
            issues = self.issue_service.list_issues({"include_archived": False})
        except Exception:
            issues = []
        if not issues:
            return QMessageBox.information(self, "Issues", "No issues are available to link.")

        labels = [
            f"{issue['issue_number']} - {issue['title']} [{issue['status']}]"
            for issue in issues
        ]
        selected, ok = QInputDialog.getItem(
            self,
            "Link File to Issue",
            "Select issue:",
            labels,
            0,
            False,
        )
        if not ok or not selected:
            return
        issue = issues[labels.index(selected)]

        file_type = self._selected_file_type()
        default_role = {
            "PDF": "exported_pdf",
            "STEP": "exported_step",
            "STP": "exported_step",
        }.get(file_type, "other")
        roles = [
            "exported_pdf",
            "exported_step",
            "inspection_report",
            "screenshot",
            "supporting_doc",
            "other",
        ]
        role, ok = QInputDialog.getItem(
            self,
            "File Role",
            "Traceability role:",
            roles,
            max(0, roles.index(default_role) if default_role in roles else 0),
            False,
        )
        if not ok:
            return
        note, _ = QInputDialog.getText(self, "Link Note", "Note (optional):")
        version_id = self._selected_version_id()
        try:
            self.traceability_service.link_issue_to_engineering_file(
                int(issue["id"]),
                int(file_id),
                int(version_id) if version_id else None,
                role,
                note or "",
            )
            self._load_related_issues_for_file(file_id, version_id)
            self._refresh_linked_issue_traceability(int(issue["id"]))
            QMessageBox.information(self, "Linked", "Engineering file linked to issue.")
        except Exception as e:
            QMessageBox.critical(self, "Link File", str(e))

    def _refresh_linked_issue_traceability(self, issue_id: int):
        try:
            issue_page = getattr(self.window(), "issue_page", None)
            if not issue_page:
                return
            if getattr(issue_page, "current_issue_id", None) == int(issue_id):
                if hasattr(issue_page, "_load_engineering_files"):
                    issue_page._load_engineering_files()
                if hasattr(issue_page, "_load_history"):
                    issue_page._load_history()
                if hasattr(issue_page, "_load_commits"):
                    issue_page._load_commits()
            elif hasattr(issue_page, "refresh"):
                # Keep the issue list counts/current cache fresh without rebuilding the whole app.
                issue_page.refresh()
        except Exception:
            pass

    def add_attachment_version(self):
        if not self.perm.can("release_files"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to manage files.")
        file_id = self._selected_attachment_id()
        if not file_id:
            QMessageBox.warning(self, "Select", "Select an attachment first.")
            return
        revision, ok = QInputDialog.getText(self, "Version Revision", "Revision (optional, e.g. A010):")
        if not ok:
            return
        revision = (revision or "").strip().upper()
        note, ok = QInputDialog.getText(self, "Version Note", "Version note (optional):")
        if not ok:
            return
        source_path, _ = QFileDialog.getOpenFileName(self, "Select new version file")
        if not source_path:
            return
        try:
            self.part_file_service.add_new_version_with_revision(
                file_id,
                source_path,
                note=note or "",
                revision=revision,
            )
            self.refresh_files_tab()
            # keep versions visible
            self.on_attachment_selected()
            if getattr(self, "current_part_id", None):
                self._refresh_part_in_tree(int(self.current_part_id))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add version:\n{e}")

    def set_active_version(self):
        if not self.perm.can("release_files"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to manage files.")
        file_id = self._selected_attachment_id()
        version_id = self._selected_version_id()
        if not file_id or not version_id:
            QMessageBox.warning(self, "Select", "Select a document and a historical version.")
            return
        try:
            self.part_file_service.set_active_version(file_id, version_id)
            self.refresh_files_tab()
            self.on_attachment_selected()
            if getattr(self, "current_part_id", None):
                self._refresh_part_in_tree(int(self.current_part_id))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to select the working document version:\n{e}")

    def open_active_attachment(self):
        entry = self._selected_managed_file() or {}
        if not entry:
            QMessageBox.warning(self, "Select", "Select managed content.")
            return
        self._open_file(str(entry.get("path") or ""), "Managed Content")

    def open_active_attachment_folder(self):
        entry = self._selected_managed_file() or {}
        if not entry:
            QMessageBox.warning(self, "Select", "Select managed content.")
            return
        path = str(entry.get("path") or "")
        resolved = self._resolve_file_path(path)
        if not resolved or not safe_exists(resolved):
            QMessageBox.warning(self, "Missing File", f"Attachment file not found:\n{resolved}")
            return
        folder = os.path.dirname(os.path.abspath(os.path.normpath(resolved)))
        if not folder or not safe_exists(folder):
            QMessageBox.warning(self, "Missing Folder", f"Managed file folder not found:\n{folder}")
            return
        try:
            safe_startfile(folder)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open managed file folder:\n{e}")

    def _infer_file_type_from_path(self, path: str) -> str:
        ext = os.path.splitext(path or "")[1].lower()
        if ext == ".pdf":
            return "PDF"
        if ext in (".step", ".stp"):
            return "STEP"
        if ext in (".dwg", ".dxf"):
            return "DWG"
        if ext in (".xlsx", ".xls"):
            return "XLSX"
        return "OTHER"

    @staticmethod
    def _is_native_creo_path(path: str) -> bool:
        return bool(re.search(r"\.(?:prt|asm|drw)(?:\.\d+)?$", str(path or ""), re.IGNORECASE))

    def _on_files_dropped(self, paths):
        if not getattr(self, "current_part_id", None):
            QMessageBox.warning(self, "Select", "Select a part first.")
            return
        revision, ok = QInputDialog.getText(self, "Version Revision", "Revision for dropped files (optional, e.g. A010):")
        if not ok:
            return
        revision = (revision or "").strip().upper()
        note, ok = QInputDialog.getText(self, "Version Note", "Version note for dropped files (optional):")
        if not ok:
            return

        added = 0
        failed = []
        # cache current attachments for quick type lookup; refresh when we create a new attachment
        attachments = self.part_file_service.list_attachments(self.current_part_id)

        for p in paths:
            try:
                if not p or not safe_exists(p):
                    raise ValueError("File not found")
                if self._is_native_creo_path(p):
                    raise ValueError("Native Creo content must be added through checkout and commit")

                file_type = self._infer_file_type_from_path(p)
                if file_type == "OTHER":
                    file_type, ok = QInputDialog.getItem(
                        self,
                        "File Type",
                        f"Select type for:\n{os.path.basename(p)}",
                        ["PDF", "STEP", "DWG", "XLSX", "OTHER"],
                        4,
                        False,
                    )
                    if not ok:
                        continue

                # If an attachment of this type already exists for the part, add as a new version
                existing = None
                try:
                    for f in (attachments or []):
                        expected_role = self.managed_file_service.role_for_type(file_type)
                        if (
                            str((f.file_type or "")).upper() == str(file_type or "").upper()
                            and str(getattr(f, "file_role", "") or "").lower() == expected_role
                        ):
                            existing = f
                            break
                except Exception:
                    existing = None

                if existing:
                    # add new version to existing attachment
                    self.part_file_service.add_new_version_with_revision(
                        existing.id,
                        p,
                        note=note or "",
                        revision=revision,
                    )
                else:
                    display_name = os.path.splitext(os.path.basename(p))[0]
                    self.part_file_service.create_attachment(
                        part_id=self.current_part_id,
                        file_type=file_type,
                        display_name=display_name,
                        description="",
                        source_path=p,
                        note=note or "",
                        revision_override=revision,
                        file_role=self.managed_file_service.role_for_type(file_type),
                    )
                    # refresh cached attachments so subsequent dropped files see the new attachment
                    try:
                        attachments = self.part_file_service.list_attachments(self.current_part_id)
                    except Exception:
                        pass
                added += 1
            except Exception as e:
                failed.append(f"{os.path.basename(p)}: {e}")

        if added:
            self.refresh_files_tab()
            self._refresh_part_in_tree(int(self.current_part_id))
        if failed:
            QMessageBox.warning(self, "Some files failed", "\n".join(failed[:15]))

    def open_selected_version(self):
        history = self._selected_history_row() or {}
        if not history:
            QMessageBox.warning(self, "Select", "Select historical content.")
            return
        self._open_file(str(history.get("path") or ""), "Historical Content")

    def delete_selected_version(self):
        if not self.perm.can("release_files"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to manage files.")
        file_id = self._selected_attachment_id()
        version_id = self._selected_version_id()
        if not file_id or not version_id:
            QMessageBox.warning(self, "Select", "Select an attachment and a version.")
            return
        confirm = QMessageBox.question(
            self, "Obsolete Version",
            "Mark the selected document version obsolete? The managed content will be preserved.",
            QMessageBox.Yes | QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return
        try:
            self.part_file_service.delete_version(file_id, version_id)
            self.refresh_files_tab()
            self.on_attachment_selected()
            if getattr(self, "current_part_id", None):
                self._refresh_part_in_tree(int(self.current_part_id))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to obsolete version:\n{e}")

    def remove_attachment(self):
        if not self.perm.can("release_files"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to manage files.")
        file_id = self._selected_attachment_id()
        if not file_id:
            QMessageBox.warning(self, "Select", "Select an attachment.")
            return
        confirm = QMessageBox.question(
            self, "Obsolete Document",
            "Mark this document and its versions obsolete? Managed content and released references will be preserved.",
            QMessageBox.Yes | QMessageBox.No,
        )
        if confirm != QMessageBox.Yes:
            return
        try:
            self.part_file_service.delete_attachment(file_id)
            self.refresh_files_tab()
            if getattr(self, "current_part_id", None):
                self._refresh_part_in_tree(int(self.current_part_id))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to remove attachment:\n{e}")

    def export_package(self):
        mode, ok = QInputDialog.getItem(
            self,
            "Export Package",
            "Export mode:",
            ["Selected part", "Selected part + children", "Choose parts (any)"],
            1,
            False,
        )
        if not ok:
            return

        part_ids = []
        include_children = False

        if mode.startswith("Selected"):
            if not getattr(self, "current_part_id", None):
                QMessageBox.warning(self, "Select", "Select a part first.")
                return
            part_ids = [self.current_part_id]
            include_children = mode.endswith("children")
        else:
            current_filter_ids = self._current_bom_view_item_ids_for_selector()
            use_current_filter = self._has_active_bom_view_filter_for_selector()
            dlg = PackagePartsDialog(
                self,
                project_id=self.session.project_id,
                preselected_ids=[],
                current_filter_ids=current_filter_ids,
                use_current_filter=use_current_filter,
                title="Export Package - Select Items",
                subtitle="Filter the EBOM scope and select all filtered Items for the delivery package.",
                kicker="DELIVERY PACKAGE",
            )
            if dlg.exec_() != QDialog.Accepted:
                return
            part_ids = dlg.selected_part_ids()
            if not part_ids:
                QMessageBox.warning(self, "Select", "No parts selected.")
                return
            include_children = QMessageBox.question(
                self,
                "Export Package",
                "Include all children of the selected parts?",
                QMessageBox.Yes | QMessageBox.No,
            ) == QMessageBox.Yes

        package_name, ok = QInputDialog.getText(self, "Package Name", "Enter package name:")
        if not ok or not package_name.strip():
            return

        dest_dir = QFileDialog.getExistingDirectory(self, "Select destination folder")
        if not dest_dir:
            return

        create_zip = (
            QMessageBox.question(
                self,
                "Export Package",
                "Create a ZIP archive too?",
                QMessageBox.Yes | QMessageBox.No,
            )
            == QMessageBox.Yes
        )

        def do_export(progress_callback):
            return self.package_export_service.export_package_for_parts(
                part_ids=part_ids,
                destination_dir=dest_dir,
                include_children=include_children,
                package_name=package_name.strip(),
                create_zip=create_zip,
                progress_callback=progress_callback,
            )

        def show_export_result(manifest):
            missing_count = len(manifest.get("missing", []))
            out_dir = manifest.get("package", {}).get("output_dir", "")
            zip_path = manifest.get("package", {}).get("zip_path", "")
            QMessageBox.information(
                self,
                "Export Complete",
                f"Package exported to:\n{out_dir}\n\nZIP: {zip_path or 'No'}\nMissing files: {missing_count}",
            )
            if out_dir:
                try:
                    safe_startfile(out_dir)
                except Exception:
                    pass

        self._run_export_with_progress(
            title="Export Package",
            initial_message="Preparing package export...",
            operation=do_export,
            on_success=show_export_result,
            error_title="Export Package Failed",
        )

    def _resolve_file_path(self, path: str) -> str:
        if not path:
            return ""
        path = path.strip()
        if not path:
            return ""
        if os.path.isabs(path):
            return path
        if getattr(self, "working_dir", None):
            candidate = os.path.join(self.working_dir, path)
            return candidate
        return path

    def _open_file(self, path: str, label: str):
        resolved = self._resolve_file_path(path)
        if not resolved:
            QMessageBox.warning(self, "Not Found", f"No {label} file associated with this part.")
            return
        if not safe_exists(resolved):
            QMessageBox.warning(self, "Missing File", f"{label} file not found:\n{resolved}")
            return
        try:
            safe_startfile(resolved)  # Windows
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open {label} file:\n{e}")

    def add_pdf_to_part(self, part_id: int):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select PDF file", "", "PDF Files (*.pdf);;All Files (*)")
        if not file_path:
            return
        try:
            self.bom_service.update_part(part_id, {"pdf_path": file_path})
            self._refresh_part_in_tree(int(part_id))
            QMessageBox.information(self, "Success", "PDF associated with part successfully.")
            self.display_details(part_id)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to attach PDF: {e}")

    def add_step_to_part(self, part_id: int):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select STEP file", "", "STEP Files (*.step *.stp);;All Files (*)")
        if not file_path:
            return
        try:
            self.bom_service.update_part(part_id, {"step_path": file_path})
            self._refresh_part_in_tree(int(part_id))
            QMessageBox.information(self, "Success", "STEP associated with part successfully.")
            self.display_details(part_id)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to attach STEP: {e}")

    def open_part_pdf(self, part_id: int):
        path = ""
        try:
            for attachment in self.part_file_service.list_attachments(int(part_id)):
                if str(attachment.file_type or "").upper() == "PDF":
                    path = self.part_file_service.resolve_active_path(attachment.id)
                    if path:
                        break
        except Exception:
            path = ""
        if not path:
            details = self.bom_service.get_part_details(part_id) or {}
            path = details.get("pdf_path", "")
        self._open_file(path, "PDF")

    def preview_part_pdf(self, part_id: int):
        """Select a BOM item, open its Files tab, and preview its active PDF."""
        self.display_details(int(part_id))
        try:
            self.tabs.setCurrentWidget(self.files_tab)
        except Exception:
            pass
        if not self.pdf_viewer.is_loaded():
            details = self.bom_service.get_part_details(int(part_id)) or {}
            path = self._resolve_file_path(details.get("pdf_path", ""))
            if path:
                self.pdf_viewer.load_pdf(path)
            else:
                self.pdf_viewer.show_error("No PDF attachment is available for this BOM item.")

    def open_part_step(self, part_id: int):
        details = self.bom_service.get_part_details(part_id) or {}
        self._open_file(details.get("step_path", ""), "STEP")

    def _open_step_in_3d_viewer(self, part_id: int):
        """Open the part's latest committed STEP file in the advanced 3D
        viewer with face-lifecycle tracking enabled.

        The viewer uses the STEP copies stored by the commit workflow
        (commits.step_file_path) — these are the same files used for
        the STEP diff engine, so face fingerprints match the lifecycle
        history."""
        try:
            latest_step = self.commit_repo.get_latest_step_commit_for_part(
                int(part_id), int(self.session.project_id)
            )
        except Exception as exc:
            QMessageBox.critical(self, "3D Viewer", f"Failed to query committed STEP:\n{exc}")
            return

        if not latest_step or not latest_step.step_file_path:
            QMessageBox.warning(
                self, "3D Viewer",
                "No committed STEP file found for this part.\n\n"
                "Attach a STEP file when committing to enable 3D viewing\n"
                "and face-level lifecycle tracking.",
            )
            return

        step_path = latest_step.step_file_path
        if not os.path.exists(step_path):
            QMessageBox.warning(
                self, "3D Viewer",
                f"Committed STEP file not found on disk:\n{step_path}\n\n"
                "The file may have been moved or deleted.",
            )
            return

        try:
            from tools.CAD.step_viewer.launcher import launch_viewer
            from config import DB_NAME
            launch_viewer(
                step_path,
                part_id=part_id,
                project_id=self.session.project_id,
                db_path=DB_NAME,
                face_map_path=getattr(latest_step, "step_face_map_path", None),
            )
        except Exception as e:
            QMessageBox.critical(self, "3D Viewer", f"Failed to open 3D viewer:\n{e}")

    def _set_files_controls_enabled(self, enabled: bool):
        for w in (
            getattr(self, "pdf_browse_btn", None),
            getattr(self, "pdf_open_btn", None),
            getattr(self, "pdf_clear_btn", None),
            getattr(self, "step_browse_btn", None),
            getattr(self, "step_open_btn", None),
            getattr(self, "step_clear_btn", None),
        ):
            if w is not None:
                w.setEnabled(enabled)

    def _refresh_files_tab(self, details: dict):
        if not details:
            if hasattr(self, "pdf_path_view"):
                self.pdf_path_view.setText("")
            if hasattr(self, "step_path_view"):
                self.step_path_view.setText("")
            self._set_files_controls_enabled(False)
            return

        self._set_files_controls_enabled(True)
        if hasattr(self, "pdf_path_view"):
            self.pdf_path_view.setText(details.get("pdf_path", "") or "")
        if hasattr(self, "step_path_view"):
            self.step_path_view.setText(details.get("step_path", "") or "")

    def browse_pdf_for_selected(self):
        if not getattr(self, "current_part_id", None):
            return
        self.add_pdf_to_part(self.current_part_id)

    def browse_step_for_selected(self):
        if not getattr(self, "current_part_id", None):
            return
        self.add_step_to_part(self.current_part_id)

    def open_pdf_for_selected(self):
        if not getattr(self, "current_part_id", None):
            return
        self.open_part_pdf(self.current_part_id)

    def open_step_for_selected(self):
        if not getattr(self, "current_part_id", None):
            return
        self.open_part_step(self.current_part_id)

    def clear_pdf_for_selected(self):
        if not getattr(self, "current_part_id", None):
            return
        try:
            self.bom_service.update_part(self.current_part_id, {"pdf_path": ""})
            self._refresh_part_in_tree(int(self.current_part_id))
            self.display_details(self.current_part_id)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to clear PDF: {e}")

    def clear_step_for_selected(self):
        if not getattr(self, "current_part_id", None):
            return
        try:
            self.bom_service.update_part(self.current_part_id, {"step_path": ""})
            self._refresh_part_in_tree(int(self.current_part_id))
            self.display_details(self.current_part_id)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to clear STEP: {e}")

    def compare_part_structure(self, part_id=None, whole_bom: bool = False):
        if isinstance(part_id, bool):
            part_id = None
        if not whole_bom and not part_id:
            part_id = getattr(self, "current_part_id", None)
        if not whole_bom and not part_id:
            return QMessageBox.warning(self, "Compare", "Select a BOM item first.")
        if not getattr(self.session, "project_id", None):
            return QMessageBox.warning(self, "Compare", "No active project selected.")
        try:
            dlg = WindchillCompareSetupDialog(
                self,
                self.bom_service,
                self.project_service,
                int(self.session.project_id),
                int(part_id) if part_id else None,
                whole_bom=whole_bom,
            )
            dlg.exec_()
        except Exception as e:
            QMessageBox.critical(self, "Compare", f"Failed to open structure compare:\n{e}")

    def compare_assembly_iterations(self, part_id=None):
        """Open the immutable occurrence comparison for one assembly."""
        if isinstance(part_id, bool):
            part_id = None
        part_id = part_id or getattr(self, "current_part_id", None)
        if not part_id:
            return QMessageBox.warning(
                self, "Compare Assembly Iterations", "Select an assembly first."
            )
        try:
            details = self.bom_service.get_part_details(int(part_id)) or {}
            if str(details.get("type") or "").strip().lower() not in {"asm", "assembly"}:
                return QMessageBox.warning(
                    self,
                    "Compare Assembly Iterations",
                    "Iteration comparison is available only for assemblies.",
                )
            iterations = self.bom_service.list_part_iterations(int(part_id)) or []
            if len(iterations) < 2:
                return QMessageBox.information(
                    self,
                    "Compare Assembly Iterations",
                    "This assembly needs at least two checked-in iterations before they can be compared.",
                )
            dialog = AssemblyIterationCompareDialog(
                self,
                self.bom_service,
                int(part_id),
                cad_file_opener=self._show_iteration_cad_files,
            )
            dialog.exec_()
        except Exception as exc:
            QMessageBox.critical(
                self,
                "Compare Assembly Iterations",
                f"Failed to compare assembly iterations:\n{exc}",
            )

    def create_assembly_configuration(self, part_id=None):
        if isinstance(part_id, bool):
            part_id = None
        part_id = part_id or getattr(self, "current_part_id", None)
        if not part_id:
            return QMessageBox.warning(
                self, "Create Configuration", "Select an assembly first."
            )
        project_id = getattr(self.session, "project_id", None)
        if not project_id:
            return QMessageBox.warning(
                self, "Create Configuration", "No active project is selected."
            )
        try:
            details = self.bom_service.get_part_details(int(part_id)) or {}
            if str(details.get("type") or "").strip().lower() not in {"asm", "assembly"}:
                return QMessageBox.warning(
                    self, "Create Configuration", "A configuration must start from an assembly."
                )
            if not self.bom_service.list_part_iterations(int(part_id)):
                return QMessageBox.information(
                    self,
                    "Create Configuration",
                    "This assembly has no checked-in iteration to freeze.",
                )
            dialog = CreateAssemblyConfigurationDialog(
                self,
                self.assembly_configuration_service,
                self.bom_service,
                int(project_id),
                int(part_id),
            )
            dialog.exec_()
        except Exception as exc:
            QMessageBox.critical(
                self, "Create Configuration", f"Failed to create configuration:\n{exc}"
            )

    def manage_assembly_configurations(self):
        project_id = getattr(self.session, "project_id", None)
        if not project_id:
            return QMessageBox.warning(
                self, "Manage Configurations", "No active project is selected."
            )
        try:
            dialog = ManageAssemblyConfigurationsDialog(
                self, self.assembly_configuration_service, int(project_id)
            )
            dialog.exec_()
        except Exception as exc:
            QMessageBox.critical(
                self, "Manage Configurations", f"Failed to open configurations:\n{exc}"
            )


    def add_child(self, parent_item=None):
        """Select a parent, then select a child directly from the BOM/search view."""
        if not self.perm.can("manage_parts"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to add child relations.")
        if isinstance(parent_item, bool) or not isinstance(parent_item, QTreeWidgetItem):
            parent_item = None
        if parent_item is not None and (
            self._is_folder_tree_item(parent_item)
            or str(parent_item.text(BOM_COL_TYPE) or "").strip().lower() not in {"asm", "assembly"}
        ):
            return QMessageBox.warning(self, "Add Child", "Only an assembly can contain child items.")

        if parent_item is None:
            state = {"phase": "select_parent", "selections": []}
            message = "Select the parent assembly in the BOM. You can use Search to locate it. Press Esc to cancel."
        else:
            state = {
                "phase": "select_child",
                "target_parent_id": int(parent_item.data(0, Qt.UserRole)),
                "target_parent_name": str(parent_item.text(BOM_COL_NAME) or ""),
                "selections": [],
            }
            message = (
                f"Parent selected: {state['target_parent_name']}. Select a child in the BOM; "
                "use Search if needed. Press Esc to cancel."
            )
        self._start_relation_selection_mode(state, message)

    def add_selected_to_parent(self, selected_items) -> None:
        """Keep selected child occurrences and wait for a target parent in the BOM."""
        if not self.perm.can("manage_parts"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to change BOM relations.")
        selections = []
        for item in selected_items or []:
            payload = self._relation_selection_for_item(item)
            if payload is not None:
                selections.append(payload)
        if not selections:
            return QMessageBox.warning(self, "Add to Parent", "Select one or more BOM items first.")
        state = {"phase": "select_target_parent", "selections": selections}
        count = len(selections)
        self._start_relation_selection_mode(
            state,
            f"{count} item{'s' if count != 1 else ''} selected. Select the target parent assembly in the BOM; "
            "use Search if needed. Press Esc to cancel.",
        )

    def _start_relation_selection_mode(self, state: dict, message: str) -> None:
        self._exit_relation_selection_mode(clear_status=False)
        self._relation_selection_state = dict(state or {})
        for tree in (self.tree, self._search_tree):
            try:
                tree.itemClicked.connect(self._on_relation_selection_clicked)
            except Exception:
                pass
        self.setCursor(Qt.PointingHandCursor)
        self.window().statusBar().showMessage(message)

    def _exit_relation_selection_mode(self, clear_status: bool = True) -> None:
        for tree in (getattr(self, "tree", None), getattr(self, "_search_tree", None)):
            if tree is None:
                continue
            try:
                tree.itemClicked.disconnect(self._on_relation_selection_clicked)
            except Exception:
                pass
        self._relation_selection_state = None
        self.setCursor(Qt.ArrowCursor)
        if clear_status:
            try:
                self.window().statusBar().showMessage("Relation operation cancelled.")
            except Exception:
                pass

    def _relation_selection_for_item(self, item: QTreeWidgetItem) -> dict | None:
        if item is None or self._is_folder_tree_item(item) or self._is_lazy_placeholder(item):
            return None
        try:
            child_id = int(item.data(0, Qt.UserRole))
        except (TypeError, ValueError):
            return None
        payload = {
            "child_id": child_id,
            "child_name": str(item.text(BOM_COL_NAME) or item.text(0) or child_id),
            "source_parent_id": None,
            "source_options": [],
        }
        if item.treeWidget() is self.tree:
            payload["source_parent_id"] = self._direct_relation_parent_for_item(item)
            return payload

        try:
            sources = self.bom_service.relation_sources_for_part(child_id) or []
        except Exception:
            sources = []
        if len(sources) == 1:
            payload["source_parent_id"] = int(sources[0]["parent_id"])
        elif len(sources) > 1:
            payload["source_options"] = list(sources)
        return payload

    def _direct_relation_parent_for_item(self, item: QTreeWidgetItem) -> int | None:
        if item is None or item.treeWidget() is not self.tree:
            return None
        parent = item.parent()
        while parent is not None and self._is_folder_tree_item(parent):
            parent = parent.parent()
        if parent is None or parent.data(0, Qt.UserRole) is None:
            return None
        try:
            return int(parent.data(0, Qt.UserRole))
        except (TypeError, ValueError):
            return None

    def _on_relation_selection_clicked(self, item: QTreeWidgetItem, _column: int = 0) -> None:
        state = getattr(self, "_relation_selection_state", None)
        if not state:
            return
        phase = str(state.get("phase") or "")
        if phase in {"select_parent", "select_target_parent"}:
            if self._is_folder_tree_item(item):
                self.window().statusBar().showMessage("Select an assembly, not an organizational folder.")
                return
            if str(item.text(BOM_COL_TYPE) or "").strip().lower() not in {"asm", "assembly"}:
                self.window().statusBar().showMessage("The target parent must be an assembly.")
                return
            target_parent_id = int(item.data(0, Qt.UserRole))
            target_parent_name = str(item.text(BOM_COL_NAME) or "")
            if phase == "select_parent":
                state["phase"] = "select_child"
                state["target_parent_id"] = target_parent_id
                state["target_parent_name"] = target_parent_name
                self.window().statusBar().showMessage(
                    f"Parent selected: {target_parent_name}. Select a child in the BOM; use Search if needed. "
                    "Press Esc to cancel."
                )
                return
            self._finish_relation_selection(target_parent_id, target_parent_name, state.get("selections") or [])
            return

        if phase == "select_child":
            selection = self._relation_selection_for_item(item)
            if selection is None:
                return
            self._finish_relation_selection(
                int(state["target_parent_id"]),
                str(state.get("target_parent_name") or ""),
                [selection],
            )

    def _ask_relation_action(self, count: int, target_name: str) -> str | None:
        box = QMessageBox(self)
        box.setWindowTitle("Add to Parent")
        box.setIcon(QMessageBox.Question)
        box.setText(
            f"Add {count} selected item{'s' if count != 1 else ''} to {target_name}?"
        )
        box.setInformativeText("Copy keeps the current occurrence. Move transfers it to the selected parent.")
        copy_button = box.addButton("Copy", QMessageBox.AcceptRole)
        move_button = box.addButton("Move", QMessageBox.ActionRole)
        box.addButton(QMessageBox.Cancel)
        box.exec_()
        clicked = box.clickedButton()
        if clicked is copy_button:
            return "copy"
        if clicked is move_button:
            return "move"
        return None

    def _resolve_relation_sources(self, selections: list[dict], action: str) -> list[dict] | None:
        resolved = []
        for selection in selections:
            row = dict(selection)
            options = list(row.pop("source_options", []) or [])
            row.pop("child_name", None)
            if options:
                labels = []
                by_label = {}
                for option in options:
                    number = str(
                        option.get("parent_part_number")
                        or option.get("parent_aes_number")
                        or ""
                    ).strip()
                    name = str(option.get("parent_name") or "").strip()
                    quantity = int(option.get("quantity") or 1)
                    label = f"{number} - {name} (Qty {quantity})" if number else f"{name} (Qty {quantity})"
                    labels.append(label)
                    by_label[label] = option
                selected_label, ok = QInputDialog.getItem(
                    self,
                    "Select Direct Occurrence",
                    f"Choose the source occurrence to {action}:",
                    labels,
                    0,
                    False,
                )
                if not ok:
                    return None
                row["source_parent_id"] = int(by_label[selected_label]["parent_id"])
            resolved.append(row)
        return resolved

    def _finish_relation_selection(
        self,
        target_parent_id: int,
        target_parent_name: str,
        selections: list[dict],
    ) -> None:
        action = self._ask_relation_action(len(selections), target_parent_name)
        if action is None:
            self._exit_relation_selection_mode()
            return
        resolved = self._resolve_relation_sources(selections, action)
        if resolved is None:
            return
        try:
            result = self.bom_service.apply_child_relation_operation(
                int(target_parent_id), resolved, action
            )
        except Exception as exc:
            QMessageBox.critical(self, "BOM Structure", str(exc))
            return

        changed = len(result.get("child_ids") or [])
        skipped = len(result.get("skipped_child_ids") or [])
        verb = "moved" if action == "move" else "copied"
        result["message"] = f"{changed} item{'s' if changed != 1 else ''} {verb}."
        if skipped:
            result["message"] += f" {skipped} already under the selected parent were skipped."
        self._exit_relation_selection_mode(clear_status=False)
        self._apply_child_relation_result(result)

    def _start_interactive_add_child(self, parent_item=None):
        """Start interactive add-child mode (select parent → child → confirm)."""
        if not self.perm.can("manage_parts"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to add child relations.")
        if isinstance(parent_item, bool):
            parent_item = None
        if parent_item is not None and not isinstance(parent_item, QTreeWidgetItem):
            parent_item = None
        if getattr(self, "add_child_mode", False):
            self._exit_add_child_mode()

        selection_tree = parent_item.treeWidget() if parent_item is not None else self._current_tree_for_filtering()
        if selection_tree is None:
            selection_tree = self.tree
        if parent_item is not None:
            if self._is_folder_tree_item(parent_item):
                return QMessageBox.warning(self, "Add Child", "An organizational folder cannot be a BOM parent.")
            part_type = str(parent_item.text(BOM_COL_TYPE) or "").strip().lower()
            if part_type not in {"asm", "assembly"}:
                return QMessageBox.warning(self, "Add Child", "Only an assembly can contain child items.")

        self.parent_item = parent_item
        self.child_item = None
        self.add_child_mode = True
        self._add_child_selection_tree = selection_tree

        # 🔹 Keep your existing design — just give a temporary focus cue
        pal = selection_tree.palette()
        self._original_base_color = pal.color(QPalette.Base)
        pal.setColor(QPalette.Base, QColor(230, 245, 255))  # very light blue background
        selection_tree.setPalette(pal)

        # Cursor + status hint
        self.setCursor(Qt.PointingHandCursor)
        self.window().statusBar().showMessage("🔹 Select the parent part in the tree (Press Esc to cancel)...")

        if parent_item is not None:
            self.window().statusBar().showMessage(
                f"Parent selected: {parent_item.text(BOM_COL_NAME)} | Now select the child part (Press Esc to cancel)..."
            )

        # Connect handlers
        selection_tree.itemClicked.connect(self._on_tree_click_for_add_child)


    def _on_tree_click_for_add_child(self, item):
        """Handle parent → child selection sequence."""
        if not self.add_child_mode:
            return

        # Step 1: select parent
        if self.parent_item is None:
            part_type = item.text(BOM_COL_TYPE).strip().lower()
            if part_type not in {"asm", "assembly"}:
                self.window().statusBar().showMessage("⚠️ Selected part is not an ASM. Please select a valid assembly.")
                return  # Don’t continue

            self.parent_item = item
            self.window().statusBar().showMessage(f"Parent selected: {item.text(BOM_COL_NAME)} | Now select the child part...")
            return

        # Step 2: select child
        if self.child_item is None:
            if item == self.parent_item:
                self.window().statusBar().showMessage("⚠️ Parent and child cannot be the same part. Select a different child or press Esc to cancel.")
                return

            self.child_item = item
            parent_name = self.parent_item.text(BOM_COL_NAME)
            child_name = self.child_item.text(BOM_COL_NAME)

            # Confirm relationship
            confirm = QMessageBox.question(
                self,
                "Confirm Relation",
                f"Add relation:\n\nParent: {parent_name}\nChild: {child_name}\n\nConfirm?",
                QMessageBox.Yes | QMessageBox.No,
            )

            if confirm == QMessageBox.Yes:
                parent_id = self.parent_item.data(0, Qt.UserRole)
                child_id = self.child_item.data(0, Qt.UserRole)
                try:
                    self.bom_service.add_child_by_id(parent_id, child_id)
                    self._add_child_relation_to_trees(int(parent_id), int(child_id), self.child_item)
                    self.window().statusBar().showMessage("Child added successfully.")
                    try:
                        self.display_details(int(parent_id))
                    except Exception:
                        pass
                except Exception as e:
                    self.window().statusBar().showMessage(f"Failed to add child: {str(e)}")
            else:
                self.window().statusBar().showMessage("Add-child operation cancelled.")

            # Exit selection mode
            self._exit_add_child_mode()


    def _exit_add_child_mode(self):
        """Clean up state, visuals, and connections."""
        selection_tree = getattr(self, "_add_child_selection_tree", None)
        if selection_tree is not None:
            try:
                selection_tree.itemClicked.disconnect(self._on_tree_click_for_add_child)
            except Exception:
                pass
            try:
                palette = selection_tree.palette()
                palette.setColor(QPalette.Base, self._original_base_color)
                selection_tree.setPalette(palette)
            except Exception:
                pass

        self.setCursor(Qt.ArrowCursor)
        #self.window().statusBar().clearMessage()
        self.parent_item = None
        self.child_item = None
        self.add_child_mode = False
        self._add_child_selection_tree = None


    def keyPressEvent(self, event: QKeyEvent):
        """Handle ESC to cancel add-child mode."""
        if getattr(self, "_pdm_structure_target_state", None) and event.key() == Qt.Key_Escape:
            self._exit_pdm_structure_target_selection()
        elif getattr(self, "_relation_selection_state", None) and event.key() == Qt.Key_Escape:
            self._exit_relation_selection_mode()
        elif getattr(self, "add_child_mode", False) and event.key() == Qt.Key_Escape:
            self.window().statusBar().showMessage("Add-child operation cancelled.")
            self._exit_add_child_mode()
        else:
            # pass to default handler
            super().keyPressEvent(event)

    def add_dwg_to_part(self, parent_aes):
        part = self.bom_service.get_part_details(parent_aes)
        if not part:
            QMessageBox.warning(self, "Not Found", f"Part with AES {parent_aes} not found.")
            return

        file_path, _ = QFileDialog.getOpenFileName(self, "Select drawing file", "", "All Files (*);;PDF Files (*.pdf);;Images (*.png *.jpg *.jpeg *.bmp)")
        if not file_path:
            return

        drawing_no, ok = QInputDialog.getText(self, "Drawing Number", "Enter drawing number (optional):", text=str(part.get("drawing_number", "")))
        if not ok:
            return

        update_payload = {
            "drawing": file_path,
            "drawing_number": drawing_no or part.get("drawing_number")
        }

        try:
            self.bom_service.update_part(parent_aes, update_payload)
            self._refresh_part_in_tree(int(parent_aes))
            QMessageBox.information(self, "Success", "Drawing associated with part successfully.")
            self.display_details(parent_aes)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to attach drawing: {e}")

    # -------------------------
    # Basic CRUD
    # -------------------------
    def _schedule_search(self, _text: str = ""):
        """Debounce search requests to avoid expensive reloads while typing."""
        try:
            self._search_timer.start()
        except Exception:
            self._perform_search_now()

    def perform_search(self):
        """Backward-compatible entrypoint (called elsewhere)."""
        self._perform_search_now()

    def _perform_search_now(self):
        query = self.search_input.text().strip()
        if getattr(self, "_bom_mode", "cad") == "cad":
            visible = self._refresh_pdm_cad_filter()
            self._tree_stack.setCurrentWidget(self._cad_tree)
            try:
                self.clear_filter_btn.setEnabled(bool(query))
                self.search_btn.setText(f"Search ({visible})" if query else "Search")
            except Exception:
                pass
            return
        if getattr(self, "_bom_mode", "cad") == "ebom":
            visible = self._refresh_ebom_filters()
            if not self._is_default_bom_advanced_filter():
                self._update_advanced_filter_button_state(visible)
            return
        if not query:
            self._exit_search_mode()
            if not self._is_default_bom_advanced_filter():
                self.apply_bom_tree_filter(self._bom_advanced_filters)
            return

        # Show spinner immediately, then fetch results in a background thread.
        self._tree_load_seq += 1
        seq = self._tree_load_seq
        self._set_tree_loading(True)

        try:
            if self._search_worker is not None:
                self._search_worker.cancel()
        except Exception:
            pass

        worker = _SearchWorker(seq, self.bom_service, query)
        thread = QThread(self)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(self._on_search_finished)
        worker.failed.connect(self._on_search_failed)
        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        self._search_worker = worker
        self._search_thread = thread
        thread.start()

    def _on_search_finished(self, seq: int, results: object) -> None:
        if getattr(self, "_bom_mode", "cad") == "ebom":
            return
        if int(seq) != int(getattr(self, "_tree_load_seq", 0)):
            return
        self._begin_search_build(int(seq), list(results or []))

    def _on_search_failed(self, seq: int, err: str) -> None:
        if int(seq) != int(getattr(self, "_tree_load_seq", 0)):
            return
        try:
            # Show empty results on error; keep UI responsive
            self._search_tree.clear()
        except Exception:
            pass
        self._set_tree_loading(False)

    def _begin_search_build(self, seq: int, results: list) -> None:
        self._search_build_seq = int(seq)
        try:
            self._search_build_timer.stop()
        except Exception:
            pass
        self._search_build_queue = deque(results or [])

        # A new basic search replaces any previous advanced flat result list.
        self._advanced_filter_flat_mode = False
        self._enter_search_mode()   # switch stack to _search_tree (index 2), clears it

        self._set_tree_loading(True)
        try:
            self._search_tree.setUpdatesEnabled(False)
        except Exception:
            pass
        self._search_build_timer.start()

    def _search_build_step(self) -> None:
        if int(self._search_build_seq) != int(getattr(self, "_tree_load_seq", 0)):
            try:
                self._search_build_timer.stop()
            except Exception:
                pass
            return

        start = time.perf_counter()
        processed = 0
        try:
            while self._search_build_queue and processed < 200 and (time.perf_counter() - start) < 0.01:
                part = self._search_build_queue.popleft()
                item = self._make_tree_item(part)
                self._search_tree.addTopLevelItem(item)
                processed += 1

            if not self._search_build_queue:
                self._search_build_timer.stop()
                try:
                    self._sync_search_tree_row_numbers()
                    self._search_tree.setUpdatesEnabled(True)
                except Exception:
                    pass
                try:
                    # Search results are flat; no need to expand deeply.
                    self._search_tree.expandToDepth(1)
                except Exception:
                    pass
                try:
                    if not self._is_default_bom_advanced_filter():
                        self.apply_bom_tree_filter(self._bom_advanced_filters)
                except Exception:
                    pass
                self._set_tree_loading(False)
        except RuntimeError:
            try:
                self._search_build_timer.stop()
            except Exception:
                pass
            return

    def _issues_for_part(self, part_id: int | None) -> set:
        issues = set()
        if part_id is None:
            return issues
        try:
            pid = int(part_id)
        except Exception:
            return issues
        for row in (self.missing_files or []):
            try:
                bom_id, issue_type, _filename = row
                if int(bom_id) == pid:
                    issues.add(str(issue_type))
            except Exception:
                continue
        return issues

    def _rebuild_missing_ids(self) -> None:
        ids = set()
        for row in (self.missing_files or []):
            try:
                ids.add(int(row[0]))
            except Exception:
                continue
        self.missing_ids = ids

    def _replace_missing_rows_for_part(self, part_id: int, rows: list | None) -> set:
        try:
            pid = int(part_id)
        except Exception:
            return set()

        kept = []
        for row in (self.missing_files or []):
            try:
                if int(row[0]) == pid:
                    continue
            except Exception:
                pass
            kept.append(row)

        new_rows = []
        for row in (rows or []):
            try:
                if int(row[0]) == pid:
                    new_rows.append(row)
            except Exception:
                continue

        self.missing_files = kept + new_rows
        self._rebuild_missing_ids()

        try:
            if self._cached_missing_map is not None:
                if new_rows:
                    self._cached_missing_map[pid] = {str(row[1]) for row in new_rows}
                else:
                    self._cached_missing_map.pop(pid, None)
        except Exception:
            pass

        return {str(row[1]) for row in new_rows}

    def _invalidate_doc_indicator(self, part_id: int | None = None) -> None:
        try:
            if part_id is None:
                self._doc_indicator_cache.clear()
            else:
                self._doc_indicator_cache.pop(int(part_id), None)
        except Exception:
            self._doc_indicator_cache = {}

    def _refresh_diagnostic_for_part(self, part_id: int) -> set:
        self._invalidate_doc_indicator(int(part_id))
        rows = []
        try:
            if self.working_dir and os.path.isdir(self.working_dir):
                rows = self.diag_service.sync_bom_part_files(self.working_dir, int(part_id)) or []
        except Exception:
            rows = []
        return self._replace_missing_rows_for_part(int(part_id), rows)

    def _make_tree_item(self, info: dict) -> QTreeWidgetItem:
        item = QTreeWidgetItem([""] * BOM_TREE_COLUMN_COUNT)
        self._apply_tree_item_data(item, info or {})
        return item

    @staticmethod
    def _is_folder_tree_item(item: QTreeWidgetItem | None) -> bool:
        return bool(item is not None and item.data(0, BOM_TREE_FOLDER_ROLE))

    def _make_folder_tree_item(self, folder: dict) -> QTreeWidgetItem:
        item = QTreeWidgetItem(["", str(folder.get("name") or "Folder"), "", "", "", "", "", ""])
        item.setData(0, Qt.UserRole, None)
        item.setData(0, BOM_TREE_FOLDER_ROLE, int(folder["id"]))
        item.setData(0, BOM_TREE_PATH_ROLE, str(folder.get("_tree_path") or ""))
        item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, True)
        item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, True)
        item.setIcon(BOM_COL_NAME, _bom_type_icon("folder"))
        item.setToolTip(BOM_COL_NAME, "Organizational folder")
        folder_id = int(folder["id"])
        row_no = (getattr(self, "_bom_folder_path_rows", {}) or {}).get(str(folder.get("_tree_path") or ""), "")
        if not row_no:
            row_no = (getattr(self, "_bom_folder_row_numbers", {}) or {}).get(folder_id, "")
        item.setText(BOM_COL_ROW, str(row_no or ""))
        item.setTextAlignment(BOM_COL_ROW, Qt.AlignCenter)
        return item

    @staticmethod
    def _is_lazy_placeholder(item: QTreeWidgetItem | None) -> bool:
        return bool(item is not None and item.data(0, BOM_TREE_PLACEHOLDER_ROLE))

    def _ensure_lazy_placeholder(self, item: QTreeWidgetItem) -> None:
        if not item.data(0, BOM_TREE_IS_ASSEMBLY_ROLE):
            return
        if item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE):
            return
        if item.childCount() and self._is_lazy_placeholder(item.child(0)):
            return
        placeholder = QTreeWidgetItem(["", "Loading...", "", "", "", "", "", ""])
        placeholder.setData(0, BOM_TREE_PLACEHOLDER_ROLE, True)
        placeholder.setDisabled(True)
        item.addChild(placeholder)

    @staticmethod
    def _container_count(container) -> int:
        return container.topLevelItemCount() if isinstance(container, QTreeWidget) else container.childCount()

    @staticmethod
    def _container_item(container, index: int) -> QTreeWidgetItem:
        return container.topLevelItem(index) if isinstance(container, QTreeWidget) else container.child(index)

    @staticmethod
    def _container_take(container, index: int) -> QTreeWidgetItem:
        return container.takeTopLevelItem(index) if isinstance(container, QTreeWidget) else container.takeChild(index)

    @staticmethod
    def _container_add(container, item: QTreeWidgetItem) -> None:
        if isinstance(container, QTreeWidget):
            container.addTopLevelItem(item)
        else:
            container.addChild(item)

    def _take_direct_bom_item(self, container, part_id: int) -> QTreeWidgetItem | None:
        for index in range(self._container_count(container)):
            item = self._container_item(container, index)
            if self._is_folder_tree_item(item):
                continue
            try:
                if int(item.data(0, Qt.UserRole)) == int(part_id):
                    return self._container_take(container, index)
            except Exception:
                continue
        return None

    def _extract_folder_bom_items(self, folder_item: QTreeWidgetItem) -> list[QTreeWidgetItem]:
        extracted = []
        while folder_item.childCount():
            child = folder_item.takeChild(0)
            if self._is_folder_tree_item(child):
                extracted.extend(self._extract_folder_bom_items(child))
            else:
                extracted.append(child)
        return extracted

    def _clear_folder_nodes_from_container(self, container) -> None:
        index = self._container_count(container) - 1
        promoted = []
        while index >= 0:
            item = self._container_item(container, index)
            if self._is_folder_tree_item(item):
                folder_item = self._container_take(container, index)
                promoted.extend(self._extract_folder_bom_items(folder_item))
            index -= 1
        for item in promoted:
            self._container_add(container, item)

    def _folder_context_containers(self, parent_bom_id):
        if parent_bom_id is None:
            return [self.tree]
        return self._find_tree_items(int(parent_bom_id), self.tree)

    def _render_folder_context(self, parent_bom_id, folders: list[dict], containers=None) -> None:
        context_folders = [
            folder for folder in folders
            if folder.get("effective_parent_bom_id") == parent_bom_id
        ]
        if not context_folders:
            return
        children_by_folder = defaultdict(list)
        roots = []
        for folder in context_folders:
            parent_folder_id = folder.get("parent_folder_id")
            if parent_folder_id is None:
                roots.append(folder)
            else:
                children_by_folder[int(parent_folder_id)].append(folder)
        roots.sort(key=lambda row: (int(row.get("sort_order") or 0), int(row["id"])))
        for values in children_by_folder.values():
            values.sort(key=lambda row: (int(row.get("sort_order") or 0), int(row["id"])))

        def add_folder(folder: dict, display_parent, base_container) -> None:
            display_path = ""
            if not isinstance(display_parent, QTreeWidget):
                display_path = str(display_parent.data(0, BOM_TREE_PATH_ROLE) or "")
            folder_path = f"{display_path}/f{int(folder['id'])}" if display_path else f"f{int(folder['id'])}"
            folder_item = self._make_folder_tree_item(dict(folder, _tree_path=folder_path))
            self._container_add(display_parent, folder_item)
            assigned_ids = {int(value) for value in (folder.get("item_ids") or [])}
            ordered_assigned_ids = []
            search_container = display_parent
            for index in range(self._container_count(search_container)):
                candidate = self._container_item(search_container, index)
                if self._is_folder_tree_item(candidate):
                    continue
                try:
                    candidate_id = int(candidate.data(0, Qt.UserRole))
                except (TypeError, ValueError):
                    continue
                if candidate_id in assigned_ids:
                    ordered_assigned_ids.append(candidate_id)
            for part_id in ordered_assigned_ids:
                bom_item = self._take_direct_bom_item(search_container, int(part_id))
                if bom_item is not None:
                    folder_item.addChild(bom_item)
            for child_folder in children_by_folder.get(int(folder["id"]), []):
                add_folder(child_folder, folder_item, search_container)
            folder_item.setExpanded(False)

        base_containers = containers if containers is not None else self._folder_context_containers(parent_bom_id)
        for base_container in base_containers:
            for folder in roots:
                add_folder(folder, base_container, base_container)

    def _apply_organizational_folders(self) -> None:
        try:
            folders = self.bom_service.list_bom_folders() or []
        except Exception:
            folders = []
        self._bom_folders_cache = list(folders)
        contexts = {folder.get("effective_parent_bom_id") for folder in folders}
        for context in sorted(contexts, key=lambda value: (-1 if value is None else int(value))):
            self._render_folder_context(context, folders)

    def _refresh_folder_context(self, parent_bom_id) -> None:
        containers = self._folder_context_containers(parent_bom_id)
        for container in containers:
            self._clear_folder_nodes_from_container(container)
        try:
            folders = self.bom_service.list_bom_folders() or []
        except Exception:
            folders = []
        self._bom_folders_cache = list(folders)
        render_containers = containers
        if getattr(self, "_lazy_tree_active", False) and parent_bom_id is not None:
            render_containers = [
                container for container in containers
                if container.data(0, BOM_TREE_CHILDREN_LOADED_ROLE)
            ]
        if render_containers:
            self._render_folder_context(parent_bom_id, folders, containers=render_containers)
        self._renumber_full_bom_tree_rows()
        self._sync_search_tree_row_numbers()

    def _apply_tree_item_data(self, item: QTreeWidgetItem, info: dict) -> None:
        part_id = info.get("id")
        existing_path = str(item.data(0, BOM_TREE_PATH_ROLE) or "")
        existing_loaded = item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE)
        existing_is_assembly = bool(item.data(0, BOM_TREE_IS_ASSEMBLY_ROLE))
        issues = self._issues_for_part(part_id)
        locked_txt = ""
        if info.get("locked"):
            who = str(info.get("locked_by_username") or "").strip()
            locked_txt = f"In Work ({who})" if who else "In Work"

        row_text = str(info.get("_tree_row_text") or info.get("_tree_row_number") or "").strip()
        if not row_text:
            row_text = item.text(BOM_COL_ROW)
        item.setText(BOM_COL_ROW, row_text)
        item.setTextAlignment(BOM_COL_ROW, Qt.AlignCenter)
        is_ebom_item_row = (
            item.treeWidget() is getattr(self, "_ebom_tree", None)
            or "bom_id" in info
            or "version_label" in info
            or "source_quantity" in info
        )
        item.setText(
            BOM_COL_NAME,
            self._windchill_item_label(info)
            if is_ebom_item_row else str(info.get("name", "") or ""),
        )
        item.setData(0, BOM_TREE_INWORK_ROLE, locked_txt)
        has_lazy_metadata = "_has_children" in info
        has_structural_children = any(
            item.child(index).data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD
            for index in range(item.childCount())
        )
        is_asm = (
            bool(info.get("_has_children"))
            if has_lazy_metadata
            else existing_is_assembly or bool((info.get("children") or [])) or has_structural_children
        )
        item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, is_asm)
        item.setData(0, BOM_TREE_PATH_ROLE, str(info.get("_tree_path") or existing_path))
        loaded_state = not bool(info.get("_has_children")) if has_lazy_metadata else existing_loaded
        if loaded_state is None:
            loaded_state = True
        item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, bool(loaded_state))
        issue_summary = self._issue_summary_cache.get(int(part_id), {}) if part_id is not None else {}
        item.setData(0, BOM_TREE_ISSUE_ROLE, issue_summary)
        category_names = info.get("category_names", info.get("categories", []))
        if isinstance(category_names, str):
            category_names = [category_names] if category_names.strip() else []
        item.setData(0, BOM_TREE_CATEGORY_ROLE, list(category_names or []))
        binding_update_count = int(info.get("binding_update_count") or 0)
        item.setData(0, BOM_TREE_BINDING_UPDATE_ROLE, binding_update_count)
        item.setData(0, BOM_TREE_POLICY_ROLE, {
            "classification": str(info.get("classification") or "PHYSICAL"),
            "default_ebom_behavior": str(
                info.get("default_ebom_behavior") or "NORMAL"
            ),
            "ebom_behavior": str(info.get("ebom_behavior") or "INHERIT"),
            "resolved_ebom_behavior": str(
                info.get("resolved_ebom_behavior")
                or info.get("default_ebom_behavior")
                or "NORMAL"
            ),
            "represented_part_id": info.get("represented_part_id"),
            "cad_control_mode": str(info.get("cad_control_mode") or "CONTROLLED"),
        })
        item.setData(0, BOM_TREE_OCCURRENCE_ROLE, {
            "usage_id": info.get("usage_id"),
            "parent_id": info.get("relation_parent_id"),
            "quantity": info.get("quantity"),
        })
        item.setData(
            0, BOM_TREE_PROMOTION_ROLE, list(info.get("promoted_through") or [])
        )
        item.setText(BOM_COL_FILES, "")
        item.setData(0, BOM_TREE_ITEM_NUMBER_ROLE, str(info.get("part_number") or ""))
        item.setData(0, BOM_TREE_AES_NUMBER_ROLE, str(info.get("aes_number") or ""))
        visible_number = (
            info.get("part_number")
            if item.treeWidget() is getattr(self, "_ebom_tree", None)
            else info.get("aes_number")
        )
        item.setText(BOM_COL_AES, str(visible_number or ""))
        if item.treeWidget() is getattr(self, "_ebom_tree", None):
            item.setToolTip(
                BOM_COL_AES,
                "Item Number (PLM identity)"
                + (
                    f"\nAES delivery reference: {info.get('aes_number')}"
                    if info.get("aes_number")
                    else "\nAES delivery reference: not assigned"
                ),
            )
        item.setText(BOM_COL_TYPE, str(info.get("type", "") or ""))
        item.setText(BOM_COL_REV, str(info.get("current_version") or info.get("revision", "") or ""))
        item.setText(BOM_COL_STATUS, str(info.get("status", "") or ""))
        item.setText(BOM_COL_INTEGRITY, "")
        item.setData(0, Qt.UserRole, part_id)
        self._apply_extra_bom_columns(item, info)

        try:
            if info.get("_defer_indicators"):
                empty_summary = {
                    "pdf": {"state": "absent", "tooltip": "PDF: checking..."},
                    "step": {"state": "absent", "tooltip": "STEP: checking..."},
                }
                item.setData(BOM_COL_FILES, BOM_TREE_FILES_ROLE, _file_badges_payload(part_id, issues, empty_summary))
                item.setData(BOM_COL_INTEGRITY, BOM_TREE_INTEGRITY_ROLE, _integrity_payload(part_id, self.missing_files, self.missing_ids))
                item.setIcon(BOM_COL_NAME, _bom_type_icon(info.get("type")))
                self._indicator_refresh_queue.append((item, part_id))
                if not self._indicator_refresh_timer.isActive():
                    self._indicator_refresh_timer.start()
                return
            summary = self._indicator_summary_for_part(part_id, issues)
            item.setData(BOM_COL_FILES, BOM_TREE_FILES_ROLE, _file_badges_payload(part_id, issues, summary))
            item.setData(BOM_COL_INTEGRITY, BOM_TREE_INTEGRITY_ROLE, _integrity_payload(part_id, self.missing_files, self.missing_ids))
            item.setIcon(BOM_COL_NAME, _bom_type_icon(info.get("type")))
            tips0 = []
            if locked_txt:
                tips0.append(locked_txt)
            direct_active = int(issue_summary.get("direct_active_count", issue_summary.get("active_count") or 0) or 0)
            inherited_active = int(issue_summary.get("inherited_active_count") or 0)
            if direct_active and inherited_active:
                tips0.append(f"{direct_active} direct active issue(s); {inherited_active} inherited active issue(s) from children")
            elif direct_active:
                tips0.append(f"{direct_active} direct active issue(s) linked to this part")
            elif inherited_active:
                tips0.append(f"{inherited_active} inherited active issue(s) from child parts")
            elif int(issue_summary.get("total_count") or 0):
                tips0.append("All issues linked to this part are resolved")
            if binding_update_count:
                tips0.append(
                    f"{binding_update_count} direct child version update(s) available"
                )
            classification = str(info.get("classification") or "PHYSICAL").upper()
            occurrence_behavior = str(info.get("ebom_behavior") or "INHERIT").upper()
            resolved_behavior = str(
                info.get("resolved_ebom_behavior")
                or info.get("default_ebom_behavior")
                or "NORMAL"
            ).upper()
            tips0.append(
                f"EBOM policy: {classification}; occurrence {occurrence_behavior}; "
                f"effective {resolved_behavior}"
            )
            if str(info.get("cad_control_mode") or "CONTROLLED").upper() == "SUPPLIER_PACKAGE":
                tips0.append(
                    "Supplier-managed CAD package: internal owned CAD dependencies are not BOM items and are excluded from individual integrity checks."
                )
            if info.get("represented_part_id"):
                tips0.append(
                    f"CAD-only representation of physical BOM item #{info.get('represented_part_id')}. "
                    "It shares that item's AES number and has no drawing, PDF, or STEP delivery output."
                )
            elif resolved_behavior == "EXCLUDE":
                tips0.append(
                    "Not for delivery: this CAD occurrence and its descendants do not appear in Released EBOM."
                )
            elif resolved_behavior == "FLATTEN":
                tips0.append(
                    "CAD grouping only: this occurrence does not appear in Released EBOM; its children are promoted."
                )
            promoted_through = list(info.get("promoted_through") or [])
            if promoted_through:
                labels = " > ".join(
                    str(
                        value.get("part_number")
                        or value.get("name")
                        or value.get("aes_number")
                        or value.get("bom_id")
                    )
                    for value in promoted_through
                )
                tips0.append(
                    f"Promoted through flattened CAD structure: {labels}"
                )
            if not info.get("represented_part_id"):
                tips0.append(str(summary["pdf"].get("tooltip", "PDF: unknown")))
                tips0.append(str(summary["step"].get("tooltip", "STEP: unknown")))
            item.setToolTip(BOM_COL_NAME, "\n".join(tips0))
            item.setToolTip(BOM_COL_FILES, "\n".join([
                str(summary["pdf"].get("tooltip", "PDF: unknown")),
                str(summary["step"].get("tooltip", "STEP: unknown")),
            ]))
        except Exception:
            pass

    def _iter_tree_items(self, tree_widget: QTreeWidget):
        stack = []
        try:
            for i in range(tree_widget.topLevelItemCount()):
                stack.append(tree_widget.topLevelItem(i))
        except Exception:
            return
        while stack:
            item = stack.pop()
            yield item
            try:
                for i in range(item.childCount() - 1, -1, -1):
                    stack.append(item.child(i))
            except Exception:
                continue

    def _renumber_tree_rows(self, tree_widget: QTreeWidget | None, update_full_map: bool = False) -> None:
        if tree_widget is None:
            return
        row_map = defaultdict(list)
        row_no = 0

        def recurse(item: QTreeWidgetItem, ancestors_visible: bool = True):
            nonlocal row_no, row_map
            visible = True if update_full_map else (ancestors_visible and not item.isHidden())
            is_related_cad = item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD
            is_placeholder = self._is_lazy_placeholder(item)
            if visible and not is_related_cad and not is_placeholder:
                row_no += 1
                item.setText(BOM_COL_ROW, str(row_no))
                if update_full_map:
                    try:
                        row_map[int(item.data(0, Qt.UserRole))].append(row_no)
                    except Exception:
                        pass
            else:
                item.setText(BOM_COL_ROW, "")
            item.setTextAlignment(BOM_COL_ROW, Qt.AlignCenter)
            for idx in range(item.childCount()):
                recurse(item.child(idx), visible)

        try:
            for idx in range(tree_widget.topLevelItemCount()):
                recurse(tree_widget.topLevelItem(idx), True)
        except Exception:
            pass
        if update_full_map:
            self._bom_row_numbers = {pid: rows for pid, rows in row_map.items()}

    def _refresh_bom_row_numbers(self) -> None:
        self._renumber_full_bom_tree_rows()
        self._sync_search_tree_row_numbers()
        self._renumber_tree_rows(getattr(self, "_ebom_tree", None))

    def _renumber_full_bom_tree_rows(self) -> None:
        if getattr(self, "_lazy_tree_active", False):
            try:
                numbering = self.bom_service.get_bom_lazy_numbering(int(self.session.project_id))
                self._bom_row_numbers = dict(numbering.get("row_numbers") or {})
                self._bom_folder_row_numbers = dict(numbering.get("folder_rows") or {})
                self._bom_folder_path_rows = dict(numbering.get("folder_path_rows") or {})
                self._bom_folders_cache = list(numbering.get("folders") or [])
                path_rows = numbering.get("path_rows") or {}
                for item in self._iter_tree_items(self.tree):
                    if self._is_lazy_placeholder(item):
                        continue
                    folder_id = item.data(0, BOM_TREE_FOLDER_ROLE)
                    if folder_id:
                        path = str(item.data(0, BOM_TREE_PATH_ROLE) or "")
                        row_no = self._bom_folder_path_rows.get(path, self._bom_folder_row_numbers.get(int(folder_id), ""))
                    else:
                        path = str(item.data(0, BOM_TREE_PATH_ROLE) or "")
                        row_no = path_rows.get(path, "")
                        if not row_no:
                            try:
                                rows = self._bom_row_numbers.get(int(item.data(0, Qt.UserRole)), [])
                                row_no = rows[0] if rows else ""
                            except Exception:
                                row_no = ""
                    item.setText(BOM_COL_ROW, str(row_no or ""))
                    item.setTextAlignment(BOM_COL_ROW, Qt.AlignCenter)
                return
            except Exception:
                pass
        self._renumber_tree_rows(getattr(self, "tree", None), update_full_map=True)

    def _full_bom_row_text_for_part(self, part_id) -> str:
        try:
            rows = (getattr(self, "_bom_row_numbers", {}) or {}).get(int(part_id), [])
        except Exception:
            rows = []
        if not rows:
            return ""
        if len(rows) <= 3:
            return ", ".join(str(r) for r in rows)
        return ", ".join(str(r) for r in rows[:3]) + "+"

    def _sync_search_tree_row_numbers(self) -> None:
        search_tree = getattr(self, "_search_tree", None)
        if search_tree is None:
            return
        try:
            for item in self._iter_tree_items(search_tree):
                item.setText(BOM_COL_ROW, self._full_bom_row_text_for_part(item.data(0, Qt.UserRole)))
                item.setTextAlignment(BOM_COL_ROW, Qt.AlignCenter)
        except Exception:
            pass

    def _on_tree_item_entered(self, item: QTreeWidgetItem, column: int) -> None:
        """Show explicit tooltip when hovering a tree item (works around platform quirks)."""
        try:
            if item is None:
                return
            if column in (BOM_COL_FILES, BOM_COL_STATUS, BOM_COL_INTEGRITY):
                return
            tip = item.toolTip(column) or item.toolTip(BOM_COL_NAME) or ""
            if tip:
                tw = item.treeWidget()
                QToolTip.showText(QCursor.pos(), tip, tw or self.tree)
        except Exception:
            pass

    def _find_tree_items(self, part_id: int, tree_widget: QTreeWidget | None = None) -> list:
        try:
            pid = int(part_id)
        except Exception:
            return []
        widgets = [tree_widget] if tree_widget is not None else [
            getattr(self, "tree", None), getattr(self, "_search_tree", None),
            getattr(self, "_ebom_tree", None),
        ]
        matches = []
        for widget in widgets:
            if widget is None:
                continue
            for item in self._iter_tree_items(widget):
                try:
                    if item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                        continue
                    payload = item.data(0, PDM_NODE_PAYLOAD_ROLE) or {}
                    occurrence = item.data(0, BOM_TREE_OCCURRENCE_ROLE) or {}
                    candidate_values = (
                        item.data(0, Qt.UserRole),
                        payload.get("id") if isinstance(payload, dict) else None,
                        payload.get("item_id") if isinstance(payload, dict) else None,
                        occurrence.get("child_item_id") if isinstance(occurrence, dict) else None,
                    )
                    if any(value is not None and int(value) == pid for value in candidate_values):
                        matches.append(item)
                except Exception:
                    continue
        return matches

    def _add_node_to_tree(self, tree_widget: QTreeWidget, info: dict, parent_item: QTreeWidgetItem | None = None) -> QTreeWidgetItem:
        item = self._make_tree_item(info or {})
        if parent_item is None:
            tree_widget.addTopLevelItem(item)
        else:
            parent_item.addChild(item)
        for child in (info or {}).get("children", []) or []:
            self._add_node_to_tree(tree_widget, child, item)
        return item

    def _node_from_tree_item(self, item: QTreeWidgetItem) -> dict:
        part_id = item.data(0, Qt.UserRole)
        try:
            node = self.bom_service.get_part_details(int(part_id)) or {}
        except Exception:
            node = {}
        if not node:
            node = {
                "id": part_id,
                "name": item.text(BOM_COL_NAME),
                "part_number": item.data(0, BOM_TREE_ITEM_NUMBER_ROLE) or "",
                "aes_number": item.data(0, BOM_TREE_AES_NUMBER_ROLE) or item.text(BOM_COL_AES),
                "type": item.text(BOM_COL_TYPE),
                "revision": item.text(BOM_COL_REV),
                "status": item.text(BOM_COL_STATUS),
            }
        node = dict(node)
        node["children"] = []
        for i in range(item.childCount()):
            node["children"].append(self._node_from_tree_item(item.child(i)))
        return node

    def _part_matches_current_search(self, info: dict) -> bool:
        try:
            query = self.search_input.text().strip().lower()
        except Exception:
            query = ""
        if not query:
            return False
        for key in ("part_number", "name", "aes_number"):
            if query in str(info.get(key, "") or "").lower():
                return True
        return False

    def _select_part_item(self, part_id: int) -> None:
        target_tree = self._search_tree if getattr(self, "_in_search_mode", False) else self.tree
        matches = self._find_tree_items(part_id, target_tree)
        if not matches and target_tree is not self.tree:
            matches = self._find_tree_items(part_id, self.tree)
            target_tree = self.tree
        if not matches:
            return
        item = matches[0]
        try:
            target_tree.setCurrentItem(item)
            target_tree.scrollToItem(item)
        except Exception:
            pass

    def _add_part_to_tree(self, part_id: int) -> dict:
        self._refresh_diagnostic_for_part(int(part_id))
        info = self.bom_service.get_part_details(int(part_id)) or {}
        if not info:
            return {}
        info = dict(info)
        info["children"] = []
        if getattr(self, "_lazy_tree_active", False):
            try:
                numbering = self.bom_service.get_bom_lazy_numbering(int(self.session.project_id))
                path = str(int(part_id))
                info["_tree_path"] = path
                info["_tree_row_number"] = (numbering.get("path_rows") or {}).get(path, "")
                info["_has_children"] = False
                info["_defer_indicators"] = True
                self._bom_row_numbers = dict(numbering.get("row_numbers") or {})
            except Exception:
                pass

        try:
            if not self._find_tree_items(int(part_id), self.tree):
                self._add_node_to_tree(self.tree, info)
        except Exception:
            pass

        try:
            if getattr(self, "_in_search_mode", False) and self._part_matches_current_search(info):
                if not self._find_tree_items(int(part_id), self._search_tree):
                    self._add_node_to_tree(self._search_tree, info)
        except Exception:
            pass

        self._renumber_full_bom_tree_rows()
        self._sync_search_tree_row_numbers()
        self._select_part_item(int(part_id))
        return info

    def _child_exists_under_item(self, parent_item: QTreeWidgetItem, child_id: int) -> bool:
        for idx in range(parent_item.childCount()):
            child = parent_item.child(idx)
            if self._is_folder_tree_item(child) and self._child_exists_under_item(child, child_id):
                return True
            try:
                if int(child.data(0, Qt.UserRole)) == int(child_id):
                    return True
            except Exception:
                continue
        return False

    def _direct_child_items(self, parent_item: QTreeWidgetItem, child_id: int) -> list[QTreeWidgetItem]:
        matches = []
        for idx in range(parent_item.childCount()):
            child = parent_item.child(idx)
            if self._is_folder_tree_item(child):
                matches.extend(self._direct_child_items(child, child_id))
                continue
            try:
                if int(child.data(0, Qt.UserRole)) == int(child_id):
                    matches.append(child)
            except Exception:
                continue
        return matches

    def _add_child_relation_to_tree_widget(
        self,
        tree_widget: QTreeWidget,
        parent_id: int,
        child_node: dict,
    ) -> int:
        if tree_widget is None or not child_node:
            return 0
        added = 0
        child_id = int(child_node.get("id"))
        try:
            tree_widget.setUpdatesEnabled(False)
        except Exception:
            pass
        try:
            for parent_item in self._find_tree_items(int(parent_id), tree_widget):
                if (
                    tree_widget is self.tree
                    and getattr(self, "_lazy_tree_active", False)
                    and not parent_item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE)
                ):
                    # The database is already updated. Let the first expansion fetch
                    # the complete direct level once instead of inserting a duplicate.
                    continue
                if self._child_exists_under_item(parent_item, child_id):
                    continue
                node = dict(child_node)
                if (
                    tree_widget is self.tree
                    and getattr(self, "_lazy_tree_active", False)
                    and parent_item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE)
                ):
                    parent_path = str(parent_item.data(0, BOM_TREE_PATH_ROLE) or "")
                    child_path = f"{parent_path}/{child_id}" if parent_path else str(child_id)
                    try:
                        numbering = self.bom_service.get_bom_lazy_numbering(int(self.session.project_id))
                        node["_tree_path"] = child_path
                        node["_tree_row_number"] = (numbering.get("path_rows") or {}).get(child_path, "")
                        node["_has_children"] = child_id in (numbering.get("has_children") or set())
                    except Exception:
                        pass
                    node["_defer_indicators"] = True
                new_item = self._add_node_to_tree(tree_widget, node, parent_item)
                self._ensure_lazy_placeholder(new_item)
                parent_item.setExpanded(True)
                added += 1
        finally:
            try:
                tree_widget.setUpdatesEnabled(True)
                tree_widget.viewport().update()
            except Exception:
                pass
        return added

    def _add_child_relation_to_trees(self, parent_id: int, child_id: int, source_item: QTreeWidgetItem | None = None) -> None:
        if source_item is not None:
            child_node = self._node_from_tree_item(source_item)
        else:
            child_node = self.bom_service.get_part_details(int(child_id)) or {}
            child_node = dict(child_node)
            child_node["children"] = []
        if not child_node:
            return

        self._add_child_relation_to_tree_widget(self.tree, int(parent_id), child_node)
        try:
            if getattr(self, "_in_search_mode", False):
                if self._part_matches_current_search(child_node):
                    self._add_child_relation_to_tree_widget(self._search_tree, int(parent_id), child_node)
        except Exception:
            pass

        try:
            self._refresh_part_in_tree(int(parent_id))
            self._refresh_part_in_tree(int(child_id))
        except Exception:
            pass
        self._renumber_full_bom_tree_rows()
        self._sync_search_tree_row_numbers()

    def _remove_child_relation_from_tree_widget(self, tree_widget: QTreeWidget, parent_id: int, child_id: int) -> int:
        if tree_widget is None:
            return 0
        removed = 0
        try:
            tree_widget.setUpdatesEnabled(False)
        except Exception:
            pass
        try:
            for parent_item in self._find_tree_items(int(parent_id), tree_widget):
                for child_item in list(self._direct_child_items(parent_item, int(child_id))):
                    try:
                        visual_parent = child_item.parent()
                        if visual_parent is not None:
                            visual_parent.removeChild(child_item)
                        else:
                            index = tree_widget.indexOfTopLevelItem(child_item)
                            if index >= 0:
                                tree_widget.takeTopLevelItem(index)
                        removed += 1
                    except Exception:
                        pass
                if (
                    tree_widget is self.tree
                    and getattr(self, "_lazy_tree_active", False)
                    and parent_item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE)
                ):
                    has_engineering_child = False
                    for index in range(parent_item.childCount()):
                        candidate = parent_item.child(index)
                        if self._is_lazy_placeholder(candidate):
                            continue
                        if self._is_folder_tree_item(candidate):
                            if candidate.childCount() > 0:
                                has_engineering_child = True
                                break
                            continue
                        has_engineering_child = True
                        break
                    parent_item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, has_engineering_child)
                    if not has_engineering_child:
                        parent_item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, True)
        finally:
            try:
                tree_widget.setUpdatesEnabled(True)
                tree_widget.viewport().update()
            except Exception:
                pass
        return removed

    def _remove_child_relation_from_trees(self, parent_id: int, child_id: int) -> None:
        self._remove_child_relation_from_tree_widget(self.tree, int(parent_id), int(child_id))
        self._remove_child_relation_from_tree_widget(self._search_tree, int(parent_id), int(child_id))
        try:
            self._refresh_part_in_tree(int(parent_id))
            self._refresh_part_in_tree(int(child_id))
        except Exception:
            pass
        self._renumber_full_bom_tree_rows()
        self._sync_search_tree_row_numbers()

    def remove_selected_children(
        self,
        parent_id: int,
        parent_name: str,
        selected_items,
    ) -> None:
        child_ids = []
        child_names = []
        for item in selected_items or []:
            try:
                child_id = int(item.data(0, Qt.UserRole))
            except (TypeError, ValueError):
                continue
            if child_id in child_ids:
                continue
            child_ids.append(child_id)
            child_names.append(str(item.text(BOM_COL_NAME) or child_id))
        if not child_ids:
            return

        count = len(child_ids)
        preview = "\n".join(f"- {name}" for name in child_names[:10])
        if count > 10:
            preview += f"\n- ... and {count - 10} more"
        reply = QMessageBox.question(
            self,
            "Remove Child Relations",
            f"Remove {count} selected child{'ren' if count != 1 else ''} from {parent_name}?\n\n"
            f"{preview}\n\n"
            "This does not delete any BOM item. If an item has no other parent, it will be moved to the top level.",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        pending_paths = set(getattr(self, "_pending_relation_expanded_paths", set()) or set())
        pending_part_ids = set(getattr(self, "_pending_relation_expanded_part_ids", set()) or set())

        def capture(item: QTreeWidgetItem, selected_subtree: bool = False) -> None:
            if item is None or self._is_lazy_placeholder(item):
                return
            try:
                item_id = int(item.data(0, Qt.UserRole))
            except (TypeError, ValueError):
                item_id = None
            in_selected_subtree = selected_subtree or (item_id in child_ids if item_id is not None else False)
            if item.isExpanded():
                path = str(item.data(0, BOM_TREE_PATH_ROLE) or "")
                if path:
                    pending_paths.add(path)
                if in_selected_subtree and item_id is not None:
                    pending_part_ids.add(item_id)
            for index in range(item.childCount()):
                capture(item.child(index), in_selected_subtree)

        for parent_item in self._find_tree_items(int(parent_id), self.tree):
            for index in range(parent_item.childCount()):
                capture(parent_item.child(index))
        self._pending_relation_expanded_paths = pending_paths
        self._pending_relation_expanded_part_ids = pending_part_ids

        try:
            result = self.bom_service.remove_children_from_parent(int(parent_id), child_ids)
        except Exception as exc:
            QMessageBox.critical(self, "Remove Child", f"Failed to remove selected children:\n{exc}")
            return

        self._cancel_lazy_requests_for_parts({int(parent_id)})
        try:
            self._bom_folders_cache = list(self.bom_service.list_bom_folders() or [])
        except Exception:
            pass
        try:
            numbering = self.bom_service.get_bom_lazy_numbering(int(self.session.project_id))
            has_children = set(numbering.get("has_children") or set())
            path_rows = dict(numbering.get("path_rows") or {})
            self._bom_row_numbers = dict(numbering.get("row_numbers") or {})
            self._bom_folder_row_numbers = dict(numbering.get("folder_rows") or {})
            self._bom_folder_path_rows = dict(numbering.get("folder_path_rows") or {})
        except Exception:
            has_children = {int(parent_id)}
            path_rows = {}

        for parent_item in self._find_tree_items(int(parent_id), self.tree):
            parent_item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, int(parent_id) in has_children)
            parent_item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, False)
            parent_item.setData(0, BOM_TREE_REPLACE_CHILDREN_ROLE, True)
            self._load_lazy_children_for_item(parent_item, asynchronous=True, expand_after=False)

        self._renumber_full_bom_tree_rows()
        for child_id in result.get("moved_to_root_ids") or []:
            root_path = str(int(child_id))
            if any(
                str(existing.data(0, BOM_TREE_PATH_ROLE) or "") == root_path
                for existing in self._find_tree_items(int(child_id), self.tree)
            ):
                continue
            info = self.bom_service.get_part_details(int(child_id)) or {}
            if not info:
                continue
            info = dict(info)
            info["children"] = []
            info["_tree_path"] = root_path
            info["_tree_row_number"] = path_rows.get(root_path, "")
            info["_has_children"] = int(child_id) in has_children
            info["_defer_indicators"] = True
            root_item = self._make_tree_item(info)
            target_row = int(info.get("_tree_row_number") or 0)
            insert_at = self.tree.topLevelItemCount()
            for index in range(self.tree.topLevelItemCount()):
                try:
                    existing_row = int(self.tree.topLevelItem(index).text(BOM_COL_ROW) or 0)
                except (TypeError, ValueError):
                    existing_row = 0
                if existing_row and target_row and existing_row > target_row:
                    insert_at = index
                    break
            self.tree.insertTopLevelItem(insert_at, root_item)
            self._ensure_lazy_placeholder(root_item)
            if int(child_id) in self._pending_relation_expanded_part_ids:
                self._pending_relation_expanded_part_ids.discard(int(child_id))
                root_item.setExpanded(True)

        self._renumber_full_bom_tree_rows()
        self._sync_search_tree_row_numbers()
        try:
            self.tree.viewport().update()
            self.display_details(int(parent_id))
        except Exception:
            pass

        moved_items = result.get("moved_to_root_items") or []
        if moved_items:
            moved_names = "\n".join(
                f"- {self._item_identity_text(row)}" for row in moved_items
            )
            QMessageBox.warning(
                self,
                "Moved to Top Level",
                "The following items had no other parent and were moved to the top level. "
                "They were not deleted:\n\n" + moved_names,
            )
        else:
            QMessageBox.information(
                self,
                "Children Removed",
                f"{count} child relation{'s' if count != 1 else ''} removed. The items still exist elsewhere in the BOM.",
            )

    def _apply_child_relation_result(self, result: dict) -> None:
        """Refresh changed relation contexts without altering expansion state."""
        if not result:
            return
        parent_ids = {int(result.get("target_parent_id"))}
        parent_ids.update(int(value) for value in (result.get("source_parent_ids") or []))
        changed_child_ids = {int(value) for value in (result.get("child_ids") or [])}

        pending_paths = set(getattr(self, "_pending_relation_expanded_paths", set()) or set())
        pending_part_ids = set(getattr(self, "_pending_relation_expanded_part_ids", set()) or set())

        def capture_expanded(item: QTreeWidgetItem, moving_subtree: bool = False) -> None:
            if item is None or self._is_lazy_placeholder(item):
                return
            try:
                item_id = int(item.data(0, Qt.UserRole))
            except (TypeError, ValueError):
                item_id = None
            in_moving_subtree = moving_subtree or (item_id in changed_child_ids if item_id is not None else False)
            if item.isExpanded():
                path = str(item.data(0, BOM_TREE_PATH_ROLE) or "")
                if path:
                    pending_paths.add(path)
                if in_moving_subtree and item_id is not None:
                    pending_part_ids.add(item_id)
            for child_index in range(item.childCount()):
                capture_expanded(item.child(child_index), in_moving_subtree)

        for parent_id in parent_ids:
            for parent_item in self._find_tree_items(parent_id, self.tree):
                for child_index in range(parent_item.childCount()):
                    capture_expanded(parent_item.child(child_index))
        if result.get("had_root_sources"):
            for child_id in changed_child_ids:
                for root_item in self._find_tree_items(child_id, self.tree):
                    if str(root_item.data(0, BOM_TREE_PATH_ROLE) or "") == str(child_id):
                        capture_expanded(root_item, True)

        self._pending_relation_expanded_paths = pending_paths
        self._pending_relation_expanded_part_ids = pending_part_ids
        self._cancel_lazy_requests_for_parts(parent_ids)

        if result.get("had_root_sources"):
            for child_id in (result.get("child_ids") or []):
                for item in list(self._find_tree_items(int(child_id), self.tree)):
                    if str(item.data(0, BOM_TREE_PATH_ROLE) or "") != str(int(child_id)):
                        continue
                    visual_parent = item.parent()
                    if visual_parent is not None:
                        visual_parent.removeChild(item)
                    else:
                        index = self.tree.indexOfTopLevelItem(item)
                        if index >= 0:
                            self.tree.takeTopLevelItem(index)

        try:
            self._bom_folders_cache = list(self.bom_service.list_bom_folders() or [])
        except Exception:
            pass
        try:
            numbering = self.bom_service.get_bom_lazy_numbering(int(self.session.project_id))
            has_children = set(numbering.get("has_children") or set())
        except Exception:
            has_children = set(parent_ids)

        for parent_id in parent_ids:
            for item in self._find_tree_items(int(parent_id), self.tree):
                if self._is_folder_tree_item(item):
                    continue
                item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, parent_id in has_children)
                item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, False)
                item.setData(0, BOM_TREE_REPLACE_CHILDREN_ROLE, True)
                if parent_id in has_children and item.childCount() == 0:
                    self._ensure_lazy_placeholder(item)
                self._load_lazy_children_for_item(
                    item, asynchronous=True, expand_after=False
                )

        self._renumber_full_bom_tree_rows()
        self._sync_search_tree_row_numbers()
        try:
            self.tree.viewport().update()
        except Exception:
            pass

        message = str(result.get("message") or "BOM structure updated.")
        try:
            self.window().statusBar().showMessage(message)
        except Exception:
            pass

    def _restore_pending_relation_expansions(self, parent_item: QTreeWidgetItem) -> None:
        pending_paths = getattr(self, "_pending_relation_expanded_paths", set()) or set()
        pending_part_ids = getattr(self, "_pending_relation_expanded_part_ids", set()) or set()

        def restore(item: QTreeWidgetItem) -> None:
            if item is None or self._is_lazy_placeholder(item):
                return
            path = str(item.data(0, BOM_TREE_PATH_ROLE) or "")
            try:
                part_id = int(item.data(0, Qt.UserRole))
            except (TypeError, ValueError):
                part_id = None
            should_expand = path in pending_paths or (part_id is not None and part_id in pending_part_ids)
            if should_expand:
                pending_paths.discard(path)
                if part_id is not None:
                    pending_part_ids.discard(part_id)
                item.setExpanded(True)
            if self._is_folder_tree_item(item):
                for index in range(item.childCount()):
                    restore(item.child(index))

        for index in range(parent_item.childCount()):
            restore(parent_item.child(index))

    def _selected_sibling_reorder_context(self):
        selected = [item for item in self.tree.selectedItems() if item is not None]
        if not selected:
            item = self.tree.currentItem()
            selected = [item] if item is not None else []
        if not selected:
            raise ValueError("Select one or more BOM items to reorder.")

        visual_parent = selected[0].parent()
        if visual_parent is None:
            raise ValueError("Top-level BOM items cannot be reordered here because they are not child associations.")
        for item in selected:
            if item.parent() is not visual_parent:
                raise ValueError("Reorder works only for items under the same parent assembly.")
        engineering_parent = visual_parent
        while engineering_parent is not None and engineering_parent.data(0, Qt.UserRole) is None:
            engineering_parent = engineering_parent.parent()
        if engineering_parent is None or engineering_parent.data(0, Qt.UserRole) is None:
            raise ValueError("The engineering parent for this visual group was not found.")
        parent_id = int(engineering_parent.data(0, Qt.UserRole))
        selected_ids = [int(item.data(0, Qt.UserRole)) for item in selected]
        current_order = [int(value) for value in self.bom_service.ordered_child_ids(parent_id)]
        selected_set = set(selected_ids)
        selected_in_order = [cid for cid in current_order if cid in selected_set]
        if not selected_in_order:
            raise ValueError("No valid selected children to reorder.")
        return parent_id, current_order, selected_in_order

    def _apply_reordered_children(self, parent_id: int, ordered_child_ids: list[int]) -> None:
        self.bom_service.reorder_children(int(parent_id), [int(x) for x in ordered_child_ids])
        self._reorder_children_in_tree_widget(self.tree, int(parent_id), ordered_child_ids)
        self._reorder_children_in_tree_widget(self._search_tree, int(parent_id), ordered_child_ids)
        self._renumber_full_bom_tree_rows()
        self._sync_search_tree_row_numbers()
        try:
            self._refresh_part_in_tree(int(parent_id))
        except Exception:
            pass

    def _reorder_children_in_tree_widget(self, tree_widget: QTreeWidget, parent_id: int, ordered_child_ids: list[int]) -> None:
        if tree_widget is None:
            return
        order = [int(x) for x in ordered_child_ids]
        order_positions = {child_id: index for index, child_id in enumerate(order)}
        try:
            tree_widget.setUpdatesEnabled(False)
        except Exception:
            pass
        try:
            for parent_item in self._find_tree_items(int(parent_id), tree_widget):
                def sort_visual_container(container: QTreeWidgetItem) -> None:
                    children = [container.takeChild(0) for _ in range(container.childCount())]
                    sortable_positions = []
                    sortable_items = []
                    for index, child in enumerate(children):
                        try:
                            child_id = int(child.data(0, Qt.UserRole))
                        except (TypeError, ValueError):
                            child_id = None
                        if child_id in order_positions:
                            sortable_positions.append(index)
                            sortable_items.append(child)
                    sortable_items.sort(
                        key=lambda child: order_positions[int(child.data(0, Qt.UserRole))]
                    )
                    for index, child in zip(sortable_positions, sortable_items):
                        children[index] = child
                    for child in children:
                        container.addChild(child)
                    for child in children:
                        if self._is_folder_tree_item(child):
                            sort_visual_container(child)

                sort_visual_container(parent_item)
        finally:
            try:
                tree_widget.setUpdatesEnabled(True)
                tree_widget.viewport().update()
            except Exception:
                pass

    def _reorder_selected_siblings(self, mode: str) -> None:
        if not self.perm.can("manage_parts"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to reorder BOM items.")
        try:
            parent_id, current_order, selected_in_order = self._selected_sibling_reorder_context()
            selected_set = set(selected_in_order)
            remaining = [cid for cid in current_order if cid not in selected_set]

            if mode == "top":
                new_order = selected_in_order + remaining
            elif mode == "bottom":
                new_order = remaining + selected_in_order
            elif mode == "up":
                new_order = list(current_order)
                selected_positions = [i for i, cid in enumerate(new_order) if cid in selected_set]
                if selected_positions and selected_positions[0] > 0:
                    before_idx = selected_positions[0] - 1
                    before = new_order.pop(before_idx)
                    insert_at = selected_positions[-1]
                    new_order.insert(insert_at, before)
            elif mode == "down":
                new_order = list(current_order)
                selected_positions = [i for i, cid in enumerate(new_order) if cid in selected_set]
                if selected_positions and selected_positions[-1] < len(new_order) - 1:
                    after_idx = selected_positions[-1] + 1
                    after = new_order.pop(after_idx)
                    insert_at = selected_positions[0]
                    new_order.insert(insert_at, after)
            elif mode == "position":
                pos, ok = QInputDialog.getInt(
                    self,
                    "Move To Position",
                    f"New position for selected item(s), 1 to {len(current_order)}:",
                    1,
                    1,
                    max(1, len(current_order)),
                    1,
                )
                if not ok:
                    return
                insert_at = max(0, min(int(pos) - 1, len(remaining)))
                new_order = remaining[:insert_at] + selected_in_order + remaining[insert_at:]
            else:
                return

            if new_order == current_order:
                return
            self._apply_reordered_children(parent_id, new_order)
            self.window().statusBar().showMessage("BOM child order updated.")
        except Exception as e:
            QMessageBox.warning(self, "Reorder", str(e))

    def _handle_tree_drag_reorder(self, selected_ids: list, target_id: int, target_parent_id: int, where: str) -> None:
        if not self.perm.can("manage_parts"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to reorder BOM items.")
        try:
            if int(target_parent_id) < 0:
                raise ValueError("Top-level BOM items cannot be reordered here.")
            selected_set = {int(x) for x in selected_ids or []}
            if int(target_id) in selected_set:
                return
            current_order = [
                int(value) for value in self.bom_service.ordered_child_ids(int(target_parent_id))
            ]
            if int(target_id) not in current_order or not selected_set.issubset(set(current_order)):
                raise ValueError("Drag reorder works only between siblings under the same parent assembly.")
            selected_in_order = [cid for cid in current_order if cid in selected_set]
            remaining = [cid for cid in current_order if cid not in selected_set]
            target_index = remaining.index(int(target_id))
            if where == "below":
                target_index += 1
            new_order = remaining[:target_index] + selected_in_order + remaining[target_index:]
            if new_order == current_order:
                return
            self._apply_reordered_children(int(target_parent_id), new_order)
            self.window().statusBar().showMessage("BOM child order updated.")
        except Exception as e:
            QMessageBox.warning(self, "Reorder", str(e))

    def _refresh_part_in_tree(self, part_id: int) -> dict:
        self._refresh_diagnostic_for_part(int(part_id))
        info = self.bom_service.get_part_details(int(part_id)) or {}
        if not info:
            return {}
        for item in self._find_tree_items(int(part_id)):
            self._apply_tree_item_data(item, info)
            if item.treeWidget() is getattr(self, "_ebom_tree", None):
                item.setData(0, PDM_OBJECT_KIND_ROLE, PDM_OBJECT_ITEM)
                item.setIcon(BOM_COL_NAME, _pdm_item_icon())
        try:
            if getattr(self, "_in_search_mode", False) and self._part_matches_current_search(info):
                if not self._find_tree_items(int(part_id), self._search_tree):
                    self._add_node_to_tree(self._search_tree, dict(info, children=[]))
        except Exception:
            pass
        try:
            if not self._is_default_bom_advanced_filter():
                self.apply_bom_tree_filter(self._bom_advanced_filters)
        except Exception:
            pass
        for tree in (
            getattr(self, "tree", None), getattr(self, "_search_tree", None),
            getattr(self, "_ebom_tree", None),
        ):
            try:
                tree.viewport().update()
            except Exception:
                pass
        return info

    def refresh_parts_after_merge(self, part_ids) -> None:
        """Refresh only BOM rows affected by a successful merge."""
        refreshed = set()
        requested = list(part_ids or [])
        try:
            requested.extend(self.bom_service.direct_parent_ids(requested))
        except Exception:
            pass
        for part_id in requested:
            try:
                pid = int(part_id)
            except Exception:
                continue
            if pid in refreshed:
                continue
            refreshed.add(pid)
            self._refresh_part_in_tree(pid)

        try:
            current = getattr(self, "current_part_id", None)
            if current is not None and int(current) in refreshed:
                self.display_details(int(current))
        except Exception:
            pass

    def _find_pdm_cad_items(self, cad_document_ids) -> list[QTreeWidgetItem]:
        wanted = set()
        for cad_id in cad_document_ids or []:
            try:
                wanted.add(int(cad_id))
            except Exception:
                continue
        if not wanted:
            return []
        matches = []
        for tree in (getattr(self, "_cad_tree", None), getattr(self, "_ebom_tree", None)):
            if tree is None:
                continue
            for item in self._iter_tree_items(tree):
                if self._is_lazy_placeholder(item):
                    continue
                if item.data(0, PDM_OBJECT_KIND_ROLE) != PDM_OBJECT_CAD:
                    continue
                try:
                    cad_id = int(item.data(0, PDM_CAD_DOCUMENT_ID_ROLE))
                except Exception:
                    continue
                if cad_id in wanted:
                    matches.append(item)
        return matches

    def refresh_cad_documents_after_merge(self, cad_document_ids) -> None:
        """Refresh only loaded CAD rows affected by a successful merge/check-in."""
        requested = set()
        for cad_id in cad_document_ids or []:
            try:
                requested.add(int(cad_id))
            except Exception:
                continue
        if not requested:
            return
        try:
            documents = {}
            for document in self.bom_service.list_pdm_cad_documents() or []:
                try:
                    cad_id = int(document.get("id"))
                except Exception:
                    continue
                if cad_id in requested:
                    documents[cad_id] = document
        except Exception:
            documents = {}
        if not documents:
            return

        touched_trees = set()
        for item in self._find_pdm_cad_items(requested):
            try:
                cad_id = int(item.data(0, PDM_CAD_DOCUMENT_ID_ROLE))
            except Exception:
                continue
            document = documents.get(cad_id)
            if not document:
                continue
            tree = item.treeWidget()
            if tree is getattr(self, "_cad_tree", None):
                self._apply_pdm_cad_tree_item_data(item, document)
            elif tree is getattr(self, "_ebom_tree", None):
                self._apply_ebom_cad_tree_item_data(item, document)
            if tree is not None:
                touched_trees.add(tree)

        try:
            associations = getattr(self, "_ebom_associations_by_item", {})
            for rows in associations.values():
                for index, row in enumerate(list(rows or [])):
                    try:
                        cad_id = int(row.get("id"))
                    except Exception:
                        continue
                    if cad_id in documents:
                        updated = dict(row)
                        updated.update(documents[cad_id])
                        rows[index] = updated
        except Exception:
            pass

        for tree in touched_trees:
            try:
                tree.viewport().update()
            except Exception:
                pass
        try:
            current_cad = getattr(self, "current_cad_document_id", None)
            if current_cad is not None and int(current_cad) in requested:
                selected = self._current_tree_for_filtering().currentItem()
                if selected and selected.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
                    self._show_pdm_cad_selection(selected)
        except Exception:
            pass
        self._sync_visual_action_states()

    def refresh_after_pdm_merge(self, part_ids=None, cad_document_ids=None) -> None:
        """Refresh only affected Item and CAD Document rows after push to master."""
        if part_ids:
            self.refresh_parts_after_merge(part_ids)
        if cad_document_ids:
            self.refresh_cad_documents_after_merge(cad_document_ids)

    def refresh_issue_indicators(self, affected_part_ids=None) -> None:
        """Repaint only BOM nodes whose propagated issue summary changed."""
        try:
            previous = dict(self._issue_summary_cache or {})
            current = self.issue_service.part_summary()
        except Exception:
            return

        candidate_ids = set(previous) | set(current)
        candidate_ids.update(
            int(part_id) for part_id in (affected_part_ids or []) if part_id is not None
        )
        changed_ids = {
            int(part_id)
            for part_id in candidate_ids
            if previous.get(int(part_id), {}) != current.get(int(part_id), {})
        }
        self._issue_summary_cache = current

        for part_id in changed_ids:
            try:
                info = self.bom_service.get_part_details(int(part_id)) or {}
                if not info:
                    continue
                for item in self._find_tree_items(int(part_id)):
                    self._apply_tree_item_data(item, info)
                    if item.treeWidget() is getattr(self, "_ebom_tree", None):
                        item.setData(0, PDM_OBJECT_KIND_ROLE, PDM_OBJECT_ITEM)
                        item.setIcon(BOM_COL_NAME, _pdm_item_icon())
            except Exception:
                continue

        try:
            score = max(
                0,
                self.issue_service.health_score()
                - len(getattr(self, "missing_ids", set()) or set()),
            )
            color = "#2e7d32" if score >= 85 else ("#a16207" if score >= 65 else "#b91c1c")
            if getattr(self, "_bom_mode", "cad") == "ebom":
                self.bom_health_label.setText(f"Item health: {score}/100")
                self.bom_health_label.setStyleSheet(
                    f"font-size:11px;font-weight:700;color:{color};background:transparent;border:none;"
                )
        except Exception:
            pass

        try:
            current_part_id = getattr(self, "current_part_id", None)
            if current_part_id is not None and int(current_part_id) in changed_ids:
                self.refresh_issues_tab()
        except Exception:
            pass

    def _remove_part_from_tree_widget(self, tree_widget: QTreeWidget, part_id: int, promote_children: bool = False) -> None:
        matches = self._find_tree_items(int(part_id), tree_widget)
        promoted = {}
        if promote_children:
            for item in matches:
                for i in range(item.childCount()):
                    node = self._node_from_tree_item(item.child(i))
                    try:
                        child_id = int(node.get("id"))
                    except Exception:
                        continue
                    if child_id != int(part_id):
                        promoted.setdefault(child_id, node)

        for item in list(matches):
            parent = item.parent()
            try:
                if parent is not None:
                    parent.removeChild(item)
                else:
                    index = tree_widget.indexOfTopLevelItem(item)
                    if index >= 0:
                        tree_widget.takeTopLevelItem(index)
            except Exception:
                continue

        if promote_children:
            for child_id, node in promoted.items():
                if self._find_tree_items(child_id, tree_widget):
                    continue
                self._add_node_to_tree(tree_widget, node)

    def _remove_part_from_trees(self, part_id: int) -> None:
        self._replace_missing_rows_for_part(int(part_id), [])
        try:
            self._remove_part_from_tree_widget(self.tree, int(part_id), promote_children=True)
        except Exception:
            pass
        try:
            self._remove_part_from_tree_widget(self._ebom_tree, int(part_id), promote_children=False)
        except Exception:
            pass
        try:
            self._remove_part_from_tree_widget(self._search_tree, int(part_id), promote_children=False)
        except Exception:
            pass
        self._renumber_full_bom_tree_rows()
        self._sync_search_tree_row_numbers()

    def add_part(self):
        if not self.perm.can("manage_parts"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to create Items.")
        dialog = PartDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            part_data = dialog.get_data()
            if not part_data["name"]:
                QMessageBox.warning(self, "Validation Error", "Name is required.")
                return
            try:
                added= self.bom_service.add_part(part_data)
                if type(added) == int:
                    self._add_part_to_tree(int(added))
                    if str(self.bom_mode_selector.currentData() or "") == "ebom":
                        self._select_item_in_ebom(int(added))
                    self.display_details(int(added))
                    created = self.bom_service.get_part_details(int(added)) or {}
                    number = created.get("part_number") or added
                    self.window().statusBar().showMessage(f"Item {number} created.", 6000)
                else:
                    QMessageBox.warning(self, "New Item", "An unexpected error occurred while creating the Item.")
                    self.window().statusBar().showMessage("Item creation failed.")
            except Exception as e:
                QMessageBox.critical(self, "New Item", f"Could not create Item:\n{str(e)}")

    def edit_part(self, id=None):
        if not self.perm.can("manage_parts"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to edit Items.")
        if not id:
            if not self.current_part_id:
                QMessageBox.warning(self, "No Selection", "Select an Item to edit.")
                return
            id = self.current_part_id

        part_data = self.bom_service.get_part_details(id)
        if not part_data:
            QMessageBox.warning(self, "Not Found", f"Item {id} was not found.")
            return

        dialog = PartDialog(self, part_data)
        if dialog.exec_() == QDialog.Accepted:
            updated_data = dialog.get_data()
            if not updated_data["name"]:
                QMessageBox.warning(self, "Validation Error", "Name is required.")
                return
            try:
                undo_record = self.undo_service.snapshot_item_update(
                    int(id), f"Edit Item {id}"
                )
                self.bom_service.update_part(id, updated_data)
                self.undo_service.push(undo_record)
                self._refresh_part_in_tree(int(id))
                self._refresh_ebom_association_rows_for_item(int(id))
                number = updated_data.get("part_number") or id
                self.window().statusBar().showMessage(
                    f"Item {number} attributes updated. Press Ctrl+Z to undo.", 6000
                )
                self.display_details(id)
            except Exception as e:
                QMessageBox.critical(self, "Edit Attributes", f"Could not update Item:\n{str(e)}")

    def delete_part(self, id=None):
        if not self.perm.can("manage_parts"):
            return QMessageBox.warning(self, "Permission", "You do not have permission to delete Items.")
        if not id:
            if not self.current_part_id:
                QMessageBox.warning(self, "No Selection", "Select an Item to delete.")
                return
            id = self.current_part_id

        # Run the checkout-only preflight before asking for irreversible
        # confirmation.  The service repeats it during deletion for safety.
        try:
            self.bom_service.assert_item_fully_checked_in_for_delete(int(id))
        except Exception as exc:
            QMessageBox.warning(
                self,
                "Delete Blocked",
                str(exc),
            )
            return

        reply = QMessageBox.question(
            self,
            "Delete Item",
            f"Delete Item {id}?\n\nYou can press Ctrl+Z immediately after deletion to restore it.",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                undo_record = self.undo_service.snapshot_item_delete(
                    int(id), f"Delete Item {id}"
                )
                try:
                    parent_ids = list(self.bom_service.direct_parent_ids([int(id)]) or [])
                except Exception:
                    parent_ids = []
                self.bom_service.delete_part(id)
                self.undo_service.push(undo_record)
                self._remove_part_from_trees(int(id))
                for parent_id in parent_ids:
                    self._refresh_loaded_part_branch(int(parent_id))
                self._remove_part_from_trees(int(id))
                self.window().statusBar().showMessage("Item deleted. Press Ctrl+Z to undo.", 6000)
                self.clear_details()
            except Exception as e:
                QMessageBox.critical(self, "Delete Item", f"Could not delete Item:\n{str(e)}")

    def _refresh_current_tree_item_lock_state(self, part_id: int | None = None):
        """Refresh the selected part row in every visible BOM tree representation."""
        try:
            pid = part_id or getattr(self, "current_part_id", None)
            if not pid:
                current_tree = self._current_tree_for_filtering()
                item = current_tree.currentItem() if current_tree else None
                pid = item.data(0, Qt.UserRole) if item else None
            if not pid:
                return
            self._refresh_part_in_tree(int(pid))
        except Exception:
            pass

    def _refresh_loaded_part_branch(self, part_id: int) -> None:
        self._refresh_part_in_tree(int(part_id))
        try:
            self._refresh_pdm_ebom_structure_branch(int(part_id))
        except Exception:
            pass
        if not getattr(self, "_lazy_tree_active", False):
            return
        for item in list(self._find_tree_items(int(part_id), self.tree)):
            if not item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE):
                continue
            parent_path = str(item.data(0, BOM_TREE_PATH_ROLE) or "")
            expanded = bool(item.isExpanded())
            nodes = self.bom_service.get_bom_lazy_children(
                int(self.session.project_id), int(part_id), parent_path
            ) or []
            item.setData(0, BOM_TREE_REPLACE_CHILDREN_ROLE, True)
            self._apply_lazy_children_result(item, int(part_id), nodes, expand_after=False)
            item.setExpanded(expanded and bool(nodes))

    def _refresh_lock_family_rows(self, part_id: int, refresh_loaded_branches: bool = False):
        try:
            part_ids = self.bom_service.part_ids_sharing_base_file(int(part_id)) or [int(part_id)]
        except Exception:
            part_ids = [int(part_id)]
        for pid in part_ids:
            if refresh_loaded_branches:
                self._refresh_loaded_part_branch(int(pid))
            else:
                self._refresh_current_tree_item_lock_state(int(pid))
            self._invalidate_doc_indicator(int(pid))
        if refresh_loaded_branches:
            self._renumber_full_bom_tree_rows()
            self._sync_search_tree_row_numbers()
        return part_ids

    def checkin_part(self, part_id=None):
        if isinstance(part_id, bool):
            part_id = None
        if part_id is None and getattr(self, "current_cad_document_id", None) is not None:
            self._checkin_pdm_cad_document(
                int(self.current_cad_document_id), getattr(self, "_current_cad_payload", {})
            )
            return
        part_id = part_id or getattr(self, "current_part_id", None)
        if not part_id:
            QMessageBox.warning(self, "Check In", "Select a checked-out BOM item.")
            return
        try:
            analysis = self.bom_service.analyze_item_checkout(int(part_id))
        except (ValueError, PermissionError) as exc:
            QMessageBox.warning(self, "Check In", str(exc))
            return
        except Exception as exc:
            QMessageBox.critical(self, "Check In", f"Could not analyze the checkout:\n{exc}")
            return

        dialog = CheckoutReviewDialog(analysis, self)
        if dialog.exec_() != QDialog.Accepted:
            return
        if dialog.action == CheckoutReviewDialog.ACTION_COMMIT:
            main_window = self.window()
            commit_page = getattr(main_window, "commit_page", None)
            if commit_page is None or not hasattr(commit_page, "prepare_checkin_from_bom"):
                QMessageBox.warning(self, "Check In", "The Commit page is not available.")
                return
            commit_page.prepare_checkin_from_bom(analysis)
            if hasattr(main_window, "switch_page"):
                main_window.switch_page(1)
            return
        if dialog.action != CheckoutReviewDialog.ACTION_CHECKIN:
            return

        try:
            result = self.bom_service.checkin_item_data(
                int(part_id), dialog.comment()
            )
            context = (result or {}).get("context") or {}
            created_iteration = bool(
                ((result or {}).get("analysis") or analysis or {}).get("has_non_cad_changes")
            )
            QMessageBox.information(
                self,
                "Check In",
                (
                    f"Checkout completed as {context.get('version_label') or analysis.get('next_version')}.\n"
                    "The Item iteration was created independently; associated CAD Documents were not checked in."
                    if created_iteration else
                    f"Checkout closed at {context.get('version_label') or analysis.get('current_version') or '-'}.\n"
                    "No Item data changes were detected, so no new Item iteration was created. "
                    "Associated CAD Documents were not checked in."
                ),
            )
            for affected_id in (result or {}).get("affected_part_ids") or [int(part_id)]:
                self._refresh_loaded_part_branch(int(affected_id))
                self._invalidate_doc_indicator(int(affected_id))
            self._renumber_full_bom_tree_rows()
            self._sync_search_tree_row_numbers()
            self.display_details(int(part_id))
        except (ValueError, PermissionError) as exc:
            QMessageBox.warning(self, "Check In", str(exc))
        except Exception as exc:
            QMessageBox.critical(self, "Check In", f"Check-in failed:\n{exc}")

    def undo_checkout(self, part_id=None):
        if isinstance(part_id, bool):
            part_id = None
        if part_id is None and getattr(self, "current_cad_document_id", None) is not None:
            self._undo_pdm_cad_checkout(int(self.current_cad_document_id))
            return
        if not part_id:
            if not self.current_part_id:
                QMessageBox.warning(self, "No Selection", "Please select a checked-out item.")
                return
            part_id = self.current_part_id
        details = self.bom_service.get_part_details(int(part_id)) or {}
        actor_user_id = self._prompt_actor_user_id(
            "Undo Item Checkout As",
            "Undo this Item checkout as:",
            default_user_id=details.get("locked_by_user_id"),
        )
        if actor_user_id is None:
            return
        answer = QMessageBox.question(
            self,
            "Undo Checkout",
            "Discard the working Item attributes/structure and release this checkout?\n\n"
            "Any associated CAD checkout owned by the same acting user will also be undone. "
            "CAD/Item associations will not be removed.\n\n"
            "This does not create an iteration and is not a check-in.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return
        try:
            self.bom_service.undo_item_checkout(int(part_id), as_user_id=actor_user_id)
            QMessageBox.information(self, "Undo Checkout", "Checkout undone. No iteration was created.")
            affected_cad_ids = []
            try:
                affected_cad_ids = [
                    int(row.get("id"))
                    for row in (self.bom_service.list_item_cad_associations(int(part_id)) or [])
                    if row.get("id") is not None
                ]
            except Exception:
                affected_cad_ids = []
            self._refresh_pdm_context_rows(
                item_ids=[int(part_id)],
                cad_ids=affected_cad_ids,
            )
            try:
                self.display_details(int(part_id))
            except Exception:
                pass
        except ValueError as e:
            QMessageBox.warning(self, "Undo Checkout", str(e))
        except PermissionError as e:
            QMessageBox.warning(self, "Permission", str(e))
        except Exception as e:
            QMessageBox.critical(self, "Undo Checkout", f"Failed to undo checkout: {str(e)}")

    def checkout_part(self, part_id=None):
        if isinstance(part_id, bool):
            part_id = None
        if part_id is None and getattr(self, "current_cad_document_id", None) is not None:
            self._checkout_pdm_cad_document(
                int(self.current_cad_document_id), getattr(self, "_current_cad_payload", {})
            )
            return
        if not part_id:
            if not self.current_part_id:
                QMessageBox.warning(self, "No Selection", "Please select a part to check out.")
                return
            part_id = self.current_part_id

        details = self.bom_service.get_part_details(int(part_id)) or {}
        checkout_origin = str(details.get("checkout_origin") or "ITEM").upper()
        lock_owner_id = details.get("locked_by_user_id")
        auto_checkout_is_mine = bool(
            details.get("locked") and checkout_origin == "CAD"
            and (
                lock_owner_id is None or self.session.user_id is None
                or int(lock_owner_id) == int(self.session.user_id)
                or self.perm.can("merge")
            )
        )
        if auto_checkout_is_mine:
            if QMessageBox.question(
                self,
                "Make Item Checkout Explicit",
                "Keep this automatically reserved Item as an explicit Item checkout?\n\n"
                "The active associated CAD Document remains checked out and must still be "
                "checked in or undone separately.",
                QMessageBox.Yes | QMessageBox.Cancel,
                QMessageBox.Yes,
            ) != QMessageBox.Yes:
                return
            try:
                self.bom_service.checkout_item(int(part_id))
            except (ValueError, PermissionError) as exc:
                QMessageBox.warning(self, "Item Checkout", str(exc))
                return
            QMessageBox.information(
                self,
                "Item Checkout",
                "The Item checkout is now explicit. The associated CAD checkout remains active.",
            )
            self._refresh_pdm_context_rows(item_ids=[int(part_id)])
            self.display_details(int(part_id))
            return

        associated_cad_rows = [
            row for row in (self.bom_service.list_item_cad_associations(int(part_id)) or [])
            if str(row.get("category") or "").upper() != "DRAWING"
            and row.get("id") is not None
        ]
        owner_cad = next(
            (
                row for row in associated_cad_rows
                if str(row.get("association_type") or "").upper() == "OWNER"
            ),
            None,
        )
        selected_cad_rows = []
        workspace = None
        workspace_descriptor = {}
        if associated_cad_rows:
            selected_cad_ids = self._select_checked_rows_dialog(
                "Select CAD Documents",
                "Select associated CAD Documents to check out with this Item. "
                "Leave all unchecked for Item-only checkout.",
                associated_cad_rows,
                lambda row: (
                    f"{row.get('file_name') or row.get('name') or row.get('id')}  "
                    f"[{row.get('association_type') or 'ASSOCIATED'}]"
                ),
                checked_ids=[
                    int(owner_cad["id"])
                ] if owner_cad and owner_cad.get("id") is not None else [-1],
            )
            if selected_cad_ids is None:
                return
            selected_set = {int(value) for value in selected_cad_ids}
            selected_cad_rows = [
                row for row in associated_cad_rows
                if int(row.get("id")) in selected_set
            ]
            if selected_cad_rows:
                copy_to_workspace, workspace = self._prompt_cad_workspace_copy(
                    "Item and CAD Checkout Workspace"
                )
                if workspace is None:
                    return
                if copy_to_workspace:
                    workspace_descriptor = self._local_cad_workspaces().checkout_descriptor(
                        workspace["id"]
                    )
        state = str(
            details.get("revision_state") or details.get("lifecycle_state") or ""
        ).strip().lower()
        released_revision_code = None
        if state == "released":
            current_version = str(details.get("current_version") or details.get("revision") or "")
            try:
                suggested_revision = self.bom_service.suggest_next_revision(int(part_id))
            except Exception:
                suggested_revision = ""
            released_revision_code, ok = QInputDialog.getText(
                self,
                "Check Out Released Item",
                f"{current_version} is Released and will remain immutable.\n"
                "The next completed commit (check-in) will create revision:",
                QLineEdit.Normal,
                suggested_revision,
            )
            if not ok:
                return
            released_revision_code = str(released_revision_code or "").strip()
            if not released_revision_code:
                QMessageBox.warning(self, "Check Out", "Enter the revision to create on commit.")
                return

        as_user_id = self._prompt_actor_user_id(
            "Check Out As",
            "Check out this Item as:",
            default_user_id=lock_owner_id,
        )
        if as_user_id is None:
            return

        confirmation = (
            f"Check out Released {details.get('current_version') or details.get('revision')} for work?\n\n"
            f"The Released iteration will not change. The next completed commit will create "
            f"{released_revision_code}.1."
            if released_revision_code else
            f"Check out Item {part_id}?\n\n"
            + (
                f"The Item and {len(selected_cad_rows)} associated CAD Document"
                f"{'s' if len(selected_cad_rows) != 1 else ''} will be reserved "
                + (
                    f"in workspace {workspace.get('name')}."
                    if workspace else
                    "without copying CAD files to a workspace."
                )
                if selected_cad_rows else
                "Only Item metadata and structure will be reserved; CAD remains checked in."
            )
        )
        reply = QMessageBox.question(
            self,
            "Confirm Check Out",
            confirmation,
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            revised_cads = []
            materialized_paths = []
            try:
                selected_item_revision_codes = {int(part_id): released_revision_code} if released_revision_code else {}
                related_item_rows_by_id = {}
                for cad_row in selected_cad_rows:
                    for related_row in self._associated_item_rows_for_cad_checkout(
                        int(cad_row["id"]), cad_row
                    ):
                        related_item_rows_by_id[int(related_row["id"])] = related_row
                extra_revision_codes = self._prompt_released_item_revision_codes(
                    [
                        row for item_key, row in related_item_rows_by_id.items()
                        if int(item_key) != int(part_id)
                    ]
                )
                if extra_revision_codes is None:
                    return
                selected_item_revision_codes.update(extra_revision_codes)
                for cad_row in selected_cad_rows:
                    if str(cad_row.get("lifecycle_state") or "").upper() == "RELEASED":
                        revised_cads.append(
                            self.bom_service.revise_pdm_cad_document(int(cad_row["id"]))
                        )
                self.bom_service.checkout_item(
                    part_id,
                    as_user_id=as_user_id,
                    released_revision_code=released_revision_code,
                    include_owner_cad=False,
                )
                materialized = None
                cad_checkout_results = []
                if selected_cad_rows:
                    try:
                        for cad_row in selected_cad_rows:
                            cad_checkout_result = self.bom_service.checkout_pdm_cad_document(
                                int(cad_row["id"]),
                                released_item_revision_codes=selected_item_revision_codes,
                                as_user_id=as_user_id,
                                **workspace_descriptor,
                            )
                            cad_checkout_results.append(cad_checkout_result)
                            if workspace:
                                materialized_files = self._local_cad_workspaces().materialize_cad_document_package(
                                    workspace["id"],
                                    int(cad_row["id"]),
                                    preserve_existing=True,
                                )
                                materialized_paths.extend(
                                    row["path"] for row in materialized_files
                                )
                    except Exception:
                        for cad_row in selected_cad_rows:
                            try:
                                self.bom_service.undo_checkout_pdm_cad_document(
                                    int(cad_row["id"]), "CAD checkout failed", as_user_id=as_user_id
                                )
                            except Exception:
                                pass
                        raise
                QMessageBox.information(
                    self, "Success",
                    "Item checked out for metadata/structure changes."
                    + (
                        "\nCAD workspace files were copied:\n"
                        + "\n".join(f"- {path}" for path in materialized_paths)
                        if materialized_paths else
                        (
                            "\nSelected CAD Documents were checked out without workspace copies."
                            if selected_cad_rows else
                            "\nCAD Documents remain checked in."
                        )
                    ),
                )
                self._refresh_pdm_context_rows(
                    item_ids=[int(part_id)] + sorted(related_item_rows_by_id),
                    cad_ids=[
                        int(row["id"]) for row in selected_cad_rows
                    ] + [
                        int(value)
                        for cad_result in locals().get("cad_checkout_results", [])
                        for value in (cad_result or {}).get("related_drawing_checkout_ids", [])
                    ],
                )
                self._refresh_current_tree_item_indicator()
                try:
                    self.display_details(int(part_id))
                except Exception:
                    pass
            except ValueError as e:
                suffix = (
                    "\n\nThe Item checkout remains active for metadata changes."
                    if selected_cad_rows and self.bom_service.lock_repo.get_by_part(int(part_id))
                    else ""
                )
                QMessageBox.warning(self, "Check Out Failed", str(e) + suffix)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to check out part: {str(e)}")
    # -------------------------
    # Tree loading / details
    # -------------------------
    def _parse_dt(self, s: str):
        if not s:
            return None
        try:
            dt = datetime.fromisoformat(str(s).replace("Z", "+00:00"))
            if getattr(dt, "tzinfo", None) is not None:
                dt = dt.replace(tzinfo=None)
            return dt
        except Exception:
            return None

    def _get_part_modified_dt(self, part_id: int):
        try:
            d = self.bom_service.get_part_details(int(part_id)) or {}
            # DB stores as sqlite datetime('now') => 'YYYY-MM-DD HH:MM:SS'
            s = d.get("modified") or d.get("modified_at") or d.get("updated")
            return self._parse_dt(str(s)) if s else None
        except Exception:
            return None

    def _commit_effective_dt(self, commit):
        try:
            raw = getattr(commit, "merged_at", None) or getattr(commit, "committed_at", None)
            return self._parse_dt(str(raw)) if raw else None
        except Exception:
            return None

    def _latest_change_commit_for_part(self, part_id: int):
        """Latest real CAD/DRW change for this part across the project family."""
        try:
            proj = self.project_service.get_project_by_id(self.session.project_id) or {}
            root_id = proj.get("root_project_id")
            commits = (
                self.diag_service.repo.get_all_commits_for_root(int(root_id))
                if root_id else self.diag_service.repo.get_all_commits(self.session.project_id)
            )
        except Exception:
            commits = []

        latest = None
        latest_key = (datetime.min, -1)
        good_statuses = {"pending", "validated", "approved"}
        for c in commits or []:
            try:
                if int(getattr(c, "part_id", 0) or 0) != int(part_id):
                    continue
                if str(getattr(c, "status", "") or "").lower() not in good_statuses:
                    continue
                dt = self._commit_effective_dt(c) or datetime.min
                key = (dt, int(getattr(c, "id", 0) or 0))
                if key > latest_key:
                    latest = c
                    latest_key = key
            except Exception:
                continue
        return latest

    def _commit_ack_token(self, commit) -> str:
        if not commit:
            return ""
        dt = str(getattr(commit, "merged_at", None) or getattr(commit, "committed_at", None) or "")
        cid = str(getattr(commit, "commit_id", "") or "")
        row_id = str(getattr(commit, "id", "") or "")
        return f"commit:{row_id}:{cid}:{dt}"

    def _ack_covers_commit(self, ack: dict | None, commit) -> bool:
        if not ack or not commit:
            return False
        value = str(ack.get("acknowledged_against") or "").strip()
        if not value:
            return False
        if value == self._commit_ack_token(commit):
            return True

        # Backward compatibility: older acknowledgements stored a timestamp.
        ack_dt = self._parse_dt(value)
        commit_dt = self._commit_effective_dt(commit)
        if ack_dt and commit_dt and ack_dt >= commit_dt:
            return True
        return False

    def _doc_ack_target(self, part_id: int, doc_type: str) -> str:
        latest_commit = self._latest_change_commit_for_part(int(part_id))
        if latest_commit:
            return self._commit_ack_token(latest_commit)
        part_mod = self._get_part_modified_dt(int(part_id))
        if part_mod:
            return part_mod.isoformat(sep=" ")
        return datetime.now().isoformat(sep=" ")

    def _legacy_doc_path(self, part_id: int, doc_type: str) -> str:
        try:
            details = self.bom_service.get_part_details(int(part_id)) or {}
        except Exception:
            details = {}
        key = "pdf_path" if str(doc_type).upper() == "PDF" else "step_path"
        return str(details.get(key) or "").strip()

    def _doc_indicator_state(self, part_id: int, doc_type: str, attachments: list | None = None, latest_commit=None) -> dict:
        doc_type = str(doc_type).upper()
        label = doc_type
        latest_commit = latest_commit if latest_commit is not None else self._latest_change_commit_for_part(int(part_id))
        commit_dt = self._commit_effective_dt(latest_commit) if latest_commit else None

        if attachments is None:
            try:
                attachments = self.part_file_service.list_attachments(int(part_id))
            except Exception:
                attachments = []

        candidates = []
        for att in attachments or []:
            try:
                if str(getattr(att, "file_type", "") or "").upper() != doc_type:
                    continue
                ver = self.part_file_service.get_active_version(att.id)
                created_dt = self._parse_dt(getattr(ver, "created_at", None) or "") if ver else None
                sort_dt = created_dt or self._parse_dt(getattr(att, "created_at", None) or "") or datetime.min
                candidates.append((sort_dt, att, ver))
            except Exception:
                continue

        if candidates:
            candidates.sort(key=lambda row: (row[0], int(getattr(row[1], "id", 0) or 0)), reverse=True)
            _sort_dt, att, ver = candidates[0]
            if not ver:
                return {"state": "bad", "tooltip": f"{label}: attachment has no active version"}

            created_dt = self._parse_dt(getattr(ver, "created_at", None) or "")
            path = self.part_file_service.resolve_version_path(ver)
            if not path or not safe_exists(path):
                return {"state": "bad", "tooltip": f"{label}: file missing on disk"}

            if doc_type == "STEP" and latest_commit:
                latest_step_path = str(getattr(latest_commit, "step_file_path", "") or "").strip()
                latest_step_name = os.path.basename(latest_step_path)
                version_name = str(getattr(ver, "original_filename", "") or "").strip()
                if latest_step_name and version_name and latest_step_name == version_name:
                    return {"state": "ok", "tooltip": f"{label}: current from compared commit STEP"}

            if latest_commit and (not created_dt or (commit_dt and commit_dt > created_dt)):
                try:
                    ack = self.part_doc_ack_service.get_ack(int(part_id), doc_type)
                except Exception:
                    ack = None
                if self._ack_covers_commit(ack, latest_commit):
                    return {"state": "ack", "tooltip": f"{label}: safe by user acknowledgement"}
                return {"state": "bad", "tooltip": f"{label}: newer commit needs review"}

            return {"state": "ok", "tooltip": f"{label}: current"}

        legacy_path = self._legacy_doc_path(int(part_id), doc_type)
        if legacy_path:
            resolved = self._resolve_file_path(legacy_path)
            if not resolved or not safe_exists(resolved):
                return {"state": "bad", "tooltip": f"{label}: file missing on disk"}
            if latest_commit:
                try:
                    ack = self.part_doc_ack_service.get_ack(int(part_id), doc_type)
                except Exception:
                    ack = None
                if self._ack_covers_commit(ack, latest_commit):
                    return {"state": "ack", "tooltip": f"{label}: legacy file safe by user acknowledgement"}
                return {"state": "bad", "tooltip": f"{label}: legacy file needs review after newer commit"}
            return {"state": "ok", "tooltip": f"{label}: legacy file exists"}

        return {"state": "absent", "tooltip": f"{label}: no attachment"}

    def _indicator_summary_for_part(self, part_id: int | None, issues: set | None = None) -> dict:
        if part_id is None:
            return {
                "pdf": {"state": "absent", "tooltip": "PDF: no attachment"},
                "step": {"state": "absent", "tooltip": "STEP: no attachment"},
            }
        try:
            pid = int(part_id)
        except Exception:
            pid = None
        if pid is None:
            return {
                "pdf": {"state": "absent", "tooltip": "PDF: no attachment"},
                "step": {"state": "absent", "tooltip": "STEP: no attachment"},
            }

        cached = self._doc_indicator_cache.get(pid)
        if cached is None:
            try:
                attachments = self.part_file_service.list_attachments(pid)
            except Exception:
                attachments = []
            latest_commit = self._latest_change_commit_for_part(pid)
            cached = {
                "pdf": self._doc_indicator_state(pid, "PDF", attachments, latest_commit),
                "step": self._doc_indicator_state(pid, "STEP", attachments, latest_commit),
            }
            self._doc_indicator_cache[pid] = cached

        summary = {
            "pdf": dict(cached.get("pdf") or {}),
            "step": dict(cached.get("step") or {}),
        }

        issues = issues or set()
        pdf_state = str(summary["pdf"].get("state") or "").lower()
        step_state = str(summary["step"].get("state") or "").lower()
        if ("missing_pdf" in issues or "outdated_pdf" in issues) and pdf_state not in ("ok", "ack"):
            summary["pdf"] = {"state": "bad", "tooltip": "PDF: file missing or outdated on disk"}
        if ("missing_step" in issues or "outdated_step" in issues) and step_state not in ("ok", "ack"):
            summary["step"] = {"state": "bad", "tooltip": "STEP: file missing or outdated on disk"}
        return summary

    def _pick_indicator_icon(self, issues: set, part_id: int | None = None) -> QIcon:
        summary = self._indicator_summary_for_part(part_id, issues)
        return self._make_indicator_icon(
            summary["pdf"].get("state", "absent"),
            summary["step"].get("state", "absent"),
        )

    def load_tree(self):
        # Show spinner immediately, then fetch tree data in a background thread.
        self._cancel_lazy_level_requests()
        self._tree_load_seq += 1
        seq = self._tree_load_seq

        pid = getattr(self.session, "project_id", None)
        if not pid:
            try:
                self.tree.clear()
            except Exception:
                pass
            self._set_tree_loading(False)
            self._mark_initial_tree_ready()
            return
        try:
            self._issue_summary_cache = self.issue_service.part_summary()
            score = max(0, self.issue_service.health_score() - len(getattr(self, "missing_ids", set()) or set()))
            color = "#2e7d32" if score >= 85 else ("#a16207" if score >= 65 else "#b91c1c")
            self.bom_health_label.setText(f"Health: {score}/100")
            self.bom_health_label.setStyleSheet(
                f"font-size:11px;font-weight:700;color:{color};background:transparent;border:none;"
            )
        except Exception:
            self._issue_summary_cache = {}

        self._set_tree_loading(True)

        try:
            if self._tree_worker is not None:
                self._tree_worker.cancel()
        except Exception:
            pass

        worker = _TreeLoadWorker(seq, self.bom_service, int(pid), list(self.missing_files or []))
        thread = QThread(self)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(self._on_tree_loaded)
        worker.failed.connect(self._on_tree_failed)
        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        self._tree_worker = worker
        self._tree_thread = thread
        thread.start()

    def _on_tree_loaded(self, seq: int, tree_data: object, missing_map: object) -> None:
        if int(seq) != int(getattr(self, "_tree_load_seq", 0)):
            return
        self._begin_tree_build(int(seq), dict(tree_data or {}), dict(missing_map or {}))

    def _on_tree_failed(self, seq: int, err: str) -> None:
        if int(seq) != int(getattr(self, "_tree_load_seq", 0)):
            return
        try:
            self.tree.clear()
        except Exception:
            pass
        self._set_tree_loading(False)
        self._mark_initial_tree_ready()

    def _mark_initial_tree_ready(self) -> None:
        """Record the first terminal tree state and emit its sticky readiness signal."""
        if self._initial_tree_ready_emitted:
            return
        self._initial_tree_ready_emitted = True
        self.initial_tree_ready.emit()

    @property
    def initial_tree_is_ready(self) -> bool:
        return bool(self._initial_tree_ready_emitted)

    def _load_tree_from_data(self, tree_data: dict, missing_map: dict) -> None:
        # Use incremental builder to keep UI responsive.
        self._begin_tree_build(int(getattr(self, "_tree_load_seq", 0)), tree_data or {}, missing_map or {})

    def _begin_tree_build(self, seq: int, tree_data: dict, missing_map: dict) -> None:
        self._tree_build_seq = int(seq)
        self._tree_build_missing_map = missing_map or {}
        self._last_tree_node_count = 0
        self._invalidate_doc_indicator()
        self._indicator_refresh_timer.stop()
        self._indicator_refresh_queue.clear()
        self._lazy_tree_active = "roots" in (tree_data or {})
        self._lazy_tree_materialized = False
        self._pending_relation_expanded_paths = set()
        self._pending_relation_expanded_part_ids = set()
        self._bom_row_numbers = dict((tree_data or {}).get("row_numbers") or {})
        self._bom_folder_row_numbers = dict((tree_data or {}).get("folder_rows") or {})
        self._bom_folder_path_rows = dict((tree_data or {}).get("folder_path_rows") or {})
        self._bom_folders_cache = list((tree_data or {}).get("folders") or [])

        # cache for restore after search
        try:
            self._cached_tree_data = tree_data
            self._cached_missing_map = missing_map
            self._full_tree_cached = True
        except Exception:
            pass

        try:
            self._tree_build_timer.stop()
        except Exception:
            pass

        self._tree_build_queue = deque()
        try:
            roots = (tree_data or {}).get("roots") if self._lazy_tree_active else (tree_data or {}).values()
            for info in (roots or []):
                self._tree_build_queue.append((None, info))
        except Exception:
            self._tree_build_queue = deque()

        try:
            self.tree.clear()
        except Exception:
            pass

        self._set_tree_loading(True)
        try:
            self.tree.setUpdatesEnabled(False)
        except Exception:
            pass
        self._tree_build_timer.start()

    def _tree_build_step(self) -> None:
        if int(self._tree_build_seq) != int(getattr(self, "_tree_load_seq", 0)):
            try:
                self._tree_build_timer.stop()
            except Exception:
                pass
            return

        start = time.perf_counter()
        processed = 0
        try:
            while self._tree_build_queue and processed < 50 and (time.perf_counter() - start) < 0.01:
                parent_item, info = self._tree_build_queue.popleft()
                item = self._make_tree_item(info)

                if parent_item is None:
                    self.tree.addTopLevelItem(item)
                else:
                    parent_item.addChild(item)

                if self._lazy_tree_active:
                    self._ensure_lazy_placeholder(item)

                children = info.get("children", [])
                if children:
                    try:
                        for c in children:
                            self._tree_build_queue.append((item, c))
                    except Exception:
                        pass

                self._last_tree_node_count += 1
                processed += 1

            if not self._tree_build_queue:
                self._tree_build_timer.stop()
                try:
                    self.tree.setUpdatesEnabled(True)
                except Exception:
                    pass
                try:
                    if self._lazy_tree_active:
                        self._render_folder_context(None, self._bom_folders_cache)
                    else:
                        self._apply_organizational_folders()
                        self._renumber_full_bom_tree_rows()
                    self._sync_search_tree_row_numbers()
                except Exception:
                    pass

                # Start every BOM level collapsed; users expand only the branches they need.
                try:
                    self.tree.collapseAll()
                except Exception:
                    pass

                self._set_tree_loading(False)
                try:
                    if not self._is_default_bom_advanced_filter():
                        self.apply_bom_tree_filter(self._bom_advanced_filters)
                except Exception:
                    pass

                # Fire once when the first current tree load reaches a terminal state.
                try:
                    self._mark_initial_tree_ready()
                except Exception:
                    pass
        except RuntimeError:
            try:
                self._tree_build_timer.stop()
            except Exception:
                pass
            try:
                self.tree.setUpdatesEnabled(True)
                self._set_tree_loading(False)
            except Exception:
                pass
            try:
                self._mark_initial_tree_ready()
            except Exception:
                pass
            return

    def _load_tree_impl(self, project_id: int) -> None:
        # Legacy internal method name kept for compatibility; now sync-calls the builder.
        if not project_id:
            return
        tree_data = self.bom_service.get_bom_lazy_tree(project_id) or {}
        missing_map = {}
        try:
            for bom_id, issue_type, _filename in (self.missing_files or []):
                missing_map.setdefault(int(bom_id), set()).add(str(issue_type))
        except Exception:
            missing_map = {}
        self._load_tree_from_data(tree_data, missing_map)

    def _enter_search_mode(self) -> None:
        """Switch the stack to the search-results view and prepare it for new results."""
        self._in_search_mode = True
        try:
            self._search_tree.clear()
        except Exception:
            pass
        try:
            self._tree_stack.setCurrentIndex(2)
        except Exception:
            pass

    def _exit_search_mode(self) -> None:
        """Restore the full-tree view (index 1). Zero data access — pure UI swap."""
        self._in_search_mode = False
        try:
            self._tree_stack.setCurrentIndex(1)
        except Exception:
            pass
        try:
            self._search_tree.clear()      # free search-result memory
        except Exception:
            pass

    def _on_bom_item_expanded(self, item: QTreeWidgetItem) -> None:
        if not getattr(self, "_lazy_tree_active", False):
            return
        if (
            item is None
            or self._is_folder_tree_item(item)
            or self._is_lazy_placeholder(item)
            or item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE)
        ):
            return

        # Keep the branch visually closed while its direct level is fetched. This
        # prevents the empty-row flash and lets the branch indicator host the spinner.
        previous_signal_state = self.tree.blockSignals(True)
        try:
            item.setExpanded(False)
        finally:
            self.tree.blockSignals(previous_signal_state)
        self._load_lazy_children_for_item(item, asynchronous=True, expand_after=True)

    def _load_lazy_children_for_item(
        self,
        item: QTreeWidgetItem,
        asynchronous: bool = True,
        expand_after: bool = False,
    ) -> None:
        if item is None or self._is_folder_tree_item(item) or self._is_lazy_placeholder(item):
            return
        if item.data(0, BOM_TREE_CHILDREN_LOADED_ROLE):
            return
        if item.data(0, BOM_TREE_LOADING_ROLE):
            return
        try:
            part_id = int(item.data(0, Qt.UserRole))
        except (TypeError, ValueError):
            return

        parent_path = str(item.data(0, BOM_TREE_PATH_ROLE) or "")
        if asynchronous:
            self._start_lazy_children_request(item, part_id, parent_path, expand_after)
            return

        try:
            nodes = self.bom_service.get_bom_lazy_children(
                int(self.session.project_id), part_id,
                parent_path,
            ) or []
            self._apply_lazy_children_result(item, part_id, nodes, expand_after=expand_after)
        except Exception as exc:
            item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, False)
            self._ensure_lazy_placeholder(item)
            self.show_alert(f"Could not load this BOM level: {exc}", "error")

    def _start_lazy_children_request(
        self,
        item: QTreeWidgetItem,
        part_id: int,
        parent_path: str,
        expand_after: bool,
    ) -> None:
        self._lazy_level_request_seq += 1
        request_id = int(self._lazy_level_request_seq)
        worker = _LazyChildrenWorker(
            request_id,
            self.bom_service,
            int(self.session.project_id),
            int(part_id),
            parent_path,
        )
        thread = QThread(self)
        worker.moveToThread(thread)
        self._lazy_level_requests[request_id] = {
            "worker": worker,
            "thread": thread,
            "item": item,
            "part_id": int(part_id),
            "parent_path": str(parent_path),
            "expand_after": bool(expand_after),
        }
        self.tree.setItemLoading(item, True)

        thread.started.connect(worker.run)
        worker.finished.connect(self._on_lazy_children_loaded)
        worker.failed.connect(self._on_lazy_children_failed)
        worker.done.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.start()

    def _lazy_request_item_is_current(self, request: dict) -> bool:
        item = request.get("item")
        try:
            return (
                item is not None
                and item.treeWidget() is self.tree
                and int(item.data(0, Qt.UserRole)) == int(request.get("part_id"))
                and str(item.data(0, BOM_TREE_PATH_ROLE) or "") == str(request.get("parent_path") or "")
            )
        except Exception:
            return False

    def _on_lazy_children_loaded(self, request_id: int, nodes: object) -> None:
        request = self._lazy_level_requests.pop(int(request_id), None)
        if not request:
            return
        item = request.get("item")
        if not self._lazy_request_item_is_current(request):
            try:
                self.tree.setItemLoading(item, False)
            except Exception:
                pass
            return
        expand_after = bool(request.get("expand_after"))
        loaded_nodes = list(nodes or [])
        try:
            self._apply_lazy_children_result(
                item,
                int(request["part_id"]),
                loaded_nodes,
                expand_after=False,
            )
        except Exception as exc:
            item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, False)
            self._ensure_lazy_placeholder(item)
            self.show_alert(f"Could not display this BOM level: {exc}", "error")
            loaded_nodes = []
        finally:
            try:
                self.tree.setItemLoading(item, False)
            except Exception:
                pass
        if expand_after and item.treeWidget() is self.tree and loaded_nodes:
            item.setExpanded(True)

    def _on_lazy_children_failed(self, request_id: int, error: str) -> None:
        request = self._lazy_level_requests.pop(int(request_id), None)
        if not request:
            return
        item = request.get("item")
        try:
            self.tree.setItemLoading(item, False)
        except Exception:
            pass
        if not self._lazy_request_item_is_current(request):
            return
        item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, False)
        self._ensure_lazy_placeholder(item)
        self.show_alert(f"Could not load this BOM level: {error}", "error")

    def _apply_lazy_children_result(
        self,
        item: QTreeWidgetItem,
        part_id: int,
        nodes: list,
        expand_after: bool = False,
    ) -> None:
        self.tree.setUpdatesEnabled(False)
        try:
            item.setData(0, BOM_TREE_CHILDREN_LOADED_ROLE, True)
            if item.data(0, BOM_TREE_REPLACE_CHILDREN_ROLE):
                while item.childCount():
                    item.takeChild(0)
                item.setData(0, BOM_TREE_REPLACE_CHILDREN_ROLE, False)
            for index in range(item.childCount() - 1, -1, -1):
                if self._is_lazy_placeholder(item.child(index)):
                    item.takeChild(index)
            for node in nodes:
                try:
                    child_id = int(node.get("id"))
                except (TypeError, ValueError):
                    child_id = None
                if child_id is not None and self._child_exists_under_item(item, child_id):
                    continue
                child = self._make_tree_item(node)
                item.addChild(child)
                self._ensure_lazy_placeholder(child)
            self._render_folder_context(part_id, self._bom_folders_cache, containers=[item])
            item.setData(0, BOM_TREE_IS_ASSEMBLY_ROLE, bool(nodes))
            self._restore_pending_relation_expansions(item)
        finally:
            try:
                self.tree.setUpdatesEnabled(True)
                self.tree.viewport().update()
            except Exception:
                pass
        if expand_after and item.treeWidget() is self.tree and bool(nodes):
            item.setExpanded(True)

    def _materialize_all_lazy_branches(self) -> None:
        if not getattr(self, "_lazy_tree_active", False) or getattr(self, "_lazy_tree_materialized", False):
            return
        self._cancel_lazy_level_requests()

        def visit(item: QTreeWidgetItem):
            if self._is_folder_tree_item(item):
                for index in range(item.childCount()):
                    visit(item.child(index))
                return
            if self._is_lazy_placeholder(item):
                return
            self._load_lazy_children_for_item(item, asynchronous=False, expand_after=False)
            for index in range(item.childCount()):
                visit(item.child(index))

        self.tree.setUpdatesEnabled(False)
        try:
            for index in range(self.tree.topLevelItemCount()):
                visit(self.tree.topLevelItem(index))
            self._lazy_tree_materialized = True
        finally:
            self.tree.setUpdatesEnabled(True)
            self.tree.viewport().update()

    def on_tree_item_clicked(self, item, column):
        if item is not None and item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
            self._show_pdm_cad_selection(item)
            self._sync_visual_action_states()
            return
        folder_id = item.data(0, BOM_TREE_FOLDER_ROLE) if item is not None else None
        if folder_id:
            self._show_folder_selection(item, int(folder_id))
            self._sync_visual_action_states()
            return
        item_id = item.data(0, Qt.UserRole)
        self._selected_pdm_kind = PDM_OBJECT_ITEM
        self.current_cad_document_id = None
        self._current_cad_payload = {}
        self.display_details(item_id)
        self._sync_visual_action_states()

    @staticmethod
    def _item_summary_field_titles() -> dict[str, str]:
        return {
            "part_number": "Number",
            "aes_number": "AES Number",
            "item_type": "Item Type",
            "source": "Source",
            "view": "View",
            "unit": "Default Unit",
            "drawing": "DRW Number",
            "type": "Type",
            "revision": "Revision / Iteration",
            "state": "Lifecycle",
            "material": "Material",
            "categories": "Categories",
        }

    def _set_summary_field_titles(self, titles: dict[str, str]) -> None:
        for key, (field_label, _value_label, _detail_keys) in self._details_summary_fields.items():
            field_label.setText(str(titles.get(key) or key).upper())

    def _show_pdm_cad_selection(self, item: QTreeWidgetItem) -> None:
        payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
        cad_id = item.data(0, PDM_CAD_DOCUMENT_ID_ROLE)
        if cad_id is None:
            return
        payload.setdefault("id", int(cad_id))
        payload.setdefault("item_id", item.data(0, PDM_ASSOCIATED_ITEM_ID_ROLE))
        payload.setdefault("association_id", item.data(0, PDM_ASSOCIATION_ID_ROLE))
        payload.setdefault("association_type", item.data(0, PDM_ASSOCIATION_TYPE_ROLE))
        self._selected_pdm_kind = PDM_OBJECT_CAD
        self.current_cad_document_id = int(cad_id)
        self.current_associated_item_id = payload.get("item_id")
        self._current_cad_payload = payload
        self.current_part_id = None
        self.current_folder_id = None
        self._current_part_details = {}
        self.details_summary_card.show()
        self.associated_files_card.hide()
        self.details_alert_frame.hide()
        self.view_full_details_btn.setEnabled(False)
        self.edit_categories_btn.setEnabled(False)
        self.details_name_label.setText(
            str(payload.get("name") or payload.get("file_name") or "CAD Document")
        )
        identity = ["CAD DOCUMENT"]
        if payload.get("file_name"):
            identity.append(str(payload["file_name"]))
        if payload.get("category"):
            identity.append(str(payload["category"]).replace("_", " "))
        self.details_identity_label.setText("  |  ".join(identity))
        self._set_summary_field_titles({
            "aes_number": "CAD File",
            "part_number": "Approved Creo File",
            "drawing": "Related DRW",
            "type": "CAD Category",
            "revision": "CAD Revision / Iteration",
            "state": "Lifecycle",
            "material": "Associations",
            "categories": "Related Items / Versions",
        })
        association_lines = self._pdm_item_association_lines(payload)
        related_item = "\n".join(association_lines) or "Not associated"
        association_summary = (
            f"{len(association_lines)} active Item association(s)"
            if association_lines else "None"
        )
        representation_lines = []
        seen_representation_keys = set()
        for association in self._pdm_document_associations(payload):
            item_value = association.get("item_id")
            try:
                item_value = int(item_value)
            except Exception:
                continue
            try:
                item_documents = self.bom_service.list_item_cad_associations(item_value) or []
            except Exception:
                item_documents = []
            item_label = self._pdm_item_association_label(association)
            representation_lines.append(f"{item_label}")
            for document in item_documents:
                if str(document.get("category") or "").upper() == "DRAWING":
                    continue
                key = (item_value, int(document.get("id") or 0), document.get("association_type"))
                if key in seen_representation_keys:
                    continue
                seen_representation_keys.add(key)
                representation_lines.append(f"  {self._cad_representation_summary_line(document)}")
        revision = f"{payload.get('revision') or 'A'}.{int(payload.get('iteration') or 1)}"
        creo_file = self._pdm_creo_file_text(payload) or "No approved file yet"
        related_drawings = list(payload.get("related_drawings") or [])
        drawing_text = "\n".join(
            f"{drawing.get('file_name') or drawing.get('name')} "
            f"({drawing.get('revision') or 'A'}.{int(drawing.get('iteration') or 1)}, "
            f"{drawing.get('lifecycle_state') or 'IN_WORK'}"
            + (
                f", Creo {self._pdm_creo_file_text(drawing)}"
                if self._pdm_creo_file_text(drawing) else ""
            )
            + ")"
            for drawing in related_drawings
        ) or "Not linked"
        values = {
            "aes_number": str(payload.get("file_name") or "-"),
            "part_number": creo_file,
            "drawing": drawing_text,
            "type": str(payload.get("category") or "OTHER"),
            "revision": f"CAD {revision}\nIndependent from linked Item version",
            "state": str(payload.get("lifecycle_state") or "-"),
            "material": association_summary,
            "categories": "\n".join(representation_lines) or related_item,
        }
        for key, (_label, value_label, _keys) in self._details_summary_fields.items():
            value_label.setText(values.get(key, "-"))
            if key in {"item_type", "source", "view", "unit"}:
                _label.hide()
                value_label.hide()
            else:
                _label.show()
                value_label.show()
        self.uses_tree.clear()
        self.where_used_tree.clear()
        self.effective_where_used_tree.clear()
        self.structure_summary_label.setText(
            "CAD membership is shown in CAD Structure; the related Item hierarchy remains independent."
        )
        self.notes_view.setText(
            "This is a managed PRT/ASM CAD model. Related native DRW files are shown in "
            "the model details and never appear as independent CAD Structure nodes. "
            "The CAD Document revision/iteration is intentionally independent from the "
            "associated EBOM Item revision/iteration; the OWNER/IMAGE/CONTENT association "
            "is the link between them. "
            "Item PDF/STEP delivery content, Item integrity and Item issues are intentionally "
            "not evaluated on this CAD structure row."
        )
        try:
            self.part_issues_table.setRowCount(0)
            self.history_panel.clear()
            self.files_table.setRowCount(0)
            self.versions_table.setRowCount(0)
        except Exception:
            pass
        self._set_engineering_item_actions_enabled(False)
        checked_out = payload.get("checked_out_by") is not None
        try:
            checked_out_by_me = (
                checked_out and self.session.user_id is not None
                and int(payload.get("checked_out_by")) == int(self.session.user_id)
            )
        except Exception:
            checked_out_by_me = False
        can_finish_cad_checkout = checked_out_by_me or self.perm.can("merge")
        can_manage = self.perm.can("manage_parts")
        self.checkout_part_btn.setEnabled(can_manage and not checked_out)
        self.checkin_part_btn.setEnabled(can_manage and checked_out and can_finish_cad_checkout)
        self.undo_checkout_btn.setEnabled(can_manage and checked_out and can_finish_cad_checkout)
        self.checkout_part_btn.setToolTip(
            "Check out this CAD Document and reserve its associated Item data as required"
        )
        self.checkin_part_btn.setToolTip("Create the next CAD Document iteration")
        self.undo_checkout_btn.setToolTip("Discard this CAD Document checkout")
        self._sync_action_ribbon_menus()
        try:
            self.tabs.setCurrentIndex(0)
        except Exception:
            pass
        self._sync_visual_action_states()

    def _set_engineering_item_actions_enabled(self, enabled: bool) -> None:
        can_manage = self.perm.can("manage_parts")
        self.edit_part_btn.setEnabled(bool(enabled and can_manage))
        self.delete_part_btn.setEnabled(bool(enabled and can_manage))
        self.add_child_btn.setEnabled(bool(enabled and can_manage))
        self.checkout_part_btn.setEnabled(bool(enabled))
        self.checkin_part_btn.setEnabled(bool(enabled))
        self.undo_checkout_btn.setEnabled(bool(enabled))
        self.set_revision_btn.setEnabled(bool(enabled and self.perm.can("set_revision")))
        self.release_revision_btn.setEnabled(bool(enabled and self.perm.can("set_revision")))
        try:
            details = getattr(self, "_current_part_details", {}) or {}
            self.compare_iterations_btn.setEnabled(
                bool(enabled) and str(details.get("type") or "").lower() in {"asm", "assembly"}
            )
            self.create_configuration_btn.setEnabled(
                bool(enabled) and str(details.get("type") or "").lower() in {"asm", "assembly"}
            )
            self.update_child_versions_btn.setEnabled(
                bool(enabled) and str(details.get("type") or "").lower() in {"asm", "assembly"}
            )
        except Exception:
            pass
        self._sync_action_ribbon_menus()

    def _show_folder_selection(self, item: QTreeWidgetItem, folder_id: int) -> None:
        self.clear_details()
        self.current_folder_id = int(folder_id)
        self.details_summary_card.hide()
        self.associated_files_card.hide()
        self.details_alert_frame.hide()
        self._set_engineering_item_actions_enabled(False)
        try:
            self.tabs.setCurrentIndex(0)
        except Exception:
            pass

    def _indicator_refresh_step(self) -> None:
        start = time.perf_counter()
        processed = 0
        while self._indicator_refresh_queue and processed < 8 and (time.perf_counter() - start) < 0.008:
            item, part_id = self._indicator_refresh_queue.popleft()
            try:
                if item.treeWidget() is None:
                    continue
                issues = self._issues_for_part(part_id)
                summary = self._indicator_summary_for_part(part_id, issues)
                item.setData(BOM_COL_FILES, BOM_TREE_FILES_ROLE, _file_badges_payload(part_id, issues, summary))
                item.setData(BOM_COL_INTEGRITY, BOM_TREE_INTEGRITY_ROLE, _integrity_payload(part_id, self.missing_files, self.missing_ids))
                item.setToolTip(BOM_COL_FILES, "\n".join([
                    str(summary["pdf"].get("tooltip", "PDF: unknown")),
                    str(summary["step"].get("tooltip", "STEP: unknown")),
                ]))
                name_tips = []
                locked_text = str(item.data(0, BOM_TREE_INWORK_ROLE) or "")
                if locked_text:
                    name_tips.append(locked_text)
                issue_summary = item.data(0, BOM_TREE_ISSUE_ROLE) or {}
                active_count = int(issue_summary.get("active_count") or 0)
                if active_count:
                    name_tips.append(f"{active_count} active issue(s) linked to this part")
                policy = item.data(0, BOM_TREE_POLICY_ROLE) or {}
                classification = str(
                    policy.get("classification") or "PHYSICAL"
                ).upper()
                occurrence_behavior = str(
                    policy.get("ebom_behavior") or "INHERIT"
                ).upper()
                resolved_behavior = str(
                    policy.get("resolved_ebom_behavior") or "NORMAL"
                ).upper()
                name_tips.append(
                    f"EBOM policy: {classification}; occurrence {occurrence_behavior}; "
                    f"effective {resolved_behavior}"
                )
                if resolved_behavior == "EXCLUDE":
                    name_tips.append(
                        "Not for delivery: excluded with all descendants from Released EBOM."
                    )
                elif resolved_behavior == "FLATTEN":
                    name_tips.append(
                        "CAD grouping only: hidden from Released EBOM; children are promoted."
                    )
                name_tips.extend([
                    str(summary["pdf"].get("tooltip", "PDF: unknown")),
                    str(summary["step"].get("tooltip", "STEP: unknown")),
                ])
                item.setToolTip(BOM_COL_NAME, "\n".join(name_tips))
            except Exception:
                pass
            processed += 1
        if not self._indicator_refresh_queue:
            self._indicator_refresh_timer.stop()
        try:
            self.tree.viewport().update()
            self._search_tree.viewport().update()
            self._ebom_tree.viewport().update()
        except Exception:
            pass

    def _clear_folder_selection(self) -> None:
        self.current_folder_id = None
        self.details_summary_card.show()
        self.associated_files_card.show()
        self._set_engineering_item_actions_enabled(False)
        self.clear_details()

    def _create_structure_relation_tree(self) -> QTreeWidget:
        tree = QTreeWidget()
        tree.setColumnCount(10)
        tree.setHeaderLabels([
            "Name", "Number", "Relation", "Qty", "Type", "Current",
            "Bound", "Latest", "Binding", "Lifecycle",
        ])
        tree.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tree.setSelectionBehavior(QAbstractItemView.SelectRows)
        tree.setSelectionMode(QAbstractItemView.SingleSelection)
        tree.setAlternatingRowColors(True)
        tree.setUniformRowHeights(True)
        tree.setIconSize(QSize(12, 12))
        tree.setIndentation(14)
        tree.setRootIsDecorated(True)
        header = tree.header()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        for column in range(1, 10):
            header.setSectionResizeMode(column, QHeaderView.ResizeToContents)
        return tree

    def _add_structure_relation_node(
        self,
        tree: QTreeWidget,
        node: dict,
        parent: QTreeWidgetItem | None = None,
    ) -> QTreeWidgetItem | None:
        if not node:
            return None
        quantity = node.get("quantity")
        item = QTreeWidgetItem([
            str(node.get("name") or ""),
            str(node.get("part_number") or node.get("aes_number") or ""),
            str(node.get("relation") or ""),
            "" if quantity is None else str(quantity),
            str(node.get("type") or ""),
            str(node.get("current_version") or node.get("revision") or ""),
            str(node.get("bound_version") or ""),
            str(node.get("latest_version") or ""),
            str(node.get("binding_status") or ""),
            str(node.get("lifecycle_state") or node.get("status") or ""),
        ])
        item.setData(0, Qt.UserRole, node.get("id"))
        item.setData(0, Qt.UserRole + 1, node.get("usage_id"))
        item.setData(0, Qt.UserRole + 2, node.get("relation_parent_id"))
        item.setData(0, STRUCTURE_CURRENT_ITERATION_ROLE, node.get("current_iteration_id"))
        item.setData(0, STRUCTURE_BOUND_ITERATION_ROLE, node.get("bound_iteration_id"))
        item.setData(0, STRUCTURE_LATEST_ITERATION_ROLE, node.get("latest_iteration_id"))
        item.setIcon(0, _bom_type_icon(node.get("type")))
        if node.get("cycle"):
            item.setToolTip(0, "Circular structure reference")
            item.setForeground(2, QBrush(QColor("#b91c1c")))
        if node.get("promotion_path"):
            item.setToolTip(
                0,
                "Effective EBOM parent after flattening through: "
                + str(node["promotion_path"]),
            )
        if parent is None:
            tree.addTopLevelItem(item)
        else:
            parent.addChild(item)
        for child in node.get("children") or []:
            self._add_structure_relation_node(tree, child, item)
        return item

    def _populate_structure_views(self, context: dict) -> None:
        uses = (context or {}).get("uses")
        where_used = (context or {}).get("where_used")
        effective_where_used = (context or {}).get("effective_where_used")
        for tree, root in (
            (self.uses_tree, uses),
            (self.where_used_tree, where_used),
            (self.effective_where_used_tree, effective_where_used),
        ):
            tree.setUpdatesEnabled(False)
            try:
                tree.clear()
                root_item = self._add_structure_relation_node(tree, root)
                if root_item is not None:
                    root_item.setExpanded(True)
            finally:
                tree.setUpdatesEnabled(True)

        uses_count = int((context or {}).get("uses_count") or 0)
        where_used_count = int((context or {}).get("where_used_count") or 0)
        effective_count = int(
            (context or {}).get("effective_where_used_count") or 0
        )
        effective_error = str(
            (context or {}).get("effective_where_used_error") or ""
        )
        self.structure_summary_label.setText(
            f"Structure relations: {uses_count} uses  |  {where_used_count} direct CAD parent(s)"
            f"  |  {effective_count} effective EBOM parent occurrence(s)"
        )
        self.structure_views.setTabText(0, f"Uses ({uses_count})")
        self.structure_views.setTabText(1, f"CAD Where Used ({where_used_count})")
        self.structure_views.setTabText(
            2, f"Effective EBOM Where Used ({effective_count})"
        )
        self.effective_where_used_tree.setToolTip(
            effective_error or "First visible EBOM parent after flattening."
        )
        try:
            details = getattr(self, "_current_part_details", {}) or {}
            self._update_lifecycle_action_states(details)
        except Exception:
            self.compare_iterations_btn.setEnabled(False)
            self.update_child_versions_btn.setEnabled(False)

    def update_child_versions(self):
        if not getattr(self, "current_part_id", None):
            return
        try:
            rows = self.bom_service.get_child_version_status(int(self.current_part_id)) or []
        except Exception as exc:
            QMessageBox.warning(self, "Child Versions", str(exc))
            return
        outdated = [row for row in rows if not row.get("is_latest")]
        if not outdated:
            QMessageBox.information(self, "Child Versions", "All direct children already use their latest iterations.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Update Child Versions")
        dialog.resize(720, 420)
        layout = QVBoxLayout(dialog)
        info = QLabel(
            "Select direct children to adopt in the checked-out assembly. "
            "The new bindings become permanent only when the assembly is committed."
        )
        info.setWordWrap(True)
        layout.addWidget(info)
        table = QTableWidget(len(rows), 5)
        table.setHorizontalHeaderLabels(["Child", "Bound", "Latest", "Status", "Source"])
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for column in range(1, 5):
            table.horizontalHeader().setSectionResizeMode(column, QHeaderView.ResizeToContents)
        for row_index, row in enumerate(rows):
            child_label = str(row.get("part_number") or row.get("aes_number") or "").strip()
            name = str(row.get("name") or "").strip()
            if child_label and name:
                child_label = f"{child_label} - {name}"
            else:
                child_label = child_label or name
            aes_number = str(row.get("aes_number") or "").strip()
            if aes_number and aes_number.casefold() != str(row.get("part_number") or "").strip().casefold():
                child_label += f"  |  AES {aes_number}"
            values = [
                child_label,
                str(row.get("bound_version") or ""),
                str(row.get("latest_version") or ""),
                "Current" if row.get("is_latest") else "Update available",
                str(row.get("binding_source") or ""),
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(value)
                if column == 0:
                    item.setData(Qt.UserRole, int(row["child_bom_id"]))
                table.setItem(row_index, column, item)
                if not row.get("is_latest"):
                    item.setSelected(True)
        layout.addWidget(table)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Ok).setText("Update Selected")
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        if dialog.exec_() != QDialog.Accepted:
            return
        selected_ids = []
        for index in table.selectionModel().selectedRows():
            item = table.item(index.row(), 0)
            if item is not None:
                selected_ids.append(int(item.data(Qt.UserRole)))
        if not selected_ids:
            QMessageBox.warning(self, "Child Versions", "Select at least one child.")
            return
        try:
            changed = self.bom_service.update_children_to_latest(
                int(self.current_part_id), selected_ids
            )
            self._refresh_part_in_tree(int(self.current_part_id))
            self.display_details(int(self.current_part_id))
            QMessageBox.information(
                self, "Child Versions", f"Updated {len(changed)} child binding(s) in the working assembly."
            )
        except Exception as exc:
            QMessageBox.warning(self, "Child Versions", str(exc))

    def _show_latest_bom_cad_files(self, part_id: int) -> None:
        try:
            details = self.bom_service.get_part_details(int(part_id)) or {}
            iteration_id = details.get("current_iteration_id")
            if not iteration_id:
                raise ValueError("This item has no current iteration.")
            self._show_iteration_cad_files(
                int(iteration_id), "Latest", str(details.get("current_version") or "")
            )
        except Exception as exc:
            QMessageBox.warning(self, "CAD Files", str(exc))

    def _show_iteration_cad_files(
        self, iteration_id: int, relation_label: str, displayed_version: str = ""
    ) -> None:
        files = self.bom_service.get_iteration_cad_files(int(iteration_id)) or {}
        if not files:
            QMessageBox.warning(self, "CAD Files", "The selected iteration was not found.")
            return
        version = str(files.get("version_label") or displayed_version or "").strip()
        dialog = QDialog(self)
        dialog.setWindowTitle(f"{relation_label} CAD Files - {version}".strip(" -"))
        dialog.resize(560, 210)
        layout = QVBoxLayout(dialog)
        heading = QLabel(f"{relation_label} iteration: {version or iteration_id}")
        heading.setStyleSheet("font-weight:700;")
        layout.addWidget(heading)
        table = QTableWidget(2, 2)
        table.setHorizontalHeaderLabels(["File Role", "Captured Creo File"])
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        rows = (
            ("Native CAD", str(files.get("filename") or "Not captured")),
            ("Drawing", str(files.get("drawing") or "Not captured")),
        )
        for row_index, values in enumerate(rows):
            for column, value in enumerate(values):
                table.setItem(row_index, column, QTableWidgetItem(value))
        layout.addWidget(table)
        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        dialog.exec_()

    def _show_structure_context_menu(self, position) -> None:
        tree = self.sender()
        if not isinstance(tree, QTreeWidget):
            return
        item = tree.itemAt(position)
        if item is None:
            return
        menu = QMenu(tree)
        entries = (
            ("Current", STRUCTURE_CURRENT_ITERATION_ROLE, item.text(5)),
            ("Bound", STRUCTURE_BOUND_ITERATION_ROLE, item.text(6)),
            ("Latest", STRUCTURE_LATEST_ITERATION_ROLE, item.text(7)),
        )
        for label, role, version in entries:
            iteration_id = item.data(0, role)
            if iteration_id is None:
                continue
            action = menu.addAction(
                f"View {label} CAD Files ({version})" if str(version or "").strip()
                else f"View {label} CAD Files"
            )
            action.triggered.connect(
                lambda _=False, iid=int(iteration_id), relation=label, ver=str(version or ""):
                self._show_iteration_cad_files(iid, relation, ver)
            )
        part_id = item.data(0, Qt.UserRole)
        is_assembly = str(item.text(4) or "").strip().lower() in {"asm", "assembly"}
        if part_id is not None and is_assembly:
            if menu.actions():
                menu.addSeparator()
            compare_action = menu.addAction("Compare Assembly Iterations...")
            compare_action.triggered.connect(
                lambda _=False, pid=int(part_id): self.compare_assembly_iterations(pid)
            )
        if menu.actions():
            menu.exec_(tree.viewport().mapToGlobal(position))

    def _open_structure_tree_item(self, item: QTreeWidgetItem, _column: int) -> None:
        part_id = item.data(0, Qt.UserRole) if item is not None else None
        if part_id is None:
            return
        self._select_part_item(int(part_id))
        self.display_details(int(part_id))

    def _on_tree_item_double_clicked(self, item, column):
        if item is not None and item.data(0, PDM_OBJECT_KIND_ROLE) == PDM_OBJECT_CAD:
            self._show_pdm_cad_selection(item)
            return
        item_id = item.data(0, Qt.UserRole)
        if not item_id:
            return
        if int(column) == BOM_COL_FILES:
            tree = item.treeWidget() or self.tree
            payload = item.data(BOM_COL_FILES, BOM_TREE_FILES_ROLE) or {}
            row_rect = tree.visualItemRect(item)
            column_rect = QRect(
                tree.header().sectionViewportPosition(BOM_COL_FILES),
                row_rect.top(),
                tree.header().sectionSize(BOM_COL_FILES),
                row_rect.height(),
            )
            pdf_rect, step_rect = _files_delegate_pill_rects(column_rect, payload)
            cursor_pos = tree.viewport().mapFromGlobal(QCursor.pos())
            if step_rect and step_rect.contains(cursor_pos):
                self.open_part_step(int(item_id))
            else:
                self.preview_part_pdf(int(item_id))
            return
        self.issue_requested.emit(int(item_id))

    def display_details(self, item_id):
        if item_id is None:
            return
        self._selected_pdm_kind = PDM_OBJECT_ITEM
        self.current_cad_document_id = None
        self.current_associated_item_id = None
        self._current_cad_payload = {}
        self._set_summary_field_titles(self._item_summary_field_titles())
        self.checkout_part_btn.setToolTip("Check out the selected Item only")
        self.checkin_part_btn.setToolTip("Review and finish the selected Item checkout")
        self.undo_checkout_btn.setToolTip("Undo the selected Item checkout")
        self.current_folder_id = None
        self.details_summary_card.show()
        self.associated_files_card.show()
        self._set_engineering_item_actions_enabled(True)
        self.current_part_id = item_id
        details = self.bom_service.get_part_details(item_id) or {}
        try:
            structure_context = self.bom_service.get_structure_context(int(item_id)) or {}
        except Exception:
            structure_context = {}
        # Detailed unified history across revisions
        history = self.bom_service.get_history_detailed(item_id, include_all_revisions=True) or []
        analytics = self.bom_service.get_history_analytics(item_id, include_all_revisions=True) or {}

        self.refresh_files_tab()

        if self.missing_files:
            missing_msgs = []
            for bom_id, issue_type, filename in self.missing_files:
                if bom_id == item_id:
                    if issue_type == 'missing_file':
                        missing_msgs.append(f"⚠️ Missing file: {filename} ")
                    elif issue_type == 'outdated_file':
                        missing_msgs.append(f"⚠️ Outdated file : {filename} is not the latest version in working directory.")
                    elif issue_type == 'missing_drawing':
                        missing_msgs.append(f"⚠️ Missing drawing: {filename}")
                    elif issue_type == 'missing_pdf':
                        missing_msgs.append(f"⚠️ Missing PDF: {filename}")
                    elif issue_type == 'missing_step':
                        missing_msgs.append(f"⚠️ Missing STEP: {filename}")
            if missing_msgs:
                self.show_alert(" | ".join(missing_msgs), "warning", "details")
            else:
                self.hide_alert("details")
        else:
            self.hide_alert("details")

        self._update_details_summary(details)

        self._populate_structure_views(structure_context)

        # History panel (genius panel)
        self._history_rows = list(history)
        self.history_panel.set_data(history, analytics)

        self.notes_view.setText(details.get("notes", ""))
        self.refresh_issues_tab()
        self._sync_visual_action_states()

    def clear_details(self):
        self.current_part_id = None
        self.current_cad_document_id = None
        self.current_associated_item_id = None
        self._current_cad_payload = {}
        self._selected_pdm_kind = None
        self._set_summary_field_titles(self._item_summary_field_titles())
        self._update_details_summary({})
        self.uses_tree.clear()
        self.where_used_tree.clear()
        self.effective_where_used_tree.clear()
        self.structure_summary_label.setText("Select a BOM item")
        self.structure_views.setTabText(0, "Uses")
        self.structure_views.setTabText(1, "CAD Where Used")
        self.structure_views.setTabText(2, "Effective EBOM Where Used")
        self.notes_view.clear()
        try:
            self.part_issues_table.setRowCount(0)
        except Exception:
            pass
        try:
            self.history_panel.clear()
            self._history_rows = []
        except Exception:
            pass
        self.refresh_files_tab()
        self._sync_visual_action_states()

    @staticmethod
    def _details_card_label(text: str, object_name: str) -> QLabel:
        label = QLabel(text)
        label.setObjectName(object_name)
        return label

    @staticmethod
    def _details_value(details: dict, keys) -> str:
        for key in keys:
            value = details.get(key)
            if value is not None and str(value).strip():
                return str(value)
        return ""

    def _update_details_summary(self, details: dict) -> None:
        self._current_part_details = dict(details or {})
        has_details = bool(self._current_part_details)
        self.view_full_details_btn.setEnabled(has_details)
        self.edit_categories_btn.setEnabled(has_details and bool(getattr(self, "current_part_id", None)))

        name = self._details_value(self._current_part_details, ("name", "part_number", "aes_number"))
        self.details_name_label.setText(name or "Select a BOM item")

        identity_parts = []
        aes_number = self._details_value(self._current_part_details, ("aes_number",))
        part_number = self._details_value(self._current_part_details, ("part_number",))
        part_type = self._details_value(self._current_part_details, ("type",))
        current_version = self._details_value(self._current_part_details, ("current_version", "revision"))
        if part_number:
            identity_parts.append(f"NUMBER {part_number}")
        if aes_number:
            identity_parts.append(f"AES {aes_number}")
        if part_type:
            identity_parts.append(part_type.upper())
        if current_version:
            identity_parts.append(f"Version {current_version}")
        classification = str(
            self._current_part_details.get("classification") or "PHYSICAL"
        ).upper()
        default_behavior = str(
            self._current_part_details.get("default_ebom_behavior") or "NORMAL"
        ).upper()
        if classification != "PHYSICAL":
            identity_parts.append(classification.replace("_", " "))
        if default_behavior == "EXCLUDE":
            identity_parts.append("Default: NOT FOR DELIVERY")
        elif default_behavior == "FLATTEN":
            identity_parts.append("Default: FLATTEN IN EBOM")
        if self._current_part_details.get("represented_part_id"):
            target_label = self._details_value(
                self._current_part_details,
                (
                    "represented_part_number",
                    "represented_part_name",
                    "represented_part_aes",
                ),
            )
            identity_parts.append(
                f"CAD REP OF {target_label or ('#' + str(self._current_part_details['represented_part_id']))}"
            )
        if str(
            self._current_part_details.get("cad_control_mode") or "CONTROLLED"
        ).upper() == "SUPPLIER_PACKAGE":
            dependency_count = int(
                self._current_part_details.get("cad_dependency_count") or 0
            )
            identity_parts.append(f"SUPPLIER PACKAGE · {dependency_count} CAD FILES")
        self.details_identity_label.setText(
            "  |  ".join(identity_parts)
            if identity_parts else "Select a row in the BOM structure to view its summary."
        )

        cad_file = self._details_value(
            self._current_part_details, ("filename", "base_file_name")
        )
        drawing_file = self._details_value(
            self._current_part_details, ("drawing", "base_drw_name", "drawing_number")
        )
        try:
            pdm_documents = self.bom_service.list_item_cad_associations(
                int(self.current_part_id)
            ) if self.current_part_id is not None else []
        except Exception:
            pdm_documents = []
        model_links = [
            self._cad_representation_summary_line(row)
            for row in pdm_documents
            if str(row.get("category") or "").upper() != "DRAWING"
        ]
        drawing_links = []
        listed_drawing_ids = set()
        for row in pdm_documents:
            if str(row.get("category") or "").upper() != "DRAWING":
                continue
            drawing_id = row.get("id")
            if drawing_id is not None:
                listed_drawing_ids.add(int(drawing_id))
            role = "PRIMARY" if row.get("is_primary_drawing") else str(
                row.get("association_type") or "CONTENT"
            ).upper().replace("_", " ")
            drawing_links.append(
                f"{row.get('file_name') or row.get('name') or 'CAD Drawing'} "
                f"[{role}; {self._pdm_cad_revision_text(row)}]"
            )
        for model in pdm_documents:
            if str(model.get("category") or "").upper() == "DRAWING":
                continue
            for drawing in self._pdm_selected_drawings(model):
                drawing_id = drawing.get("id")
                if drawing_id is not None and int(drawing_id) in listed_drawing_ids:
                    continue
                label = (
                    f"{drawing.get('file_name') or drawing.get('name') or 'CAD Drawing'} "
                    f"[{'PRIMARY' if drawing.get('is_primary_drawing') else 'SUPPORTING'}; "
                    f"MODEL {model.get('file_name') or model.get('name') or 'CAD Model'}; "
                    f"{self._pdm_cad_revision_text(drawing)}]"
                )
                drawing_links.append(label)
                if drawing_id is not None:
                    listed_drawing_ids.add(int(drawing_id))
        self.associated_cad_file_label.setText(
            "\n".join(model_links) if model_links else (cad_file or "Not linked")
        )
        drawing_display = (
            "\n".join(drawing_links)
            if drawing_links else (
                "No drawing assigned to this Item"
                if model_links else (drawing_file or "Not linked")
            )
        )
        self.associated_drawing_file_label.setText(
            drawing_display
        )

        for _field_key, (field_label, value_label, detail_keys) in self._details_summary_fields.items():
            value = self._details_value(self._current_part_details, detail_keys)
            if _field_key in {"item_type", "source", "view"} and value:
                value = value.replace("_", " ").title()
            elif _field_key == "unit" and value.upper() == "EA":
                value = "each"
            field_label.show()
            value_label.show()
            value_label.setText(value)
        self._update_lifecycle_action_states(self._current_part_details)

    def _update_lifecycle_action_states(self, details: dict) -> None:
        has_item = bool(details and getattr(self, "current_part_id", None))
        locked = bool(details.get("locked")) if has_item else False
        state = str(
            details.get("revision_state") or details.get("lifecycle_state") or ""
        ).strip().lower()
        released = state == "released"
        obsolete = state == "obsolete"
        pending_revision = str(details.get("pending_revision_code") or "").strip()
        editable_checkout = not released or bool(pending_revision)
        is_assembly = str(details.get("type") or "").strip().lower() in {"asm", "assembly"}
        can_manage = self.perm.can("manage_parts")
        can_revision = self.perm.can("set_revision")
        # Item Structure is an engineering workspace, not a read-only report.
        # Item locks remain independent from CAD Document locks.
        read_only_ebom = False
        try:
            active_cad = (
                self.bom_service.checked_out_cad_for_item(int(self.current_part_id))
                if has_item else []
            )
        except Exception:
            active_cad = []
        checkout_origin = str(details.get("checkout_origin") or "ITEM").upper()
        lock_is_mine = (
            details.get("locked_by_user_id") is None
            or self.session.user_id is None
            or int(details.get("locked_by_user_id")) == int(self.session.user_id)
        )
        can_promote_auto_checkout = bool(
            locked and checkout_origin == "CAD" and lock_is_mine
        )
        can_finish_checkout = bool(lock_is_mine or self.perm.can("merge"))
        self.checkout_part_btn.setEnabled(
            has_item and (not locked or can_promote_auto_checkout)
            and not obsolete and not read_only_ebom
        )
        self.checkin_part_btn.setEnabled(
            has_item and locked and can_finish_checkout and not read_only_ebom
        )
        self.undo_checkout_btn.setEnabled(
            has_item and locked and can_finish_checkout and not read_only_ebom
        )
        if active_cad:
            labels = ", ".join(
                str(row.get("file_name") or row.get("name") or "CAD")
                for row in active_cad[:3]
            )
            self.checkin_part_btn.setToolTip(
                f"Check in Item data only; associated CAD stays checked out: {labels}"
            )
            self.undo_checkout_btn.setToolTip(
                f"Undo Item checkout and automatically undo associated CAD: {labels}"
            )
        elif can_promote_auto_checkout:
            self.checkout_part_btn.setToolTip(
                "Convert the automatic CAD-origin reservation into an explicit Item checkout"
            )
        self.edit_part_btn.setEnabled(
            has_item and locked and can_finish_checkout and can_manage
            and editable_checkout and not read_only_ebom
        )
        self.delete_part_btn.setEnabled(
            has_item and can_manage and editable_checkout and not read_only_ebom
        )
        self.add_child_btn.setEnabled(
            has_item and locked and can_finish_checkout and can_manage and is_assembly
            and editable_checkout and not read_only_ebom
        )
        self.compare_iterations_btn.setEnabled(has_item and is_assembly)
        self.create_configuration_btn.setEnabled(
            has_item and is_assembly and not read_only_ebom
        )
        self.update_child_versions_btn.setEnabled(
            has_item and locked and can_finish_checkout and is_assembly
            and editable_checkout and not read_only_ebom
        )
        self.set_revision_btn.setEnabled(
            has_item and not locked and released and not pending_revision
            and can_revision and not read_only_ebom
        )
        self.release_revision_btn.setEnabled(
            has_item and not locked and not released and can_revision and not read_only_ebom
        )
        self._sync_action_ribbon_menus()

    def _edit_current_part_categories(self) -> None:
        part_id = getattr(self, "current_part_id", None)
        if not part_id:
            return

        try:
            available_categories = self.bom_service.list_categories()
            assigned_categories = set(self.bom_service.categories_for_part(int(part_id)))
        except Exception as exc:
            QMessageBox.critical(self, "Categories", f"Could not load categories:\n{exc}")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Assign Categories")
        dialog.resize(440, 430)
        layout = QVBoxLayout(dialog)
        layout.setSpacing(8)
        layout.addWidget(QLabel("Assign this BOM item to one or more project categories."))

        category_list = QListWidget()
        category_list.setAlternatingRowColors(True)

        def add_category_item(name: str, checked: bool = False, category_id=None) -> None:
            clean_name = " ".join(str(name or "").split())
            if not clean_name:
                return
            for row in range(category_list.count()):
                if category_list.item(row).text().casefold() == clean_name.casefold():
                    if checked:
                        category_list.item(row).setCheckState(Qt.Checked)
                    if category_id is not None:
                        category_list.item(row).setData(Qt.UserRole, int(category_id))
                    return
            item = QListWidgetItem(clean_name)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if checked else Qt.Unchecked)
            item.setData(Qt.UserRole, int(category_id) if category_id is not None else None)
            category_list.addItem(item)

        for category in available_categories:
            name = str(category.get("name") or "")
            add_category_item(name, name in assigned_categories, category.get("id"))
        layout.addWidget(category_list, 1)

        def confirm_assigned_category_deletion(category_name: str, parts: list) -> bool:
            confirmation = QDialog(dialog)
            confirmation.setWindowTitle("Delete Assigned Category")
            confirmation.resize(620, 430)
            confirmation_layout = QVBoxLayout(confirmation)
            warning = QLabel(
                f"Category '{category_name}' is assigned to {len(parts)} BOM item(s).\n"
                "Deleting it will remove the category from every item listed below."
            )
            warning.setWordWrap(True)
            warning.setStyleSheet(
                "color:#991b1b;background:#fee2e2;border:1px solid #f87171;"
                "border-radius:5px;padding:8px;font-weight:700;"
            )
            confirmation_layout.addWidget(warning)

            parts_table = QTableWidget()
            parts_table.setColumnCount(4)
            parts_table.setHorizontalHeaderLabels(["Number", "Name", "AES Number", "Type"])
            parts_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            for column in (0, 2, 3):
                parts_table.horizontalHeader().setSectionResizeMode(column, QHeaderView.ResizeToContents)
            parts_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            parts_table.setSelectionBehavior(QAbstractItemView.SelectRows)
            parts_table.setRowCount(len(parts))
            for row, part in enumerate(parts):
                values = (
                    part.get("part_number") or "",
                    part.get("name") or "",
                    part.get("aes_number") or "",
                    part.get("type") or "",
                )
                for column, value in enumerate(values):
                    parts_table.setItem(row, column, QTableWidgetItem(str(value)))
            confirmation_layout.addWidget(parts_table, 1)

            confirmation_buttons = QDialogButtonBox()
            delete_button = confirmation_buttons.addButton(
                "Delete Category", QDialogButtonBox.DestructiveRole
            )
            cancel_button = confirmation_buttons.addButton(QDialogButtonBox.Cancel)
            delete_button.setObjectName("danger")
            delete_button.clicked.connect(confirmation.accept)
            cancel_button.clicked.connect(confirmation.reject)
            confirmation_layout.addWidget(confirmation_buttons)
            return confirmation.exec_() == QDialog.Accepted

        def delete_category_item(item: QListWidgetItem) -> None:
            if item is None:
                return
            category_name = item.text()
            category_id = item.data(Qt.UserRole)
            if category_id is None:
                category_list.takeItem(category_list.row(item))
                return
            try:
                usage = self.bom_service.category_usage(int(category_id))
                parts = list(usage.get("parts") or [])
            except Exception as exc:
                QMessageBox.critical(dialog, "Delete Category", f"Could not inspect category usage:\n{exc}")
                return

            if parts:
                confirmed = confirm_assigned_category_deletion(category_name, parts)
            else:
                confirmed = QMessageBox.question(
                    dialog,
                    "Delete Category",
                    f"Delete the empty project category '{category_name}'?",
                    QMessageBox.Yes | QMessageBox.Cancel,
                    QMessageBox.Cancel,
                ) == QMessageBox.Yes
            if not confirmed:
                return

            try:
                result = self.bom_service.delete_category(int(category_id))
                affected_parts = list(result.get("parts") or [])
                category_list.takeItem(category_list.row(item))
                deleted_key = category_name.casefold()
                for part in affected_parts:
                    affected_id = int(part["id"])
                    for tree_item in self._find_tree_items(affected_id):
                        remaining = [
                            name for name in (tree_item.data(0, BOM_TREE_CATEGORY_ROLE) or [])
                            if str(name).casefold() != deleted_key
                        ]
                        tree_item.setData(0, BOM_TREE_CATEGORY_ROLE, remaining)

                if any(int(part["id"]) == int(part_id) for part in affected_parts):
                    assigned = self.bom_service.categories_for_part(int(part_id))
                    self._current_part_details["categories"] = ", ".join(assigned)
                    self._update_details_summary(self._current_part_details)

                active_categories = list(self._bom_advanced_filters.get("categories") or [])
                filtered_categories = [
                    name for name in active_categories if str(name).casefold() != deleted_key
                ]
                if filtered_categories != active_categories:
                    self._bom_advanced_filters["categories"] = filtered_categories
                    self.apply_bom_tree_filter(self._bom_advanced_filters)
            except Exception as exc:
                QMessageBox.critical(dialog, "Delete Category", f"Could not delete category:\n{exc}")

        def show_category_context_menu(position) -> None:
            item = category_list.itemAt(position)
            if item is None:
                return
            menu = QMenu(category_list)
            delete_action = menu.addAction("Delete Category")
            delete_action.triggered.connect(lambda: delete_category_item(item))
            menu.exec_(category_list.viewport().mapToGlobal(position))

        category_list.setContextMenuPolicy(Qt.CustomContextMenu)
        category_list.customContextMenuRequested.connect(show_category_context_menu)

        new_category_row = QHBoxLayout()
        new_category_input = QLineEdit()
        new_category_input.setPlaceholderText("New project category")
        add_category_btn = QPushButton("Add Category")

        def add_new_category() -> None:
            name = new_category_input.text()
            if not name.strip():
                return
            add_category_item(name, checked=True)
            new_category_input.clear()

        add_category_btn.clicked.connect(add_new_category)
        new_category_input.returnPressed.connect(add_new_category)
        new_category_row.addWidget(new_category_input, 1)
        new_category_row.addWidget(add_category_btn)
        layout.addLayout(new_category_row)

        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)

        def save_categories() -> None:
            add_new_category()
            names = [
                category_list.item(row).text()
                for row in range(category_list.count())
                if category_list.item(row).checkState() == Qt.Checked
            ]
            try:
                assigned = self.bom_service.set_part_categories(int(part_id), names)
                self._current_part_details["categories"] = ", ".join(assigned)
                self._update_details_summary(self._current_part_details)
                for item in self._find_tree_items(int(part_id)):
                    item.setData(0, BOM_TREE_CATEGORY_ROLE, list(assigned))
                if not self._is_default_bom_advanced_filter():
                    self.apply_bom_tree_filter(self._bom_advanced_filters)
                dialog.accept()
            except Exception as exc:
                QMessageBox.critical(dialog, "Categories", f"Could not save categories:\n{exc}")

        buttons.accepted.connect(save_categories)
        buttons.rejected.connect(dialog.reject)
        dialog.exec_()

    def _open_full_bom_details(self) -> None:
        details = dict(getattr(self, "_current_part_details", {}) or {})
        if not details:
            return

        number = self._details_value(details, ("part_number",))
        name = self._details_value(details, ("name",))
        title = " — ".join(value for value in (number, name) if value) or "Item"
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Full Details - {title}")
        dlg.resize(820, 560)

        layout = QVBoxLayout(dlg)
        heading = QLabel(title)
        heading.setStyleSheet("font-size:13px;font-weight:700;color:#172033;")
        layout.addWidget(heading)

        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Property", "Value"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setWordWrap(True)
        table.setRowCount(len(details))
        for row, (key, value) in enumerate(details.items()):
            table.setItem(row, 0, QTableWidgetItem(str(key)))
            table.setItem(row, 1, QTableWidgetItem(str(value if value is not None else "")))
        table.resizeRowsToContents()
        layout.addWidget(table, 1)

        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(dlg.reject)
        buttons.accepted.connect(dlg.accept)
        layout.addWidget(buttons)
        dlg.exec_()

    def refresh_issues_tab(self):
        if not getattr(self, "current_part_id", None):
            self.part_issues_table.setRowCount(0)
            return
        try:
            issues = self.issue_service.issues_for_part(int(self.current_part_id))
        except Exception:
            issues = []
        self.part_issues_table.setRowCount(len(issues))
        for row, issue in enumerate(issues):
            values = [issue["issue_number"], issue["title"], issue["status"], issue["priority"]]
            for col, value in enumerate(values):
                cell = QTableWidgetItem(str(value))
                cell.setData(Qt.UserRole, int(issue["id"]))
                if col == 3:
                    cell.setForeground(QBrush(QColor("#b91c1c" if issue["priority"] == "Critical" else "#374151")))
                self.part_issues_table.setItem(row, col, cell)

    def _open_current_part_issues(self):
        if getattr(self, "current_part_id", None):
            self.issue_requested.emit(int(self.current_part_id))

    def _create_issue_for_current_part(self):
        if getattr(self, "current_part_id", None):
            self.create_issue_requested.emit(int(self.current_part_id))

    # ═══════════════════════════════════════════════════════════════════════
    #  History Panel Callbacks
    # ═══════════════════════════════════════════════════════════════════════

    def _open_history_details_dialog(self, ev_data: dict):
        """Open a rich details dialog for a history event.
        Called by HistoryPanel.open_details_requested signal with the full event dict."""
        if not ev_data:
            return

        ev_type = str(ev_data.get("event", "") or "")
        commit_id = str(ev_data.get("commit_id", "") or "")
        if ev_type.upper() == "COMMIT" and commit_id:
            try:
                self._open_commit_history_details_dialog(ev_data, commit_id)
                return
            except Exception as exc:
                QMessageBox.warning(self, "Commit Details", f"Could not load full commit details:\n{exc}")

        ts = str(ev_data.get("timestamp", "") or "")
        user = str(ev_data.get("user", "") or "")
        details_text = str(ev_data.get("details", "") or "")
        style = _style_for(ev_type)

        dlg = QDialog(self)
        dlg.setWindowTitle(f"{style['icon']}  History Details — {style['label']}")
        dlg.setMinimumSize(740, 480)
        layout = QVBoxLayout(dlg)
        layout.setSpacing(12)

        # ── Header card ───────────────────────────────────────────────
        header = QFrame()
        header.setStyleSheet(f"""
            QFrame {{
                background: {style['bg']};
                border: 1px solid {style['color']}40;
                border-left: 5px solid {style['color']};
                border-radius: 8px;
                padding: 10px;
            }}
        """)
        h_lay = QVBoxLayout(header)
        h_lay.setSpacing(4)

        badge_lbl = QLabel(f'  {style["icon"]}  {style["label"].upper()}  ')
        badge_lbl.setStyleSheet(f"""
            background: {style['color']}; color: #ffffff;
            border-radius: 10px; padding: 3px 12px;
            font-size: 12px; font-weight: bold;
        """)
        badge_lbl.setFixedWidth(badge_lbl.sizeHint().width() + 16)
        h_lay.addWidget(badge_lbl)

        rel = _relative_time(ts)
        ts_lbl = QLabel(f"📅  {ts}    ({rel})" if rel else f"📅  {ts}")
        ts_lbl.setStyleSheet(f"color: {style['color']}; font-size: 13px; font-weight: 600; border: none; background: transparent;")
        h_lay.addWidget(ts_lbl)

        info_parts = []
        if user:
            info_parts.append(f"👤 {user}")
        object_version = str(ev_data.get("object_version", "") or "")
        if object_version:
            info_parts.append(f"Revision / iteration: {object_version}")
        proj = str(ev_data.get("project", "") or "")
        ver = str(ev_data.get("version", "") or "")
        if proj:
            info_parts.append(f"📁 {proj}" + (f" ({ver})" if ver else ""))
        commit_id = str(ev_data.get("commit_id", "") or "")
        
        if commit_id:
            info_parts.append(f"🏷 {commit_id[:16]}")
        if info_parts:
            info_lbl = QLabel("    ".join(info_parts))
            info_lbl.setStyleSheet("color: #4b5563; font-size: 12px; border: none; background: transparent;")
            h_lay.addWidget(info_lbl)
        commit_unique_id = str(ev_data.get("commit_unique_id", "") or "")
        db_commit_info = None
        if commit_unique_id and commit_unique_id != "":
            try:
                project_id = getattr(self.session, "project_id", None)
                if project_id:
                    db_commit = self.commit_repo.get_commit_by_commitid(commit_unique_id, project_id)
                    if db_commit:
                        db_commit_info = db_commit
            except Exception:
                db_commit_info = None


        # --- Add extra commit-related info ---
        shown_keys = {"timestamp", "event", "object_version", "user", "project", "version", "details", "commit_id", "step_diff_status", "step_diff_summary", "step_error", "step_file_path", "step_prev_file_path"}
        extra_commit_info = []
        for k, v in (ev_data.items() if isinstance(ev_data, dict) else []):
            if k in shown_keys:
                continue
            if v is not None and v != "":
                extra_commit_info.append(f"<b>{k}</b>: {v}")
        if extra_commit_info:
            extra_lbl = QLabel("<br>".join(extra_commit_info))
            extra_lbl.setStyleSheet("color: #374151; font-size: 11px; border: none; background: transparent;")
            extra_lbl.setTextFormat(Qt.RichText)
            h_lay.addWidget(extra_lbl)
            
        if db_commit_info:
            from dataclasses import asdict
            shown_keys = {"id", "commit_id", "part_id", "project_id", "message"}
            extra_commit_info = []
            for k, v in asdict(db_commit_info).items():
                if k in shown_keys:
                    continue
                if v is not None and v != "":
                    extra_commit_info.append(f"<b>{k}</b>: {v}")
            if extra_commit_info:
                extra_lbl = QLabel("<br>".join(extra_commit_info))
                extra_lbl.setStyleSheet("color: #374151; font-size: 11px; border: none; background: transparent;")
                extra_lbl.setTextFormat(Qt.RichText)
                h_lay.addWidget(extra_lbl)
        

        layout.addWidget(header)

        # ── Details text ──────────────────────────────────────────────
        layout.addWidget(QLabel("Details:"))
        step_status = str(ev_data.get("step_diff_status") or "").strip()
        step_summary = str(ev_data.get("step_diff_summary") or "").strip()
        step_error = str(ev_data.get("step_error") or "").strip()
        extra = []
        if step_status:
            extra.append(f"STEP status: {step_status}")
        if step_summary:
            extra.append(f"STEP summary:\n{step_summary}")
        if step_error:
            extra.append(f"STEP error: {step_error}")
        full_text = details_text if not extra else f"{details_text}\n\n{'─' * 40}\n" + "\n".join(extra)
        if looks_like_html(full_text):
            txt = QTextEdit()
            txt.setReadOnly(True)
            txt.setStyleSheet("""
                QTextEdit {
                    background: #f9fafb; border: 1px solid #e5e7eb;
                    border-radius: 6px; font-size: 11px; padding: 8px;
                }
            """)
            txt.setHtml(full_text)
        else:
            txt = QPlainTextEdit()
            txt.setReadOnly(True)
            txt.setStyleSheet("""
                QPlainTextEdit {
                    background: #f9fafb; border: 1px solid #e5e7eb;
                    border-radius: 6px; font-family: 'Consolas', 'Cascadia Mono', monospace;
                    font-size: 11px; padding: 8px;
                }
            """)
            txt.setPlainText(full_text)
        layout.addWidget(txt)

        # ── Action buttons ────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)

        if step_status and ev_type.upper() == "COMMIT":
            step_path = str(ev_data.get("step_file_path") or "").strip()
            if step_path:
                open_step = QPushButton("🔬  Open STEP in Viewer")
                open_step.setObjectName("neutral")
                open_step.clicked.connect(lambda: self._open_associated_step_for_history_event(ev_data))
                btn_row.addWidget(open_step)

            if step_status.upper() == "COMPARED":
                diff_btn = QPushButton("🔍  Show STEP Diff Zones")
                diff_btn.setObjectName("primary")
                diff_btn.clicked.connect(lambda: self._show_step_diff_for_history_event(ev_data))
                btn_row.addWidget(diff_btn)

        # Copy button
        copy_btn = QPushButton("📋  Copy to Clipboard")
        copy_btn.setObjectName("neutral")
        copy_btn.clicked.connect(lambda: self._copy_event_to_clipboard(ev_data))
        btn_row.addWidget(copy_btn)

        btn_row.addStretch()

        close_btn = QPushButton("Close")
        close_btn.setObjectName("neutral")
        close_btn.clicked.connect(dlg.close)
        btn_row.addWidget(close_btn)

        layout.addLayout(btn_row)
        dlg.exec_()

    def _open_commit_history_details_dialog(self, ev_data: dict, commit_id: str):
        project_id = getattr(self.session, "project_id", None)
        details = CommitService().get_commit_group_details(
            str(commit_id),
            int(project_id) if project_id is not None else None,
        )
        files = details.get("files") or []
        if not files:
            raise ValueError(f"No commit rows found for {commit_id}.")

        first = files[0]
        status = details.get("status") or first.get("status") or ""
        style = _style_for("COMMIT")
        if str(status).lower() == "approved":
            style = {"icon": "✅", "label": "Approved", "color": "#16a34a", "bg": "#dcfce7"}
        elif str(status).lower() == "validated":
            style = {"icon": "🔵", "label": "Validated", "color": "#2563eb", "bg": "#dbeafe"}
        elif str(status).lower() == "pending":
            style = {"icon": "🟡", "label": "Pending", "color": "#ca8a04", "bg": "#fef9c3"}

        dlg = QDialog(self)
        dlg.setWindowTitle(f"{style['icon']}  Commit Details — {details.get('title') or commit_id}")
        dlg.setMinimumSize(900, 680)
        screen = QApplication.primaryScreen().availableGeometry()
        dlg.resize(min(1040, int(screen.width() * 0.88)), min(820, int(screen.height() * 0.88)))
        root = QVBoxLayout(dlg)
        root.setContentsMargins(10, 10, 10, 10)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setStyleSheet("QScrollArea { border: none; }")
        content = QWidget()
        layout = QVBoxLayout(content)
        layout.setContentsMargins(4, 4, 12, 4)
        layout.setSpacing(12)

        header = QFrame()
        header.setStyleSheet(f"""
            QFrame {{
                background: {style['bg']};
                border: 1px solid {style['color']}40;
                border-left: 5px solid {style['color']};
                border-radius: 8px;
            }}
        """)
        h_lay = QVBoxLayout(header)
        h_lay.setSpacing(5)
        badge = QLabel(f"  {style['icon']}  {style['label'].upper()}  ")
        badge.setStyleSheet(f"""
            background: {style['color']}; color: #ffffff;
            border-radius: 10px; padding: 3px 12px;
            font-size: 12px; font-weight: bold;
        """)
        badge.setFixedWidth(badge.sizeHint().width() + 20)
        h_lay.addWidget(badge)
        h_lay.addWidget(QLabel(f"<b style='font-size:13px'>{details.get('title') or first.get('filename') or commit_id}</b>"))
        meta = []
        if details.get("author"):
            meta.append(f"👤 {details.get('author')}")
        if details.get("checker"):
            meta.append(f"🔍 {details.get('checker')}")
        if details.get("committed_at"):
            meta.append(f"🗓 {str(details.get('committed_at'))[:19]} ({_relative_time(str(details.get('committed_at')))})")
        meta.append(f"🏷 {commit_id[:16]}")
        meta_lbl = QLabel("    ".join(meta))
        meta_lbl.setWordWrap(True)
        meta_lbl.setStyleSheet("color:#4b5563;font-size:11px;border:none;background:transparent;")
        h_lay.addWidget(meta_lbl)
        layout.addWidget(header)

        message = details.get("message") or ev_data.get("details") or ""
        if message:
            msg = QTextEdit()
            msg.setReadOnly(True)
            msg.setMinimumHeight(140)
            msg.setMaximumHeight(280)
            if looks_like_html(str(message)):
                msg.setHtml(str(message))
            else:
                msg.setPlainText(str(message))
            layout.addWidget(msg)

        display_details = dict(details)
        display_details["object_version"] = ev_data.get("object_version", "")
        self._add_commit_key_value_table(layout, display_details, first)
        self._add_commit_files_tree(layout, files)
        self._add_commit_issues_table(layout, details.get("issues") or [])
        self._add_commit_doc_table(layout, details.get("engineering_files") or [], "Vaulted Engineering Outputs")
        self._add_commit_doc_table(layout, details.get("validation_docs") or [], "Validation Documents")

        scroll.setWidget(content)
        root.addWidget(scroll, 1)

        buttons = QHBoxLayout()
        copy_btn = QPushButton("📋  Copy to Clipboard")
        copy_btn.setObjectName("neutral")
        copy_btn.clicked.connect(lambda: QApplication.clipboard().setText(json.dumps(details, indent=2, default=str)))
        buttons.addWidget(copy_btn)
        buttons.addStretch()
        close_btn = QPushButton("Close")
        close_btn.setObjectName("neutral")
        close_btn.clicked.connect(dlg.accept)
        buttons.addWidget(close_btn)
        root.addLayout(buttons)
        dlg.exec_()

    def _add_commit_key_value_table(self, layout, details: dict, first: dict):
        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Field", "Value"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.setMinimumHeight(190)
        table.setMaximumHeight(330)
        rows = [
            ("Status", details.get("status")),
            ("Commit ID", details.get("commit_id")),
            ("Project", details.get("project_name")),
            ("Project Version", details.get("project_version_label")),
            ("Revision / Iteration", details.get("object_version")),
            ("Author", details.get("author")),
            ("Designer", details.get("designer")),
            ("Checker", details.get("checker")),
            ("Committed At", details.get("committed_at")),
            ("Merged By", details.get("merged_by")),
            ("Merged At", details.get("merged_at")),
            ("Merge ID", details.get("merge_id")),
            ("Merge Message", details.get("merge_message")),
            ("Approved Version", details.get("approved_version")),
            ("PR Path", details.get("pr_path")),
            ("Signature", details.get("signature")),
            ("Item Number", first.get("part_number")),
            ("AES Number", first.get("aes_number")),
            ("Selected Item", first.get("part_name") or first.get("part_id")),
        ]
        rows = [(k, v) for k, v in rows if v not in (None, "")]
        table.setRowCount(len(rows))
        for r, (key, value) in enumerate(rows):
            key_item = QTableWidgetItem(str(key))
            key_item.setFont(QFont("Segoe UI", 10, QFont.Bold))
            table.setItem(r, 0, key_item)
            val_item = QTableWidgetItem(str(value))
            val_item.setToolTip(str(value))
            table.setItem(r, 1, val_item)
        layout.addWidget(table)

    def _add_commit_files_tree(self, layout, files: list):
        layout.addWidget(QLabel(f"<b>Files in this commit ({len(files)})</b>"))
        tree = QTreeWidget()
        tree.setColumnCount(2)
        tree.setHeaderLabels(["File / Field", "Value"])
        tree.header().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        tree.setAlternatingRowColors(True)
        tree.setMinimumHeight(260)
        for item in files:
            filename = str(item.get("filename") or "")
            top = QTreeWidgetItem([filename, f"Status: {item.get('status') or ''} | Approved version: {item.get('approved_version') or ''}"])
            font = top.font(0)
            font.setBold(True)
            top.setFont(0, font)
            tree.addTopLevelItem(top)
            for label, value in (
                ("Row ID", item.get("id")),
                ("Commit ID", item.get("commit_id")),
                ("Status", item.get("status")),
                ("Type", item.get("type")),
                ("Item", item.get("part_name") or item.get("part_id")),
                ("Item Number", item.get("part_number")),
                ("AES Number", item.get("aes_number")),
                ("Internal Item ID", item.get("part_id")),
                ("Revision", item.get("part_revision")),
                ("Source Path", item.get("file_path")),
                ("PR Path", item.get("pr_path")),
                ("Approved Version", item.get("approved_version")),
                ("Merged At", item.get("merged_at")),
                ("STEP Status", item.get("step_diff_status")),
                ("STEP File", item.get("step_file_path")),
                ("STEP Diff", item.get("step_diff_path")),
            ):
                if value not in (None, ""):
                    child = QTreeWidgetItem([str(label), str(value)])
                    child.setToolTip(1, str(value))
                    top.addChild(child)
            top.setExpanded(True)
        layout.addWidget(tree)

    def _add_commit_issues_table(self, layout, issues: list):
        if not issues:
            return
        layout.addWidget(QLabel(f"<b>Linked Issues ({len(issues)})</b>"))
        table = QTableWidget()
        table.setColumnCount(7)
        table.setHorizontalHeaderLabels(["Issue", "Title", "Status", "Priority", "Relation", "Validation", "Resolution"])
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setMinimumHeight(170)
        table.setMaximumHeight(300)
        table.setRowCount(len(issues))
        for r, issue in enumerate(issues):
            values = [
                issue.get("issue_number"), issue.get("title"), issue.get("status"),
                issue.get("priority"), issue.get("relation_type"),
                issue.get("validation_status"), issue.get("resolution_comment"),
            ]
            for c, value in enumerate(values):
                table.setItem(r, c, QTableWidgetItem(str(value or "")))
        layout.addWidget(table)

    def _add_commit_doc_table(self, layout, docs: list, title: str):
        if not docs:
            return
        layout.addWidget(QLabel(f"<b>{title} ({len(docs)})</b>"))
        table = QTableWidget()
        table.setColumnCount(8)
        table.setHorizontalHeaderLabels(["Role", "Type", "Filename", "Part", "Version", "Revision", "Exists", "Path"])
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(7, QHeaderView.Stretch)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setMinimumHeight(180)
        table.setMaximumHeight(320)
        table.setRowCount(len(docs))
        for r, doc in enumerate(docs):
            path = doc.get("source_path") or doc.get("stored_path") or ""
            exists = bool(path and safe_exists(path))
            values = [
                doc.get("doc_role") or doc.get("file_role"),
                doc.get("file_type"),
                doc.get("original_filename") or doc.get("filename") or doc.get("display_name"),
                doc.get("part_name") or doc.get("part_id"),
                doc.get("version_no") or doc.get("resolved_version_id"),
                doc.get("revision"),
                "Yes" if exists else "Missing",
                path,
            ]
            for c, value in enumerate(values):
                cell = QTableWidgetItem(str(value or ""))
                cell.setToolTip(str(value or ""))
                cell.setData(Qt.UserRole, path)
                table.setItem(r, c, cell)
        layout.addWidget(table)

        row = QHBoxLayout()
        row.addStretch()
        open_btn = QPushButton("Open Selected Doc")
        open_btn.setObjectName("neutral")
        open_btn.clicked.connect(lambda _c=False, t=table: self._open_selected_commit_doc(t))
        row.addWidget(open_btn)
        layout.addLayout(row)

    def _open_selected_commit_doc(self, table):
        row = table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Open Document", "Select a document first.")
            return
        item = table.item(row, table.columnCount() - 1)
        path = item.data(Qt.UserRole) if item else ""
        if not path or not safe_exists(path):
            QMessageBox.warning(self, "Open Document", f"File not found:\n{path}")
            return
        try:
            safe_startfile(path)
        except Exception as exc:
            QMessageBox.critical(self, "Open Document", f"Failed to open file:\n{exc}")

    def _copy_event_to_clipboard(self, ev_data: dict):
        from PyQt5.QtWidgets import QApplication
        lines = []
        for k in ("timestamp", "event", "user", "project", "version", "details",
                   "commit_id", "step_diff_status", "step_diff_summary"):
            v = ev_data.get(k)
            if v:
                lines.append(f"{k}: {v}")
        QApplication.clipboard().setText("\n".join(lines))

    def _open_associated_step_for_history_event(self, event_row: dict):
        step_path = str(event_row.get("step_file_path") or "").strip()
        if not step_path:
            QMessageBox.warning(self, "STEP Viewer", "No STEP file associated with this history row.")
            return
        if not os.path.exists(step_path):
            QMessageBox.warning(self, "STEP Viewer", f"STEP file not found:\n{step_path}")
            return

        try:
            from tools.CAD.step_viewer.launcher import launch_viewer
            launch_viewer(step_path)
        except Exception as e:
            QMessageBox.critical(self, "STEP Viewer", f"Failed to open STEP viewer:\n{e}")

    def _show_step_diff_for_history_event(self, event_row: dict):
        status = str(event_row.get("step_diff_status") or "").strip().upper()
        prev_path = str(event_row.get("step_prev_file_path") or "").strip()
        current_path = str(event_row.get("step_file_path") or "").strip()

        if status != "COMPARED":
            QMessageBox.information(
                self,
                "STEP Diff",
                "This commit has no previous STEP to compare (baseline or unavailable compare).",
            )
            return

        if not prev_path or not current_path:
            QMessageBox.warning(self, "STEP Diff", "STEP paths are missing for this commit.")
            return
        if not os.path.exists(prev_path) or not os.path.exists(current_path):
            QMessageBox.warning(self, "STEP Diff", "One or both STEP files are missing on disk.")
            return

        try:
            commit_id = str(event_row.get("commit_id") or "commit")
            from tools.CAD.step_viewer.launcher import launch_diff_viewer
            launch_diff_viewer(
                prev_path, current_path,
                commit_a=f"{commit_id}_prev", commit_b=commit_id,
            )
        except Exception as e:
            QMessageBox.critical(self, "STEP Diff", f"Failed to visualize STEP diff:\n{e}")

    def export_history_csv(self):
        rows = getattr(self, "_history_rows", []) or []
        if not rows:
            # also try from the panel
            try:
                rows = self.history_panel.get_all_rows()
            except Exception:
                pass
        if not rows:
            QMessageBox.information(self, "Export History", "No history rows to export.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Export History CSV", "history.csv", "CSV Files (*.csv)")
        if not path:
            return

        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["timestamp", "project", "version", "object_version", "event", "user", "details",
                            "commit_id", "step_diff_status", "step_diff_summary"])
                for ev in rows:
                    w.writerow(
                        [
                            ev.get("timestamp", ""),
                            ev.get("project", ""),
                            ev.get("version", ""),
                            ev.get("object_version", ""),
                            ev.get("event", ""),
                            ev.get("user", ""),
                            html_to_plain_text(str(ev.get("details", "") or "")),
                            ev.get("commit_id", ""),
                            ev.get("step_diff_status", ""),
                            ev.get("step_diff_summary", ""),
                        ]
                    )
            QMessageBox.information(self, "Export History", f"✅ Exported {len(rows)} events → {path}")
        except Exception as e:
            QMessageBox.critical(self, "Export History", f"Failed to export CSV:\n{e}")


    def show_alert(self, message: str, alert_type: str = "error", location: str = "global"):
        """Show an alert box with message and style depending on type."""
        if location == "global":
            target_frame = self.alert_frame
            target_label = self.alert_label
        elif location == "details":
            target_frame = self.details_alert_frame
            target_label = self.details_alert_label

        if alert_type == "error":
            target_frame.setStyleSheet("""
                QFrame { background-color: #fee2e2; border: 1px solid #f87171; border-radius: 6px; }
                QLabel { color: #991b1b; font-weight: bold; padding: 6px; }
            """)
        elif alert_type == "success":
            target_frame.setStyleSheet("""
                QFrame { background-color: #dcfce7; border: 1px solid #4ade80; border-radius: 6px; }
                QLabel { color: #166534; font-weight: bold; padding: 6px; }
            """)
        elif alert_type == "warning":
            target_frame.setStyleSheet("""
                QFrame { background-color: #fef9c3; border: 1px solid #facc15; border-radius: 6px; }
                QLabel { color: #713f12; font-weight: bold; padding: 6px; }
            """)
        else:
            target_frame.setStyleSheet("")  # Default

        target_label.setText(message)
        target_frame.show()

    def hide_alert(self, location: str = "global"):
        """Completely hide the alert frame."""
        if location == "global":
            self.alert_frame.hide()
        elif location == "details":
            self.details_alert_frame.hide()


    # -------------------------
    # Export
    # -------------------------
    def _doc_export_text(self, item: QTreeWidgetItem, doc_key: str) -> str:
        payload = item.data(BOM_COL_FILES, BOM_TREE_FILES_ROLE) or {}
        value = payload.get(doc_key)
        if isinstance(value, (tuple, list)) and value:
            state = str(value[0] or "na").lower()
            labels = {
                "ok": "OK",
                "outdated": "Outdated",
                "missing": "Missing",
                "na": "Not attached",
                "bad": "Missing",
                "ack": "OK",
            }
            return labels.get(state, state.title() if state else "Unknown")
        return "Unknown"

    def _doc_export_detail_text(self, item: QTreeWidgetItem, doc_key: str) -> str:
        payload = item.data(BOM_COL_FILES, BOM_TREE_FILES_ROLE) or {}
        value = payload.get(doc_key)
        if isinstance(value, (tuple, list)) and len(value) > 1:
            return str(value[1] or "").strip()
        return ""

    def _integrity_export_text(self, item: QTreeWidgetItem) -> str:
        payload = item.data(BOM_COL_INTEGRITY, BOM_TREE_INTEGRITY_ROLE) or {}
        state = str(payload.get("state") or "ok").lower()
        return "Has issues" if state == "warn" else "Healthy"

    def _issue_export_text(self, item: QTreeWidgetItem) -> str:
        summary = item.data(0, BOM_TREE_ISSUE_ROLE) or {}
        active = int(summary.get("active_count") or 0)
        total = int(summary.get("total_count") or 0)
        if active:
            return f"{active} active / {total} linked"
        if total:
            return f"0 active / {total} linked"
        return "No linked issues"

    def _active_filter_summary(self) -> list[tuple[str, str]]:
        filters = dict(getattr(self, "_bom_advanced_filters", {}) or self._default_bom_advanced_filters())
        defaults = self._default_bom_advanced_filters()
        labels = {
            "text": "Contains",
            "text_match_mode": "Text match mode",
            "work_state": "Work state",
            "work_owner": "Owner",
            "status": "Status",
            "type": "Type",
            "revision": "Revision",
            "categories": "Categories",
            "structure": "Structure",
            "pdf": "PDF",
            "step": "STEP",
            "integrity": "Integrity",
            "issues": "Issues",
            "remove_duplicates": "Duplicate items",
            "show_parent_matches": "Show parent branches",
            "expand_matches": "Expand matches",
        }
        rows = []
        for key, label in labels.items():
            value = filters.get(key, defaults.get(key, ""))
            if key == "text_match_mode" and not str(filters.get("text") or "").strip():
                continue
            if value != defaults.get(key, ""):
                if isinstance(value, (list, tuple, set)):
                    rows.append((label, ", ".join(str(item) for item in value)))
                elif key == "text_match_mode":
                    rows.append((label, "Match whole word" if value == "whole_word" else "Normal filter"))
                elif key == "remove_duplicates":
                    rows.append((label, "First occurrence only"))
                else:
                    rows.append((label, str(value)))
        search_query = ""
        try:
            search_query = self.search_input.text().strip()
        except Exception:
            pass
        if search_query:
            rows.append(("Search mode", "Current search results"))
        return rows or [("Filter", "None - all visible BOM rows")]

    def _collect_visible_bom_export_rows(self) -> list[dict]:
        if getattr(self, "_bom_mode", "cad") == "ebom":
            tree = self._current_pdm_tree() or getattr(self, "_ebom_tree", self.tree)
        else:
            tree = self._current_tree_for_filtering()
        rows: list[dict] = []

        def recurse(item: QTreeWidgetItem, level: int, ancestor_visible: bool = True):
            visible = ancestor_visible and not item.isHidden()
            if not visible:
                return
            part_id = item.data(0, Qt.UserRole)
            issue_text = self._issue_export_text(item)
            integrity_payload = item.data(BOM_COL_INTEGRITY, BOM_TREE_INTEGRITY_ROLE) or {}
            details = []
            for tip in (item.toolTip(BOM_COL_NAME), item.toolTip(BOM_COL_FILES), integrity_payload.get("tooltip")):
                tip = str(tip or "").strip()
                if tip and tip not in details:
                    details.append(tip)
            rows.append({
                "level": level,
                "name": item.text(BOM_COL_NAME),
                "part_number": item.data(0, BOM_TREE_ITEM_NUMBER_ROLE) or "",
                "aes_number": item.data(0, BOM_TREE_AES_NUMBER_ROLE) or "",
                "type": item.text(BOM_COL_TYPE),
                "revision": item.text(BOM_COL_REV),
                "categories": ", ".join(item.data(0, BOM_TREE_CATEGORY_ROLE) or []),
                "status": item.text(BOM_COL_STATUS),
                "work_state": item.data(0, BOM_TREE_INWORK_ROLE) or "Checked In",
                "pdf_status": self._doc_export_text(item, "pdf"),
                "pdf_details": self._doc_export_detail_text(item, "pdf"),
                "step_status": self._doc_export_text(item, "step"),
                "step_details": self._doc_export_detail_text(item, "step"),
                "integrity": self._integrity_export_text(item),
                "issues": issue_text,
                "part_id": "" if part_id is None else str(part_id),
                "details": "\n".join(details),
            })
            if not item.isExpanded():
                return
            for child_index in range(item.childCount()):
                recurse(item.child(child_index), level + 1, visible)

        for top_index in range(tree.topLevelItemCount()):
            recurse(tree.topLevelItem(top_index), 0)
        return rows

    def _export_visible_bom_csv(self, file_path: str, rows: list[dict]) -> None:
        fieldnames = [
            "level", "part_number", "name", "aes_number", "type", "revision", "categories", "status",
            "work_state", "pdf_status", "pdf_details", "step_status", "step_details",
            "integrity", "issues", "part_id", "details",
        ]
        with open(file_path, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    def _export_visible_bom_xlsx(self, file_path: str, rows: list[dict]) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError(
                "Excel export requires openpyxl. Install dependencies from requirements.txt, then try again."
            ) from exc

        headers = [
            ("level", "Level"),
            ("part_number", "Number"),
            ("name", "Name"),
            ("aes_number", "AES Number"),
            ("type", "Type"),
            ("revision", "Revision"),
            ("categories", "Categories"),
            ("status", "Status"),
            ("work_state", "Work State"),
            ("pdf_status", "PDF Status"),
            ("pdf_details", "PDF Details"),
            ("step_status", "STEP Status"),
            ("step_details", "STEP Details"),
            ("integrity", "Integrity"),
            ("issues", "Issues"),
            ("part_id", "Internal Item ID"),
            ("details", "Details"),
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Filtered BOM"

        title = "BOM Export"
        if not self._is_default_bom_advanced_filter() or getattr(self, "_in_search_mode", False):
            title = "Filtered BOM Export"
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14, color="111827")
        ws["A2"] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A3"] = f"Rows: {len(rows)}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))

        header_row = 5
        for col, (_, label) in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_index, row in enumerate(rows, start=header_row + 1):
            for col, (key, _) in enumerate(headers, start=1):
                value = row.get(key, "")
                cell = ws.cell(row=row_index, column=col, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=(key == "details"))
                if key == "name":
                    cell.alignment = Alignment(indent=min(int(row.get("level") or 0), 10), vertical="top")
                    if int(row.get("level") or 0) == 0:
                        cell.font = Font(bold=True)
            try:
                ws.row_dimensions[row_index].outlineLevel = min(int(row.get("level") or 0), 7)
            except Exception:
                pass

        ws.freeze_panes = "A6"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(header_row, header_row + len(rows))}"
        widths = {
            "A": 8, "B": 32, "C": 16, "D": 12, "E": 12, "F": 28, "G": 14,
            "H": 18, "I": 14, "J": 14, "K": 32, "L": 16, "M": 32, "N": 20,
            "O": 18, "P": 10, "Q": 54,
        }
        for column, width in widths.items():
            ws.column_dimensions[column].width = width

        summary = wb.create_sheet("Export Filter")
        summary["A1"] = "Filter"
        summary["B1"] = "Value"
        summary["A1"].font = Font(bold=True, color="FFFFFF")
        summary["B1"].font = Font(bold=True, color="FFFFFF")
        summary["A1"].fill = PatternFill("solid", fgColor="1F2937")
        summary["B1"].fill = PatternFill("solid", fgColor="1F2937")
        for idx, (label, value) in enumerate(self._active_filter_summary(), start=2):
            summary.cell(row=idx, column=1, value=label)
            summary.cell(row=idx, column=2, value=value)
        summary.column_dimensions["A"].width = 24
        summary.column_dimensions["B"].width = 48
        wb.save(file_path)

    def export_bom(self):
        if getattr(self, "_bom_mode", "cad") == "ebom":
            default_name = (
                "item_structure_visible.csv"
                if self._is_default_bom_advanced_filter()
                else "item_structure_filtered.csv"
            )
            file_path, selected_filter = QFileDialog.getSaveFileName(
                self,
                "Export Visible Item Structure",
                default_name,
                "CSV Files (*.csv);;Excel Workbook (*.xlsx);;All Files (*)",
            )
            if not file_path:
                return
            try:
                rows = self._collect_visible_bom_export_rows()
                if not rows:
                    QMessageBox.warning(
                        self,
                        "Export Visible Item Structure",
                        "There are no visible Item rows to export.",
                    )
                    return
                lower = file_path.lower()
                wants_xlsx = "excel" in selected_filter.lower() or lower.endswith(".xlsx")
                if wants_xlsx:
                    if not lower.endswith(".xlsx"):
                        file_path += ".xlsx"
                    self._export_visible_bom_xlsx(file_path, rows)
                else:
                    if not lower.endswith(".csv"):
                        file_path += ".csv"
                    self._export_visible_bom_csv(file_path, rows)
                QMessageBox.information(
                    self,
                    "Export Visible Item Structure",
                    f"Exported {len(rows)} visible Item row(s) to {file_path}.",
                )
            except Exception as exc:
                QMessageBox.critical(self, "Export Visible Item Structure", str(exc))
            return
        if getattr(self, "_bom_mode", "cad") == "cad":
            file_path, _selected_filter = QFileDialog.getSaveFileName(
                self,
                "Export CAD Structure",
                "cad_structure.csv",
                "CSV Files (*.csv)",
            )
            if not file_path:
                return
            if not file_path.lower().endswith(".csv"):
                file_path += ".csv"
            rows = []

            def collect(item: QTreeWidgetItem, level: int, parent_path: str) -> None:
                if item.isHidden():
                    return
                payload = dict(item.data(0, PDM_CAD_PAYLOAD_ROLE) or {})
                label = str(item.text(CAD_COL_NAME) or "")
                path = f"{parent_path}/{label}" if parent_path else label
                rows.append({
                    "level": level,
                    "path": path,
                    "cad_file": item.text(CAD_COL_NAME),
                    "description": item.text(CAD_COL_DESCRIPTION),
                    "category": item.text(CAD_COL_CATEGORY),
                    "cad_creo_version": item.text(CAD_COL_REV),
                    "lifecycle": item.text(CAD_COL_STATE),
                    "related_item": item.text(CAD_COL_ASSOCIATION),
                    "checkout": item.text(CAD_COL_CHECKOUT),
                    "build": item.text(CAD_COL_BUILD),
                    "quantity": item.text(CAD_COL_QTY),
                })
                for index in range(item.childCount()):
                    collect(item.child(index), level + 1, path)

            for index in range(self._cad_tree.topLevelItemCount()):
                collect(self._cad_tree.topLevelItem(index), 0, "")
            if not rows:
                QMessageBox.warning(self, "Export CAD Structure", "There are no visible CAD rows to export.")
                return
            try:
                with open(file_path, "w", newline="", encoding="utf-8-sig") as handle:
                    writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
                    writer.writeheader()
                    writer.writerows(rows)
                QMessageBox.information(
                    self, "Export CAD Structure",
                    f"Exported {len(rows)} CAD Document occurrence(s) to {file_path}.",
                )
            except Exception as exc:
                QMessageBox.critical(self, "Export CAD Structure", str(exc))
            return
        default_name = "bom_filtered.xlsx" if not self._is_default_bom_advanced_filter() else "bom.xlsx"
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Export BOM",
            default_name,
            "Excel Workbook (*.xlsx);;CSV Files (*.csv);;All Files (*)",
        )
        if not file_path:
            return
        try:
            rows = self._collect_visible_bom_export_rows()
            if not rows:
                QMessageBox.warning(self, "Export BOM", "There are no visible BOM rows to export.")
                return

            lower = file_path.lower()
            wants_csv = "csv" in selected_filter.lower() or lower.endswith(".csv")
            if wants_csv:
                if not lower.endswith(".csv"):
                    file_path += ".csv"
                self._export_visible_bom_csv(file_path, rows)
            else:
                if not lower.endswith(".xlsx"):
                    file_path += ".xlsx"
                self._export_visible_bom_xlsx(file_path, rows)
            QMessageBox.information(self, "Success", f"BOM exported to {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export BOM: {str(e)}")

    def _make_indicator_icon(self, pdf_ok, step_ok, step_present: bool = True) -> QIcon:
        """Create a clearer 2-part badge icon: left=PDF, right=STEP.

        Each side is a small rounded badge labeled with 'PDF' / 'STEP'.
        Colors: ok=green, ack=blue, bad=red, absent=gray. Keeps cache for speed.
        """
        def _state(value):
            if isinstance(value, bool):
                return "ok" if value else "bad"
            value = str(value or "").lower()
            return value if value in {"ok", "ack", "bad", "absent"} else "absent"

        pdf_state = _state(pdf_ok)
        step_state = _state(step_ok)
        cache_key = (pdf_state, step_state, bool(step_present))
        try:
            icon = getattr(self, "_indicator_icon_cache", {}).get(cache_key)
            if icon:
                return icon
        except Exception:
            self._indicator_icon_cache = {}

        # Badge dimensions (wide enough for short label)
        h = 12
        badge_w = 24
        gap = 2
        total_w = badge_w * 2 + gap if step_present else badge_w
        pm = QPixmap(total_w, h)
        pm.fill(Qt.transparent)

        painter = QPainter(pm)
        painter.setRenderHint(QPainter.Antialiasing, True)

        colors = {
            "ok": QColor(34, 197, 94),    # green
            "ack": QColor(14, 165, 233),  # blue
            "bad": QColor(239, 68, 68),   # red
            "absent": QColor(156, 163, 175),
        }

        def draw_badge(x: int, label: str, state: str):
            bg = colors.get(state, colors["absent"])
            painter.setPen(Qt.NoPen)
            painter.setBrush(bg)
            rect = QRect(x, 0, badge_w, h)
            painter.drawRoundedRect(rect, 3, 3)

            # Text color: white for colored states, dark for absent
            if state == "absent":
                text_color = QColor(55, 65, 81)
            else:
                text_color = QColor(255, 255, 255)

            f = painter.font()
            f.setPointSize(6)
            f.setBold(True)
            painter.setFont(f)
            painter.setPen(text_color)
            painter.drawText(rect, Qt.AlignCenter, label)

            # ACK state: small inner ring to indicate acknowledged
            if state == "ack":
                painter.setPen(QColor(255, 255, 255, 200))
                painter.setBrush(Qt.NoBrush)
                painter.drawRoundedRect(rect.adjusted(3, 3, -3, -3), 4, 4)

        # Draw PDF badge (left)
        draw_badge(0, "PDF", pdf_state)

        # Draw STEP badge (right) if present
        if step_present:
            draw_badge(badge_w + gap, "STEP", step_state)

        painter.end()
        icon = QIcon(pm)
        try:
            self._indicator_icon_cache[cache_key] = icon
        except Exception:
            self._indicator_icon_cache = {cache_key: icon}
        return icon

    def _refresh_current_tree_item_indicator(self):
        """Refresh the selected row's PDF/STEP indicator in all visible trees."""
        try:
            part_id = getattr(self, "current_part_id", None)
            if not part_id:
                current_tree = self._current_tree_for_filtering()
                item = current_tree.currentItem() if current_tree else None
                part_id = item.data(0, Qt.UserRole) if item else None
            if not part_id:
                return

            self._refresh_part_in_tree(int(part_id))
            # also refresh details warning banner (if open)
            try:
                if getattr(self, "current_part_id", None) == part_id:
                    self.display_details(int(part_id))
            except Exception:
                pass
        except Exception:
            pass

    def show_files_context_menu(self, position):
        if not getattr(self, "current_part_id", None):
            return
        file_id = self._selected_attachment_id()
        if not file_id:
            return

        file_type = self._selected_file_type()

        menu = QMenu(self)
        if file_type in ("PDF", "STEP"):
            act = QAction(f"Acknowledge {file_type} as safe", self)

            def _do():
                ack_target = self._doc_ack_target(int(self.current_part_id), file_type)
                self.part_doc_ack_service.mark_up_to_date(
                    int(self.current_part_id), file_type, ack_target
                )
                # fast refresh (avoid full tree rebuild)
                self._refresh_current_tree_item_indicator()

            act.triggered.connect(_do)
            menu.addAction(act)

        menu.exec_(self.files_table.viewport().mapToGlobal(position))

    def show_versions_context_menu(self, position):
        if not getattr(self, "current_part_id", None):
            return
        version_id = self._selected_version_id()
        if not version_id:
            return

        ver = self.part_file_service.repo.get_version_by_id(int(version_id))
        if not ver:
            return
        pf = self.part_file_service.repo.get_file_by_id(int(getattr(ver, "file_id", 0) or 0))
        if not pf:
            return

        file_type = str(getattr(pf, "file_type", "") or "").upper()
        menu = QMenu(self)
        edit_revision_act = QAction("Edit Revision", self)
        edit_note_act = QAction("Edit Note", self)

        def _edit_revision():
            current = str(getattr(ver, "revision", "") or "").strip()
            revision, ok = QInputDialog.getText(
                self,
                "Edit Version Revision",
                "Revision (blank allowed, e.g. A010):",
                text=current,
            )
            if not ok:
                return
            try:
                self.part_file_service.update_version_revision(int(version_id), revision)
                self.on_attachment_selected()
            except Exception as exc:
                QMessageBox.critical(self, "Edit Revision", f"Failed to update revision:\n{exc}")

        def _edit_note():
            current = str(getattr(ver, "note", "") or "")
            note, ok = QInputDialog.getText(
                self,
                "Edit Version Note",
                "Note:",
                text=current,
            )
            if not ok:
                return
            try:
                self.part_file_service.update_version_note(int(version_id), note)
                self.on_attachment_selected()
            except Exception as exc:
                QMessageBox.critical(self, "Edit Note", f"Failed to update note:\n{exc}")

        edit_revision_act.triggered.connect(_edit_revision)
        edit_note_act.triggered.connect(_edit_note)
        can_edit = self.perm.can("release_files")
        edit_revision_act.setEnabled(can_edit)
        edit_note_act.setEnabled(can_edit)
        menu.addAction(edit_revision_act)
        menu.addAction(edit_note_act)
        menu.addSeparator()

        if file_type in ("PDF", "STEP"):
            act = QAction(f"Acknowledge {file_type} as safe", self)

            def _do():
                ack_target = self._doc_ack_target(int(self.current_part_id), file_type)
                self.part_doc_ack_service.mark_up_to_date(
                    int(self.current_part_id), file_type, ack_target
                )
                # fast refresh (avoid full tree rebuild)
                self._refresh_current_tree_item_indicator()

            act.triggered.connect(_do)
            menu.addAction(act)

        menu.exec_(self.versions_table.viewport().mapToGlobal(position))

