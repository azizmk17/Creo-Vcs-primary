import json
import os
import re
import shutil
import zipfile
from datetime import datetime
from html import escape
from typing import Iterable, Optional

from core.repositories.issue_repository import IssueRepository
from core.repositories.traceability_repository import TraceabilityRepository
from core.services.base_service import BaseService
from core.services.part_file_service import PartFileService


class TraceabilityService(BaseService):
    """Application service for engineering traceability links and reports."""

    def __init__(self, repo: Optional[TraceabilityRepository] = None):
        super().__init__()
        self.repo = repo or TraceabilityRepository()
        self.issue_repo = IssueRepository()
        self.part_file_service = PartFileService()

    def link_jira(self, issue_id: int, jira_key: str = "", jira_url: str = "",
                  jira_summary: str = "", jira_status: str = "") -> int:
        return self.repo.link_jira(
            issue_id=issue_id,
            jira_key=jira_key,
            jira_url=jira_url,
            actor_id=self.user_id,
            jira_summary=(jira_summary or None),
            jira_status=(jira_status or None),
        )

    def jira_links(self, issue_id: int) -> list[dict]:
        return self.repo.jira_links(issue_id)

    def link_issue_to_commit(self, issue_ids: Iterable[int], commit_id: str,
                             relation_type: str = "solves", note: str = ""):
        self.repo.link_issue_to_commit(issue_ids, commit_id, self.user_id, relation_type, note)

    def link_issue_to_engineering_file(self, issue_id: int, part_file_id: int,
                                       version_id: Optional[int] = None,
                                       role: str = "other", note: str = "") -> int:
        return self.repo.link_issue_to_engineering_file(
            issue_id=issue_id,
            part_file_id=part_file_id,
            version_id=version_id,
            role=role,
            actor_id=self.user_id,
            note=note,
        )

    def engineering_files_for_issue(self, issue_id: int) -> list[dict]:
        return self.repo.engineering_files_for_issue(issue_id)

    def issues_for_engineering_file(self, part_file_id: int) -> list[dict]:
        return self.repo.issues_for_engineering_file(part_file_id)

    def commit_links_for_issue(self, issue_id: int) -> list[dict]:
        return self.repo.commit_links_for_issue(issue_id)

    def mark_commit_reverted(self, commit_id: str, project_id: Optional[int] = None, note: str = "") -> bool:
        return self.repo.mark_commit_reverted(commit_id, project_id, self.user_id, note)

    def get_issue_traceability(self, issue_id: int) -> dict:
        return self.repo.get_issue_traceability(issue_id)

    def export_issue_traceability(self, filters: Optional[dict] = None,
                                  include_engineering_files: bool = True) -> list[dict]:
        return self.repo.export_issue_traceability(
            int(self.project_id),
            filters or {},
            include_engineering_files=include_engineering_files,
        )

    def export_issue_traceability_json(self, destination_path: str, filters: Optional[dict] = None,
                                       include_engineering_files: bool = True) -> str:
        if not destination_path:
            raise ValueError("Destination path is required")
        folder = os.path.dirname(destination_path)
        if folder:
            os.makedirs(folder, exist_ok=True)
        data = self.export_issue_traceability(filters, include_engineering_files)
        with open(destination_path, "w", encoding="utf-8") as handle:
            json.dump(data, handle, indent=2, default=str)
        return destination_path

    def export_issue_traceability_package(self, issue_id: int, destination_dir: str) -> dict:
        if not issue_id:
            raise ValueError("Issue ID is required")
        if not destination_dir:
            raise ValueError("Destination folder is required")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception:
            Workbook = None
            Alignment = Font = PatternFill = get_column_letter = None

        report = self.get_issue_traceability(int(issue_id))
        issue = report.get("issue") or {}
        issue_number = issue.get("issue_number") or f"issue_{issue_id}"
        title = issue.get("title") or "untitled"
        package_name = self._safe_filename(
            f"{issue_number}_{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        )
        package_dir = os.path.join(destination_dir, package_name)
        input_dir = os.path.join(package_dir, "input_data")
        output_dir = os.path.join(package_dir, "output_engineering_files")
        os.makedirs(input_dir, exist_ok=True)
        os.makedirs(output_dir, exist_ok=True)

        engineering_links = self._expanded_engineering_links(report.get("engineering_files") or [])
        copied_inputs = self._copy_issue_attachments(int(issue_id), input_dir)
        copied_outputs = self._copy_engineering_files(engineering_links, output_dir)

        workbook_path = os.path.join(package_dir, "issue_traceability.xlsx")
        sheets = []
        wb = Workbook() if Workbook else None
        if wb:
            wb.remove(wb.active)

        def add_sheet(name, headers, rows, description=""):
            sheets.append((name, headers, rows))
            if not wb:
                return None
            ws = wb.create_sheet(name[:31])
            ws.sheet_view.showGridLines = False
            ws.append([name])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
            title_cell = ws.cell(row=1, column=1)
            title_cell.font = Font(bold=True, size=16, color="FFFFFF")
            title_cell.fill = PatternFill("solid", fgColor="1F4E78")
            title_cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[1].height = 24
            if description:
                ws.append([description])
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(headers)))
                ws.cell(row=2, column=1).font = Font(italic=True, color="666666")
                ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            else:
                ws.append([""])
            ws.append(headers)
            header_fill = PatternFill("solid", fgColor="D9EAF7")
            for cell in ws[3]:
                cell.font = Font(bold=True, color="1F2937")
                cell.fill = header_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            for row in rows:
                ws.append([self._cell_value(v) for v in row])
            ws.freeze_panes = "A4"
            if rows:
                ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{ws.max_row}"
            for col_idx, header in enumerate(headers, start=1):
                width = min(max(len(str(header)) + 2, 14), 48)
                for row_idx in range(4, min(ws.max_row, 42) + 1):
                    value = ws.cell(row=row_idx, column=col_idx).value
                    if value:
                        width = min(max(width, min(len(str(value)) + 2, 80)), 80)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            return ws

        add_sheet(
            "Overview",
            ["Metric", "Value"],
            [
                ["Issue", f"{issue.get('issue_number') or issue_id} - {issue.get('title') or ''}"],
                ["Status", issue.get("status")],
                ["Priority", issue.get("priority")],
                ["Input attachments copied", len(copied_inputs)],
                ["Engineering outputs copied", sum(1 for x in copied_outputs if x.get("exists"))],
                ["Engineering output links", len(copied_outputs)],
                ["Linked commits", len(report.get("linked_commits") or [])],
                ["Native Creo files", len(report.get("native_creo_files") or [])],
                ["Created at", datetime.now().isoformat(timespec="seconds")],
                ["Package folder", package_dir],
            ],
            "Package summary for the exported issue traceability bundle.",
        )
        add_sheet("Issue", ["Field", "Value"], sorted((issue or {}).items()),
                  "Complete internal issue fields including title, description, status, dates, and users.")
        add_sheet(
            "Jira",
            ["ID", "Jira Key", "Jira URL", "Summary", "Status", "Last Checked", "Created At"],
            [[j.get("id"), j.get("jira_key"), j.get("jira_url"), j.get("jira_summary"),
              j.get("jira_status"), j.get("last_checked_at"), j.get("created_at")]
             for j in report.get("jira_links") or []],
            "External Jira references attached to the internal issue.",
        )
        add_sheet(
            "Linked Commits",
            ["Commit ID", "Relation", "Validation", "Title", "Message", "Author", "Committed At",
             "Group Status", "Commit Status", "Merged By", "Merged At", "Reverted At", "Revert Note"],
            [[c.get("commit_id"), c.get("relation_type"), c.get("validation_status"),
              c.get("title"), self._plain_text(c.get("message")), c.get("author_name"),
              c.get("committed_at"), c.get("group_status"), c.get("commit_status"),
              c.get("merged_by_name"), c.get("merged_at"), c.get("reverted_at"), c.get("revert_note")]
             for c in report.get("linked_commits") or []],
            "Logical commits linked to the issue and their resolution state.",
        )
        commit_file_rows = []
        for commit in report.get("linked_commits") or []:
            for f in commit.get("files_changed") or []:
                commit_file_rows.append([
                    commit.get("commit_id"), f.get("commit_row_id"), f.get("filename"),
                    f.get("change_type"), f.get("type"), f.get("status"),
                    f.get("part_id"), f.get("part_name"), f.get("aes_number"),
                    f.get("base_file_name"), f.get("file_path"), f.get("committed_at"),
                ])
        add_sheet(
            "Commit Files",
            ["Commit ID", "Row ID", "Filename", "Change Type", "Type", "Status", "Part ID",
             "Part Name", "AES", "Base File", "Source Path", "Committed At"],
            commit_file_rows,
            "Every file row changed by linked commits.",
        )
        add_sheet(
            "Native Creo Files",
            ["Part ID", "Name", "Type", "Filename", "Drawing", "Base CAD", "Base DRW",
             "AES", "Part Number", "Drawing Number", "Revision", "Lifecycle"],
            [[p.get("id"), p.get("name"), p.get("type"), p.get("filename"), p.get("drawing"),
              p.get("base_file_name"), p.get("base_drw_name"), p.get("aes_number"),
             p.get("part_number"), p.get("drawing_number"), p.get("revision"), p.get("lifecycle_state")]
             for p in report.get("native_creo_files") or []],
            "Native Creo files affected by or related to the issue.",
        )
        add_sheet(
            "Input Attachments",
            ["Attachment ID", "Filename", "Source Path", "Package Path", "Exists", "Created At", "Uploaded By"],
            [[a.get("id"), a.get("file_name"), a.get("source_path"), a.get("package_path"),
              a.get("exists"), a.get("created_at"), a.get("uploaded_by")]
             for a in copied_inputs],
            "Issue attachments copied as input/reference data.",
        )
        add_sheet(
            "Engineering Outputs",
            ["Link ID", "Role", "Part File ID", "Version ID", "Display Name", "File Type",
             "Original Filename", "Package Path", "Source Path", "Exists", "Version No",
             "Revision", "SHA256", "Size", "Part ID", "Part Name", "AES", "Linked At", "Note",
             "Source Candidates"],
            [[e.get("id"), e.get("file_role"), e.get("part_file_id"), e.get("resolved_version_id"),
              e.get("display_name"), e.get("file_type"), e.get("original_filename"),
              e.get("package_path"), e.get("source_path"), e.get("exists"), e.get("version_no"),
              e.get("revision"), e.get("sha256"), e.get("size_bytes"), e.get("part_id"), e.get("part_name"),
              e.get("aes_number"), e.get("linked_at"), e.get("note"),
              "\n".join(e.get("source_candidates") or [])]
             for e in copied_outputs],
            "Linked vaulted engineering output files copied into output_engineering_files.",
        )
        add_sheet(
            "Timeline",
            ["ID", "Action", "Field", "Old Value", "New Value", "Details", "User", "Created At"],
            [[h.get("id"), h.get("action"), h.get("field_name"), h.get("old_value"),
              h.get("new_value"), h.get("details_json"), h.get("username"), h.get("created_at")]
             for h in report.get("timeline") or []],
            "Chronological issue history and traceability events.",
        )
        if wb:
            wb.save(workbook_path)
        else:
            self._write_minimal_xlsx(workbook_path, sheets)

        manifest = {
            "package_dir": package_dir,
            "workbook": workbook_path,
            "input_data_dir": input_dir,
            "output_engineering_files_dir": output_dir,
            "issue_id": int(issue_id),
            "issue_number": issue_number,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "input_files": copied_inputs,
            "output_files": copied_outputs,
        }
        with open(os.path.join(package_dir, "manifest.json"), "w", encoding="utf-8") as handle:
            json.dump(manifest, handle, indent=2, default=str)
        return manifest

    def _copy_issue_attachments(self, issue_id: int, destination_dir: str) -> list[dict]:
        copied = []
        for attachment in self.issue_repo.attachments(issue_id):
            src = str(attachment.get("file_path") or "")
            dst = ""
            exists = bool(src and os.path.exists(src))
            if exists:
                dst = self._unique_path(destination_dir, attachment.get("file_name") or os.path.basename(src))
                shutil.copy2(src, dst)
            copied.append({
                **attachment,
                "source_path": src,
                "package_path": dst,
                "exists": exists,
            })
        return copied

    def _expanded_engineering_links(self, links: list[dict]) -> list[dict]:
        expanded = []
        seen = set()
        for link in links:
            part_file_id = link.get("part_file_id")
            explicit_version_id = link.get("part_file_version_id")
            versions = []
            if part_file_id and not explicit_version_id:
                try:
                    versions = self.part_file_service.repo.get_versions(int(part_file_id)) or []
                except Exception:
                    versions = []
            if not versions:
                key = (link.get("id"), link.get("resolved_version_id") or explicit_version_id)
                if key not in seen:
                    seen.add(key)
                    expanded.append(dict(link))
                continue
            for version in versions:
                key = (link.get("id"), version.id)
                if key in seen:
                    continue
                seen.add(key)
                row = dict(link)
                row.update({
                    "resolved_version_id": version.id,
                    "version_no": version.version_no,
                    "original_filename": version.original_filename,
                    "vault_rel_path": version.vault_rel_path,
                    "sha256": version.sha256,
                    "size_bytes": version.size_bytes,
                    "version_created_at": version.created_at,
                    "revision": getattr(version, "revision", None) or link.get("revision"),
                    "expanded_from_file_level_link": True,
                })
                expanded.append(row)
        return expanded

    def _copy_engineering_files(self, links: list[dict], destination_dir: str) -> list[dict]:
        copied = []
        for link in links:
            candidates = self._engineering_source_candidates(link)
            src = next((path for path in candidates if path and os.path.exists(path)), "")
            exists = bool(src and os.path.exists(src))
            dst = ""
            if exists:
                role = self._safe_filename(str(link.get("file_role") or "engineering_file"))
                part = self._safe_filename(str(link.get("part_name") or link.get("part_id") or "part"))
                filename = link.get("original_filename") or link.get("display_name") or os.path.basename(src)
                dst = self._unique_path(os.path.join(destination_dir, role), f"{part}_{filename}")
                os.makedirs(os.path.dirname(dst), exist_ok=True)
                shutil.copy2(src, dst)
            copied.append({
                **link,
                "source_path": src,
                "package_path": dst,
                "exists": exists,
                "source_candidates": candidates,
            })
        return copied

    def _engineering_source_candidates(self, link: dict) -> list[str]:
        candidates = []

        def add(path):
            if not path:
                return
            path = os.path.normpath(str(path))
            if path not in candidates:
                candidates.append(path)

        version = None
        version_id = link.get("resolved_version_id") or link.get("part_file_version_id")
        if version_id:
            try:
                version = self.part_file_service.repo.get_version_by_id(int(version_id))
            except Exception:
                version = None
        if version:
            add(self.part_file_service.resolve_version_path(version))
            vault_rel_path = getattr(version, "vault_rel_path", None)
            original_filename = getattr(version, "original_filename", None)
            root_project_id = getattr(version, "root_project_id", None)
            project_version_label = getattr(version, "project_version_label", None)
        else:
            vault_rel_path = link.get("vault_rel_path")
            original_filename = link.get("original_filename") or link.get("display_name")
            root_project_id = link.get("root_project_id")
            project_version_label = link.get("project_version_label")

        if vault_rel_path:
            if os.path.isabs(str(vault_rel_path)):
                add(str(vault_rel_path))
            working_dirs = []
            for root_id, label in (
                (root_project_id, project_version_label),
                (root_project_id, "A"),
                (None, None),
            ):
                try:
                    if root_id and label:
                        project = self.part_file_service.project_service.get_project_by_root_and_label(
                            int(root_id), str(label)
                        )
                        wd = (project or {}).get("working_directory") or ""
                    else:
                        wd = self.part_file_service._working_dir()
                except Exception:
                    wd = ""
                if wd and wd not in working_dirs:
                    working_dirs.append(wd)
            for wd in working_dirs:
                add(os.path.join(wd, str(vault_rel_path)))
                add(os.path.join(wd, "vault", os.path.basename(str(vault_rel_path))))

        for found in self._find_vault_file_candidates(original_filename, link.get("sha256")):
            add(found)
        return candidates

    def _find_vault_file_candidates(self, filename: str, sha256: str = "") -> list[str]:
        filename = os.path.basename(str(filename or ""))
        if not filename:
            return []
        roots = []
        try:
            wd = self.part_file_service._working_dir()
            if wd:
                roots.append(os.path.join(wd, "vault"))
                roots.append(wd)
        except Exception:
            pass
        matches = []
        target_hash = str(sha256 or "").strip().lower()
        for root in roots:
            if not root or not os.path.isdir(root):
                continue
            for dirpath, _dirnames, filenames in os.walk(root):
                if filename not in filenames:
                    continue
                path = os.path.join(dirpath, filename)
                if target_hash:
                    try:
                        if self.part_file_service._hash_file_sha256(path).lower() != target_hash:
                            continue
                    except Exception:
                        pass
                matches.append(path)
        return matches

    def _safe_filename(self, value: str) -> str:
        value = self._plain_text(value or "").strip() or "export"
        value = re.sub(r"\s+", "_", value)
        return "".join(ch if ch.isalnum() or ch in "._-()" else "_" for ch in value)[:150]

    def _unique_path(self, folder: str, filename: str) -> str:
        os.makedirs(folder, exist_ok=True)
        safe = self._safe_filename(os.path.basename(filename))
        stem, ext = os.path.splitext(safe)
        candidate = os.path.join(folder, safe)
        idx = 2
        while os.path.exists(candidate):
            candidate = os.path.join(folder, f"{stem}_{idx}{ext}")
            idx += 1
        return candidate

    def _plain_text(self, value) -> str:
        text = str(value or "")
        text = re.sub(r"<[^>]+>", " ", text)
        return re.sub(r"\s+", " ", text).strip()

    def _cell_value(self, value):
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, default=str)
        return value

    def _xml_text(self, value) -> str:
        text = "" if value is None else str(value)
        text = "".join(
            ch for ch in text
            if ch in ("\t", "\n", "\r") or ord(ch) >= 32
        )
        return escape(text, quote=True)

    def _write_minimal_xlsx(self, destination_path: str, sheets: list[tuple]):
        """Write a dependency-free XLSX workbook with inline strings."""
        def col_ref(index: int) -> str:
            result = ""
            while index:
                index, rem = divmod(index - 1, 26)
                result = chr(65 + rem) + result
            return result

        def sheet_xml(name, headers, rows, description):
            all_rows = (
                [[name]]
                + [[description or ""]]
                + [headers]
                + [[self._cell_value(v) for v in row] for row in rows]
            )
            xml_rows = []
            for r_idx, row in enumerate(all_rows, start=1):
                cells = []
                for c_idx, value in enumerate(row, start=1):
                    ref = f"{col_ref(c_idx)}{r_idx}"
                    text = self._xml_text(value)
                    style = ' s="1"' if r_idx == 1 else (' s="2"' if r_idx == 3 else "")
                    cells.append(f'<c r="{ref}"{style} t="inlineStr"><is><t>{text}</t></is></c>')
                xml_rows.append(f'<row r="{r_idx}">{"".join(cells)}</row>')
            widths = "".join(
                f'<col min="{idx}" max="{idx}" width="{min(max(len(str(header)) + 4, 14), 42)}" customWidth="1"/>'
                for idx, header in enumerate(headers, start=1)
            )
            merge_ref = f"A1:{col_ref(max(1, len(headers)))}1"
            auto_filter = f'<autoFilter ref="A3:{col_ref(max(1, len(headers)))}{max(3, len(all_rows))}"/>' if rows else ""
            merge_cells = (
                f'<mergeCells count="1"><mergeCell ref="{merge_ref}"/></mergeCells>'
                if len(headers) > 1 else ""
            )
            return (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                '<sheetViews><sheetView workbookViewId="0" showGridLines="0">'
                '<pane ySplit="3" topLeftCell="A4" activePane="bottomLeft" state="frozen"/>'
                '</sheetView></sheetViews>'
                f'<cols>{widths}</cols>'
                f'<sheetData>{"".join(xml_rows)}</sheetData>'
                f'{auto_filter}'
                f'{merge_cells}'
                '</worksheet>'
            )

        safe_sheets = []
        used = set()
        normalized_sheets = []
        for raw in sheets:
            if len(raw) == 4:
                normalized_sheets.append(raw)
            else:
                name, headers, rows = raw
                normalized_sheets.append((name, headers, rows, ""))

        for idx, (name, headers, rows, description) in enumerate(normalized_sheets, start=1):
            clean = re.sub(r"[\[\]:*?/\\]", "_", str(name or f"Sheet{idx}"))[:31] or f"Sheet{idx}"
            base = clean
            suffix = 2
            while clean in used:
                clean = f"{base[:28]}_{suffix}"
                suffix += 1
            used.add(clean)
            safe_sheets.append((clean, headers, rows, description))

        workbook_sheets = "".join(
            f'<sheet name="{self._xml_text(name)}" sheetId="{idx}" r:id="rId{idx}"/>'
            for idx, (name, _, _, _) in enumerate(safe_sheets, start=1)
        )
        workbook_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<workbookPr date1904="0"/>'
            '<bookViews><workbookView activeTab="0"/></bookViews>'
            f'<sheets>{workbook_sheets}</sheets></workbook>'
        )
        rels = "".join(
            f'<Relationship Id="rId{idx}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{idx}.xml"/>'
            for idx in range(1, len(safe_sheets) + 1)
        )
        rels += (
            f'<Relationship Id="rId{len(safe_sheets) + 1}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
            'Target="styles.xml"/>'
        )
        workbook_rels = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f'{rels}</Relationships>'
        )
        overrides = "".join(
            f'<Override PartName="/xl/worksheets/sheet{idx}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            for idx in range(1, len(safe_sheets) + 1)
        )
        content_types = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
            '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
            f'{overrides}</Types>'
        )
        root_rels = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
            '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
            '</Relationships>'
        )
        created = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
        app_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
            'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
            '<Application>Creo VCS</Application></Properties>'
        )
        core_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
            'xmlns:dc="http://purl.org/dc/elements/1.1/" '
            'xmlns:dcterms="http://purl.org/dc/terms/" '
            'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
            'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
            '<dc:creator>Creo VCS</dc:creator>'
            '<cp:lastModifiedBy>Creo VCS</cp:lastModifiedBy>'
            f'<dcterms:created xsi:type="dcterms:W3CDTF">{created}</dcterms:created>'
            f'<dcterms:modified xsi:type="dcterms:W3CDTF">{created}</dcterms:modified>'
            '</cp:coreProperties>'
        )

        with zipfile.ZipFile(destination_path, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", content_types)
            archive.writestr("_rels/.rels", root_rels)
            archive.writestr("docProps/app.xml", app_xml)
            archive.writestr("docProps/core.xml", core_xml)
            archive.writestr("xl/workbook.xml", workbook_xml)
            archive.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
            archive.writestr("xl/styles.xml", self._minimal_xlsx_styles())
            for idx, (name, headers, rows, description) in enumerate(safe_sheets, start=1):
                archive.writestr(f"xl/worksheets/sheet{idx}.xml", sheet_xml(name, headers, rows, description))

    def _minimal_xlsx_styles(self) -> str:
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<fonts count="3">'
            '<font><sz val="11"/><name val="Calibri"/></font>'
            '<font><b/><sz val="16"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font>'
            '<font><b/><sz val="11"/><color rgb="FF1F2937"/><name val="Calibri"/></font>'
            '</fonts>'
            '<fills count="4">'
            '<fill><patternFill patternType="none"/></fill>'
            '<fill><patternFill patternType="gray125"/></fill>'
            '<fill><patternFill patternType="solid"><fgColor rgb="FF1F4E78"/><bgColor indexed="64"/></patternFill></fill>'
            '<fill><patternFill patternType="solid"><fgColor rgb="FFD9EAF7"/><bgColor indexed="64"/></patternFill></fill>'
            '</fills>'
            '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
            '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
            '<cellXfs count="3">'
            '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf>'
            '<xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyFont="1" applyFill="1" applyAlignment="1"><alignment vertical="center"/></xf>'
            '<xf numFmtId="0" fontId="2" fillId="3" borderId="0" xfId="0" applyFont="1" applyFill="1" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf>'
            '</cellXfs>'
            '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
            '</styleSheet>'
        )
